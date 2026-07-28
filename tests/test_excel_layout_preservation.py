from __future__ import annotations

import unittest
import tempfile
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage

from core.excel_layout_export import apply_excel_layout
from core.full_workbook_sync import TAB_COLORS, synchronize_project_to_workbook
from core.workbook_status_sync import project_hash
from tests.test_excel_layout_export import layout_page


class ExcelLayoutPreservationTests(unittest.TestCase):
    def test_unrelated_sheet_content_survives(self) -> None:
        wb = Workbook()
        target = wb.active
        target.title = "EMS TEST"
        source = wb.create_sheet("REFERENCE")
        source["A1"] = "=1+2"
        source.merge_cells("B2:C3")
        source.column_dimensions["A"].width = 27
        apply_excel_layout(target, layout_page())
        self.assertEqual(source["A1"].value, "=1+2")
        self.assertIn("B2:C3", {str(item) for item in source.merged_cells.ranges})
        self.assertEqual(source.column_dimensions["A"].width, 27)

    def test_layout_and_tab_color_are_conflict_hash_inputs(self) -> None:
        page = layout_page()
        project = {"metadata": {}, "worksheets": [], "pages": [page]}
        before = project_hash(project)
        page["excelLayout"]["tables"][0]["columnWidths"][0] += 1
        self.assertNotEqual(before, project_hash(project))
        before = project_hash(project)
        page["excelLayout"]["tabColor"] = "#112233"
        self.assertNotEqual(before, project_hash(project))

    def test_atomic_full_mirror_preserves_unmatched_sheet_content(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            workbook_path = root / "fixture.xlsx"
            image_path = root / "pixel.png"
            PilImage.new("RGB", (2, 2), "red").save(image_path)
            wb = Workbook()
            wb.active.title = "00_INDEX"
            wb.create_sheet("EMS TEST")
            ref = wb.create_sheet("REFERENCE")
            ref["A1"] = "=SUM(1,2)"
            ref.merge_cells("B2:C3")
            ref.column_dimensions["A"].width = 27
            ref.row_dimensions[1].height = 31
            ref.add_image(Image(image_path), "E2")
            wb.save(workbook_path)
            page = layout_page()
            page.update({
                "order": 1, "sheetCode": "EMS T.1", "displaySheetCode": "EMS T.1",
                "sheetTitle": "Sanitized Test", "publishStatus": "YES",
                "issueStatus": "draft", "pageType": "canvas", "canvasObjects": [],
            })
            project = {"id": "fixture", "metadata": {}, "worksheets": [], "pages": [page]}
            synchronize_project_to_workbook(
                workbook_path, "fixture", project, SimpleNamespace(docs=root), app_hash="fixture-hash"
            )
            reopened = load_workbook(workbook_path, data_only=False)
            self.assertEqual(reopened["REFERENCE"]["A1"].value, "=SUM(1,2)")
            self.assertIn("B2:C3", {str(item) for item in reopened["REFERENCE"].merged_cells.ranges})
            self.assertEqual(reopened["REFERENCE"].column_dimensions["A"].width, 27)
            self.assertEqual(reopened["REFERENCE"].row_dimensions[1].height, 31)
            self.assertEqual(len(reopened["REFERENCE"]._images), 1)
            self.assertEqual(reopened["EMS TEST"].page_setup.orientation, "landscape")
            self.assertGreater(len(reopened["EMS TEST"].merged_cells.ranges), 4)
            reopened.close()
            page["publishStatus"] = "NO"
            page["include"] = False
            synchronize_project_to_workbook(
                workbook_path, "fixture", project, SimpleNamespace(docs=root), app_hash="fixture-hash-2"
            )
            excluded = load_workbook(workbook_path)["EMS TEST"]
            self.assertEqual(excluded.sheet_properties.tabColor.rgb[-6:], TAB_COLORS["excluded"][-6:])


if __name__ == "__main__":
    unittest.main()
