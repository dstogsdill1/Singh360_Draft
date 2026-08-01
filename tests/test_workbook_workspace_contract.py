from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

import server
from core.full_workbook_sync import ensure_index_state_controls
from core.project_workspace import WorkbookDocumentStore
from core.workbook_importer import import_workbook
from core.workbook_status_sync import project_hash
from core.workbook_workspace import (
    apply_controlled_default_validations,
    apply_source_sheet_contract,
    detect_table_regions,
    edit_overlaps_protected,
)
from tests.generated_fixtures import isolate_server_runtime


def source_sheet() -> dict:
    return {
        "id": "src-equipment",
        "name": "SRC R-2.0 Equipment",
        "cells": {
            "A1": {"v": "Old title"},
            "A2": {"v": "Old instruction"},
            "A4": {"v": "Field"},
            "B4": {"v": "Value"},
            "C4": {"v": "Notes"},
            "A5": {"v": "Source Role"},
            "B5": {"v": "Equipment"},
            "C5": {"v": "Generated template metadata"},
            "A7": {"v": "Equipment"},
            "B7": {"v": "Model"},
            "A8": {"v": "Rack A"},
            "B8": {"v": "Fixture"},
            "D7": {"v": "Spacer should remain blank"},
            "F7": {"v": "Condenser"},
            "G7": {"v": "Model"},
            "F8": {"v": "Fixture"},
            "G8": {"f": "=B8"},
        },
        "styles": {
            "A7": {"bold": True, "fill": "#DDEBF7"},
            "G8": {"fill": "#FFFFFF"},
        },
        "merges": [],
        "rowHeights": {"8": 22.25},
        "columnWidths": {"A": 18.125, "B": 13.75, "F": 21.5},
        "defaultColumnWidth": 8.43,
        "defaultRowHeight": 15.0,
        "hiddenRows": [],
        "hiddenColumns": [],
        "archived": False,
        "tabColor": None,
        "protectedRanges": [],
        "dataValidations": [],
        "conditionalFormats": [],
        "tableRegions": [],
        "tableLayout": "side_by_side",
        "annotations": [],
    }


class WorkbookWorkspaceContractTests(unittest.TestCase):
    def test_source_contract_moves_only_template_metadata_and_protects_headers(self) -> None:
        original = source_sheet()
        contracted = apply_source_sheet_contract(
            original,
            {
                "sheetCode": "R-2.0",
                "title": "Refrigeration Equipment",
                "pageType": "Source",
                "include": "NO",
            },
        )

        self.assertEqual("Field", original["cells"]["A4"]["v"])
        self.assertNotIn("A4", contracted["cells"])
        self.assertNotIn("A5", contracted["cells"])
        self.assertNotIn("B5", contracted["cells"])
        self.assertNotIn("C5", contracted["cells"])
        self.assertNotIn("A3", contracted["cells"])
        self.assertEqual("Rack A", contracted["cells"]["A8"]["v"])
        self.assertEqual("Refrigeration Equipment", contracted["cells"]["A1"]["v"])
        self.assertIn("A1:H1", contracted["merges"])
        self.assertIn("A2:H2", contracted["merges"])
        self.assertEqual(3, contracted["sourceSetup"]["editableStartRow"])
        self.assertIn("equipment", contracted["sourceSetup"]["purpose"].lower())
        self.assertEqual("Source Role", contracted["sourceSetup"]["metadata"][0]["field"])
        self.assertEqual("#F28C28", contracted["styles"]["A1"]["fill"])
        self.assertTrue(contracted["styles"]["A1"]["locked"])
        self.assertEqual("#D9DEE3", contracted["styles"]["G8"]["fill"])
        self.assertTrue(edit_overlaps_protected("A1:C4", contracted["protectedRanges"]))
        self.assertFalse(edit_overlaps_protected("A3:C3", contracted["protectedRanges"]))
        contracted["cells"]["A3"] = {"v": "Field"}
        contracted["cells"]["B3"] = {"v": "Value"}
        contracted["cells"]["C3"] = {"v": "Notes"}
        contracted["cells"]["A4"] = {"v": "Customer Row"}
        repeated = apply_source_sheet_contract(contracted)
        self.assertEqual("Field", repeated["cells"]["A3"]["v"])
        self.assertEqual("Customer Row", repeated["cells"]["A4"]["v"])

    def test_table_detection_preserves_two_exact_regions_and_geometry(self) -> None:
        sheet = source_sheet()
        sheet["cells"].pop("D7")
        regions = detect_table_regions(sheet, start_row=3)
        self.assertEqual(["A4:C5", "A7:B8", "F7:G8"], [item["range"] for item in regions])
        contracted = apply_source_sheet_contract(sheet)
        regions = detect_table_regions(contracted, start_row=3)
        self.assertEqual(["A7:B8", "F7:G8"], [item["range"] for item in regions])
        self.assertEqual(18.125, contracted["columnWidths"]["A"])
        self.assertEqual(22.25, contracted["rowHeights"]["8"])
        self.assertEqual("=B8", contracted["cells"]["G8"]["f"])

    def test_controlled_defaults_preserve_existing_status_rule(self) -> None:
        sheet = {
            **source_sheet(),
            "name": "00_INDEX",
            "cells": {
                "A4": {"v": "Include"},
                "B4": {"v": "Required"},
                "C4": {"v": "Manual Paste Needed"},
                "D4": {"v": "Convert?"},
                "E4": {"v": "Status"},
                "F4": {"v": "Lifecycle"},
            },
            "dataValidations": [{
                "id": "workbook-status",
                "ranges": ["E5:E500"],
                "type": "list",
                "values": ["Workbook Defined"],
            }],
            "conditionalFormats": [],
        }
        controlled = apply_controlled_default_validations(sheet)
        by_range = {
            item["ranges"][0]: item
            for item in controlled["dataValidations"]
        }
        self.assertEqual(["YES", "NO", "VERIFY"], by_range["A5:A500"]["values"])
        self.assertEqual(["YES", "NO", "VERIFY"], by_range["B5:B500"]["values"])
        self.assertEqual(["YES", "NO", "VERIFY"], by_range["C5:C500"]["values"])
        self.assertEqual(["YES", "NO"], by_range["D5:D500"]["values"])
        self.assertEqual(["Workbook Defined"], by_range["E5:E500"]["values"])
        self.assertEqual(
            ["Draft", "Draft Confirmed", "Public", "Public Confirmed"],
            by_range["F5:F500"]["values"],
        )
        self.assertTrue(any(
            item.get("source") == "singh360-publish-state"
            for item in controlled["conditionalFormats"]
        ))

    def test_project_local_save_reload_preserves_geometry_and_layout(self) -> None:
        with tempfile.TemporaryDirectory(prefix="s360-workspace-contract-") as root:
            project = {
                "id": "aaaaaaaaaaaaaaaa",
                "metadata": {"projectName": "Sanitized Source Contract"},
                "pages": [{
                    "id": "p-src",
                    "sheetTab": "SRC R-2.0 Equipment",
                    "sheetCode": "R-2.0",
                    "sheetTitle": "Equipment",
                    "pageType": "data-grid",
                    "include": False,
                    "publishStatus": "NO",
                }],
                "worksheets": [],
            }
            store = WorkbookDocumentStore(Path(root))
            document = {
                "revision": 0,
                "updatedAt": "",
                "sheets": [source_sheet()],
            }
            saved = store.save(project, 0, document)
            reloaded = store.load(project)
            source = reloaded["sheets"][0]
            self.assertEqual(saved["revision"], reloaded["revision"])
            self.assertEqual({"A": 18.125, "B": 13.75, "F": 21.5}, source["columnWidths"])
            self.assertEqual({"8": 22.25}, source["rowHeights"])
            self.assertEqual("side_by_side", source["tableLayout"])
            self.assertEqual("=B8", source["cells"]["G8"]["f"])

    def test_import_keeps_excluded_source_page_and_only_yes_publishes(self) -> None:
        with tempfile.TemporaryDirectory(prefix="s360-import-contract-") as root:
            path = Path(root) / "fixture.xlsx"
            workbook = Workbook()
            metadata = workbook.active
            metadata.title = "00_PROJECT_META"
            metadata.append(["Project Name", "Sanitized Fixture"])
            index = workbook.create_sheet("00_INDEX")
            index.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Page Type"])
            index.append(["VERIFY", 1, "R-2.0", "SRC R-2.0 Equipment", "Equipment", "Source"])
            source = workbook.create_sheet("SRC R-2.0 Equipment")
            source["A3"] = "Fixture"
            workbook.save(path)
            workbook.close()
            project = import_workbook(path, project_id="fixture-contract")
            page = next(item for item in project["pages"] if item["sheetTab"] == "SRC R-2.0 Equipment")
            self.assertFalse(page["include"])
            self.assertEqual("VERIFY", page["publishStatus"])

    def test_disposable_excel_controls_round_trip_without_customer_data(self) -> None:
        with tempfile.TemporaryDirectory(prefix="s360-writeback-contract-") as root:
            path = Path(root) / "disposable.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "00_INDEX"
            for column, value in enumerate(
                ["Include", "Issue Status", "Sheet Tab", "Page Title"],
                1,
            ):
                sheet.cell(4, column, value)
            headers = {
                "include": 1,
                "issue status": 2,
                "sheet tab": 3,
                "page title": 4,
            }
            ensure_index_state_controls(sheet, 4, headers)
            workbook.save(path)
            workbook.close()
            reloaded = load_workbook(path)
            index = reloaded["00_INDEX"]
            formulas = {
                validation.formula1
                for validation in index.data_validations.dataValidation
            }
            self.assertIn('"YES,NO,VERIFY"', formulas)
            self.assertIn('"Draft,Draft Confirmed,Public,Public Confirmed"', formulas)
            rule_count = sum(
                len(rules)
                for rules in index.conditional_formatting._cf_rules.values()
            )
            self.assertGreaterEqual(rule_count, 7)
            reloaded.close()

    def test_new_workspace_metadata_does_not_invalidate_legacy_sync_baseline(self) -> None:
        project = {
            "metadata": {"projectName": "Sanitized Hash Fixture"},
            "pages": [{
                "id": "page-1",
                "order": 1,
                "include": False,
                "sheetCode": "R-2.0",
                "sheetTitle": "Fixture",
                "sheetTab": "SRC Fixture",
                "issueStatus": "draft",
                "canvasObjects": [],
                "assets": [],
                "notes": "",
            }],
            "worksheets": [{
                "id": "sheet-1",
                "name": "SRC Fixture",
                "grid": [["Fixture"]],
                "styles": {},
                "formulas": {},
                "mergedCells": [],
            }],
        }
        baseline = project_hash(project)
        enriched = deepcopy(project)
        enriched["pages"][0]["publishStatus"] = "VERIFY"
        enriched["worksheets"][0].update({
            "protectedRanges": ["A1:H2"],
            "dataValidations": [{"ranges": ["A3:A20"], "type": "list"}],
            "conditionalFormats": [{"ranges": ["A3:A20"], "type": "text"}],
            "tableRegions": [{"id": "table-1", "range": "A3:B9"}],
            "tableLayout": "side_by_side",
            "annotations": [{"id": "note", "text": "Fixture", "placement": "right"}],
        })
        enriched["dataWorkspace"] = {"revision": 2, "signature": "fixture"}
        self.assertEqual(baseline, project_hash(enriched))


class DataWorkspaceStateEndpointTests(unittest.TestCase):
    def setUp(self) -> None:
        self.runtime = isolate_server_runtime(server)
        self.client = server.app.test_client()
        self.project_id = "bbbbbbbbbbbbbbbb"
        self.project = {
            "id": self.project_id,
            "metadata": {"projectName": "Sanitized Dirty State"},
            "worksheets": [{
                "id": "sheet-1",
                "name": "Fixture",
                "grid": [["Header"], ["Value"]],
                "styles": {},
                "mergedCells": [],
            }],
            "pages": [],
            "sources": [],
            "workbookSync": {"mode": "disabled", "status": "disabled", "warning": ""},
        }
        server.store.save(self.project_id, deepcopy(self.project))

    def tearDown(self) -> None:
        self.runtime.cleanup()

    def test_local_data_workspace_save_preserves_standalone_sync_metadata(self) -> None:
        initial = self.client.get(f"/api/projects/{self.project_id}/data-workspace")
        self.assertEqual(200, initial.status_code)
        document = initial.get_json()
        document["sheets"][0]["cells"]["A2"] = {"v": "Changed only in app"}
        response = self.client.put(
            f"/api/projects/{self.project_id}/data-workspace",
            json={"expectedRevision": document["revision"], "document": document},
        )
        self.assertEqual(200, response.status_code, response.get_data(as_text=True))
        saved_project = server.store.load(self.project_id)
        self.assertEqual(
            self.project["workbookSync"],
            saved_project["workbookSync"],
        )
        self.assertEqual(1, saved_project["dataWorkspace"]["revision"])
        reopened_project = self.client.get(
            f"/api/projects/{self.project_id}"
        ).get_json()
        self.assertEqual(1, reopened_project["dataWorkspace"]["revision"])

        stale = self.client.put(
            f"/api/projects/{self.project_id}/data-workspace",
            json={"expectedRevision": 0, "document": document},
        )
        self.assertEqual(409, stale.status_code)
        current = self.client.get(f"/api/projects/{self.project_id}/data-workspace").get_json()
        self.assertEqual("Changed only in app", current["sheets"][0]["cells"]["A2"]["v"])


if __name__ == "__main__":
    unittest.main()
