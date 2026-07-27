"""Generated, sanitized fixtures for Singh360 Draft regression tests."""
from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any


INDEX_HEADERS = [
    "Include",
    "Order",
    "Sheet Code",
    "Sheet Tab",
    "Page Title",
    "Family",
    "Page Type",
    "Notes",
]


def isolate_server_runtime(server: Any) -> TemporaryDirectory[str]:
    """Point an imported server module at a disposable .docs project store."""
    from core.project_store import ProjectStore

    runtime = TemporaryDirectory(prefix="s360_smoke_runtime_")
    docs_dir = Path(runtime.name) / ".docs"
    docs_dir.mkdir(parents=True, exist_ok=True)
    server.DOCS_DIR = docs_dir
    server.store = ProjectStore(docs_dir)
    return runtime


def _add_control_sheets(workbook: Any, project_name: str) -> Any:
    metadata = workbook.active
    metadata.title = "00_PROJECT_META"
    metadata.append(["Project Name", project_name])
    metadata.append(["Workbook Schema Version", "2"])
    metadata.append(["Help Version", "test"])

    index = workbook.create_sheet("00_INDEX")
    index["A1"] = "SINGH360 DRAFT — SANITIZED REGRESSION FIXTURE"
    for column, value in enumerate(INDEX_HEADERS, start=1):
        index.cell(4, column, value)
    return index


def _append_index(index: Any, values: list[Any]) -> None:
    index.append(values)


def write_workbook(path: Path) -> Path:
    from openpyxl import Workbook

    workbook = Workbook()
    index = _add_control_sheets(workbook, "Sanitized Regression Project")
    _append_index(index, ["YES", 1, "EMS 1.0", "Assets", "Asset Schedule", "Equipment", "table", ""])
    _append_index(index, ["YES", 2, "EMS 2.0", "Network", "Network Schedule", "Network", "table", ""])

    assets = workbook.create_sheet("Assets")
    assets.append(["Tag", "Description", "Quantity"])
    assets.append(["CTRL-01", "Generic Controller", 1])
    assets.append(["SENS-01", "Generic Temperature Sensor", 2])

    network = workbook.create_sheet("Network")
    network.append(["Device", "Protocol", "Address"])
    network.append(["CTRL-01", "BACnet/IP", "10.0.0.10"])
    workbook.save(path)
    return path


def write_sa31_regression_workbook(path: Path) -> Path:
    """Generate the sheet-code, index, and continuation shape used by SA31 tests."""
    from openpyxl import Workbook

    workbook = Workbook()
    index = _add_control_sheets(workbook, "Sanitized Multi-Sheet Regression")
    rows = [
        ["YES", 1, "EMS 0.0", "Cover", "Cover / Project Info", "Front Matter", "cover", ""],
        ["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""],
        ["YES", 5, "EMS 0.4", "Scope", "Project Scope", "Front Matter", "text", ""],
        ["YES", 12, "EMS 1.1", "IDF Network", "IDF Network Table", "Network", "table", ""],
        ["YES", 15, "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "Lighting", "io-table", ""],
    ]
    for row in rows:
        _append_index(index, row)

    cover = workbook.create_sheet("Cover")
    cover["A1"] = "SANITIZED EMS PROJECT"

    scope = workbook.create_sheet("Scope")
    scope.append(["Section", "Scope Language", "Status", "Notes"])
    scope.append(["Executive Summary", "Generated regression scope.", "Review", ""])

    network = workbook.create_sheet("IDF Network")
    network.append([
        "Port",
        "Label",
        "Device / Drop",
        "From",
        "To",
        "Cable",
        "Notes",
        "Controller ID",
        "IP Address",
    ])
    for port in range(1, 49):
        network.append([
            str(port),
            f"L{port}",
            f"Generic Device {port}",
            "IDF",
            f"Drop {port}",
            "CAT6",
            "",
            "CTRL-01",
            f"10.10.0.{port}",
        ])

    panel = workbook.create_sheet("LCP Panel Schedule")
    panel.append(["RO#", "Description", "Type", "DI#", "Status", "Notes"])
    for row_number in range(1, 61):
        panel.append([
            str(row_number),
            f"Generated Relay {row_number}",
            "NO",
            str(row_number),
            "Review",
            "",
        ])
        panel.row_dimensions[row_number + 1].height = 30

    workbook.save(path)
    return path


def write_829_regression_workbook(path: Path) -> Path:
    """Generate the 7-column IDF switch-matrix shape used by 829 regressions."""
    from openpyxl import Workbook

    workbook = Workbook()
    index = _add_control_sheets(workbook, "Sanitized Switch-Matrix Regression")
    rows = [
        ["YES", 1, "EMS 0.0", "Cover", "Cover / Project Info", "Front Matter", "cover", ""],
        ["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""],
        ["YES", 3, "EMS 13.2", "EMS 13.2 IDF #2", "IDF #2 Port / Network Table", "Network", "table", ""],
        ["NO", 4, "EMS 13.2a", "EMS 13.2a IDF #2 Layout", "Reserved IDF Layout", "Network", "drawing", ""],
    ]
    for row in rows:
        _append_index(index, row)

    cover = workbook.create_sheet("Cover")
    cover["A1"] = "SANITIZED SWITCH-MATRIX PROJECT"

    matrix = workbook.create_sheet("EMS 13.2 IDF #2")
    matrix.append(["IDF #2 TABLE (SWITCH 3 & 4)", "", "Controller ID Legend", "IP Network"])
    matrix.append(["MODEL", "GENERATED", "001-050: Generic controller", "Location"])
    matrix.append(["", "", "", ""])
    matrix.append([
        "Label #",
        "Description",
        "Controller ID",
        "IP Address",
        "IDF#",
        "Switch#",
        "Port#",
    ])
    for switch_number in (3, 4, 5):
        for port in range(1, 49):
            matrix.append([
                f"L{switch_number}{port:02d}",
                f"Generated switch {switch_number} device {port}",
                200 + switch_number * 50 + port,
                "-",
                2,
                switch_number,
                port,
            ])

    layout = workbook.create_sheet("EMS 13.2a IDF #2 Layout")
    layout["A1"] = "RESERVED GENERATED LAYOUT CODE"

    workbook.save(path)
    return path


def write_geometry_browser_workbook(path: Path) -> Path:
    """Generate the browser/PDF geometry fixture without customer content."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    index = _add_control_sheets(workbook, "Disposable Geometry Browser Smoke")
    page_codes = (
        "EMS 3.0",
        "EMS 4.0",
        "EMS 7.0",
        "EMS 8.0",
        "EMS 13.0",
        "EMS 23.0",
        "EMS 24.0",
        "EMS 24.1",
        "EMS 24.2",
    )
    orange = PatternFill(fill_type="solid", fgColor="F4B183")
    header = PatternFill(fill_type="solid", fgColor="FCE4D6")
    widths = (14.0, 34.0, 18.0, 24.0, 48.0, 20.0)

    for order, code in enumerate(page_codes, start=1):
        tab = code.replace("EMS ", "Geometry ")
        _append_index(index, ["YES", order, code, tab, f"{code} Geometry Proof", "Generated", "table", ""])
        sheet = workbook.create_sheet(tab)
        sheet.sheet_format.defaultColWidth = 11.5
        sheet.sheet_format.defaultRowHeight = 18.0
        for column, width in zip("ABCDEF", widths):
            sheet.column_dimensions[column].width = width

        sheet.merge_cells("A1:F1")
        sheet["A1"] = f"{code} DISPOSABLE GEOMETRY INSTRUCTION BAND"
        sheet["A1"].fill = orange
        sheet["A1"].font = Font(bold=True, size=12)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 31.5

        headings = ("Tag", "Description", "Type", "Signal", "Instructions", "Status")
        for column, value in enumerate(headings, start=1):
            cell = sheet.cell(2, column, value)
            cell.fill = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[2].height = 24.0

        row_count = 78 if code == "EMS 24.2" else 14
        for row in range(3, row_count + 3):
            values = (
                f"T-{row - 2:03d}",
                f"Generated equipment description {row - 2}",
                "PR0650CD-TDB",
                "0-10VDC",
                "Normal words remain readable and wrap only at spaces across the preserved workbook width map.",
                "Review",
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=column in (2, 5))
            sheet.row_dimensions[row].height = 27.75 if row == 3 else 22.5

        # Exercise hidden geometry without hiding the visible proof range.
        sheet.row_dimensions[row_count + 4].hidden = True
        sheet.column_dimensions["H"].hidden = True
        sheet["A3"] = "=1+1"

    workbook.save(path)
    return path


def write_pdf(path: Path) -> Path:
    import fitz

    document = fitz.open()
    page = document.new_page(width=11 * 72, height=8.5 * 72)
    page.insert_text((72, 72), "Sanitized Singh360 Draft PDF fixture")
    page.draw_rect(fitz.Rect(72, 100, 576, 360), color=(0, 0, 0))
    document.save(path)
    document.close()
    return path
