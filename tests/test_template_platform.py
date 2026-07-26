from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from core.project_data_compiler import apply_compile, preview_compile
from core.project_store import ProjectStore
from core.project_template_service import ProjectTemplateService
from core.template_platform import (
    ProfileRegistry, RevisionConflict, SourceLibrary, TemplatePlatformError,
    TemplateRegistry, WorkbookDocumentStore, sha256_file,
)

ROOT = Path(__file__).resolve().parents[1]


def base_workbook(path: Path) -> None:
    profile = ProfileRegistry(ROOT / "defaults/project_templates/project_profiles.json").get("BASE_CORE")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in profile["dataSheets"]:
        sheet = workbook.create_sheet(name)
        sheet.append([name])
        sheet.append(["Field", "Value"])
    workbook.save(path)


class TemplatePlatformTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.docs = self.root / ".docs"
        self.docs.mkdir()
        self.profiles = ProfileRegistry(ROOT / "defaults/project_templates/project_profiles.json")
        self.templates = TemplateRegistry(self.docs)
        self.store = ProjectStore(self.docs)
        self.staged = self.docs / "template_staging" / "base.xlsx"
        self.staged.parent.mkdir()
        base_workbook(self.staged)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_profile_inheritance_and_registry(self) -> None:
        profile = self.profiles.get("EMS_FULL")
        self.assertIn("00_PROJECT_META", profile["dataSheets"])
        self.assertIn("11_NETWORK_PORTS", profile["dataSheets"])
        self.assertEqual(len(self.profiles.list()), 7)
        self.assertEqual(self.profiles.get("EMS_LIGHTING")["extends"], "BASE_CORE")

    def test_template_validation_registration_and_copy(self) -> None:
        validation = self.templates.validate(self.staged)
        self.assertTrue(validation["valid"], validation)
        record = self.templates.register(self.staged, [item["id"] for item in self.profiles.list()])
        runtime = Path(record["absoluteRuntimePath"])
        self.assertTrue(runtime.is_file())
        self.assertNotEqual(runtime, self.staged)
        self.assertEqual(sha256_file(runtime), record["sha256"])

    def test_create_project_profile_and_physical_sheets(self) -> None:
        record = self.templates.register(self.staged, [item["id"] for item in self.profiles.list()])
        service = ProjectTemplateService(self.store, self.profiles, self.templates)
        project = service.create({"profileId": "EMS_FULL", "templateId": record["templateId"], "metadata": {"projectName": "Test Project", "client": "Example"}})
        self.assertEqual(project["schemaVersion"], 2)
        self.assertRegex(project["id"], r"^[a-f0-9]{16}$")
        folder = self.store.find_dir(project["id"])
        self.assertIsNotNone(folder)
        workbook = load_workbook(folder / "sources/workbook/base.xlsx")
        self.assertIn("11_NETWORK_PORTS", workbook.sheetnames)
        self.assertIn("19_BILL_OF_MATERIALS", workbook.sheetnames)
        self.assertEqual(workbook["00_PROJECT_META"]["B3"].value, project["id"])
        workbook.close()

    def test_create_transaction_rolls_back_only_new_folder(self) -> None:
        record = self.templates.register(self.staged, [item["id"] for item in self.profiles.list()])
        service = ProjectTemplateService(self.store, self.profiles, self.templates)
        with patch.object(service, "_apply_profile", side_effect=RuntimeError("fixture failure")):
            with self.assertRaises(RuntimeError):
                service.create({"profileId": "EMS_FULL", "templateId": record["templateId"], "metadata": {"projectName": "Rollback Test"}})
        self.assertEqual([path for path in self.store.projects_dir.iterdir() if path.is_dir()], [])
        self.assertTrue(list((self.docs / "failure_logs").glob("project_create_*.json")))

    def test_source_manifest_checksum_version_archive_and_traversal(self) -> None:
        project = self.store.projects_dir / "Sources__1234567890abcdef"
        self.store.ensure_folders(project)
        library = SourceLibrary(project)
        first = library.upload(io.BytesIO(b"a,b\n1,2\n"), "fixture.csv")
        second = library.upload(io.BytesIO(b"a,b\n3,4\n"), "fixture.csv", {"supersedes": first["id"]})
        self.assertEqual(second["version"], 2)
        self.assertEqual(library.load()["sources"][0]["status"], "superseded")
        self.assertEqual(library.archive(second["id"])["status"], "archived")
        with self.assertRaises(TemplatePlatformError):
            library.upload(io.BytesIO(b"MZ"), "../../unsafe.exe")
        with self.assertRaises(TemplatePlatformError):
            library.resolve("../../etc/passwd")

    def test_workbook_revision_conflict_and_history(self) -> None:
        project = self.store.projects_dir / "Workbook__1234567890abcdef"
        document = {"revision": 1, "updatedAt": "", "sheets": []}
        storage = WorkbookDocumentStore(project)
        storage.create(document)
        saved = storage.save(1, document)
        self.assertEqual(saved["revision"], 2)
        with self.assertRaises(RevisionConflict):
            storage.save(1, document)
        self.assertEqual(len(list(storage.history.glob("workbook_*.json"))), 1)

    def test_compile_stable_ids_and_manual_preservation(self) -> None:
        profile = self.profiles.get("EMS_FULL")
        index_cells = {}
        headers = ["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Page ID", "Source Mode"]
        for column, value in enumerate(headers, start=1):
            index_cells[f"{chr(64 + column)}1"] = {"v": value}
        rows = [
            ["YES", 1, "EMS 1.0", "00_PROJECT_META", "Cover", "Front Matter", "Cover", "cover", "generated"],
            ["YES", 2, "EMS 2.0", "00_INDEX", "Sheet Index / TOC", "Front Matter", "Sheet Index", "index", "generated"],
            ["YES", 3, "EMS 13.0", "11_NETWORK_PORTS", "RDM / IDF Network Table", "Network / Data", "Network Table", "network", "canonical"],
        ]
        for row_number, row in enumerate(rows, start=2):
            for column, value in enumerate(row, start=1):
                index_cells[f"{chr(64 + column)}{row_number}"] = {"v": value}
        workbook = {"revision": 1, "sheets": [
            {"id": "index", "name": "00_INDEX", "cells": index_cells, "styles": {}, "merges": [], "rowHeights": {}, "columnWidths": {}},
            {"id": "network", "name": "11_NETWORK_PORTS", "cells": {"A1": {"v": "NETWORK"}, "A2": {"v": "Port"}, "A3": {"v": "1"}}, "styles": {}, "merges": [], "rowHeights": {}, "columnWidths": {}},
        ]}
        project = {"id": "1234567890abcdef", "pages": [], "projectProfileId": "EMS_FULL"}
        first, preview = apply_compile(project, workbook, profile)
        self.assertTrue(any(item["family"] == "RDM / IDF Network Table" for item in preview["operations"]))
        network = next(page for page in first["pages"] if page["pageFamily"] == "RDM / IDF Network Table")
        network["canvasObjects"] = [{"id": "manual-note", "type": "textbox", "text": "Keep me"}]
        network["assets"] = [{"id": "manual-image"}]
        network["underlay"] = {"source": "manual.pdf"}
        workbook["sheets"][1]["cells"]["A3"] = {"v": "2"}
        second, _ = apply_compile(first, workbook, profile)
        updated = next(page for page in second["pages"] if page["id"] == network["id"])
        self.assertEqual(updated["canvasObjects"], network["canvasObjects"])
        self.assertEqual(updated["assets"], network["assets"])
        self.assertEqual(updated["underlay"], network["underlay"])
        self.assertEqual(updated["blocks"][0]["rows"][0][0], "2")
        self.assertEqual(network["id"], updated["id"])
        self.assertEqual(first["pages"][0]["pageFamily"], "Cover")
        self.assertEqual(first["pages"][1]["pageFamily"], "Sheet Index / TOC")
        self.assertEqual(preview_compile(first, workbook, profile)["projectId"], project["id"])


if __name__ == "__main__":
    unittest.main()
