from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path
import re
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from core.full_workbook_sync import verify_synchronized_workbook
from core.project_workspace import (
    ProjectWorkspaceError,
    WorkbookDocumentStore,
    drawing_workspace_sequence,
    project_base_drawing_manifest,
    reconcile_project_workbook_order,
    reconcile_workbook_document_order,
)
from core.sheet_index_sync import sync_project_sheet_index
from core.workbook_link_manager import _s360_workbook_projection_hash
from core.workbook_status_sync import (
    WorkbookSyncError,
    _s360_index_manifest,
    sync_project_to_workbook,
)


HEADERS = [
    "Include",
    "Order",
    "Sheet Code",
    "Sheet Tab",
    "Page Title",
    "Family",
    "Page Type",
    "Notes",
    "Render Profile",
    "Split Mode",
    "Page ID",
    "Parent Page ID",
    "Issue Status",
    "Source Mode",
    "Sync Direction",
]


class DisposableStore:
    def __init__(self, root: Path):
        self.docs = root / "runtime"
        self.docs.mkdir()
        self.saved: dict[str, dict] = {}

    def dir_for(self, project_id: str, project: dict) -> Path:
        path = self.docs / "projects" / project_id
        path.mkdir(parents=True, exist_ok=True)
        return path

    def sources_dir(self, project_id: str, kind: str, project: dict) -> Path:
        path = self.dir_for(project_id, project) / "sources" / kind
        path.mkdir(parents=True, exist_ok=True)
        return path

    def assets_excel_dir(self, project_id: str, project: dict) -> Path:
        path = self.dir_for(project_id, project) / "assets" / "excel"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def save(self, project_id: str, project: dict) -> None:
        self.saved[project_id] = deepcopy(project)
        (self.dir_for(project_id, project) / "project.json").write_text(
            json.dumps(project, indent=2),
            encoding="utf-8",
        )

    def load(self, project_id: str) -> dict | None:
        value = self.saved.get(project_id)
        return deepcopy(value) if value is not None else None


def page(
    page_id: str,
    order: int,
    code: str,
    tab: str,
    title: str,
    worksheet_id: str,
    page_type: str = "data-grid",
) -> dict:
    return {
        "id": page_id,
        "order": order,
        "include": True,
        "publishStatus": "YES",
        "sheetCode": code,
        "displaySheetCode": code,
        "sheetTab": tab,
        "sheetTitle": title,
        "pageType": page_type,
        "linkedWorksheetId": worksheet_id,
        "canvasObjects": [{"type": "text", "text": f"manual-{page_id}"}],
        "notes": "",
    }


def worksheet(sheet_id: str, name: str, grid: list[list[object]]) -> dict:
    return {
        "id": sheet_id,
        "name": name,
        "grid": grid,
        "styles": {},
        "formulas": {},
        "mergedCells": [],
        "rowHeights": {},
        "columnWidths": {},
        "visible": True,
    }


def fixture_project(path: Path) -> dict:
    index_grid = [HEADERS]
    stale = [
        ("page-cover", "EMS 0.0", "Cover", "Cover / Project Info"),
        ("page-index", "EMS 0.1", "00_INDEX", "Sheet Index"),
        ("page-alpha", "EMS 1.0", "ALPHA", "Alpha"),
        ("page-beta", "EMS 2.0", "BETA", "Beta"),
        ("page-gamma", "EMS 3.0", "GAMMA", "Gamma"),
    ]
    for order, (page_id, code, tab, title) in enumerate(stale, start=1):
        index_grid.append(
            ["YES", order, code, tab, title, "", "data-grid", "", "", "",
             page_id, "", "Draft", "Workbook", "Both"]
        )
    pages = [
        page("page-cover", 1, "EMS 0.0", "Cover", "Cover / Project Info", "ws-cover", "cover"),
        page("page-index", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "ws-index", "index"),
        page("page-gamma", 3, "EMS 3.0", "GAMMA", "Gamma", "ws-gamma"),
        page("page-alpha", 4, "EMS 1.0", "ALPHA", "Alpha", "ws-alpha"),
        page("page-beta", 5, "EMS 2.0", "BETA", "Beta", "ws-beta"),
        {
            **page(
                "page-beta__continuation_1",
                6,
                "EMS 2.0a",
                "BETA",
                "Beta — Continued",
                "ws-beta",
            ),
            "generatedContinuation": True,
            "continuationOf": "page-beta",
            "continuationIndex": 1,
        },
    ]
    project = {
        "id": "sync-order-fixture",
        "metadata": {"projectName": "Sanitized Order Fixture"},
        "pages": pages,
        "worksheets": [
            worksheet("ws-index", "00_INDEX", index_grid),
            worksheet("ws-alpha", "ALPHA", [["alpha"]]),
            worksheet("ws-beta", "BETA", [["beta"]]),
            worksheet("ws-cover", "Cover", [["cover"]]),
            worksheet("ws-gamma", "GAMMA", [["gamma"]]),
            worksheet("ws-source", "SRC NOTES", [["preserve source"]]),
        ],
        "sources": [],
        "workbookSync": {
            "workbook": str(path),
            "status": "app_changed",
        },
    }
    return sync_project_sheet_index(project)


def write_fixture_workbook(path: Path) -> None:
    workbook = Workbook()
    metadata = workbook.active
    metadata.title = "00_PROJECT_META"
    metadata["A1"] = "Project Name"
    metadata["B1"] = "Sanitized Order Fixture"
    index = workbook.create_sheet("00_INDEX")
    for column, label in enumerate(HEADERS, start=1):
        index.cell(1, column, label)
    rows = [
        ("page-cover", "EMS 0.0", "Cover", "Cover / Project Info"),
        ("page-index", "EMS 0.1", "00_INDEX", "Sheet Index"),
        ("page-alpha", "EMS 1.0", "ALPHA", "Alpha"),
        ("page-beta", "EMS 2.0", "BETA", "Beta"),
        ("page-gamma", "EMS 3.0", "GAMMA", "Gamma"),
    ]
    for row_number, (page_id, code, tab, title) in enumerate(rows, start=2):
        values = [
            "YES",
            row_number - 1,
            code,
            tab,
            title,
            "",
            "data-grid",
            "",
            "",
            "",
            page_id,
            "",
            "Draft",
            "Workbook",
            "Both",
        ]
        for column, value in enumerate(values, start=1):
            index.cell(row_number, column, value)
    workbook.create_sheet("00_HELP")
    for name in ("ALPHA", "BETA", "Cover", "GAMMA", "SRC NOTES"):
        sheet = workbook.create_sheet(name)
        sheet["A1"] = f"payload-{name}"
    workbook.save(path)
    workbook.close()


def document_for(project: dict) -> dict:
    sheets = []
    for source in project["worksheets"]:
        cells = {}
        for row_number, row in enumerate(source["grid"], start=1):
            for column_number, value in enumerate(row, start=1):
                if value not in (None, ""):
                    letter = chr(64 + column_number)
                    cells[f"{letter}{row_number}"] = {"v": value}
        sheets.append(
            {
                "id": source["id"],
                "name": source["name"],
                "cells": cells,
                "styles": {"A1": {"bold": True, "fill": "#DDEBF7"}},
                "merges": ["A7:B7"] if source["name"] == "ALPHA" else [],
                "rowHeights": {"1": 22.25},
                "columnWidths": {"A": 18.125},
                "defaultColumnWidth": 8.43,
                "defaultRowHeight": 15.0,
                "hiddenRows": [],
                "hiddenColumns": [],
                "archived": False,
                "tabColor": None,
                "role": "source" if source["name"].startswith("SRC") else None,
                "sourceSetup": {},
                "protectedRanges": [],
                "dataValidations": [],
                "conditionalFormats": [],
                "tableRegions": (
                    [{"id": "table-alpha", "range": "A1:B4", "label": "Alpha table"}]
                    if source["name"] == "ALPHA"
                    else []
                ),
                "tableLayout": "side_by_side" if source["name"] == "ALPHA" else "single",
                "annotations": (
                    [{"id": "note-alpha", "placement": "right", "text": "Keep annotation"}]
                    if source["name"] == "ALPHA"
                    else []
                ),
            }
        )
    return {"revision": 0, "updatedAt": "", "sheets": sheets}


class SyncOrderReconciliationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(prefix="s360-sync-order-")
        self.root = Path(self.temp.name)
        self.path = self.root / "disposable.xlsx"
        write_fixture_workbook(self.path)
        self.store = DisposableStore(self.root)
        self.project = fixture_project(self.path)
        self.document_store = WorkbookDocumentStore(
            self.store.dir_for(self.project["id"], self.project)
        )
        self.document_store.save(
            self.project,
            0,
            document_for(self.project),
        )

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_reorder_is_identical_everywhere_and_preserves_workspace_payload(self) -> None:
        synced = sync_project_to_workbook(
            self.project["id"],
            deepcopy(self.project),
            self.store,
        )
        self.assertEqual("in_sync", synced["workbookSync"]["status"])
        self.assertTrue(synced["workbookSync"]["verified"])
        self.assertEqual(
            "verified",
            synced["workbookSync"]["verification"]["status"],
        )

        expected = project_base_drawing_manifest(synced)
        expected_ids = [item["pageId"] for item in expected]
        expected_tabs = [item["sheetTab"] for item in expected]
        self.assertEqual(
            ["page-cover", "page-index", "page-gamma", "page-alpha", "page-beta"],
            expected_ids,
        )

        index = _s360_index_manifest(self.path)
        self.assertEqual(expected_ids, [item["id"] for item in index])
        self.assertNotIn("page-beta__continuation_1", [item["id"] for item in index])

        workbook = load_workbook(self.path, read_only=True)
        try:
            drawing_keys = {tab.casefold() for tab in expected_tabs}
            physical = [
                sheet.title
                for sheet in workbook.worksheets
                if sheet.title.casefold() in drawing_keys
            ]
            self.assertEqual(expected_tabs, physical)
            self.assertIn("SRC NOTES", workbook.sheetnames)
        finally:
            workbook.close()

        reloaded = self.document_store.load(synced)
        workspace = drawing_workspace_sequence(reloaded)
        self.assertEqual(expected_ids, [item["pageId"] for item in workspace])
        self.assertEqual(expected_tabs, [item["sheetTab"] for item in workspace])
        alpha = next(sheet for sheet in reloaded["sheets"] if sheet["name"] == "ALPHA")
        self.assertEqual({"bold": True, "fill": "#DDEBF7"}, alpha["styles"]["A1"])
        self.assertEqual(["A7:B7"], alpha["merges"])
        self.assertEqual({"1": 22.25}, alpha["rowHeights"])
        self.assertEqual({"A": 18.125}, alpha["columnWidths"])
        self.assertEqual("Alpha table", alpha["tableRegions"][0]["label"])
        self.assertEqual("Keep annotation", alpha["annotations"][0]["text"])

        worksheet_names = [item["name"] for item in synced["worksheets"]]
        self.assertEqual(expected_tabs, worksheet_names[: len(expected_tabs)])
        index_worksheet = next(
            item for item in synced["worksheets"] if item["name"] == "00_INDEX"
        )
        self.assertEqual(
            expected_tabs,
            [str(row[3]) for row in index_worksheet["grid"][1:6]],
        )
        continuation = next(
            page
            for page in synced["pages"]
            if page["id"] == "page-beta__continuation_1"
        )
        self.assertEqual(
            [{"type": "text", "text": "manual-page-beta__continuation_1"}],
            continuation["canvasObjects"],
        )

    def test_blank_page_creates_matching_project_worksheet_and_index_row(self) -> None:
        candidate = deepcopy(self.project)
        candidate["pages"].append(
            {
                "id": "page-blank",
                "order": 6,
                "include": True,
                "sheetCode": "",
                "displaySheetCode": "",
                "sheetTab": "",
                "sheetTitle": "New Sheet",
                "pageType": "canvas",
                "blocks": [],
                "canvasObjects": [],
            }
        )

        reconciled, manifest = reconcile_project_workbook_order(candidate)
        blank = next(page for page in reconciled["pages"] if page["id"] == "page-blank")
        self.assertEqual("", blank["sheetCode"])
        self.assertEqual("", blank["displaySheetCode"])
        self.assertEqual("New Sheet", blank["sheetTab"])
        self.assertTrue(blank["linkedWorksheetId"].startswith("worksheet_"))

        matching = next(
            item
            for item in reconciled["worksheets"]
            if item["id"] == blank["linkedWorksheetId"]
        )
        self.assertEqual("New Sheet", matching["name"])
        self.assertEqual("", matching["sourceSetup"]["sheetCode"])
        self.assertEqual("New Sheet", matching["sourceSetup"]["title"])

        index = next(
            item for item in reconciled["worksheets"] if item["name"] == "00_INDEX"
        )
        header = {str(value): column for column, value in enumerate(index["grid"][0])}
        row = next(
            row
            for row in index["grid"][1:]
            if row[header["Page ID"]] == "page-blank"
        )
        self.assertEqual("", row[header["Sheet Code"]])
        self.assertEqual("New Sheet", row[header["Sheet Tab"]])
        self.assertEqual("New Sheet", row[header["Page Title"]])
        self.assertEqual(
            ("page-blank", "", "New Sheet"),
            (
                manifest[-1]["pageId"],
                manifest[-1]["sheetCode"],
                manifest[-1]["sheetTab"],
            ),
        )

    def test_blank_page_worksheet_tab_is_valid_and_unique(self) -> None:
        candidate = deepcopy(self.project)
        candidate["worksheets"].append(worksheet("source-new", "New Sheet", [["source"]]))
        candidate["pages"].append(
            {
                "id": "page-blank",
                "order": 6,
                "include": True,
                "sheetCode": "",
                "displaySheetCode": "",
                "sheetTab": "",
                "sheetTitle": "New/Sheet:*?[] With A Very Long Worksheet Title",
                "pageType": "canvas",
                "blocks": [],
                "canvasObjects": [],
            }
        )

        reconciled, _ = reconcile_project_workbook_order(candidate)
        blank = next(page for page in reconciled["pages"] if page["id"] == "page-blank")
        self.assertLessEqual(len(blank["sheetTab"]), 31)
        self.assertNotRegex(blank["sheetTab"], r"[\[\]:*?/\\]")
        names = [item["name"].casefold() for item in reconciled["worksheets"]]
        self.assertEqual(len(names), len(set(names)))

    def test_injected_mismatch_reports_the_first_exact_difference(self) -> None:
        synced = sync_project_to_workbook(
            self.project["id"],
            deepcopy(self.project),
            self.store,
        )
        mismatch = deepcopy(synced)
        page_gamma = next(
            page for page in mismatch["pages"] if page["id"] == "page-gamma"
        )
        page_gamma["sheetTitle"] = "Injected mismatch"
        with self.assertRaisesRegex(
            RuntimeError,
            r"00_INDEX item 3 Page Title mismatch: expected 'Injected mismatch', found 'Gamma'\.",
        ):
            verify_synchronized_workbook(
                self.path,
                self.project["id"],
                mismatch,
                self.store,
            )

    def test_sync_status_is_not_set_when_verification_fails(self) -> None:
        candidate = deepcopy(self.project)
        with patch(
            "core.full_workbook_sync.verify_synchronized_workbook",
            side_effect=RuntimeError(
                "00_INDEX item 1 Order mismatch: expected 1, found 2."
            ),
        ):
            with self.assertRaisesRegex(
                WorkbookSyncError,
                r"00_INDEX item 1 Order mismatch",
            ):
                sync_project_to_workbook(
                    candidate["id"],
                    candidate,
                    self.store,
                )
        saved = self.store.load(candidate["id"])
        self.assertFalse(
            saved
            and saved.get("workbookSync", {}).get("status") == "in_sync"
        )

    def test_ambiguous_sheet_code_is_rejected_instead_of_silently_matched(self) -> None:
        project = {
            "pages": [
                {
                    "id": "page-one",
                    "order": 1,
                    "include": True,
                    "sheetCode": "EMS 1.0",
                    "sheetTab": "MISSING",
                    "sheetTitle": "Missing",
                }
            ]
        }
        duplicate = {
            "cells": {},
            "styles": {},
            "merges": [],
            "rowHeights": {},
            "columnWidths": {},
            "defaultColumnWidth": 8.43,
            "defaultRowHeight": 15,
            "hiddenRows": [],
            "hiddenColumns": [],
            "archived": False,
            "tabColor": None,
            "role": None,
            "sourceSetup": {"sheetCode": "EMS 1.0"},
            "protectedRanges": [],
            "dataValidations": [],
            "conditionalFormats": [],
            "tableRegions": [],
            "tableLayout": "single",
            "annotations": [],
        }
        document = {
            "revision": 0,
            "updatedAt": "",
            "sheets": [
                {**deepcopy(duplicate), "id": "one", "name": "ONE"},
                {**deepcopy(duplicate), "id": "two", "name": "TWO"},
            ],
        }
        with self.assertRaisesRegex(
            ProjectWorkspaceError,
            "Ambiguous worksheet sheet code EMS 1.0",
        ):
            reconcile_workbook_document_order(document, project)

    def test_page_reorder_changes_the_workbook_projection(self) -> None:
        before = deepcopy(self.project)
        after = deepcopy(before)
        base = [
            page
            for page in after["pages"]
            if not page.get("generatedContinuation")
        ]
        base[2]["order"], base[3]["order"] = (
            base[3]["order"],
            base[2]["order"],
        )
        self.assertNotEqual(
            _s360_workbook_projection_hash(before),
            _s360_workbook_projection_hash(after),
        )

    def test_frontend_guard_validation_and_tooltip_contract_is_wired(self) -> None:
        root = Path(__file__).resolve().parents[1]
        workspace = (
            root / "frontend/src/workspace/DataWorkspace.tsx"
        ).read_text(encoding="utf-8")
        provider = (
            root / "frontend/src/components/help/TooltipProvider.tsx"
        ).read_text(encoding="utf-8")
        tooltip = (
            root / "frontend/src/components/help/AppTooltip.tsx"
        ).read_text(encoding="utf-8")
        styles = (
            root / "frontend/src/styles/tooltips.css"
        ).read_text(encoding="utf-8")

        host = re.search(
            r'<div\s+ref=\{containerRef\}.*?/>',
            workspace,
            flags=re.DOTALL,
        )
        self.assertIsNotNone(host)
        self.assertNotIn("data-help-id", host.group(0))
        self.assertNotIn("data-tooltip-body", host.group(0))
        for required in (
            "baselineValidationErrorsRef",
            "strictValidationDetail",
            "newly introduced strict-dropdown values",
            "await restoreDocument(confirmed",
            "worksheet.getTabColor()",
            "event.type !== CommandType.COMMAND",
            "TAB_COLOR_COMMAND_IDS.has(event.id)",
            'data-testid="data-workspace-shell"',
            'data-testid="drawing-pages-strip"',
            'data-testid="drawing-page-tab"',
            'data-help-id="workspace.save"',
            'data-help-id="workspace.discard"',
            'data-help-id="dialog.cancel"',
            "setNavigation(null)",
        ):
            self.assertIn(required, workspace)
        self.assertIn("HOVER_DELAY_MS = 650", provider)
        self.assertGreaterEqual(provider.count("modalIsOpen()"), 3)
        self.assertIn('[role="dialog"][aria-modal="true"]', tooltip)
        self.assertIn("max-width: min(300px", styles)


if __name__ == "__main__":
    unittest.main()
