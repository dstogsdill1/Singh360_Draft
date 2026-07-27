from __future__ import annotations

from io import BytesIO
from hashlib import sha256
import json
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

import server
from core.workbook_link_manager import status_payload, workbook_metadata
from tests.generated_fixtures import isolate_server_runtime, write_workbook


class ProjectImportTransactionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.runtime = isolate_server_runtime(server)
        self.client = server.app.test_client()
        self.source_root = tempfile.TemporaryDirectory(prefix="s360_import_source_")
        self.source = write_workbook(
            Path(self.source_root.name) / "Sanitized_EMS_Workbook.xlsx"
        )
        workbook = load_workbook(self.source)
        metadata = workbook["00_PROJECT_META"]
        metadata.append(["Linked Project ID", "aaaaaaaaaaaaaaaa"])
        workbook.save(self.source)
        workbook.close()
        self.source_bytes = self.source.read_bytes()

    def tearDown(self) -> None:
        self.source_root.cleanup()
        self.runtime.cleanup()

    def _create(self):
        return self.client.post(
            "/api/projects/new",
            data={
                "profile": "ems",
                "file": (BytesIO(self.source_bytes), self.source.name),
            },
            content_type="multipart/form-data",
        )

    def test_creation_is_atomic_and_establishes_internal_authority(self) -> None:
        response = self._create()
        self.assertEqual(200, response.status_code, response.get_data(as_text=True))
        project_id = response.get_json()["id"]

        project_dirs = list(server.store.projects_dir.glob("*__*"))
        self.assertEqual(1, len(project_dirs))
        project_dir = project_dirs[0]
        project = json.loads((project_dir / "project.json").read_text("utf-8"))
        workbook_path = (
            project_dir / "sources" / "workbook" / self.source.name
        )

        self.assertEqual(project_id, project["id"])
        self.assertEqual("ems", project["projectProfile"])
        self.assertGreater(len(project["pages"]), 0)
        self.assertTrue(workbook_path.is_file())
        self.assertEqual(self.source.name, project["metadata"]["sourceFile"])
        self.assertEqual(self.source.name, project["sourceWorkbookName"])
        self.assertNotIn("temp_", json.dumps(project))
        self.assertEqual(str(workbook_path), project["sources"][0]["path"])
        self.assertEqual(str(workbook_path), project["workbookSync"]["workbook"])
        self.assertEqual("in_sync", project["workbookSync"]["status"])
        self.assertEqual(project_id, workbook_metadata(workbook_path)["projectId"])
        self.assertEqual(
            "in_sync",
            status_payload(project_id, project, server.store)["status"],
        )
        opened = self.client.get(f"/api/projects/{project_id}")
        self.assertEqual(200, opened.status_code)
        self.assertEqual(self.source.name, opened.get_json()["sourceWorkbookName"])
        self.assertEqual(self.source_bytes, self.source.read_bytes())

    def test_failed_creation_leaves_no_project_card_or_shell(self) -> None:
        invalid = Workbook()
        invalid.active.title = "Not A Singh360 Workbook"
        invalid_bytes = BytesIO()
        invalid.save(invalid_bytes)
        invalid.close()
        invalid_bytes.seek(0)

        response = self.client.post(
            "/api/projects/new",
            data={
                "profile": "ems",
                "file": (invalid_bytes, "invalid.xlsx"),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(400, response.status_code)
        payload = response.get_json()
        self.assertIn("00_PROJECT_META and 00_INDEX", payload["detail"])
        self.assertEqual([], list(server.store.projects_dir.glob("*__*")))
        projects = self.client.get("/api/projects").get_json()["projects"]
        self.assertEqual([], projects)
        staging = server.DOCS_DIR / ".project_staging"
        self.assertFalse(staging.exists() and any(staging.iterdir()))

    def test_existing_project_upload_keeps_id_and_manual_objects(self) -> None:
        created = self._create()
        self.assertEqual(200, created.status_code, created.get_data(as_text=True))
        project_id = created.get_json()["id"]
        project_dir_before = server.store.find_dir(project_id)
        project = server.store.load(project_id)
        self.assertIsNotNone(project)
        project["pages"][0]["canvasObjects"] = [
            {"type": "textbox", "text": "manual object must survive"}
        ]
        server.store.save(project_id, project)

        response = self.client.post(
            f"/api/projects/{project_id}/reimport",
            data={
                "replacePageIds": "[]",
                "file": (BytesIO(self.source_bytes), self.source.name),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(200, response.status_code, response.get_data(as_text=True))
        self.assertEqual(project_id, response.get_json()["id"])
        self.assertEqual(project_dir_before, server.store.find_dir(project_id))
        self.assertEqual(1, len(list(server.store.projects_dir.glob("*__*"))))
        updated = server.store.load(project_id)
        self.assertEqual(
            "manual object must survive",
            updated["pages"][0]["canvasObjects"][0]["text"],
        )
        self.assertGreater(len(updated["worksheets"]), 0)
        self.assertEqual(self.source.name, updated["metadata"]["sourceFile"])
        workbook_files = [
            *project_dir_before.joinpath("sources", "workbook").glob("*.xlsx"),
            *project_dir_before.joinpath("sources", "workbook").glob("*.xlsm"),
        ]
        self.assertEqual([self.source.name], [path.name for path in workbook_files])
        self.assertEqual(
            project_id,
            workbook_metadata(workbook_files[0])["projectId"],
        )
        self.assertEqual(200, self.client.get(f"/api/projects/{project_id}").status_code)
        staging = server.DOCS_DIR / ".workbook_staging"
        self.assertFalse(staging.exists() and any(staging.iterdir()))

    def test_existing_project_rejects_a_different_project_workbook(self) -> None:
        created = self._create()
        self.assertEqual(200, created.status_code, created.get_data(as_text=True))
        project_id = created.get_json()["id"]
        project_dir = server.store.find_dir(project_id)
        project_before = (project_dir / "project.json").read_bytes()
        workbook_path = project_dir / "sources" / "workbook" / self.source.name
        workbook_hash_before = sha256(workbook_path.read_bytes()).hexdigest()

        wrong = load_workbook(BytesIO(self.source_bytes))
        wrong["00_PROJECT_META"]["B1"] = "A Different Sanitized Project"
        wrong_bytes = BytesIO()
        wrong.save(wrong_bytes)
        wrong.close()
        wrong_bytes.seek(0)
        response = self.client.post(
            f"/api/projects/{project_id}/reimport",
            data={
                "replacePageIds": "[]",
                "file": (wrong_bytes, "different_project.xlsx"),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(400, response.status_code)
        self.assertIn(
            "appears to belong to a different project",
            response.get_json()["detail"],
        )
        self.assertEqual(project_before, (project_dir / "project.json").read_bytes())
        self.assertEqual(
            workbook_hash_before,
            sha256(workbook_path.read_bytes()).hexdigest(),
        )
        self.assertEqual(1, len(list(server.store.projects_dir.glob("*__*"))))


if __name__ == "__main__":
    unittest.main()
