# S360 CRITICAL SYNC V44
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.drawing_page_mirror import (
    DRAWING_MANIFEST_SHEET,
    is_generated_mirror_sheet,
    rebuild_drawing_page_mirrors,
    sync_base_index_sheet_tabs,
)
from core.workbook_importer import import_workbook


TAB_COLORS = {
    "draft": "F28C28",
    "draft_confirmed": "76B852",
    "public": "2D7DD2",
    "public_confirmed": "14845A",
    "excluded": "9AA3AB",
    "control": "252C34",
}


def _base_workbook(path: Path) -> None:
    wb = Workbook()
    meta = wb.active
    meta.title = "00_PROJECT_META"
    meta["A1"] = "Linked Project ID"
    meta["B1"] = "0123456789abcdef"

    index = wb.create_sheet("00_INDEX")
    headers = [
        "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
        "Page Type", "Notes", "Render Profile", "Split Mode", "Page ID",
        "Parent Page ID", "Issue Status", "Source Mode", "Sync Direction",
        "Last Sync UTC", "Workbook Hash", "App Hash",
    ]
    for column, header in enumerate(headers, start=1):
        index.cell(4, column, header)
    values = [
        "YES", 1, "EMS 1.0", "EMS 1.0 Cover", "Cover / Project Info",
        "Front", "Cover", "", "", "none", "page_cover", "", "Draft",
        "Workbook", "Both", "", "", "",
    ]
    for column, value in enumerate(values, start=1):
        index.cell(5, column, value)

    # S360 CRITICAL SYNC V44I — this fixture exercises a project containing
    # both a cover and a real base schedule. Exact physical-tab synchronization
    # requires every base page to have its stable Page ID represented in
    # 00_INDEX, exactly as the full workbook mirror writes it before drawing
    # mirrors are rebuilt.
    schedule_values = [
        "YES", 2, "EMS 12.3", "EMS 12.3 Source",
        "Refrigeration Circuit Schedule", "Refrigeration", "data-grid",
        "", "", "auto_rows", "page_schedule", "", "Draft",
        "Workbook", "Both", "", "", "",
    ]
    for column, value in enumerate(schedule_values, start=1):
        index.cell(6, column, value)

    cover = wb.create_sheet("EMS 1.0 Cover")
    cover["A1"] = "Cover"
    wb.save(path)


def test_generated_mirror_tabs_are_complete_and_ignored_on_import(tmp_path: Path) -> None:
    path = tmp_path / "mirror.xlsx"
    _base_workbook(path)

    wb = load_workbook(path)
    cover = wb["EMS 1.0 Cover"]
    project = {
        "id": "0123456789abcdef",
        "metadata": {"projectName": "Test"},
        "pages": [
            {
                "id": "page_cover",
                "order": 1,
                "include": True,
                "sheetCode": "EMS 1.0",
                "displaySheetCode": "EMS 1.0",
                "sheetTab": "EMS 1.0 Cover",
                "sheetTitle": "Cover / Project Info",
                "pageType": "cover",
                "blocks": [],
            },
            {
                "id": "page_schedule",
                "order": 2,
                "include": True,
                "sheetCode": "EMS 12.3",
                "displaySheetCode": "EMS 12.3",
                "sheetTab": "EMS 12.3 Source",
                "sheetTitle": "Refrigeration Circuit Schedule",
                "pageType": "data-grid",
                "blocks": [
                    {
                        "type": "excelRange",
                        "grid": [["SYSTEM", "VALUE"], ["A", "1"]],
                        "styles": {"0:0": {"bold": True}},
                        "mergedCells": [],
                        "colWidths": [90, 90],
                        "rowHeights": [20, 20],
                    }
                ],
            },
            {
                "id": "page_schedule_c1",
                "order": 3,
                "include": True,
                "sheetCode": "EMS 12.3a",
                "displaySheetCode": "EMS 12.3a",
                "sheetTab": "EMS 12.3 Source",
                "sheetTitle": "Refrigeration Circuit Schedule — CONTINUED",
                "pageType": "data-grid",
                "generatedContinuation": True,
                "continuationOf": "page_schedule",
                "continuationIndex": 1,
                "blocks": [
                    {
                        "type": "excelRange",
                        "grid": [["SYSTEM", "VALUE"], ["B", "2"]],
                        "styles": {"0:0": {"bold": True}},
                        "mergedCells": [],
                        "colWidths": [90, 90],
                        "rowHeights": [20, 20],
                        "srcRows": [0, 2],
                    }
                ],
            },
        ],
    }

    source = wb.create_sheet("EMS 12.3 Source")
    source["A1"] = "SYSTEM"
    source["B1"] = "VALUE"
    source["A2"] = "A"
    source["B2"] = "1"
    source["A3"] = "B"
    source["B3"] = "2"

    result = rebuild_drawing_page_mirrors(
        wb,
        project,
        {"page_cover": cover, "page_schedule": source},
        project["id"],
        TAB_COLORS,
    )
    assert result["drawingPageCount"] == 3
    assert result["generatedMirrorCount"] == 1
    assert DRAWING_MANIFEST_SHEET in wb.sheetnames

    wb.save(path)
    wb.close()

    check = load_workbook(path)
    mirror_names = [
        ws.title for ws in check.worksheets
        if is_generated_mirror_sheet(ws)
    ]
    check.close()
    assert len(mirror_names) == 1

    imported = import_workbook(path, project_id="fedcba9876543210")
    imported_names = {
        worksheet.get("name")
        for worksheet in imported.get("worksheets") or []
    }
    assert DRAWING_MANIFEST_SHEET not in imported_names
    assert not (set(mirror_names) & imported_names)

from scripts.repair_critical_sync_v44 import (
    block_fingerprint,
    verify_continuations,
    verify_workbook,
    workbook_tab_lookup,
    resolve_workbook_tab,
)


def _generated_index_page(
    page_id: str,
    code: str,
    grid: list[list[str]],
    *,
    continuation_of: str = "",
    continuation_index: int = 0,
) -> dict:
    page = {
        "id": page_id,
        "order": continuation_index + 1,
        "include": True,
        "sheetCode": code,
        "displaySheetCode": code,
        "sheetTitle": "Sheet Index / TOC" + (" — CONTINUED" if continuation_of else ""),
        "pageType": "index",
        "blocks": [
            {
                "type": "excelRange",
                "srcRows": [0, 1, 2, 3],
                "grid": grid,
            }
        ],
    }
    if continuation_of:
        page.update(
            {
                "continuationOf": continuation_of,
                "pageGroupId": continuation_of,
                "continuationIndex": continuation_index,
                "generatedContinuation": True,
                "indexContinuation": True,
                "generatedIndexContinuation": True,
            }
        )
    return page


def test_generated_index_chunks_with_shared_src_rows_are_not_duplicates() -> None:
    base = _generated_index_page(
        "index_base",
        "EMS 2.0",
        [["PAGE", "SHEET CODE"], ["1", "EMS 1.0"]],
    )
    cont_a = _generated_index_page(
        "index_a",
        "EMS 2.0a",
        [["PAGE", "SHEET CODE"], ["47", "EMS 12.3"]],
        continuation_of="index_base",
        continuation_index=1,
    )
    cont_b = _generated_index_page(
        "index_b",
        "EMS 2.0b",
        [["PAGE", "SHEET CODE"], ["93", "EMS 24.3"]],
        continuation_of="index_base",
        continuation_index=2,
    )

    assert block_fingerprint(base) != block_fingerprint(cont_a)
    assert block_fingerprint(cont_a) != block_fingerprint(cont_b)
    result = verify_continuations({"pages": [base, cont_a, cont_b]})
    assert result["duplicateContinuationCount"] == 0


def test_true_repeated_continuation_slice_is_rejected() -> None:
    base = _generated_index_page(
        "schedule_base",
        "EMS 12.2",
        [["ITEM", "VALUE"], ["A", "1"]],
    )
    cont_a = _generated_index_page(
        "schedule_a",
        "EMS 12.2a",
        [["ITEM", "VALUE"], ["B", "2"]],
        continuation_of="schedule_base",
        continuation_index=1,
    )
    cont_b = _generated_index_page(
        "schedule_b",
        "EMS 12.2b",
        [["ITEM", "VALUE"], ["B", "2"]],
        continuation_of="schedule_base",
        continuation_index=2,
    )

    try:
        verify_continuations({"pages": [base, cont_a, cont_b]})
    except RuntimeError as exc:
        assert "duplicates schedule_a" in str(exc)
    else:
        raise AssertionError("A true repeated continuation slice was not rejected.")

def test_workbook_tab_lookup_preserves_excel_whitespace_identity() -> None:
    wb = Workbook()
    first = wb.active
    first.title = "EMS 12.1a Refrigeration System "
    second = wb.create_sheet("---  BLANK - PAGE END  ---  TEM")

    lookup = workbook_tab_lookup(wb)
    assert resolve_workbook_tab(
        "EMS 12.1a Refrigeration System",
        lookup,
    ) == first.title
    assert resolve_workbook_tab(
        "--- BLANK - PAGE END --- TEM",
        lookup,
    ) == second.title
    wb.close()


def test_verify_workbook_accepts_exact_excel_whitespace_preserved_tabs(
    tmp_path: Path,
) -> None:
    path = tmp_path / "whitespace_mirror.xlsx"
    wb = Workbook()

    meta = wb.active
    meta.title = "00_PROJECT_META"

    index = wb.create_sheet("00_INDEX")
    index_headers = [
        "Include",
        "Order",
        "Sheet Code",
        "Sheet Tab",
        "Page Title",
        "Page ID",
    ]
    for column, header in enumerate(index_headers, start=1):
        index.cell(4, column, header)
    index_values = [
        "NO",
        1,
        "EMS 89.0",
        "---  BLANK - PAGE END  ---  TEM",
        "Blank / Page-End Template",
        "page_blank",
    ]
    for column, value in enumerate(index_values, start=1):
        index.cell(5, column, value)

    manifest = wb.create_sheet(DRAWING_MANIFEST_SHEET)
    manifest_headers = [
        "Page",
        "Include",
        "Sheet Code",
        "Page Title",
        "Excel Tab",
        "Page Kind",
        "Source Base Tab",
        "Page ID",
    ]
    for column, header in enumerate(manifest_headers, start=1):
        manifest.cell(4, column, header)

    manifest_rows = [
        [
            "",
            "NO",
            "EMS 89.0",
            "Blank / Page-End Template",
            "---  BLANK - PAGE END  ---  TEM",
            "Base worksheet",
            "---  BLANK - PAGE END  ---  TEM",
            "page_blank",
        ],
        [
            1,
            "YES",
            "EMS 12.1a",
            "Refrigeration System Table — CONTINUED",
            "EMS 12.1a Refrigeration System ",
            "Generated continuation",
            "EMS 12.1 Refrig Sys Table",
            "page_cont",
        ],
    ]
    for row_number, values in enumerate(manifest_rows, start=5):
        for column, value in enumerate(values, start=1):
            manifest.cell(row_number, column, value)

    blank_sheet = wb.create_sheet("---  BLANK - PAGE END  ---  TEM")
    blank_sheet["A1"] = "Template"

    generated = wb.create_sheet("EMS 12.1a Refrigeration System ")
    generated.sheet_properties.codeName = "S360GEN_TEST"
    generated["A1"] = "Continuation"

    wb.save(path)
    wb.close()

    project = {
        "pages": [
            {
                "id": "page_blank",
                "order": 1,
                "include": False,
                "sheetCode": "EMS 89.0",
                "sheetTitle": "Blank / Page-End Template",
            },
            {
                "id": "page_cont",
                "order": 2,
                "include": True,
                "sheetCode": "EMS 12.1a",
                "sheetTitle": "Refrigeration System Table — CONTINUED",
                "generatedContinuation": True,
                "continuationOf": "page_base",
                "continuationIndex": 1,
            },
        ]
    }

    result = verify_workbook(path, project, tmp_path)
    assert result["drawingPageRows"] == 2
    assert result["generatedMirrorTabs"] == 1

def test_rebuild_syncs_base_index_tab_to_exact_generated_physical_tab() -> None:
    wb = Workbook()
    meta = wb.active
    meta.title = "00_PROJECT_META"

    index = wb.create_sheet("00_INDEX")
    headers = [
        "Include",
        "Order",
        "Sheet Code",
        "Sheet Tab",
        "Page Title",
        "Page ID",
    ]
    for column, header in enumerate(headers, start=1):
        index.cell(4, column, header)
    values = [
        "YES",
        1,
        "EMS 2.0",
        "EMS 2.0 Sheet Index",
        "Sheet Index / TOC",
        "page_index",
    ]
    for column, value in enumerate(values, start=1):
        index.cell(5, column, value)

    stale = wb.create_sheet("EMS 2.0 Sheet Index")
    stale.sheet_properties.codeName = "S360GEN_STALE"

    index_grid = [
        [],
        [],
        [],
        headers.copy(),
        values.copy(),
    ]
    project = {
        "id": "0123456789abcdef",
        "worksheets": [
            {
                "id": "ws_index",
                "name": "00_INDEX",
                "sourceSheet": "00_INDEX",
                "grid": index_grid,
            }
        ],
        "pages": [
            {
                "id": "page_index",
                "order": 1,
                "include": True,
                "sheetCode": "EMS 2.0",
                "displaySheetCode": "EMS 2.0",
                "sheetTab": "EMS 2.0 Sheet Index",
                "sheetTitle": "Sheet Index / TOC",
                "pageType": "index",
                "linkedWorksheetId": "ws_index",
                "blocks": [
                    {
                        "type": "excelRange",
                        "grid": [
                            ["PAGE", "SHEET CODE", "PAGE TITLE"],
                            [1, "EMS 1.0", "Cover"],
                        ],
                        "styles": {},
                        "mergedCells": [],
                        "colWidths": [70, 120, 240],
                        "rowHeights": [20, 20],
                    }
                ],
            }
        ],
    }

    result = rebuild_drawing_page_mirrors(
        wb,
        project,
        {},
        project["id"],
        TAB_COLORS,
    )
    physical = project["pages"][0]["workbookMirrorTab"]

    assert physical == "EMS 2.0 Sheet Index - TOC"
    assert physical in wb.sheetnames
    assert index.cell(5, 4).value == physical
    assert project["pages"][0]["sheetTab"] == physical
    assert project["worksheets"][0]["grid"][4][3] == physical
    assert result["indexTabSync"]["workbookRowsUpdated"] == 1
    assert result["indexTabSync"]["projectPagesUpdated"] == 1
    assert result["indexTabSync"]["projectIndexRowsUpdated"] == 1

    manifest = wb[DRAWING_MANIFEST_SHEET]
    assert manifest.cell(5, 5).value == physical
    wb.close()


def test_verify_workbook_rejects_00_index_tab_text_drift(
    tmp_path: Path,
) -> None:
    path = tmp_path / "index_tab_drift.xlsx"
    wb = Workbook()
    wb.active.title = "00_PROJECT_META"

    index = wb.create_sheet("00_INDEX")
    headers = [
        "Include",
        "Order",
        "Sheet Code",
        "Sheet Tab",
        "Page Title",
        "Page ID",
    ]
    for column, header in enumerate(headers, start=1):
        index.cell(4, column, header)
    values = [
        "NO",
        1,
        "EMS 89.0",
        "--- BLANK - PAGE END --- TEM",
        "Blank / Page-End Template",
        "page_blank",
    ]
    for column, value in enumerate(values, start=1):
        index.cell(5, column, value)

    manifest = wb.create_sheet(DRAWING_MANIFEST_SHEET)
    manifest_headers = [
        "Page",
        "Include",
        "Sheet Code",
        "Page Title",
        "Excel Tab",
        "Page Kind",
        "Source Base Tab",
        "Page ID",
    ]
    for column, header in enumerate(manifest_headers, start=1):
        manifest.cell(4, column, header)
    manifest_values = [
        "",
        "NO",
        "EMS 89.0",
        "Blank / Page-End Template",
        "---  BLANK - PAGE END  ---  TEM",
        "Base worksheet",
        "---  BLANK - PAGE END  ---  TEM",
        "page_blank",
    ]
    for column, value in enumerate(manifest_values, start=1):
        manifest.cell(5, column, value)

    wb.create_sheet("---  BLANK - PAGE END  ---  TEM")
    wb.create_sheet("Generated Check").sheet_properties.codeName = "S360GEN_CHECK"
    wb.save(path)
    wb.close()

    project = {
        "pages": [
            {
                "id": "page_blank",
                "order": 1,
                "include": False,
                "sheetCode": "EMS 89.0",
                "sheetTitle": "Blank / Page-End Template",
            }
        ]
    }

    try:
        verify_workbook(path, project, tmp_path)
    except RuntimeError as exc:
        assert "does not exactly match" in str(exc)
    else:
        raise AssertionError(
            "00_INDEX worksheet-title drift was not rejected."
        )

def test_server_accepts_large_production_project_saves() -> None:
    import server

    limit = int(server.app.config.get("MAX_CONTENT_LENGTH") or 0)
    assert limit == 1024 * 1024 * 1024

