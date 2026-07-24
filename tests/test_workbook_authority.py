from __future__ import annotations

import ast
from copy import deepcopy
import json
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from core.heb_idf_switch_matrix import next_available_continuation_code, sheet_code_key
from core.page_composer import continuation_code
from core.page_identity import is_sheet_index_page
from core.sheet_index_sync import sync_project_sheet_index
from core.workbook_status_sync import (
    _normalize_continuation_identities,
    sync_project_from_workbook,
    sync_project_to_workbook,
)


ROOT = Path(__file__).resolve().parents[1]


class TempStore:
    def __init__(self, root: Path):
        self.docs = root / "docs"
        self.docs.mkdir()

    def dir_for(self, project_id: str, project: dict) -> Path:
        path = self.docs / project_id
        path.mkdir(exist_ok=True)
        return path

    def sources_dir(self, project_id: str, kind: str, project: dict) -> Path:
        path = self.dir_for(project_id, project) / "sources" / kind
        path.mkdir(parents=True, exist_ok=True)
        return path

    def assets_excel_dir(self, project_id: str, project: dict) -> Path:
        path = self.dir_for(project_id, project) / "assets" / "excel"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def save(self, project_id: str, project: dict) -> None:
        (self.dir_for(project_id, project) / "project.json").write_text(
            json.dumps(project, default=str), encoding="utf-8"
        )


def index_page() -> dict:
    return {
        "id": "page-index",
        "order": 2,
        "include": True,
        "sheetCode": "EMS 2.0",
        "displaySheetCode": "EMS 2.0",
        "sheetTab": "EMS 2.0 Sheet Index",
        "sheetTitle": "Sheet Index / TOC",
        "pageType": "Sheet Index",
        "renderMode": "generated_index",
        "blocks": [{
            "type": "excelRange",
            "grid": [
                ["SINGH360 EMS — SHEET INDEX / TOC", "", "", "", ""],
                ["Only explicit YES rows publish.", "", "", "", ""],
                ["", "", "", "", ""],
                ["INCLUDE", "ORDER", "SHEET CODE", "SHEET TAB", "PAGE TITLE"],
            ],
            "rowHeights": [20, 20, 20, 20],
        }],
    }


class WorkbookAuthorityTests(unittest.TestCase):
    def test_no_duplicate_top_level_functions(self) -> None:
        for relative in ("core/workbook_link_manager.py", "core/workbook_status_sync.py"):
            tree = ast.parse((ROOT / relative).read_text(encoding="utf-8"))
            names = [
                node.name for node in tree.body
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
            ]
            duplicates = sorted({name for name in names if names.count(name) > 1})
            self.assertEqual([], duplicates, f"{relative}: {duplicates}")

    def test_sheet_index_predicate_accepts_controlled_spellings(self) -> None:
        for value in ("index", "Sheet Index", "sheet-index", "sheet_index"):
            self.assertTrue(is_sheet_index_page({"pageType": value}))
        self.assertTrue(is_sheet_index_page({
            "pageType": "Front Matter",
            "sheetCode": "EMS 2.0",
            "sheetTab": "EMS 2.0 Sheet Index",
            "sheetTitle": "TOC",
        }))

    def test_sheet_index_paginates_distinct_complete_chunks_idempotently(self) -> None:
        pages = [
            {
                "id": "page-cover", "order": 1, "include": True,
                "sheetCode": "EMS 1.0", "displaySheetCode": "EMS 1.0",
                "sheetTab": "EMS 1.0 Cover", "sheetTitle": "Cover", "pageType": "cover",
            },
            index_page(),
        ]
        pages.extend({
            "id": f"page-{number}", "order": number + 2, "include": True,
            "sheetCode": f"EMS X{number}", "displaySheetCode": f"EMS X{number}",
            "sheetTab": f"Sheet {number}", "sheetTitle": f"Page {number}",
            "pageType": "data-grid",
        } for number in range(1, 99))
        project = {"id": "test", "pages": pages, "worksheets": []}

        first = sync_project_sheet_index(deepcopy(project))
        index_pages = [p for p in first["pages"] if is_sheet_index_page(p)]
        self.assertGreater(len(index_pages), 1)
        chunks = [
            next(b for b in p["blocks"] if b["type"] == "excelRange")["grid"][4:]
            for p in index_pages
        ]
        self.assertTrue(all(chunks))
        self.assertEqual(len(chunks), len({chunk[0][1] for chunk in chunks}))
        listed = [row[2] for chunk in chunks for row in chunk]
        published = [
            p.get("displaySheetCode") or p.get("sheetCode")
            for p in first["pages"] if p.get("include", True)
        ]
        self.assertEqual(published, listed)
        self.assertEqual("EMS 1.0", published[0])
        self.assertEqual("EMS 2.0", published[1])

        second = sync_project_sheet_index(deepcopy(first))
        identity = lambda doc: [
            (p["id"], p.get("sheetCode"), p.get("continuationOf"))
            for p in doc["pages"]
        ]
        self.assertEqual(identity(first), identity(second))

    def test_continuation_suffix_advances_and_skips_reserved_codes(self) -> None:
        self.assertEqual("EMS 8.0b", continuation_code("EMS 8.0a", 1))
        used = {sheet_code_key("EMS 8.0b"), sheet_code_key("EMS 8.0c")}
        self.assertEqual(
            "EMS 8.0d",
            next_available_continuation_code("EMS 8.0a", 1, used),
        )

    def test_continuation_ids_remap_to_stable_base_without_collision(self) -> None:
        pages = [
            {"id": "stable-base", "pageGroupId": "import-base"},
            {"id": "import-base_c1", "generatedContinuation": True, "continuationOf": "import-base", "continuationIndex": 1, "sheetTitle": "Table — CONTINUED — CONTINUED"},
            {"id": "stable-base__continuation_1", "pageGroupId": "other-import"},
        ]
        _normalize_continuation_identities(
            pages,
            {"import-base": "stable-base", "other-import": "stable-base__continuation_1"},
        )
        self.assertEqual("stable-base", pages[1]["continuationOf"])
        self.assertEqual("stable-base__continuation_1_2", pages[1]["id"])
        self.assertEqual("Table — CONTINUED", pages[1]["sheetTitle"])
        self.assertEqual(len(pages), len({page["id"] for page in pages}))

    def test_workbook_writer_never_adds_generated_index_rows(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            path = root / "book.xlsx"
            wb = Workbook()
            meta = wb.active
            meta.title = "00_PROJECT_META"
            index = wb.create_sheet("00_INDEX")
            wb.create_sheet("00_HELP")
            wb.create_sheet("EMS 1.0 Cover")
            wb.create_sheet("EMS 2.0 Sheet Index")
            headers = [
                "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
                "Page Type", "Notes", "Render Profile", "Split Mode", "Page ID",
                "Parent Page ID", "Issue Status", "Source Mode", "Sync Direction",
                "Last Sync UTC", "Workbook Hash", "App Hash",
            ]
            for col, value in enumerate(headers, 1):
                index.cell(4, col, value)
            rows = [
                ("YES", 1, "EMS 1.0", "EMS 1.0 Cover", "Cover", "cover", "cover", "", "", "none", "cover-id"),
                ("YES", 2, "EMS 2.0", "EMS 2.0 Sheet Index", "Sheet Index / TOC", "index", "Sheet Index", "", "", "none", "index-id"),
            ]
            for row_number, values in enumerate(rows, 5):
                for col, value in enumerate(values, 1):
                    index.cell(row_number, col, value)
            wb.save(path)
            wb.close()

            project = {
                "id": "test", "metadata": {}, "worksheets": [],
                "workbookSync": {"workbook": str(path)},
                "pages": [
                    {"id": "cover-id", "order": 1, "include": True, "sheetCode": "EMS 1.0", "sheetTab": "EMS 1.0 Cover"},
                    {"id": "index-id", "order": 2, "include": True, "sheetCode": "EMS 2.0", "sheetTab": "EMS 2.0 Sheet Index", "pageType": "Sheet Index"},
                    {"id": "index-id__index_cont_1", "order": 3, "include": True, "sheetCode": "EMS 2.0a", "sheetTab": "EMS 2.0 Sheet Index", "generatedContinuation": True, "continuationOf": "index-id"},
                ],
            }
            sync_project_to_workbook("test", project, TempStore(root))
            check = load_workbook(path, read_only=True)
            data = list(check["00_INDEX"].iter_rows(min_row=5, values_only=True))
            check.close()
            nonempty = [row for row in data if any(value not in (None, "") for value in row)]
            self.assertEqual(2, len(nonempty))
            self.assertEqual([1, 2], [row[1] for row in nonempty])
            self.assertNotIn("EMS 2.0a", [row[2] for row in nonempty])

    def test_workbook_page_id_replaces_legacy_id_without_losing_manual_data(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            path = root / "book.xlsx"
            wb = Workbook()
            wb.active.title = "00_PROJECT_META"
            index = wb.create_sheet("00_INDEX")
            headers = ["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Page ID"]
            for col, value in enumerate(headers, 1):
                index.cell(4, col, value)
            values = ["YES", 1, "EMS 9.1", "EMS 9.1 RDM Parts", "RDM Parts", "stable-page-id"]
            for col, value in enumerate(values, 1):
                index.cell(5, col, value)
            wb.create_sheet("EMS 9.1 RDM Parts")
            wb.save(path)
            wb.close()

            imported = {
                "metadata": {}, "worksheets": [],
                "pages": [{
                    "id": "import-id", "order": 1, "include": True,
                    "sheetCode": "EMS 9.1", "sheetTab": "EMS 9.1 RDM Parts",
                    "sheetTitle": "RDM Parts", "pageType": "data-grid",
                }],
            }
            project = {
                "id": "test", "metadata": {}, "worksheets": [],
                "workbookSync": {"workbook": str(path)},
                "pages": [{
                    "id": "legacy-id", "order": 1, "include": True,
                    "sheetCode": "EMS 9.1", "sheetTab": "EMS 9.1 RDM Parts",
                    "canvasObjects": [{"type": "text", "text": "keep"}],
                }],
            }
            with patch("core.workbook_importer.import_workbook", return_value=imported):
                result = sync_project_from_workbook("test", project, TempStore(root))
            self.assertEqual("stable-page-id", result["pages"][0]["id"])
            self.assertEqual(project["pages"][0]["canvasObjects"], result["pages"][0]["canvasObjects"])


if __name__ == "__main__":
    unittest.main()
