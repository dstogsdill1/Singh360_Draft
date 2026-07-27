from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill

from core.project_workspace import WorkbookDocumentStore, workbook_file_to_document
from core.workbook_geometry import (
    excel_column_width_to_pixels,
    pdf_points_to_pixels,
    pixels_to_excel_column_width,
    pixels_to_pdf_points,
    pixels_to_row_height_points,
    row_height_points_to_pixels,
    unchanged_excel_width_or_converted,
    unchanged_row_height_or_converted,
)
from core.page_composer import _split_excel_range_block
from core.workbook_importer import (
    _apply_table_geometry,
    _excel_range_block,
    _worksheet_payload,
    import_workbook,
)
from core.workbook_status_sync import _s360_apply_worksheet_payload
from core.worksheet_export import export_worksheet_xlsx


class WorkbookGeometryTests(unittest.TestCase):
    def test_canonical_unit_conversions_and_exact_unchanged_values(self) -> None:
        self.assertEqual(64, excel_column_width_to_pixels(8.43))
        self.assertEqual(225, excel_column_width_to_pixels(31.5))
        self.assertEqual(31.5, unchanged_excel_width_or_converted(225, 31.5))
        resized_width = pixels_to_excel_column_width(240)
        self.assertEqual(240, excel_column_width_to_pixels(resized_width))

        self.assertEqual(20.0, row_height_points_to_pixels(15.0))
        self.assertEqual(24.0, row_height_points_to_pixels(18.0))
        self.assertEqual(
            18.0,
            unchanged_row_height_or_converted(24.0, 18.0),
        )
        self.assertEqual(18.0, pixels_to_row_height_points(24.0))
        self.assertEqual(72.0, pixels_to_pdf_points(96.0))
        self.assertEqual(96.0, pdf_points_to_pixels(72.0))

    def test_import_and_project_local_persistence_keep_full_geometry(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "fixture.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Geometry"
            sheet.sheet_format.defaultColWidth = 9.25
            sheet.sheet_format.defaultRowHeight = 16.5
            sheet.column_dimensions["A"].width = 31.5
            sheet.column_dimensions["B"].width = 11.25
            sheet.column_dimensions["B"].hidden = True
            sheet.row_dimensions[1].height = 24.0
            sheet.row_dimensions[2].hidden = True
            sheet.merge_cells("A1:C1")
            sheet["A1"] = "ORANGE INSTRUCTION BAND"
            sheet["A1"].fill = PatternFill("solid", fgColor="F4B183")
            sheet["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            sheet["A2"] = "=1+1"
            workbook.save(source)
            workbook.close()

            document = workbook_file_to_document(source)
            imported = document["sheets"][0]
            self.assertEqual(9.25, imported["defaultColumnWidth"])
            self.assertEqual(16.5, imported["defaultRowHeight"])
            self.assertEqual(31.5, imported["columnWidths"]["A"])
            self.assertEqual(24.0, imported["rowHeights"]["1"])
            self.assertEqual(["B"], imported["hiddenColumns"])
            self.assertEqual([2], imported["hiddenRows"])
            self.assertEqual(["A1:C1"], imported["merges"])
            self.assertEqual("=1+1", imported["cells"]["A2"]["f"])
            self.assertTrue(imported["styles"]["A1"]["wrap"])
            self.assertEqual("center", imported["styles"]["A1"]["hAlign"])
            self.assertEqual("#F4B183", imported["styles"]["A1"]["fill"])

            project = {
                "id": "geometry-fixture",
                "metadata": {"projectName": "Geometry Fixture"},
                "worksheets": [],
            }
            store = WorkbookDocumentStore(root)
            saved = store.save(project, 0, document)
            reloaded = store.load(project)
            self.assertEqual(saved, reloaded)
            self.assertEqual(31.5, reloaded["sheets"][0]["columnWidths"]["A"])
            self.assertEqual(24.0, reloaded["sheets"][0]["rowHeights"]["1"])
            self.assertEqual(["B"], reloaded["sheets"][0]["hiddenColumns"])
            self.assertEqual(["A1:C1"], reloaded["sheets"][0]["merges"])

            formula_workbook = load_workbook(source, data_only=False)
            value_workbook = load_workbook(source, data_only=True)
            payload = _worksheet_payload(
                formula_workbook["Geometry"],
                value_workbook["Geometry"],
            )
            formula_workbook.close()
            value_workbook.close()
            self.assertEqual(225, payload["colWidthsPx"][0])
            self.assertEqual(32.0, payload["rowHeightsPx"][0])
            self.assertEqual([1], payload["hiddenColumns"])
            self.assertEqual([1], payload["hiddenRows"])
            self.assertEqual("workbook-v1", payload["geometryAuthority"])

            project_import = import_workbook(source, project_id="geometry-import")
            project_sheet = next(
                item
                for item in project_import["worksheets"]
                if item["name"] == "Geometry"
            )
            self.assertEqual(9.25, project_sheet["defaultColumnWidth"])
            self.assertEqual(16.5, project_sheet["defaultRowHeight"])
            self.assertEqual([1], project_sheet["hiddenColumns"])
            self.assertEqual([1], project_sheet["hiddenRows"])
            self.assertEqual("workbook-v1", project_sheet["geometryAuthority"])

            block = _excel_range_block(project_sheet, "hidden-geometry")
            self.assertEqual(
                ["ORANGE INSTRUCTION BAND", ""],
                block["grid"][0],
            )
            self.assertEqual([225, 70], block["colWidths"])
            self.assertEqual([32.0], block["rowHeights"])
            self.assertEqual([0], block["srcRows"])
            self.assertEqual(
                [
                    {
                        "startRow": 0,
                        "startCol": 0,
                        "endRow": 0,
                        "endCol": 1,
                    }
                ],
                block["mergedCells"],
            )

    def test_excel_writeback_and_standalone_export_round_trip_geometry(self) -> None:
        project = {
            "worksheets": [
                {
                    "id": "geometry-sheet",
                    "name": "Geometry",
                    "sourceSheet": "Geometry",
                    "grid": [
                        ["ORANGE INSTRUCTION BAND", "", ""],
                        ["Wrapped words stay together", "2", ""],
                    ],
                    "formulas": {"B2": "=1+1"},
                    "styles": {
                        "A1": {
                            "fill": "#F4B183",
                            "bold": True,
                            "hAlign": "center",
                            "vAlign": "center",
                            "wrap": True,
                        },
                        "A2": {"wrap": True, "hAlign": "left"},
                    },
                    "mergedCells": [
                        {
                            "startRow": 0,
                            "startCol": 0,
                            "endRow": 0,
                            "endCol": 2,
                        }
                    ],
                    "columnWidths": {"A": 31.5, "B": 11.25, "C": 8.43},
                    "rowHeights": {"1": 24.0, "2": 18.0},
                    "colWidthsPx": [225, 84, 64],
                    "rowHeightsPx": [32.0, 24.0],
                    "defaultColumnWidth": 9.25,
                    "defaultRowHeight": 16.5,
                    "hiddenRows": [1],
                    "hiddenColumns": [2],
                    "geometryAuthority": "workbook-v1",
                }
            ]
        }

        workbook = Workbook()
        workbook.active.title = "Geometry"
        _s360_apply_worksheet_payload(workbook, project)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "writeback.xlsx"
            workbook.save(path)
            workbook.close()
            checked = load_workbook(path, data_only=False)
            sheet = checked["Geometry"]
            self.assertEqual(31.5, sheet.column_dimensions["A"].width)
            self.assertEqual(24.0, sheet.row_dimensions[1].height)
            self.assertTrue(sheet.row_dimensions[2].hidden)
            self.assertTrue(sheet.column_dimensions["C"].hidden)
            self.assertIn("A1:C1", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual("=1+1", sheet["B2"].value)
            self.assertTrue(sheet["A1"].alignment.wrap_text)
            self.assertEqual("F4B183", sheet["A1"].fill.fgColor.rgb[-6:])
            checked.close()

            data = export_worksheet_xlsx(project["worksheets"][0])
            exported = Path(tmp) / "standalone.xlsx"
            exported.write_bytes(data)
            checked = load_workbook(exported, data_only=False)
            sheet = checked["Geometry"]
            self.assertEqual(31.5, sheet.column_dimensions["A"].width)
            self.assertEqual(24.0, sheet.row_dimensions[1].height)
            self.assertTrue(sheet.row_dimensions[2].hidden)
            self.assertTrue(sheet.column_dimensions["C"].hidden)
            self.assertIn("A1:C1", {str(item) for item in sheet.merged_cells.ranges})
            checked.close()

    def test_drawing_geometry_is_uniform_and_continuations_reuse_width_map(self) -> None:
        widths = [72, 216, 432]
        block = {
            "id": "geometry-block",
            "type": "excelRange",
            "grid": [
                ["ORANGE INSTRUCTION BAND", "", ""],
                *[
                    [
                        str(index),
                        "Normal words wrap only at spaces",
                        "Readable continuation content " * 3,
                    ]
                    for index in range(1, 61)
                ],
            ],
            "styles": {
                "0:0": {"fill": "#F4B183", "bold": True, "wrap": True},
                **{
                    f"{row}:2": {"wrap": True}
                    for row in range(1, 61)
                },
            },
            "mergedCells": [
                {
                    "startRow": 0,
                    "startCol": 0,
                    "endRow": 0,
                    "endCol": 2,
                }
            ],
            "colWidths": list(widths),
            "rowHeights": [24, *([30] * 60)],
            "headerRowCount": 1,
            "repeatRows": [0],
            "splitMode": "auto_rows",
            "minScale": 0.5,
            "allowContinuation": True,
            "scaleMode": "fit_body",
        }
        _apply_table_geometry(
            block,
            family="text",
            page_type="data-grid",
            layout_profile="instruction_table",
        )
        self.assertEqual(widths, block["colWidths"])
        parts = _split_excel_range_block(block)
        self.assertGreater(len(parts), 1)
        for part in parts:
            self.assertEqual(widths, part["colWidths"])
            self.assertEqual(
                [{"startRow": 0, "startCol": 0, "endRow": 0, "endCol": 2}],
                part["mergedCells"],
            )


if __name__ == "__main__":
    unittest.main()
