from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from core.ems_workbook_contract import CANONICAL_REPOSITORY
from core.project_preflight import compute_project_preflight
from core.workbook_importer import import_workbook
from core.workbook_reimport import apply_reimport


def _recipe(workbook: Workbook, tab: str, code: str, title: str) -> None:
    ws = workbook.create_sheet(tab)
    ws.append([title])
    ws.append([code, "page recipe"])
    ws.append([])
    ws.append(["Field", "Value", "Notes"])
    ws.append(["Current Content", "Placeholder", "Recipe rows never publish."])


def _canonical(workbook: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(name)
    ws.append([name])
    ws.append(["Canonical editable table"])
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)


def write_contract_workbook(path: Path, *, stale: bool = False) -> Path:
    workbook = Workbook()
    meta = workbook.active
    meta.title = "00_PROJECT_META"
    meta.append(["SINGH360 EMS - PROJECT METADATA"])
    meta.append(["Metadata authority"])
    meta.append([])
    meta.append(["Field", "Value", "Authority", "Notes"])
    rows = [
        ["Project Name", "Sanitized 829 Contract Test", "Project", ""],
        ["Drawing Package File Name", "SANITIZED_829_PACKAGE", "Project", ""],
        ["Project Revision", "V3 TEMPLATE" if stale else "R1", "Project", ""],
        ["Template Version", "1.0.0", "Template", ""],
        ["Location", "100 Test Way", "Project", ""],
        ["Linked Project ID", "stale-project" if stale else "", "Application", ""],
        ["Repository", "dstogsdill1/Singh360_SmartDraw" if stale else CANONICAL_REPOSITORY, "Application", ""],
    ]
    for row in rows:
        meta.append(row)

    index = workbook.create_sheet("00_INDEX")
    index.append(["SINGH360 EMS - PAGE MANIFEST"])
    index.append(["Only YES publishes"])
    index.append([])
    index.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    page_rows = [
        ["YES", 1, "EMS 1.0", "EMS 1.0 Cover", "Cover / Project Info", "cover", "cover", ""],
        ["YES", 2, "EMS 2.0", "EMS 2.0 Sheet Index", "Sheet Index / TOC", "Front Matter", "index", ""],
        ["YES", 3, "EMS 3.0", "EMS 3.0 Guidelines", "Guidelines", "Front Matter", "data-grid", ""],
        ["YES", 4, "EMS 4.0", "EMS 4.0 Abbrev", "Abbreviations / Symbol Key", "Front Matter", "data-grid", ""],
        ["YES", 5, "EMS 5.0", "EMS 5.0 Directory", "Project Directory / Contacts", "Front Matter", "data-grid", ""],
        ["YES", 6, "EMS 7.0", "EMS 7.0 Workflow", "Project Workflow / Milestones", "Front Matter", "data-grid", ""],
        ["YES", 7, "EMS 8.0", "EMS 8.0 Resp Matrix", "Responsibility Matrix", "Front Matter", "data-grid", ""],
        ["YES", 8, "EMS 12.0", "EMS 12.0 Overall Layout", "Overall Layout", "Layout", "canvas", ""],
        ["YES", 9, "EMS 13.0", "EMS 13.0 Network Summary", "WICP / IDF / Network Summary", "Network", "data-grid", ""],
        ["YES", 10, "EMS 13.1", "EMS 13.1 IDF 1", "IDF #1 Port / Network Table", "Network", "data-grid", ""],
        ["YES", 11, "EMS 13.2", "EMS 13.2 IDF 2", "IDF #2 Port / Network Table", "Network", "data-grid", ""],
        ["YES", 12, "EMS 13.3", "EMS 13.3 IDF 3", "IDF #3 Port / Network Table", "Network", "data-grid", ""],
        ["YES", 13, "EMS 18.0", "EMS 18.0 Rack A IO", "Rack A I/O Schedule", "I/O", "data-grid", ""],
        ["YES", 14, "EMS 19.0", "EMS 19.0 Rack B IO", "Rack B I/O Schedule", "I/O", "data-grid", ""],
        ["YES", 15, "EMS 20.0", "EMS 20.0 Rack C IO", "Rack C I/O Schedule", "I/O", "data-grid", ""],
        ["YES", 16, "EMS 21.0", "EMS 21.0 WICP Summary", "WICP Count Summary", "Panels", "data-grid", ""],
        ["YES", 17, "EMS 22.0", "EMS 22.0 WICP IO", "WICP I/O Schedule", "I/O", "data-grid", ""],
        ["YES", 18, "EMS 23.0", "EMS 23.0 Case Controllers", "Case Controller Schedule", "Panels", "data-grid", ""],
        ["YES", 19, "EMS 24.0", "EMS 24.0 Lighting Matrix", "Lighting Output Matrix", "Lighting", "data-grid", ""],
        ["YES", 20, "EMS 24.1", "EMS 24.1 Lighting IO", "Lighting TDB I/O Schedule", "Lighting", "data-grid", ""],
        ["YES", 21, "EMS 24.2", "EMS 24.2 Lighting Schedule", "Lighting Control / Dimming Schedule", "Lighting", "data-grid", ""],
    ]
    for row in page_rows:
        index.append(row)

    network_rows = [
        [1, 1, 1, "PORT-1", "Generic test device"],
        [1, 1, 2, "PORT-2", "Generic test device"],
        [2, 3, 1, "PORT-3", "Generic test device"],
        [3, 4, 1, "PORT-4", "Generic test device"],
    ]
    if stale:
        network_rows[0][-1] = "Spare network port - sample from template"
    _canonical(workbook, "21_NETWORK_PORTS", ["IDF", "Switch", "Port", "Label", "Device / Drop"], network_rows)
    _canonical(
        workbook,
        "20_CONTROLLERS",
        [
            "Controller ID", "Controller Label", "Controller Type", "Panel / Location",
            "Network / IDF", "IP Address", "Source ID", "Status", "Notes",
        ],
        [
            ["CC-1", "Case Controller 1", "Case Controller", "Rack A", "IDF 1", "", "SRC-1", "VERIFY", ""],
            ["601", "Lighting Controller", "Lighting", "LCP-1", "IDF 1", "", "SRC-2", "VERIFY", ""],
        ],
    )
    _canonical(
        workbook,
        "22_PANELS",
        ["Panel ID", "Panel Type", "Panel Name", "Rack/System", "Location", "Status"],
        [
            ["WICP-01", "WICP", "WICP #1", "", "Sales Floor", "VERIFY"],
            ["WICP-02", "WICP", "WICP #2", "", "Back Room", "VERIFY"],
            ["LCP-1", "LCP", "LCP-1", "Lighting", "", "VERIFY"],
        ],
    )
    panel_io_rows = [
        ["WICP-01", "CC-1", "WICP", "1", "DI", "POINT-1", "WICP input", "", "", "VERIFY"],
        ["RACK-A-IO", "CC-A", "Rack A", "1", "DO", "RA-1", "Rack A output", "Rack A", "", "VERIFY"],
        ["RACK-B-IO", "CC-B", "Rack B", "1", "DO", "RB-1", "Rack B output", "Rack B", "", "VERIFY"],
    ]
    if not stale:
        panel_io_rows.append(
            ["RACK-C-IO", "CC-C", "Rack C", "1", "DO", "RC-1", "Rack C output", "Rack C", "", "VERIFY"]
        )
    _canonical(
        workbook,
        "23_PANEL_IO",
        [
            "Panel ID", "Controller ID", "I/O Group", "Point No.", "Point Type",
            "Point Label", "Description", "Device / Case", "Cable / Terminal", "Status",
        ],
        panel_io_rows,
    )
    _canonical(
        workbook,
        "26_LIGHTING_OUTPUTS",
        [
            "Output / Zone", "Controller ID", "Panel", "Point", "Output Type",
            "Description", "Area / Fixture Group", "Schedule / Time", "Source ID", "Status", "Notes",
        ],
        [
            ["C1", "601", "LCP-1", "RO1", "Relay", "Display lights", "Sales", "Open", "SRC-E", "VERIFY", ""],
            ["D1", "601", "LCP-1", "AIO1", "0-10V Dimming", "Dimming zone", "Sales", "Open", "SRC-E", "VERIFY", ""],
        ],
    )
    _canonical(workbook, "32_GUIDELINES", ["Guideline ID", "Topic", "Requirement"], [["G-1", "Testing", "Verify the generated package"]])
    _canonical(
        workbook,
        "00_HELP",
        ["Step", "Workflow", "What goes here", "Rule"],
        [["1", "Review", "Review canonical data before export.", "Do not publish recipe rows."]],
    )
    _canonical(
        workbook,
        "03_SCOPE_AND_PLAN",
        ["Phase", "Task", "Output", "Status", "Notes"],
        [["1 - Intake", "Collect sources", "Source library", "NOT STARTED", "Use verified sources."]],
    )

    for row in page_rows:
        _recipe(workbook, row[3], row[2], row[4])
    workbook.save(path)
    workbook.close()
    return path


class EmsWorkbookContractTests(unittest.TestCase):
    def test_recipe_pages_render_canonical_tables_and_filter_idf(self) -> None:
        with TemporaryDirectory(prefix="s360_contract_") as raw:
            workbook = write_contract_workbook(Path(raw) / "contract.xlsx")
            project = import_workbook(workbook, project_id="contract-project")

        pages = {page["sheetCode"]: page for page in project["pages"] if not page.get("generatedContinuation")}
        self.assertEqual("Sanitized 829 Contract Test", project["metadata"]["projectName"])
        self.assertEqual("SANITIZED_829_PACKAGE", project["metadata"]["drawingPackageFileName"])
        self.assertEqual("R1", project["metadata"]["revision"])
        self.assertEqual("1.0.0", project["metadata"]["templateVersion"])

        all_blocks = str([page.get("blocks") for page in pages.values()])
        self.assertNotIn("Current Content", all_blocks)
        self.assertNotIn("Placeholder", all_blocks)

        idf1 = pages["EMS 13.1"]
        idf2 = pages["EMS 13.2"]
        self.assertEqual("21_NETWORK_PORTS", idf1["sourceSheet"])
        self.assertEqual(2, idf1["canonicalDataRowCount"])
        self.assertEqual(1, idf2["canonicalDataRowCount"])
        self.assertEqual(
            {"1"},
            {
                str(row[0])
                for row in idf1["blocks"][0]["grid"][4:]
                if row and str(row[0]).strip()
            },
        )
        self.assertEqual(
            {"2"},
            {
                str(row[0])
                for row in idf2["blocks"][0]["grid"][4:]
                if row and str(row[0]).strip()
            },
        )
        self.assertEqual("32_GUIDELINES", pages["EMS 3.0"]["sourceSheet"])
        for code, canonical in (
            ("EMS 4.0", "33_ABBREVIATIONS"),
            ("EMS 5.0", "34_PROJECT_DIRECTORY"),
            ("EMS 7.0", "36_WORKFLOW_MILESTONES"),
            ("EMS 8.0", "37_RESPONSIBILITY_MATRIX"),
        ):
            self.assertEqual(canonical, pages[code]["sourceSheet"])
            self.assertGreater(pages[code]["canonicalDataRowCount"], 0)

        network_summary = pages["EMS 13.0"]
        self.assertEqual(
            ["network_summary", "wicp_count_summary"],
            [block["canonicalView"] for block in network_summary["blocks"]],
        )
        self.assertEqual(
            {"21_NETWORK_PORTS", "22_PANELS"},
            {block["canonicalSourceSheet"] for block in network_summary["blocks"]},
        )
        self.assertFalse(
            any(
                page.get("generatedContinuation")
                and str(page.get("sheetCode") or "").startswith("EMS 13.0")
                for page in project["pages"]
            )
        )

        for code, rack in (("EMS 18.0", "A"), ("EMS 19.0", "B"), ("EMS 20.0", "C")):
            page = pages[code]
            block = page["blocks"][0]
            self.assertEqual("23_PANEL_IO", block["canonicalSourceSheet"])
            self.assertEqual("rack_io", block["canonicalView"])
            self.assertEqual(rack, block["canonicalViewFilter"])
            self.assertTrue(
                all(f"RACK-{rack}" in str(row[0]) for row in block["grid"][4:])
            )

        wicp_summary = pages["EMS 21.0"]["blocks"][0]
        self.assertEqual("22_PANELS", wicp_summary["canonicalSourceSheet"])
        self.assertEqual(["Panel Type", "Count", "Panel IDs", "Location", "Status"], wicp_summary["grid"][3])
        self.assertEqual("2", str(wicp_summary["grid"][4][1]))

        wicp_io = pages["EMS 22.0"]["blocks"][0]
        self.assertEqual("23_PANEL_IO", wicp_io["canonicalSourceSheet"])
        self.assertTrue(all(str(row[0]).startswith("WICP") for row in wicp_io["grid"][4:]))

        case_schedule = pages["EMS 23.0"]["blocks"][0]
        self.assertEqual("20_CONTROLLERS", case_schedule["canonicalSourceSheet"])
        self.assertEqual(["CC-1"], [str(row[0]) for row in case_schedule["grid"][4:]])

        lighting = [pages[code]["blocks"][0] for code in ("EMS 24.0", "EMS 24.1", "EMS 24.2")]
        self.assertEqual(
            ["lighting_matrix", "lighting_io", "lighting_dimming"],
            [block["canonicalView"] for block in lighting],
        )
        self.assertEqual(3, len({tuple(block["grid"][3]) for block in lighting}))
        self.assertEqual(["D1"], [str(row[0]) for row in lighting[2]["grid"][4:]])
        self.assertTrue(all(block["canonicalSourceSheet"] == "26_LIGHTING_OUTPUTS" for block in lighting))

        cover_text = str(pages["EMS 1.0"]["blocks"])
        self.assertIn("SANITIZED_829_PACKAGE", cover_text)
        self.assertIn("Project Revision: R1", cover_text)
        self.assertNotIn("Template Version", cover_text)

    def test_preflight_reports_empty_mapping_and_stale_template_data(self) -> None:
        with TemporaryDirectory(prefix="s360_preflight_") as raw:
            workbook = write_contract_workbook(Path(raw) / "stale.xlsx", stale=True)
            project = import_workbook(workbook, project_id="contract-project")

        codes = {issue["code"] for issue in compute_project_preflight(project)}
        self.assertIn("included_page_no_mapped_data", codes)
        self.assertIn("stale_linked_project_id", codes)
        self.assertIn("stale_repository_name", codes)
        self.assertIn("template_text_in_project_revision", codes)
        self.assertIn("sample_engineering_rows", codes)
        self.assertNotIn("included_page_recipe_only", codes)

    def test_manual_canvas_objects_survive_canonical_reimport(self) -> None:
        with TemporaryDirectory(prefix="s360_contract_reimport_") as raw:
            workbook = write_contract_workbook(Path(raw) / "contract.xlsx")
            project = import_workbook(workbook, project_id="contract-project")
            layout = next(page for page in project["pages"] if page["sheetCode"] == "EMS 12.0")
            layout["canvasObjects"] = [{"type": "textbox", "text": "manual survives"}]
            cover = next(page for page in project["pages"] if page["sheetCode"] == "EMS 1.0")
            cover["canvasObjects"] = [{"type": "image", "src": "manual-logo.png"}]
            editable = load_workbook(workbook)
            try:
                meta = editable["00_PROJECT_META"]
                for row in meta.iter_rows(min_col=1, max_col=2):
                    if row[0].value == "Project Revision":
                        row[1].value = "R2"
                        break
                editable.save(workbook)
            finally:
                editable.close()
            updated, summary = apply_reimport(
                project,
                workbook,
                project_id="contract-project",
                source_filename=workbook.name,
            )

        updated_layout = next(page for page in updated["pages"] if page["sheetCode"] == "EMS 12.0")
        self.assertEqual("manual survives", updated_layout["canvasObjects"][0]["text"])
        self.assertIn("EMS 12.0", summary["preserved"])
        updated_cover = next(page for page in updated["pages"] if page["sheetCode"] == "EMS 1.0")
        self.assertEqual("manual-logo.png", updated_cover["canvasObjects"][0]["src"])
        self.assertIn("Project Revision: R2", str(updated_cover["blocks"]))
        self.assertIn("EMS 1.0", summary["updated"])
        updated_abbreviations = next(
            page for page in updated["pages"] if page["sheetCode"] == "EMS 4.0"
        )
        self.assertGreater(updated_abbreviations["canonicalDataRowCount"], 0)
        self.assertEqual([], updated_abbreviations["missingCanonicalSources"])

    def test_tracked_runtime_template_is_clean_and_styled(self) -> None:
        template = (
            Path(__file__).resolve().parents[1]
            / "defaults"
            / "runtime_templates"
            / "Singh360_BASE_Project_Workbook_Template_V1.xlsx"
        )
        workbook = load_workbook(template)
        try:
            for required in (
                "00_PROJECT_META", "00_INDEX", "21_NETWORK_PORTS",
                "31_OPEN_ITEMS", "32_GUIDELINES", "37_RESPONSIBILITY_MATRIX",
            ):
                self.assertIn(required, workbook.sheetnames)
            for name in ("21_NETWORK_PORTS", "32_GUIDELINES"):
                ws = workbook[name]
                self.assertEqual("00D97706", ws["A1"].fill.fgColor.rgb)
                self.assertEqual("001F4E78", ws["A2"].fill.fgColor.rgb)
                self.assertIsNone(ws["A3"].value)
                self.assertEqual("00A6A6A6", ws["A4"].fill.fgColor.rgb)
                self.assertTrue(all(ws.cell(row, 1).value in (None, "") for row in range(5, 20)))
            metadata = {
                str(row[0].value): str(row[1].value or "")
                for row in workbook["00_PROJECT_META"].iter_rows(min_row=5)
                if row[0].value
            }
            self.assertEqual("", metadata["Linked Project ID"])
            self.assertEqual(CANONICAL_REPOSITORY, metadata["Repository"])
            self.assertEqual("TBD", metadata["Project Revision"])
            self.assertEqual("1.0.0", metadata["Template Version"])
            blob = " ".join(str(cell.value or "") for ws in workbook for row in ws for cell in row)
            self.assertNotIn("SmartDraw", blob)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
