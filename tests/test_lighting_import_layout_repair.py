from __future__ import annotations

from copy import deepcopy
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage

from core.page_composer import compose_pages
from core.project_store import ProjectStore
from core.sheet_importer import import_workbook_sheets
from core.spreadsheet_layout import exact_source_layout, semantic_layout


def semantic_fixture() -> dict:
    rows, cols = 82, 34
    grid = [["" for _ in range(cols)] for _ in range(rows)]
    styles = {}
    for start, end, title in ((0, 11, "LIGHTING CONTROL SECTIONS"), (13, 28, "LIGHTING CONTROLS SCHEDULE")):
        grid[0][start] = title
        for col in range(start, end + 1):
            styles[f"0:{col}"] = {"fill": "#F4B183", "bold": True, "fontSize": 11}
            grid[1][col] = "Description" if col == start + 1 else f"IO{col - start + 1}"
            styles[f"1:{col}"] = {"fill": "#D9E1F2", "bold": True, "borders": {"bottom": {"style": "thin"}}}
        for row in range(2, 62 if start == 0 else 80):
            grid[row][start] = f"R{row:02d}"
            grid[row][start + 1] = "Long wrapped lighting control description for a generated disposable fixture"
            grid[row][start + 2] = f"DI-{row:02d}"
    for row, title in ((16, "LCP2"), (31, "LCP3"), (46, "LCP4")):
        grid[row][0] = title
        for col in range(12):
            styles[f"{row}:{col}"] = {"fill": "#F4B183", "bold": True, "fontSize": 11}
    # Formatting-only tails and an empty merged range must not inflate output.
    styles["81:33"] = {"fontName": "Arial"}
    return {
        "id": "fixture_excel",
        "type": "excelRange",
        "grid": grid,
        "styles": styles,
        "mergedCells": [
            {"startRow": 0, "startCol": 0, "endRow": 0, "endCol": 11},
            {"startRow": 0, "startCol": 13, "endRow": 0, "endCol": 28},
            {"startRow": 16, "startCol": 0, "endRow": 16, "endCol": 11},
            {"startRow": 31, "startCol": 0, "endRow": 31, "endCol": 11},
            {"startRow": 46, "startCol": 0, "endRow": 46, "endCol": 11},
            {"startRow": 81, "startCol": 30, "endRow": 81, "endCol": 33},
        ],
        "colWidths": [260 if col in {1, 14} else 70 for col in range(cols)],
        "rowHeights": [22 for _ in range(rows)],
        "srcRows": list(range(rows)),
        "sourceWorksheetId": "ws_fixture",
    }


class LightingImportLayoutRepairTests(unittest.TestCase):
    def test_exact_source_default_preserves_geometry_and_does_not_false_split(self) -> None:
        source = semantic_fixture()
        blocks, diagnostics = exact_source_layout(source, override="auto")
        self.assertEqual(diagnostics["selectedArrangement"], "source")
        self.assertEqual(diagnostics["blockCount"], 1)
        self.assertEqual(diagnostics["detectedColumnRanges"], [])
        self.assertTrue(diagnostics["sourceGeometryPreserved"])
        self.assertLess(diagnostics["effectiveRows"], diagnostics["rawRows"])
        self.assertLess(diagnostics["effectiveColumns"], diagnostics["rawColumns"])
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["colWidths"], source["colWidths"][:29])
        self.assertEqual(blocks[0]["rowHeights"], source["rowHeights"][:80])
        self.assertTrue(all(cell == "" for cell in (row[12] for row in blocks[0]["grid"])))
        self.assertNotIn("nowrapColumns", blocks[0])
        self.assertEqual(blocks[0]["styles"]["16:0"]["fill"], "#F4B183")

        pages = compose_pages([{
            "id": "page_fixture", "sheetCode": "EMS 12.0", "displaySheetCode": "EMS 12.0",
            "sheetTitle": "LIGHTING CONTROL IO", "pageGroupId": "page_fixture",
            "renderMode": "excel_exact", "layoutProfile": "exact_source_excel",
            "tableLayout": "source", "blocks": blocks, "include": True,
            "issueStatus": "draft", "publishStatus": "",
        }])
        self.assertGreater(len(pages), 1)
        self.assertEqual(pages[0]["id"], "page_fixture")
        self.assertTrue(all(not warning for page in pages for warning in (page.get("layoutWarnings") or [])))
        self.assertTrue(all(page.get("pageGroupId") == "page_fixture" for page in pages))
        self.assertTrue(all(page.get("issueStatus") == "draft" for page in pages))
        self.assertTrue(all(page["blocks"][0]["colWidths"] == blocks[0]["colWidths"] for page in pages))
        continuation_starts = [
            next(row for row in page["blocks"][0]["srcRows"] if row != 0)
            for page in pages[1:]
        ]
        self.assertTrue(all(start in {16, 31, 46, 62} or start > 62 for start in continuation_starts))

    def test_page_layout_overrides_are_deterministic(self) -> None:
        exact, exact_diag = exact_source_layout(semantic_fixture(), override="exact_source")
        legacy_auto, legacy_diag = exact_source_layout(semantic_fixture(), override="auto")
        two, two_diag = exact_source_layout(semantic_fixture(), override="two_columns")
        one_page, one_page_diag = exact_source_layout(semantic_fixture(), override="keep_one_page")
        self.assertEqual((len(exact), exact_diag["selectedArrangement"]), (1, "source"))
        self.assertEqual((len(legacy_auto), legacy_diag["selectedArrangement"]), (1, "source"))
        self.assertEqual((len(two), two_diag["selectedArrangement"]), (2, "side_by_side"))
        self.assertFalse(one_page[0]["allowContinuation"])
        self.assertEqual(one_page_diag["layoutOverride"], "keep_one_page")
        self.assertTrue(two_diag["manualOverride"])

    def test_import_preserves_authoritative_workbook_and_embedded_source_image(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            workbook = root / "Lighting Controls.xlsx"
            pixel = root / "fixture.png"
            PilImage.new("RGB", (8, 8), "orange").save(pixel)
            wb = Workbook()
            ws = wb.active
            ws.title = "LIGHTING CONTROL IO"
            ws.merge_cells("A1:L1"); ws["A1"] = "LIGHTING CONTROL SECTIONS"
            ws.merge_cells("N1:AC1"); ws["N1"] = "LIGHTING CONTROLS SCHEDULE"
            for cell in (*ws["A1:L1"][0], *ws["N1:AC1"][0]):
                cell.fill = __import__("openpyxl").styles.PatternFill("solid", fgColor="F4B183")
            for row in range(2, 48):
                ws.cell(row, 1, f"R{row}"); ws.cell(row, 2, "Long description"); ws.cell(row, 14, f"DI-{row}")
            ws["AH80"].font = __import__("openpyxl").styles.Font(name="Arial")
            ws.add_image(Image(pixel), "B4")
            wb.save(workbook)
            wb.close()
            project = {
                "id": "fixture", "metadata": {"projectName": "Fixture"},
                "sourceWorkbookName": "Authoritative.xlsx",
                "workbookSync": {"workbook": r"C:\authoritative\Authoritative.xlsx"},
                "worksheets": [], "sources": [],
                "pages": [{"id": "before", "sheetCode": "EMS 11.0", "displaySheetCode": "EMS 11.0", "sheetTitle": "Before"},
                          {"id": "after", "sheetCode": "EMS 13.0", "displaySheetCode": "EMS 13.0", "sheetTitle": "After"}],
            }
            updated, pages = import_workbook_sheets(
                deepcopy(project), workbook, ["LIGHTING CONTROL IO"], insert_after_page_id="before",
                assets_dir=root / "assets", asset_url_prefix="/fixture/assets", source_filename=workbook.name,
                source_sha256="1" * 64,
                project_local_path=f"sources/workbook/src_fixture_{workbook.name}",
            )
            self.assertEqual(updated["sourceWorkbookName"], "Authoritative.xlsx")
            self.assertEqual(updated["workbookSync"]["workbook"], r"C:\authoritative\Authoritative.xlsx")
            self.assertEqual(pages[0]["id"], pages[0]["pageGroupId"])
            self.assertEqual(pages[0]["sheetCode"], "EMS 12.0")
            self.assertEqual(pages[0]["issueStatus"], "draft")
            self.assertEqual(pages[0]["importedFrom"]["sourceFile"], workbook.name)
            self.assertEqual(pages[0]["sourceImport"]["sha256"], "1" * 64)
            self.assertEqual(
                pages[0]["sourceImport"]["projectLocalPath"],
                f"sources/workbook/src_fixture_{workbook.name}",
            )
            self.assertEqual(
                pages[0]["sourceImport"]["selectedWorksheet"],
                "LIGHTING CONTROL IO",
            )
            self.assertTrue(pages[0]["createdAt"])
            self.assertEqual(pages[0]["createdAt"], pages[0]["modifiedAt"])
            imported_ws = next(item for item in updated["worksheets"] if item["id"] == pages[0]["linkedWorksheetId"])
            self.assertEqual(len(imported_ws["embeddedImages"]), 1)
            self.assertTrue(any((root / "assets").iterdir()))

            replacement_input = deepcopy(updated)
            replacement_page = next(item for item in replacement_input["pages"] if item["id"] == pages[0]["id"])
            replacement_page["createdAt"] = "2025-01-02T03:04:05Z"
            replacement_page["modifiedAt"] = "2025-01-02T03:04:05Z"
            _, replacements = import_workbook_sheets(
                replacement_input, workbook, ["LIGHTING CONTROL IO"], replace_page_id=pages[0]["id"],
                assets_dir=root / "assets", asset_url_prefix="/fixture/assets", source_filename=workbook.name,
                source_sha256="2" * 64,
                project_local_path=f"sources/workbook/src_revised_{workbook.name}",
            )
            self.assertEqual(pages[0]["id"], replacements[0]["id"])
            self.assertEqual("2025-01-02T03:04:05Z", replacements[0]["createdAt"])
            self.assertNotEqual("2025-01-02T03:04:05Z", replacements[0]["modifiedAt"])
            self.assertEqual("2" * 64, replacements[0]["sourceImport"]["sha256"])

    def test_project_save_is_atomic_validated_and_keeps_recovery(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            store = ProjectStore(Path(raw))
            original = {"id": "fixture", "metadata": {"projectName": "Fixture"}, "pages": [{"id": "p1"}]}
            store.save("fixture", deepcopy(original))
            updated = deepcopy(original); updated["pages"].append({"id": "p2"})
            store.save("fixture", updated)
            self.assertEqual([page["id"] for page in store.load("fixture")["pages"]], ["p1", "p2"])
            backups = list((store.find_dir("fixture") / "backups").glob("project_*.json"))
            self.assertTrue(backups)
            self.assertFalse(list(store.find_dir("fixture").glob("*.tmp")))

    def test_project_history_rotation_archives_instead_of_deleting(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            store = ProjectStore(Path(raw)); store._MAX_BACKUPS = 2
            project = {"id": "fixture", "metadata": {"projectName": "Fixture"}, "pages": []}
            for revision in range(4):
                project["revision"] = revision
                store.save("fixture", deepcopy(project))
            backups = store.find_dir("fixture") / "backups"
            retained = list(backups.glob("project_*.json"))
            archived = list((backups / "project_json_archive").glob("project_*.json"))
            self.assertEqual(len(retained), 2)
            self.assertEqual(len(archived), 1)

if __name__ == "__main__":
    unittest.main()
