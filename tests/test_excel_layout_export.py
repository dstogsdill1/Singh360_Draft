from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from core.worksheet_export import export_excel_layout_page_xlsx


def layout_page() -> dict:
    def table(table_id: str, y: int, widths: list[int], title: str) -> dict:
        return {
            "id": table_id, "x": 80, "y": y, "width": 1000, "height": 150,
            "rows": [["POINT", "DESCRIPTION", "VALUE"], ["P-1", "Neutral test value", "1"]],
            "columnWidths": widths, "rowHeights": [26, 24], "merges": [], "title": title,
            "titleStyle": {"fill": "#F4B183", "bold": True, "align": "center"},
            "headerStyle": {"fill": "#D9EAF7", "bold": True, "align": "center"},
            "bodyStyle": {"fill": "#FFFFFF", "wrap": True},
            "keepTogether": True, "splitRows": True,
            "repeatTitle": True, "repeatHeaders": True,
        }
    return {
        "id": "layout-page", "sheetTab": "EMS TEST", "include": True,
        "excelLayout": {
            "version": 1, "pageWidth": 1632, "pageHeight": 1056,
            "printableMargin": 48, "snapSize": 8, "tabColor": "#00AA77",
            "tables": [
                table("a", 90, [180, 600, 220], "UPPER SCHEDULE"),
                table("b", 360, [450, 250, 300], "LOWER SCHEDULE"),
            ],
        },
    }


class ExcelLayoutExportTests(unittest.TestCase):
    def test_real_cells_titles_geometry_and_print_setup(self) -> None:
        page = layout_page()
        page["excelLayout"]["tables"][1]["y"] = 1120
        wb = load_workbook(BytesIO(export_excel_layout_page_xlsx(page)))
        ws = wb["EMS TEST"]
        self.assertEqual(ws.page_setup.orientation, "landscape")
        self.assertEqual(str(ws.page_setup.paperSize), str(ws.PAPERSIZE_TABLOID))
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertTrue(ws.print_area)
        self.assertEqual(len(ws.row_breaks.brk), 1)
        self.assertEqual(ws.sheet_properties.tabColor.rgb[-6:], "00AA77")
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
        self.assertIn("UPPER SCHEDULE", values)
        self.assertIn("LOWER SCHEDULE", values)
        self.assertIn("Neutral test value", values)
        merged = {str(item) for item in ws.merged_cells.ranges}
        self.assertGreater(len(merged), 6)
        title_cells = [cell for row in ws.iter_rows() for cell in row if cell.value in {"UPPER SCHEDULE", "LOWER SCHEDULE"}]
        self.assertTrue(all(cell.coordinate in {str(rng).split(":")[0] for rng in merged} for cell in title_cells))
        # Unrelated table geometry produces different first-cell merge spans.
        point_cells = [cell for row in ws.iter_rows() for cell in row if cell.value == "POINT"]
        self.assertEqual(len(point_cells), 2)
        spans = []
        for cell in point_cells:
            spans.append(next(str(rng) for rng in ws.merged_cells.ranges if rng.min_row == cell.row and rng.min_col == cell.column))
        self.assertNotEqual(spans[0].split(":")[1][0], spans[1].split(":")[1][0])


if __name__ == "__main__":
    unittest.main()
