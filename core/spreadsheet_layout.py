"""Source-faithful spreadsheet range layout for ANSI-B drawing pages.

The source worksheet remains authoritative.  This module derives printable
blocks and diagnostics without changing source cells, formulas, or geometry.
Automatic layout keeps the source as one uniformly scaled region.  The older
semantic column splitter remains available only for an explicit Two Columns
override.
"""
from __future__ import annotations

from copy import deepcopy
import re
from typing import Any

from openpyxl.utils.cell import get_column_letter


# Keep a small measurement margin above the 7.5pt acceptance floor because
# Chromium's font metrics and PDF point rounding can render ~0.2pt below the
# nominal CSS transform.
MIN_BODY_FONT_PT = 7.8
TARGET_BODY_FONT_PT = 9.0
TWO_COLUMN_NATURAL_WIDTH = 650.0
BLANK_ROW_HEIGHT = 8.0
DEFAULT_ROW_HEIGHT = 21.0
MIN_COLUMN_WIDTH = 26.0
# ExcelRangeRenderer's actual vertical fit area after the 64px page band,
# padding, and title-block safety gap (700 - 20 - 64 - 20).
SEMANTIC_SAFE_BODY_HEIGHT = 596.0


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _natural_size(block: dict[str, Any]) -> tuple[float, float]:
    grid = block.get("grid") or []
    widths = block.get("colWidths") or []
    heights = block.get("rowHeights") or []
    cols = max((len(row) for row in grid), default=0)
    return (
        sum(float(widths[c] if c < len(widths) else 64.0) for c in range(cols)),
        sum(float(heights[r] if r < len(heights) else DEFAULT_ROW_HEIGHT) for r in range(len(grid))),
    )


def effective_used_bounds(block: dict[str, Any]) -> dict[str, int]:
    """Return raw/effective dimensions, ignoring format-only outer tails."""
    grid = block.get("grid") or []
    raw_rows = len(grid)
    raw_cols = max((len(row) for row in grid), default=0)
    occupied = [
        (r, c)
        for r, row in enumerate(grid)
        for c, value in enumerate(row)
        if _text(value)
    ]
    if not occupied:
        return {
            "startRow": 0,
            "startCol": 0,
            "endRow": max(0, raw_rows - 1),
            "endCol": max(0, raw_cols - 1),
            "rawRows": raw_rows,
            "rawCols": raw_cols,
            "effectiveRows": raw_rows,
            "effectiveCols": raw_cols,
        }

    min_row = min(r for r, _ in occupied)
    max_row = max(r for r, _ in occupied)
    min_col = min(c for _, c in occupied)
    max_col = max(c for _, c in occupied)

    # A merge is meaningful only when its anchor/value lies in the occupied
    # content envelope. Empty decorative merges outside it do not expand output.
    for merge in block.get("mergedCells") or []:
        anchor = (
            int(merge.get("startRow", 0)),
            int(merge.get("startCol", 0)),
        )
        if anchor in occupied:
            max_row = max(max_row, int(merge.get("endRow", max_row)))
            max_col = max(max_col, int(merge.get("endCol", max_col)))

    return {
        "startRow": min_row,
        "startCol": min_col,
        "endRow": max_row,
        "endCol": max_col,
        "rawRows": raw_rows,
        "rawCols": raw_cols,
        "effectiveRows": max_row - min_row + 1,
        "effectiveCols": max_col - min_col + 1,
    }


def crop_excel_block(
    block: dict[str, Any],
    *,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    block_id: str,
) -> dict[str, Any]:
    """Crop a normalized r:c-style Excel block without altering source data."""
    out = deepcopy(block)
    grid = block.get("grid") or []
    out["id"] = block_id
    out["grid"] = [
        list(row[start_col : end_col + 1])
        for row in grid[start_row : end_row + 1]
    ]
    out["colWidths"] = list((block.get("colWidths") or [])[start_col : end_col + 1])
    out["rowHeights"] = list((block.get("rowHeights") or [])[start_row : end_row + 1])
    out["srcRows"] = list((block.get("srcRows") or list(range(len(grid))))[start_row : end_row + 1])

    styles: dict[str, Any] = {}
    for key, style in (block.get("styles") or {}).items():
        try:
            raw_row, raw_col = str(key).split(":", 1)
            row, col = int(raw_row), int(raw_col)
        except (TypeError, ValueError):
            continue
        if start_row <= row <= end_row and start_col <= col <= end_col:
            styles[f"{row - start_row}:{col - start_col}"] = deepcopy(style)
    out["styles"] = styles

    merges: list[dict[str, int]] = []
    for merge in block.get("mergedCells") or []:
        sr = int(merge.get("startRow", 0)); er = int(merge.get("endRow", sr))
        sc = int(merge.get("startCol", 0)); ec = int(merge.get("endCol", sc))
        if sr >= start_row and er <= end_row and sc >= start_col and ec <= end_col:
            merges.append({
                "startRow": sr - start_row,
                "endRow": er - start_row,
                "startCol": sc - start_col,
                "endCol": ec - start_col,
            })
    out["mergedCells"] = merges

    repeat = [
        row - start_row
        for row in (block.get("repeatRows") or [])
        if start_row <= row <= end_row
    ]
    out["repeatRows"] = repeat
    out["headerRowCount"] = len(repeat)
    out["sourceRange"] = (
        f"{get_column_letter(start_col + 1)}{start_row + 1}:"
        f"{get_column_letter(end_col + 1)}{end_row + 1}"
    )
    return out


def _style_has_visual_bridge(style: dict[str, Any] | None) -> bool:
    if not style:
        return False
    return bool(style.get("fill") or style.get("borders"))


def _has_header_role(block: dict[str, Any], start_col: int, end_col: int) -> bool:
    grid = block.get("grid") or []
    styles = block.get("styles") or {}
    header_limit = min(6, len(grid))
    for row in range(header_limit):
        populated = sum(
            1 for col in range(start_col, end_col + 1)
            if col < len(grid[row]) and _text(grid[row][col])
        )
        styled = sum(
            1 for col in range(start_col, end_col + 1)
            if _style_has_visual_bridge(styles.get(f"{row}:{col}"))
        )
        if populated and styled:
            return True
    return False


def detect_independent_column_blocks(block: dict[str, Any]) -> list[tuple[int, int]]:
    """Detect top-level horizontal regions separated by a real blank gap."""
    grid = block.get("grid") or []
    n_cols = max((len(row) for row in grid), default=0)
    if n_cols < 3:
        return [(0, max(0, n_cols - 1))]
    styles = block.get("styles") or {}
    populated_columns = {
        col for col in range(n_cols)
        if any(col < len(row) and _text(row[col]) for row in grid)
    }
    if not populated_columns:
        return [(0, n_cols - 1)]

    first, last = min(populated_columns), max(populated_columns)
    blank_segments: list[tuple[int, int]] = []
    start: int | None = None
    for col in range(first, last + 1):
        blank = col not in populated_columns
        visually_bridged = any(
            _style_has_visual_bridge(styles.get(f"{row}:{col}"))
            for row in range(len(grid))
        )
        is_gap = blank and not visually_bridged
        if is_gap and start is None:
            start = col
        elif not is_gap and start is not None:
            blank_segments.append((start, col - 1))
            start = None
    if start is not None:
        blank_segments.append((start, last))

    boundaries = [first]
    ends: list[int] = []
    for gap_start, gap_end in blank_segments:
        left_start = boundaries[-1]
        left_end = gap_start - 1
        right_start = gap_end + 1
        if left_end < left_start or right_start > last:
            continue
        if not _has_header_role(block, left_start, left_end):
            continue
        if not _has_header_role(block, right_start, last):
            continue
        if not any(
            c < len(row) and _text(row[c])
            for row in grid[1:]
            for c in range(left_start, left_end + 1)
        ):
            continue
        if not any(
            c < len(row) and _text(row[c])
            for row in grid[1:]
            for c in range(right_start, last + 1)
        ):
            continue
        ends.append(left_end)
        boundaries.append(right_start)

    if not ends:
        return [(first, last)]
    ranges = [(boundaries[i], ends[i]) for i in range(len(ends))]
    ranges.append((boundaries[-1], last))
    return ranges


_SHORT_CODE_HEADER = re.compile(
    r"(^|\b)(ro|di|ti|aio|io|id|type|poles?|from|to|offset|control|contactor)(#|\b)",
    re.I,
)
_PART_OR_CODE = re.compile(r"^[A-Z0-9][A-Z0-9 ._+/#:-]{0,17}$", re.I)


def _fit_widths(block: dict[str, Any]) -> tuple[list[float], set[int]]:
    grid = block.get("grid") or []
    n_cols = max((len(row) for row in grid), default=0)
    widths: list[float] = []
    nowrap: set[int] = set()
    for col in range(n_cols):
        values = [_text(row[col]) for row in grid if col < len(row) and _text(row[col])]
        # Only the first populated label is a column header. Including the
        # first data value made a description containing words like "control"
        # incorrectly force the entire description column to no-wrap.
        header = values[0] if values else ""
        max_chars = max((len(value) for value in values), default=1)
        short_values = bool(values) and all(
            len(value) <= 18 and _PART_OR_CODE.fullmatch(value)
            for value in values[1:] or values
        )
        if _SHORT_CODE_HEADER.search(header) or short_values:
            width = max(30.0, min(76.0, max_chars * 5.5 + 12.0))
            nowrap.add(col)
        elif max_chars > 22:
            width = max(88.0, min(150.0, min(max_chars, 34) * 4.0 + 14.0))
        else:
            width = max(42.0, min(108.0, max_chars * 5.4 + 14.0))
        widths.append(width)

    total = sum(widths)
    if total > TWO_COLUMN_NATURAL_WIDTH:
        reducible = sum(max(0.0, width - MIN_COLUMN_WIDTH) for width in widths)
        reduction = total - TWO_COLUMN_NATURAL_WIDTH
        if reducible > 0:
            ratio = min(1.0, reduction / reducible)
            widths = [
                max(MIN_COLUMN_WIDTH, width - (width - MIN_COLUMN_WIDTH) * ratio)
                for width in widths
            ]
    return widths, nowrap


def autofit_excel_block(block: dict[str, Any]) -> dict[str, Any]:
    """Compact columns/rows while maintaining a 7.5pt rendered floor."""
    out = deepcopy(block)
    grid = out.get("grid") or []
    widths, nowrap = _fit_widths(out)
    out["colWidths"] = [round(width, 2) for width in widths]
    out["nowrapColumns"] = sorted(nowrap)
    out["minScale"] = MIN_BODY_FONT_PT / TARGET_BODY_FONT_PT
    out["allowContinuation"] = True
    out["splitMode"] = "auto_rows"
    out["scaleMode"] = "fit_body"
    out["trimBlankRows"] = True
    out["trimBlankColumns"] = True
    out["renderProfile"] = "semantic_excel_table"
    out["safeBodyHeight"] = SEMANTIC_SAFE_BODY_HEIGHT
    # Browser table layout can grow beyond explicit row minima because CSS
    # wraps complete words. Reserve that measured growth during pagination.
    out["heightSafetyFactor"] = 1.05

    styles = deepcopy(out.get("styles") or {})
    for row, values in enumerate(grid):
        for col, value in enumerate(values):
            if not _text(value):
                continue
            key = f"{row}:{col}"
            style = deepcopy(styles.get(key) or {})
            style["fontSize"] = max(float(style.get("fontSize") or TARGET_BODY_FONT_PT), TARGET_BODY_FONT_PT)
            if col not in nowrap and len(_text(value)) > 18:
                style["wrap"] = True
            elif col in nowrap:
                style["wrap"] = False
            styles[key] = style
    out["styles"] = styles

    heights: list[float] = []
    old_heights = out.get("rowHeights") or []

    def wrapped_lines(text: str, capacity: int) -> int:
        """Conservative CSS pre-wrap line count using complete words."""
        total = 0
        for paragraph in text.splitlines() or [text]:
            words = paragraph.split() or [""]
            lines = 1
            used = 0
            for word in words:
                needed = len(word) if used == 0 else len(word) + 1
                if used and used + needed > capacity:
                    lines += 1
                    used = len(word)
                else:
                    used += needed
            total += lines
        return max(1, total)

    for row, values in enumerate(grid):
        if not any(_text(value) for value in values):
            heights.append(BLANK_ROW_HEIGHT)
            continue
        required_height = 15.0
        for col, value in enumerate(values):
            text = _text(value)
            if not text or col in nowrap or col >= len(widths):
                continue
            if len(text) > 18:
                style = styles.get(f"{row}:{col}") or {}
                font_px = max(TARGET_BODY_FONT_PT, float(style.get("fontSize") or TARGET_BODY_FONT_PT)) * 4.0 / 3.0
                chars_per_line = max(4, int((widths[col] - 8) / max(6.0, font_px * 0.62)))
                lines = wrapped_lines(text, chars_per_line)
                required_height = max(required_height, 6.0 + lines * font_px * 1.15)
        source_height = float(old_heights[row] if row < len(old_heights) else DEFAULT_ROW_HEIGHT)
        heights.append(round(max(source_height, required_height), 2))
    out["rowHeights"] = heights
    natural_w, natural_h = _natural_size(out)
    out["layoutMetrics"] = {
        "naturalWidth": round(natural_w, 2),
        "naturalHeight": round(natural_h, 2),
        "minimumBodyFontPt": MIN_BODY_FONT_PT,
        "targetBodyFontPt": TARGET_BODY_FONT_PT,
    }
    return out


def canonical_layout_override(value: str | None) -> str:
    """Normalize legacy saved values without preserving their old reflow."""
    normalized = (value or "exact_source").strip().casefold().replace("-", "_")
    if normalized in {"auto", "exact", "exact_source", "one_column", "single", "continue_blocks", "stacked"}:
        return "exact_source"
    if normalized in {"two_columns", "side_by_side"}:
        return "two_columns"
    if normalized in {"keep_one_page", "one_page", "fit_one_page"}:
        return "keep_one_page"
    raise ValueError(f"Unknown Excel layout override: {value}")


def exact_source_layout(
    full_block: dict[str, Any],
    *,
    override: str = "exact_source",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Keep source geometry intact unless Two Columns is explicitly chosen.

    Internal blank rows/columns, widths, heights, merges, styles, and relative
    positions remain unchanged.  Only format-only outer tails outside an
    explicit print area are removed.
    """
    canonical = canonical_layout_override(override)
    bounds = effective_used_bounds(full_block)
    has_print_area = bool(str(full_block.get("printArea") or "").strip())
    if has_print_area:
        exact = deepcopy(full_block)
    else:
        exact = crop_excel_block(
            full_block,
            start_row=bounds["startRow"],
            end_row=bounds["endRow"],
            start_col=bounds["startCol"],
            end_col=bounds["endCol"],
            block_id=f"{full_block.get('id', 'excel')}_source",
        )

    if canonical == "two_columns":
        blocks, diagnostics = semantic_layout(exact, override="two_columns")
        diagnostics["layoutOverride"] = canonical
        diagnostics["sourceGeometryPreserved"] = False
        diagnostics["manualOverride"] = True
        return blocks, diagnostics

    exact.update({
        "allowContinuation": canonical != "keep_one_page",
        "splitMode": "auto_rows" if canonical != "keep_one_page" else "none",
        "scaleMode": "fit_body",
        "trimBlankRows": False,
        "trimBlankColumns": False,
        "renderProfile": "exact_source_excel",
        "layoutProfile": "exact_source_excel",
    })
    natural_w, natural_h = _natural_size(exact)
    diagnostics = {
        "rawUsedRange": f"{bounds['rawRows']}x{bounds['rawCols']}",
        "effectiveUsedRange": f"{bounds['effectiveRows']}x{bounds['effectiveCols']}",
        "rawRows": bounds["rawRows"],
        "rawColumns": bounds["rawCols"],
        "effectiveRows": bounds["effectiveRows"],
        "effectiveColumns": bounds["effectiveCols"],
        "blockCount": 1,
        "detectedColumnRanges": [],
        "selectedArrangement": "source",
        "layoutOverride": canonical,
        "sourceGeometryPreserved": True,
        "manualOverride": canonical == "keep_one_page",
        "naturalWidth": round(natural_w, 2),
        "naturalHeight": round(natural_h, 2),
        "overflowCount": 0,
        "clippedCellCount": 0,
        "titleBlockOverlapCount": 0,
    }
    return [exact], diagnostics


def semantic_layout(
    full_block: dict[str, Any],
    *,
    override: str = "auto",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Return printable blocks plus machine-readable layout diagnostics."""
    override = (override or "auto").strip().casefold().replace("-", "_")
    bounds = effective_used_bounds(full_block)
    effective = crop_excel_block(
        full_block,
        start_row=bounds["startRow"],
        end_row=bounds["endRow"],
        start_col=bounds["startCol"],
        end_col=bounds["endCol"],
        block_id=f"{full_block.get('id', 'excel')}_effective",
    )
    detected = detect_independent_column_blocks(effective)
    if override in {"one_column", "single"}:
        ranges = [(0, max((len(row) for row in effective.get("grid") or []), default=1) - 1)]
        arrangement = "single"
    elif override in {"two_columns", "side_by_side"} and len(detected) >= 2:
        ranges = detected[:2]
        arrangement = "side_by_side"
    elif override in {"continue_blocks", "stacked"}:
        ranges = detected
        arrangement = "stacked"
    elif len(detected) == 2:
        ranges = detected
        arrangement = "side_by_side"
    elif len(detected) > 2:
        ranges = detected
        arrangement = "stacked"
    else:
        ranges = detected
        arrangement = "single"

    blocks: list[dict[str, Any]] = []
    for index, (start_col, end_col) in enumerate(ranges):
        block_rows = [
            row for row, values in enumerate(effective.get("grid") or [])
            if any(_text(values[col]) for col in range(start_col, min(end_col + 1, len(values))))
        ]
        end_row = max(block_rows) if block_rows else len(effective.get("grid") or []) - 1
        cropped = crop_excel_block(
            effective,
            start_row=0,
            end_row=max(0, end_row),
            start_col=start_col,
            end_col=end_col,
            block_id=f"{full_block.get('id', 'excel')}_region_{index + 1}",
        )
        cropped["layoutColumn"] = index
        cropped["preferredColumnShare"] = round(1.0 / max(1, len(ranges)), 3)
        blocks.append(autofit_excel_block(cropped))

    diagnostics = {
        "rawUsedRange": f"{bounds['rawRows']}x{bounds['rawCols']}",
        "effectiveUsedRange": f"{bounds['effectiveRows']}x{bounds['effectiveCols']}",
        "rawRows": bounds["rawRows"],
        "rawColumns": bounds["rawCols"],
        "effectiveRows": bounds["effectiveRows"],
        "effectiveColumns": bounds["effectiveCols"],
        "blockCount": len(blocks),
        "detectedColumnRanges": [
            f"{get_column_letter(start + 1)}:{get_column_letter(end + 1)}"
            for start, end in detected
        ],
        "selectedArrangement": arrangement,
        "layoutOverride": override,
        "minimumRenderedFontPt": MIN_BODY_FONT_PT,
        "overflowCount": 0,
        "clippedCellCount": 0,
        "titleBlockOverlapCount": 0,
    }
    return blocks, diagnostics
