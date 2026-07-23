from __future__ import annotations

from copy import copy
from datetime import datetime, timezone
from hashlib import sha1
import json
import os
from pathlib import Path
import shutil
import tempfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties


REQUIRED_HEADERS = [
    "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
    "Page Type", "Notes", "Render Profile", "Split Mode", "Page ID",
    "Parent Page ID", "Issue Status", "Source Mode", "Sync Direction",
    "Last Sync UTC", "Workbook Hash", "App Hash",
]
CONTROL_SHEETS = ["00_PROJECT_META", "00_INDEX", "00_HELP"]
STATUS_COLORS = {
    "draft": "F28C28",
    "draft confirmed": "76B852",
    "public": "2D7DD2",
    "public confirmed": "14845A",
    "excluded": "9AA3AB",
    "control": "252C34",
    "help": "276FA8",
}


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _stable_id(tab: str) -> str:
    return "S360-" + sha1(tab.encode("utf-8")).hexdigest()[:12].upper()


def _header_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, max(ws.max_column or 1, len(REQUIRED_HEADERS)) + 1):
        value = str(ws.cell(4, col).value or "").strip()
        if value:
            result[value.casefold()] = col
    return result


def _iter_index_rows(ws, headers: dict[str, int]):
    tab_col = headers.get("sheet tab")
    if not tab_col:
        return
    for row in range(5, (ws.max_row or 4) + 1):
        tab = str(ws.cell(row, tab_col).value or "").strip()
        if tab:
            yield row, tab


def _formula_errors(wb, limit: int = 200) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "e" or (
                    isinstance(value, str)
                    and value.upper().startswith(("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A"))
                ):
                    errors.append({"sheet": ws.title, "cell": cell.coordinate, "value": str(value)})
                    if len(errors) >= limit:
                        return errors
    return errors


def audit_workbook(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=False, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        issues: list[dict[str, Any]] = []
        missing_control = [name for name in CONTROL_SHEETS[:2] if name not in wb.sheetnames]
        if missing_control:
            issues.append({"severity": "critical", "code": "missing_control", "message": f"Missing control sheets: {', '.join(missing_control)}"})

        indexed_tabs: list[str] = []
        duplicate_codes: list[str] = []
        duplicate_orders: list[str] = []
        duplicate_ids: list[str] = []
        missing_tabs: list[str] = []
        invalid_include: list[str] = []
        invalid_status: list[str] = []
        missing_headers: list[str] = []
        header_map: dict[str, int] = {}
        index_rows = 0

        if "00_INDEX" in wb.sheetnames:
            ws = wb["00_INDEX"]
            header_map = _header_map(ws)
            missing_headers = [header for header in REQUIRED_HEADERS if header.casefold() not in header_map]
            if missing_headers:
                issues.append({"severity": "error", "code": "missing_headers", "message": f"00_INDEX is missing {len(missing_headers)} required columns.", "items": missing_headers})

            seen_codes: set[str] = set()
            seen_orders: set[str] = set()
            seen_ids: set[str] = set()
            for row, tab in _iter_index_rows(ws, header_map) or []:
                index_rows += 1
                indexed_tabs.append(tab)
                code = str(ws.cell(row, header_map.get("sheet code", 3)).value or "").strip()
                order = str(ws.cell(row, header_map.get("order", 2)).value or "").strip()
                page_id = str(ws.cell(row, header_map.get("page id", 11)).value or "").strip()
                include = str(ws.cell(row, header_map.get("include", 1)).value or "").strip().upper()
                status = str(ws.cell(row, header_map.get("issue status", 13)).value or "").strip().casefold()

                if code and code in seen_codes:
                    duplicate_codes.append(code)
                seen_codes.add(code)
                if order and order in seen_orders:
                    duplicate_orders.append(order)
                seen_orders.add(order)
                if page_id and page_id in seen_ids:
                    duplicate_ids.append(page_id)
                seen_ids.add(page_id)
                if tab not in wb.sheetnames:
                    missing_tabs.append(tab)
                if include not in {"YES", "NO"}:
                    invalid_include.append(tab)
                if status not in {"draft", "draft confirmed", "public", "public confirmed"}:
                    invalid_status.append(tab)

        if duplicate_codes:
            issues.append({"severity": "warning", "code": "duplicate_codes", "message": "Duplicate sheet codes were found.", "items": sorted(set(duplicate_codes))})
        if duplicate_orders:
            issues.append({"severity": "warning", "code": "duplicate_orders", "message": "Duplicate page-order values were found.", "items": sorted(set(duplicate_orders))})
        if duplicate_ids:
            issues.append({"severity": "error", "code": "duplicate_page_ids", "message": "Duplicate permanent Page IDs were found.", "items": sorted(set(duplicate_ids))})
        if missing_tabs:
            issues.append({"severity": "error", "code": "missing_tabs", "message": "00_INDEX refers to workbook tabs that do not exist.", "items": missing_tabs})
        if invalid_include:
            issues.append({"severity": "warning", "code": "invalid_include", "message": "Rows contain invalid Include values.", "items": invalid_include})
        if invalid_status:
            issues.append({"severity": "warning", "code": "invalid_status", "message": "Rows contain invalid Issue Status values.", "items": invalid_status})

        unindexed = [
            name for name in wb.sheetnames
            if name not in CONTROL_SHEETS and name not in indexed_tabs
        ]
        if unindexed:
            issues.append({"severity": "info", "code": "unindexed_sheets", "message": f"{len(unindexed)} workbook sheets are not represented in 00_INDEX.", "items": unindexed[:100]})

        errors = _formula_errors(wb)
        if errors:
            issues.append({"severity": "error", "code": "formula_errors", "message": f"{len(errors)} formula/error cells require review.", "items": errors[:50]})

        font_styles: dict[str, int] = {}
        for ws in wb.worksheets:
            if ws.title in CONTROL_SHEETS:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in (None, ""):
                        continue
                    key = f"{cell.font.name or 'None'} / {cell.font.sz or 'None'}"
                    font_styles[key] = font_styles.get(key, 0) + 1
        if len(font_styles) > 8:
            issues.append({
                "severity": "info",
                "code": "font_variation",
                "message": f"{len(font_styles)} font name/size combinations were found. Strict formatting can normalize indexed table sheets.",
                "items": [{"style": key, "cells": value} for key, value in sorted(font_styles.items(), key=lambda item: -item[1])[:20]],
            })

        counts = {
            "sheets": len(wb.sheetnames),
            "indexRows": index_rows,
            "includedRows": 0,
            "excludedRows": 0,
            "unindexedSheets": len(unindexed),
            "formulaErrors": len(errors),
            "critical": sum(1 for issue in issues if issue["severity"] == "critical"),
            "errors": sum(1 for issue in issues if issue["severity"] == "error"),
            "warnings": sum(1 for issue in issues if issue["severity"] == "warning"),
        }
        if "00_INDEX" in wb.sheetnames and header_map:
            include_col = header_map.get("include", 1)
            for row, _tab in _iter_index_rows(wb["00_INDEX"], header_map) or []:
                include = str(wb["00_INDEX"].cell(row, include_col).value or "").strip().upper()
                if include == "YES":
                    counts["includedRows"] += 1
                else:
                    counts["excludedRows"] += 1

        return {
            "ok": True,
            "path": str(path),
            "filename": path.name,
            "modifiedUtc": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "counts": counts,
            "issues": issues,
            "safeRepairAvailable": True,
            "strictRepairAvailable": True,
        }
    finally:
        wb.close()


def _backup(path: Path, store: Any, project_id: str) -> Path:
    root = store.docs / "backups" / "workbook_quality" / project_id
    root.mkdir(parents=True, exist_ok=True)
    target = root / f"{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}_{path.name}"
    shutil.copy2(path, target)
    backups = sorted(root.glob("*"), key=lambda item: item.stat().st_mtime, reverse=True)
    for stale in backups[20:]:
        stale.unlink(missing_ok=True)
    return target


def _ensure_help(wb) -> None:
    if "00_HELP" not in wb.sheetnames:
        ws = wb.create_sheet("00_HELP", min(2, len(wb.sheetnames)))
        ws["A1"] = "SINGH360 DRAFT — QUICK HELP"
        ws.merge_cells("A1:J2")
        ws["A3"] = "Open the app Help Center for the current workflow instructions."
        ws.merge_cells("A3:J3")
    wb["00_HELP"].sheet_properties.tabColor = STATUS_COLORS["help"]


def _ensure_index(wb):
    if "00_PROJECT_META" not in wb.sheetnames:
        wb.create_sheet("00_PROJECT_META", 0)
    if "00_INDEX" not in wb.sheetnames:
        wb.create_sheet("00_INDEX", 1)
    _ensure_help(wb)

    for name in ("00_PROJECT_META", "00_INDEX"):
        wb[name].sheet_properties.tabColor = STATUS_COLORS["control"]

    ws = wb["00_INDEX"]
    headers = _header_map(ws)
    for col, header in enumerate(REQUIRED_HEADERS, start=1):
        if header.casefold() not in headers:
            ws.cell(4, col, header)
    headers = _header_map(ws)
    return ws, headers


def _copy_font(cell, *, name: str | None = None, size: float | None = None):
    current = cell.font
    cell.font = Font(
        name=name or current.name,
        sz=size or current.sz,
        b=current.b,
        i=current.i,
        vertAlign=current.vertAlign,
        u=current.u,
        strike=current.strike,
        color=copy(current.color),
        outline=current.outline,
        shadow=current.shadow,
        condense=current.condense,
        extend=current.extend,
        family=current.family,
        charset=current.charset,
        scheme=current.scheme,
    )


def repair_workbook(path: Path, project_id: str, store: Any, mode: str = "safe") -> dict[str, Any]:
    if mode not in {"safe", "strict"}:
        raise ValueError("Repair mode must be safe or strict.")
    backup = _backup(path, store, project_id)
    wb = load_workbook(path, read_only=False, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    changes: list[str] = []
    try:
        ws, headers = _ensure_index(wb)
        changes.append("Verified 00_PROJECT_META, 00_INDEX, and 00_HELP.")

        rows_by_tab: dict[str, int] = {}
        max_row = max(ws.max_row or 4, 4)
        for row, tab in _iter_index_rows(ws, headers) or []:
            rows_by_tab[tab] = row

        # Register every physical worksheet without publishing it automatically.
        for sheet_name in wb.sheetnames:
            if sheet_name in CONTROL_SHEETS or sheet_name in rows_by_tab:
                continue
            max_row += 1
            row = max_row
            ws.cell(row, headers["include"], "NO")
            ws.cell(row, headers["order"], row - 4)
            ws.cell(row, headers["sheet code"], "TBD")
            ws.cell(row, headers["sheet tab"], sheet_name)
            ws.cell(row, headers["page title"], sheet_name)
            ws.cell(row, headers["family"], "Unclassified")
            ws.cell(row, headers["page type"], "data-grid")
            ws.cell(row, headers["notes"], "Automatically registered by Workbook Inspector. Verify before publishing.")
            ws.cell(row, headers["page id"], _stable_id(sheet_name))
            ws.cell(row, headers["issue status"], "Draft")
            ws.cell(row, headers["source mode"], "Workbook")
            ws.cell(row, headers["sync direction"], "Both")
            rows_by_tab[sheet_name] = row
            changes.append(f"Registered unindexed sheet: {sheet_name}")

        # Normalize manifest fields and sequential physical order.
        ordered_tabs: list[str] = []
        for sequence, (row, tab) in enumerate(_iter_index_rows(ws, headers) or [], start=1):
            ordered_tabs.append(tab)
            include = str(ws.cell(row, headers["include"]).value or "").strip().upper()
            if include not in {"YES", "NO"}:
                include = "NO"
                ws.cell(row, headers["include"], include)
            ws.cell(row, headers["order"], sequence)
            if not str(ws.cell(row, headers["page id"]).value or "").strip():
                ws.cell(row, headers["page id"], _stable_id(tab))
            status = str(ws.cell(row, headers["issue status"]).value or "").strip().casefold()
            if status not in {"draft", "draft confirmed", "public", "public confirmed"}:
                ws.cell(row, headers["issue status"], "Draft")
                status = "draft"
            if not str(ws.cell(row, headers["source mode"]).value or "").strip():
                ws.cell(row, headers["source mode"], "Workbook")
            if not str(ws.cell(row, headers["sync direction"]).value or "").strip():
                ws.cell(row, headers["sync direction"], "Both")

            if tab in wb.sheetnames:
                color = STATUS_COLORS["excluded"] if include != "YES" else STATUS_COLORS.get(status, STATUS_COLORS["draft"])
                wb[tab].sheet_properties.tabColor = color
            else:
                ws.cell(row, headers["include"], "NO")
                note_col = headers["notes"]
                note = str(ws.cell(row, note_col).value or "").strip()
                marker = "VERIFY: referenced workbook tab is missing."
                if marker not in note:
                    ws.cell(row, note_col, (note + " " + marker).strip())

        # Physical tab order follows 00_INDEX; unmatched tabs stay at the end.
        controls = [name for name in CONTROL_SHEETS if name in wb.sheetnames]
        indexed = [name for name in ordered_tabs if name in wb.sheetnames and name not in controls]
        remaining = [name for name in wb.sheetnames if name not in controls and name not in indexed]
        wb._sheets = [wb[name] for name in controls + indexed + remaining]
        changes.append("Reordered workbook tabs to match 00_INDEX.")

        # Control/index formatting.
        dark = PatternFill("solid", fgColor="252C34")
        orange = PatternFill("solid", fgColor="F28C28")
        gray = PatternFill("solid", fgColor="D0D4D8")
        thin = Side(style="thin", color="A7AFB5")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, header in enumerate(REQUIRED_HEADERS, start=1):
            cell = ws.cell(4, col)
            cell.value = header
            cell.fill = gray
            cell.font = Font(name="Arial", size=9, bold=True, color="20262D")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A5"
        ws.sheet_properties.tabColor = STATUS_COLORS["control"]
        wb["00_PROJECT_META"].sheet_properties.tabColor = STATUS_COLORS["control"]
        wb["00_HELP"].sheet_properties.tabColor = STATUS_COLORS["help"]

        widths = [11, 8, 14, 34, 38, 24, 20, 40, 20, 15, 22, 22, 20, 20, 15, 22, 20, 20]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(4, col).column_letter].width = width

        status_letter = ws.cell(4, headers["issue status"]).column_letter
        include_letter = ws.cell(4, headers["include"]).column_letter
        include_dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
        status_dv = DataValidation(type="list", formula1='"Draft,Draft Confirmed,Public,Public Confirmed"', allow_blank=False)
        ws.add_data_validation(include_dv)
        ws.add_data_validation(status_dv)
        include_dv.add(f"{include_letter}5:{include_letter}500")
        status_dv.add(f"{status_letter}5:{status_letter}500")

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row or 5, min_col=1, max_col=len(REQUIRED_HEADERS)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                _copy_font(cell, name="Arial", size=8)

        if mode == "strict":
            table_tabs: set[str] = set()
            for row, tab in _iter_index_rows(ws, headers) or []:
                page_type = str(ws.cell(row, headers["page type"]).value or "").casefold()
                if any(token in page_type for token in ("table", "schedule", "data-grid", "matrix")):
                    table_tabs.add(tab)
            for sheet_name in table_tabs:
                if sheet_name not in wb.sheetnames:
                    continue
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value in (None, ""):
                            continue
                        _copy_font(cell, name="Arial", size=8)
                        if not any(r.min_row <= cell.row <= r.max_row and r.min_col <= cell.column <= r.max_col for r in sheet.merged_cells.ranges):
                            cell.border = border
                        cell.alignment = copy(cell.alignment)
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical or "center",
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=True if cell.alignment.wrap_text is None else cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )
            changes.append(f"Strictly normalized fonts/borders on {len(table_tabs)} indexed table/schedule sheets.")

        if wb.calculation is None:
            wb.calculation = CalcProperties(
                calcMode="auto",
                fullCalcOnLoad=True,
                forceFullCalc=True,
            )
        else:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"

        temp_fd, temp_name = tempfile.mkstemp(prefix=path.stem + "_", suffix=path.suffix, dir=path.parent)
        os.close(temp_fd)
        temp_path = Path(temp_name)
        try:
            wb.save(temp_path)
            os.replace(temp_path, path)
        finally:
            temp_path.unlink(missing_ok=True)
    finally:
        wb.close()

    return {
        "ok": True,
        "mode": mode,
        "backup": str(backup),
        "changes": changes,
        "audit": audit_workbook(path),
        "message": "Workbook repaired from a safety backup. Formula errors are reported but never guessed or silently rewritten.",
    }
