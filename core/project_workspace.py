"""Project-local file explorer and workbook workspace services.

This module restores the useful, project-local parts of the historical
template-platform implementation without requiring a schema-v2 migration.
Existing project packages remain authoritative and their source files are
indexed non-destructively into a virtual folder tree.
"""
from __future__ import annotations

import csv
from copy import deepcopy
import hashlib
import json
import os
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from werkzeug.utils import secure_filename

from core.workbook_geometry import (
    DEFAULT_COLUMN_WIDTH_UNITS,
    DEFAULT_ROW_HEIGHT_POINTS,
    pixels_to_excel_column_width,
    pixels_to_row_height_points,
)
from core.workbook_importer import _cell_style, _color_hex
from core.workbook_workspace import (
    apply_controlled_default_validations,
    apply_source_sheet_contract,
    is_source_sheet,
)


STANDARD_FOLDERS = [
    "Drawings",
    "Converted Schedules",
    "Survey Photos",
    "References",
    "Assets",
    "Symbol-Mapped Drawings",
    "Manual Layouts",
    "Programming",
    "Commissioning",
    "Archive",
]
SAFE_EXTENSIONS = {
    ".pdf": "pdf",
    ".png": "images",
    ".jpg": "images",
    ".jpeg": "images",
    ".webp": "images",
    ".svg": "images",
    ".xlsx": "spreadsheets",
    ".xlsm": "spreadsheets",
    ".csv": "csv",
    ".txt": "text",
    ".md": "text",
    ".json": "text",
    ".log": "text",
    ".doc": "documents",
    ".docx": "documents",
    ".rtf": "text",
    ".odt": "documents",
    ".ods": "spreadsheets",
    ".ppt": "documents",
    ".pptx": "documents",
}
MAX_FILE_BYTES = 100 * 1024 * 1024
MAX_ZIP_BYTES = 500 * 1024 * 1024
MAX_ZIP_MEMBERS = 2_000
FILE_ID_RE = re.compile(r"^[a-f0-9]{16}$")


class ProjectWorkspaceError(ValueError):
    """User-correctable project workspace error."""


class WorkbookRevisionConflict(ProjectWorkspaceError):
    """The browser tried to save an outdated workbook document."""


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def atomic_json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_virtual_path(value: str) -> str:
    raw = str(value or "").replace("\\", "/").strip("/")
    if not raw:
        return ""
    parts = [part.strip() for part in raw.split("/") if part.strip()]
    if any(
        part in {".", ".."}
        or not re.fullmatch(r'[^<>:"|?*\x00-\x1f]+', part)
        or part.endswith((".", " "))
        for part in parts
    ):
        raise ProjectWorkspaceError("The project folder path is unsafe.")
    return "/".join(parts)


def linked_project_root(project: dict[str, Any] | None) -> str:
    """Return the explicitly bound physical Project Files root, if any."""
    if not isinstance(project, dict):
        return ""
    metadata = project.get("metadata")
    metadata = metadata if isinstance(metadata, dict) else {}
    candidates = (
        project.get("linkedProjectRoot"),
        project.get("projectRoot"),
        project.get("EXACT_LINKED_PROJECT_ROOT"),
        metadata.get("linkedProjectRoot"),
        metadata.get("projectRoot"),
        metadata.get("EXACT_LINKED_PROJECT_ROOT"),
    )
    return next((str(item).strip() for item in candidates if str(item or "").strip()), "")


def _safe_item_name(value: str, *, label: str) -> str:
    name = str(value or "").strip()
    if (
        not name
        or name in {".", ".."}
        or Path(name.replace("\\", "/")).name != name
        or safe_virtual_path(name) != name
    ):
        raise ProjectWorkspaceError(f"A safe {label} name is required.")
    return name


def _conflict_safe_path(path: Path) -> Path:
    """Return path or a sibling '(n)' name without overwriting anything."""
    if not path.exists():
        return path
    suffix = path.suffix if path.is_file() or path.suffix else ""
    stem = path.name[: -len(suffix)] if suffix else path.name
    for index in range(1, 10_000):
        candidate = path.with_name(f"{stem} ({index}){suffix}")
        if not candidate.exists():
            return candidate
    raise ProjectWorkspaceError(f"Could not create a conflict-safe name for {path.name}.")


def _legacy_virtual_folder(relative_path: Path) -> str:
    first = relative_path.parts[0].casefold() if relative_path.parts else ""
    suffix = relative_path.suffix.casefold()
    if first == "workbook" or first == "spreadsheets" or suffix in {
        ".xlsx",
        ".xlsm",
        ".csv",
        ".ods",
    }:
        return "Converted Schedules"
    if first == "pdf":
        return "Drawings"
    if first == "images" and suffix == ".pdf":
        return "Symbol-Mapped Drawings"
    if first == "images":
        return "Assets"
    if first == "documents":
        return "References"
    return "References"


class LegacyProjectFileLibrary:
    """Virtual project file tree backed by files inside one project package."""

    def __init__(self, project_dir: Path):
        self.project_dir = Path(project_dir).resolve()
        self.manifest_path = self.project_dir / "source_library.json"
        self.sources_root = self.project_dir / "sources"

    @staticmethod
    def empty() -> dict[str, Any]:
        return {
            "schemaVersion": 3,
            "folders": list(STANDARD_FOLDERS),
            "archivedFolders": [],
            "files": [],
            "conversionQueue": [],
            "importReports": [],
        }

    def _read_manifest(self) -> dict[str, Any]:
        if not self.manifest_path.is_file():
            return self.empty()
        try:
            payload = json.loads(self.manifest_path.read_text("utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ProjectWorkspaceError(
                "The Project Files index is unreadable."
            ) from exc
        payload["schemaVersion"] = 3
        payload.setdefault("folders", [])
        payload.setdefault("archivedFolders", [])
        payload.setdefault("files", payload.pop("sources", []))
        payload.setdefault("conversionQueue", [])
        payload.setdefault("importReports", [])
        payload["folders"] = list(
            dict.fromkeys([*STANDARD_FOLDERS, *payload["folders"]])
        )
        for record in payload["files"]:
            record.setdefault("fileType", record.pop("sourceType", "other"))
            record.setdefault("virtualPath", "")
            record.setdefault(
                "relativePath",
                f"{record['virtualPath']}/{record.get('originalFileName', '')}".strip(
                    "/"
                ),
            )
        return payload

    def _discover_existing(self, known_paths: set[str]) -> list[dict[str, Any]]:
        if not self.sources_root.is_dir():
            return []
        records: list[dict[str, Any]] = []
        for path in sorted(self.sources_root.rglob("*")):
            if not path.is_file():
                continue
            local_path = path.relative_to(self.project_dir).as_posix()
            if local_path in known_paths:
                continue
            relative = path.relative_to(self.sources_root)
            virtual_path = _legacy_virtual_folder(relative)
            suffix = path.suffix.casefold()
            record_id = hashlib.sha256(
                f"project-file:{local_path}".encode("utf-8")
            ).hexdigest()[:16]
            records.append(
                {
                    "id": record_id,
                    "originalFileName": path.name,
                    "storedFileName": path.name,
                    "mediaType": suffix.lstrip(".") or "file",
                    "fileType": SAFE_EXTENSIONS.get(suffix, "other"),
                    "size": path.stat().st_size,
                    "sha256": sha256_file(path),
                    "dateAdded": datetime.fromtimestamp(
                        path.stat().st_mtime, timezone.utc
                    ).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "version": 1,
                    "status": "active",
                    "virtualPath": virtual_path,
                    "relativePath": f"{virtual_path}/{path.name}",
                    "localProjectPath": local_path,
                    "tags": [],
                    "notes": "",
                    "discovered": True,
                }
            )
        return records

    def load(self) -> dict[str, Any]:
        payload = self._read_manifest()
        known = {
            str(item.get("localProjectPath") or "")
            for item in payload["files"]
        }
        payload["files"] = [
            *payload["files"],
            *self._discover_existing(known),
        ]
        discovered_folders = [
            str(item.get("virtualPath") or "") for item in payload["files"]
        ]
        payload["folders"] = sorted(
            dict.fromkeys(
                folder
                for folder in [*STANDARD_FOLDERS, *payload["folders"], *discovered_folders]
                if folder
            ),
            key=str.casefold,
        )
        return payload

    def _save(self, payload: dict[str, Any]) -> None:
        payload["schemaVersion"] = 3
        atomic_json_write(self.manifest_path, payload)

    def create_folder(self, virtual_path: str) -> str:
        folder = safe_virtual_path(virtual_path)
        if not folder:
            raise ProjectWorkspaceError("A folder name is required.")
        payload = self.load()
        if folder not in payload["folders"]:
            payload["folders"].append(folder)
            payload["folders"].sort(key=str.casefold)
            self._save(payload)
        return folder

    def upload(
        self,
        stream: BinaryIO,
        original_name: str,
        virtual_path: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        clean_name = secure_filename(Path(str(original_name)).name)
        suffix = Path(clean_name).suffix.casefold()
        file_type = SAFE_EXTENSIONS.get(suffix)
        if not clean_name or not file_type:
            raise ProjectWorkspaceError(
                f"Unsupported or unsafe project file: {original_name}"
            )
        folder = safe_virtual_path(virtual_path)
        # Materialize existing records before creating the destination. Otherwise
        # the just-written content-addressed file would be rediscovered as a
        # second legacy record during this same upload.
        payload = self.load()
        record_id = uuid.uuid4().hex[:16]
        destination_dir = self.sources_root / "library" / folder
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / f"{record_id}__{clean_name}"
        digest = hashlib.sha256()
        total = 0
        try:
            with destination.open("xb") as handle:
                while True:
                    chunk = stream.read(1024 * 1024)
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > MAX_FILE_BYTES:
                        raise ProjectWorkspaceError(
                            "Project files may not exceed 100 MB each."
                        )
                    digest.update(chunk)
                    handle.write(chunk)
        except Exception:
            destination.unlink(missing_ok=True)
            raise

        existing = [
            item
            for item in payload["files"]
            if item.get("originalFileName") == original_name
            and item.get("virtualPath", "") == folder
        ]
        record = {
            "id": record_id,
            "originalFileName": original_name,
            "storedFileName": destination.name,
            "mediaType": suffix.lstrip("."),
            "fileType": file_type,
            "size": total,
            "sha256": digest.hexdigest(),
            "dateAdded": utcnow(),
            "addedBy": str((metadata or {}).get("addedBy") or ""),
            "version": len(existing) + 1,
            "status": "active",
            "virtualPath": folder,
            "relativePath": f"{folder}/{original_name}".strip("/"),
            "localProjectPath": destination.relative_to(
                self.project_dir
            ).as_posix(),
            "tags": list((metadata or {}).get("tags") or []),
            "notes": str((metadata or {}).get("notes") or ""),
        }
        for old in existing:
            if old.get("status") == "active":
                old["status"] = "superseded"
                old["supersededBy"] = record_id
                record["supersedes"] = old["id"]
        payload["files"].append(record)
        if folder and folder not in payload["folders"]:
            payload["folders"].append(folder)
        self._save(payload)
        return record

    def import_zip(
        self, stream: BinaryIO, original_name: str, virtual_path: str = ""
    ) -> dict[str, Any]:
        root = safe_virtual_path(virtual_path)
        fd, temp_name = tempfile.mkstemp(
            prefix=".project-files-", suffix=".zip", dir=self.project_dir
        )
        os.close(fd)
        archive_path = Path(temp_name)
        imported: list[dict[str, Any]] = []
        rejected: list[dict[str, str]] = []
        created_folders: set[str] = set()
        try:
            with archive_path.open("wb") as handle:
                shutil.copyfileobj(stream, handle)
            if archive_path.stat().st_size > MAX_ZIP_BYTES:
                raise ProjectWorkspaceError("ZIP imports may not exceed 500 MB.")
            with zipfile.ZipFile(archive_path) as bundle:
                members = bundle.infolist()
                if len(members) > MAX_ZIP_MEMBERS:
                    raise ProjectWorkspaceError("ZIP contains too many entries.")
                if sum(item.file_size for item in members) > MAX_ZIP_BYTES:
                    raise ProjectWorkspaceError(
                        "ZIP expands beyond the 500 MB safety limit."
                    )
                for member in members:
                    try:
                        member_path = safe_virtual_path(member.filename)
                        if not member_path:
                            continue
                        member_folder = Path(member_path).parent.as_posix()
                        if member_folder == ".":
                            member_folder = ""
                        destination_folder = "/".join(
                            part for part in (root, member_folder) if part
                        )
                        if member.is_dir():
                            if destination_folder:
                                self.create_folder(destination_folder)
                                created_folders.add(destination_folder)
                            continue
                        if member.flag_bits & 0x1:
                            raise ProjectWorkspaceError(
                                "Encrypted ZIP entries are not supported."
                            )
                        if member.file_size > MAX_FILE_BYTES:
                            raise ProjectWorkspaceError(
                                "ZIP member exceeds the 100 MB file limit."
                            )
                        with bundle.open(member) as source:
                            imported.append(
                                self.upload(
                                    source,
                                    Path(member_path).name,
                                    destination_folder,
                                    {
                                        "notes": (
                                            f"Imported from {original_name}:"
                                            f"{member.filename}"
                                        )
                                    },
                                )
                            )
                    except Exception as exc:
                        rejected.append(
                            {"path": member.filename, "reason": str(exc)}
                        )
            payload = self.load()
            report = {
                "id": uuid.uuid4().hex[:16],
                "archive": original_name,
                "imported": len(imported),
                "createdFolders": sorted(created_folders),
                "rejected": rejected,
                "createdAt": utcnow(),
            }
            payload["importReports"].append(report)
            self._save(payload)
            return {"report": report, "files": imported}
        finally:
            archive_path.unlink(missing_ok=True)

    def _file(self, payload: dict[str, Any], file_id: str) -> dict[str, Any]:
        if not FILE_ID_RE.fullmatch(file_id):
            raise ProjectWorkspaceError("Invalid project file ID.")
        for record in payload["files"]:
            if record.get("id") == file_id:
                return record
        raise ProjectWorkspaceError("Project file was not found.")

    def resolve(self, file_id: str) -> tuple[dict[str, Any], Path]:
        payload = self.load()
        record = self._file(payload, file_id)
        path = (self.project_dir / str(record["localProjectPath"])).resolve()
        try:
            path.relative_to(self.project_dir)
        except ValueError as exc:
            raise ProjectWorkspaceError(
                "Project file path escaped the project package."
            ) from exc
        if not path.is_file():
            raise ProjectWorkspaceError("Project file is missing.")
        return record, path

    def rename_file(self, file_id: str, name: str) -> dict[str, Any]:
        safe_name = Path(str(name).replace("\\", "/")).name.strip()
        if (
            not safe_name
            or safe_name in {".", ".."}
            or safe_name != str(name).strip()
            or not safe_virtual_path(safe_name)
        ):
            raise ProjectWorkspaceError("A safe file name is required.")
        payload = self.load()
        record = self._file(payload, file_id)
        record["originalFileName"] = safe_name
        record["relativePath"] = (
            f"{record.get('virtualPath', '')}/{safe_name}".strip("/")
        )
        record["renamedAt"] = utcnow()
        self._save(payload)
        return record

    def move_file(self, file_id: str, destination: str) -> dict[str, Any]:
        folder = safe_virtual_path(destination)
        payload = self.load()
        record = self._file(payload, file_id)
        record["virtualPath"] = folder
        record["relativePath"] = (
            f"{folder}/{record['originalFileName']}".strip("/")
        )
        record["movedAt"] = utcnow()
        if folder and folder not in payload["folders"]:
            payload["folders"].append(folder)
        self._save(payload)
        return record

    def archive_file(self, file_id: str) -> dict[str, Any]:
        payload = self.load()
        record = self._file(payload, file_id)
        record["status"] = "archived"
        record["archivedAt"] = utcnow()
        self._save(payload)
        return record

    def restore_file(self, file_id: str) -> dict[str, Any]:
        payload = self.load()
        record = self._file(payload, file_id)
        record["status"] = "active"
        record.pop("archivedAt", None)
        record["restoredAt"] = utcnow()
        self._save(payload)
        return record

    def _relocate_folder(
        self, payload: dict[str, Any], old_path: str, new_path: str
    ) -> None:
        old = safe_virtual_path(old_path)
        new = safe_virtual_path(new_path)
        if not old or not new or old == new:
            raise ProjectWorkspaceError(
                "Choose a different source and destination folder."
            )
        if old not in payload["folders"]:
            raise ProjectWorkspaceError("Project folder was not found.")
        if new == old or new.startswith(f"{old}/"):
            raise ProjectWorkspaceError(
                "A folder cannot be moved inside itself."
            )
        rewritten: list[str] = []
        for folder in payload["folders"]:
            if folder == old or folder.startswith(f"{old}/"):
                rewritten.append(f"{new}{folder[len(old):]}")
            else:
                rewritten.append(folder)
        payload["folders"] = sorted(dict.fromkeys(rewritten), key=str.casefold)
        for record in payload["files"]:
            current = str(record.get("virtualPath") or "")
            if current == old or current.startswith(f"{old}/"):
                record["virtualPath"] = f"{new}{current[len(old):]}"
                record["relativePath"] = (
                    f"{record['virtualPath']}/{record['originalFileName']}"
                )

    def rename_folder(self, path: str, name: str) -> str:
        current = safe_virtual_path(path)
        parent = Path(current).parent.as_posix()
        if parent == ".":
            parent = ""
        safe_name = safe_virtual_path(name)
        if "/" in safe_name:
            raise ProjectWorkspaceError("Enter one folder name.")
        destination = "/".join(
            part for part in (parent, safe_name) if part
        )
        payload = self.load()
        self._relocate_folder(payload, current, destination)
        self._save(payload)
        return destination

    def move_folder(self, path: str, destination: str) -> str:
        current = safe_virtual_path(path)
        parent = safe_virtual_path(destination)
        moved = "/".join(
            part for part in (parent, Path(current).name) if part
        )
        payload = self.load()
        self._relocate_folder(payload, current, moved)
        self._save(payload)
        return moved

    def archive_folder(self, path: str) -> str:
        current = safe_virtual_path(path)
        if current == "Archive" or current.startswith("Archive/"):
            raise ProjectWorkspaceError("This folder is already archived.")
        destination = f"Archive/{current}"
        payload = self.load()
        self._relocate_folder(payload, current, destination)
        for record in payload["files"]:
            folder = str(record.get("virtualPath") or "")
            if folder == destination or folder.startswith(f"{destination}/"):
                record["status"] = "archived"
                record["archivedAt"] = utcnow()
        payload["archivedFolders"].append(
            {"path": destination, "restorePath": current, "archivedAt": utcnow()}
        )
        self._save(payload)
        return destination

    def restore_folder(self, path: str) -> str:
        archived = safe_virtual_path(path)
        payload = self.load()
        entry = next(
            (
                item
                for item in payload["archivedFolders"]
                if item.get("path") == archived
            ),
            None,
        )
        if not entry:
            raise ProjectWorkspaceError(
                "This folder does not have a restore location."
            )
        destination = safe_virtual_path(str(entry["restorePath"]))
        self._relocate_folder(payload, archived, destination)
        for record in payload["files"]:
            folder = str(record.get("virtualPath") or "")
            if folder == destination or folder.startswith(f"{destination}/"):
                record["status"] = "active"
                record.pop("archivedAt", None)
        payload["archivedFolders"].remove(entry)
        self._save(payload)
        return destination

    def preview(self, file_id: str) -> dict[str, Any]:
        record, path = self.resolve(file_id)
        result: dict[str, Any] = {
            "file": record,
            "kind": record["fileType"],
            "previewError": "",
        }
        try:
            suffix = path.suffix.casefold()
            if suffix in {".xlsx", ".xlsm"}:
                workbook = load_workbook(
                    path,
                    read_only=True,
                    data_only=False,
                    keep_vba=suffix == ".xlsm",
                )
                result["sheets"] = workbook.sheetnames
                sheet = workbook[workbook.sheetnames[0]]
                result["grid"] = [
                    ["" if value is None else str(value) for value in row]
                    for row in sheet.iter_rows(
                        max_row=50, max_col=20, values_only=True
                    )
                ]
                workbook.close()
            elif suffix == ".csv":
                text = path.read_text("utf-8", errors="replace")[:100_000]
                result["text"] = text
                result["grid"] = list(csv.reader(text.splitlines()))[:100]
            elif suffix in {".txt", ".md", ".json", ".log", ".rtf"}:
                result["text"] = path.read_text(
                    "utf-8", errors="replace"
                )[:100_000]
            elif suffix == ".pdf":
                import fitz

                pdf = fitz.open(path)
                result["pageCount"] = pdf.page_count
                result["pageSizes"] = [
                    {"width": page.rect.width, "height": page.rect.height}
                    for page in list(pdf)[:25]
                ]
                pdf.close()
            elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
                from PIL import Image

                with Image.open(path) as image:
                    result["dimensions"] = {
                        "width": image.width,
                        "height": image.height,
                    }
            else:
                result["fallback"] = "metadata"
        except Exception as exc:
            result["previewError"] = f"{type(exc).__name__}: {exc}"
        return result


class LinkedProjectFileLibrary:
    """Live, exact Project Files view backed by one validated physical root."""

    def __init__(self, project_dir: Path, root: str | Path):
        self.project_dir = Path(project_dir).resolve()
        self.root = Path(root).expanduser().resolve()
        self.audit_path = self.project_dir / "linked_project_files.json"

    def _require_root(self) -> Path:
        if not self.root.is_dir():
            raise ProjectWorkspaceError(
                f"The linked project root is unavailable: {self.root}"
            )
        if self.root == Path(self.root.anchor):
            raise ProjectWorkspaceError(
                "A drive root cannot be used as a linked project root."
            )
        return self.root

    def _path(
        self,
        relative_path: str,
        *,
        must_exist: bool = True,
    ) -> Path:
        root = self._require_root()
        relative = safe_virtual_path(relative_path)
        candidate = (
            root
            if not relative
            else root.joinpath(*relative.split("/"))
        ).resolve(strict=must_exist)
        if candidate != root and root not in candidate.parents:
            raise ProjectWorkspaceError(
                "The linked project path escaped the bound physical root."
            )
        return candidate

    def _relative(self, path: Path) -> str:
        resolved = path.resolve(strict=True)
        root = self._require_root()
        try:
            return resolved.relative_to(root).as_posix()
        except ValueError as exc:
            raise ProjectWorkspaceError(
                "The linked project path escaped the bound physical root."
            ) from exc

    def _read_audit(self) -> dict[str, Any]:
        if not self.audit_path.is_file():
            return {"schemaVersion": 1, "archives": []}
        try:
            payload = json.loads(self.audit_path.read_text("utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ProjectWorkspaceError(
                "The linked Project Files audit history is unreadable."
            ) from exc
        payload.setdefault("schemaVersion", 1)
        payload.setdefault("archives", [])
        return payload

    def _save_audit(self, payload: dict[str, Any]) -> None:
        payload["schemaVersion"] = 1
        atomic_json_write(self.audit_path, payload)

    def _rewrite_archive_paths(self, old_path: str, new_path: str) -> None:
        audit = self._read_audit()
        changed = False
        for item in audit["archives"]:
            archived = str(item.get("archivePath") or "")
            if archived == old_path or archived.startswith(f"{old_path}/"):
                item["archivePath"] = f"{new_path}{archived[len(old_path):]}"
                changed = True
        if changed:
            self._save_audit(audit)

    def _record(
        self,
        path: Path,
        *,
        status: str | None = None,
    ) -> dict[str, Any]:
        relative = self._relative(path)
        stat = path.stat()
        parent = Path(relative).parent.as_posix()
        if parent == ".":
            parent = ""
        suffix = path.suffix.casefold()
        modified = datetime.fromtimestamp(
            stat.st_mtime, timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%SZ")
        return {
            "id": hashlib.sha256(
                f"linked-project-file:{relative}".encode("utf-8")
            ).hexdigest()[:16],
            "originalFileName": path.name,
            "storedFileName": path.name,
            "mediaType": suffix.lstrip(".") or "file",
            "fileType": SAFE_EXTENSIONS.get(suffix, "other"),
            "size": stat.st_size,
            "sha256": "",
            "dateAdded": modified,
            "modifiedAt": modified,
            "version": 1,
            "status": status or "active",
            "virtualPath": parent,
            "relativePath": relative,
            "localProjectPath": str(path),
            "physicalPath": str(path),
            "tags": [],
            "notes": "",
            "linked": True,
        }

    def load(self) -> dict[str, Any]:
        root = self._require_root()
        audit = self._read_audit()
        archived_files = {
            str(item.get("archivePath") or "")
            for item in audit["archives"]
            if item.get("kind") == "file"
        }
        archived_folders = [
            str(item.get("archivePath") or "")
            for item in audit["archives"]
            if item.get("kind") == "folder"
        ]
        folders: list[str] = []
        files: list[dict[str, Any]] = []
        try:
            for current, directory_names, file_names in os.walk(
                root, topdown=True, followlinks=False
            ):
                current_path = Path(current)
                self._relative(current_path)
                directory_names.sort(key=str.casefold)
                file_names.sort(key=str.casefold)
                for name in directory_names:
                    child = current_path / name
                    relative = self._relative(child)
                    folders.append(relative)
                for name in file_names:
                    child = current_path / name
                    relative = self._relative(child)
                    files.append(
                        self._record(
                            child,
                            status=(
                                "archived"
                                if relative in archived_files
                                or any(
                                    relative.startswith(f"{folder}/")
                                    for folder in archived_folders
                                )
                                else "active"
                            ),
                        )
                    )
        except OSError as exc:
            raise ProjectWorkspaceError(
                f"The linked project root could not be enumerated: {exc}"
            ) from exc
        return {
            "schemaVersion": 4,
            "mode": "linked",
            "linked": True,
            "rootPath": str(root),
            "rootName": root.name,
            "folders": sorted(folders, key=str.casefold),
            "archivedFolders": [
                {
                    "path": str(item.get("archivePath") or ""),
                    "restorePath": str(item.get("originalPath") or ""),
                    "archivedAt": str(item.get("archivedAt") or ""),
                }
                for item in audit["archives"]
                if item.get("kind") == "folder"
                and self._path(
                    str(item.get("archivePath") or ""),
                    must_exist=False,
                ).is_dir()
            ],
            "files": sorted(
                files,
                key=lambda item: str(item["relativePath"]).casefold(),
            ),
            "conversionQueue": [],
            "importReports": [],
        }

    def create_folder(self, virtual_path: str) -> str:
        relative = safe_virtual_path(virtual_path)
        if not relative:
            raise ProjectWorkspaceError("A folder name is required.")
        destination = self._path(relative, must_exist=False)
        destination = _conflict_safe_path(destination)
        destination.mkdir(parents=True, exist_ok=False)
        return self._relative(destination)

    def upload(
        self,
        stream: BinaryIO,
        original_name: str,
        virtual_path: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        name = _safe_item_name(Path(str(original_name)).name, label="file")
        folder = safe_virtual_path(virtual_path)
        destination_dir = self._path(folder, must_exist=False)
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = _conflict_safe_path(destination_dir / name)
        total = 0
        try:
            with destination.open("xb") as handle:
                while True:
                    chunk = stream.read(1024 * 1024)
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > MAX_FILE_BYTES:
                        raise ProjectWorkspaceError(
                            "Project files may not exceed 100 MB each."
                        )
                    handle.write(chunk)
        except Exception:
            destination.unlink(missing_ok=True)
            raise
        modified_ms = (metadata or {}).get("modifiedTimeMs")
        if isinstance(modified_ms, (int, float)) and modified_ms > 0:
            modified_seconds = float(modified_ms) / 1000
            os.utime(destination, (modified_seconds, modified_seconds))
        return self._record(destination)

    def import_zip(
        self, stream: BinaryIO, original_name: str, virtual_path: str = ""
    ) -> dict[str, Any]:
        root = safe_virtual_path(virtual_path)
        fd, temp_name = tempfile.mkstemp(
            prefix=".singh360-project-files-", suffix=".zip"
        )
        os.close(fd)
        archive_path = Path(temp_name)
        imported: list[dict[str, Any]] = []
        rejected: list[dict[str, str]] = []
        created_folders: set[str] = set()
        try:
            with archive_path.open("wb") as handle:
                shutil.copyfileobj(stream, handle)
            if archive_path.stat().st_size > MAX_ZIP_BYTES:
                raise ProjectWorkspaceError("ZIP imports may not exceed 500 MB.")
            with zipfile.ZipFile(archive_path) as bundle:
                members = bundle.infolist()
                if len(members) > MAX_ZIP_MEMBERS:
                    raise ProjectWorkspaceError("ZIP contains too many entries.")
                if sum(item.file_size for item in members) > MAX_ZIP_BYTES:
                    raise ProjectWorkspaceError(
                        "ZIP expands beyond the 500 MB safety limit."
                    )
                for member in members:
                    try:
                        member_path = safe_virtual_path(member.filename)
                        if not member_path:
                            continue
                        if (member.external_attr >> 16) & 0o170000 == 0o120000:
                            raise ProjectWorkspaceError(
                                "ZIP symbolic links are not supported."
                            )
                        parent = Path(member_path).parent.as_posix()
                        if parent == ".":
                            parent = ""
                        destination_folder = "/".join(
                            part for part in (root, parent) if part
                        )
                        if member.is_dir():
                            folder_path = self._path(
                                "/".join(
                                    part
                                    for part in (root, member_path)
                                    if part
                                ),
                                must_exist=False,
                            )
                            if not folder_path.exists():
                                folder_path.mkdir(parents=True)
                                created_folders.add(
                                    self._relative(folder_path)
                                )
                            continue
                        if member.flag_bits & 0x1:
                            raise ProjectWorkspaceError(
                                "Encrypted ZIP entries are not supported."
                            )
                        if member.file_size > MAX_FILE_BYTES:
                            raise ProjectWorkspaceError(
                                "ZIP member exceeds the 100 MB file limit."
                            )
                        modified_at = datetime(*member.date_time).timestamp() * 1000
                        with bundle.open(member) as source:
                            imported.append(
                                self.upload(
                                    source,
                                    Path(member_path).name,
                                    destination_folder,
                                    {"modifiedTimeMs": modified_at},
                                )
                            )
                    except Exception as exc:
                        rejected.append(
                            {"path": member.filename, "reason": str(exc)}
                        )
            report = {
                "id": uuid.uuid4().hex[:16],
                "archive": original_name,
                "imported": len(imported),
                "createdFolders": sorted(created_folders),
                "rejected": rejected,
                "createdAt": utcnow(),
            }
            return {"report": report, "files": imported}
        finally:
            archive_path.unlink(missing_ok=True)

    def _file(self, file_id: str) -> dict[str, Any]:
        if not FILE_ID_RE.fullmatch(file_id):
            raise ProjectWorkspaceError("Invalid project file ID.")
        for record in self.load()["files"]:
            if record.get("id") == file_id:
                return record
        raise ProjectWorkspaceError("Project file was not found.")

    def resolve(self, file_id: str) -> tuple[dict[str, Any], Path]:
        record = self._file(file_id)
        path = self._path(str(record["relativePath"]))
        if not path.is_file():
            raise ProjectWorkspaceError("Project file is missing.")
        return record, path

    def rename_file(self, file_id: str, name: str) -> dict[str, Any]:
        record, source = self.resolve(file_id)
        safe_name = _safe_item_name(name, label="file")
        destination = _conflict_safe_path(source.with_name(safe_name))
        source.rename(destination)
        self._rewrite_archive_paths(
            str(record["relativePath"]), self._relative(destination)
        )
        return self._record(
            destination,
            status=str(record.get("status") or "active"),
        )

    def move_file(self, file_id: str, destination: str) -> dict[str, Any]:
        record, source = self.resolve(file_id)
        folder = self._path(safe_virtual_path(destination))
        if not folder.is_dir():
            raise ProjectWorkspaceError("The destination folder was not found.")
        target = _conflict_safe_path(folder / source.name)
        shutil.move(str(source), str(target))
        self._rewrite_archive_paths(
            str(record["relativePath"]), self._relative(target)
        )
        return self._record(
            target,
            status=str(record.get("status") or "active"),
        )

    def archive_file(self, file_id: str) -> dict[str, Any]:
        record, source = self.resolve(file_id)
        relative = str(record["relativePath"])
        if relative == "Archive" or relative.startswith("Archive/"):
            raise ProjectWorkspaceError("This file is already archived.")
        target = self._path(f"Archive/{relative}", must_exist=False)
        target.parent.mkdir(parents=True, exist_ok=True)
        target = _conflict_safe_path(target)
        shutil.move(str(source), str(target))
        audit = self._read_audit()
        audit["archives"].append(
            {
                "kind": "file",
                "archivePath": self._relative(target),
                "originalPath": relative,
                "archivedAt": utcnow(),
            }
        )
        self._save_audit(audit)
        return self._record(target, status="archived")

    def restore_file(self, file_id: str) -> dict[str, Any]:
        record, source = self.resolve(file_id)
        audit = self._read_audit()
        entry = next(
            (
                item
                for item in audit["archives"]
                if item.get("kind") == "file"
                and item.get("archivePath") == record["relativePath"]
            ),
            None,
        )
        if not entry:
            raise ProjectWorkspaceError(
                "This file does not have a recorded restore location."
            )
        target = self._path(
            str(entry["originalPath"]), must_exist=False
        )
        target.parent.mkdir(parents=True, exist_ok=True)
        target = _conflict_safe_path(target)
        shutil.move(str(source), str(target))
        audit["archives"].remove(entry)
        self._save_audit(audit)
        return self._record(target)

    def rename_folder(self, path: str, name: str) -> str:
        current = safe_virtual_path(path)
        source = self._path(current)
        if source == self.root or not source.is_dir():
            raise ProjectWorkspaceError("Project folder was not found.")
        target = _conflict_safe_path(
            source.with_name(_safe_item_name(name, label="folder"))
        )
        source.rename(target)
        self._rewrite_archive_paths(current, self._relative(target))
        return self._relative(target)

    def move_folder(self, path: str, destination: str) -> str:
        current = safe_virtual_path(path)
        source = self._path(current)
        parent = self._path(safe_virtual_path(destination))
        if source == self.root or not source.is_dir() or not parent.is_dir():
            raise ProjectWorkspaceError("Project folder was not found.")
        if parent == source or source in parent.parents:
            raise ProjectWorkspaceError("A folder cannot be moved inside itself.")
        target = _conflict_safe_path(parent / source.name)
        shutil.move(str(source), str(target))
        self._rewrite_archive_paths(current, self._relative(target))
        return self._relative(target)

    def archive_folder(self, path: str) -> str:
        relative = safe_virtual_path(path)
        source = self._path(relative)
        if source == self.root or not source.is_dir():
            raise ProjectWorkspaceError("Project folder was not found.")
        if relative == "Archive" or relative.startswith("Archive/"):
            raise ProjectWorkspaceError("This folder is already archived.")
        target = self._path(f"Archive/{relative}", must_exist=False)
        target.parent.mkdir(parents=True, exist_ok=True)
        target = _conflict_safe_path(target)
        shutil.move(str(source), str(target))
        audit = self._read_audit()
        audit["archives"].append(
            {
                "kind": "folder",
                "archivePath": self._relative(target),
                "originalPath": relative,
                "archivedAt": utcnow(),
            }
        )
        self._save_audit(audit)
        return self._relative(target)

    def restore_folder(self, path: str) -> str:
        archived = safe_virtual_path(path)
        source = self._path(archived)
        audit = self._read_audit()
        entry = next(
            (
                item
                for item in audit["archives"]
                if item.get("kind") == "folder"
                and item.get("archivePath") == archived
            ),
            None,
        )
        if not entry:
            raise ProjectWorkspaceError(
                "This folder does not have a recorded restore location."
            )
        target = self._path(
            str(entry["originalPath"]), must_exist=False
        )
        target.parent.mkdir(parents=True, exist_ok=True)
        target = _conflict_safe_path(target)
        shutil.move(str(source), str(target))
        audit["archives"].remove(entry)
        self._save_audit(audit)
        return self._relative(target)

    def resolve_folder(self, path: str) -> Path:
        folder = self._path(safe_virtual_path(path))
        if not folder.is_dir():
            raise ProjectWorkspaceError("Project folder was not found.")
        return folder

    def preview(self, file_id: str) -> dict[str, Any]:
        record, path = self.resolve(file_id)
        record = {**record, "sha256": sha256_file(path)}
        result: dict[str, Any] = {
            "file": record,
            "kind": record["fileType"],
            "previewError": "",
        }
        try:
            suffix = path.suffix.casefold()
            if suffix in {".xlsx", ".xlsm"}:
                workbook = load_workbook(
                    path,
                    read_only=True,
                    data_only=False,
                    keep_vba=suffix == ".xlsm",
                )
                result["sheets"] = workbook.sheetnames
                sheet = workbook[workbook.sheetnames[0]]
                result["grid"] = [
                    ["" if value is None else str(value) for value in row]
                    for row in sheet.iter_rows(
                        max_row=50, max_col=20, values_only=True
                    )
                ]
                workbook.close()
            elif suffix == ".csv":
                text = path.read_text("utf-8", errors="replace")[:100_000]
                result["text"] = text
                result["grid"] = list(csv.reader(text.splitlines()))[:100]
            elif suffix in {".txt", ".md", ".json", ".log", ".rtf"}:
                result["text"] = path.read_text(
                    "utf-8", errors="replace"
                )[:100_000]
            elif suffix == ".pdf":
                import fitz

                with fitz.open(path) as pdf:
                    result["pageCount"] = pdf.page_count
                    result["pageSizes"] = [
                        {"width": page.rect.width, "height": page.rect.height}
                        for page in list(pdf)[:25]
                    ]
            elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
                from PIL import Image

                with Image.open(path) as image:
                    result["dimensions"] = {
                        "width": image.width,
                        "height": image.height,
                    }
            else:
                result["fallback"] = "metadata"
        except Exception as exc:
            result["previewError"] = f"{type(exc).__name__}: {exc}"
        return result


def ProjectFileLibrary(
    project_dir: Path,
    project: dict[str, Any] | None = None,
) -> LegacyProjectFileLibrary | LinkedProjectFileLibrary:
    """Select exact linked-root mode or the unchanged legacy virtual mode."""
    root = linked_project_root(project)
    if root:
        return LinkedProjectFileLibrary(project_dir, root)
    return LegacyProjectFileLibrary(project_dir)


def open_local_path(path: Path) -> None:
    """Open an already validated real file in its Windows default application."""
    if os.name != "nt" or not hasattr(os, "startfile"):
        raise ProjectWorkspaceError(
            "Opening a local file requires the Windows Singh360 server."
        )
    os.startfile(str(path))  # type: ignore[attr-defined]


def reveal_local_path(path: Path, *, select: bool) -> None:
    """Open Explorer at an already validated file or folder path."""
    if os.name != "nt":
        raise ProjectWorkspaceError(
            "Showing a local path requires the Windows Singh360 server."
        )
    arguments = (
        ["explorer.exe", f"/select,{path}"]
        if select
        else ["explorer.exe", str(path)]
    )
    subprocess.Popen(  # noqa: S603 - fixed executable, validated path
        arguments,
        close_fds=True,
    )


def _json_cell_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _data_validations_payload(sheet: Any) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    container = getattr(sheet, "data_validations", None)
    for index, validation in enumerate(
        getattr(container, "dataValidation", None) or []
    ):
        ranges = [
            str(item)
            for item in getattr(getattr(validation, "ranges", None), "ranges", [])
        ]
        if not ranges:
            continue
        formula1 = getattr(validation, "formula1", None)
        values: list[str] = []
        if (
            str(getattr(validation, "type", "") or "").casefold() == "list"
            and isinstance(formula1, str)
            and len(formula1) >= 2
            and formula1.startswith('"')
            and formula1.endswith('"')
        ):
            values = [
                value.strip()
                for value in formula1[1:-1].split(",")
            ]
        output.append(
            {
                "id": str(
                    getattr(validation, "uid", None)
                    or f"excel-validation-{index + 1}"
                ),
                "ranges": ranges,
                "type": str(getattr(validation, "type", "") or ""),
                "operator": str(getattr(validation, "operator", "") or ""),
                "formula1": formula1,
                "formula2": getattr(validation, "formula2", None),
                "values": values,
                "allowBlank": bool(getattr(validation, "allow_blank", False)),
                "showDropdown": not bool(
                    getattr(validation, "showDropDown", False)
                ),
                "showErrorMessage": bool(
                    getattr(validation, "showErrorMessage", True)
                ),
                "error": str(getattr(validation, "error", "") or ""),
                "errorTitle": str(
                    getattr(validation, "errorTitle", "") or ""
                ),
            }
        )
    return output


def _conditional_formats_payload(sheet: Any) -> list[dict[str, Any]]:
    """Capture portable conditional-format metadata without mutating rules."""
    output: list[dict[str, Any]] = []
    rules_by_range = getattr(
        getattr(sheet, "conditional_formatting", None),
        "_cf_rules",
        {},
    )
    for range_key, rules in rules_by_range.items():
        sqref = str(getattr(range_key, "sqref", None) or range_key)
        for rule in rules or []:
            fill = None
            font_color = None
            differential = getattr(rule, "dxf", None)
            if differential is not None:
                pattern = getattr(getattr(differential, "fill", None), "fgColor", None)
                fill = _color_hex(pattern)
                font_color = _color_hex(
                    getattr(getattr(differential, "font", None), "color", None)
                )
            output.append(
                {
                    "ranges": [value for value in sqref.split() if value],
                    "type": str(getattr(rule, "type", "") or ""),
                    "operator": str(getattr(rule, "operator", "") or ""),
                    "formula": [
                        str(value)
                        for value in (getattr(rule, "formula", None) or [])
                    ],
                    "priority": int(getattr(rule, "priority", 0) or 0),
                    "stopIfTrue": bool(
                        getattr(rule, "stopIfTrue", False)
                    ),
                    "fill": fill,
                    "fontColor": font_color,
                }
            )
    return output


def _project_index_entries(project: dict[str, Any]) -> dict[str, dict[str, Any]]:
    index = next(
        (
            sheet
            for sheet in (project.get("worksheets") or [])
            if str(sheet.get("name") or "").strip().casefold() == "00_index"
        ),
        None,
    )
    grid = index.get("grid") if isinstance(index, dict) else None
    if not isinstance(grid, list):
        return {}
    header_row = -1
    headers: dict[str, int] = {}
    for row_number, row in enumerate(grid[:30]):
        found = {
            str(value or "").strip().casefold(): column
            for column, value in enumerate(row if isinstance(row, list) else [])
            if str(value or "").strip()
        }
        if {"include", "sheet tab", "page title"}.issubset(found):
            header_row = row_number
            headers = found
            break
    if header_row < 0:
        return {}

    def value(row: list[Any], label: str) -> Any:
        column = headers.get(label)
        return row[column] if column is not None and column < len(row) else ""

    output: dict[str, dict[str, Any]] = {}
    for raw in grid[header_row + 1 :]:
        row = raw if isinstance(raw, list) else []
        tab = str(value(row, "sheet tab") or "").strip()
        if not tab:
            continue
        output[tab.casefold()] = {
            "include": value(row, "include"),
            "sheetCode": value(row, "sheet code"),
            "title": value(row, "page title"),
            "pageType": value(row, "page type"),
            "role": value(row, "source role"),
        }
    return output


def _apply_document_contract(
    document: dict[str, Any],
    project: dict[str, Any],
) -> dict[str, Any]:
    output = dict(document)
    entries = _project_index_entries(project)
    index_sheet = next(
        (
            sheet
            for sheet in (output.get("sheets") or [])
            if str(sheet.get("name") or "").strip().casefold() == "00_index"
        ),
        None,
    )
    if isinstance(index_sheet, dict):
        rows: dict[int, dict[int, str]] = {}
        for coordinate, cell in (index_sheet.get("cells") or {}).items():
            match = re.fullmatch(r"([A-Z]+)(\d+)", str(coordinate), re.I)
            if not match or not isinstance(cell, dict):
                continue
            column_number = 0
            for character in match.group(1).upper():
                column_number = column_number * 26 + ord(character) - 64
            parsed = (int(match.group(2)), column_number)
            text = str(cell.get("v") or "").strip()
            if text:
                rows.setdefault(parsed[0], {})[parsed[1]] = text
        header_row = 0
        headers: dict[str, int] = {}
        for row_number, values in rows.items():
            labels = {value.casefold(): column for column, value in values.items()}
            if (
                any(key in labels for key in ("sheet tab", "sheet name", "tab"))
                and any(key in labels for key in ("page title", "sheet title", "title"))
            ):
                header_row = row_number
                headers = labels
                break
        if header_row and headers:
            def column(*aliases: str) -> int:
                return next((headers[item] for item in aliases if item in headers), -1)

            tab_column = column("sheet tab", "sheet name", "tab")
            for row_number, values in rows.items():
                if row_number <= header_row or tab_column < 0:
                    continue
                tab = values.get(tab_column, "").strip()
                if not tab:
                    continue
                def value(*aliases: str) -> str:
                    index = column(*aliases)
                    return values.get(index, "") if index >= 0 else ""
                entries[tab.casefold()] = {
                    "include": value("include", "include / publish", "publish"),
                    "sheetCode": value("sheet code", "sheet no.", "sheet no", "code"),
                    "title": value("page title", "sheet title", "title"),
                    "pageType": value("page type", "type"),
                    "role": value("source role", "role"),
                }
    sheets: list[dict[str, Any]] = []
    for raw in output.get("sheets") or []:
        sheet = dict(raw)
        entry = entries.get(str(sheet.get("name") or "").casefold(), {})
        if is_source_sheet(
            sheet.get("name"),
            entry.get("pageType"),
            entry.get("role"),
        ):
            sheet = apply_source_sheet_contract(sheet, entry)
        sheet = apply_controlled_default_validations(sheet)
        sheets.append(sheet)
    output["sheets"] = sheets
    return output


def workbook_file_to_document(path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        path,
        data_only=False,
        keep_vba=path.suffix.casefold() == ".xlsm",
    )
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        cells: dict[str, Any] = {}
        styles: dict[str, dict[str, Any]] = {}
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, 5_000),
            max_col=min(sheet.max_column, 200),
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    value = _json_cell_value(cell.value)
                    cells[cell.coordinate] = (
                        {"f": value}
                        if isinstance(value, str) and value.startswith("=")
                        else {"v": value}
                    )
                if cell.has_style:
                    style = _cell_style(cell)
                    if style:
                        styles[cell.coordinate] = style
        default_column_width = (
            getattr(sheet.sheet_format, "defaultColWidth", None)
            or DEFAULT_COLUMN_WIDTH_UNITS
        )
        default_row_height = (
            getattr(sheet.sheet_format, "defaultRowHeight", None)
            or DEFAULT_ROW_HEIGHT_POINTS
        )
        tab_color = _color_hex(getattr(sheet.sheet_properties, "tabColor", None))
        sheets.append(
            {
                "id": uuid.uuid4().hex[:16],
                "name": sheet.title,
                "cells": cells,
                "styles": styles,
                "merges": [str(item) for item in sheet.merged_cells.ranges],
                "rowHeights": {
                    str(index): dimension.height
                    for index, dimension in sheet.row_dimensions.items()
                    if dimension.height is not None
                },
                "columnWidths": {
                    key: dimension.width
                    for key, dimension in sheet.column_dimensions.items()
                    if dimension.width is not None
                },
                "defaultColumnWidth": float(default_column_width),
                "defaultRowHeight": float(default_row_height),
                "hiddenRows": sorted(
                    int(index)
                    for index, dimension in sheet.row_dimensions.items()
                    if dimension.hidden
                ),
                "hiddenColumns": sorted(
                    str(key)
                    for key, dimension in sheet.column_dimensions.items()
                    if dimension.hidden
                ),
                "archived": sheet.sheet_state != "visible",
                "tabColor": tab_color,
                "protectedRanges": [],
                "dataValidations": _data_validations_payload(sheet),
                "conditionalFormats": _conditional_formats_payload(sheet),
                "tableRegions": [],
                "tableLayout": "single",
                "annotations": [],
            }
        )
    workbook.close()
    return {"revision": 0, "updatedAt": utcnow(), "sheets": sheets}


def project_to_workbook_document(project: dict[str, Any]) -> dict[str, Any]:
    def coordinate(row: int, column: int) -> str:
        current = column + 1
        letters = ""
        while current:
            current, remainder = divmod(current - 1, 26)
            letters = chr(65 + remainder) + letters
        return f"{letters}{row + 1}"

    def dimensions(
        explicit: Any,
        pixels: Any,
        *,
        columns: bool,
    ) -> dict[str, float]:
        output: dict[str, float] = {}
        value = explicit
        if isinstance(value, dict):
            output.update({
                str(key): float(item)
                for key, item in value.items()
                if isinstance(item, (int, float)) and item > 0
            })
        elif isinstance(value, list):
            for index, item in enumerate(value):
                if not isinstance(item, (int, float)) or item <= 0:
                    continue
                key = coordinate(0, index)[:-1] if columns else str(index + 1)
                output[key] = (
                    pixels_to_excel_column_width(item)
                    if columns
                    else pixels_to_row_height_points(item)
                )
        if isinstance(pixels, list):
            for index, item in enumerate(pixels):
                if not isinstance(item, (int, float)) or item <= 0:
                    continue
                key = coordinate(0, index)[:-1] if columns else str(index + 1)
                output.setdefault(
                    key,
                    (
                        pixels_to_excel_column_width(item)
                        if columns
                        else pixels_to_row_height_points(item)
                    ),
                )
        return output

    sheets: list[dict[str, Any]] = []
    for index, worksheet in enumerate(project.get("worksheets") or []):
        cells: dict[str, Any] = {}
        formulas = (
            worksheet.get("formulas")
            if isinstance(worksheet.get("formulas"), dict)
            else {}
        )
        for row_index, row in enumerate(worksheet.get("grid") or [], 1):
            for column_index, value in enumerate(row, 1):
                current = column_index
                letters = ""
                while current:
                    current, remainder = divmod(current - 1, 26)
                    letters = chr(65 + remainder) + letters
                cell_address = f"{letters}{row_index}"
                formula = formulas.get(cell_address)
                if isinstance(formula, str) and formula.startswith("="):
                    cells[cell_address] = {"f": formula}
                elif value not in (None, ""):
                    cells[cell_address] = {"v": value}
        hidden_rows = sorted(
            {
                int(row) + 1
                for row in (worksheet.get("hiddenRows") or [])
                if isinstance(row, int) and row >= 0
            }
        )
        hidden_columns = sorted(
            {
                coordinate(0, int(column))[:-1]
                for column in (worksheet.get("hiddenColumns") or [])
                if isinstance(column, int) and column >= 0
            }
        )
        sheets.append(
            {
                "id": str(
                    worksheet.get("id")
                    or hashlib.sha256(
                        f"worksheet:{index}".encode("utf-8")
                    ).hexdigest()[:16]
                ),
                "name": str(worksheet.get("name") or f"Sheet {index + 1}"),
                "cells": cells,
                "styles": dict(worksheet.get("styles") or {}),
                "merges": [
                    (
                        f"{coordinate(int(item['startRow']), int(item['startCol']))}:"
                        f"{coordinate(int(item['endRow']), int(item['endCol']))}"
                    )
                    if isinstance(item, dict)
                    and all(
                        key in item
                        for key in ("startRow", "startCol", "endRow", "endCol")
                    )
                    else str(item)
                    for item in (worksheet.get("mergedCells") or [])
                ],
                "rowHeights": dimensions(
                    worksheet.get("rowHeights") or {},
                    worksheet.get("rowHeightsPx") or [],
                    columns=False,
                ),
                "columnWidths": dimensions(
                    worksheet.get("columnWidths") or {},
                    worksheet.get("colWidthsPx") or [],
                    columns=True,
                ),
                "defaultColumnWidth": float(
                    worksheet.get("defaultColumnWidth")
                    or DEFAULT_COLUMN_WIDTH_UNITS
                ),
                "defaultRowHeight": float(
                    worksheet.get("defaultRowHeight")
                    or DEFAULT_ROW_HEIGHT_POINTS
                ),
                "hiddenRows": hidden_rows,
                "hiddenColumns": hidden_columns,
                "archived": not bool(worksheet.get("visible", True)),
                "tabColor": worksheet.get("tabColor"),
                "role": worksheet.get("role"),
                "sourceSetup": deepcopy(
                    worksheet.get("sourceSetup") or {}
                ),
                "protectedRanges": list(
                    worksheet.get("protectedRanges") or []
                ),
                "dataValidations": deepcopy(
                    worksheet.get("dataValidations") or []
                ),
                "conditionalFormats": deepcopy(
                    worksheet.get("conditionalFormats") or []
                ),
                "tableRegions": deepcopy(
                    worksheet.get("tableRegions") or []
                ),
                "tableLayout": worksheet.get("tableLayout") or "single",
                "annotations": deepcopy(
                    worksheet.get("annotations") or []
                ),
            }
        )
    return _apply_document_contract(
        {"revision": 0, "updatedAt": utcnow(), "sheets": sheets},
        project,
    )


def _positive_dimensions(value: Any) -> dict[str, float]:
    if not isinstance(value, dict):
        return {}
    output: dict[str, float] = {}
    for key, raw in value.items():
        try:
            number = float(raw)
        except (TypeError, ValueError):
            continue
        if number > 0:
            output[str(key)] = number
    return output


def _normalize_workbook_document(document: dict[str, Any]) -> dict[str, Any]:
    """Read legacy workbook mirrors without rewriting their project package."""
    normalized = dict(document)
    normalized["revision"] = int(normalized.get("revision") or 0)
    normalized.setdefault("updatedAt", utcnow())
    sheets: list[dict[str, Any]] = []
    for index, raw_sheet in enumerate(normalized.get("sheets") or []):
        if not isinstance(raw_sheet, dict):
            continue
        sheet = dict(raw_sheet)
        sheet["id"] = str(sheet.get("id") or uuid.uuid4().hex[:16])
        sheet["name"] = str(sheet.get("name") or f"Sheet {index + 1}")
        sheet["cells"] = (
            dict(sheet.get("cells")) if isinstance(sheet.get("cells"), dict) else {}
        )
        sheet["styles"] = (
            dict(sheet.get("styles"))
            if isinstance(sheet.get("styles"), dict)
            else {}
        )
        sheet["merges"] = [
            str(item) for item in (sheet.get("merges") or []) if str(item)
        ]
        sheet["rowHeights"] = _positive_dimensions(sheet.get("rowHeights"))
        sheet["columnWidths"] = _positive_dimensions(
            sheet.get("columnWidths")
        )
        sheet["defaultColumnWidth"] = float(
            sheet.get("defaultColumnWidth") or DEFAULT_COLUMN_WIDTH_UNITS
        )
        sheet["defaultRowHeight"] = float(
            sheet.get("defaultRowHeight") or DEFAULT_ROW_HEIGHT_POINTS
        )
        sheet["hiddenRows"] = sorted(
            {
                int(item)
                for item in (sheet.get("hiddenRows") or [])
                if isinstance(item, int) and item >= 1
            }
        )
        sheet["hiddenColumns"] = sorted(
            {
                str(item).upper()
                for item in (sheet.get("hiddenColumns") or [])
                if re.fullmatch(r"[A-Za-z]+", str(item))
            }
        )
        sheet["archived"] = bool(sheet.get("archived"))
        sheet["tabColor"] = sheet.get("tabColor") or None
        sheet["role"] = str(sheet.get("role") or "") or None
        sheet["sourceSetup"] = (
            dict(sheet.get("sourceSetup"))
            if isinstance(sheet.get("sourceSetup"), dict)
            else {}
        )
        for field in (
            "protectedRanges",
            "dataValidations",
            "conditionalFormats",
            "tableRegions",
            "annotations",
        ):
            sheet[field] = (
                deepcopy(sheet.get(field))
                if isinstance(sheet.get(field), list)
                else []
            )
        layout = str(sheet.get("tableLayout") or "single")
        sheet["tableLayout"] = (
            layout
            if layout in {"single", "side_by_side", "stacked"}
            else "single"
        )
        sheets.append(sheet)
    normalized["sheets"] = sheets
    return normalized


_DRAWING_INDEX_HEADERS = [
    "Include",
    "Order",
    "Sheet Code",
    "Sheet Tab",
    "Page Title",
    "Family",
    "Page Type",
    "Notes",
    "Render Profile",
    "Split Mode",
    "Page ID",
    "Parent Page ID",
    "Issue Status",
    "Source Mode",
    "Sync Direction",
]


def _generated_drawing_page(page: dict[str, Any]) -> bool:
    if (
        page.get("generatedContinuation")
        or page.get("indexContinuation")
        or page.get("generatedIndexContinuation")
    ):
        return True
    page_id = str(page.get("id") or "")
    source_mode = str(page.get("sourceMode") or "").strip().casefold()
    return bool(
        page.get("continuationOf")
        and source_mode != "workbook"
        and (
            "__continuation_" in page_id
            or re.search(r"_c\d+(?:_\d+)?$", page_id)
        )
    )


_INVALID_WORKSHEET_TITLE = re.compile(r"[\[\]:*?/\\]")


def _unique_worksheet_tab(
    title: Any,
    used_tabs: set[str],
) -> str:
    """Return a valid, case-insensitively unique Excel worksheet title."""
    base = _INVALID_WORKSHEET_TITLE.sub(" ", str(title or "")).strip().strip("'")
    base = re.sub(r"\s+", " ", base) or "New Sheet"
    base = base[:31].rstrip() or "New Sheet"
    candidate = base
    suffix_number = 2
    while candidate.casefold() in used_tabs:
        suffix = f" ({suffix_number})"
        candidate = f"{base[: 31 - len(suffix)].rstrip()}{suffix}"
        suffix_number += 1
    used_tabs.add(candidate.casefold())
    return candidate


def _ensure_base_page_worksheet_identities(project: dict[str, Any]) -> None:
    """Give new base pages a stable worksheet identity before reconciliation.

    Sheet code is intentionally not part of this requirement. A user may leave
    the optional code blank until the explicit renumber workflow.
    """
    pages = [
        page
        for page in project.get("pages") or []
        if isinstance(page, dict) and not _generated_drawing_page(page)
    ]
    used_tabs = {
        str(page.get("sheetTab") or "").strip().casefold()
        for page in pages
        if str(page.get("sheetTab") or "").strip()
    }
    used_tabs.update(
        str(worksheet.get("name") or "").strip().casefold()
        for worksheet in project.get("worksheets") or []
        if isinstance(worksheet, dict)
        and str(worksheet.get("name") or "").strip()
    )
    for page in pages:
        page_id = str(page.get("id") or "").strip()
        if not str(page.get("sheetTab") or "").strip():
            page["sheetTab"] = _unique_worksheet_tab(
                page.get("sheetTitle") or "New Sheet",
                used_tabs,
            )
        if page_id and not str(page.get("linkedWorksheetId") or "").strip():
            page["linkedWorksheetId"] = "worksheet_" + hashlib.sha256(
                page_id.encode("utf-8")
            ).hexdigest()[:16]


def project_base_drawing_manifest(
    project: dict[str, Any],
) -> list[dict[str, Any]]:
    """Return the exact base drawing manifest controlled by project ``00_INDEX``."""
    ordered = [
        (position, page)
        for position, page in enumerate(project.get("pages") or [])
        if isinstance(page, dict) and not _generated_drawing_page(page)
    ]
    ordered.sort(
        key=lambda item: (
            int(item[1].get("order") or 10**9),
            item[0],
        )
    )
    manifest: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_tabs: set[str] = set()
    for order, (_, page) in enumerate(ordered, start=1):
        page_id = str(page.get("id") or "").strip()
        sheet_tab = str(
            page.get("sheetTab") or page.get("sourceSheet") or ""
        ).strip()
        sheet_code = str(
            page.get("displaySheetCode") or page.get("sheetCode") or ""
        ).strip()
        if not page_id:
            raise ProjectWorkspaceError(
                f"Base drawing page at order {order} has no stable Page ID."
            )
        if page_id.casefold() in seen_ids:
            raise ProjectWorkspaceError(
                f"Duplicate base Page ID prevents order reconciliation: {page_id}."
            )
        if not sheet_tab:
            raise ProjectWorkspaceError(
                f"Base drawing page {page_id} has no worksheet tab name."
            )
        if sheet_tab.casefold() in seen_tabs:
            raise ProjectWorkspaceError(
                "Duplicate drawing worksheet tab prevents order reconciliation: "
                f"{sheet_tab}."
            )
        seen_ids.add(page_id.casefold())
        seen_tabs.add(sheet_tab.casefold())
        publish = str(page.get("publishStatus") or "").strip().upper()
        if publish not in {"YES", "NO", "VERIFY"}:
            publish = "YES" if page.get("include", True) else "NO"
        manifest.append(
            {
                "pageId": page_id,
                "worksheetId": str(page.get("linkedWorksheetId") or "").strip(),
                "sheetCode": sheet_code,
                "sheetTab": sheet_tab,
                "title": str(page.get("sheetTitle") or sheet_tab).strip(),
                "include": publish == "YES",
                "publishStatus": publish,
                "order": order,
                "family": str(page.get("pageFamily") or ""),
                "pageType": str(page.get("pageType") or ""),
                "notes": str(page.get("notes") or ""),
                "renderProfile": str(
                    page.get("renderProfile")
                    or page.get("layoutProfile")
                    or ""
                ),
                "splitMode": str(page.get("splitMode") or ""),
                "parentPageId": str(page.get("parentPageId") or ""),
                "issueStatus": str(page.get("issueStatus") or "draft"),
                "sourceMode": str(page.get("sourceMode") or ""),
                "syncDirection": str(page.get("syncDirection") or "Both"),
            }
        )
    return manifest


def _index_row_values(item: dict[str, Any]) -> dict[str, Any]:
    status = str(item.get("issueStatus") or "draft").strip()
    status = status.replace("_", " ").replace("-", " ").title()
    return {
        "Include": item.get("publishStatus") or (
            "YES" if item.get("include") else "NO"
        ),
        "Order": int(item.get("order") or 0),
        "Sheet Code": item.get("sheetCode") or "",
        "Sheet Tab": item.get("sheetTab") or "",
        "Page Title": item.get("title") or "",
        "Family": item.get("family") or "",
        "Page Type": item.get("pageType") or "",
        "Notes": item.get("notes") or "",
        "Render Profile": item.get("renderProfile") or "",
        "Split Mode": item.get("splitMode") or "",
        "Page ID": item.get("pageId") or "",
        "Parent Page ID": item.get("parentPageId") or "",
        "Issue Status": status,
        "Source Mode": item.get("sourceMode") or "",
        "Sync Direction": item.get("syncDirection") or "Both",
    }


def _reconcile_index_grid(
    worksheet: dict[str, Any],
    manifest: list[dict[str, Any]],
) -> None:
    grid = worksheet.get("grid")
    if not isinstance(grid, list):
        return
    header_row = -1
    headers: dict[str, int] = {}
    for row_number, raw in enumerate(grid[:30]):
        row = raw if isinstance(raw, list) else []
        found = {
            str(value or "").strip().casefold(): column
            for column, value in enumerate(row)
            if str(value or "").strip()
        }
        if {"include", "sheet tab", "page title"}.issubset(found):
            header_row = row_number
            headers = found
            break
    if header_row < 0:
        return
    header = list(grid[header_row]) if isinstance(grid[header_row], list) else []
    for label in _DRAWING_INDEX_HEADERS:
        key = label.casefold()
        if key not in headers:
            headers[key] = len(header)
            header.append(label)
    grid[header_row] = header
    max_column = max(headers.values(), default=-1) + 1
    required_rows = header_row + len(manifest) + 1
    while len(grid) < required_rows:
        grid.append([])
    for offset, item in enumerate(manifest, start=1):
        row_number = header_row + offset
        row = list(grid[row_number]) if isinstance(grid[row_number], list) else []
        if len(row) < max_column:
            row.extend([""] * (max_column - len(row)))
        for label, value in _index_row_values(item).items():
            row[headers[label.casefold()]] = value
        grid[row_number] = row
    for row_number in range(required_rows, len(grid)):
        row = list(grid[row_number]) if isinstance(grid[row_number], list) else []
        if len(row) < max_column:
            row.extend([""] * (max_column - len(row)))
        for label in _DRAWING_INDEX_HEADERS:
            row[headers[label.casefold()]] = ""
        grid[row_number] = row
    worksheet["grid"] = grid


def _column_letters(number: int) -> str:
    output = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        output = chr(65 + remainder) + output
    return output


def _reconcile_index_cells(
    sheet: dict[str, Any],
    manifest: list[dict[str, Any]],
) -> None:
    cells = sheet.setdefault("cells", {})
    rows: dict[int, dict[int, str]] = {}
    for coordinate, cell in cells.items():
        match = re.fullmatch(r"([A-Z]+)(\d+)", str(coordinate), re.I)
        if not match or not isinstance(cell, dict):
            continue
        column = 0
        for character in match.group(1).upper():
            column = column * 26 + ord(character) - 64
        value = str(cell.get("v") or "").strip()
        if value:
            rows.setdefault(int(match.group(2)), {})[column] = value
    header_row = 0
    headers: dict[str, int] = {}
    for row_number in sorted(rows):
        if row_number > 30:
            break
        found = {
            value.casefold(): column
            for column, value in rows[row_number].items()
        }
        if {"include", "sheet tab", "page title"}.issubset(found):
            header_row = row_number
            headers = found
            break
    if not header_row:
        return
    next_column = max(headers.values(), default=0) + 1
    for label in _DRAWING_INDEX_HEADERS:
        key = label.casefold()
        if key in headers:
            continue
        headers[key] = next_column
        cells[f"{_column_letters(next_column)}{header_row}"] = {"v": label}
        next_column += 1
    for offset, item in enumerate(manifest, start=1):
        row_number = header_row + offset
        for label, value in _index_row_values(item).items():
            cells[f"{_column_letters(headers[label.casefold()])}{row_number}"] = {
                "v": value
            }
    last_manifest_row = header_row + len(manifest)
    controlled_columns = set(headers.values())
    for coordinate in list(cells):
        match = re.fullmatch(r"([A-Z]+)(\d+)", str(coordinate), re.I)
        if not match or int(match.group(2)) <= last_manifest_row:
            continue
        column = 0
        for character in match.group(1).upper():
            column = column * 26 + ord(character) - 64
        if column in controlled_columns:
            cells.pop(coordinate, None)


def _identity_match(
    payloads: list[dict[str, Any]],
    item: dict[str, Any],
    used: set[int],
    *,
    document: bool,
) -> dict[str, Any] | None:
    available = [payload for payload in payloads if id(payload) not in used]
    stable_id = str(item.get("worksheetId") or "").strip().casefold()
    if stable_id:
        matches = [
            payload
            for payload in available
            if str(payload.get("id") or "").strip().casefold() == stable_id
        ]
        if len(matches) > 1:
            raise ProjectWorkspaceError(
                f"Ambiguous worksheet identity {item.get('worksheetId')}."
            )
        if matches:
            return matches[0]
    tab = str(item.get("sheetTab") or "").strip().casefold()
    if tab:
        matches = [
            payload
            for payload in available
            if str(payload.get("name") or "").strip().casefold() == tab
        ]
        if len(matches) > 1:
            raise ProjectWorkspaceError(
                f"Ambiguous worksheet tab {item.get('sheetTab')}."
            )
        if matches:
            return matches[0]
    code = str(item.get("sheetCode") or "").strip().casefold()
    if code:
        matches = []
        for payload in available:
            setup = payload.get("sourceSetup")
            payload_code = (
                setup.get("sheetCode")
                if isinstance(setup, dict)
                else payload.get("sheetCode")
            )
            if str(payload_code or "").strip().casefold() == code:
                matches.append(payload)
        if len(matches) > 1:
            raise ProjectWorkspaceError(
                f"Ambiguous worksheet sheet code {item.get('sheetCode')}."
            )
        if matches:
            return matches[0]
    return None


def _new_project_worksheet(item: dict[str, Any]) -> dict[str, Any]:
    stable = str(item.get("worksheetId") or "").strip()
    if not stable:
        stable = "worksheet_" + hashlib.sha256(
            str(item.get("pageId") or "").encode("utf-8")
        ).hexdigest()[:16]
    is_index = str(item.get("sheetTab") or "").strip().casefold() == "00_index"
    return {
        "id": stable,
        "name": item.get("sheetTab") or "Drawing Page",
        "sourceSheet": item.get("sheetTab") or "Drawing Page",
        "grid": [list(_DRAWING_INDEX_HEADERS)] if is_index else [],
        "styles": {},
        "formulas": {},
        "mergedCells": [],
        "rowHeights": {},
        "columnWidths": {},
        "defaultColumnWidth": DEFAULT_COLUMN_WIDTH_UNITS,
        "defaultRowHeight": DEFAULT_ROW_HEIGHT_POINTS,
        "hiddenRows": [],
        "hiddenColumns": [],
        "visible": True,
        "role": "drawing",
        "sourceSetup": {
            "authority": "00_INDEX",
            "sheetCode": item.get("sheetCode") or "",
            "title": item.get("title") or "",
        },
        "protectedRanges": [],
        "dataValidations": [],
        "conditionalFormats": [],
        "tableRegions": [],
        "tableLayout": "single",
        "annotations": [],
    }


def reconcile_project_workbook_order(
    project: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Reconcile project pages, the project ``00_INDEX``, and worksheet order."""
    reconciled = deepcopy(project)
    _ensure_base_page_worksheet_identities(reconciled)
    manifest = project_base_drawing_manifest(reconciled)
    page_by_id = {
        str(page.get("id") or ""): page
        for page in reconciled.get("pages") or []
        if isinstance(page, dict)
    }
    for item in manifest:
        page = page_by_id[item["pageId"]]
        page["order"] = item["order"]
        page["publishStatus"] = item["publishStatus"]
        page["include"] = item["include"]

    worksheets = [
        worksheet
        for worksheet in reconciled.get("worksheets") or []
        if isinstance(worksheet, dict)
    ]
    index = next(
        (
            worksheet
            for worksheet in worksheets
            if str(worksheet.get("name") or "").strip().casefold()
            == "00_index"
        ),
        None,
    )
    if index is not None:
        _reconcile_index_grid(index, manifest)

    drawing_worksheets: list[dict[str, Any]] = []
    used: set[int] = set()
    for item in manifest:
        worksheet = _identity_match(
            worksheets, item, used, document=False
        )
        if worksheet is None:
            worksheet = _new_project_worksheet(item)
            worksheets.append(worksheet)
            if str(worksheet.get("name") or "").casefold() == "00_index":
                _reconcile_index_grid(worksheet, manifest)
        worksheet["name"] = item["sheetTab"]
        worksheet["sourceSheet"] = item["sheetTab"]
        worksheet["role"] = "drawing"
        source_setup = worksheet.get("sourceSetup")
        if not isinstance(source_setup, dict):
            source_setup = {}
        source_setup.update(
            {
                "authority": "00_INDEX",
                "sheetCode": item["sheetCode"],
                "title": item["title"],
            }
        )
        worksheet["sourceSetup"] = source_setup
        used.add(id(worksheet))
        drawing_worksheets.append(worksheet)
    reconciled["worksheets"] = [
        *drawing_worksheets,
        *[worksheet for worksheet in worksheets if id(worksheet) not in used],
    ]
    return reconciled, manifest


def _new_drawing_document_sheet(item: dict[str, Any]) -> dict[str, Any]:
    stable = str(item.get("worksheetId") or "").strip()
    if not stable:
        stable = "drawing_" + hashlib.sha256(
            str(item.get("pageId") or "").encode("utf-8")
        ).hexdigest()[:16]
    return {
        "id": stable,
        "name": item.get("sheetTab") or "Drawing Page",
        "cells": {},
        "styles": {},
        "merges": [],
        "rowHeights": {},
        "columnWidths": {},
        "defaultColumnWidth": DEFAULT_COLUMN_WIDTH_UNITS,
        "defaultRowHeight": DEFAULT_ROW_HEIGHT_POINTS,
        "hiddenRows": [],
        "hiddenColumns": [],
        "archived": False,
        "tabColor": None,
        "role": "drawing",
        "sourceSetup": {},
        "protectedRanges": [],
        "dataValidations": [],
        "conditionalFormats": [],
        "tableRegions": [],
        "tableLayout": "single",
        "annotations": [],
    }


def reconcile_workbook_document_order(
    document: dict[str, Any],
    project: dict[str, Any],
    manifest: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Reorder only worksheet identities while preserving every sheet payload."""
    reconciled = _normalize_workbook_document(deepcopy(document))
    final_manifest = manifest or project_base_drawing_manifest(project)
    sheets = [
        sheet
        for sheet in reconciled.get("sheets") or []
        if isinstance(sheet, dict)
    ]
    index = next(
        (
            sheet
            for sheet in sheets
            if str(sheet.get("name") or "").strip().casefold() == "00_index"
        ),
        None,
    )
    if index is not None:
        _reconcile_index_cells(index, final_manifest)

    drawing_sheets: list[dict[str, Any]] = []
    used: set[int] = set()
    for item in final_manifest:
        sheet = _identity_match(sheets, item, used, document=True)
        if sheet is None:
            sheet = _new_drawing_document_sheet(item)
            sheets.append(sheet)
        used.add(id(sheet))
        sheet["name"] = item["sheetTab"]
        sheet["workspaceSection"] = "drawing"
        sheet["drawingPageId"] = item["pageId"]
        sheet["drawingSheetCode"] = item["sheetCode"]
        sheet["drawingOrder"] = item["order"]
        sheet["drawingTitle"] = item["title"]
        sheet["drawingInclude"] = item["include"]
        sheet["drawingPublishStatus"] = item["publishStatus"]
        drawing_sheets.append(sheet)
    remaining = [sheet for sheet in sheets if id(sheet) not in used]
    for sheet in remaining:
        name = str(sheet.get("name") or "").strip().casefold()
        sheet["workspaceSection"] = (
            "control" if name.startswith("00_") else "source"
        )
        sheet.pop("drawingPageId", None)
        sheet.pop("drawingSheetCode", None)
        sheet.pop("drawingOrder", None)
        sheet.pop("drawingTitle", None)
        sheet.pop("drawingInclude", None)
        sheet.pop("drawingPublishStatus", None)
    reconciled["sheets"] = [*drawing_sheets, *remaining]
    return _apply_document_contract(reconciled, project)


def drawing_workspace_sequence(
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    return [
        {
            "pageId": str(sheet.get("drawingPageId") or ""),
            "sheetCode": str(sheet.get("drawingSheetCode") or ""),
            "sheetTab": str(sheet.get("name") or ""),
            "title": str(sheet.get("drawingTitle") or ""),
            "include": bool(sheet.get("drawingInclude")),
            "publishStatus": str(sheet.get("drawingPublishStatus") or ""),
            "order": int(sheet.get("drawingOrder") or 0),
        }
        for sheet in document.get("sheets") or []
        if isinstance(sheet, dict)
        and sheet.get("workspaceSection") == "drawing"
    ]


class WorkbookDocumentStore:
    """JSON workbook mirror used by the browser-only Data Workspace."""

    def __init__(self, project_dir: Path):
        self.path = Path(project_dir) / "data" / "workbook.json"
        self.history = Path(project_dir) / "backups" / "workbook"

    def load(self, project: dict[str, Any]) -> dict[str, Any]:
        if self.path.is_file():
            return _apply_document_contract(
                _normalize_workbook_document(
                    json.loads(self.path.read_text("utf-8"))
                ),
                project,
            )
        return _apply_document_contract(
            _normalize_workbook_document(
                project_to_workbook_document(project)
            ),
            project,
        )

    def save(
        self,
        project: dict[str, Any],
        expected_revision: int,
        document: dict[str, Any],
    ) -> dict[str, Any]:
        current = self.load(project)
        if int(current.get("revision") or 0) != int(expected_revision):
            raise WorkbookRevisionConflict(
                "Data Workspace revision conflict: "
                f"expected {expected_revision}, current {current.get('revision', 0)}."
            )
        if self.path.is_file():
            self.history.mkdir(parents=True, exist_ok=True)
            shutil.copy2(
                self.path,
                self.history
                / f"workbook_{datetime.now().strftime('%Y%m%d-%H%M%S-%f')[:-3]}.json",
            )
        saved = reconcile_workbook_document_order(
            document,
            project,
        )
        saved["revision"] = int(current.get("revision") or 0) + 1
        saved["updatedAt"] = utcnow()
        atomic_json_write(self.path, saved)
        return saved

    def reconcile_order(
        self,
        project: dict[str, Any],
        manifest: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        prepared, changed = self.prepare_reconciled_order(project, manifest)
        if not changed:
            return prepared
        return self.commit_reconciled_order(prepared)

    def prepare_reconciled_order(
        self,
        project: dict[str, Any],
        manifest: list[dict[str, Any]] | None = None,
    ) -> tuple[dict[str, Any], bool]:
        """Build and validate the next workspace document without writing it."""
        current = self.load(project)
        reconciled = reconcile_workbook_document_order(
            current, project, manifest
        )
        comparable_current = {
            key: value
            for key, value in current.items()
            if key not in {"revision", "updatedAt"}
        }
        comparable_reconciled = {
            key: value
            for key, value in reconciled.items()
            if key not in {"revision", "updatedAt"}
        }
        if comparable_current == comparable_reconciled and self.path.is_file():
            return current, False
        reconciled["revision"] = int(current.get("revision") or 0) + 1
        reconciled["updatedAt"] = utcnow()
        return reconciled, True

    def commit_reconciled_order(
        self,
        reconciled: dict[str, Any],
    ) -> dict[str, Any]:
        """Atomically replace workbook.json with a prevalidated document."""
        if self.path.is_file():
            self.history.mkdir(parents=True, exist_ok=True)
            shutil.copy2(
                self.path,
                self.history
                / f"workbook_{datetime.now().strftime('%Y%m%d-%H%M%S-%f')[:-3]}.json",
            )
        atomic_json_write(self.path, reconciled)
        return reconciled

    def import_file(
        self,
        project: dict[str, Any],
        source_path: Path,
        display_name: str,
    ) -> dict[str, Any]:
        current = self.load(project)
        suffix = source_path.suffix.casefold()
        if suffix in {".xlsx", ".xlsm"}:
            imported = workbook_file_to_document(source_path)
        elif suffix == ".csv":
            with source_path.open(
                "r", encoding="utf-8-sig", errors="replace", newline=""
            ) as handle:
                rows = list(csv.reader(handle))
            cells: dict[str, Any] = {}
            for row_index, row in enumerate(rows[:5_000], 1):
                for column_index, value in enumerate(row[:200], 1):
                    current_column = column_index
                    letters = ""
                    while current_column:
                        current_column, remainder = divmod(
                            current_column - 1, 26
                        )
                        letters = chr(65 + remainder) + letters
                    if value != "":
                        cells[f"{letters}{row_index}"] = {"v": value}
            imported = {
                "sheets": [
                    {
                        "id": uuid.uuid4().hex[:16],
                        "name": Path(display_name).stem[:31],
                        "cells": cells,
                        "styles": {},
                        "merges": [],
                        "rowHeights": {},
                        "columnWidths": {},
                        "defaultColumnWidth": DEFAULT_COLUMN_WIDTH_UNITS,
                        "defaultRowHeight": DEFAULT_ROW_HEIGHT_POINTS,
                        "hiddenRows": [],
                        "hiddenColumns": [],
                        "archived": False,
                        "tabColor": None,
                        "protectedRanges": [],
                        "dataValidations": [],
                        "conditionalFormats": [],
                        "tableRegions": [],
                        "tableLayout": "single",
                        "annotations": [],
                    }
                ]
            }
        else:
            raise ProjectWorkspaceError(
                "Only XLSX, XLSM, and CSV files can be sent to Data Workspace."
            )

        existing_names = {
            str(sheet.get("name") or "").casefold()
            for sheet in current.get("sheets") or []
        }
        for sheet in imported.get("sheets") or []:
            base = str(sheet.get("name") or "Imported")[:27]
            candidate = base
            counter = 2
            while candidate.casefold() in existing_names:
                candidate = f"{base[:26]} {counter}"
                counter += 1
            sheet["name"] = candidate
            sheet["id"] = uuid.uuid4().hex[:16]
            existing_names.add(candidate.casefold())
            current.setdefault("sheets", []).append(sheet)
        return self.save(
            project,
            int(current.get("revision") or 0),
            current,
        )
