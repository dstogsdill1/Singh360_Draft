"""Safe Data Workspace contracts shared by import, persistence, and tests.

This module contains no runtime project paths.  It operates only on sanitized
workbook-document dictionaries so callers can use it for live projects without
ever opening or writing the linked Excel workbook.
"""
from __future__ import annotations

from copy import deepcopy
from hashlib import sha256
import json
import re
from typing import Any, Iterable

from openpyxl.utils import column_index_from_string, get_column_letter


SOURCE_TAB_COLOR = "7F8C8D"
EXCLUDED_TAB_COLOR = "9AA3AB"
TITLE_FILL = "F28C28"
INSTRUCTION_FILL = "E6EEF5"
LOCKED_FILL = "D9DEE3"
YES_FILL = "C6E0B4"
NO_FILL = "D9D9D9"
VERIFY_FILL = "FFE699"

INCLUDE_VALUES = ("YES", "NO", "VERIFY")
YES_NO_VALUES = ("YES", "NO")
LIFECYCLE_VALUES = (
    "Draft",
    "Draft Confirmed",
    "Public",
    "Public Confirmed",
)

_COORDINATE_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)
_RANGE_RE = re.compile(
    r"^(?:'[^']+'|[^!]+)!?"  # optional sheet prefix
    r"([A-Z]+)(\d+):([A-Z]+)(\d+)$",
    re.IGNORECASE,
)


def _openpyxl_color(value: Any) -> str | None:
    if value is None:
        return None
    color_type = getattr(value, "type", None)
    raw = getattr(value, "rgb", None)
    if color_type == "rgb" and raw:
        text = str(raw).upper().replace("#", "")
        return text[-6:] if len(text) >= 6 else None
    return None


def openpyxl_data_validations(sheet: Any) -> list[dict[str, Any]]:
    """Serialize worksheet validation rules without changing the workbook."""
    output: list[dict[str, Any]] = []
    container = getattr(sheet, "data_validations", None)
    for index, validation in enumerate(
        getattr(container, "dataValidation", None) or []
    ):
        ranges = [
            str(item)
            for item in getattr(getattr(validation, "ranges", None), "ranges", [])
        ]
        if not ranges:
            continue
        formula1 = getattr(validation, "formula1", None)
        values: list[str] = []
        if (
            str(getattr(validation, "type", "") or "").casefold() == "list"
            and isinstance(formula1, str)
            and formula1.startswith('"')
            and formula1.endswith('"')
        ):
            values = [item.strip() for item in formula1[1:-1].split(",")]
        output.append(
            {
                "id": f"excel-validation-{index + 1}",
                "ranges": ranges,
                "type": str(getattr(validation, "type", "") or ""),
                "operator": str(getattr(validation, "operator", "") or ""),
                "formula1": formula1,
                "formula2": getattr(validation, "formula2", None),
                "values": values,
                "allowBlank": bool(getattr(validation, "allow_blank", False)),
                "showDropdown": not bool(
                    getattr(validation, "showDropDown", False)
                ),
                "showErrorMessage": bool(
                    getattr(validation, "showErrorMessage", True)
                ),
                "error": str(getattr(validation, "error", "") or ""),
                "errorTitle": str(
                    getattr(validation, "errorTitle", "") or ""
                ),
            }
        )
    return output


def openpyxl_conditional_formats(sheet: Any) -> list[dict[str, Any]]:
    """Serialize portable conditional-format metadata for local round-trips."""
    output: list[dict[str, Any]] = []
    rules_by_range = getattr(
        getattr(sheet, "conditional_formatting", None),
        "_cf_rules",
        {},
    )
    for range_key, rules in rules_by_range.items():
        sqref = str(getattr(range_key, "sqref", None) or range_key)
        for rule in rules or []:
            differential = getattr(rule, "dxf", None)
            output.append(
                {
                    "ranges": [item for item in sqref.split() if item],
                    "type": str(getattr(rule, "type", "") or ""),
                    "operator": str(getattr(rule, "operator", "") or ""),
                    "formula": [
                        str(value)
                        for value in (getattr(rule, "formula", None) or [])
                    ],
                    "priority": int(getattr(rule, "priority", 0) or 0),
                    "stopIfTrue": bool(
                        getattr(rule, "stopIfTrue", False)
                    ),
                    "fill": _openpyxl_color(
                        getattr(
                            getattr(differential, "fill", None),
                            "fgColor",
                            None,
                        )
                    ),
                    "fontColor": _openpyxl_color(
                        getattr(
                            getattr(differential, "font", None),
                            "color",
                            None,
                        )
                    ),
                }
            )
    return output


def normalize_publish_value(value: Any) -> str:
    """Return the strict publication state; blanks and unknowns stay excluded."""
    raw = str(value or "").strip().upper()
    return raw if raw in INCLUDE_VALUES else ""


def publishes(value: Any) -> bool:
    """Only an explicit YES publishes."""
    return normalize_publish_value(value) == "YES"


def is_source_sheet(name: Any, page_type: Any = "", role: Any = "") -> bool:
    tab = str(name or "").strip().casefold()
    kind = f"{page_type or ''} {role or ''}".strip().casefold()
    return (
        tab.startswith("src")
        or kind == "source"
        or "source / reference" in kind
        or "attachment / reference" in kind
    )


def source_purpose(name: Any) -> str:
    """Give guarded source tabs a purpose without inventing engineering data."""
    tab = str(name or "").strip()
    key = tab.casefold()
    if "r-2.0" in key and "equipment" in key:
        return (
            "Paste refrigeration equipment, rack manufacturer/model, and "
            "condenser source data here."
        )
    if "r-2.1" in key and "circuit" in key:
        return "Paste refrigeration circuit source data here."
    if "r-2.2" in key and "disclosure" in key:
        return "Paste refrigerant disclosure source data here."
    return f"Paste source/reference data for {tab or 'this sheet'} here."


def _coordinate(value: str) -> tuple[int, int] | None:
    match = _COORDINATE_RE.fullmatch(str(value or "").strip())
    if not match:
        return None
    return int(match.group(2)), column_index_from_string(match.group(1))


def _a1(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _range_bounds(value: str) -> tuple[int, int, int, int] | None:
    raw = str(value or "").strip()
    if "!" in raw:
        raw = raw.rsplit("!", 1)[1]
    raw = raw.replace("$", "").strip("'")
    if ":" not in raw:
        raw = f"{raw}:{raw}"
    match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", raw, re.I)
    if not match:
        return None
    start_column = column_index_from_string(match.group(1))
    start_row = int(match.group(2))
    end_column = column_index_from_string(match.group(3))
    end_row = int(match.group(4))
    return (
        min(start_row, end_row),
        min(start_column, end_column),
        max(start_row, end_row),
        max(start_column, end_column),
    )


def ranges_overlap(left: str, right: str) -> bool:
    a = _range_bounds(left)
    b = _range_bounds(right)
    if not a or not b:
        return False
    return not (
        a[2] < b[0]
        or b[2] < a[0]
        or a[3] < b[1]
        or b[3] < a[1]
    )


def edit_overlaps_protected(
    edit_range: str,
    protected_ranges: Iterable[str],
) -> bool:
    return any(ranges_overlap(edit_range, item) for item in protected_ranges)


def _cell_value(sheet: dict[str, Any], coordinate: str) -> Any:
    cell = (sheet.get("cells") or {}).get(coordinate)
    return cell.get("v") if isinstance(cell, dict) else None


def _metadata_block(sheet: dict[str, Any]) -> tuple[int, int] | None:
    """Return the exact generated Field/Value/Notes block, if present."""
    setup_fields = {
        "source role",
        "page mapping",
        "sheet code",
        "sheet title",
        "page title",
        "page type",
        "render profile",
        "include",
        "publish",
        "purpose",
        "required",
        "manual paste needed",
        "convert?",
    }
    for row in range(1, 21):
        labels = [
            str(_cell_value(sheet, f"{column}{row}") or "").strip().casefold()
            for column in ("A", "B", "C")
        ]
        if labels != ["field", "value", "notes"]:
            continue
        next_field = str(_cell_value(sheet, f"A{row + 1}") or "").strip().casefold()
        if next_field not in setup_fields:
            continue
        end = row
        for candidate in range(row + 1, min(row + 80, 5001)):
            values = [
                _cell_value(sheet, f"{column}{candidate}")
                for column in ("A", "B", "C")
            ]
            if not any(value not in (None, "") for value in values):
                # The generated setup table is contiguous. Its first blank row
                # is the hard boundary so a user table beginning after one
                # spacer row can never be mistaken for metadata.
                break
            end = candidate
        return row, end
    return None


def _active_width(sheet: dict[str, Any], minimum: int = 8) -> int:
    maximum = minimum
    for coordinate in {
        *(sheet.get("cells") or {}).keys(),
        *(sheet.get("styles") or {}).keys(),
    }:
        parsed = _coordinate(coordinate)
        if parsed:
            maximum = max(maximum, parsed[1])
    for column in (sheet.get("columnWidths") or {}):
        try:
            maximum = max(maximum, column_index_from_string(str(column)))
        except ValueError:
            continue
    for merged in sheet.get("merges") or []:
        bounds = _range_bounds(str(merged))
        if bounds:
            maximum = max(maximum, bounds[3])
    return maximum


def _remove_generated_metadata(
    sheet: dict[str, Any],
) -> list[dict[str, str]]:
    block = _metadata_block(sheet)
    if not block:
        return []
    start, end = block
    metadata: list[dict[str, str]] = []
    for row in range(start + 1, end + 1):
        field = str(_cell_value(sheet, f"A{row}") or "").strip()
        value = str(_cell_value(sheet, f"B{row}") or "").strip()
        notes = str(_cell_value(sheet, f"C{row}") or "").strip()
        if field or value or notes:
            metadata.append({"field": field, "value": value, "notes": notes})

    for collection_name in ("cells", "styles"):
        collection = sheet.get(collection_name)
        if not isinstance(collection, dict):
            continue
        for coordinate in list(collection):
            parsed = _coordinate(coordinate)
            if parsed and start <= parsed[0] <= end and parsed[1] <= 3:
                collection.pop(coordinate, None)
    sheet["merges"] = [
        value
        for value in (sheet.get("merges") or [])
        if not (
            (bounds := _range_bounds(str(value)))
            and start <= bounds[0] <= end
            and bounds[2] <= end
            and bounds[3] <= 3
        )
    ]
    return metadata


def source_sheet_setup(
    sheet: dict[str, Any],
    index_entry: dict[str, Any] | None = None,
) -> dict[str, Any]:
    entry = index_entry or {}
    title = str(
        entry.get("title")
        or entry.get("pageTitle")
        or _cell_value(sheet, "A1")
        or sheet.get("name")
        or "Source Sheet"
    ).strip()
    sheet_code = str(
        entry.get("sheetCode")
        or entry.get("code")
        or ""
    ).strip()
    return {
        "authority": "00_INDEX",
        "sheetCode": sheet_code,
        "title": title,
        "pageType": str(entry.get("pageType") or "Source"),
        "publish": normalize_publish_value(
            entry.get("publish", entry.get("include", ""))
        ),
        "purpose": source_purpose(sheet.get("name")),
        "instruction": (
            f"{source_purpose(sheet.get('name'))} Editable paste canvas starts "
            "at A3. Gray cells are locked."
        ),
        "editableStartRow": 3,
    }


def apply_source_sheet_contract(
    sheet: dict[str, Any],
    index_entry: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Move only the generated metadata block and install protected row bands."""
    output = deepcopy(sheet)
    existing_setup = output.get("sourceSetup")
    setup = dict(existing_setup or {})
    setup = {**source_sheet_setup(output, index_entry), **setup}
    metadata = _remove_generated_metadata(output) if not existing_setup else []
    if metadata and not setup.get("metadata"):
        setup["metadata"] = metadata

    width = _active_width(output)
    last_column = get_column_letter(width)
    title_range = f"A1:{last_column}1"
    instruction_range = f"A2:{last_column}2"
    output["role"] = "Source"
    output["sourceSetup"] = setup
    output["tabColor"] = (
        EXCLUDED_TAB_COLOR
        if normalize_publish_value(setup.get("publish")) != "YES"
        else SOURCE_TAB_COLOR
    )
    cells = dict(output.get("cells") or {})
    styles = dict(output.get("styles") or {})
    cells["A1"] = {"v": setup["title"]}
    cells["A2"] = {"v": setup["instruction"]}
    for column in range(1, width + 1):
        title_cell = _a1(1, column)
        instruction_cell = _a1(2, column)
        styles[title_cell] = {
            **dict(styles.get(title_cell) or {}),
            "bold": True,
            "fontColor": "#FFFFFF",
            "fill": f"#{TITLE_FILL}",
            "hAlign": "center",
            "vAlign": "center",
            "locked": True,
        }
        styles[instruction_cell] = {
            **dict(styles.get(instruction_cell) or {}),
            "fill": f"#{INSTRUCTION_FILL}",
            "hAlign": "left",
            "vAlign": "center",
            "wrap": True,
            "locked": True,
        }
    output["cells"] = cells
    output["styles"] = styles
    output["merges"] = [
        value
        for value in (output.get("merges") or [])
        if not ranges_overlap(str(value), title_range)
        and not ranges_overlap(str(value), instruction_range)
    ] + [title_range, instruction_range]
    protected = [title_range, instruction_range]
    for coordinate, cell in cells.items():
        if isinstance(cell, dict) and str(cell.get("f") or "").startswith("="):
            protected.append(f"{coordinate}:{coordinate}")
            styles[coordinate] = {
                **dict(styles.get(coordinate) or {}),
                "fill": f"#{LOCKED_FILL}",
                "locked": True,
            }
    output["styles"] = styles
    output["protectedRanges"] = list(dict.fromkeys(protected))
    output.setdefault("tableRegions", [])
    output.setdefault("tableLayout", "single")
    output.setdefault("annotations", [])
    return output


def _document_headers(
    sheet: dict[str, Any],
    *,
    max_rows: int = 30,
) -> tuple[int, dict[str, int]]:
    rows: dict[int, dict[int, str]] = {}
    for coordinate, cell in (sheet.get("cells") or {}).items():
        parsed = _coordinate(coordinate)
        if not parsed or parsed[0] > max_rows or not isinstance(cell, dict):
            continue
        value = str(cell.get("v") or "").strip()
        if value:
            rows.setdefault(parsed[0], {})[parsed[1]] = value
    best_row = 0
    best: dict[str, int] = {}
    best_score = 0
    for row_number, values in rows.items():
        headers = {
            value.casefold(): column
            for column, value in values.items()
        }
        score = sum(
            label in headers
            for label in (
                "include",
                "publish",
                "sheet tab",
                "page title",
                "required",
                "manual paste needed",
                "convert?",
                "status",
                "lifecycle",
                "issue status",
            )
        )
        if score > best_score:
            best_row = row_number
            best = headers
            best_score = score
    return best_row, best


def apply_controlled_default_validations(
    sheet: dict[str, Any],
) -> dict[str, Any]:
    """Add strict controlled lists only where the workbook supplied no rule."""
    output = deepcopy(sheet)
    header_row, headers = _document_headers(output)
    if not header_row or not headers:
        return output
    validations = list(output.get("dataValidations") or [])

    def has_rule(target: str) -> bool:
        return any(
            ranges_overlap(target, value)
            for validation in validations
            if isinstance(validation, dict)
            for value in (validation.get("ranges") or [])
        )

    fields = (
        (("include", "include / publish", "publish"), INCLUDE_VALUES, True),
        (("required",), INCLUDE_VALUES, True),
        (("manual paste needed",), INCLUDE_VALUES, True),
        (("convert?", "convert"), YES_NO_VALUES, True),
        (("lifecycle", "issue status"), LIFECYCLE_VALUES, False),
    )
    for aliases, values, allow_blank in fields:
        column = next(
            (headers[alias] for alias in aliases if alias in headers),
            None,
        )
        if column is None:
            continue
        target = (
            f"{get_column_letter(column)}{header_row + 1}:"
            f"{get_column_letter(column)}500"
        )
        if has_rule(target):
            continue
        validations.append(
            {
                "id": f"default-{re.sub('[^a-z0-9]+', '-', aliases[0])}",
                "ranges": [target],
                "type": "list",
                "values": list(values),
                "formula1": f'"{",".join(values)}"',
                "allowBlank": allow_blank,
                "showDropdown": True,
                "showErrorMessage": True,
                "error": f"Choose one of: {', '.join(values)}.",
                "strict": True,
                "source": "singh360-default",
            }
        )
    output["dataValidations"] = validations

    include_column = next(
        (
            headers[alias]
            for alias in ("include", "include / publish", "publish")
            if alias in headers
        ),
        None,
    )
    lifecycle_column = headers.get("lifecycle", headers.get("issue status"))
    formats = list(output.get("conditionalFormats") or [])
    if include_column is not None and not any(
        isinstance(item, dict)
        and item.get("source") == "singh360-publish-state"
        for item in formats
    ):
        include_letter = get_column_letter(include_column)
        last_column = get_column_letter(max(headers.values(), default=include_column))
        body = f"A{header_row + 1}:{last_column}500"
        formats.extend(
            [
                {
                    "ranges": [
                        f"{include_letter}{header_row + 1}:"
                        f"{include_letter}500"
                    ],
                    "type": "text",
                    "operator": "equal",
                    "value": "YES",
                    "fill": f"#{YES_FILL}",
                    "source": "singh360-publish-state",
                },
                {
                    "ranges": [body],
                    "type": "formula",
                    "formula": [
                        f'=${include_letter}{header_row + 1}="NO"'
                    ],
                    "fill": f"#{NO_FILL}",
                    "source": "singh360-publish-state",
                },
                {
                    "ranges": [body],
                    "type": "formula",
                    "formula": [
                        f'=${include_letter}{header_row + 1}="VERIFY"'
                    ],
                    "fill": f"#{VERIFY_FILL}",
                    "source": "singh360-publish-state",
                },
                {
                    "ranges": [body],
                    "type": "formula",
                    "formula": [
                        f'=${include_letter}{header_row + 1}=""'
                    ],
                    "fill": f"#{NO_FILL}",
                    "source": "singh360-publish-state",
                },
            ]
        )
    if lifecycle_column is not None and not any(
        isinstance(item, dict)
        and item.get("source") == "singh360-lifecycle"
        for item in formats
    ):
        letter = get_column_letter(lifecycle_column)
        target = f"{letter}{header_row + 1}:{letter}500"
        for value, fill in (
            ("Draft", "#F28C28"),
            ("Draft Confirmed", "#76B852"),
            ("Public", "#2D7DD2"),
            ("Public Confirmed", "#14845A"),
        ):
            formats.append(
                {
                    "ranges": [target],
                    "type": "text",
                    "operator": "equal",
                    "value": value,
                    "fill": fill,
                    "source": "singh360-lifecycle",
                }
            )
    output["conditionalFormats"] = formats
    return output


def detect_table_regions(
    sheet: dict[str, Any],
    *,
    start_row: int = 3,
) -> list[dict[str, Any]]:
    """Detect value/formula islands split by wholly blank rows or columns."""
    occupied: set[tuple[int, int]] = set()
    for coordinate, cell in (sheet.get("cells") or {}).items():
        parsed = _coordinate(coordinate)
        if not parsed or parsed[0] < start_row:
            continue
        if isinstance(cell, dict) and (
            cell.get("v") not in (None, "")
            or str(cell.get("f") or "").startswith("=")
        ):
            occupied.add(parsed)
    if not occupied:
        return []
    min_row = min(row for row, _ in occupied)
    max_row = max(row for row, _ in occupied)
    min_column = min(column for _, column in occupied)
    max_column = max(column for _, column in occupied)

    nonblank_rows = {row for row, _ in occupied}
    nonblank_columns = {column for _, column in occupied}

    def segments(values: set[int], first: int, last: int) -> list[tuple[int, int]]:
        output: list[tuple[int, int]] = []
        begin: int | None = None
        for value in range(first, last + 1):
            if value in values and begin is None:
                begin = value
            if value not in values and begin is not None:
                output.append((begin, value - 1))
                begin = None
        if begin is not None:
            output.append((begin, last))
        return output

    row_segments = segments(nonblank_rows, min_row, max_row)
    column_segments = segments(nonblank_columns, min_column, max_column)
    regions: list[dict[str, Any]] = []
    for row_start, row_end in row_segments:
        for column_start, column_end in column_segments:
            island = {
                item
                for item in occupied
                if row_start <= item[0] <= row_end
                and column_start <= item[1] <= column_end
            }
            if not island:
                continue
            actual_row_start = min(item[0] for item in island)
            actual_row_end = max(item[0] for item in island)
            actual_column_start = min(item[1] for item in island)
            actual_column_end = max(item[1] for item in island)
            range_a1 = (
                f"{_a1(actual_row_start, actual_column_start)}:"
                f"{_a1(actual_row_end, actual_column_end)}"
            )
            regions.append(
                {
                    "id": f"table-{len(regions) + 1}",
                    "range": range_a1,
                    "label": f"Table {len(regions) + 1}",
                }
            )
    regions.sort(
        key=lambda item: _range_bounds(str(item["range"])) or (0, 0, 0, 0)
    )
    for index, region in enumerate(regions, start=1):
        region["id"] = f"table-{index}"
        region["label"] = f"Table {index}"
    return regions


def workbook_document_signature(document: dict[str, Any]) -> str:
    payload = deepcopy(document)
    payload.pop("revision", None)
    payload.pop("updatedAt", None)
    return sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
    ).hexdigest()
