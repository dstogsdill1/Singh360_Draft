from __future__ import annotations

from contextlib import contextmanager
from copy import deepcopy
from datetime import datetime, timezone
from hashlib import sha256
import json
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from core.page_identity import is_sheet_index_page
from core.workbook_geometry import (
    DEFAULT_COLUMN_WIDTH_UNITS,
    DEFAULT_ROW_HEIGHT_POINTS,
    unchanged_excel_width_or_converted,
    unchanged_row_height_or_converted,
)

HELP_VERSION = "2026.07.22-status-sync-1"
SCHEMA_VERSION = "5.0"
STATUS_LABELS = {
    "draft": "Draft",
    "draft_confirmed": "Draft Confirmed",
    "public": "Public",
    "public_confirmed": "Public Confirmed",
}
TAB_COLORS = {
    "draft": "F28C28",
    "draft_confirmed": "76B852",
    "public": "2D7DD2",
    "public_confirmed": "14845A",
    "excluded": "9AA3AB",
    "control": "252C34",
    "help": "276FA8",
}
INDEX_HEADERS = [
    "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
    "Page Type", "Notes", "Render Profile", "Split Mode", "Page ID",
    "Parent Page ID", "Issue Status", "Source Mode", "Sync Direction",
    "Last Sync UTC", "Workbook Hash", "App Hash",
]


class WorkbookSyncError(RuntimeError):
    pass


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def normalize_status(value: Any) -> str:
    raw = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return raw if raw in STATUS_LABELS else "draft"


def file_hash(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()




def _workbook_path(store: Any, project_id: str, project: dict[str, Any]) -> Path | None:
    # S360 external-workbook-link path priority.
    sync = project.get('workbookSync') if isinstance(project.get('workbookSync'), dict) else {}
    configured = str(sync.get('workbook') or '').strip()
    if configured:
        return Path(os.path.expandvars(os.path.expanduser(configured)))
    folder = store.sources_dir(project_id, "workbook", project)
    preferred = str(project.get("sourceWorkbookName") or project.get("metadata", {}).get("sourceFile") or "").strip()
    if preferred:
        candidate = folder / Path(preferred).name
        if candidate.is_file():
            return candidate
    files = sorted([*folder.glob("*.xlsx"), *folder.glob("*.xlsm")], key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


# S360 DEAD LOCK SELF-HEAL V13
def _pid_is_running(pid: int) -> bool:
    # Return True only when the recorded lock owner still exists.
    if pid <= 0:
        return False
    if os.name == "nt":
        try:
            import ctypes

            process_query_limited_information = 0x1000
            handle = ctypes.windll.kernel32.OpenProcess(
                process_query_limited_information,
                False,
                pid,
            )
            if not handle:
                return False
            ctypes.windll.kernel32.CloseHandle(handle)
            return True
        except Exception:
            return False
    try:
        os.kill(pid, 0)
        return True
    except (OSError, ProcessLookupError, PermissionError):
        return False




def _backup_workbook(path: Path, store: Any, project_id: str) -> Path:
    backup_dir = store.docs / "backups" / "workbook_status_sync" / project_id
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")[:-3]
    target = backup_dir / f"{stamp}_{path.name}"
    shutil.copy2(path, target)
    backups = sorted(backup_dir.glob(f"*_{path.name}"), key=lambda p: p.name)
    for old in backups[:-20]:
        old.unlink(missing_ok=True)
    return target


def _headers(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, max(ws.max_column, len(INDEX_HEADERS)) + 1):
        value = str(ws.cell(4, col).value or "").strip()
        if value:
            result[value.casefold()] = col
    for idx, header in enumerate(INDEX_HEADERS, start=1):
        if header.casefold() not in result:
            ws.cell(4, idx, header)
            result[header.casefold()] = idx
    return result


def _ensure_help_sheet(wb) -> None:
    if "00_HELP" in wb.sheetnames:
        ws = wb["00_HELP"]
    else:
        ws = wb.create_sheet("00_HELP", 2)
    ws.sheet_properties.tabColor = TAB_COLORS["help"]
    ws["A1"] = "SINGH360 DRAFT — QUICK HELP"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="252C34")
    if "A1:J2" not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells("A1:J2")
    ws["A3"] = f"Help version {HELP_VERSION}. This control tab never publishes."
    if "A3:J3" not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells("A3:J3")
    ws["A5"] = "PAGE STATUS — FOUR DISTINCT STAGES"
    if "A5:J5" not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells("A5:J5")
    ws["A5"].fill = PatternFill("solid", fgColor="F28C28")
    ws["A5"].font = Font(bold=True, color="FFFFFF")
    ws["A6"], ws["B6"], ws["C6"] = "Status", "Meaning", "Workbook tab"
    data = [
        ("Draft", "Initial creation and active development.", "Orange"),
        ("Draft Confirmed", "Engineer reviewed and confirmed the draft.", "Light green"),
        ("Public", "Approved to go out for bid or external review.", "Blue"),
        ("Public Confirmed", "Final approved publication before as-builts.", "Dark green"),
    ]
    for row, item in enumerate(data, start=7):
        for col, value in enumerate(item, start=1):
            ws.cell(row, col, value)
    ws["A12"] = "INCLUDE / EXCLUDE IS SEPARATE"
    if "A12:J12" not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells("A12:J12")
    ws["A12"].fill = PatternFill("solid", fgColor="276FA8")
    ws["A12"].font = Font(bold=True, color="FFFFFF")
    ws["A13"] = "Include = YES"
    ws["B13"] = "Appears in Sheet Index, Page X of Y, and export."
    ws["A14"] = "Include = NO"
    ws["B14"] = "Remains visible/editable, turns gray, and is omitted from export."
    ws["A17"] = "Open detailed app help"
    ws["B17"] = "http://127.0.0.1:8766/app?help=1"
    ws["B17"].hyperlink = ws["B17"].value
    ws["B17"].style = "Hyperlink"
    ws["A19"] = "Documentation rule"
    ws["B19"] = "The app and workbook Help Version must match. Tests fail when workflow code changes without help updates."
    for col, width in {"A": 24, "B": 84, "C": 18}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=1, max_row=24, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _ensure_meta(wb, project: dict[str, Any]) -> None:
    ws = wb["00_PROJECT_META"] if "00_PROJECT_META" in wb.sheetnames else wb.create_sheet("00_PROJECT_META", 0)
    ws.sheet_properties.tabColor = TAB_COLORS["control"]
    existing: dict[str, int] = {}
    for row in range(1, max(ws.max_row, 30) + 1):
        key = str(ws.cell(row, 1).value or "").strip()
        if key:
            existing[key.casefold()] = row
    values = {
        "Workbook Schema Version": SCHEMA_VERSION,
        "Help Version": HELP_VERSION,
        "Issue Status Model": "Draft | Draft Confirmed | Public | Public Confirmed",
        "Drawing Set Control": "Include / Exclude is separate from Issue Status",
        "Sync Mode": "One User / Two-Way Page Manifest",
        "Linked Project ID": project.get("id", ""),
        "Last Sync UTC": utcnow(),
        "Last Sync Status": "Synchronized",
    }
    next_row = max(ws.max_row + 1, 17)
    for key, value in values.items():
        row = existing.get(key.casefold())
        if row is None:
            row = next_row
            next_row += 1
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)
    row = existing.get("open help", next_row)
    ws.cell(row, 1, "Open Help")
    ws.cell(row, 2, "OPEN HELP — PAGE 3")
    ws.cell(row, 2).hyperlink = "#'00_HELP'!A1"
    ws.cell(row, 2).style = "Hyperlink"


def _ensure_index(wb):
    ws = wb["00_INDEX"] if "00_INDEX" in wb.sheetnames else wb.create_sheet("00_INDEX", 1)
    ws.sheet_properties.tabColor = TAB_COLORS["control"]
    headers = _headers(ws)
    ws["A2"] = "Only explicit YES rows publish. Issue Status is separate. Excluded pages remain visible/editable and gray."
    ws["K2"] = "OPEN HELP — PAGE 3"
    ws["K2"].hyperlink = "#'00_HELP'!A1"
    ws["K2"].style = "Hyperlink"
    status_col = headers["issue status"]
    include_col = headers["include"]
    status_letter = ws.cell(4, status_col).column_letter
    include_letter = ws.cell(4, include_col).column_letter
    status_dv = DataValidation(type="list", formula1='"Draft,Draft Confirmed,Public,Public Confirmed"', allow_blank=False)
    include_dv = DataValidation(
        type="list",
        formula1='"YES,NO,VERIFY"',
        allow_blank=True,
        errorStyle="stop",
        showErrorMessage=True,
        error="Choose YES, NO, or VERIFY. Only YES publishes.",
    )
    ws.add_data_validation(status_dv)
    ws.add_data_validation(include_dv)
    status_dv.add(f"{status_letter}5:{status_letter}500")
    include_dv.add(f"{include_letter}5:{include_letter}500")
    return ws, headers


def _row_maps(ws, headers: dict[str, int]):
    by_id: dict[str, int] = {}
    by_tab: dict[str, int] = {}
    for row in range(5, ws.max_row + 1):
        page_id = str(ws.cell(row, headers["page id"]).value or "").strip()
        tab = str(ws.cell(row, headers["sheet tab"]).value or "").strip()
        if page_id:
            by_id[page_id] = row
        if tab:
            by_tab[tab.casefold()] = row
    return by_id, by_tab


def _tab_color(include: bool, status: str) -> str:
    if not include:
        return TAB_COLORS["excluded"]
    return TAB_COLORS.get(status, TAB_COLORS["draft"])


def _safe_sheet_name(wb, requested: str) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", requested or "APP PAGE").strip()[:31] or "APP PAGE"
    if base not in wb.sheetnames:
        return base
    index = 2
    while True:
        suffix = f" {index}"
        candidate = (base[: 31 - len(suffix)] + suffix).strip()
        if candidate not in wb.sheetnames:
            return candidate
        index += 1


def _write_companion_sheet(wb, page: dict[str, Any], project_id: str) -> str:
    requested = str(page.get("sheetTab") or page.get("sheetTitle") or "APP PAGE")
    tab = _safe_sheet_name(wb, requested)
    ws = wb.create_sheet(tab)
    ws["A1"] = "SINGH360 APP PAGE COMPANION"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="252C34")
    ws.merge_cells("A1:H2")
    values = [
        ("Sheet Code", page.get("displaySheetCode") or page.get("sheetCode") or "NEW"),
        ("Page Title", page.get("sheetTitle") or "Untitled Sheet"),
        ("Page ID", page.get("id") or ""),
    ]
    for row, (label, value) in enumerate(values, start=4):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="E6EBEF")
    ws["A8"] = "This page is managed in Singh360 Draft. Drawings, highlights, crops, and canvas objects remain app-owned."
    ws.merge_cells("A8:H10")
    ws["A12"] = "Open project"
    ws["B12"] = f"http://127.0.0.1:8766/app?project={project_id}"
    ws["B12"].hyperlink = ws["B12"].value
    ws["B12"].style = "Hyperlink"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65
    return tab




# S360 WORKBOOK AUTHORITY V15
# The linked workbook owns workbook-driven structure and source worksheets.
# App-only canvas/assets remain protected in project.json and are merged back
# after every workbook refresh.

from contextlib import contextmanager as _s360_contextmanager
from copy import copy as _s360_copy
from datetime import date as _s360_date, datetime as _s360_datetime
import threading as _s360_threading

_S360_PROJECT_LOCKS: dict[str, _s360_threading.Lock] = {}


def project_hash(project: dict[str, Any]) -> str:
    """Hash all persisted editor state, not just the page manifest."""
    payload = {
        "metadata": project.get("metadata", {}),
        "pages": [
            {
                "id": p.get("id"),
                "order": p.get("order"),
                "include": p.get("include", True),
                "publishStatus": p.get("publishStatus", ""),
                "sheetCode": p.get("sheetCode"),
                "displaySheetCode": p.get("displaySheetCode"),
                "sheetTitle": p.get("sheetTitle"),
                "sheetTab": p.get("sheetTab"),
                "issueStatus": normalize_status(p.get("issueStatus")),
                "canvasObjects": p.get("canvasObjects", []),
                "assets": p.get("assets", []),
                "notes": p.get("notes", ""),
            }
            for p in project.get("pages", [])
            if isinstance(p, dict)
        ],
        "worksheets": [
            {
                "id": w.get("id"),
                "name": w.get("name"),
                "grid": w.get("grid", []),
                "formulas": w.get("formulas", {}),
                "styles": w.get("styles", {}),
                "mergedCells": w.get("mergedCells", []),
                "rowHeights": w.get("rowHeights", {}),
                "columnWidths": w.get("columnWidths", {}),
                "rowHeightsPx": w.get("rowHeightsPx", []),
                "colWidthsPx": w.get("colWidthsPx", []),
                "defaultColumnWidth": w.get("defaultColumnWidth"),
                "defaultRowHeight": w.get("defaultRowHeight"),
                "hiddenRows": w.get("hiddenRows", []),
                "hiddenColumns": w.get("hiddenColumns", []),
                "geometryAuthority": w.get("geometryAuthority"),
                "protectedRanges": w.get("protectedRanges", []),
                "dataValidations": w.get("dataValidations", []),
                "conditionalFormats": w.get("conditionalFormats", []),
                "tableRegions": w.get("tableRegions", []),
                "tableLayout": w.get("tableLayout", "single"),
                "annotations": w.get("annotations", []),
            }
            for w in project.get("worksheets", [])
            if isinstance(w, dict)
        ],
        "dataWorkspace": project.get("dataWorkspace", {}),
    }
    return sha256(
        json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str).encode("utf-8")
    ).hexdigest()


@_s360_contextmanager
def _project_lock(store: Any, project_id: str, project: dict[str, Any]):
    """One in-process sync at a time; remove stale crash locks safely."""
    gate = _S360_PROJECT_LOCKS.setdefault(project_id, _s360_threading.Lock())
    if not gate.acquire(timeout=180):
        raise WorkbookSyncError("Workbook save is still running. Wait a moment and save again.")
    project_dir = store.dir_for(project_id, project)
    project_dir.mkdir(parents=True, exist_ok=True)
    lock = project_dir / ".workbook-status-sync.lock"
    try:
        if lock.exists():
            age = datetime.now().timestamp() - lock.stat().st_mtime
            pid = 0
            try:
                pid = int(json.loads(lock.read_text(encoding="utf-8")).get("pid") or 0)
            except Exception:
                pid = 0
            # The in-process gate proves another V15 thread is not active. A lock
            # from this PID or an old process is stale. A fresh foreign PID gets a
            # brief safety window before it is treated as stale.
            if pid != os.getpid() and age < 120:
                raise WorkbookSyncError(f"Workbook save is already running. Lock: {lock}")
            lock.unlink(missing_ok=True)
        fd = os.open(str(lock), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(json.dumps({"pid": os.getpid(), "created": utcnow()}, indent=2))
        yield
    finally:
        lock.unlink(missing_ok=True)
        gate.release()




def _s360_page_match(page: dict[str, Any], manifest: list[dict[str, Any]]) -> dict[str, Any] | None:
    tab = str(page.get("sheetTab") or "").strip().casefold()
    code = str(page.get("sheetCode") or "").strip().casefold()
    for item in manifest:
        if tab and code and tab == str(item.get("sheetTab") or "").casefold() and code == str(item.get("sheetCode") or "").casefold():
            return item
    for item in manifest:
        if tab and tab == str(item.get("sheetTab") or "").casefold():
            return item
    for item in manifest:
        if code and code == str(item.get("sheetCode") or "").casefold():
            return item
    return None


def _normalize_continuation_identities(
    pages: list[dict[str, Any]],
    base_id_remap: dict[str, str],
) -> None:
    """Attach every generated page to its final controlled base Page ID."""
    reserved = {
        str(page.get("id") or "")
        for page in pages
        if not page.get("generatedContinuation") and not page.get("continuationOf")
    }
    used = set(reserved)
    for page in pages:
        if not page.get("generatedContinuation") and not page.get("continuationOf"):
            page["pageGroupId"] = page.get("id")
            page["continuationOf"] = None
            page["continuationIndex"] = 0
            continue

        old_base = str(page.get("continuationOf") or page.get("pageGroupId") or "")
        base_id = base_id_remap.get(old_base, old_base)
        page["pageGroupId"] = base_id
        page["continuationOf"] = base_id
        index = max(1, int(page.get("continuationIndex") or 1))
        candidate = f"{base_id}__continuation_{index}"
        discriminator = 2
        while candidate in used:
            candidate = f"{base_id}__continuation_{index}_{discriminator}"
            discriminator += 1
        page["id"] = candidate
        used.add(candidate)

        title = str(page.get("sheetTitle") or "Untitled Sheet")
        title = re.sub(
            r"(?:\s*[—-]\s*)?CONTINUED(?:\s*[—-]\s*CONTINUED)*\s*$",
            "",
            title,
            flags=re.IGNORECASE,
        ).strip(" —-")
        page["sheetTitle"] = f"{title} — CONTINUED"


def sync_project_from_workbook(project_id: str, project: dict[str, Any], store: Any) -> dict[str, Any]:
    """Import the complete workbook, including brand-new 00_INDEX pages."""
    path = _workbook_path(store, project_id, project)
    if path is None:
        raise WorkbookSyncError("No workbook is linked to this project.")
    if not path.is_file():
        raise WorkbookSyncError(f"The linked workbook was not found: {path}")

    with _project_lock(store, project_id, project):
        from core.project_model import ensure_project_shape, recalc_page_numbers
        from core.sheet_index_sync import sync_project_sheet_index
        from core.workbook_importer import import_workbook

        try:
            imported = import_workbook(
                path,
                project_id=project_id,
                assets_dir=store.assets_excel_dir(project_id, project),
                asset_url_prefix=f"/api/assets/{project_id}",
            )
        except PermissionError as exc:
            raise WorkbookSyncError(
                "The linked workbook is open or locked. Close Excel, then reopen the project."
            ) from exc
        except Exception as exc:
            raise WorkbookSyncError(
                f"The linked workbook could not rebuild the project: {type(exc).__name__}: {exc}"
            ) from exc

        manifest = _s360_index_manifest(path)
        old_pages = [deepcopy(p) for p in project.get("pages", []) if isinstance(p, dict)]
        old_worksheets = [deepcopy(w) for w in project.get("worksheets", []) if isinstance(w, dict)]
        old_by_id = {str(p.get("id") or ""): p for p in old_pages if p.get("id")}
        old_by_pair = {
            (str(p.get("sheetTab") or "").strip().casefold(), str(p.get("sheetCode") or "").strip().casefold()): p
            for p in old_pages
            if p.get("sheetTab") or p.get("sheetCode")
        }
        old_by_tab = {str(p.get("sheetTab") or "").strip().casefold(): p for p in old_pages if p.get("sheetTab")}
        old_by_code = {str(p.get("sheetCode") or "").strip().casefold(): p for p in old_pages if p.get("sheetCode")}
        old_ws_name_by_id = {
            str(w.get("id") or ""): str(w.get("name") or w.get("sourceSheet") or "")
            for w in old_worksheets
        }

        app_owned_fields = {
            "canvasObjects", "assets", "underlay", "underlays", "background", "overlays",
            "annotations", "pastedImages", "imageCrop", "crop", "crops", "masks",
            "highlightedCells", "manualObjects", "lockedObjects", "connectors",
        }
        workbook_owned_fields = {
            "order", "include", "publishStatus", "sheetCode", "displaySheetCode", "sheetTitle",
            "sheetTab", "pageType", "pageFamily", "layoutProfile", "renderMode",
            "renderProfile", "sourceSheet", "sourceRange", "printArea", "splitMode",
            "repeatRows", "minScale", "allowContinuation", "scaleMode", "orientation",
            "template", "templateId", "linkedWorksheetId", "blocks", "sourceRevision",
            "notes", "pageNumber", "pageTotal", "pageGroupId", "continuationOf",
            "continuationIndex", "generatedContinuation", "layoutWarnings", "issueStatus",
            "parentPageId", "sourceMode", "syncDirection",
        }

        imported_pages: list[dict[str, Any]] = []
        base_id_remap: dict[str, str] = {}
        used_old: set[int] = set()
        for raw in imported.get("pages", []):
            if not isinstance(raw, dict):
                continue
            page = deepcopy(raw)
            imported_id = str(page.get("id") or "")
            item = _s360_page_match(page, manifest) if not page.get("continuationOf") else None
            if item:
                for key in (
                    "include", "publishStatus", "order", "sheetCode", "displaySheetCode", "sheetTab",
                    "sheetTitle", "pageFamily", "notes", "renderProfile", "splitMode",
                    "parentPageId", "issueStatus", "sourceMode", "syncDirection",
                ):
                    if item.get(key) not in (None, ""):
                        page[key] = deepcopy(item[key])
                if item.get("pageType"):
                    page["pageType"] = item["pageType"]
                if item.get("id"):
                    page["id"] = item["id"]

            tab = str(page.get("sheetTab") or "").strip().casefold()
            code = str(page.get("sheetCode") or "").strip().casefold()
            existing = (
                old_by_id.get(str(page.get("id") or ""))
                or old_by_pair.get((tab, code))
                or old_by_tab.get(tab)
                or old_by_code.get(code)
            )
            if existing is not None and id(existing) not in used_old:
                used_old.add(id(existing))
                # Workbook Page IDs are stable base identities. Preserve an old
                # ID only for legacy rows whose controlled manifest has none.
                if existing.get("id") and not (item and item.get("id")):
                    page["id"] = existing["id"]
                for key, value in existing.items():
                    if key not in workbook_owned_fields and key not in page:
                        page[key] = deepcopy(value)
                for key in app_owned_fields:
                    if existing.get(key) not in (None, "", [], {}):
                        page[key] = deepcopy(existing[key])
            if not page.get("generatedContinuation") and not page.get("continuationOf"):
                base_id_remap[imported_id] = str(page.get("id") or imported_id)
            imported_pages.append(page)

        _normalize_continuation_identities(imported_pages, base_id_remap)
        for page in imported_pages:
            code_key = str(
                page.get("displaySheetCode") or page.get("sheetCode") or ""
            ).strip().casefold()
            if code_key.startswith("src ") or code_key == "template":
                page["include"] = False

        archived_manual: list[dict[str, Any]] = []
        # Preserve unmatched app-only pages. If a removed workbook page carries
        # manual work, archive it explicitly instead of silently dropping it or
        # publishing it as a duplicate base page.
        for existing in old_pages:
            if id(existing) in used_old:
                continue
            has_manual = any(existing.get(k) not in (None, "", [], {}) for k in app_owned_fields)
            app_only = (
                not existing.get("linkedWorksheetId")
                or str(existing.get("sourceMode") or "").strip().casefold() in {"app", "manual", "app-only"}
                or str(existing.get("pageType") or "").strip().casefold() in {"canvas", "underlay"}
            )
            if app_only:
                preserved = deepcopy(existing)
                code_key = str(
                    preserved.get("displaySheetCode")
                    or preserved.get("sheetCode")
                    or ""
                ).strip().casefold()
                if code_key.startswith("src ") or code_key == "template":
                    preserved["include"] = False
                preserved["order"] = len(imported_pages) + 1
                imported_pages.append(preserved)
                continue
            if has_manual:
                archived = deepcopy(existing)
                archived["archivedReason"] = (
                    "Workbook refresh could not safely map this page to a controlled base Page ID."
                )
                archived["archivedAt"] = utcnow()
                archived_manual.append(archived)

        old_ws_by_name = {
            str(w.get("name") or w.get("sourceSheet") or "").strip().casefold(): w
            for w in old_worksheets
        }
        merged_worksheets: list[dict[str, Any]] = []
        for raw in imported.get("worksheets", []):
            if not isinstance(raw, dict):
                continue
            worksheet = deepcopy(raw)
            old = old_ws_by_name.get(str(worksheet.get("name") or worksheet.get("sourceSheet") or "").strip().casefold())
            if old:
                for field in ("hiddenRows", "hiddenColumns", "hiddenCells"):
                    if old.get(field) not in (None, "", [], {}):
                        worksheet[field] = deepcopy(old[field])
            merged_worksheets.append(worksheet)

        merged = deepcopy(project)
        metadata = deepcopy(project.get("metadata") or {})
        metadata.update(deepcopy(imported.get("metadata") or {}))
        metadata["sourceFile"] = path.name
        merged["id"] = project_id
        merged["metadata"] = metadata
        merged["worksheets"] = merged_worksheets
        merged["pages"] = imported_pages
        merged["archivedPages"] = [
            *deepcopy(project.get("archivedPages") or []),
            *archived_manual,
        ]
        recalc_page_numbers(merged)
        merged["sourceWorkbookName"] = path.name
        merged["projectDisplayName"] = (
            imported.get("projectDisplayName")
            or metadata.get("projectName")
            or merged.get("projectDisplayName")
            or path.stem
        )
        merged = sync_project_sheet_index(ensure_project_shape(merged))

        # S360 GENERATED INDEX VALIDATION V15.3
        # Every explicitly included normal worksheet must import. The Sheet
        # Index is different: Singh360 regenerates it from the final page
        # manifest and may assign generated continuation codes/tabs such as
        # EMS 2.0a. Validate that an actual generated index page exists rather
        # than requiring the exact source worksheet pair.
        actual_pages = [
            item for item in merged.get("pages", [])
            if isinstance(item, dict)
        ]
        actual_pairs = {
            (
                str(
                    item.get("sheetCode")
                    or item.get("displaySheetCode")
                    or ""
                ).strip().casefold(),
                str(item.get("sheetTab") or "").strip().casefold(),
            )
            for item in actual_pages
        }

        has_generated_index = any(is_sheet_index_page(item) for item in actual_pages)
        missing: list[str] = []
        for item in manifest:
            if not item.get("include"):
                continue

            code = str(item.get("sheetCode") or "").strip()
            tab = str(item.get("sheetTab") or "").strip()
            title = str(item.get("sheetTitle") or "").strip()
            page_type = str(item.get("pageType") or "").strip()

            code_key = code.casefold()
            tab_key = tab.casefold()
            if is_sheet_index_page(item):
                if not has_generated_index:
                    missing.append(f"{code} / {tab}")
                continue

            if (code_key, tab_key) not in actual_pairs:
                missing.append(f"{code} / {tab}")

        if missing:
            raise WorkbookSyncError(
                "Workbook import omitted included 00_INDEX pages: "
                + "; ".join(missing[:20])
            )

        sync = dict(merged.get("workbookSync") or {})
        sync.update({
            "mode": "external-workbook-link",
            "workbook": str(path),
            "status": "in_sync",
            "warning": "",
            "lastSyncUtc": utcnow(),
            "workbookHash": file_hash(path),
            "appHash": project_hash(merged),
            "authority": "workbook",
            "lastAuthorityAction": "workbook_to_app",
        })
        merged["workbookSync"] = sync
        store.save(project_id, merged)
        return merged


def _s360_cell_display(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (_s360_datetime, _s360_date)):
        return value.isoformat()
    return str(value)


def _s360_coerce(value: Any, existing: Any) -> Any:
    text = "" if value is None else str(value)
    if text == "":
        return None
    if isinstance(existing, bool):
        return text.strip().casefold() in {"true", "yes", "y", "1", "x"}
    if isinstance(existing, int) and not isinstance(existing, bool):
        try:
            return int(float(text.replace(",", "")))
        except Exception:
            return text
    if isinstance(existing, float):
        try:
            return float(text.replace(",", ""))
        except Exception:
            return text
    if isinstance(existing, (_s360_datetime, _s360_date)):
        try:
            return _s360_datetime.fromisoformat(text)
        except Exception:
            return text
    stripped = text.strip()
    if stripped.casefold() in {"true", "false"}:
        return stripped.casefold() == "true"
    try:
        if stripped and re.fullmatch(r"[-+]?\d+", stripped.replace(",", "")):
            return int(stripped.replace(",", ""))
        if stripped and re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+)", stripped.replace(",", "")):
            return float(stripped.replace(",", ""))
    except Exception:
        pass
    return text


def _s360_hex(value: Any) -> str | None:
    text = str(value or "").strip().lstrip("#")
    if len(text) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in text):
        return text.upper()
    return None


def _s360_apply_worksheet_payload(wb: Any, project: dict[str, Any]) -> None:
    from openpyxl.cell.cell import MergedCell as _S360MergedCell
    from openpyxl.styles import Border as _S360Border, PatternFill as _S360Fill, Side as _S360Side
    from openpyxl.utils import get_column_letter as _s360_col

    for payload in project.get("worksheets", []):
        if not isinstance(payload, dict):
            continue
        name = str(payload.get("sourceSheet") or payload.get("name") or "").strip()
        if not name or name in {"00_PROJECT_META", "00_INDEX", "00_HELP"} or name not in wb.sheetnames:
            continue
        ws = wb[name]
        grid = payload.get("grid") if isinstance(payload.get("grid"), list) else []
        formulas = payload.get("formulas") if isinstance(payload.get("formulas"), dict) else {}
        styles = payload.get("styles") if isinstance(payload.get("styles"), dict) else {}
        for row_number, row in enumerate(grid, start=1):
            if not isinstance(row, list):
                continue
            for col_number, display in enumerate(row, start=1):
                cell = ws.cell(row_number, col_number)
                if isinstance(cell, _S360MergedCell):
                    continue
                address = f"{_s360_col(col_number)}{row_number}"
                if address in formulas and str(formulas[address] or "").startswith("="):
                    desired = formulas[address]
                    if cell.value != desired:
                        cell.value = desired
                else:
                    desired_display = "" if display is None else str(display)
                    if _s360_cell_display(cell) != desired_display:
                        cell.value = _s360_coerce(display, cell.value)

                spec = styles.get(address)
                if not isinstance(spec, dict):
                    continue
                font = _s360_copy(cell.font)
                changed_font = False
                mapping = {
                    "bold": "bold", "italic": "italic", "underline": "underline",
                    "fontSize": "size", "fontName": "name",
                }
                for source_key, target_key in mapping.items():
                    if source_key in spec:
                        setattr(font, target_key, spec[source_key])
                        changed_font = True
                color = _s360_hex(spec.get("fontColor")) if "fontColor" in spec else None
                if color:
                    font.color = color
                    changed_font = True
                if changed_font:
                    cell.font = font

                alignment = _s360_copy(cell.alignment)
                changed_alignment = False
                for source_key, target_key in (
                    ("hAlign", "horizontal"), ("vAlign", "vertical"),
                    ("wrap", "wrap_text"), ("rotation", "text_rotation"),
                    ("indent", "indent"),
                ):
                    if source_key in spec:
                        setattr(alignment, target_key, spec[source_key])
                        changed_alignment = True
                if changed_alignment:
                    cell.alignment = alignment

                fill = _s360_hex(spec.get("fill")) if "fill" in spec else None
                if fill:
                    cell.fill = _S360Fill("solid", fgColor=fill)

                border_spec = spec.get("borders")
                if isinstance(border_spec, dict):
                    old = cell.border
                    sides: dict[str, Any] = {}
                    for side_name in ("left", "right", "top", "bottom"):
                        side_data = border_spec.get(side_name)
                        if isinstance(side_data, dict) and side_data.get("style"):
                            sides[side_name] = _S360Side(
                                style=str(side_data.get("style")),
                                color=_s360_hex(side_data.get("color")) or "000000",
                            )
                        else:
                            sides[side_name] = getattr(old, side_name)
                    cell.border = _S360Border(
                        left=sides["left"], right=sides["right"],
                        top=sides["top"], bottom=sides["bottom"],
                        diagonal=old.diagonal, diagonal_direction=old.diagonal_direction,
                        diagonalUp=old.diagonalUp, diagonalDown=old.diagonalDown,
                        outline=old.outline, vertical=old.vertical, horizontal=old.horizontal,
                    )

        row_heights = payload.get("rowHeights") if isinstance(payload.get("rowHeights"), dict) else {}
        for key, value in row_heights.items():
            try:
                ws.row_dimensions[int(key)].height = float(value)
            except Exception:
                pass
        column_widths = payload.get("columnWidths") if isinstance(payload.get("columnWidths"), dict) else {}
        for key, value in column_widths.items():
            try:
                ws.column_dimensions[str(key)].width = float(value)
            except Exception:
                pass

        ws.sheet_format.defaultColWidth = float(
            payload.get("defaultColumnWidth") or DEFAULT_COLUMN_WIDTH_UNITS
        )
        ws.sheet_format.defaultRowHeight = float(
            payload.get("defaultRowHeight") or DEFAULT_ROW_HEIGHT_POINTS
        )
        for index, pixels in enumerate(payload.get("rowHeightsPx") or [], start=1):
            try:
                ws.row_dimensions[index].height = unchanged_row_height_or_converted(
                    pixels, row_heights.get(str(index))
                )
            except (TypeError, ValueError):
                continue
        for index, pixels in enumerate(payload.get("colWidthsPx") or [], start=1):
            letter = _s360_col(index)
            try:
                ws.column_dimensions[letter].width = unchanged_excel_width_or_converted(
                    pixels, column_widths.get(letter)
                )
            except (TypeError, ValueError):
                continue

        geometry_authoritative = payload.get("geometryAuthority") == "workbook-v1"
        if geometry_authoritative:
            for dimension in ws.row_dimensions.values():
                dimension.hidden = False
            for dimension in ws.column_dimensions.values():
                dimension.hidden = False
        for row in payload.get("hiddenRows") or []:
            if isinstance(row, int) and row >= 0:
                ws.row_dimensions[row + 1].hidden = True
        for column in payload.get("hiddenColumns") or []:
            if isinstance(column, int) and column >= 0:
                ws.column_dimensions[_s360_col(column + 1)].hidden = True

        desired_merges: set[str] = set()
        for merged in payload.get("mergedCells") or []:
            if not isinstance(merged, dict):
                continue
            try:
                desired_merges.add(
                    f"{_s360_col(int(merged['startCol']) + 1)}"
                    f"{int(merged['startRow']) + 1}:"
                    f"{_s360_col(int(merged['endCol']) + 1)}"
                    f"{int(merged['endRow']) + 1}"
                )
            except (KeyError, TypeError, ValueError):
                continue
        current_merges = {str(item) for item in ws.merged_cells.ranges}
        if geometry_authoritative:
            for merged in sorted(current_merges - desired_merges):
                ws.unmerge_cells(merged)
        for merged in sorted(desired_merges - current_merges):
            ws.merge_cells(merged)

        tab_color = _s360_hex(payload.get("tabColor"))
        if tab_color:
            ws.sheet_properties.tabColor = tab_color


def _s360_find_index_header(ws: Any) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(ws.max_row or 0, 25) + 1):
        headers: dict[str, int] = {}
        for col in range(1, max(ws.max_column or 0, len(INDEX_HEADERS)) + 1):
            label = str(ws.cell(row_number, col).value or "").strip().casefold()
            if label:
                headers[label] = col
        if {"include", "sheet tab", "page title"}.issubset(headers):
            next_col = max(headers.values(), default=0) + 1
            for label in INDEX_HEADERS:
                key = label.casefold()
                if key not in headers:
                    ws.cell(row_number, next_col, label)
                    headers[key] = next_col
                    next_col += 1
            return row_number, headers
    row_number = 1
    headers = {}
    for col, label in enumerate(INDEX_HEADERS, start=1):
        ws.cell(row_number, col, label)
        headers[label.casefold()] = col
    return row_number, headers



# S360 FULL BASE-PAGE MANIFEST MIRROR V25
def sync_project_to_workbook(
    project_id: str,
    project: dict[str, Any],
    store: Any,
) -> dict[str, Any]:
    """Mirror every app base page into 00_INDEX and workbook tab order.

    Generated continuation pages remain app/PDF output only. Existing workbook
    cells, formulas, images, merges, and unmatched source sheets are preserved.
    """
    path = _workbook_path(store, project_id, project)
    if path is None:
        raise WorkbookSyncError("No workbook is linked to this project.")
    if not path.is_file():
        raise WorkbookSyncError(f"The linked workbook was not found: {path}")

    with _project_lock(store, project_id, project):
        try:
            from core.full_workbook_sync import synchronize_project_to_workbook

            updated = synchronize_project_to_workbook(
                path,
                project_id,
                project,
                store,
                app_hash=project_hash(project),
            )
        except PermissionError as exc:
            raise WorkbookSyncError(
                "The linked workbook is open or Drive has it locked. "
                "Close Excel and Google Sheets, then sync again."
            ) from exc
        except WorkbookSyncError:
            raise
        except Exception as exc:
            raise WorkbookSyncError(
                f"Full workbook mirror failed: {type(exc).__name__}: {exc}"
            ) from exc

        stamp = utcnow()
        sync = dict(updated.get("workbookSync") or {})
        sync.update(
            {
                "mode": "external-workbook-link",
                "workbook": str(path),
                "status": "in_sync",
                "warning": "",
                "pendingReason": "",
                "lastSyncUtc": stamp,
                "workbookHash": file_hash(path),
                "appHash": project_hash(updated),
                "authority": "workbook",
                "lastAuthorityAction": "full_app_to_workbook_mirror",
                "syncEngineVersion": "V25",
            }
        )
        updated["workbookSync"] = sync
        store.save(project_id, updated)
        return updated



# S360 ROBUST INDEX SELECTOR V15.1
def _s360_scan_index_sheet(ws) -> tuple[list[dict[str, Any]], int]:
    # Read one candidate 00_INDEX without relying on dimension metadata.
    raw_rows: list[tuple[Any, ...]] = []
    header_offset = -1
    headers: dict[str, int] = {}

    try:
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = tuple(row)
            raw_rows.append(values)
            if header_offset < 0 and row_number <= 75:
                found: dict[str, int] = {}
                for index, value in enumerate(values):
                    label = str(value or "").strip().casefold()
                    if label:
                        found[label] = index
                if {"include", "sheet tab", "page title"}.issubset(found):
                    header_offset = len(raw_rows) - 1
                    headers = found
            if row_number >= 10000:
                break
    except Exception as exc:
        raise WorkbookSyncError(
            f"Could not stream candidate 00_INDEX sheet: {type(exc).__name__}: {exc}"
        ) from exc

    if header_offset < 0:
        raise WorkbookSyncError(
            "Candidate 00_INDEX does not contain Include, Sheet Tab, and Page Title headers."
        )

    def cell(row: tuple[Any, ...], name: str, default: Any = "") -> Any:
        index = headers.get(name)
        if index is None or index >= len(row):
            return default
        value = row[index]
        return default if value is None else value

    result: list[dict[str, Any]] = []
    blank_run = 0
    for values in raw_rows[header_offset + 1:]:
        tab = str(cell(values, "sheet tab")).strip()
        title = str(cell(values, "page title")).strip()
        code = str(cell(values, "sheet code")).strip()
        page_id = str(cell(values, "page id")).strip()
        if not tab and not title and not code and not page_id:
            blank_run += 1
            if blank_run >= 50:
                break
            continue
        blank_run = 0
        include_raw = str(cell(values, "include")).strip().upper()
        order_raw = cell(values, "order", None)
        try:
            order = int(float(order_raw))
        except Exception:
            order = len(result) + 1

        def value(name: str, default: str = "") -> str:
            return str(cell(values, name, default)).strip()

        result.append({
            "include": include_raw == "YES",
            "publishStatus": (
                include_raw
                if include_raw in {"YES", "NO", "VERIFY"}
                else ""
            ),
            "order": order,
            "sheetCode": code,
            "displaySheetCode": code,
            "sheetTab": tab,
            "sheetTitle": title or tab,
            "pageFamily": value("family"),
            "pageType": value("page type"),
            "notes": value("notes"),
            "renderProfile": value("render profile"),
            "splitMode": value("split mode"),
            "id": page_id,
            "parentPageId": value("parent page id"),
            "issueStatus": normalize_status(value("issue status", "Draft")),
            "sourceMode": value("source mode", "Workbook"),
            "syncDirection": value("sync direction", "Both"),
        })

    score = (
        len(result) * 1000
        + sum(1 for item in result if item.get("include")) * 10
        + sum(1 for item in result if item.get("sheetTab"))
    )
    return result, score


def _s360_index_candidates(wb) -> list[tuple[int, int, Any, list[dict[str, Any]]]]:
    candidates: list[tuple[int, int, Any, list[dict[str, Any]]]] = []
    diagnostics: list[str] = []
    for position, ws in enumerate(wb.worksheets):
        if str(ws.title or "").strip().casefold() != "00_index":
            continue
        try:
            manifest, score = _s360_scan_index_sheet(ws)
            candidates.append((score, position, ws, manifest))
        except WorkbookSyncError as exc:
            diagnostics.append(f"sheet#{position + 1}: {exc}")
    if not candidates:
        detail = "; ".join(diagnostics) if diagnostics else "No 00_INDEX worksheet exists."
        raise WorkbookSyncError("No usable 00_INDEX worksheet was found. " + detail)
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return candidates


def _s360_select_index_sheet(wb, writable: bool = False):
    # Select the populated control sheet when malformed files contain duplicates.
    try:
        return _s360_index_candidates(wb)[0][2]
    except WorkbookSyncError:
        if not writable:
            raise
        if "00_INDEX" in wb.sheetnames:
            return wb["00_INDEX"]
        return wb.create_sheet("00_INDEX", 1)


def _s360_index_manifest(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=path.suffix.lower() == ".xlsm",
    )
    try:
        return _s360_index_candidates(wb)[0][3]
    finally:
        wb.close()
