"""Persisted Singh360 Excel-layout pagination and editable-cell workbook export."""
from __future__ import annotations

from copy import copy
from math import ceil, floor
from typing import Any, Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

PAGE_WIDTH = 1632.0
PAGE_HEIGHT = 1056.0
GRID_COLUMNS = 136
GRID_ROWS_PER_PAGE = 88
DEFAULT_MARGIN = 48.0
DEFAULT_TITLE_FILL = "F4B183"


def is_excel_layout_page(page: dict[str, Any]) -> bool:
    layout = page.get("excelLayout")
    return isinstance(layout, dict) and isinstance(layout.get("tables"), list)


def _positive(value: Any, fallback: float) -> float:
    try:
        number = float(value)
        return number if number > 0 else fallback
    except (TypeError, ValueError):
        return fallback


def _hex(value: Any, fallback: str) -> str:
    text = str(value or "").strip().lstrip("#").upper()
    if len(text) == 8:
        text = text[-6:]
    return text if len(text) == 6 and all(c in "0123456789ABCDEF" for c in text) else fallback


def _tables(page: dict[str, Any]) -> list[dict[str, Any]]:
    layout = page.get("excelLayout") or {}
    return [item for item in layout.get("tables", []) if isinstance(item, dict)]


def page_count(page: dict[str, Any]) -> int:
    layout = page.get("excelLayout") or {}
    height = _positive(layout.get("pageHeight"), PAGE_HEIGHT)
    bottom = height
    for table in _tables(page):
        bottom = max(bottom, float(table.get("y") or 0) + _positive(table.get("height"), 120))
    return max(1, int(ceil(bottom / height)))


def paginate_table(table: dict[str, Any], page_height: float = PAGE_HEIGHT,
                   margin: float = DEFAULT_MARGIN) -> list[dict[str, Any]]:
    """Split on row boundaries; continuation fragments retain deterministic identity."""
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    heights = table.get("rowHeights") if isinstance(table.get("rowHeights"), list) else []
    y = max(0.0, float(table.get("y") or 0))
    start_page = int(floor(y / page_height))
    local_y = y - start_page * page_height
    title_height = 28.0 if str(table.get("title") or "") else 0.0
    header_height = _positive(heights[0] if heights else None, 24.0) if rows else 0.0
    keep = bool(table.get("keepTogether"))
    split = bool(table.get("splitRows", True))
    total_height = title_height + sum(_positive(heights[i] if i < len(heights) else None, 22.0) for i in range(len(rows)))
    if keep and total_height <= page_height - margin * 2 and local_y + total_height > page_height - margin:
        start_page += 1
        local_y = margin
    if not split or not rows:
        return [{"page": start_page, "rowStart": 0, "rowEnd": len(rows), "y": local_y,
                 "continuationIndex": 0, "id": f"{table.get('id', 'table')}:0"}]
    fragments: list[dict[str, Any]] = []
    row_start = 0
    page = start_page
    cursor_y = local_y
    while row_start < len(rows):
        repeated = (title_height if row_start == 0 or table.get("repeatTitle") else 0.0)
        if row_start > 0 and table.get("repeatHeaders") and rows:
            repeated += header_height
        available = page_height - margin - cursor_y - repeated
        row_end = row_start
        used = 0.0
        while row_end < len(rows):
            rh = _positive(heights[row_end] if row_end < len(heights) else None, 22.0)
            if row_end > row_start and used + rh > available:
                break
            if row_end == row_start and rh > available:
                break
            used += rh
            row_end += 1
        if row_end == row_start:
            page += 1
            cursor_y = margin
            continue
        index = len(fragments)
        fragments.append({"page": page, "rowStart": row_start, "rowEnd": row_end,
                          "y": cursor_y, "continuationIndex": index,
                          "id": f"{table.get('id', 'table')}:{index}"})
        row_start = row_end
        page += 1
        cursor_y = margin
    return fragments


def paginate_layout(page: dict[str, Any]) -> list[dict[str, Any]]:
    layout = page.get("excelLayout") or {}
    height = _positive(layout.get("pageHeight"), PAGE_HEIGHT)
    margin = _positive(layout.get("printableMargin"), DEFAULT_MARGIN)
    result: list[dict[str, Any]] = []
    for table in _tables(page):
        result.extend({"tableId": table.get("id"), **part}
                      for part in paginate_table(table, height, margin))
    return result


def _style(cell: Any, style: dict[str, Any], *, default_fill: str,
           default_bold: bool = False) -> None:
    fill = _hex(style.get("fill"), default_fill)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(
        color=_hex(style.get("fontColor"), "000000"),
        size=_positive(style.get("fontSize"), 10),
        bold=bool(style.get("bold", default_bold)),
    )
    cell.alignment = Alignment(
        horizontal=str(style.get("align") or "left"),
        vertical="center",
        wrap_text=bool(style.get("wrap", True)),
    )
    border_style = str(style.get("borderStyle") or "thin")
    side = Side(style=None if border_style == "none" else border_style,
                color=_hex(style.get("borderColor"), "000000"))
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _span_boundaries(widths: Iterable[Any], total: int = GRID_COLUMNS) -> list[int]:
    values = [_positive(value, 1.0) for value in widths]
    if not values:
        return [0, total]
    scale = total / sum(values)
    result = [0]
    for value in values[:-1]:
        result.append(max(result[-1] + 1, min(total - 1, round(result[-1] + value * scale))))
    result.append(total)
    return result


def apply_excel_layout(ws: Any, page: dict[str, Any]) -> None:
    """Render independent tables on a fine worksheet grid using real merged cells."""
    if not is_excel_layout_page(page):
        return
    layout = page["excelLayout"]
    page_width = _positive(layout.get("pageWidth"), PAGE_WIDTH)
    page_height = _positive(layout.get("pageHeight"), PAGE_HEIGHT)
    pages = page_count(page)
    # A fine uniform worksheet grid lets every table express unrelated visible widths.
    for col in range(1, GRID_COLUMNS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 1.15
    for row in range(1, pages * GRID_ROWS_PER_PAGE + 1):
        ws.row_dimensions[row].height = 9
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    # Clear only app-managed page area. Unrelated sheets remain untouched.
    for row in ws.iter_rows(min_row=1, max_row=pages * GRID_ROWS_PER_PAGE,
                            min_col=1, max_col=GRID_COLUMNS):
        for cell in row:
            cell.value = None
            cell._style = copy(ws["A1"]._style)
    for table in _tables(page):
        rows = table.get("rows") if isinstance(table.get("rows"), list) else []
        col_count = max((len(row) for row in rows if isinstance(row, list)), default=1)
        widths = list(table.get("columnWidths") or [1] * col_count)
        while len(widths) < col_count:
            widths.append(1)
        x = max(0.0, float(table.get("x") or 0))
        width = min(page_width - x, _positive(table.get("width"), page_width - x))
        left = max(1, min(GRID_COLUMNS, 1 + round(x / page_width * GRID_COLUMNS)))
        grid_width = max(col_count, round(width / page_width * GRID_COLUMNS))
        right = min(GRID_COLUMNS, left + grid_width - 1)
        spans = _span_boundaries(widths[:col_count], right - left + 1)
        for fragment in paginate_table(
            table, page_height, _positive(layout.get("printableMargin"), DEFAULT_MARGIN)
        ):
            excel_row = fragment["page"] * GRID_ROWS_PER_PAGE + max(
                1, 1 + round(fragment["y"] / page_height * GRID_ROWS_PER_PAGE)
            )
            title = str(table.get("title") or "")
            if title and (fragment["continuationIndex"] == 0 or table.get("repeatTitle")):
                ws.merge_cells(start_row=excel_row, start_column=left,
                               end_row=excel_row + 1, end_column=right)
                title_cell = ws.cell(excel_row, left, title)
                _style(title_cell, table.get("titleStyle") or {},
                       default_fill=DEFAULT_TITLE_FILL, default_bold=True)
                excel_row += 2
            indices = list(range(fragment["rowStart"], fragment["rowEnd"]))
            if fragment["continuationIndex"] and table.get("repeatHeaders") and rows:
                indices.insert(0, 0)
            for source_row in indices:
                row_values = rows[source_row] if source_row < len(rows) and isinstance(rows[source_row], list) else []
                row_start = excel_row
                row_end = excel_row + max(0, round(_positive(
                    (table.get("rowHeights") or [])[source_row]
                    if source_row < len(table.get("rowHeights") or []) else None, 22
                ) / 9) - 1)
                for col_index in range(col_count):
                    c1 = left + spans[col_index]
                    c2 = left + spans[col_index + 1] - 1
                    if row_end > row_start or c2 > c1:
                        ws.merge_cells(start_row=row_start, start_column=c1,
                                       end_row=row_end, end_column=c2)
                    cell = ws.cell(row_start, c1, row_values[col_index] if col_index < len(row_values) else "")
                    style = table.get("headerStyle") if source_row == 0 else table.get("bodyStyle")
                    default_fill = "D9EAF7" if source_row == 0 else (
                        _hex(table.get("alternatingFill"), "FFFFFF")
                        if source_row % 2 == 0 and table.get("alternatingFill") else "FFFFFF"
                    )
                    _style(cell, style or {}, default_fill=default_fill,
                           default_bold=source_row == 0)
                excel_row = row_end + 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False
    ws.page_margins.left = ws.page_margins.right = 0.25
    ws.page_margins.top = ws.page_margins.bottom = 0.25
    ws.print_area = f"A1:{get_column_letter(GRID_COLUMNS)}{pages * GRID_ROWS_PER_PAGE}"
    ws.row_breaks.brk = []
    for index in range(1, pages):
        ws.row_breaks.append(Break(id=index * GRID_ROWS_PER_PAGE))
    custom = _hex(layout.get("tabColor"), "")
    if custom:
        ws.sheet_properties.tabColor = custom
