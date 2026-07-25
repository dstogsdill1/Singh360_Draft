# S360 FULL WORKBOOK MIRROR ENGINE V25
from __future__ import annotations

from copy import copy, deepcopy
from datetime import datetime, timezone
import os
from pathlib import Path
import re
import shutil
import tempfile
import time
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties


INDEX_HEADERS = [
    "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
    "Page Type", "Notes", "Render Profile", "Split Mode", "Page ID",
    "Parent Page ID", "Issue Status", "Source Mode", "Sync Direction",
    "Last Sync UTC", "Workbook Hash", "App Hash",
]

TAB_COLORS = {
    "draft": "F28C28",
    "draft_confirmed": "76B852",
    "public": "2D7DD2",
    "public_confirmed": "14845A",
    "excluded": "9AA3AB",
    "control": "252C34",
    "help": "276FA8",
}

STATUS_LABELS = {
    "draft": "Draft",
    "draft_confirmed": "Draft Confirmed",
    "public": "Public",
    "public_confirmed": "Public Confirmed",
}


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def normalize_status(value: Any) -> str:
    raw = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return raw if raw in STATUS_LABELS else "draft"


def generated_page(page: dict[str, Any]) -> bool:
    if (
        page.get("generatedContinuation")
        or page.get("indexContinuation")
        or page.get("generatedIndexContinuation")
    ):
        return True
    page_id = str(page.get("id") or "")
    source_mode = str(page.get("sourceMode") or "").strip().casefold()
    # Legacy generated pages sometimes carry only a generated ID pattern plus
    # continuationOf. Physical workbook subpages such as EMS 13.1a remain base
    # pages because they have real worksheet tabs.
    return bool(
        page.get("continuationOf")
        and source_mode != "workbook"
        and (
            "__continuation_" in page_id
            or re.search(r"_c\d+(?:_\d+)?$", page_id)
        )
    )


def safe_sheet_name(
    wb: Any,
    requested: str,
    current_sheet: Any | None = None,
) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", str(requested or "APP PAGE")).strip()
    base = base[:31] or "APP PAGE"
    used = {
        str(sheet.title)
        for sheet in wb.worksheets
        if current_sheet is None or sheet is not current_sheet
    }
    if base not in used:
        return base
    index = 2
    while True:
        suffix = f"~{index}"
        candidate = (base[: 31 - len(suffix)] + suffix).strip()
        if candidate not in used:
            return candidate
        index += 1


def best_index_sheet(wb: Any):
    candidates: list[tuple[int, int, Any]] = []
    for position, ws in enumerate(wb.worksheets):
        if str(ws.title or "").strip().casefold() != "00_index":
            continue
        score = 0
        max_row = min(int(ws.max_row or 0), 5000)
        max_col = min(max(int(ws.max_column or 0), 18), 40)
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row or 1,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            if any(value not in (None, "") for value in row):
                score += 1
        candidates.append((score, position, ws))
    if candidates:
        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        return candidates[0][2]
    return wb.create_sheet("00_INDEX", 1)


def find_or_create_headers(ws: Any) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(int(ws.max_row or 1), 75) + 1):
        found: dict[str, int] = {}
        for column in range(1, max(int(ws.max_column or 1), len(INDEX_HEADERS)) + 1):
            value = str(ws.cell(row_number, column).value or "").strip()
            if value:
                found[value.casefold()] = column
        if {"include", "sheet tab", "page title"}.issubset(found):
            for index, header in enumerate(INDEX_HEADERS, start=1):
                if header.casefold() not in found:
                    ws.cell(row_number, index, header)
                    found[header.casefold()] = index
            return row_number, found

    header_row = 4
    for index, header in enumerate(INDEX_HEADERS, start=1):
        ws.cell(header_row, index, header)
    return header_row, {
        header.casefold(): index
        for index, header in enumerate(INDEX_HEADERS, start=1)
    }


def read_existing_rows(
    ws: Any,
    header_row: int,
    headers: dict[str, int],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_run = 0
    for row_number in range(header_row + 1, min(int(ws.max_row or header_row), 10000) + 1):
        values = {
            header: ws.cell(row_number, column).value
            for header, column in headers.items()
        }
        tab = str(values.get("sheet tab") or "").strip()
        code = str(values.get("sheet code") or "").strip()
        title = str(values.get("page title") or "").strip()
        page_id = str(values.get("page id") or "").strip()
        if not any((tab, code, title, page_id)):
            blank_run += 1
            if blank_run >= 100:
                break
            continue
        blank_run = 0
        try:
            order = int(float(values.get("order") or len(rows) + 1))
        except Exception:
            order = len(rows) + 1
        rows.append(
            {
                "row": row_number,
                "Include": str(values.get("include") or "").strip().upper(),
                "Order": order,
                "Sheet Code": code,
                "Sheet Tab": tab,
                "Page Title": title,
                "Family": str(values.get("family") or ""),
                "Page Type": str(values.get("page type") or ""),
                "Notes": str(values.get("notes") or ""),
                "Render Profile": str(values.get("render profile") or ""),
                "Split Mode": str(values.get("split mode") or ""),
                "Page ID": page_id,
                "Parent Page ID": str(values.get("parent page id") or ""),
                "Issue Status": str(values.get("issue status") or "Draft"),
                "Source Mode": str(values.get("source mode") or "Workbook"),
                "Sync Direction": str(values.get("sync direction") or "Both"),
                "Last Sync UTC": str(values.get("last sync utc") or ""),
                "Workbook Hash": str(values.get("workbook hash") or ""),
                "App Hash": str(values.get("app hash") or ""),
            }
        )
    return rows


def base_pages(project: dict[str, Any]) -> list[dict[str, Any]]:
    pages = [
        page
        for page in project.get("pages", [])
        if isinstance(page, dict) and not generated_page(page)
    ]
    pages.sort(key=lambda page: int(page.get("order") or 10**9))

    ids = [str(page.get("id") or "").strip() for page in pages]
    if not all(ids):
        raise RuntimeError("A base page has no Page ID.")
    if len(ids) != len(set(ids)):
        duplicates = sorted({page_id for page_id in ids if ids.count(page_id) > 1})
        raise RuntimeError(
            "Duplicate base Page IDs prevent workbook synchronization: "
            + ", ".join(duplicates[:20])
        )
    return pages


def desired_sheet_tab(page: dict[str, Any]) -> str:
    code = str(page.get("displaySheetCode") or page.get("sheetCode") or "").strip()
    title = str(page.get("sheetTitle") or "").strip()
    tab = str(page.get("sheetTab") or page.get("sourceSheet") or "").strip()
    source_mode = str(page.get("sourceMode") or "").strip().casefold()

    generic_source_tabs = {
        "symbol map",
        "symbol map / source reference",
        "app page",
        "new page",
    }
    app_managed = source_mode == "app" or not page.get("linkedWorksheetId")
    if (
        app_managed
        and code.upper().startswith("EMS ")
        and title
        and (
            not tab
            or tab.strip().casefold() in generic_source_tabs
            or tab.upper().startswith("SRC ")
        )
    ):
        return title
    return tab or title or code or "APP PAGE"


def page_row_values(
    page: dict[str, Any],
    order: int,
    *,
    stamp: str,
    app_hash: str,
) -> dict[str, Any]:
    include = bool(page.get("include", True))
    status = normalize_status(page.get("issueStatus"))
    return {
        "Include": "YES" if include else "NO",
        "Order": order,
        "Sheet Code": str(
            page.get("displaySheetCode") or page.get("sheetCode") or "NEW"
        ).strip(),
        "Sheet Tab": str(page.get("sheetTab") or "").strip(),
        "Page Title": str(page.get("sheetTitle") or "Untitled Sheet").strip(),
        "Family": str(page.get("pageFamily") or ""),
        "Page Type": str(page.get("pageType") or ""),
        "Notes": str(page.get("notes") or ""),
        "Render Profile": str(
            page.get("renderProfile") or page.get("layoutProfile") or ""
        ),
        "Split Mode": str(page.get("splitMode") or ""),
        "Page ID": str(page.get("id") or ""),
        "Parent Page ID": str(page.get("parentPageId") or ""),
        "Issue Status": STATUS_LABELS[status],
        "Source Mode": str(
            page.get("sourceMode")
            or ("Workbook" if page.get("linkedWorksheetId") else "App")
        ),
        "Sync Direction": str(page.get("syncDirection") or "Both"),
        "Last Sync UTC": stamp,
        "Workbook Hash": "",
        "App Hash": app_hash,
    }


def initialize_companion_sheet(
    ws: Any,
    page: dict[str, Any],
    project_id: str,
) -> None:
    if any(
        cell.value not in (None, "")
        for row in ws.iter_rows(min_row=1, max_row=min(int(ws.max_row or 1), 20), min_col=1, max_col=8)
        for cell in row
    ):
        return

    dark = PatternFill("solid", fgColor="252C34")
    gray = PatternFill("solid", fgColor="E6EBEF")
    ws["A1"] = "SINGH360 APP-MANAGED PAGE"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = dark
    ws.merge_cells("A1:H2")

    details = [
        ("Sheet Code", page.get("displaySheetCode") or page.get("sheetCode") or "NEW"),
        ("Page Title", page.get("sheetTitle") or "Untitled Sheet"),
        ("Page ID", page.get("id") or ""),
        ("Published", "YES" if page.get("include", True) else "NO"),
    ]
    for row_number, (label, value) in enumerate(details, start=4):
        ws.cell(row_number, 1, label)
        ws.cell(row_number, 2, value)
        ws.cell(row_number, 1).font = Font(bold=True)
        ws.cell(row_number, 1).fill = gray

    ws["A9"] = (
        "This worksheet mirrors the Singh360 page identity and order. "
        "Canvas objects, pasted images, connectors, crops, symbols, and overlays "
        "remain app-owned and are rendered in the Singh360 PDF package."
    )
    ws.merge_cells("A9:H12")
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A14"] = "Open Singh360 Project"
    ws["B14"] = f"http://127.0.0.1:8766/app?project={project_id}&mode=editor"
    ws["B14"].hyperlink = ws["B14"].value
    ws["B14"].style = "Hyperlink"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72


def copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int) -> None:
    if source_row == target_row:
        return
    for column in range(1, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def update_meta(wb: Any, project: dict[str, Any], stamp: str) -> None:
    if "00_PROJECT_META" in wb.sheetnames:
        ws = wb["00_PROJECT_META"]
    else:
        ws = wb.create_sheet("00_PROJECT_META", 0)

    ws.sheet_properties.tabColor = TAB_COLORS["control"]
    existing: dict[str, int] = {}
    for row in range(1, max(int(ws.max_row or 1), 40) + 1):
        key = str(ws.cell(row, 1).value or "").strip()
        if key:
            existing[key.casefold()] = row

    values = {
        "Linked Project ID": str(project.get("id") or ""),
        "Last Sync UTC": stamp,
        "Last Sync Status": "Synchronized",
        "Sync Mode": "One User / Full Base-Page Manifest Mirror",
    }
    next_row = max(int(ws.max_row or 1) + 1, 17)
    for key, value in values.items():
        row = existing.get(key.casefold())
        if row is None:
            row = next_row
            next_row += 1
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)


def synchronize_project_to_workbook(
    path: Path,
    project_id: str,
    project: dict[str, Any],
    store: Any,
    *,
    app_hash: str,
) -> dict[str, Any]:
    if not path.is_file():
        raise RuntimeError(f"The linked workbook was not found: {path}")

    backup_dir = (
        Path(store.docs)
        / "backups"
        / "full_manifest_sync"
        / project_id
    )
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp_file = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_path = backup_dir / f"{stamp_file}_{path.name}"
    shutil.copy2(path, backup_path)

    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(
        path,
        data_only=False,
        read_only=False,
        keep_vba=keep_vba,
    )
    temp_path: Path | None = None
    try:
        stamp = utcnow()
        index_ws = best_index_sheet(wb)
        index_ws.sheet_properties.tabColor = TAB_COLORS["control"]
        header_row, headers = find_or_create_headers(index_ws)
        existing_rows = read_existing_rows(index_ws, header_row, headers)

        existing_by_id = {
            row["Page ID"]: row
            for row in existing_rows
            if row["Page ID"]
        }
        existing_by_tab = {
            row["Sheet Tab"].casefold(): row
            for row in existing_rows
            if row["Sheet Tab"]
        }
        existing_by_code: dict[str, list[dict[str, Any]]] = {}
        for row in existing_rows:
            code_key = row["Sheet Code"].casefold()
            if code_key:
                existing_by_code.setdefault(code_key, []).append(row)

        pages = base_pages(project)
        app_ids = {str(page.get("id") or "") for page in pages}
        app_rows: list[dict[str, Any]] = []
        app_sheet_objects: list[Any] = []
        used_sheet_objects: set[int] = set()
        used_row_numbers: set[int] = set()

        for order, page in enumerate(pages, start=1):
            page_id = str(page.get("id") or "")
            current_tab = str(page.get("sheetTab") or "").strip()
            code = str(
                page.get("displaySheetCode") or page.get("sheetCode") or ""
            ).strip()
            matched_row = existing_by_id.get(page_id)
            if matched_row is not None and int(matched_row["row"]) in used_row_numbers:
                matched_row = None
            if matched_row is None and current_tab:
                candidate_row = existing_by_tab.get(current_tab.casefold())
                if candidate_row is not None and int(candidate_row["row"]) not in used_row_numbers:
                    matched_row = candidate_row
            if matched_row is None and code:
                code_matches = [
                    row for row in existing_by_code.get(code.casefold(), [])
                    if int(row["row"]) not in used_row_numbers
                ]
                if len(code_matches) == 1:
                    matched_row = code_matches[0]
            if matched_row is not None:
                used_row_numbers.add(int(matched_row["row"]))

            matched_sheet = None
            candidate_tabs = [
                current_tab,
                str(matched_row.get("Sheet Tab") if matched_row else "").strip(),
            ]
            for candidate in candidate_tabs:
                if candidate and candidate in wb.sheetnames:
                    sheet_candidate = wb[candidate]
                    if id(sheet_candidate) not in used_sheet_objects:
                        matched_sheet = sheet_candidate
                        break

            requested_tab = desired_sheet_tab(page)
            if matched_sheet is None:
                actual_tab = safe_sheet_name(wb, requested_tab)
                matched_sheet = wb.create_sheet(actual_tab)
                initialize_companion_sheet(matched_sheet, page, project_id)
            else:
                actual_tab = safe_sheet_name(
                    wb,
                    requested_tab or matched_sheet.title,
                    matched_sheet,
                )
                if matched_sheet.title != actual_tab:
                    matched_sheet.title = actual_tab

            page["sheetTab"] = matched_sheet.title
            include = bool(page.get("include", True))
            status = normalize_status(page.get("issueStatus"))
            matched_sheet.sheet_properties.tabColor = (
                TAB_COLORS.get(status, TAB_COLORS["draft"])
                if include
                else TAB_COLORS["excluded"]
            )

            app_rows.append(
                page_row_values(
                    page,
                    order,
                    stamp=stamp,
                    app_hash=app_hash,
                )
            )
            app_sheet_objects.append(matched_sheet)
            used_sheet_objects.add(id(matched_sheet))

        # Preserve unmatched workbook-only rows and sheets as excluded source rows.
        preserved_rows: list[dict[str, Any]] = []
        preserved_sheet_objects: list[Any] = []
        next_order = len(app_rows) + 1
        for row in sorted(existing_rows, key=lambda item: (item["Order"], item["row"])):
            if int(row["row"]) in used_row_numbers:
                continue
            if row["Page ID"] and row["Page ID"] in app_ids:
                continue
            tab = row["Sheet Tab"]
            if not tab or tab not in wb.sheetnames:
                continue
            sheet = wb[tab]
            if id(sheet) in used_sheet_objects:
                continue
            preserved = dict(row)
            preserved.pop("row", None)
            preserved["Include"] = "NO"
            preserved["Order"] = next_order
            preserved["Last Sync UTC"] = stamp
            preserved["App Hash"] = app_hash
            preserved_rows.append(preserved)
            preserved_sheet_objects.append(sheet)
            used_sheet_objects.add(id(sheet))
            sheet.sheet_properties.tabColor = TAB_COLORS["excluded"]
            next_order += 1

        manifest_rows = app_rows + preserved_rows
        template_row = header_row + 1
        max_col = max(headers.values())
        old_last_row = max(
            [int(row["row"]) for row in existing_rows] or [template_row]
        )
        new_last_row = header_row + len(manifest_rows)

        for offset, values in enumerate(manifest_rows, start=1):
            target_row = header_row + offset
            if target_row > int(index_ws.max_row or 0):
                copy_row_style(index_ws, template_row, target_row, max_col)
            for header in INDEX_HEADERS:
                index_ws.cell(
                    target_row,
                    headers[header.casefold()],
                    values.get(header, ""),
                )

        for row_number in range(new_last_row + 1, old_last_row + 1):
            for column in range(1, max_col + 1):
                index_ws.cell(row_number, column).value = None

        controls: list[Any] = []
        for control_name in ("00_PROJECT_META", index_ws.title, "00_HELP"):
            if control_name in wb.sheetnames:
                sheet = wb[control_name]
                if sheet not in controls:
                    controls.append(sheet)
        if "00_HELP" in wb.sheetnames:
            wb["00_HELP"].sheet_properties.tabColor = TAB_COLORS["help"]

        remaining = [
            sheet
            for sheet in wb.worksheets
            if sheet not in controls
            and id(sheet) not in used_sheet_objects
        ]
        wb._sheets = (
            controls
            + app_sheet_objects
            + preserved_sheet_objects
            + remaining
        )

        update_meta(wb, project, stamp)

        if wb.calculation is None:
            wb.calculation = CalcProperties(
                calcMode="auto",
                fullCalcOnLoad=True,
                forceFullCalc=True,
            )
        else:
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True

        fd, temp_name = tempfile.mkstemp(
            prefix=path.stem + "_sync_",
            suffix=path.suffix,
            dir=path.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)
        wb.save(temp_path)
        wb.close()

        last_error: Exception | None = None
        for attempt in range(1, 6):
            try:
                os.replace(temp_path, path)
                temp_path = None
                break
            except PermissionError as exc:
                last_error = exc
                if attempt >= 5:
                    raise RuntimeError(
                        "The workbook is open or Drive has it locked. "
                        "Close Excel/Google Sheets and sync again."
                    ) from exc
                time.sleep(0.8 * attempt)
        if temp_path is not None:
            raise RuntimeError(
                f"Atomic workbook replacement did not complete: {last_error}"
            )

        updated = deepcopy(project)
        updated.setdefault("workbookSync", {})["lastManifestRowCount"] = len(
            manifest_rows
        )
        updated["workbookSync"]["lastBasePageCount"] = len(app_rows)
        updated["workbookSync"]["lastPreservedWorkbookOnlyCount"] = len(
            preserved_rows
        )
        updated["workbookSync"]["lastFullMirrorBackup"] = str(backup_path)
        return updated
    finally:
        try:
            wb.close()
        except Exception:
            pass
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
