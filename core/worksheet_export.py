"""Export one project worksheet to a standalone workbook."""
from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.workbook_geometry import (
    DEFAULT_COLUMN_WIDTH_UNITS,
    DEFAULT_ROW_HEIGHT_POINTS,
    unchanged_excel_width_or_converted,
    unchanged_row_height_or_converted,
)


def _safe_sheet_title(name: str) -> str:
    t = re.sub(r"[\[\]:*?/\\]", "_", (name or "Sheet").strip())[:31]
    return t or "Sheet"


def export_source_worksheet_xlsx(source_path: str | Path, sheet_name: str) -> bytes:
    """Extract the exact original sheet, retaining its styles, images and print setup."""
    path = Path(source_path)
    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Worksheet {sheet_name!r} was not found in the source workbook.")
        for ws in list(wb.worksheets):
            if ws.title != sheet_name:
                wb.remove(ws)
        wb.active = 0
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()
    finally:
        wb.close()


def _side(style: dict[str, Any] | None) -> Side | None:
    if not style:
        return None
    return Side(style=style.get("style", "thin"), color=style.get("color", "000000"))


def export_worksheet_xlsx(ws: dict[str, Any]) -> bytes:
    """Fallback reconstruction when the original source workbook is unavailable."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = _safe_sheet_title(str(ws.get("name") or ws.get("sourceSheet") or "Sheet"))
    grid = ws.get("grid") if isinstance(ws.get("grid"), list) else []
    styles = ws.get("styles") if isinstance(ws.get("styles"), dict) else {}
    col_widths = ws.get("colWidthsPx") if isinstance(ws.get("colWidthsPx"), list) else []
    row_heights = ws.get("rowHeightsPx") if isinstance(ws.get("rowHeightsPx"), list) else []
    exact_col_widths = ws.get("columnWidths") if isinstance(ws.get("columnWidths"), dict) else {}
    exact_row_heights = ws.get("rowHeights") if isinstance(ws.get("rowHeights"), dict) else {}
    formulas = ws.get("formulas") if isinstance(ws.get("formulas"), dict) else {}
    for r, row in enumerate(grid):
        if not isinstance(row, list): continue
        for c, val in enumerate(row):
            a1 = f"{get_column_letter(c + 1)}{r + 1}"
            cell = sheet.cell(row=r + 1, column=c + 1, value=formulas.get(a1, val))
            st = styles.get(a1)
            if not isinstance(st, dict): continue
            cell.font = Font(bold=bool(st.get("bold")), italic=bool(st.get("italic")), underline="single" if st.get("underline") else None, size=st.get("fontSize") or 11, name=st.get("fontName") or None, color=(st.get("fontColor") or "000000").lstrip("#"))
            fill_color = st.get("fill")
            if fill_color: cell.fill = PatternFill("solid", fgColor=fill_color.lstrip("#"))
            borders = st.get("borders") if isinstance(st.get("borders"), dict) else {}
            cell.border = Border(left=_side(borders.get("left")), right=_side(borders.get("right")), top=_side(borders.get("top")), bottom=_side(borders.get("bottom")))
            cell.alignment = Alignment(horizontal=st.get("hAlign") or "general", vertical=st.get("vAlign") or "bottom", wrap_text=bool(st.get("wrap")), text_rotation=int(st.get("rotation") or 0), indent=int(st.get("indent") or 0))
    sheet.sheet_format.defaultColWidth = float(
        ws.get("defaultColumnWidth") or DEFAULT_COLUMN_WIDTH_UNITS
    )
    sheet.sheet_format.defaultRowHeight = float(
        ws.get("defaultRowHeight") or DEFAULT_ROW_HEIGHT_POINTS
    )
    for letter, units in exact_col_widths.items():
        try:
            if float(units) > 0:
                sheet.column_dimensions[str(letter)].width = float(units)
        except (TypeError, ValueError):
            continue
    for row, points in exact_row_heights.items():
        try:
            if float(points) > 0:
                sheet.row_dimensions[int(row)].height = float(points)
        except (TypeError, ValueError):
            continue
    for c, width in enumerate(col_widths):
        if width:
            letter = get_column_letter(c + 1)
            sheet.column_dimensions[letter].width = unchanged_excel_width_or_converted(
                width, exact_col_widths.get(letter)
            )
    for r, height in enumerate(row_heights):
        if height:
            key = str(r + 1)
            sheet.row_dimensions[r + 1].height = unchanged_row_height_or_converted(
                height, exact_row_heights.get(key)
            )
    for row in ws.get("hiddenRows") or []:
        if isinstance(row, int) and row >= 0:
            sheet.row_dimensions[row + 1].hidden = True
    for column in ws.get("hiddenColumns") or []:
        if isinstance(column, int) and column >= 0:
            sheet.column_dimensions[get_column_letter(column + 1)].hidden = True
    for merge in ws.get("mergedCells") or []:
        if not isinstance(merge, dict): continue
        sheet.merge_cells(start_row=int(merge.get("startRow", 0)) + 1, start_column=int(merge.get("startCol", 0)) + 1, end_row=int(merge.get("endRow", 0)) + 1, end_column=int(merge.get("endCol", 0)) + 1)
    if ws.get("printArea"):
        try: sheet.print_area = str(ws["printArea"])
        except Exception: pass
    bio = BytesIO(); wb.save(bio); return bio.getvalue()


def export_excel_layout_page_xlsx(page: dict[str, Any]) -> bytes:
    """Standalone export for an app-managed Excel-layout drawing page."""
    from core.excel_layout_export import apply_excel_layout

    wb = Workbook()
    sheet = wb.active
    sheet.title = _safe_sheet_title(
        str(page.get("sheetTab") or page.get("sheetTitle") or "Layout")
    )
    apply_excel_layout(sheet, page)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
