"""Import exactly selected worksheets into an existing Singh360 project.

The one-sheet workflow preserves the worksheet's Excel geometry and styling and
never rebuilds the full project package.  The source workbook remains untouched.
"""
from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from core.project_model import classify_page_type, sanitize_json
from core.page_normalizer import normalize_page
from core.page_composer import compose_pages, page_family
from core.spreadsheet_layout import canonical_layout_override, exact_source_layout
from core.workbook_importer import (
    EXCEL_MIN_SCALE,
    _excel_range_block,
    _extract_embedded_images,
    _parse_index,
    _safe_name,
    _worksheet_payload,
)


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _new_id(prefix: str = "p") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def preview_workbook_sheets(xlsx_path: str | Path) -> list[dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    out: list[dict[str, Any]] = []
    try:
        index_name = next((n for n in wb.sheetnames if re.sub(r"[^a-z0-9]", "", n.lower()) in {"00index", "index"}), None)
        index_entries = _parse_index(wb, index_name) if index_name else []
        by_tab = {str(item.get("sheetTab") or "").strip().lower(): item for item in index_entries}
        for name in wb.sheetnames:
            ws = wb[name]
            meta = by_tab.get(name.strip().lower(), {})
            out.append({
                "sheetName": name,
                "rowEstimate": ws.max_row or 0,
                "colEstimate": ws.max_column or 0,
                "detectedPageType": classify_page_type(name, str(meta.get("sheetTitle") or name), ""),
                "sheetCode": str(meta.get("sheetCodeRaw") or "").strip(),
                "pageTitle": str(meta.get("sheetTitle") or name).strip(),
                "listedInIndex": bool(meta),
                "printArea": str(getattr(ws, "print_area", "") or ""),
            })
    finally:
        wb.close()
    return out


def _index_meta(wb, sheet_name: str) -> dict[str, Any]:
    index_name = next((n for n in wb.sheetnames if re.sub(r"[^a-z0-9]", "", n.lower()) in {"00index", "index"}), None)
    if not index_name:
        return {}
    for item in _parse_index(wb, index_name):
        if str(item.get("sheetTab") or "").strip().lower() == sheet_name.strip().lower():
            return item
    return {}


def _parse_print_area(value: Any) -> tuple[int, int, int, int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    # Use the first area only.  Sheet names may be quoted and are discarded.
    text = text.split(",", 1)[0]
    if "!" in text:
        text = text.split("!", 1)[1]
    text = text.replace("$", "")
    if ":" not in text:
        return None
    start, end = text.split(":", 1)
    try:
        scol, srow = coordinate_from_string(start)
        ecol, erow = coordinate_from_string(end)
        return srow - 1, column_index_from_string(scol) - 1, erow - 1, column_index_from_string(ecol) - 1
    except Exception:
        return None


def _slice_payload_to_print_area(payload: dict[str, Any]) -> dict[str, Any]:
    area = _parse_print_area(payload.get("printArea"))
    if area is None:
        return payload
    sr, sc, er, ec = area
    grid = payload.get("grid") or []
    if not grid:
        return payload
    er = min(er, len(grid) - 1)
    ec = min(ec, max((len(row) for row in grid), default=1) - 1)
    if er < sr or ec < sc:
        return payload
    out = dict(payload)
    out["grid"] = [list(row[sc : ec + 1]) for row in grid[sr : er + 1]]
    out["colWidthsPx"] = list((payload.get("colWidthsPx") or [])[sc : ec + 1])
    out["rowHeightsPx"] = list((payload.get("rowHeightsPx") or [])[sr : er + 1])
    out["hiddenRows"] = [row - sr for row in (payload.get("hiddenRows") or []) if sr <= row <= er]
    out["hiddenColumns"] = [col - sc for col in (payload.get("hiddenColumns") or []) if sc <= col <= ec]

    styles: dict[str, Any] = {}
    for key, style in (payload.get("styles") or {}).items():
        match = re.fullmatch(r"([A-Z]+)(\d+)", str(key))
        if not match:
            continue
        col = column_index_from_string(match.group(1)) - 1
        row = int(match.group(2)) - 1
        if sr <= row <= er and sc <= col <= ec:
            # Small local A1 converter.
            n = col - sc + 1
            letters = ""
            while n:
                n, rem = divmod(n - 1, 26)
                letters = chr(65 + rem) + letters
            styles[f"{letters}{row - sr + 1}"] = style
    out["styles"] = styles

    merges: list[dict[str, int]] = []
    for merge in payload.get("mergedCells") or []:
        msr = int(merge.get("startRow", 0)); msc = int(merge.get("startCol", 0))
        mer = int(merge.get("endRow", msr)); mec = int(merge.get("endCol", msc))
        if msr >= sr and msc >= sc and mer <= er and mec <= ec:
            merges.append({
                "startRow": msr - sr, "startCol": msc - sc,
                "endRow": mer - sr, "endCol": mec - sc,
            })
    out["mergedCells"] = merges
    out["sourceRange"] = f"PRINT_AREA:{payload.get('printArea')}"
    return out


def _exact_excel_block(ws_data: dict[str, Any], ws_id: str) -> dict[str, Any]:
    """Compatibility wrapper returning the exact-source base block."""
    blocks, _diagnostics = _exact_excel_layout(ws_data, ws_id)
    return blocks[0]


def _exact_excel_layout(
    ws_data: dict[str, Any],
    ws_id: str,
    *,
    layout_override: str = "exact_source",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data = _slice_payload_to_print_area(ws_data)
    full = _excel_range_block(
        {**data, "id": ws_id},
        f"{ws_id}_excel_exact",
        {
            "splitMode": "auto_rows",
            "allowContinuation": True,
            "scaleMode": "fit_body",
            "minScale": max(EXCEL_MIN_SCALE, 7.5 / 9.0),
            "trimBlankRows": False,
            "trimBlankColumns": False,
        },
    )
    canonical = canonical_layout_override(layout_override)
    blocks, diagnostics = exact_source_layout(full, override=canonical)
    for block in blocks:
        block.update({
            "sourceWorksheetId": ws_id,
            "sourceSheet": data.get("sourceSheet") or data.get("name") or "",
            "editable": True,
            "layoutProfile": "semantic_excel_table" if canonical == "two_columns" else "exact_source_excel",
        })
    return blocks, diagnostics


_EMS_CODE_RE = re.compile(r"^EMS\s+(\d+)(?:\.(\d+))?$", re.I)


def _stable_import_sheet_code(
    pages: list[dict[str, Any]],
    insert_at: int,
    explicit: str,
) -> str:
    """Use 00_INDEX when present; otherwise fill an unambiguous EMS gap."""
    explicit = explicit.strip()
    if explicit and explicit.upper() not in {"NEW", "TBD"}:
        return explicit
    used = {
        str(page.get("displaySheetCode") or page.get("sheetCode") or "").strip().casefold()
        for page in pages
    }

    def nearest(items: list[dict[str, Any]]) -> tuple[int, int] | None:
        for page in items:
            match = _EMS_CODE_RE.fullmatch(
                str(page.get("displaySheetCode") or page.get("sheetCode") or "").strip()
            )
            if match:
                return int(match.group(1)), int(match.group(2) or 0)
        return None

    before = nearest(list(reversed(pages[:insert_at])))
    after = nearest(pages[insert_at:])
    candidates: list[tuple[int, int]] = []
    if before and after and after[0] - before[0] >= 2:
        candidates.append((before[0] + 1, 0))
    if before:
        candidates.extend((before[0], minor) for minor in range(before[1] + 1, 100))
        candidates.append((before[0] + 1, 0))
    if after and after[0] > 1:
        candidates.append((after[0] - 1, 0))
    for major, minor in candidates:
        candidate = f"EMS {major}.{minor}"
        if candidate.casefold() not in used:
            return candidate
    # No engineering sequence can be inferred safely. A stable draft identity
    # is preferable to the temporary word NEW and does not invent project data.
    return "DRAFT-IMPORT"


def import_workbook_sheets(
    project: dict[str, Any],
    xlsx_path: str | Path,
    sheet_names: list[str],
    *,
    insert_after_page_id: str | None = None,
    replace_page_id: str | None = None,
    append: bool = False,
    template_override: str | None = None,
    preserve_exact: bool = True,
    layout_override: str = "auto",
    assets_dir=None,
    asset_url_prefix: str | None = None,
    source_filename: str | None = None,
    source_sha256: str = "",
    project_local_path: str = "",
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Import only the requested worksheets; never rebuild the full package."""
    if preserve_exact and len(sheet_names) != 1:
        raise ValueError("Add One Formatted Sheet requires exactly one worksheet.")

    xlsx_path = Path(xlsx_path)
    keep_vba = xlsx_path.suffix.lower() == ".xlsm"
    wb = load_workbook(filename=str(xlsx_path), data_only=False, keep_vba=keep_vba)
    wb_values = load_workbook(filename=str(xlsx_path), data_only=True, keep_vba=keep_vba)
    src_filename = source_filename or xlsx_path.name
    imported_at = _ts()
    try:
        src_id = _new_id("src")
        project.setdefault("sources", []).append({
            "id": src_id,
            "type": "imported-workbook",
            "name": src_filename,
            "originalFileName": src_filename,
            "path": str(xlsx_path),
            "projectLocalPath": project_local_path,
            "sha256": source_sha256,
            "sourceType": "excel_workbook",
            "selectedWorksheets": list(sheet_names),
            "importMode": "one_time_editable_table",
            "importedAt": imported_at,
        })

        existing_ws_names = {str(ws.get("name") or "") for ws in project.get("worksheets", [])}
        ws_id_by_name: dict[str, str] = {}
        meta_by_name: dict[str, dict[str, Any]] = {}
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            payload = _worksheet_payload(wb[sheet_name], wb_values[sheet_name])
            if assets_dir is not None:
                Path(assets_dir).mkdir(parents=True, exist_ok=True)
            embedded = _extract_embedded_images(wb[sheet_name], assets_dir, asset_url_prefix or "", sheet_name)
            ws_id = _new_id("ws")
            ws_id_by_name[sheet_name] = ws_id
            meta_by_name[sheet_name] = _index_meta(wb, sheet_name)
            unique_name = sheet_name
            if unique_name in existing_ws_names:
                unique_name = f"{sheet_name} (imported {_ts()[:10]})"
            project.setdefault("worksheets", []).append({
                "id": ws_id,
                "name": unique_name,
                "sourceId": src_id,
                "visible": True,
                "classHint": "excel_exact" if preserve_exact else "imported",
                **payload,
                "embeddedImages": embedded,
                "provenance": {
                    "sheet": sheet_name,
                    "sourceFile": src_filename,
                    "sourcePath": str(xlsx_path),
                    "projectLocalPath": project_local_path,
                    "sha256": source_sha256,
                    "sourceType": "excel_workbook",
                    "importMode": "one_time_editable_table",
                    "importedAt": imported_at,
                },
            })
            existing_ws_names.add(unique_name)

        pages: list[dict[str, Any]] = project.get("pages", [])

        if insert_after_page_id:
            ref_idx = next((i for i, page in enumerate(pages) if page.get("id") == insert_after_page_id), None)
            insert_at = ref_idx + 1 if ref_idx is not None else len(pages)
        elif replace_page_id:
            insert_at = next((i for i, page in enumerate(pages) if page.get("id") == replace_page_id), len(pages))
        else:
            insert_at = len(pages)

        def build_page(sheet_name: str, ws_id: str, target: dict[str, Any] | None = None) -> dict[str, Any]:
            ws_data = next(w for w in project.get("worksheets", []) if w.get("id") == ws_id)
            meta = meta_by_name.get(sheet_name) or {}
            title = str(meta.get("sheetTitle") or (target or {}).get("sheetTitle") or sheet_name).strip()
            explicit_code = str(meta.get("sheetCodeRaw") or (target or {}).get("displaySheetCode") or (target or {}).get("sheetCode") or "").strip()
            code = _stable_import_sheet_code(pages, insert_at, explicit_code)
            detected = template_override or classify_page_type(sheet_name, title, "")
            # A deliberately formatted worksheet is a worksheet page, even when
            # its title contains words such as Layout or Location.
            page_type = detected if detected in {"cover", "index"} else ("data-grid" if preserve_exact else detected)
            family = page_family(sheet_name, title, "")
            if preserve_exact:
                blocks, layout_diagnostics = _exact_excel_layout(
                    ws_data,
                    ws_id,
                    layout_override=layout_override,
                )
            else:
                blocks = normalize_page(ws_data, ws_id, page_type, title)
                layout_diagnostics = {}
            base = dict(target or {})
            base.update({
                "linkedWorksheetId": ws_id,
                "sheetCode": code,
                "displaySheetCode": code,
                "sheetTitle": title,
                "sheetTab": sheet_name,
                "pageType": page_type,
                "pageFamily": family,
                "renderMode": "excel_exact" if preserve_exact else base.get("renderMode", "normalized"),
                "layoutProfile": ("semantic_excel_table" if canonical_layout_override(layout_override) == "two_columns" else "exact_source_excel") if preserve_exact else base.get("layoutProfile", ""),
                "layoutOverride": canonical_layout_override(layout_override) if preserve_exact else base.get("layoutOverride", "exact_source"),
                "layoutDiagnostics": layout_diagnostics,
                "tableLayout": layout_diagnostics.get("selectedArrangement", "single"),
                "splitMode": ("none" if canonical_layout_override(layout_override) == "keep_one_page" else "auto_rows") if preserve_exact else base.get("splitMode", "auto_rows"),
                "allowContinuation": canonical_layout_override(layout_override) != "keep_one_page" if preserve_exact else base.get("allowContinuation", True),
                "minScale": max(EXCEL_MIN_SCALE, 7.5 / 9.0) if preserve_exact else base.get("minScale", EXCEL_MIN_SCALE),
                "scaleMode": "fit_body",
                "trimBlankRows": False if preserve_exact else base.get("trimBlankRows", True),
                "trimBlankColumns": False if preserve_exact else base.get("trimBlankColumns", True),
                "blocks": blocks,
                "issueStatus": base.get("issueStatus") or "draft",
                "publishStatus": base.get("publishStatus") or "",
                "notes": str(meta.get("notes") or base.get("notes") or ""),
                "createdAt": base.get("createdAt") or imported_at,
                "modifiedAt": imported_at,
                "importedFrom": {
                    "sourceFile": src_filename,
                    "sheetName": sheet_name,
                    "projectLocalPath": project_local_path,
                    "sha256": source_sha256,
                    "importedAt": imported_at,
                    "preservedExcelStyle": bool(preserve_exact),
                },
                "sourceImport": {
                    "sourceId": src_id,
                    "sourceType": "excel_workbook",
                    "originalFileName": src_filename,
                    "sha256": source_sha256,
                    "selectedWorksheet": sheet_name,
                    "importMode": "one_time_editable_table",
                    "projectLocalPath": project_local_path,
                    "importedAt": imported_at,
                },
            })
            return base

        if replace_page_id:
            sheet_name = sheet_names[0]
            ws_id = ws_id_by_name.get(sheet_name)
            target_idx = next((i for i, page in enumerate(pages) if page.get("id") == replace_page_id), None)
            if target_idx is None or not ws_id:
                raise ValueError("The current page or selected worksheet could not be found.")
            target = pages[target_idx]
            group_id = target.get("pageGroupId") or target.get("id")
            pages = [p for p in pages if not (p.get("generatedContinuation") and (p.get("continuationOf") == group_id or p.get("pageGroupId") == group_id) and p.get("id") != replace_page_id)]
            target_idx = next(i for i, page in enumerate(pages) if page.get("id") == replace_page_id)
            updated = build_page(sheet_name, ws_id, pages[target_idx])
            updated["pageGroupId"] = updated["id"]
            updated["continuationOf"] = None
            updated["continuationIndex"] = 0
            updated["generatedContinuation"] = False
            replacements = compose_pages([updated])
            pages[target_idx : target_idx + 1] = replacements
            for i, page in enumerate(pages): page["order"] = i + 1
            project["pages"] = pages
            project.setdefault("importHistory", []).append({"sourceFile": src_filename, "sheetNames": sheet_names, "importedAt": _ts(), "pagesAdded": max(0, len(replacements) - 1), "replacedPageId": replace_page_id})
            return sanitize_json(project), replacements

        new_pages: list[dict[str, Any]] = []
        for sheet_name in sheet_names:
            ws_id = ws_id_by_name.get(sheet_name)
            if not ws_id:
                continue
            page = build_page(sheet_name, ws_id)
            page.update({
                "id": _new_id("page"), "order": 0, "include": True,
                "templateId": "ansi-b-standard", "canvasObjects": [], "assets": [], "underlays": [],
                "revisionRows": [], "pageGroupId": "", "continuationOf": None,
                "continuationIndex": 0, "generatedContinuation": False, "layoutWarnings": [],
            })
            page["pageGroupId"] = page["id"]
            new_pages.extend(compose_pages([page]))

        for i, page in enumerate(new_pages): pages.insert(insert_at + i, page)
        for i, page in enumerate(pages): page["order"] = i + 1
        project["pages"] = pages
        project.setdefault("importHistory", []).append({"sourceFile": src_filename, "sheetNames": sheet_names, "importedAt": _ts(), "pagesAdded": len(new_pages), "preservedExcelStyle": bool(preserve_exact)})
        project["renumberSuggested"] = False
        return sanitize_json(project), new_pages
    finally:
        wb.close()
        wb_values.close()


def repair_imported_excel_page(
    project: dict[str, Any],
    page_id: str,
    *,
    layout_override: str = "exact_source",
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Recompose one source-linked imported page while preserving its identity."""
    pages = list(project.get("pages") or [])
    index = next((i for i, page in enumerate(pages) if page.get("id") == page_id), None)
    if index is None:
        raise ValueError(f"Page not found: {page_id}")
    page = pages[index]
    if page.get("generatedContinuation"):
        raise ValueError("Select the base imported worksheet page, not a continuation.")
    if not isinstance(page.get("importedFrom"), dict):
        raise ValueError("Only an imported worksheet page can use source layout controls.")
    worksheet_id = str(page.get("linkedWorksheetId") or "")
    worksheet = next(
        (item for item in project.get("worksheets") or [] if item.get("id") == worksheet_id),
        None,
    )
    if worksheet is None:
        raise ValueError("The imported source worksheet is no longer attached.")

    canonical = canonical_layout_override(layout_override)
    blocks, diagnostics = _exact_excel_layout(
        worksheet,
        worksheet_id,
        layout_override=canonical,
    )
    code = _stable_import_sheet_code(pages, index, str(page.get("displaySheetCode") or page.get("sheetCode") or ""))
    updated = {
        **page,
        "sheetCode": code,
        "displaySheetCode": code,
        "renderMode": "excel_exact",
        "renderProfile": "semantic_excel_table" if canonical == "two_columns" else "exact_source_excel",
        "layoutProfile": "semantic_excel_table" if canonical == "two_columns" else "exact_source_excel",
        "layoutOverride": canonical,
        "layoutDiagnostics": diagnostics,
        "tableLayout": diagnostics.get("selectedArrangement", "single"),
        "splitMode": "auto_rows",
        "allowContinuation": canonical != "keep_one_page",
        "minScale": max(EXCEL_MIN_SCALE, 7.5 / 9.0),
        "scaleMode": "fit_body",
        "trimBlankRows": False,
        "trimBlankColumns": False,
        "blocks": blocks,
        "issueStatus": page.get("issueStatus") or "draft",
        "publishStatus": page.get("publishStatus") or "",
        "pageGroupId": page_id,
        "continuationOf": None,
        "continuationIndex": 0,
        "generatedContinuation": False,
        "layoutWarnings": [],
    }
    group_id = str(page.get("pageGroupId") or page_id)
    pages = [
        existing for existing in pages
        if not (
            existing.get("generatedContinuation")
            and (existing.get("continuationOf") == group_id or existing.get("pageGroupId") == group_id)
        )
    ]
    index = next(i for i, existing in enumerate(pages) if existing.get("id") == page_id)
    replacements = compose_pages([updated])
    pages[index : index + 1] = replacements
    for order, existing in enumerate(pages, start=1):
        existing["order"] = order
    project["pages"] = pages
    project["renumberSuggested"] = False
    return sanitize_json(project), replacements
