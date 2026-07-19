from __future__ import annotations

import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

from core.metadata_inference import infer_metadata_from_labeled_grid
from core.project_model import classify_page_type, default_project, recalc_page_numbers, sanitize_json
from core.page_normalizer import normalize_page
from core.page_composer import (
    BODY_BUDGET,
    BODY_W,
    EXCEL_EXACT_FAMILIES,
    EXCEL_MIN_SCALE,
    compose_pages,
    continuation_code,
    log_render_diagnostics,
    page_family,
)
from core.table_style_profile import (
    ABSOLUTE_MIN_FONT_SIZE,
    DENSE_FONT_SIZE,
    NARRATIVE_FONT_SIZE,
    NARRATIVE_MIN_FONT_SIZE,
    NARRATIVE_RENDER_PROFILE,
    RENDER_PROFILE,
    apply_singh360_profile,
)

# Default Singh360 normalized header style applied to every non-cover page.
DEFAULT_HEADER_STYLE = "orange"

_INDEX_ALIASES = {
    "include": {"include", "inc", "include?", "use?", "selected"},
    # Plain sequential Order/Page column only. Never treat "Sheet No." as Order —
    # that string is reserved for the canonical drawing sheet code (Phase B).
    "order": {"order", "num", "seq", "page", "page no", "page no.", "page #"},
    # The real drawing-package sheet number (e.g. "EMS 0.2") — separate from
    # the plain sequential "Order" column. FINAL RENDER POLISH 4G, Phase A:
    # the title block must show this, not the sequential output order.
    # "Suggested EMS Code" is the SA38/Kyle-workbook alias (00_APP_INDEX).
    "sheet_code": {
        "sheet code", "code", "drawing no", "drawing no.", "drawing number",
        "dwg no", "dwg no.", "dwg code", "sheet #", "sheet no", "sheet no.",
        "suggested ems code",
    },
    # "Original Tab" is the SA38/Kyle-workbook alias (00_APP_INDEX).
    "sheet_tab": {"sheet tab", "sheet", "worksheet", "tab", "tab name", "original tab"},
    # "Normalized Page Title" is the SA38/Kyle-workbook alias (00_APP_INDEX).
    "sheet_title": {"page title", "title", "sheet title", "page name", "name", "normalized page title"},
    "use_source": {"use", "source", "use / source", "use/source", "type"},
    "family": {"family", "page family", "discipline"},
    "page_type": {"page type", "kind"},
    "notes": {"notes", "remarks", "description", "comment"},
}

# Matches a sheet-code prefix baked into the worksheet tab name itself, e.g.
# "EMS 0.2 Abbrev" -> "EMS 0.2". Used only as a fallback when the index sheet
# has no dedicated Sheet Code column (Phase A).
_TAB_SHEET_CODE_RE = re.compile(r"^([A-Za-z]{2,8})\s*[-_ ]?\s*(\d{1,3}\.\d{1,3})\b")


def _sheet_code_from_tab(name: str) -> str:
    """Best-effort sheet code parsed from a worksheet tab name (fallback only;
    the index sheet's own Sheet Code column always takes precedence)."""
    m = _TAB_SHEET_CODE_RE.match((name or "").strip())
    if not m:
        return ""
    prefix, num = m.groups()
    return f"{prefix.upper()} {num}"


def _norm(v: Any) -> str:
    if v is None:
        return ""
    text = " ".join(str(v).split()).strip()
    return "" if text.lower() in {"nan", "nat", "<na>", "none"} else text


def _normalized_sheet_key(name: str) -> str:
    return (name or "").replace(" ", "").replace("_", "").upper()


def _is_metadata_sheet_name(name: str) -> bool:
    """True for a project-metadata/control sheet (e.g. ``00_PROJECT_META``).

    Metadata sheets are never an index candidate and are never rendered as an
    output page (Kyle/SA38 workbook import, Phase A)."""
    key = _normalized_sheet_key(name)
    return "PROJECTMETA" in key


def _find_metadata_sheet(workbook) -> str | None:
    for name in workbook.sheetnames:
        if _is_metadata_sheet_name(name):
            return name
    return None


def _find_index_sheet(workbook) -> str | None:
    """Pick the controlling index sheet.

    Kyle/SA38-style workbooks may ship both a canonical ``00_INDEX`` and a
    richer ``00_APP_INDEX`` alias sheet (Phase A). ``00_INDEX`` always wins
    when both exist; metadata/control sheets (``00_PROJECT_META``) are never
    treated as an index candidate even though their name may contain other
    index-like substrings.
    """
    candidates = [
        name
        for name in workbook.sheetnames
        if "INDEX" in _normalized_sheet_key(name) and not _is_metadata_sheet_name(name)
    ]
    if not candidates:
        return None
    for name in candidates:
        if _normalized_sheet_key(name) in ("00INDEX", "INDEX"):
            return name
    return candidates[0]


def _find_alias_index_sheets(workbook, winning_index: str | None) -> set[str]:
    """Other index-like sheets that lost the Phase A preference (e.g.
    ``00_APP_INDEX`` when ``00_INDEX`` exists) — always excluded from output
    pages regardless of any Include flag."""
    return {
        name
        for name in workbook.sheetnames
        if "INDEX" in _normalized_sheet_key(name)
        and not _is_metadata_sheet_name(name)
        and name != winning_index
    }


def _header_map(header_row: list[str]) -> dict[str, int]:
    normed = [h.lower() for h in header_row]
    out: dict[str, int] = {}
    for key, aliases in _INDEX_ALIASES.items():
        out[key] = -1
        for i, cell in enumerate(normed):
            if cell in aliases:
                out[key] = i
                break
    return out


def _included(raw: str, sheet_title: str, use_source: str, *, in_index: bool = True) -> bool:
    """Index include column is law: only explicit YES/TRUE/include renders output."""
    text = (raw or "").strip().lower()
    cls_blob = f"{sheet_title} {use_source}".lower()
    if text in {"n", "no", "false", "0", "exclude", "off", "excluded", "disabled"}:
        return False
    if text in {"y", "yes", "true", "1", "include", "x", "✓", "on"}:
        return True
    if "disabled" in cls_blob or "not included" in cls_blob or "optional" in cls_blob:
        return False
    if "template" in cls_blob or "utility" in cls_blob:
        return False
    if in_index:
        # Blank or unrecognized include cell on an index row → excluded.
        return False
    return False


# Excel unit conversion. Column width is in "characters" of the default font;
# the classic Excel/openpyxl approximation is px ~= width * 7 + 5. Row height is
# in points; CSS px = pt * 96/72. Defaults mirror Excel's out-of-the-box sheet.
_DEFAULT_COL_PX = 64
_DEFAULT_ROW_PX = 20


def _col_width_px(width: float | None) -> int:
    if not width or width <= 0:
        return _DEFAULT_COL_PX
    return max(8, int(round(width * 7 + 5)))


def _row_height_px(height: float | None) -> int:
    if not height or height <= 0:
        return _DEFAULT_ROW_PX
    return max(8, int(round(height * 4 / 3)))


def _color_hex(color) -> str | None:
    """Return an ARGB/RGB openpyxl color as #RRGGBB, or None for default/none."""
    try:
        if color is None or getattr(color, "type", None) != "rgb":
            return None
        rgb = color.rgb
        if not rgb or not isinstance(rgb, str):
            return None
        if rgb in ("00000000", "FFFFFFFF"):
            return None
        if len(rgb) == 8:
            return "#" + rgb[2:]
        if len(rgb) == 6:
            return "#" + rgb
    except Exception:
        return None
    return None


def _cell_fill_hex(cell) -> str | None:
    """Return a solid fill color as #RRGGBB, or None."""
    try:
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            return None
        return _color_hex(fill.fgColor)
    except Exception:
        return None


def _side_spec(side) -> dict[str, str] | None:
    """Return {style,color} for a single border side, or None when absent."""
    try:
        if side is None or not side.style:
            return None
        return {"style": side.style, "color": _color_hex(side.color) or "#000000"}
    except Exception:
        return None


def _cell_borders(cell) -> dict[str, Any] | None:
    """Return per-side border specs (top/right/bottom/left), or None."""
    try:
        b = cell.border
        out: dict[str, Any] = {}
        for name, side in (("top", b.top), ("right", b.right), ("bottom", b.bottom), ("left", b.left)):
            spec = _side_spec(side)
            if spec:
                out[name] = spec
        return out or None
    except Exception:
        return None


def _cell_style(cell) -> dict[str, Any]:
    """Full per-cell style for exact-range rendering. Falsy/None keys pruned."""
    font = cell.font
    align = cell.alignment
    raw: dict[str, Any] = {
        "bold": bool(getattr(font, "bold", False)),
        "italic": bool(getattr(font, "italic", False)),
        "underline": bool(getattr(font, "underline", None)),
        "fontSize": getattr(font, "size", None),
        "fontName": getattr(font, "name", None),
        "fontColor": _color_hex(getattr(font, "color", None)),
        "hAlign": getattr(align, "horizontal", None),
        "vAlign": getattr(align, "vertical", None),
        "wrap": bool(getattr(align, "wrapText", False)),
        "rotation": int(getattr(align, "textRotation", 0) or 0),
        "indent": int(getattr(align, "indent", 0) or 0),
        "fill": _cell_fill_hex(cell),
        "borders": _cell_borders(cell),
    }
    return {k: v for k, v in raw.items() if v not in (None, False, 0, "")}


def _worksheet_payload(ws, ws_data=None) -> dict[str, Any]:
    """Extract a worksheet into a render-ready payload.

    ``ws`` is loaded with formulas (data_only=False); ``ws_data`` (data_only=True)
    supplies cached calculated values so formulas display as values.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid: list[list[str]] = []
    formulas: dict[str, str] = {}
    styles: dict[str, dict[str, Any]] = {}

    for r in range(1, max_row + 1):
        row_vals: list[str] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            col_letter = get_column_letter(c)
            raw_value = cell.value
            display = raw_value
            if isinstance(raw_value, str) and raw_value.startswith("="):
                formulas[f"{col_letter}{r}"] = raw_value
                cached = ws_data.cell(r, c).value if ws_data is not None else None
                display = cached if cached is not None else ""
            value = _norm(display)
            row_vals.append(value)
            if isinstance(cell, MergedCell):
                # Placeholder cell in a merged range — keep grid value, skip style extraction.
                continue
            if cell.has_style:
                s = _cell_style(cell)
                if s:
                    styles[f"{col_letter}{r}"] = s
        grid.append(row_vals)

    # Trailing rows with no values are dropped by default, but never a row that
    # still carries a meaningful fill/border — a decorative ruled/shaded row with
    # no text is real content, not usedRange bloat (FINAL RENDER POLISH 4G,
    # Phase B: this legacy pop predates the dedicated trim step below and must
    # honor the same "meaningful style survives" rule so it can't silently
    # discard a blocked/shaded row before that step ever sees it).
    def _row_has_meaningful_style(row_idx0: int) -> bool:
        row_num = row_idx0 + 1
        return any(
            _meaningful_style(styles.get(f"{get_column_letter(c)}{row_num}"))
            for c in range(1, max_col + 1)
        )

    while grid and not any(grid[-1]) and not _row_has_meaningful_style(len(grid) - 1):
        grid.pop()

    n_rows = len(grid)
    n_cols = max((len(r) for r in grid), default=0)

    merges = []
    for merged in ws.merged_cells.ranges:
        merges.append(
            {
                "startRow": merged.min_row - 1,
                "startCol": merged.min_col - 1,
                "endRow": merged.max_row - 1,
                "endCol": merged.max_col - 1,
            }
        )

    default_col = getattr(getattr(ws, "sheet_format", None), "defaultColWidth", None)
    default_row = getattr(getattr(ws, "sheet_format", None), "defaultRowHeight", None)

    col_widths_px: list[int] = []
    for c in range(1, n_cols + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        width = dim.width if dim and dim.width else default_col
        col_widths_px.append(_col_width_px(width))

    row_heights_px: list[int] = []
    for r in range(1, n_rows + 1):
        dim = ws.row_dimensions.get(r)
        height = dim.height if dim and dim.height else default_row
        row_heights_px.append(_row_height_px(height))

    # Legacy dict forms (kept for back-compat) plus px arrays for exact rendering.
    row_heights = {
        str(idx): dim.height
        for idx, dim in ws.row_dimensions.items()
        if dim.height is not None
    }
    column_widths = {
        str(idx): dim.width
        for idx, dim in ws.column_dimensions.items()
        if dim.width is not None
    }

    print_area = None
    try:
        print_area = ws.print_area or None
        if isinstance(print_area, (list, tuple)):
            print_area = print_area[0] if print_area else None
    except Exception:
        print_area = None

    source_range = f"A1:{get_column_letter(max(n_cols, 1))}{max(n_rows, 1)}" if n_rows else ""

    return {
        "name": ws.title,
        "grid": grid,
        "formulas": formulas,
        "styles": styles,
        "mergedCells": merges,
        "rowHeights": row_heights,
        "columnWidths": column_widths,
        "colWidthsPx": col_widths_px,
        "rowHeightsPx": row_heights_px,
        "sourceSheet": ws.title,
        "sourceRange": source_range,
        "printArea": print_area,
    }


def _parse_index(workbook, index_sheet_name: str | None) -> list[dict[str, Any]]:
    if not index_sheet_name:
        return []

    ws = workbook[index_sheet_name]
    rows = [[_norm(c.value) for c in row] for row in ws.iter_rows()]
    rows = [r for r in rows if any(r)]
    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows[:20]):
        low = {x.lower() for x in row if x}
        if low & _INDEX_ALIASES["sheet_tab"] and low & _INDEX_ALIASES["sheet_title"]:
            header_idx = i
            break

    header = rows[header_idx]
    col = _header_map(header)

    entries = []
    for row in rows[header_idx + 1 :]:
        tab = row[col["sheet_tab"]] if 0 <= col["sheet_tab"] < len(row) else ""
        title = row[col["sheet_title"]] if 0 <= col["sheet_title"] < len(row) else ""
        use_source = row[col["use_source"]] if 0 <= col["use_source"] < len(row) else ""
        notes = row[col["notes"]] if 0 <= col["notes"] < len(row) else ""
        include_raw = row[col["include"]] if 0 <= col["include"] < len(row) else ""
        order_raw = row[col["order"]] if 0 <= col["order"] < len(row) else ""
        sheet_code_raw = row[col["sheet_code"]] if 0 <= col["sheet_code"] < len(row) else ""

        if not tab and not title:
            continue

        entries.append(
            {
                "sheetTab": tab,
                "sheetTitle": title or tab,
                "useSource": use_source,
                "notes": notes,
                "orderRaw": order_raw,
                "sheetCodeRaw": sheet_code_raw,
                "include": _included(include_raw, title or tab, use_source),
            }
        )

    return entries


def _infer_metadata_from_labeled_grid(ws: dict[str, Any] | None) -> dict[str, str]:
    """Backward-compatible alias for workbook import paths."""
    return infer_metadata_from_labeled_grid(ws)


def _remap_a1_styles(styles: dict[str, Any], row_map: dict[int, int]) -> dict[str, Any]:
    """Copy A1-keyed styles for retained rows, remapping row numbers."""
    out: dict[str, Any] = {}
    for key, val in (styles or {}).items():
        m = re.fullmatch(r"([A-Z]+)(\d+)", key)
        if not m:
            continue
        old_r = int(m.group(2)) - 1
        if old_r not in row_map:
            continue
        out[f"{m.group(1)}{row_map[old_r] + 1}"] = val
    return out


def _filter_index_payload_for_output(ws: dict[str, Any]) -> dict[str, Any]:
    """Return an index worksheet payload containing included rows only.

    Source tabs keep the original workbook unchanged; this filtered payload is
    used only for the normalized/output Sheet Index page.
    """
    grid = ws.get("grid") or []
    if not grid:
        return ws

    header_idx = 0
    for i, row in enumerate(grid[:20]):
        low = {str(x).lower() for x in row if x}
        if low & _INDEX_ALIASES["sheet_tab"] and low & _INDEX_ALIASES["sheet_title"]:
            header_idx = i
            break
    col = _header_map([str(x) for x in grid[header_idx]])
    include_col = col.get("include", -1)
    if include_col < 0:
        return ws

    keep: list[int] = list(range(header_idx + 1))
    for r in range(header_idx + 1, len(grid)):
        row = grid[r]
        include_raw = row[include_col] if include_col < len(row) else ""
        title = row[col["sheet_title"]] if 0 <= col["sheet_title"] < len(row) else ""
        use_source = row[col["use_source"]] if 0 <= col["use_source"] < len(row) else ""
        if _included(include_raw, title, use_source):
            keep.append(r)

    row_map = {old: new for new, old in enumerate(keep)}
    out = dict(ws)
    out["grid"] = [list(grid[r]) for r in keep if r < len(grid)]
    out["styles"] = _remap_a1_styles(ws.get("styles") or {}, row_map)
    out["rowHeightsPx"] = [
        (ws.get("rowHeightsPx") or [])[r] if r < len(ws.get("rowHeightsPx") or []) else _DEFAULT_ROW_PX
        for r in keep
    ]
    # Keep only merges whose full row span survived the filter.
    merges = []
    for m in ws.get("mergedCells") or []:
        rows = range(m.get("startRow", 0), m.get("endRow", 0) + 1)
        if all(r in row_map for r in rows):
            nm = dict(m)
            ns = [row_map[r] for r in rows]
            nm["startRow"], nm["endRow"] = min(ns), max(ns)
            merges.append(nm)
    out["mergedCells"] = merges
    return out


def _safe_name(text: str) -> str:
    out = re.sub(r"[^A-Za-z0-9._-]+", "_", (text or "sheet").strip())
    return out[:40] or "sheet"


def _extract_embedded_images(ws, assets_dir, url_prefix: str, sheet_name: str) -> list[dict[str, Any]]:
    """Extract embedded worksheet images to disk; return metadata list.

    Never raises — unsupported/unknown image types are skipped.
    """
    if assets_dir is None or not url_prefix:
        return []
    out: list[dict[str, Any]] = []
    images = getattr(ws, "_images", None) or []
    for i, img in enumerate(images):
        try:
            data = None
            getter = getattr(img, "_data", None)
            if callable(getter):
                data = getter()
            if not data:
                ref = getattr(img, "ref", None)
                if hasattr(ref, "read"):
                    data = ref.read()
            if not data:
                continue
            fmt = (getattr(img, "format", None) or "png").lower()
            if fmt not in ("png", "jpg", "jpeg", "gif", "bmp", "webp"):
                fmt = "png"
            fname = f"{_safe_name(sheet_name)}_{i + 1}.{fmt}"
            (Path(assets_dir) / fname).write_bytes(data)
            anchor_cell = ""
            try:
                frm = img.anchor._from  # type: ignore[attr-defined]
                anchor_cell = f"{get_column_letter(frm.col + 1)}{frm.row + 1}"
            except Exception:
                anchor_cell = ""
            out.append(
                {
                    "id": f"xlimg_{_safe_name(sheet_name)}_{i + 1}",
                    "sheetName": sheet_name,
                    "anchorCell": anchor_cell,
                    "width": getattr(img, "width", None),
                    "height": getattr(img, "height", None),
                    "name": fname,
                    "url": f"{url_prefix}/{fname}",
                }
            )
        except Exception:
            continue
    return out


_ASSET_IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp")


def _asset_slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (text or "").lower())


def _match_page_asset(project_id: str, sheet_code: str, title: str) -> dict[str, Any] | None:
    """Best-effort match of a page-specific reference screenshot in the
    project's ``assets/screenshots`` scaffold folder, by sheet code or
    slugified title (Phase C rule: populate blank drawing/layout pages).

    This folder is normally empty on a fresh import — it primarily supports
    a future "drop a reference PNG next to the import" workflow. Returns
    ``None`` (never raises) when no match exists.
    """
    if not project_id:
        return None
    try:
        from core.project_store import ProjectStore

        docs_dir = Path(__file__).resolve().parents[1] / ".docs"
        screenshots_dir = ProjectStore(docs_dir).dir_for(project_id) / "assets" / "screenshots"
    except Exception:
        return None
    if not screenshots_dir.is_dir():
        return None
    code_key = _asset_slug(sheet_code)
    title_key = _asset_slug(title)
    try:
        candidates = sorted(p for p in screenshots_dir.iterdir() if p.is_file())
    except OSError:
        return None
    for f in candidates:
        if f.suffix.lower() not in _ASSET_IMAGE_EXTS:
            continue
        name_key = _asset_slug(f.stem)
        if (code_key and code_key in name_key) or (title_key and len(title_key) > 3 and title_key in name_key):
            return {"name": f.name, "url": f"/api/assets/{project_id}/{f.name}"}
    return None


def _blank_page_placeholder_message(sheet_tab: str, title: str) -> str:
    """Export-visible placeholder text for a blank canvas/drawing page
    (FINAL SA31 POLISH 4I, Phase E). ``NormalizedPage.tsx`` renders this
    exact text at export time when the page has no image block and no user
    overlay content — never a silently blank page.

    Location / layout / schematic blanks all use the same clear draft note.
    """
    return "DRAWING TO BE INSERTED"


def _canonical_sheet_code(idx: dict[str, Any] | None, sheet_tab: str, order_cursor: int) -> str:
    """Resolve the title-block / index Sheet Code (FINAL SA31 POLISH 4I).

    Precedence:
      1. Index Sheet Code column (canonical engineering code).
      2. Sheet-code prefix baked into the worksheet tab name.
      3. ``EMS {order}`` last-resort placeholder — never bare ``{order}.0``,
         which previously leaked sequential Order into SHEET NO.
    """
    raw = (idx.get("sheetCodeRaw") if idx else "") or ""
    if raw.strip():
        return raw.strip()
    from_tab = _sheet_code_from_tab(sheet_tab)
    if from_tab:
        return from_tab
    return f"EMS {order_cursor}"


def _continuation_index_title(page: dict[str, Any]) -> str:
    """Prefer a clean '— Continued' title for index rows (Phase A)."""
    title = (page.get("sheetTitle") or "").strip()
    base = re.sub(r"\s*[—-]\s*CONTINUED\s*$", "", title, flags=re.IGNORECASE).strip()
    if "lcp" in base.lower() and "panel" in base.lower():
        return "LCP Panel Schedule — Continued"
    if base:
        return f"{base} — Continued"
    return title or "Continued"


def _continuation_index_page_type(page: dict[str, Any]) -> str:
    """Stable Page Type string for a generated continuation index row."""
    existing = (page.get("pageType") or "").strip().lower()
    if existing in {"io-table-continuation", "io-table", "matrix"}:
        return "io-table-continuation"
    family = (page.get("pageFamily") or "").strip().lower()
    if family in {"paneldetail", "ioschedule", "matrix", "panel detail"}:
        return "io-table-continuation"
    title = (page.get("sheetTitle") or "").lower()
    if "lcp" in title or "panel schedule" in title:
        return "io-table-continuation"
    return "continuation"


def _find_index_header_row(grid: list[list[str]]) -> int:
    for i, row in enumerate(grid[:20]):
        low = {str(x).lower() for x in row if x}
        if low & _INDEX_ALIASES["sheet_tab"] and low & _INDEX_ALIASES["sheet_title"]:
            return i
    return 0


def _index_row_sheet_code(row: list[str], code_col: int) -> str:
    if code_col < 0 or code_col >= len(row):
        return ""
    return str(row[code_col] or "").strip()


def _append_continuation_rows_to_index(project: dict[str, Any]) -> None:
    """Keep the rendered Sheet Index in exact output order (FINAL SA31 POLISH 4I).

    ``compose_pages`` may split an oversized sheet (e.g. LCP) into a base page
    plus ``generatedContinuation`` pages (``EMS 1.4`` -> ``EMS 1.4a``) *after*
    the index was rendered from workbook rows. Without this step the index
    under-counts real output and appends incomplete continuation rows at the
    bottom. Call after ``compose_pages``, before ``recalc_page_numbers``.

    Rules:
      - Insert each continuation immediately after its base sheet-code row.
      - Fill Sheet Code / Page Title / Family / Page Type / Include / Order.
      - Re-number Order to match physical export sequence for included rows.
      - Never duplicate an already-present continuation sheet code.
    """
    index_page = next(
        (p for p in project.get("pages", []) if p.get("pageType") == "index" and p.get("renderMode") == "excel_exact"),
        None,
    )
    if not index_page:
        return
    block = next((b for b in (index_page.get("blocks") or []) if b.get("type") == "excelRange"), None)
    if not block:
        return
    grid = block.get("grid") or []
    if not grid:
        return

    header_idx = _find_index_header_row(grid)
    col = _header_map([str(x) for x in grid[header_idx]])
    code_col = col.get("sheet_code", -1)
    title_col = col.get("sheet_title", -1)
    include_col = col.get("include", -1)
    order_col = col.get("order", -1)
    tab_col = col.get("sheet_tab", -1)
    family_col = col.get("family", -1)
    page_type_col = col.get("page_type", -1)
    notes_col = col.get("notes", -1)
    if title_col < 0:
        return

    n_cols = len(grid[header_idx])
    row_heights = list(block.get("rowHeights") or [])
    while len(row_heights) < len(grid):
        row_heights.append(_DEFAULT_ROW_PX)
    default_row_h = row_heights[-1] if row_heights else _DEFAULT_ROW_PX

    # Map sheet code -> last matching body-row index (for insertion).
    def code_row_map() -> dict[str, int]:
        out: dict[str, int] = {}
        for r in range(header_idx + 1, len(grid)):
            code = _index_row_sheet_code(grid[r], code_col)
            if code:
                out[code] = r
        return out

    included_pages = [p for p in project.get("pages", []) if p.get("include", True)]
    for page in included_pages:
        if not page.get("generatedContinuation"):
            continue
        cont_code = (page.get("displaySheetCode") or page.get("sheetCode") or "").strip()
        if not cont_code:
            continue
        codes = code_row_map()
        if cont_code in codes:
            # Already listed — refresh metadata in place rather than duplicating.
            r = codes[cont_code]
            if title_col >= 0:
                grid[r][title_col] = _continuation_index_title(page)
            if include_col >= 0:
                grid[r][include_col] = "YES"
            if family_col >= 0 and not str(grid[r][family_col] if family_col < len(grid[r]) else "").strip():
                family = page.get("pageFamily") or ""
                grid[r][family_col] = "Lighting" if family in ("panelDetail", "matrix", "ioSchedule") else family
            if page_type_col >= 0:
                grid[r][page_type_col] = _continuation_index_page_type(page)
            continue

        # Prefer base page's sheet code; fall back to stripping trailing letter.
        base_page = next(
            (p for p in project.get("pages", []) if p.get("id") == page.get("continuationOf")),
            None,
        )
        base_code = (
            ((base_page.get("displaySheetCode") or base_page.get("sheetCode")) if base_page else "")
            or re.sub(r"[a-z]$", "", cont_code, flags=re.IGNORECASE)
        ).strip()

        insert_at = codes.get(base_code, len(grid) - 1) + 1
        # Skip past any other already-inserted continuations for this base.
        while insert_at < len(grid):
            existing = _index_row_sheet_code(grid[insert_at], code_col)
            if existing.startswith(base_code) and existing != base_code and len(existing) == len(base_code) + 1:
                insert_at += 1
                continue
            break

        new_row = [""] * n_cols
        if include_col >= 0:
            new_row[include_col] = "YES"
        if code_col >= 0:
            new_row[code_col] = cont_code
        if title_col >= 0:
            new_row[title_col] = _continuation_index_title(page)
        if tab_col >= 0:
            new_row[tab_col] = page.get("sheetTab") or (base_page.get("sheetTab") if base_page else "") or ""
        if family_col >= 0:
            family = page.get("pageFamily") or (base_page.get("pageFamily") if base_page else "") or ""
            if family in ("panelDetail", "matrix", "ioSchedule"):
                new_row[family_col] = "Lighting"
            elif family:
                new_row[family_col] = str(family)
            else:
                # Copy family from the base index row when available.
                base_row_i = codes.get(base_code)
                if base_row_i is not None and family_col < len(grid[base_row_i]):
                    new_row[family_col] = str(grid[base_row_i][family_col] or "")
        if page_type_col >= 0:
            new_row[page_type_col] = _continuation_index_page_type(page)
        if notes_col >= 0:
            new_row[notes_col] = ""

        grid.insert(insert_at, new_row)
        row_heights.insert(insert_at, default_row_h)

    # Re-stamp Order to physical included-page sequence so index Order matches
    # "SHEET X OF Y" and the actual exported page order.
    order_by_code = {
        (p.get("displaySheetCode") or p.get("sheetCode") or "").strip(): int(p.get("order") or 0)
        for p in included_pages
        if (p.get("displaySheetCode") or p.get("sheetCode") or "").strip()
    }
    if order_col >= 0:
        for r in range(header_idx + 1, len(grid)):
            code = _index_row_sheet_code(grid[r], code_col)
            if code and code in order_by_code:
                while len(grid[r]) < n_cols:
                    grid[r].append("")
                grid[r][order_col] = str(order_by_code[code])

    block["grid"] = grid
    block["rowHeights"] = row_heights
    # Keep preferred widths in sync when column geometry was already applied.
    if block.get("colWidths") and len(block["colWidths"]) != n_cols:
        family = index_page.get("pageFamily") or "index"
        _apply_table_geometry(block, family if family != "index" else "table", "index")


def _tabular_enough(ws: dict[str, Any]) -> bool:
    """True when a worksheet has a real grid worth rendering as an exact range."""
    grid = ws.get("grid") or []
    if len(grid) < 2:
        return False
    ncols = max((len(r) for r in grid), default=0)
    if ncols < 2:
        return False
    filled = sum(1 for row in grid for c in row if (c or "").strip())
    return filled >= 3


def _should_use_excel_exact(page_type: str, family: str, ws: dict[str, Any]) -> bool:
    """Render the worksheet range verbatim (excel_exact) for all tabular EMS pages."""
    if family == "companyInfo":
        return False
    if page_type in ("cover", "canvas", "hybrid", "underlay"):
        return False
    if not _tabular_enough(ws):
        return False
    if page_type == "index":
        return True
    if family in EXCEL_EXACT_FAMILIES or family == "text":
        return True
    return False


def _looks_company_info(sheet_tab: str, title: str, use_source: str = "") -> bool:
    blob = f"{sheet_tab} {title} {use_source}".lower()
    return "company info" in blob or "singh360 company" in blob or "company/reference" in blob


def _split_settings_for_page(family: str, page_type: str, use_exact: bool) -> dict[str, Any]:
    """Default continuation/pagination settings per page family.

    Index/cover/scope/workflow never auto-split. Dense tables (matrix, IDF) prefer
    scaling down to minScale before row-splitting."""
    if not use_exact:
        return {
            "splitMode": "none",
            "allowContinuation": False,
            "minScale": EXCEL_MIN_SCALE,
            "scaleMode": "fit_body",
            "trimBlankRows": True,
            "trimBlankColumns": True,
        }

    settings: dict[str, Any] = {
        "splitMode": "auto_rows",
        "allowContinuation": True,
        "minScale": EXCEL_MIN_SCALE,
        "scaleMode": "fit_body",
        # Normalized/export tables trim trailing blank worksheet columns/rows
        # by default; the Source tab is unaffected (FINAL RENDER POLISH 4G,
        # Phase B/H). Per-page override lives on the page dict, not here.
        "trimBlankRows": True,
        "trimBlankColumns": True,
    }
    if family == "matrix":
        settings["minScale"] = EXCEL_MIN_SCALE
    elif family == "idfTable":
        # Fallback safety net for IDF-ish sheets that aren't detected as a
        # Port/Label network table (those get the dedicated two-up layout
        # instead); keeps this rare fallback above the 6.5pt readable floor.
        settings["minScale"] = 0.75
    elif family in ("ioSchedule", "panelDetail", "rackLayout"):
        settings["minScale"] = 0.75
    elif page_type == "index":
        # A Sheet Index / TOC never spills onto a continuation page.
        settings.update({"splitMode": "none", "allowContinuation": False, "scaleMode": "fit_body"})
    elif family == "text":
        # PHASE B fix: text-family pages (guideline/instruction/scope/
        # workflow/notes) used to hard-disable continuation, which forced a
        # too-tall table to shrink uniformly below the readable floor instead
        # of splitting (the "tiny unreadable strip" + TABLE OVERFLOW bug).
        # Continuation is now allowed as a last resort — a table that already
        # fits at >= minScale behaves exactly as before (no split happens),
        # but a genuinely too-tall table now falls back to a continuation
        # page (e.g. EMS 17.0 -> EMS 17.0a) instead of being crushed unreadable.
        settings.update(
            {
                "splitMode": "auto_rows",
                "allowContinuation": True,
                "scaleMode": "fit_body",
                "minScale": max(EXCEL_MIN_SCALE, NARRATIVE_MIN_FONT_SIZE / 9.0),
            }
        )
    return settings


def _layout_profile_for(family: str, page_type: str, blob: str) -> str:
    """Rendering-options profile per page type (TABLE STYLE 4F / SA31 4I)."""
    if family == "companyInfo":
        return "company_info"
    if family == "idfTable":
        return "network_48_port"
    if family in ("matrix", "ioSchedule", "panelDetail", "rackLayout"):
        return "io_table"
    if family == "text" and "instruction" in blob:
        return "instruction_table"
    # Front-matter narrative pages with long Section / Scope Language cells.
    if family == "text" and any(k in blob for k in ("scope", "workflow", "milestone")):
        return "front_matter_narrative_table"
    return "front_matter_table"


# --------------------------------------------------------------------------
# RDM / IDF network table — special two-up layout (TABLE STYLE 4F, Phase B).
#
# Default is always ONE full-width table (no rotation). Only when the real
# port count makes a single stack unreadable does the layout switch to a
# two-up (ports 1-N / N+1-total) side-by-side layout with essential columns
# only. Never invents data: a column with no source match renders blank.
# --------------------------------------------------------------------------
_IDF_TARGET_FONT = DENSE_FONT_SIZE  # 7.0pt
_IDF_PREFERRED_FONT = 7.5  # FINAL SA31 POLISH 4I Phase D preferred readable size
_IDF_MIN_FONT = ABSOLUTE_MIN_FONT_SIZE  # 6.5pt
_IDF_ROW_H = 20
_IDF_HEADER_H = 24
# Two-up scale-up targets (Phase D): grow row height / font until the table
# uses roughly 65–75% of the safe body when it would otherwise look tiny.
_IDF_SCALE_TARGET_MIN = 0.65
_IDF_SCALE_TARGET_MAX = 0.75
# PHASE C fix (SA31 export tables): a proper 48-port switch schedule always
# shows Controller ID, IP Address, and Network as their own columns — never
# merged into one "Controller / IP" column and never folded into Notes.
# Widths below are tuned so all 10 required columns fit one two-up half
# (~751px, see ``_build_idf_network_block``'s ``two_up_usable``) at a
# readable font; "Terminated By" only renders when explicitly requested via
# ``show_terminated_by`` and only if it still fits.
_IDF_COL_W = {
    "Port": 40,
    "Label": 64,
    "Device / Drop": 130,
    "Controller ID": 70,
    "IP Address": 82,
    "Network": 62,
    "From": 60,
    "To": 60,
    "Path": 170,
    "Cable": 66,
    "Notes": 117,
    "Terminated By": 70,
}
# Required output columns (Phase C) — always present when the source header
# maps to them, in this exact left-to-right order.
_IDF_REQUIRED_COLS = (
    "Port",
    "Label",
    "Device / Drop",
    "Controller ID",
    "IP Address",
    "Network",
    "From",
    "To",
    "Cable",
    "Notes",
)


def _idf_col_index(headers: list[str], *keys: str) -> int | None:
    low = [_normalize_header_cell(h) for h in headers]
    for i, h in enumerate(low):
        if not h:
            continue
        if any(k in h for k in keys):
            return i
    return None


def _idf_header_row(grid: list[list[str]]) -> int | None:
    """First row that looks like a Port/Label or Port/Controller-ID network header."""
    for r, row in enumerate(grid[:12]):
        low = [_normalize_header_cell(c) for c in row]
        has_port = any("port" in c for c in low if c)
        has_label = any(("label" in c) or ("device" in c) or ("drop" in c) for c in low if c)
        has_ctrl = any(("controller" in c) or ("ip address" in c) or ("ip addr" in c) for c in low if c)
        has_network = any("network" in c for c in low if c)
        if has_port and (has_label or has_ctrl or has_network):
            return r
        if has_ctrl and has_network:
            return r
    return None


def _is_idf_network_table(ws: dict[str, Any], family: str) -> tuple[bool, int | None]:
    if family != "idfTable":
        return False, None
    header_row = _idf_header_row(ws.get("grid") or [])
    return header_row is not None, header_row


def _idf_columns(
    headers: list[str], *, show_terminated_by: bool = False
) -> list[tuple[str, tuple[int, ...]]]:
    """Map source headers onto the required RDM/IDF network columns.

    PHASE C fix (SA31 export tables): Controller ID, IP Address, and Network
    are always their own separate output columns — never merged into a
    single "Controller / IP" column and never folded into Notes. Notes only
    ever contains the source workbook's actual Notes/Remarks/Comment column.
    "Terminated By" is detected but only emitted when ``show_terminated_by``
    is set (default hidden per spec); callers may still drop it afterwards
    if it does not fit (see ``_build_idf_network_block``).
    """
    idx = {
        "port": _idf_col_index(headers, "port"),
        "label": _idf_col_index(headers, "label"),
        "device": _idf_col_index(headers, "device", "drop", "location"),
        "controllerId": _idf_col_index(headers, "controller id", "controller"),
        "ip": _idf_col_index(headers, "ip address", "ip addr", "ip"),
        "network": _idf_col_index(headers, "network", "vlan"),
        "from": _idf_col_index(headers, "from"),
        "to": _idf_col_index(headers, "to"),
        "cable": _idf_col_index(headers, "cable"),
        "notes": _idf_col_index(headers, "notes", "remark", "comment"),
        "terminated": _idf_col_index(headers, "terminated by", "terminated"),
    }
    cols: list[tuple[str, tuple[int, ...]]] = [
        ("Port", (idx["port"],) if idx["port"] is not None else ()),
        ("Label", (idx["label"],) if idx["label"] is not None else ()),
        ("Device / Drop", (idx["device"],) if idx["device"] is not None else ()),
        ("Controller ID", (idx["controllerId"],) if idx["controllerId"] is not None else ()),
        ("IP Address", (idx["ip"],) if idx["ip"] is not None else ()),
        ("Network", (idx["network"],) if idx["network"] is not None else ()),
        ("From", (idx["from"],) if idx["from"] is not None else ()),
        ("To", (idx["to"],) if idx["to"] is not None else ()),
        ("Cable", (idx["cable"],) if idx["cable"] is not None else ()),
        ("Notes", (idx["notes"],) if idx["notes"] is not None else ()),
    ]
    # Only keep a required column when the source maps to it OR it is one of
    # the default RDM/IDF export columns (always show the column shell).
    cols = [c for c in cols if c[1] or c[0] in _IDF_REQUIRED_COLS]

    if show_terminated_by and idx["terminated"] is not None:
        cols.append(("Terminated By", (idx["terminated"],)))

    # Only combine From/To -> Path if the column count is still too wide
    # (kept as a last-resort fallback; Controller ID/IP/Network are exempt).
    if len(cols) > 11 and idx["from"] is not None and idx["to"] is not None:
        cols = [c for c in cols if c[0] not in ("From", "To")]
        insert_at = min(6, len(cols))
        cols.insert(insert_at, ("Path", (idx["from"], idx["to"])))
    return cols


_DEVICE_ABBREVIATIONS = {
    "controller": "Ctrl",
    "connection": "Conn",
    "connector": "Conn",
    "distribution": "Dist",
    "management": "Mgmt",
    "network": "Net",
    "wireless": "WiFi",
    "equipment": "Eqp",
}


def _abbreviate_device_text(text: str, max_chars: int) -> str:
    """Best-effort shrink of a Device / Drop value that is still too wide
    after Terminated By has already been hidden (Phase C rule 8: abbreviate
    Device / Drop only if required, never touch Controller ID/IP/Network).
    Never invents new text — only shortens known long words and truncates
    with an ellipsis as a last resort.
    """
    if len(text) <= max_chars:
        return text
    out = text
    for long_word, short in _DEVICE_ABBREVIATIONS.items():
        if long_word in out.lower():
            idx_ = out.lower().index(long_word)
            out = out[:idx_] + short + out[idx_ + len(long_word):]
    if len(out) <= max_chars:
        return out
    return out[: max(1, max_chars - 1)].rstrip() + "…"


def _idf_cell_value(row: list[str], spec: tuple[int, ...]) -> str:
    parts = [str(row[i]).strip() for i in spec if i < len(row) and str(row[i] or "").strip()]
    return " / ".join(parts)


def _build_idf_network_block(
    ws: dict[str, Any],
    header_row: int,
    block_id: str,
    *,
    row_slice: tuple[int, int] | None = None,
    show_terminated_by: bool = False,
) -> dict[str, Any]:
    """Build the special RDM/IDF network table block (single full-width table
    by default; two-up ports 1-N / N+1-total only when a single stack would
    fall below the readable font floor).

    ``row_slice`` (start, end-exclusive) restricts the block to a subset of
    data rows — used only by the Phase E hard-split fallback below, which
    turns an oversized network table into two balanced *single*-layout pages
    instead of ever rendering two-up below the 6.5pt floor.

    FINAL SA31 POLISH 4I Phase D: when two-up content uses <55% of safe body
    height, grow row height / prefer 7–7.5pt fonts (never below 6.5pt) until
    the block fills ~65–75%.

    PHASE C fix (SA31 export tables): Controller ID, IP Address, and Network
    are always real columns — they are never folded into Notes. When the
    assembled table is too wide for two-up, the fallback order is: hide
    Terminated By first (default hidden anyway), then abbreviate Device /
    Drop text; Controller ID / IP Address / Network / Notes are never
    touched.
    """
    grid = ws.get("grid") or []
    headers_src = grid[header_row] if header_row < len(grid) else []

    def _assemble(
        *, terminated_by: bool, device_max_chars: int | None = None
    ) -> tuple[list[str], list[int], list[list[str]], int]:
        cols = _idf_columns(headers_src, show_terminated_by=terminated_by)
        hdrs = [c[0] for c in cols]
        widths = [_IDF_COL_W.get(h, 90) for h in hdrs]
        device_i = hdrs.index("Device / Drop") if "Device / Drop" in hdrs else None
        rows_all: list[list[str]] = []
        for row in grid[header_row + 1:]:
            vals = [_idf_cell_value(row, spec) for _, spec in cols]
            if device_i is not None and device_max_chars:
                vals[device_i] = _abbreviate_device_text(vals[device_i], device_max_chars)
            if any(v for v in vals):
                rows_all.append(vals)
        return hdrs, widths, rows_all, sum(widths)

    usable_w = BODY_W - 80
    two_up_usable = (usable_w - 18) // 2  # gap between the two stacks
    fits_budget = usable_w if row_slice is not None else two_up_usable

    # 1) All required columns + Terminated By if explicitly requested.
    headers, col_widths, all_data_rows, single_w = _assemble(terminated_by=show_terminated_by)

    # 2) Too wide -> hide Terminated By first (Phase C rule 8), even if the
    # caller asked for it — "enough room" gates it too.
    if single_w > fits_budget and show_terminated_by:
        headers, col_widths, all_data_rows, single_w = _assemble(terminated_by=False)

    # 3) Still too wide -> abbreviate Device / Drop text (never touch
    # Controller ID / IP Address / Network / Notes).
    if single_w > fits_budget:
        base_device_w = _IDF_COL_W.get("Device / Drop", 130)
        overflow = single_w - fits_budget
        shrunk_w = max(70, base_device_w - overflow)
        max_chars = max(8, int(shrunk_w / 7.2))
        headers, col_widths, all_data_rows, single_w = _assemble(
            terminated_by=False, device_max_chars=max_chars
        )
        if "Device / Drop" in headers:
            col_widths[headers.index("Device / Drop")] = shrunk_w
            single_w = sum(col_widths)

    total_n = len(all_data_rows)

    title_lines = []
    for row in grid[:header_row]:
        line = " ".join(str(c).strip() for c in row if str(c or "").strip())
        if line:
            title_lines.append(line)
    section_title = " — ".join(title_lines)

    if row_slice is not None:
        start, end = row_slice
        data_rows = all_data_rows[start:end]
        port_offset = start
    else:
        data_rows = all_data_rows
        port_offset = 0
    n = len(data_rows)

    row_h = _IDF_ROW_H
    header_h = _IDF_HEADER_H
    single_h = header_h + n * row_h
    half = (n + 1) // 2 if n else 0
    two_up_h = header_h + half * row_h

    warnings: list[str] = []
    needs_hard_split = False
    if single_w <= usable_w and single_h <= BODY_BUDGET:
        layout_mode = "single"
        font_size = _IDF_PREFERRED_FONT
    elif row_slice is None and two_up_h > BODY_BUDGET:
        needs_hard_split = True
        layout_mode = "two_up"
        font_size = _IDF_MIN_FONT
    else:
        layout_mode = "two_up"
        font_size = _IDF_PREFERRED_FONT

    # Phase D scale-up: if two-up (or single) uses <55% of safe body, grow
    # row height / keep preferred font until ~65–75% fill, without exceeding
    # BODY_BUDGET or dropping below the 6.5pt floor.
    content_h = single_h if (layout_mode == "single" or needs_hard_split) else two_up_h
    fill_ratio = content_h / BODY_BUDGET if BODY_BUDGET else 1.0
    if not needs_hard_split and fill_ratio < 0.55 and n > 0:
        target_h = int(BODY_BUDGET * _IDF_SCALE_TARGET_MIN)
        rows_for_h = n if layout_mode == "single" else max(1, half)
        grown_row = max(row_h, int((target_h - header_h) / rows_for_h))
        # Cap growth so we stay within the 75% band.
        max_h = int(BODY_BUDGET * _IDF_SCALE_TARGET_MAX)
        while header_h + grown_row * rows_for_h > max_h and grown_row > row_h:
            grown_row -= 1
        row_h = grown_row
        single_h = header_h + n * row_h
        two_up_h = header_h + half * row_h
        content_h = single_h if layout_mode == "single" else two_up_h
        font_size = max(_IDF_MIN_FONT, min(_IDF_PREFERRED_FONT, font_size))

    left_rows = data_rows[:half] if layout_mode == "two_up" and not needs_hard_split else []
    right_rows = data_rows[half:] if layout_mode == "two_up" and not needs_hard_split else []
    port_left = f"{port_offset + 1}–{port_offset + half}" if half else ""
    port_right = f"{port_offset + half + 1}–{port_offset + n}" if n > half else ""

    return {
        "id": block_id,
        "type": "idfNetworkTable",
        "sourceWorksheetId": ws["id"],
        "sourceSheet": ws.get("sourceSheet") or ws.get("name", ""),
        "sourceRange": ws.get("sourceRange", ""),
        "renderMode": "excel_exact",
        "layoutMode": "single" if needs_hard_split else layout_mode,
        "sectionTitle": section_title,
        "headers": headers,
        "rows": data_rows if (layout_mode == "single" or needs_hard_split) else [],
        "leftRows": left_rows,
        "rightRows": right_rows,
        "portRangeLeft": port_left if not needs_hard_split else "",
        "portRangeRight": port_right if not needs_hard_split else "",
        "portRangeLabel": f"{port_offset + 1}–{port_offset + n}" if needs_hard_split else "",
        "colWidths": col_widths,
        "rowHeight": row_h,
        "headerHeight": header_h,
        "fontSize": _IDF_TARGET_FONT if needs_hard_split else font_size,
        "contentWidth": single_w if layout_mode == "single" or needs_hard_split else min(usable_w, single_w * 2 + 18),
        "contentHeight": content_h,
        "sourceRowCount": n,
        "totalRowCount": total_n,
        "bodyRowFillMode": "none",
        "gridLines": True,
        "styleRole": "network-two-up",
        "splitMode": "none",
        "allowContinuation": False,
        "minScale": 1.0,
        "scaleMode": "fit_body",
        "orientation": "landscape",
        "editable": False,
        "layoutWarnings": warnings,
        "needsHardSplit": needs_hard_split,
        "hardSplitBoundary": _idf_hard_split_boundary(total_n) if (needs_hard_split and row_slice is None) else None,
        "scaledUp": bool(row_h > _IDF_ROW_H),
    }


def _idf_hard_split_boundary(total_n: int) -> int:
    """Balanced row split point for the Phase E hard-split fallback."""
    return (total_n + 1) // 2


_NARROW_COL_HEAD_KEYWORDS = ("no", "#", "id", "qty", "type", "addr", "i/o", "io", "point", "ckt", "step")

# PHASE D fix (SA31 export tables): I/O and LCP panel schedule columns holding
# short technical tokens (0-10VDC, 10K2, NO, NO*, NC, DI, AIO1, PR0650CD-TDB,
# PR0663, ...) used to wrap character-by-character in a too-narrow fixed
# column. A token is any single "word" (no internal spaces) of letters,
# digits, and the punctuation these part numbers/ranges actually use.
_TECH_TOKEN_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-/*.]{0,17}$")
_TECH_TOKEN_FAMILIES = ("ioSchedule", "panelDetail", "rackLayout")


def _is_tech_token(text: str) -> bool:
    t = (text or "").strip()
    if not t or " " in t:
        return False
    return bool(_TECH_TOKEN_RE.match(t))


def _technical_token_columns(grid: list[list[str]], family: str, header_rows: int) -> dict[int, int]:
    """Columns dominated by short nowrap-worthy technical tokens.

    Returns ``{col_index: longest_token_length}`` for columns where at least
    60% of non-empty data-row values (excluding the header band) are
    single-token technical values — never guesses on prose/description
    columns, since those values contain spaces and are filtered out by
    ``_is_tech_token`` immediately.
    """
    if family not in _TECH_TOKEN_FAMILIES or not grid:
        return {}
    n_cols = max((len(r) for r in grid), default=0)
    out: dict[int, int] = {}
    for c in range(n_cols):
        values = [
            str(row[c]).strip()
            for r, row in enumerate(grid)
            if r >= header_rows and c < len(row) and str(row[c] or "").strip()
        ]
        if len(values) < 2:
            continue
        token_like = [v for v in values if _is_tech_token(v)]
        if len(token_like) / len(values) >= 0.6:
            out[c] = max(len(v) for v in token_like)
    return out


def _token_column_min_width(longest_token_len: int, font_px: int = 12) -> int:
    """Minimum column width (px) that fits ``longest_token_len`` characters on
    one line at ``font_px`` without character-by-character wrapping."""
    avg_char_w = max(5.5, font_px * 0.52)
    return int(round(longest_token_len * avg_char_w)) + 16
_STATUS_COL_HEAD_KEYWORDS = ("status", "state", "phase")
_SECTION_COL_HEAD_KEYWORDS = ("section", "phase", "category", "item", "milestone", "topic", "step")
_NARRATIVE_COL_HEAD_KEYWORDS = (
    "scope", "language", "description", "instruction", "deliverable", "detail", "narrative", "guideline",
)
_NOTES_COL_HEAD_KEYWORDS = ("notes", "remark", "comment")


def _find_header_row_index(grid: list[list[str]]) -> int:
    """The real column-header row within a normalized excel_exact grid.

    Many real sheets (SA31 instruction/BOM/matrix pages confirmed by
    inspection) lead with title/subtitle/blank rows above the actual column
    header (e.g. "EC FIELD INSTRUCTIONS" / project subtitle / blank / blank /
    "Step | Instruction"). Those title rows populate only a single leading
    cell; a real header row populates most/all columns with short text. Scan
    the first 10 rows for that shape and use it instead of always assuming
    row 0 is the header (FINAL RELEASE CLEANUP 4H+SA38, Phase D fix — column
    width/weight classification was silently inert on any sheet whose header
    isn't literally the first row).
    """
    n_cols = max((len(r) for r in grid), default=0)
    if not n_cols:
        return 0
    for r in range(min(len(grid), 10)):
        row = grid[r]
        non_empty = [str(v).strip() for v in row if str(v or "").strip()]
        if len(non_empty) < max(2, n_cols - 1):
            continue
        if any(len(v) > 60 for v in non_empty):
            continue
        return r
    return 0


def _col_header_class(head: str) -> str:
    h = (head or "").strip().lower()
    if any(k in h for k in _SECTION_COL_HEAD_KEYWORDS) and not any(k in h for k in _NARRATIVE_COL_HEAD_KEYWORDS):
        return "section"
    if any(k in h for k in _STATUS_COL_HEAD_KEYWORDS):
        return "status"
    if any(k in h for k in _NOTES_COL_HEAD_KEYWORDS):
        return "notes"
    if any(k in h for k in _NARRATIVE_COL_HEAD_KEYWORDS):
        return "narrative"
    return "other"


def _notes_column_is_sparse(grid: list[list[str]], notes_cols: list[int], header_row: int) -> bool:
    """True when Notes cells under the header are empty / mostly empty."""
    if not notes_cols:
        return True
    filled = 0
    total = 0
    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        for c in notes_cols:
            total += 1
            if c < len(row) and str(row[c] or "").strip():
                filled += 1
    if total == 0:
        return True
    return (filled / total) < 0.15


def _preferred_narrative_col_widths(grid: list[list[str]]) -> list[int]:
    """Column widths for front_matter_narrative_table (FINAL SA31 POLISH 4I).

    Targets ~85–95% of printable body width with priorities:
      Section 20–24%, Scope Language 56–62%, Status 8–10%, Notes 0–12%.
    Empty/sparse Notes shrink to ~10% or disappear into leftover space.
    """
    n_cols = max((len(r) for r in grid), default=0)
    if not n_cols:
        return []
    header_row = _find_header_row_index(grid)
    header = grid[header_row] if grid else []
    classes = [_col_header_class(str(header[c]) if c < len(header) else "") for c in range(n_cols)]
    notes_idxs = [i for i, c in enumerate(classes) if c == "notes"]
    hide_notes = _notes_column_is_sparse(grid, notes_idxs, header_row)

    # Share of BODY_W (leave ~8% margin → ~92% usable).
    target = int(BODY_W * 0.92)
    shares: list[float] = []
    for cls in classes:
        if cls == "section":
            shares.append(0.22)
        elif cls == "narrative":
            shares.append(0.58)
        elif cls == "status":
            shares.append(0.09)
        elif cls == "notes":
            shares.append(0.0 if hide_notes else 0.10)
        else:
            shares.append(0.12)
    if hide_notes:
        # Reallocate Notes share into Scope Language (or the widest non-status).
        reassigned = False
        for i, cls in enumerate(classes):
            if cls == "narrative":
                shares[i] += 0.10
                reassigned = True
                break
        if not reassigned:
            for i, cls in enumerate(classes):
                if cls == "section":
                    shares[i] += 0.05
                elif cls != "notes":
                    shares[i] += 0.05 / max(1, n_cols - len(notes_idxs))
    total = sum(shares) or 1.0
    widths = [max(48, int(round(target * s / total))) for s in shares]
    if hide_notes:
        for i in notes_idxs:
            widths[i] = max(36, int(BODY_W * 0.08))
    # Clamp to 85–95% body width band.
    lo, hi = int(BODY_W * 0.85), int(BODY_W * 0.95)
    cur = sum(widths)
    if cur < lo:
        grow = lo - cur
        order = sorted(range(n_cols), key=lambda i: shares[i], reverse=True)
        for i in order:
            if classes[i] == "notes" and hide_notes:
                continue
            add = max(1, grow // max(1, n_cols - len(notes_idxs)))
            widths[i] += add
            grow -= add
            if grow <= 0:
                break
    elif cur > hi:
        shrink = cur - hi
        order = sorted(range(n_cols), key=lambda i: shares[i])
        for i in order:
            min_w = 36 if classes[i] == "notes" else 64
            can = max(0, widths[i] - min_w)
            take = min(can, shrink)
            widths[i] -= take
            shrink -= take
            if shrink <= 0:
                break
    return widths


def _preferred_col_widths(
    grid: list[list[str]],
    family: str,
    page_type: str,
    layout_profile: str = "",
) -> list[int]:
    """Deterministic normalized column sizing before placement/export."""
    if layout_profile == "front_matter_narrative_table":
        return _preferred_narrative_col_widths(grid)
    if layout_profile in ("front_matter_table", "instruction_table") and family == "text":
        return _preferred_text_instruction_col_widths(grid)
    n_cols = max((len(r) for r in grid), default=0)
    if not n_cols:
        return []
    target = BODY_W - (96 if family not in ("ioSchedule", "panelDetail", "idfTable") else 48)
    min_w = 52 if family in ("ioSchedule", "panelDetail", "idfTable") else 76
    # PHASE B fix: a 2-3 column instruction/guideline table used to cap every
    # column at 360px, which could never reach the ~1500px body width no
    # matter how much leftover width the redistribute loop below had to give
    # away — that undersized natural width is what made Guidelines/Field
    # Instructions render as a tiny left-aligned strip (and, combined with
    # the resulting inflated wrapped-row height, trip the TABLE OVERFLOW
    # floor). Text-family wide/instruction columns may now grow to the full
    # target width; the compact Step/No narrow column below is unaffected.
    max_w = target if family == "text" else (360 if family == "index" else 300)
    # A compact narrow column (Step/No/#/ID counter) on a "text" family sheet
    # (instruction pages) should stay small and not stretch proportionally
    # with the rest of the table (FINAL RELEASE CLEANUP 4H+SA38, Phase D).
    narrow_min_w = 50
    narrow_max_w = 64
    weights: list[float] = []
    is_narrow_col: list[bool] = []
    header = grid[_find_header_row_index(grid)] if grid else []
    for c in range(n_cols):
        values = [str(row[c]) if c < len(row) else "" for row in grid]
        max_len = max((len(v) for v in values), default=1)
        head = (str(header[c]) if c < len(header) else "").lower()
        weight = max(6.0, min(float(max_len), 48.0))
        is_wide = any(k in head for k in ("description", "notes", "instruction", "scope", "remarks", "location"))
        is_narrow = (not is_wide) and any(k in head for k in _NARROW_COL_HEAD_KEYWORDS)
        if is_wide:
            weight *= 1.9
        elif is_narrow:
            weight *= 0.75
        weights.append(weight)
        is_narrow_col.append(is_narrow and family == "text")
    total = sum(weights) or 1.0
    raw = [int(round(target * w / total)) for w in weights]
    widths = [
        max(narrow_min_w, min(narrow_max_w, w)) if is_narrow_col[i] else max(min_w, min(max_w, w))
        for i, w in enumerate(raw)
    ]
    # PHASE D fix: a short technical-token column (Type/Range/Signal, etc.)
    # must be wide enough to hold its longest token on one line — the flat
    # min_w floor above was letting these columns clamp down to 52px, which
    # is narrower than tokens like "PR0650CD-TDB" or "0-10VDC" need, forcing
    # the browser to wrap them character-by-character.
    token_cols = _technical_token_columns(grid, family, _find_header_row_index(grid) + 1)
    for c, longest in token_cols.items():
        if c < len(widths):
            widths[c] = max(widths[c], _token_column_min_width(longest))
    # Distribute leftover width so small tables use the page instead of
    # floating — never grow the compact narrow (Step/No/#) column.
    diff = target - sum(widths)
    guard = 0
    while diff > 0 and guard < 2000:
        changed = False
        for i in sorted(range(n_cols), key=lambda x: weights[x], reverse=True):
            if is_narrow_col[i]:
                continue
            if widths[i] < max_w:
                widths[i] += 1
                diff -= 1
                changed = True
                if diff <= 0:
                    break
        if not changed:
            break
        guard += 1
    return widths


def _estimated_row_heights(
    grid: list[list[str]],
    col_widths: list[int],
    family: str,
    header_rows: int,
    *,
    font_px: int | None = None,
) -> list[int]:
    if font_px is None:
        font_px = 12 if family in ("matrix", "idfTable", "ioSchedule", "panelDetail", "rackLayout") else 13
    line_h = int(round(font_px * 1.22))
    out: list[int] = []
    for r, row in enumerate(grid):
        max_lines = 1
        for c, val in enumerate(row):
            text = " ".join(str(val or "").split())
            if not text:
                continue
            w = max(36, (col_widths[c] if c < len(col_widths) else _DEFAULT_COL_PX) - 8)
            chars = max(7, int(w / max(5.5, font_px * 0.48)))
            words = text.split()
            lines = 1
            cur = 0
            for word in words:
                wl = len(word)
                if cur and cur + 1 + wl > chars:
                    lines += 1
                    cur = wl
                else:
                    cur = cur + (1 if cur else 0) + wl
            # Narrative sections may wrap scope language across several lines.
            max_lines = max(max_lines, min(lines, 12 if family == "text" else 8))
        base = 28 if r < header_rows else 24
        out.append(max(base, line_h * max_lines + 8))
    return out


def _apply_table_geometry(
    block: dict[str, Any],
    family: str,
    page_type: str,
    layout_profile: str = "",
) -> None:
    """Auto-size columns and rows for normalized/output table geometry."""
    grid = block.get("grid") or []
    if not grid:
        return
    profile = layout_profile or block.get("layoutProfile") or ""
    widths = _preferred_col_widths(grid, family, page_type, profile)
    if widths:
        block["colWidths"] = widths
    header_rows = int(block.get("headerRowCount") or 1)
    font_px = None
    if profile == "front_matter_narrative_table":
        # 8.5pt ≈ 11.3 CSS px; keep Section labels from stacking word-by-word.
        font_px = 12
        block["bodyFontPx"] = 12
        block["bodyFontPt"] = NARRATIVE_FONT_SIZE
        block["minFontPt"] = NARRATIVE_MIN_FONT_SIZE
        block["renderProfile"] = NARRATIVE_RENDER_PROFILE
        block["layoutProfile"] = "front_matter_narrative_table"
        block["minScale"] = max(float(block.get("minScale") or EXCEL_MIN_SCALE), NARRATIVE_MIN_FONT_SIZE / 9.0)
    elif profile in ("front_matter_table", "instruction_table") and family == "text":
        font_px = 12
        block["bodyFontPx"] = 12
        block["bodyFontPt"] = NARRATIVE_FONT_SIZE
        block["minFontPt"] = NARRATIVE_MIN_FONT_SIZE
        block["minScale"] = max(float(block.get("minScale") or EXCEL_MIN_SCALE), NARRATIVE_MIN_FONT_SIZE / 9.0)
        header = grid[_find_header_row_index(grid)] if grid else []
        section_cols = [
            i for i in range(len(widths))
            if _col_header_class(str(header[i]) if i < len(header) else "") in ("section", "other")
            and i < len(widths) and widths[i] <= 120
        ]
        narrative_cols = [
            i for i in range(len(widths))
            if _col_header_class(str(header[i]) if i < len(header) else "") == "narrative"
        ]
        nowrap = set(section_cols)
        block["nowrapColumns"] = sorted(nowrap)
        block["preventStackedLabels"] = True
        block["wordWrapColumns"] = sorted(narrative_cols) if narrative_cols else sorted(
            i for i in range(len(widths)) if i not in nowrap
        )
    block["rowHeights"] = _estimated_row_heights(
        grid, block.get("colWidths") or widths, family, header_rows, font_px=font_px,
    )
    block["pageFamily"] = family
    block["bodyRowFillMode"] = "none"
    block["gridLines"] = True
    # Section column: prefer keep-together / nowrap cue for the frontend.
    if profile == "front_matter_narrative_table" and widths:
        header = grid[_find_header_row_index(grid)]
        section_cols = [
            i for i in range(len(widths))
            if _col_header_class(str(header[i]) if i < len(header) else "") == "section"
        ]
        block["nowrapColumns"] = section_cols
        block["preventStackedLabels"] = True
    elif family in _TECH_TOKEN_FAMILIES:
        # PHASE D fix: I/O / LCP panel schedule columns dominated by short
        # technical tokens (0-10VDC, 10K2, NO, NC, DI, AIO1, PR0650CD-TDB...)
        # must not wrap character-by-character — the column width above is
        # already sized to fit the longest token; nowrap keeps it on one
        # line instead of the browser breaking mid-token.
        token_cols = _technical_token_columns(grid, family, header_rows)
        if token_cols:
            block["nowrapColumns"] = sorted(token_cols.keys())
            block["preventStackedLabels"] = True


def _meaningful_style(style: dict[str, Any] | None) -> bool:
    """True when a style carries a real fill or border — i.e. the cell is
    visually meaningful even with no text (a blocked/shaded cell, a ruled
    line) and must never be trimmed as "blank" (Phase B)."""
    if not style:
        return False
    if style.get("fill"):
        return True
    borders = style.get("borders") or {}
    return any(borders.get(side) for side in ("top", "bottom", "left", "right"))


def _print_area_last_col_row(print_area: str | None) -> tuple[int, int] | None:
    """0-based (lastCol, lastRow) of an explicit print-area override, or None.

    Trailing-blank trimming (Phase B) must never eat into a range the source
    workbook explicitly defined as the print area, even if its tail is blank.
    """
    if not print_area:
        return None
    try:
        rng = str(print_area).split(",")[0].split("!")[-1].replace("$", "")
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        return max_col - 1, max_row - 1
    except Exception:
        return None


def _normalize_header_cell(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip()).lower()


def _drop_fully_blank_columns(
    grid: list[list[str]],
    styles_rc: dict[str, Any],
    merges: list[dict[str, Any]],
    header_rows: int,
) -> tuple[list[list[str]], dict[str, Any], list[dict[str, Any]]]:
    """Drop interior/trailing columns that are blank across the full grid."""
    n_rows = len(grid)
    n_cols = max((len(r) for r in grid), default=0)
    if n_cols <= 1:
        return grid, styles_rc, merges

    keep: list[int] = []
    for c in range(n_cols):
        has_value = any(
            c < len(grid[r]) and str(grid[r][c] or "").strip()
            for r in range(n_rows)
        )
        has_style = any(_meaningful_style(styles_rc.get(f"{r}:{c}")) for r in range(n_rows))
        if has_value or has_style:
            keep.append(c)

    if len(keep) == n_cols:
        return grid, styles_rc, merges

    col_map = {old: new for new, old in enumerate(keep)}
    new_grid = [[row[c] if c < len(row) else "" for c in keep] for row in grid]
    new_styles: dict[str, Any] = {}
    for key, val in styles_rc.items():
        try:
            rs, cs = key.split(":")
            r, c = int(rs), int(cs)
        except (ValueError, AttributeError):
            continue
        if c in col_map:
            new_styles[f"{r}:{col_map[c]}"] = val
    new_merges: list[dict[str, Any]] = []
    for m in merges:
        sc, ec = m.get("startCol", 0), m.get("endCol", 0)
        if sc not in col_map or ec not in col_map:
            continue
        nm = dict(m)
        nm["startCol"] = col_map[sc]
        nm["endCol"] = col_map[ec]
        new_merges.append(nm)
    return new_grid, new_styles, new_merges


def _drop_blank_spacer_rows(
    grid: list[list[str]],
    styles_rc: dict[str, Any],
    header_rows: int,
) -> list[list[str]]:
    """Drop fully blank spacer rows below the header band (not section dividers)."""
    if len(grid) <= header_rows + 1:
        return grid
    out = grid[:header_rows]
    for r in range(header_rows, len(grid)):
        row = grid[r]
        if any(str(v or "").strip() for v in row):
            out.append(row)
            continue
        if any(_meaningful_style(styles_rc.get(f"{r}:{c}")) for c in range(len(row))):
            out.append(row)
    return out


def _compact_text_instruction_block(block: dict[str, Any]) -> None:
    """Normalize guideline/instruction tables: drop blank columns and spacers."""
    grid = block.get("grid") or []
    if not grid:
        return
    styles_rc = block.get("styles") or {}
    merges = block.get("mergedCells") or []
    header_rows = int(block.get("headerRowCount") or 1)
    grid, styles_rc, merges = _drop_fully_blank_columns(grid, styles_rc, merges, header_rows)
    grid = _drop_blank_spacer_rows(grid, styles_rc, header_rows)
    block["grid"] = grid
    block["styles"] = styles_rc
    block["mergedCells"] = merges


def _preferred_text_instruction_col_widths(grid: list[list[str]]) -> list[int]:
    """Topic/Step 15–25%, instruction/guideline text 65–80% of body width."""
    n_cols = max((len(r) for r in grid), default=0)
    if not n_cols:
        return []
    target = int(BODY_W * 0.92)
    header_row = _find_header_row_index(grid)
    header = grid[header_row] if grid else []
    classes = [_col_header_class(str(header[c]) if c < len(header) else "") for c in range(n_cols)]

    if n_cols == 2:
        narrow, wide = int(target * 0.22), int(target * 0.70)
        c0 = classes[0]
        if c0 in ("narrative", "other") and classes[1] in ("section", "other"):
            widths = [wide, narrow]
        else:
            widths = [narrow, wide]
    else:
        shares: list[float] = []
        for cls in classes:
            if cls == "section":
                shares.append(0.20)
            elif cls == "narrative":
                shares.append(0.72)
            elif cls == "status":
                shares.append(0.08)
            elif cls == "notes":
                shares.append(0.10)
            else:
                shares.append(0.15)
        total = sum(shares) or 1.0
        widths = [max(48, int(round(target * s / total))) for s in shares]
        diff = target - sum(widths)
        if diff and widths:
            widest = max(range(n_cols), key=lambda i: shares[i])
            widths[widest] += diff

    lo = int(BODY_W * 0.85)
    if sum(widths) < lo:
        widths[-1] += lo - sum(widths)
    return widths


def _trim_trailing_blank_ranges(
    grid: list[list[str]],
    styles_rc: dict[str, Any],
    merges: list[dict[str, Any]],
    header_rows: int,
    *,
    trim_rows: bool = True,
    trim_cols: bool = True,
    print_area: str | None = None,
) -> tuple[list[list[str]], dict[str, Any], list[dict[str, Any]], int, int]:
    """Deterministically drop trailing blank worksheet columns/rows from a
    normalized excel_exact render block (FINAL RENDER POLISH 4G, Phase B).

    A trailing row/column is only dropped when, across its full extent, every
    cell has no value AND no meaningful fill/border AND is not the tail of a
    real merged cell. Never trims past an explicit print-area override, and
    never trims below the repeated header band. Returns the trimmed
    grid/styles/merges plus the pre-trim row/col counts (for diagnostics).
    """
    n_rows = len(grid)
    n_cols = max((len(r) for r in grid), default=0)
    rows_before, cols_before = n_rows, n_cols
    if n_rows == 0 or n_cols == 0:
        return grid, styles_rc, merges, rows_before, cols_before

    keep_col, keep_row = _print_area_last_col_row(print_area) or (-1, -1)

    # A merge continuation cell never carries its own text or style (the
    # anchor top-left cell holds both) — so a decorative full-width title
    # band merge does not, by itself, make its trailing columns non-blank.
    # Any merge that survives a trim simply has its endRow/endCol clamped
    # below, so shrinking a banner's span loses no content.
    def col_is_blank(c: int) -> bool:
        for r in range(n_rows):
            row = grid[r] if r < len(grid) else []
            if c < len(row) and str(row[c] or "").strip():
                return False
            if _meaningful_style(styles_rc.get(f"{r}:{c}")):
                return False
        return True

    def row_is_blank(r: int) -> bool:
        row = grid[r] if r < len(grid) else []
        if any(str(v or "").strip() for v in row):
            return False
        for c in range(max(n_cols, len(row))):
            if _meaningful_style(styles_rc.get(f"{r}:{c}")):
                return False
        return True

    if trim_cols:
        c = n_cols - 1
        while c > 0 and c > keep_col and col_is_blank(c):
            c -= 1
        n_cols = c + 1

    if trim_rows:
        floor = max(0, header_rows - 1)
        r = n_rows - 1
        while r > floor and r > keep_row and row_is_blank(r):
            r -= 1
        n_rows = r + 1

    if n_rows == rows_before and n_cols == cols_before:
        return grid, styles_rc, merges, rows_before, cols_before

    new_grid = [list(row[:n_cols]) for row in grid[:n_rows]]
    new_styles: dict[str, Any] = {}
    for key, val in styles_rc.items():
        try:
            rs, cs = key.split(":")
            r, c = int(rs), int(cs)
        except (ValueError, AttributeError):
            continue
        if r < n_rows and c < n_cols:
            new_styles[key] = val
    new_merges: list[dict[str, Any]] = []
    for m in merges:
        if m.get("startRow", 0) >= n_rows or m.get("startCol", 0) >= n_cols:
            continue
        nm = dict(m)
        nm["endRow"] = min(nm.get("endRow", 0), n_rows - 1)
        nm["endCol"] = min(nm.get("endCol", 0), n_cols - 1)
        new_merges.append(nm)

    return new_grid, new_styles, new_merges, rows_before, cols_before


def _excel_range_block(ws: dict[str, Any], block_id: str, split_settings: dict[str, Any] | None = None) -> dict[str, Any]:
    """Build a self-contained excel_exact block (grid + 0-based styles + dims)."""
    grid = [list(r) for r in (ws.get("grid") or [])]
    n_rows = len(grid)
    n_cols = max((len(r) for r in grid), default=0)
    grid = [r + [""] * (n_cols - len(r)) for r in grid]

    # A1-keyed source styles -> 0-based "r:c" keys (slice/render friendly).
    styles_rc: dict[str, Any] = {}
    for key, val in (ws.get("styles") or {}).items():
        m = re.fullmatch(r"([A-Z]+)(\d+)", key)
        if not m:
            continue
        col = column_index_from_string(m.group(1)) - 1
        row = int(m.group(2)) - 1
        if 0 <= row < n_rows and 0 <= col < n_cols:
            styles_rc[f"{row}:{col}"] = val

    col_widths = list(ws.get("colWidthsPx") or [])
    col_widths = (col_widths + [64] * n_cols)[:n_cols]
    row_heights = list(ws.get("rowHeightsPx") or [])
    row_heights = (row_heights + [20] * n_rows)[:n_rows]
    merged_cells = ws.get("mergedCells") or []

    # Header band: leading consecutive rows that are bold or filled (the yellow
    # controller / gray column headers). Repeated on continuation pages.
    header_rows = 0
    for r in range(min(n_rows, 4)):
        styled = any(
            (styles_rc.get(f"{r}:{c}", {}).get("bold") or styles_rc.get(f"{r}:{c}", {}).get("fill"))
            for c in range(n_cols)
        )
        has_text = any((grid[r][c] or "").strip() for c in range(n_cols))
        if styled and has_text:
            header_rows = r + 1
        else:
            break
    if header_rows == 0 and n_rows:
        header_rows = 1

    ss = split_settings or {}

    # Trailing-blank-range trim (FINAL RENDER POLISH 4G, Phase B). Normalized
    # output only — the Source tab keeps the untouched worksheet grid because
    # it is built from ``ws["grid"]`` directly, not from this block.
    rows_before, cols_before = n_rows, n_cols
    if ss.get("trimBlankRows", True) or ss.get("trimBlankColumns", True):
        grid, styles_rc, merged_cells, rows_before, cols_before = _trim_trailing_blank_ranges(
            grid,
            styles_rc,
            merged_cells,
            header_rows,
            trim_rows=ss.get("trimBlankRows", True),
            trim_cols=ss.get("trimBlankColumns", True),
            print_area=ws.get("printArea"),
        )
        n_rows = len(grid)
        n_cols = max((len(r) for r in grid), default=0)
        col_widths = col_widths[:n_cols]
        row_heights = row_heights[:n_rows]

    return {
        "id": block_id,
        "type": "excelRange",
        "sourceWorksheetId": ws["id"],
        "sourceSheet": ws.get("sourceSheet") or ws.get("name", ""),
        "sourceRange": ws.get("sourceRange", ""),
        "printArea": ws.get("printArea"),
        "renderMode": "excel_exact",
        "grid": grid,
        "styles": styles_rc,
        "mergedCells": merged_cells,
        "colWidths": col_widths,
        "rowHeights": row_heights,
        "srcRows": list(range(n_rows)),
        "headerRowCount": header_rows,
        "repeatRows": list(range(header_rows)),
        "splitMode": ss.get("splitMode", "auto_rows"),
        "minScale": ss.get("minScale", EXCEL_MIN_SCALE),
        "allowContinuation": ss.get("allowContinuation", True),
        "scaleMode": ss.get("scaleMode", "fit_body"),
        "orientation": "landscape",
        "styleRole": "excel-exact",
        "bodyRowFillMode": "none",
        "gridLines": True,
        "editable": True,
        "rowsBeforeTrim": rows_before,
        "colsBeforeTrim": cols_before,
        "rowsAfterTrim": n_rows,
        "colsAfterTrim": n_cols,
    }


def _company_info_block(ws: dict[str, Any], block_id: str) -> dict[str, Any]:
    rows = []
    for row in ws.get("grid") or []:
        vals = [str(c).strip() for c in row if str(c).strip()]
        if vals:
            rows.append(vals)
    return {
        "id": block_id,
        "type": "companyInfo",
        "sourceWorksheetId": ws["id"],
        "sourceSheet": ws.get("sourceSheet") or ws.get("name", ""),
        "rows": rows,
        "text": "Singh360 Company Info",
        "styleRole": "company-info",
        "editable": True,
    }


def import_workbook(
    path: str | Path,
    project_id: str | None = None,
    assets_dir=None,
    asset_url_prefix: str | None = None,
) -> dict[str, Any]:
    xlsx = Path(path)
    # Macro-enabled workbooks are parsed without altering the original file.
    # The untouched .xlsm is copied into the project's source-workbook folder.
    keep_vba = xlsx.suffix.lower() == ".xlsm"
    wb = load_workbook(filename=xlsx, data_only=False, keep_vba=keep_vba)
    try:
        wb_data = load_workbook(filename=xlsx, data_only=True, keep_vba=keep_vba)
    except Exception:
        wb_data = None

    project = default_project(project_id)
    project["metadata"]["sourceFile"] = xlsx.name

    project["sources"].append(
        {
            "id": f"src_{project['id']}_xlsx",
            "type": "workbook",
            "name": xlsx.name,
            "path": str(xlsx),
        }
    )

    sheet_payloads = [
        _worksheet_payload(
            wb[sheet_name],
            wb_data[sheet_name] if wb_data is not None and sheet_name in wb_data.sheetnames else None,
        )
        for sheet_name in wb.sheetnames
    ]
    embedded_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        embedded_by_sheet[sheet_name] = _extract_embedded_images(
            wb[sheet_name], assets_dir, asset_url_prefix or "", sheet_name
        )

    for i, ws in enumerate(sheet_payloads):
        project["worksheets"].append(
            {
                "id": f"ws_{i+1}",
                "name": ws["name"],
                "sourceId": f"src_{project['id']}_xlsx",
                "visible": True,
                "classHint": "unknown",
                "grid": ws["grid"],
                "formulas": ws["formulas"],
                "styles": ws["styles"],
                "mergedCells": ws["mergedCells"],
                "rowHeights": ws["rowHeights"],
                "columnWidths": ws["columnWidths"],
                "colWidthsPx": ws["colWidthsPx"],
                "rowHeightsPx": ws["rowHeightsPx"],
                "sourceSheet": ws["sourceSheet"],
                "sourceRange": ws["sourceRange"],
                "printArea": ws["printArea"],
                "embeddedImages": embedded_by_sheet.get(ws["name"], []),
                "provenance": {"sheet": ws["name"]},
            }
        )

    index_sheet_name = _find_index_sheet(wb)
    index_entries = _parse_index(wb, index_sheet_name)
    has_index = bool(index_sheet_name and index_entries)
    index_lookup = {e["sheetTab"].lower(): e for e in index_entries if e.get("sheetTab")}

    # Kyle/SA38 workbook import, Phase A rule 7: 00_INDEX, 00_APP_INDEX, and
    # 00_PROJECT_META are metadata/control sheets, never drawing/output pages —
    # regardless of any Include flag a hand-edited index might carry for them.
    metadata_sheet_name = _find_metadata_sheet(wb)
    never_output_sheets = _find_alias_index_sheets(wb, index_sheet_name)
    if metadata_sheet_name:
        never_output_sheets.add(metadata_sheet_name)

    # Title-block metadata precedence (Phase B): project properties (already
    # set by the caller before import — none at a fresh import) > workbook
    # metadata sheet > cover-page key/value inference. Never overwrites a
    # value that is already populated, and never invents a value that isn't
    # literally present in the workbook.
    meta_ws = next((w for w in project["worksheets"] if w["name"] == metadata_sheet_name), None) if metadata_sheet_name else None
    cover_ws = None
    for w in project["worksheets"]:
        if w["name"] in never_output_sheets:
            continue
        w_idx = index_lookup.get(w["name"].lower())
        w_title = w_idx["sheetTitle"] if w_idx else w["name"]
        w_use_source = w_idx["useSource"] if w_idx else ""
        if classify_page_type(w["name"], w_title, w_use_source) == "cover":
            cover_ws = w
            break

    inferred_metadata = {
        **_infer_metadata_from_labeled_grid(cover_ws),
        **_infer_metadata_from_labeled_grid(meta_ws),
    }
    for field, value in inferred_metadata.items():
        if value and not project["metadata"].get(field):
            project["metadata"][field] = value

    pages = []
    order_cursor = 1
    for i, ws in enumerate(project["worksheets"]):
        if ws["name"] in never_output_sheets:
            continue

        idx = index_lookup.get(ws["name"].lower())
        title = idx["sheetTitle"] if idx else ws["name"]
        is_index_tab = has_index and ws["name"] == index_sheet_name
        if idx:
            include = idx["include"]
        elif is_index_tab:
            # Phase A fix: a Sheet Index / TOC almost never lists itself as a
            # row inside its own table, so falling back to "not has_index"
            # here (the rule for every other un-indexed sheet) silently
            # dropped the index page from every export whenever the index
            # didn't self-reference. The index sheet defaults to included
            # unless it has an explicit self-row saying otherwise (handled
            # by the `if idx:` branch above).
            include = True
            if not idx:
                title = "Sheet Index / TOC"
            project.setdefault("importWarnings", []).append(
                {
                    "sheetTab": ws["name"],
                    "issue": "Sheet Index / TOC has no row for itself inside 00_INDEX; "
                    "defaulted to included so it is not silently dropped from export.",
                }
            )
        else:
            include = not has_index
        use_source = idx["useSource"] if idx else ""

        # Index include/exclude is law: no output page/tab/PDF for excluded sheets.
        if not include:
            continue

        page_type = classify_page_type(ws["name"], title, use_source)
        family = page_family(ws["name"], title, use_source)
        if _looks_company_info(ws["name"], title, use_source):
            family = "companyInfo"

        use_exact = _should_use_excel_exact(page_type, family, ws)
        split_settings = _split_settings_for_page(family, page_type, use_exact)
        is_idf_network, idf_header_row = _is_idf_network_table(ws, family)
        layout_profile = _layout_profile_for(family, page_type, f"{ws['name']} {title} {use_source}".lower())

        # Cover keeps its own look; every other page gets the Singh360 profile.
        header_style = "source" if page_type == "cover" else DEFAULT_HEADER_STYLE

        if family == "companyInfo":
            exact_block = None
            blocks = [_company_info_block(ws, f"{ws['id']}_company")]
        elif is_idf_network and idf_header_row is not None:
            split_settings = {"splitMode": "none", "allowContinuation": False, "minScale": 1.0, "scaleMode": "fit_body"}
            exact_block = _build_idf_network_block(ws, idf_header_row, f"{ws['id']}_idf")
            use_exact = True
            blocks = [exact_block]
        elif use_exact:
            render_ws = _filter_index_payload_for_output(ws) if page_type == "index" else ws
            exact_block = _excel_range_block(render_ws, f"{ws['id']}_xr", split_settings)
            if family == "text" and layout_profile in ("front_matter_table", "instruction_table"):
                _compact_text_instruction_block(exact_block)
            if "lighting" in f"{ws['name']} {title}".lower():
                exact_block["minScale"] = max(float(exact_block.get("minScale") or EXCEL_MIN_SCALE), 0.58)
            _apply_table_geometry(exact_block, family, page_type, layout_profile)
            apply_singh360_profile(exact_block, header_style)
            if layout_profile == "front_matter_narrative_table":
                # Keep narrative profile markers after the standard recolor pass.
                exact_block["renderProfile"] = NARRATIVE_RENDER_PROFILE
                exact_block["layoutProfile"] = "front_matter_narrative_table"
            if layout_profile == "instruction_table":
                # Phase D: compact 9pt (~12px) body font for Step/Instruction
                # pages — overrides the per-cell Excel font size captured on
                # import so long instruction rows never force clipping/scaling.
                exact_block["bodyFontPx"] = 12
                # PHASE B fix (was "noGrow=True"): the old fix for this page
                # kept the table pinned to its natural (narrow) size to avoid
                # growing height past the safe render area — but that is
                # exactly what produced the "tiny unreadable strip" bug,
                # since the natural size was itself too narrow. Now that
                # _preferred_col_widths lets this profile's columns reach the
                # full body width (so height no longer inflates from
                # excessive word-wrap) and _split_settings_for_page allows a
                # genuine overflow to fall back to a continuation page
                # (EMS 17.0 -> EMS 17.0a) instead of over-shrinking, this
                # page can safely grow to fill the body width like any other
                # table.
            blocks = [exact_block]
        else:
            exact_block = None
            blocks = normalize_page(ws, ws["id"], page_type, title)

            # Embedded workbook images → real image blocks (rendered, not placeholders).
            for j, emb in enumerate(ws.get("embeddedImages", []) or []):
                blocks.append(
                    {
                        "id": f"{ws['id']}_emb_{j}",
                        "type": "imagePlaceholder",
                        "sourceWorksheetId": ws["id"],
                        "sourceRange": emb.get("anchorCell", ""),
                        "filename": emb.get("name", ""),
                        "url": emb.get("url", ""),
                        "text": emb.get("name", ""),
                        "styleRole": "note",
                        "editable": False,
                    }
                )

            # Best-effort reference-screenshot match for a blank drawing/
            # canvas page with no embedded workbook image (Phase C). A no-op
            # (returns None) when the project's assets/screenshots scaffold
            # folder is empty or missing — the normal case on a fresh import.
            if page_type == "canvas" and not ws.get("embeddedImages"):
                preliminary_code = (idx.get("sheetCodeRaw") if idx else "") or _sheet_code_from_tab(ws["name"])
                matched = _match_page_asset(project["id"], preliminary_code, title)
                if matched:
                    blocks.append(
                        {
                            "id": f"{ws['id']}_asset_match",
                            "type": "imagePlaceholder",
                            "sourceWorksheetId": ws["id"],
                            "sourceRange": "",
                            "filename": matched["name"],
                            "url": matched["url"],
                            "text": matched["name"],
                            "styleRole": "note",
                            "editable": False,
                        }
                    )

        # Sheet number precedence (FINAL SA31 POLISH 4I / 4G Phase A): the
        # workbook/index "Sheet Code" column always wins — never bare Order
        # values like "5.0". Fall back to a code embedded in the tab name,
        # then ``EMS {order}`` so SHEET NO. never looks like physical order.
        sheet_code = _canonical_sheet_code(idx, ws["name"], order_cursor)
        if is_index_tab and not idx:
            sheet_code = f"EMS {order_cursor}.0" if order_cursor > 1 else "EMS 2.0"

        has_image_block = any(b.get("type") in ("imagePlaceholder", "underlayPlaceholder") for b in blocks)
        blank_page_placeholder = (
            _blank_page_placeholder_message(ws["name"], title) if page_type == "canvas" and not has_image_block else ""
        )

        page_render_profile = (
            NARRATIVE_RENDER_PROFILE
            if layout_profile == "front_matter_narrative_table"
            else RENDER_PROFILE
        )

        page = {
            "id": f"page_{i+1}",
            "order": order_cursor,
            "include": include,
            "sheetCode": sheet_code,
            "displaySheetCode": sheet_code,
            "sheetTitle": title,
            "sheetTab": ws["name"],
            "pageType": page_type,
            "pageFamily": family,
            "blankPagePlaceholder": blank_page_placeholder,
            "layoutProfile": layout_profile,
            "twoUp": bool(exact_block and exact_block.get("layoutMode") == "two_up"),
            "renderMode": "excel_exact" if use_exact else "normalized",
            "renderProfile": page_render_profile,
            "normalizedHeaderStyle": header_style,
            "sourceSheet": ws.get("sourceSheet") or ws["name"],
            "sourceRange": ws.get("sourceRange", "") if use_exact else "",
            "printArea": ws.get("printArea") if use_exact else None,
            "splitMode": split_settings.get("splitMode", "none"),
            "repeatRows": (exact_block.get("repeatRows", []) if exact_block else []),
            "minScale": split_settings.get("minScale", EXCEL_MIN_SCALE),
            "allowContinuation": split_settings.get("allowContinuation", False),
            "scaleMode": split_settings.get("scaleMode", "fit_width"),
            "trimBlankRows": split_settings.get("trimBlankRows", True),
            "trimBlankColumns": split_settings.get("trimBlankColumns", True),
            "orientation": "landscape",
            "templateId": "ansi-b-standard",
            "linkedWorksheetId": ws["id"],
            "blocks": blocks,
            "canvasObjects": [],
            "assets": [],
            "underlays": [],
            "notes": idx["notes"] if idx else "",
            "revisionRows": [],
            "pageGroupId": f"page_{i+1}",
            "continuationOf": None,
            "continuationIndex": 0,
            "generatedContinuation": False,
            "layoutWarnings": [],
            "showTerminatedBy": False,
        }

        # Phase E rule 4 fallback: a network table so dense that two-up would
        # still fall below the readable floor gets split here (at import time,
        # since idfNetworkTable blocks opt out of compose_pages's generic
        # continuation splitter) into two balanced full-width single pages.
        if exact_block is not None and exact_block.get("needsHardSplit") and exact_block.get("hardSplitBoundary"):
            boundary = exact_block["hardSplitBoundary"]
            total_n = exact_block.get("totalRowCount", 0)
            left_block = _build_idf_network_block(
                ws, idf_header_row, f"{ws['id']}_idf_a", row_slice=(0, boundary)
            )
            right_block = _build_idf_network_block(
                ws, idf_header_row, f"{ws['id']}_idf_b", row_slice=(boundary, total_n)
            )
            page["blocks"] = [left_block]
            page["twoUp"] = False
            page["layoutWarnings"] = [
                f"Network table has {total_n} ports; split into balanced pages "
                f"1–{boundary} and {boundary + 1}–{total_n} instead of an "
                "illegibly dense two-up layout."
            ]
            pages.append(page)
            order_cursor += 1

            cont_page = deepcopy(page)
            cont_page["id"] = f"page_{i+1}_b"
            cont_page["order"] = order_cursor
            cont_page["sheetCode"] = continuation_code(sheet_code, 1)
            cont_page["displaySheetCode"] = cont_page["sheetCode"]
            if "continued" not in title.lower():
                cont_page["sheetTitle"] = f"{title} — CONTINUED"
            cont_page["blocks"] = [right_block]
            cont_page["pageGroupId"] = page["pageGroupId"]
            cont_page["continuationOf"] = page["id"]
            cont_page["continuationIndex"] = 1
            cont_page["generatedContinuation"] = True
            pages.append(cont_page)
            order_cursor += 1
            continue

        pages.append(page)
        order_cursor += 1

    pages = sorted(pages, key=lambda p: p["order"])
    project["pages"] = compose_pages(pages)
    _append_continuation_rows_to_index(project)
    project["paginationLocked"] = True
    recalc_page_numbers(project)
    try:
        log_render_diagnostics(project["pages"])
    except Exception:
        pass
    return sanitize_json(project)
