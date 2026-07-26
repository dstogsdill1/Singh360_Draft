"""Template-driven schema-V2 project services.

Legacy projects never enter these code paths unless they explicitly declare
``schemaVersion: 2``.
"""
from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import shutil
import tempfile
import uuid
import zipfile
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.utils import secure_filename

PROJECT_ID_RE = re.compile(r"^[a-f0-9]{16}$")
SOURCE_ID_RE = re.compile(r"^[a-f0-9]{16}$")
SAFE_SOURCE_EXTENSIONS = {
    ".pdf": "pdf", ".png": "images", ".jpg": "images", ".jpeg": "images",
    ".webp": "images", ".svg": "images", ".xlsx": "spreadsheets",
    ".xlsm": "spreadsheets", ".csv": "csv", ".txt": "documents",
    ".doc": "documents", ".docx": "documents", ".rtf": "documents",
    ".odt": "documents", ".ods": "spreadsheets",
}
MAX_SOURCE_BYTES = 100 * 1024 * 1024
WORKBOOK_HISTORY_LIMIT = 20
SOURCE_FOLDERS = [
    "Drawings/CD Set", "Drawings/Bid Set", "Drawings/Construction Set",
    "Drawings/Bulletins", "Drawings/As-Builts",
    "Converted Schedules/Original Excel", "Converted Schedules/Cleaned Data",
    "Survey Photos", "Reference Documents", "Manufacturer Data", "Drawing Assets",
    "Manual Layouts", "Programming", "Commissioning", "Archive",
]
REQUIRED_BASE_SHEETS = {
    "00_PROJECT_META", "00_TEMPLATE_PROFILE", "00_INDEX", "00_STYLE_GUIDE",
    "01_SOURCE_LIBRARY", "02_CONVERSION_QUEUE", "03_SCOPE_AND_PLAN",
    "04_PROJECT_DIRECTORY", "05_RESPONSIBILITY_MATRIX",
    "06_WORKFLOW_MILESTONES", "07_OPEN_ITEMS",
}


class TemplatePlatformError(ValueError):
    pass


class RevisionConflict(TemplatePlatformError):
    pass


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def atomic_json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class ProfileRegistry:
    def __init__(self, path: Path):
        self.path = Path(path)
        raw = json.loads(self.path.read_text("utf-8"))
        if raw.get("schemaVersion") != 1 or not isinstance(raw.get("profiles"), list):
            raise TemplatePlatformError("Project profile registry schemaVersion must be 1.")
        self._raw = {profile["id"]: profile for profile in raw["profiles"]}
        if len(self._raw) != len(raw["profiles"]):
            raise TemplatePlatformError("Project profile IDs must be unique.")
        for profile in raw["profiles"]:
            self._validate(profile)
        self._resolved = {key: self._resolve(key, set()) for key in self._raw}

    def _validate(self, profile: dict[str, Any]) -> None:
        required = {
            "id", "displayName", "description", "extends", "styleProfile",
            "sourceSlots", "dataSheets", "pageRecipes", "defaultIncludedFamilies",
            "optionalFamilies", "validationRules", "version",
        }
        missing = sorted(required - set(profile))
        if missing:
            raise TemplatePlatformError(f"Profile {profile.get('id', '?')} is missing: {', '.join(missing)}")
        if not re.fullmatch(r"[A-Z][A-Z0-9_]+", str(profile["id"])):
            raise TemplatePlatformError(f"Invalid profile ID: {profile['id']}")
        parent = profile.get("extends")
        if parent and parent not in self._raw:
            raise TemplatePlatformError(f"Profile {profile['id']} extends unknown profile {parent}.")

    def _resolve(self, profile_id: str, visiting: set[str]) -> dict[str, Any]:
        if profile_id in visiting:
            raise TemplatePlatformError(f"Profile inheritance cycle at {profile_id}.")
        profile = copy.deepcopy(self._raw[profile_id])
        parent_id = profile.get("extends")
        if not parent_id:
            return profile
        parent = self._resolve(parent_id, visiting | {profile_id})
        for key in ("sourceSlots", "dataSheets", "pageRecipes", "defaultIncludedFamilies", "optionalFamilies"):
            profile[key] = list(dict.fromkeys(parent.get(key, []) + profile.get(key, [])))
        profile["validationRules"] = {**parent.get("validationRules", {}), **profile.get("validationRules", {})}
        return profile

    def list(self) -> list[dict[str, Any]]:
        return [copy.deepcopy(self._resolved[key]) for key in self._raw]

    def get(self, profile_id: str) -> dict[str, Any]:
        if profile_id not in self._resolved:
            raise TemplatePlatformError(f"Unknown project profile: {profile_id}")
        return copy.deepcopy(self._resolved[profile_id])


class TemplateRegistry:
    def __init__(self, docs_dir: Path):
        self.root = Path(docs_dir) / "library" / "workbook_templates"
        self.base_dir = self.root / "base"
        self.manifest_path = self.root / "manifest.json"

    def list(self) -> list[dict[str, Any]]:
        if not self.manifest_path.is_file():
            return []
        payload = json.loads(self.manifest_path.read_text("utf-8"))
        return payload.get("templates", [])

    def get(self, template_id: str, active_only: bool = True) -> dict[str, Any]:
        for record in self.list():
            if record.get("templateId") == template_id and (not active_only or record.get("active")):
                path = Path(record.get("absoluteRuntimePath", ""))
                if not path.is_file() or sha256_file(path) != record.get("sha256"):
                    raise TemplatePlatformError(f"Runtime template {template_id} is missing or its checksum changed.")
                return record
        raise TemplatePlatformError(f"No active runtime workbook template is registered as {template_id}.")

    def validate(self, path: Path) -> dict[str, Any]:
        path = Path(path)
        result: dict[str, Any] = {"valid": False, "path": str(path.resolve()), "errors": [], "warnings": []}
        if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.is_file():
            result["errors"].append("Template must be an existing XLSX or XLSM workbook.")
            return result
        try:
            wb = load_workbook(path, read_only=False, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
            names = wb.sheetnames
            missing = sorted(REQUIRED_BASE_SHEETS - set(names))
            if missing:
                result["errors"].append(f"Missing required worksheets: {', '.join(missing)}")
            result.update({"sheetNames": names, "sheetCount": len(names), "sha256": sha256_file(path)})
            result["valid"] = not result["errors"]
            wb.close()
        except Exception as exc:
            result["errors"].append(f"Workbook could not be opened: {exc}")
        return result

    def register(self, staged_path: Path, supported_profiles: list[str]) -> dict[str, Any]:
        validation = self.validate(staged_path)
        if not validation["valid"]:
            raise TemplatePlatformError("; ".join(validation["errors"]))
        self.base_dir.mkdir(parents=True, exist_ok=True)
        destination = self.base_dir / Path(staged_path).name
        if Path(staged_path).resolve() != destination.resolve():
            shutil.copy2(staged_path, destination)
        digest = sha256_file(destination)
        record = {
            "templateId": "SINGH360_BASE_V1",
            "displayName": "Singh360 Base Project Workbook Template V1",
            "version": "1.0.0",
            "fileName": destination.name,
            "absoluteRuntimePath": str(destination.resolve()),
            "sha256": digest,
            "supportedProfiles": supported_profiles,
            "active": True,
            "registeredUtc": utcnow(),
            "sourceStagingPath": str(Path(staged_path).resolve()),
            "workbookValidation": validation,
        }
        records = [r for r in self.list() if r.get("templateId") != record["templateId"]]
        records.append(record)
        atomic_json_write(self.manifest_path, {"schemaVersion": 1, "templates": records})
        return record


def _cell_style(cell: Any) -> dict[str, Any]:
    style: dict[str, Any] = {}
    if cell.font:
        style.update({
            "bold": bool(cell.font.bold), "italic": bool(cell.font.italic),
            "underline": bool(cell.font.underline),
            "fontSize": cell.font.sz, "fontName": cell.font.name,
            "fontColor": cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
        })
    if cell.fill and cell.fill.fill_type:
        style["fill"] = cell.fill.fgColor.rgb
    if cell.alignment:
        style.update({"hAlign": cell.alignment.horizontal, "vAlign": cell.alignment.vertical, "wrap": bool(cell.alignment.wrap_text)})
    if cell.number_format and cell.number_format != "General":
        style["numberFormat"] = cell.number_format
    return {key: value for key, value in style.items() if value not in (None, False, "")}


def workbook_to_document(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        cells: dict[str, Any] = {}
        styles: dict[str, Any] = {}
        max_row = min(ws.max_row, 5000)
        max_col = min(ws.max_column, 200)
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    value = cell.value.isoformat() if isinstance(cell.value, (date, datetime, time)) else cell.value
                    cells[cell.coordinate] = {"v": value}
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cells[cell.coordinate] = {"f": cell.value}
                style = _cell_style(cell)
                if style:
                    styles[cell.coordinate] = style
        sheets.append({
            "id": uuid.uuid5(uuid.NAMESPACE_URL, f"singh360:{ws.title}").hex[:16],
            "name": ws.title, "cells": cells, "styles": styles,
            "tabColor": (
                f"#{ws.sheet_properties.tabColor.rgb[-6:]}"
                if ws.sheet_properties.tabColor and ws.sheet_properties.tabColor.type == "rgb"
                else None
            ),
            "merges": [str(item) for item in ws.merged_cells.ranges],
            "rowHeights": {str(i): dim.height for i, dim in ws.row_dimensions.items() if dim.height},
            "columnWidths": {key: dim.width for key, dim in ws.column_dimensions.items() if dim.width},
            "archived": ws.sheet_state != "visible",
        })
    wb.close()
    return {"revision": 1, "updatedAt": utcnow(), "sheets": sheets}


class WorkbookDocumentStore:
    def __init__(self, project_dir: Path):
        self.path = Path(project_dir) / "data" / "workbook.json"
        self.history = Path(project_dir) / "backups" / "workbook"

    def load(self) -> dict[str, Any]:
        if not self.path.is_file():
            raise TemplatePlatformError("This project does not have a Data Workspace document.")
        return json.loads(self.path.read_text("utf-8"))

    def create(self, document: dict[str, Any]) -> None:
        if self.path.exists():
            raise TemplatePlatformError("Workbook document already exists.")
        atomic_json_write(self.path, document)

    def save(self, expected_revision: int, document: dict[str, Any]) -> dict[str, Any]:
        current = self.load()
        if current.get("revision") != expected_revision:
            raise RevisionConflict(f"Workbook revision conflict: expected {expected_revision}, current {current.get('revision')}.")
        self.history.mkdir(parents=True, exist_ok=True)
        backup = self.history / f"workbook_r{current['revision']}_{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S-%f')}.json"
        shutil.copy2(self.path, backup)
        for old in sorted(self.history.glob("workbook_*.json"))[:-WORKBOOK_HISTORY_LIMIT]:
            old.unlink(missing_ok=True)
        saved = copy.deepcopy(document)
        saved["revision"] = expected_revision + 1
        saved["updatedAt"] = utcnow()
        atomic_json_write(self.path, saved)
        return saved


class SourceLibrary:
    def __init__(self, project_dir: Path):
        self.project_dir = Path(project_dir)
        self.manifest_path = self.project_dir / "source_library.json"

    def load(self) -> dict[str, Any]:
        if not self.manifest_path.is_file():
            return {
                "schemaVersion": 2, "folders": SOURCE_FOLDERS,
                "sources": [], "conversionQueue": [], "importReports": [],
            }
        payload = json.loads(self.manifest_path.read_text("utf-8"))
        payload["schemaVersion"] = 2
        payload.setdefault("folders", SOURCE_FOLDERS)
        payload.setdefault("importReports", [])
        return payload

    @staticmethod
    def safe_virtual_path(value: str) -> str:
        raw = str(value or "").replace("\\", "/").strip("/")
        if not raw:
            return ""
        parts = [part.strip() for part in raw.split("/") if part.strip()]
        if any(
            part in {".", ".."}
            or not re.fullmatch(r'[^<>:"|?*\x00-\x1f]+', part)
            or part.endswith((".", " "))
            for part in parts
        ):
            raise TemplatePlatformError("Source virtual path contains traversal or an absolute path.")
        return "/".join(parts)

    def upload(self, stream: BinaryIO, original_name: str, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
        if Path(str(original_name).replace("\\", "/")).name != str(original_name).replace("\\", "/").split("/")[-1]:
            raise TemplatePlatformError("Source filename is unsafe.")
        safe_name = secure_filename(Path(original_name).name)
        suffix = Path(safe_name).suffix.lower()
        kind = SAFE_SOURCE_EXTENSIONS.get(suffix)
        if not safe_name or not kind:
            raise TemplatePlatformError(f"Unsupported or unsafe source file: {original_name}")
        source_id = uuid.uuid4().hex[:16]
        virtual_path = self.safe_virtual_path((metadata or {}).get("virtualPath", ""))
        destination_dir = self.project_dir / "sources" / "library" / virtual_path
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / f"{source_id}__{safe_name}"
        digest = hashlib.sha256()
        size = 0
        with destination.open("xb") as handle:
            while True:
                chunk = stream.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_SOURCE_BYTES:
                    handle.close()
                    destination.unlink(missing_ok=True)
                    raise TemplatePlatformError("Source file exceeds the 100 MB limit.")
                digest.update(chunk)
                handle.write(chunk)
        payload = self.load()
        existing_versions = [
            s for s in payload["sources"]
            if s.get("originalFileName") == original_name and s.get("virtualPath", "") == virtual_path
        ]
        record = {
            "id": source_id, "originalFileName": original_name, "storedFileName": destination.name,
            "mediaType": suffix.lstrip("."), "sourceType": kind,
            "discipline": (metadata or {}).get("discipline", ""), "size": size,
            "sha256": digest.hexdigest(), "dateAdded": utcnow(),
            "addedBy": (metadata or {}).get("addedBy", ""),
            "version": len(existing_versions) + 1, "status": "active",
            "virtualPath": virtual_path,
            "relativePath": f"{virtual_path}/{safe_name}".strip("/"),
            "localProjectPath": str(destination.relative_to(self.project_dir)).replace("\\", "/"),
            "originalLocation": (metadata or {}).get("originalLocation", ""),
            "backupLocation": (metadata or {}).get("backupLocation", ""),
            "tags": (metadata or {}).get("tags", []), "notes": (metadata or {}).get("notes", ""),
            "derivedFiles": [], "supersedes": (metadata or {}).get("supersedes"),
            "supersededBy": None,
        }
        if record["supersedes"]:
            for old in payload["sources"]:
                if old["id"] == record["supersedes"]:
                    old["supersededBy"] = source_id
                    old["status"] = "superseded"
        payload["sources"].append(record)
        if virtual_path and virtual_path not in payload["folders"]:
            payload["folders"].append(virtual_path)
        atomic_json_write(self.manifest_path, payload)
        return record

    def create_folder(self, virtual_path: str) -> str:
        safe = self.safe_virtual_path(virtual_path)
        if not safe:
            raise TemplatePlatformError("Folder path is required.")
        payload = self.load()
        if safe not in payload["folders"]:
            payload["folders"].append(safe)
            payload["folders"].sort()
            atomic_json_write(self.manifest_path, payload)
        return safe

    def import_zip(self, archive: BinaryIO, original_name: str, virtual_path: str = "") -> dict[str, Any]:
        root = self.safe_virtual_path(virtual_path)
        debug_dir = self.project_dir / "debug"
        debug_dir.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(suffix=".zip", dir=debug_dir)
        os.close(fd)
        temp = Path(temp_name)
        imported: list[dict[str, Any]] = []
        rejected: list[dict[str, str]] = []
        try:
            with temp.open("wb") as handle:
                shutil.copyfileobj(archive, handle)
            with zipfile.ZipFile(temp) as bundle:
                for member in bundle.infolist():
                    if member.is_dir():
                        continue
                    try:
                        member_path = self.safe_virtual_path(member.filename)
                        folder, name = str(Path(member_path).parent).replace("\\", "/"), Path(member_path).name
                        if folder == ".":
                            folder = ""
                        target_folder = "/".join(part for part in (root, folder) if part)
                        with bundle.open(member) as stream:
                            imported.append(self.upload(stream, name, {
                                "virtualPath": target_folder,
                                "originalLocation": f"{original_name}:{member.filename}",
                            }))
                    except Exception as exc:
                        rejected.append({"path": member.filename, "reason": str(exc)})
            payload = self.load()
            report = {
                "id": uuid.uuid4().hex[:16], "archive": original_name,
                "imported": len(imported), "rejected": rejected, "createdAt": utcnow(),
            }
            payload["importReports"].append(report)
            atomic_json_write(self.manifest_path, payload)
            return {"report": report, "sources": imported}
        finally:
            temp.unlink(missing_ok=True)

    def archive(self, source_id: str) -> dict[str, Any]:
        if not SOURCE_ID_RE.fullmatch(source_id):
            raise TemplatePlatformError("Invalid source ID.")
        payload = self.load()
        for source in payload["sources"]:
            if source["id"] == source_id:
                source["status"] = "archived"
                source["archivedAt"] = utcnow()
                atomic_json_write(self.manifest_path, payload)
                return source
        raise TemplatePlatformError("Source not found.")

    def restore(self, source_id: str) -> dict[str, Any]:
        payload = self.load()
        for source in payload["sources"]:
            if source["id"] == source_id:
                source["status"] = "active"
                source.pop("archivedAt", None)
                atomic_json_write(self.manifest_path, payload)
                return source
        raise TemplatePlatformError("Source not found.")

    def preview(self, source_id: str) -> dict[str, Any]:
        source, path = self.resolve(source_id)
        result: dict[str, Any] = {
            "source": source, "kind": source["sourceType"], "previewError": "",
        }
        try:
            suffix = path.suffix.lower()
            if suffix in {".xlsx", ".xlsm"}:
                wb = load_workbook(path, read_only=True, data_only=False)
                result["sheets"] = wb.sheetnames
                ws = wb[wb.sheetnames[0]]
                result["grid"] = [
                    ["" if value is None else str(value) for value in row]
                    for row in ws.iter_rows(max_row=50, max_col=20, values_only=True)
                ]
                wb.close()
            elif suffix in {".csv", ".txt", ".rtf"}:
                text = path.read_text("utf-8", errors="replace")[:100_000]
                result["text"] = text
                if suffix == ".csv":
                    result["grid"] = [line.split(",")[:30] for line in text.splitlines()[:100]]
            elif suffix == ".pdf":
                import fitz
                pdf = fitz.open(path)
                result["pageCount"] = pdf.page_count
                result["pageSizes"] = [
                    {"width": page.rect.width, "height": page.rect.height}
                    for page in list(pdf)[:25]
                ]
                pdf.close()
            elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
                from PIL import Image
                with Image.open(path) as image:
                    result["dimensions"] = {"width": image.width, "height": image.height}
            else:
                result["fallback"] = "metadata"
        except Exception as exc:
            result["previewError"] = f"{type(exc).__name__}: {exc}"
        return result

    def resolve(self, source_id: str) -> tuple[dict[str, Any], Path]:
        if not SOURCE_ID_RE.fullmatch(source_id):
            raise TemplatePlatformError("Invalid source ID.")
        for source in self.load()["sources"]:
            if source["id"] == source_id:
                path = (self.project_dir / source["localProjectPath"]).resolve()
                if self.project_dir.resolve() not in path.parents or not path.is_file():
                    raise TemplatePlatformError("Source path is invalid or missing.")
                return source, path
        raise TemplatePlatformError("Source not found.")


def apply_standard_sheet_style(ws: Any) -> None:
    orange = PatternFill("solid", fgColor="F47C20")
    charcoal = PatternFill("solid", fgColor="262626")
    blue = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="777777")
    for cell in ws[1]:
        cell.fill = orange
        cell.font = Font(bold=True, color="FFFFFF", size=14)
    if ws.max_row >= 2:
        for cell in ws[2]:
            cell.fill = charcoal
            cell.font = Font(bold=True, color="FFFFFF")
    if ws.max_row >= 3:
        for cell in ws[3]:
            cell.fill = blue
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def apply_document_cell_style(cell: Any, style: dict[str, Any]) -> None:
    fill = style.get("fill") or (style.get("bg") or {}).get("rgb")
    font_color = style.get("fontColor") or (style.get("cl") or {}).get("rgb")
    font_name = style.get("fontName") or style.get("ff")
    font_size = style.get("fontSize") or style.get("fs")
    bold = style.get("bold", style.get("bl"))
    italic = style.get("italic", style.get("it"))
    underline = style.get("underline") or style.get("ul")
    if fill:
        cell.fill = PatternFill("solid", fgColor=str(fill).lstrip("#"))
    cell.font = Font(
        name=font_name or cell.font.name, size=font_size or cell.font.sz,
        bold=bool(bold), italic=bool(italic),
        underline="single" if underline else None,
        color=str(font_color).lstrip("#") if font_color else cell.font.color,
    )
    cell.alignment = Alignment(
        horizontal=style.get("hAlign") or style.get("ht") or cell.alignment.horizontal,
        vertical=style.get("vAlign") or style.get("vt") or cell.alignment.vertical,
        wrap_text=bool(style.get("wrap", style.get("tb", cell.alignment.wrap_text))),
    )
    number_format = style.get("numberFormat")
    if number_format:
        cell.number_format = str(number_format)


def write_document_to_workbook(document: dict[str, Any], workbook_path: Path, project: dict[str, Any]) -> dict[str, Any]:
    backup_dir = workbook_path.parent.parent.parent / "backups" / "workbook_mirror"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{workbook_path.stem}_{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S-%f')}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup)
    wb = load_workbook(workbook_path, keep_vba=workbook_path.suffix.lower() == ".xlsm")
    desired_order: list[str] = []
    for sheet in document.get("sheets", []):
        name = str(sheet.get("name", ""))[:31]
        if not name:
            continue
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        desired_order.append(name)
        # Imported workbooks can contain overlapping merge metadata. Calling
        # unmerge_cells range-by-range can then delete the same placeholder
        # twice inside openpyxl. Reset the merge registry and placeholders as
        # one operation before applying the document's canonical merge list.
        ws.merged_cells.ranges.clear()
        for key, cell in list(ws._cells.items()):
            if isinstance(cell, MergedCell):
                del ws._cells[key]
        for coord, payload in sheet.get("cells", {}).items():
            value = payload.get("f") or payload.get("v") if isinstance(payload, dict) else payload
            ws[coord] = value
            style = sheet.get("styles", {}).get(coord)
            if isinstance(style, dict):
                apply_document_cell_style(ws[coord], style)
        for merged in sheet.get("merges", []):
            ws.merge_cells(str(merged))
        for row, height in sheet.get("rowHeights", {}).items():
            ws.row_dimensions[int(row)].height = height
        for col, width in sheet.get("columnWidths", {}).items():
            ws.column_dimensions[col].width = width
        ws.sheet_state = "hidden" if sheet.get("archived") else "visible"
    if "00_PROJECT_META" in wb.sheetnames:
        ws = wb["00_PROJECT_META"]
        metadata = project.get("metadata", {})
        rows = [("Project ID", project["id"]), ("Project Name", metadata.get("projectName", "")),
                ("Project Profile", project.get("projectProfileId", "")),
                ("Template Version", project.get("projectTemplateVersion", ""))]
        label_rows = {
            str(ws.cell(row, 1).value or "").strip().lower(): row
            for row in range(1, ws.max_row + 1)
        }
        for label, value in rows:
            row = label_rows.get(label.lower())
            if row is None:
                row = ws.max_row + 1
                ws.cell(row, 1, label)
            ws.cell(row, 2, value)
    ordered = [wb[name] for name in desired_order if name in wb.sheetnames]
    ordered.extend(ws for ws in wb.worksheets if ws.title not in desired_order)
    wb._sheets = ordered
    temp = workbook_path.with_name(f".{workbook_path.name}.tmp{workbook_path.suffix}")
    wb.save(temp)
    wb.close()
    os.replace(temp, workbook_path)
    return {"workbookPath": str(workbook_path), "backupPath": str(backup), "sha256": sha256_file(workbook_path), "sheetOrder": desired_order}
