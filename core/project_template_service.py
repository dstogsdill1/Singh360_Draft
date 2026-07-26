"""Transactional creation and backup services for schema-V2 projects."""
from __future__ import annotations

import json
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .project_store import ProjectStore, slugify
from .template_platform import (
    ProfileRegistry, TemplatePlatformError, TemplateRegistry,
    SOURCE_FOLDERS, WorkbookDocumentStore, apply_standard_sheet_style, atomic_json_write,
    sha256_file, utcnow, workbook_to_document,
)


class ProjectTemplateService:
    def __init__(self, store: ProjectStore, profiles: ProfileRegistry, templates: TemplateRegistry):
        self.store = store
        self.profiles = profiles
        self.templates = templates

    def create(self, body: dict[str, Any], source_workbook: Path | None = None) -> dict[str, Any]:
        profile = self.profiles.get(str(body.get("profileId") or ""))
        template = self.templates.get(str(body.get("templateId") or ""))
        if source_workbook is None and profile["id"] not in template.get("supportedProfiles", []):
            raise TemplatePlatformError(f"Template does not support profile {profile['id']}.")
        metadata = body.get("metadata") if isinstance(body.get("metadata"), dict) else {}
        name = str(metadata.get("projectName") or "").strip()
        if not name:
            raise TemplatePlatformError("Project name is required.")
        project_id = uuid.uuid4().hex[:16]
        project_dir = self.store.projects_dir / f"{slugify(name)}__{project_id}"
        if project_dir.exists():
            raise TemplatePlatformError("Generated project folder already exists; retry creation.")
        try:
            self.store.ensure_folders(project_dir)
            source_template = Path(source_workbook) if source_workbook else Path(template["absoluteRuntimePath"])
            if not source_template.is_file():
                raise TemplatePlatformError(f"Project workbook was not found: {source_template}")
            workbook_path = project_dir / "sources" / "workbook" / source_template.name
            shutil.copy2(source_template, workbook_path)
            self._apply_profile(workbook_path, profile, metadata, project_id)
            document = workbook_to_document(workbook_path)
            WorkbookDocumentStore(project_dir).create(document)
            source_manifest = {
                "schemaVersion": 2, "folders": SOURCE_FOLDERS,
                "sources": [], "conversionQueue": [], "importReports": [],
            }
            atomic_json_write(project_dir / "source_library.json", source_manifest)
            project = {
                "id": project_id, "schemaVersion": 2, "metadata": metadata,
                "projectProfileId": profile["id"], "projectTemplateId": template["templateId"],
                "projectTemplateVersion": template["version"], "projectTemplateHash": template["sha256"],
                "workbookDocument": {"path": "data/workbook.json", "revision": document["revision"]},
                "sourceLibrary": {"path": "source_library.json"}, "worksheets": [],
                "pages": [], "sources": [], "lastCompile": None, "compileWarnings": [],
                "sourceWorkbookName": workbook_path.name,
                "workbookSync": {"mode": "project-workbook", "workbook": str(workbook_path), "status": "in_sync", "workbookHash": sha256_file(workbook_path)},
                "createdAt": utcnow(),
            }
            self.store.save(project_id, project)
            return project
        except Exception as exc:
            failure_root = self.store.docs / "failure_logs"
            failure_root.mkdir(parents=True, exist_ok=True)
            atomic_json_write(failure_root / f"project_create_{project_id}_{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.json", {
                "projectId": project_id, "projectFolder": str(project_dir), "error": str(exc), "failedAt": utcnow(),
            })
            if project_dir.exists():
                shutil.rmtree(project_dir)
            raise

    def _apply_profile(self, workbook_path: Path, profile: dict[str, Any], metadata: dict[str, Any], project_id: str) -> None:
        wb = load_workbook(workbook_path)
        for name in profile["dataSheets"]:
            existed = name in wb.sheetnames
            ws = wb[name] if existed else wb.create_sheet(name)
            if not existed:
                ws.append([name.replace("_", " ")])
                ws.append(["Field", "Value", "Source / Notes"])
                ws.append(["", "", ""])
                apply_standard_sheet_style(ws)
        meta = wb["00_PROJECT_META"]
        mapping = [("Project ID", project_id), ("Project Name", metadata.get("projectName", "")),
                   ("Store / Site Number", metadata.get("storeNumber", "")), ("Client", metadata.get("client", "")),
                   ("Location", metadata.get("location", "")), ("Address", metadata.get("address", "")),
                   ("Project Profile", profile["id"]), ("Purpose", metadata.get("purpose", "")),
                   ("Scope Summary", metadata.get("scopeSummary", "")), ("Drawing Prefix", metadata.get("drawingPrefix", "")),
                   ("Revision", metadata.get("revision", "")), ("Drawn By", metadata.get("drawnBy", "")),
                   ("Checked By", metadata.get("checkedBy", ""))]
        self._update_named_values(meta, mapping)

        index = wb["00_INDEX"]
        headers = self._find_headers(index)
        if headers:
            profile_col = headers.get("template profile")
            include_col = headers["include"]
            family_col = headers.get("family")
            for row in range(headers["_row"] + 1, index.max_row + 1):
                if profile_col:
                    index.cell(row, profile_col, profile["id"])
                if family_col and not index.cell(row, include_col).value:
                    family = str(index.cell(row, family_col).value or "")
                    index.cell(
                        row,
                        include_col,
                        "YES" if family in profile["defaultIncludedFamilies"] else "NO",
                    )
        else:
            self._seed_index(index, profile)

        profile_ws = wb["00_TEMPLATE_PROFILE"]
        self._update_named_values(profile_ws, [
            ("Profile ID", profile["id"]),
            ("Profile Version", profile["version"]),
            ("Style Profile", profile["styleProfile"]),
        ])
        wb.save(workbook_path)
        wb.close()

    @staticmethod
    def _update_named_values(ws: Any, mapping: list[tuple[str, Any]]) -> None:
        labels = {str(label).strip().casefold(): value for label, value in mapping}
        found: set[str] = set()
        for row in ws.iter_rows():
            for cell in row:
                key = str(cell.value or "").strip().casefold()
                if key in labels:
                    ws.cell(cell.row, min(cell.column + 1, ws.max_column + 1), labels[key])
                    found.add(key)
                    break
        for label, value in mapping:
            if label.casefold() not in found:
                ws.append([label, value, "User supplied"])

    @staticmethod
    def _find_headers(ws: Any) -> dict[str, int] | None:
        for row_number in range(1, min(ws.max_row, 75) + 1):
            found = {
                str(ws.cell(row_number, column).value or "").strip().casefold(): column
                for column in range(1, ws.max_column + 1)
                if ws.cell(row_number, column).value not in (None, "")
            }
            if {"include", "sheet tab", "page title"}.issubset(found):
                found["_row"] = row_number
                return found
        return None

    @staticmethod
    def _seed_index(ws: Any, profile: dict[str, Any]) -> None:
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
        ws.delete_rows(1, ws.max_row)
        headers = [
            "Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family",
            "Page Type", "Notes", "Render Profile", "Split Mode", "Template Profile",
            "Required", "Issue Status", "Page ID", "Source Mode", "Sync Direction",
            "Color", "Data State",
        ]
        ws.append(headers)
        for order, family in enumerate(profile["defaultIncludedFamilies"], start=1):
            page_id = uuid.uuid5(uuid.NAMESPACE_URL, f"{profile['id']}|{family}|base").hex[:20]
            ws.append([
                "YES", order, "", "", family, family, "Table / Schedule", "",
                "front_matter_table", "auto", profile["id"], "YES", "draft",
                f"generated-{page_id}", "canonical", "app_to_workbook", "GOLD", "READY",
            ])
        apply_standard_sheet_style(ws)
