# S360 CRITICAL SYNC V44
"""Physical Excel mirror tabs for every Singh360 drawing page.

00_INDEX remains the base-page authority. 00_DRAWING_PAGES is the complete
page manifest, including generated continuations. Generated continuation tabs
are tagged through the worksheet codeName and are ignored on workbook import.
"""
from __future__ import annotations

from copy import copy
from hashlib import sha1
from typing import Any
import re

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.page_identity import is_sheet_index_page


MIRROR_CODE_PREFIX = "S360GEN_"
DRAWING_MANIFEST_SHEET = "00_DRAWING_PAGES"
CONTROL_SHEETS = {
    "00_PROJECT_META",
    "00_INDEX",
    "00_HELP",
    "00_AI_GUIDE",
    DRAWING_MANIFEST_SHEET,
}


def is_generated_mirror_sheet(ws: Any) -> bool:
    """Return True only for a Singh360-generated drawing-page mirror sheet."""
    code_name = str(getattr(getattr(ws, "sheet_properties", None), "codeName", "") or "")
    return code_name.upper().startswith(MIRROR_CODE_PREFIX)


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _safe_sheet_name(wb: Any, requested: str) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", _clean_text(requested) or "DRAWING PAGE")
    base = base[:31] or "DRAWING PAGE"
    if base not in wb.sheetnames:
        return base
    index = 2
    while True:
        suffix = f"~{index}"
        candidate = (base[: 31 - len(suffix)] + suffix).strip()
        if candidate not in wb.sheetnames:
            return candidate
        index += 1


def _generated_page(page: dict[str, Any]) -> bool:
    return bool(
        page.get("generatedContinuation")
        or page.get("indexContinuation")
        or page.get("generatedIndexContinuation")
        or page.get("continuationOf")
    )


def _status_key(value: Any) -> str:
    raw = _clean_text(value).lower().replace("-", "_").replace(" ", "_")
    return raw if raw in {"draft", "draft_confirmed", "public", "public_confirmed"} else "draft"


def _tab_color(page: dict[str, Any], tab_colors: dict[str, str]) -> str:
    if page.get("include", True) is False:
        return tab_colors.get("excluded", "9AA3AB")
    return tab_colors.get(_status_key(page.get("issueStatus")), tab_colors.get("draft", "F28C28"))


def _page_code(page: dict[str, Any]) -> str:
    return _clean_text(page.get("displaySheetCode") or page.get("sheetCode") or "PAGE")


def _page_title(page: dict[str, Any]) -> str:
    return _clean_text(page.get("sheetTitle") or "Untitled Drawing Page")


def _mirror_sheet_name(wb: Any, page: dict[str, Any]) -> str:
    code = _page_code(page)
    title = re.sub(r"\s*[—-]\s*CONTINUED\s*$", "", _page_title(page), flags=re.IGNORECASE)
    title = re.sub(r"\s*[—-]\s*SOURCE\s+[^—]+$", "", title, flags=re.IGNORECASE)
    words = [word for word in title.split() if word]
    short = " ".join(words[:4])
    requested = f"{code} {short}".strip()
    return _safe_sheet_name(wb, requested)


def _index_header_map_from_worksheet(ws: Any) -> tuple[int, dict[str, int]]:
    max_row = min(max(int(ws.max_row or 1), 1), 75)
    max_col = max(int(ws.max_column or 1), 18)
    for row_number in range(1, max_row + 1):
        found = {
            _clean_text(ws.cell(row_number, column).value).casefold(): column
            for column in range(1, max_col + 1)
            if _clean_text(ws.cell(row_number, column).value)
        }
        if {"sheet tab", "page id"}.issubset(found):
            return row_number, found
    raise RuntimeError(
        "00_INDEX is missing the Sheet Tab / Page ID header row required "
        "for exact drawing-page tab synchronization."
    )


def _index_header_map_from_grid(
    grid: list[list[Any]],
) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(grid[:75]):
        found = {
            _clean_text(value).casefold(): column_index
            for column_index, value in enumerate(row)
            if _clean_text(value)
        }
        if {"sheet tab", "page id"}.issubset(found):
            return row_index, found
    return None


def _project_index_worksheet(project: dict[str, Any]) -> dict[str, Any] | None:
    pages = [
        page
        for page in project.get("pages", [])
        if isinstance(page, dict)
        and is_sheet_index_page(page)
        and not _generated_page(page)
    ]
    linked_id = _clean_text(
        pages[0].get("linkedWorksheetId")
        if pages
        else ""
    )
    worksheets = [
        worksheet
        for worksheet in project.get("worksheets", [])
        if isinstance(worksheet, dict)
    ]
    if linked_id:
        linked = next(
            (
                worksheet
                for worksheet in worksheets
                if _clean_text(worksheet.get("id")) == linked_id
            ),
            None,
        )
        if linked is not None:
            return linked

    for worksheet in worksheets:
        names = (
            worksheet.get("name"),
            worksheet.get("sourceSheet"),
            worksheet.get("sheetTab"),
        )
        if any(
            re.sub(r"[\s_]+", "", _clean_text(name)).casefold()
            == "00index"
            for name in names
        ):
            return worksheet
    return None


def sync_base_index_sheet_tabs(
    wb: Any,
    project: dict[str, Any],
    page_tabs: list[tuple[dict[str, Any], str, str]],
) -> dict[str, Any]:
    """Make 00_INDEX and the saved project point to exact physical base tabs.

    The full workbook engine writes the base-only 00_INDEX before generated
    drawing-page mirrors are created. A generated base page such as the Sheet
    Index can therefore receive a different final physical tab title
    ("EMS 2.0 Sheet Index - TOC") after 00_INDEX was already written. This
    function closes that gap after mirror creation, using stable Page IDs.
    """
    base_tabs = {
        _clean_text(page.get("id")): str(excel_tab)
        for page, excel_tab, _kind in page_tabs
        if _clean_text(page.get("id"))
        and not _generated_page(page)
    }
    if not base_tabs:
        return {
            "basePageCount": 0,
            "workbookRowsUpdated": 0,
            "projectPagesUpdated": 0,
            "projectIndexRowsUpdated": 0,
        }

    index_sheet = next(
        (
            sheet
            for sheet in wb.worksheets
            if re.sub(
                r"[\s_]+",
                "",
                _clean_text(sheet.title),
            ).casefold()
            == "00index"
        ),
        None,
    )
    if index_sheet is None:
        raise RuntimeError(
            "00_INDEX was not found while synchronizing exact base-page tabs."
        )

    header_row, headers = _index_header_map_from_worksheet(index_sheet)
    tab_column = headers["sheet tab"]
    page_id_column = headers["page id"]
    matched_page_ids: set[str] = set()
    workbook_updates = 0
    blank_run = 0

    for row_number in range(
        header_row + 1,
        min(int(index_sheet.max_row or header_row), 10000) + 1,
    ):
        page_id = _clean_text(
            index_sheet.cell(row_number, page_id_column).value
        )
        sheet_tab = _clean_text(
            index_sheet.cell(row_number, tab_column).value
        )
        if not page_id and not sheet_tab:
            blank_run += 1
            if blank_run >= 100:
                break
            continue
        blank_run = 0

        physical_tab = base_tabs.get(page_id)
        if physical_tab is None:
            continue
        matched_page_ids.add(page_id)
        if index_sheet.cell(row_number, tab_column).value != physical_tab:
            index_sheet.cell(row_number, tab_column, physical_tab)
            workbook_updates += 1

    missing = sorted(set(base_tabs) - matched_page_ids)
    if missing:
        raise RuntimeError(
            "00_INDEX is missing base Page ID rows required for exact physical "
            "tab synchronization: " + ", ".join(missing[:20])
        )

    project_page_updates = 0
    for page in project.get("pages", []):
        if not isinstance(page, dict):
            continue
        page_id = _clean_text(page.get("id"))
        physical_tab = base_tabs.get(page_id)
        if physical_tab is None:
            continue
        if page.get("sheetTab") != physical_tab:
            page["sheetTab"] = physical_tab
            project_page_updates += 1
        page["workbookMirrorTab"] = physical_tab

    project_index_updates = 0
    project_index = _project_index_worksheet(project)
    if project_index is not None:
        grid = project_index.get("grid")
        if isinstance(grid, list):
            header_data = _index_header_map_from_grid(grid)
            if header_data is not None:
                grid_header_row, grid_headers = header_data
                grid_tab_column = grid_headers["sheet tab"]
                grid_page_id_column = grid_headers["page id"]
                for row in grid[grid_header_row + 1 :]:
                    if not isinstance(row, list):
                        continue
                    while len(row) <= max(grid_tab_column, grid_page_id_column):
                        row.append("")
                    page_id = _clean_text(row[grid_page_id_column])
                    physical_tab = base_tabs.get(page_id)
                    if physical_tab is None:
                        continue
                    if row[grid_tab_column] != physical_tab:
                        row[grid_tab_column] = physical_tab
                        project_index_updates += 1

    return {
        "basePageCount": len(base_tabs),
        "workbookRowsUpdated": workbook_updates,
        "projectPagesUpdated": project_page_updates,
        "projectIndexRowsUpdated": project_index_updates,
    }


def _parse_style_key(key: str) -> tuple[int, int] | None:
    text = str(key or "").strip()
    if re.fullmatch(r"\d+:\d+", text):
        row, col = text.split(":", 1)
        return int(row), int(col)
    match = re.fullmatch(r"([A-Z]+)(\d+)", text, re.IGNORECASE)
    if not match:
        return None
    letters, row_text = match.groups()
    col = 0
    for char in letters.upper():
        col = col * 26 + ord(char) - ord("A") + 1
    return int(row_text) - 1, col - 1


def _hex(value: Any) -> str | None:
    text = str(value or "").strip().lstrip("#")
    if len(text) == 8:
        text = text[-6:]
    if len(text) == 6 and all(char in "0123456789abcdefABCDEF" for char in text):
        return text.upper()
    return None


def _apply_style(cell: Any, spec: dict[str, Any]) -> None:
    if not isinstance(spec, dict):
        return

    font = copy(cell.font)
    changed_font = False
    for source, target in (
        ("bold", "bold"),
        ("italic", "italic"),
        ("underline", "underline"),
        ("fontSize", "size"),
        ("fontName", "name"),
    ):
        if source in spec:
            setattr(font, target, spec[source])
            changed_font = True
    font_color = _hex(spec.get("fontColor"))
    if font_color:
        font.color = font_color
        changed_font = True
    if changed_font:
        cell.font = font

    alignment = copy(cell.alignment)
    changed_alignment = False
    for source, target in (
        ("hAlign", "horizontal"),
        ("vAlign", "vertical"),
        ("wrap", "wrap_text"),
        ("rotation", "text_rotation"),
        ("indent", "indent"),
    ):
        if source in spec:
            setattr(alignment, target, spec[source])
            changed_alignment = True
    if changed_alignment:
        cell.alignment = alignment

    fill = _hex(spec.get("fill"))
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

    border_spec = spec.get("borders")
    if isinstance(border_spec, dict):
        old = cell.border
        sides: dict[str, Any] = {}
        for side_name in ("left", "right", "top", "bottom"):
            side_data = border_spec.get(side_name)
            if isinstance(side_data, dict) and side_data.get("style"):
                sides[side_name] = Side(
                    style=str(side_data.get("style")),
                    color=_hex(side_data.get("color")) or "000000",
                )
            else:
                sides[side_name] = getattr(old, side_name)
        cell.border = Border(
            left=sides["left"],
            right=sides["right"],
            top=sides["top"],
            bottom=sides["bottom"],
            diagonal=old.diagonal,
            diagonal_direction=old.diagonal_direction,
            diagonalUp=old.diagonalUp,
            diagonalDown=old.diagonalDown,
            outline=old.outline,
            vertical=old.vertical,
            horizontal=old.horizontal,
        )


def _write_grid(
    ws: Any,
    grid: list[list[Any]],
    *,
    start_row: int = 1,
    styles: dict[str, Any] | None = None,
    merges: list[dict[str, Any]] | None = None,
    col_widths: list[Any] | None = None,
    row_heights: list[Any] | None = None,
) -> tuple[int, int]:
    max_columns = max((len(row) for row in grid), default=1)
    max_rows = max(len(grid), 1)

    for row_index, row in enumerate(grid, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(start_row + row_index - 1, col_index, value)

    for key, spec in (styles or {}).items():
        position = _parse_style_key(str(key))
        if position is None:
            continue
        row, col = position
        if row < 0 or col < 0:
            continue
        _apply_style(ws.cell(start_row + row, col + 1), spec)

    for merge in merges or []:
        try:
            merge_start_row = int(merge.get("startRow", 0)) + 1
            merge_start_col = int(merge.get("startCol", 0)) + 1
            merge_end_row = int(merge.get("endRow", merge_start_row - 1)) + 1
            merge_end_col = int(merge.get("endCol", merge_start_col - 1)) + 1
            if merge_end_row >= merge_start_row and merge_end_col >= merge_start_col:
                ws.merge_cells(
                    start_row=start_row + merge_start_row - 1,
                    start_column=merge_start_col,
                    end_row=start_row + merge_end_row - 1,
                    end_column=merge_end_col,
                )
        except Exception:
            continue

    for index, width in enumerate(col_widths or [], start=1):
        try:
            pixels = float(width)
            ws.column_dimensions[get_column_letter(index)].width = max(2.0, min(120.0, (pixels - 5.0) / 7.0))
        except Exception:
            continue

    for index, height in enumerate(row_heights or [], start=1):
        try:
            pixels = float(height)
            ws.row_dimensions[start_row + index - 1].height = max(8.0, min(240.0, pixels * 0.75))
        except Exception:
            continue

    return start_row + max_rows - 1, max_columns


def _table_grid(block: dict[str, Any]) -> list[list[Any]]:
    headers = block.get("headers")
    rows = block.get("rows")
    grid: list[list[Any]] = []
    if isinstance(headers, list) and headers:
        grid.append(list(headers))
    if isinstance(rows, list):
        for row in rows:
            if isinstance(row, list):
                grid.append(list(row))
            elif isinstance(row, dict):
                keys = list(row)
                if not grid:
                    grid.append(keys)
                grid.append([row.get(key, "") for key in keys])
            else:
                grid.append([row])
    return grid


def _render_page(ws: Any, page: dict[str, Any]) -> None:
    blocks = [block for block in page.get("blocks", []) if isinstance(block, dict)]
    start_row = 1
    rendered = False

    for block in blocks:
        block_type = str(block.get("type") or "")
        if block_type == "excelRange" and isinstance(block.get("grid"), list):
            rows, cols = _write_grid(
                ws,
                [list(row) for row in block.get("grid") or []],
                styles=block.get("styles") if isinstance(block.get("styles"), dict) else {},
                merges=block.get("mergedCells") if isinstance(block.get("mergedCells"), list) else [],
                col_widths=block.get("colWidths") if isinstance(block.get("colWidths"), list) else [],
                row_heights=block.get("rowHeights") if isinstance(block.get("rowHeights"), list) else [],
                start_row=start_row,
            )
            start_row = rows + 3
            rendered = True
            continue

        if block_type in {"table", "matrix", "indexTable", "idfNetworkTable"}:
            grid = _table_grid(block)
            if grid:
                rows, cols = _write_grid(ws, grid, start_row=start_row)
                header_row = start_row
                if grid:
                    for cell in ws[header_row]:
                        if cell.column <= cols:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill("solid", fgColor="252C34")
                            cell.alignment = Alignment(wrap_text=True, vertical="center")
                start_row = rows + 3
                rendered = True
                continue

    if not rendered:
        ws["A1"] = _page_code(page)
        ws["A2"] = _page_title(page)
        ws["A4"] = _clean_text(page.get("notes"))
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="252C34")
        ws["A2"].font = Font(size=14, bold=True)
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 100


def _prepare_mirror_sheet(ws: Any, page: dict[str, Any], tab_colors: dict[str, str]) -> None:
    page_id = _clean_text(page.get("id"))
    digest = sha1(page_id.encode("utf-8")).hexdigest()[:12].upper()
    ws.sheet_properties.codeName = f"{MIRROR_CODE_PREFIX}{digest}"
    ws.sheet_properties.tabColor = _tab_color(page, tab_colors)
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = None
    ws.oddHeader.center.text = f"{_page_code(page)}  {_page_title(page)}"
    ws.oddFooter.center.text = "Generated drawing-page mirror — edit the source worksheet or Singh360 project."
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def _clear_sheet(ws: Any) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for key in list(ws.column_dimensions):
        del ws.column_dimensions[key]
    for key in list(ws.row_dimensions):
        del ws.row_dimensions[key]


def _write_manifest(
    wb: Any,
    project: dict[str, Any],
    page_tabs: list[tuple[dict[str, Any], str, str]],
    tab_colors: dict[str, str],
) -> Any:
    if DRAWING_MANIFEST_SHEET in wb.sheetnames:
        ws = wb[DRAWING_MANIFEST_SHEET]
        _clear_sheet(ws)
    else:
        ws = wb.create_sheet(DRAWING_MANIFEST_SHEET)

    ws.sheet_properties.codeName = "S360_DRAWING_PAGES"
    ws.sheet_properties.tabColor = tab_colors.get("control", "252C34")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws["A1"] = "SINGH360 DRAWING PAGE MANIFEST"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="252C34")
    ws.merge_cells("A1:K2")
    ws["A3"] = (
        "00_INDEX controls base pages. This sheet lists every Singh360 drawing page, "
        "including generated continuations and the exact physical Excel mirror tab."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:K3")

    headers = [
        "Page", "Include", "Sheet Code", "Page Title", "Excel Tab",
        "Page Kind", "Source Base Tab", "Page ID", "Parent / Group ID",
        "Issue Status", "Notes",
    ]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(4, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="276FA8")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_number, (page, excel_tab, kind) in enumerate(page_tabs, start=5):
        values = [
            page.get("pageNumber") if page.get("include", True) else "",
            "YES" if page.get("include", True) else "NO",
            _page_code(page),
            _page_title(page),
            excel_tab,
            kind,
            _clean_text(page.get("sheetTab") or page.get("sourceSheet")),
            _clean_text(page.get("id")),
            _clean_text(page.get("continuationOf") or page.get("pageGroupId")),
            _clean_text(page.get("issueStatus") or "Draft"),
            _clean_text(page.get("notes")),
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if page.get("include", True) is False:
            for column in range(1, len(headers) + 1):
                ws.cell(row_number, column).fill = PatternFill("solid", fgColor="E6EBEF")

    widths = [10, 10, 18, 44, 31, 15, 31, 38, 38, 20, 60]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    if page_tabs:
        ws.auto_filter.ref = f"A4:K{4 + len(page_tabs)}"
    return ws


def rebuild_drawing_page_mirrors(
    wb: Any,
    project: dict[str, Any],
    page_sheet_by_id: dict[str, Any],
    project_id: str,
    tab_colors: dict[str, str],
) -> dict[str, Any]:
    """Rebuild tagged continuation/index mirrors and the full page manifest."""
    removed: list[str] = []
    for ws in list(wb.worksheets):
        if is_generated_mirror_sheet(ws):
            removed.append(str(ws.title))
            wb.remove(ws)

    pages = [
        page for page in project.get("pages", [])
        if isinstance(page, dict)
    ]
    pages.sort(key=lambda page: int(page.get("order") or 10**9))

    page_sheets: list[Any] = []
    page_tabs: list[tuple[dict[str, Any], str, str]] = []
    generated_count = 0
    used_sheet_ids: set[int] = set()

    for page in pages:
        page_id = _clean_text(page.get("id"))
        use_generated_mirror = _generated_page(page) or is_sheet_index_page(page)
        sheet = None if use_generated_mirror else page_sheet_by_id.get(page_id)

        if sheet is None:
            title = _mirror_sheet_name(wb, page)
            sheet = wb.create_sheet(title)
            _prepare_mirror_sheet(sheet, page, tab_colors)
            _render_page(sheet, page)
            generated_count += 1
            kind = "Generated continuation" if _generated_page(page) else "Generated page mirror"
        else:
            kind = "Base worksheet"

        if id(sheet) in used_sheet_ids:
            title = _mirror_sheet_name(wb, page)
            sheet = wb.create_sheet(title)
            _prepare_mirror_sheet(sheet, page, tab_colors)
            _render_page(sheet, page)
            generated_count += 1
            kind = "Generated page mirror"

        used_sheet_ids.add(id(sheet))
        page["workbookMirrorTab"] = str(sheet.title)
        page_sheets.append(sheet)
        page_tabs.append((page, str(sheet.title), kind))

    index_tab_sync = sync_base_index_sheet_tabs(
        wb,
        project,
        page_tabs,
    )
    manifest_sheet = _write_manifest(wb, project, page_tabs, tab_colors)

    return {
        "pageSheets": page_sheets,
        "manifestSheet": manifest_sheet,
        "generatedMirrorCount": generated_count,
        "drawingPageCount": len(page_tabs),
        "removedMirrorTabs": removed,
        "indexTabSync": index_tab_sync,
        "pageTabs": [
            {
                "pageId": _clean_text(page.get("id")),
                "sheetCode": _page_code(page),
                "excelTab": tab,
                "kind": kind,
            }
            for page, tab, kind in page_tabs
        ],
    }
