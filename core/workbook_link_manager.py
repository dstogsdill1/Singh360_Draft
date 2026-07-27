from __future__ import annotations

from datetime import datetime, timezone
from hashlib import sha256
import json
import os
from pathlib import Path
import shutil
import subprocess
import threading
import time
import traceback
import uuid
from typing import Any

from openpyxl import load_workbook

from core.workbook_status_sync import (
    WorkbookSyncError,
    file_hash,
    project_hash,
    sync_project_from_workbook,
    sync_project_to_workbook,
)


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def normalize_path(value: str) -> Path:
    expanded = os.path.expandvars(os.path.expanduser(str(value or "").strip().strip('"')))
    if not expanded:
        raise WorkbookSyncError("No workbook path was provided.")
    return Path(expanded)


def internal_workbook_path(store: Any, project_id: str, project: dict[str, Any]) -> Path | None:
    folder = store.sources_dir(project_id, "workbook", project)
    preferred = str(project.get("sourceWorkbookName") or project.get("metadata", {}).get("sourceFile") or "").strip()
    if preferred:
        candidate = folder / Path(preferred).name
        if candidate.is_file():
            return candidate
    files = sorted(
        [*folder.glob("*.xlsx"), *folder.glob("*.xlsm")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


def configured_workbook_path(store: Any, project_id: str, project: dict[str, Any]) -> tuple[Path | None, str]:
    sync = project.get("workbookSync") if isinstance(project.get("workbookSync"), dict) else {}
    configured = str(sync.get("workbook") or "").strip()
    if configured:
        return normalize_path(configured), "external"
    internal = internal_workbook_path(store, project_id, project)
    return internal, "internal" if internal else "none"


def workbook_metadata(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise WorkbookSyncError(f"The linked workbook was not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise WorkbookSyncError("The linked file must be an .xlsx or .xlsm workbook.")
    try:
        wb = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    except PermissionError as exc:
        raise WorkbookSyncError("The linked workbook is open or locked. Close Excel and try again.") from exc
    except Exception as exc:
        raise WorkbookSyncError(f"The linked workbook could not be read: {exc}") from exc
    try:
        if "00_PROJECT_META" not in wb.sheetnames or "00_INDEX" not in wb.sheetnames:
            raise WorkbookSyncError("This is not a Singh360 workbook. It must contain 00_PROJECT_META and 00_INDEX.")
        meta: dict[str, str] = {}
        ws = wb["00_PROJECT_META"]
        for row_number, row in enumerate(
            ws.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True),
            start=1,
        ):
            if row_number > 80:
                break
            key = str(row[0] or "").strip()
            if key:
                meta[key.casefold()] = str(row[1] or "").strip()
        return {
            "path": str(path),
            "filename": path.name,
            "sheetCount": len(wb.sheetnames),
            "projectId": meta.get("linked project id", ""),
            "schemaVersion": meta.get("workbook schema version", ""),
            "helpVersion": meta.get("help version", ""),
            "projectName": meta.get("project name", "") or meta.get("project", ""),
            "modified": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "size": path.stat().st_size,
            "sha256": file_hash(path),
        }
    finally:
        wb.close()


def claim_workbook_for_project(path: Path, project_id: str) -> dict[str, Any]:
    """Assign a copied workbook to ``project_id`` without touching its source.

    New projects receive their own workbook copy before this function runs.
    The copy is saved through a sibling temporary file and atomically replaced,
    so a failed openpyxl save cannot leave a corrupt authority workbook.
    """
    path = Path(path)
    workbook_metadata(path)  # validates extension, readability, and control sheets
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, data_only=False, keep_vba=keep_vba)
    temp_path = path.with_name(
        f".{path.stem}.authority-{uuid.uuid4().hex[:8]}{path.suffix.lower()}"
    )
    saved = False
    try:
        sheet = workbook["00_PROJECT_META"]
        linked_row = None
        last_sync_row = None
        last_status_row = None
        for row_number in range(1, max(sheet.max_row, 30) + 1):
            key = str(sheet.cell(row_number, 1).value or "").strip().casefold()
            if key == "linked project id":
                linked_row = row_number
            elif key == "last sync utc":
                last_sync_row = row_number
            elif key == "last sync status":
                last_status_row = row_number
        next_row = max(sheet.max_row + 1, 17)
        if linked_row is None:
            linked_row = next_row
            next_row += 1
        if last_sync_row is None:
            last_sync_row = next_row
            next_row += 1
        if last_status_row is None:
            last_status_row = next_row

        sheet.cell(linked_row, 1, "Linked Project ID")
        sheet.cell(linked_row, 2, project_id)
        sheet.cell(last_sync_row, 1, "Last Sync UTC")
        sheet.cell(last_sync_row, 2, utcnow())
        sheet.cell(last_status_row, 1, "Last Sync Status")
        sheet.cell(last_status_row, 2, "Synchronized")
        workbook.save(temp_path)
        saved = True
    finally:
        workbook.close()
        if not saved:
            temp_path.unlink(missing_ok=True)
    try:
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)

    metadata = workbook_metadata(path)
    if metadata.get("projectId") != project_id:
        raise WorkbookSyncError(
            "The project workbook copy could not be assigned to the new project ID."
        )
    return metadata


def initialize_internal_workbook_link(
    project_id: str,
    project: dict[str, Any],
    path: Path,
    *,
    configured_path: Path | None = None,
) -> dict[str, Any]:
    """Attach a package-owned workbook and establish a clean authority baseline."""
    path = Path(path)
    authority_path = Path(configured_path) if configured_path is not None else path
    metadata = workbook_metadata(path)
    linked_id = str(metadata.get("projectId") or "")
    if linked_id != project_id:
        raise WorkbookSyncError(
            f"The workbook is linked to project {linked_id or '(none)'}, not {project_id}."
        )

    project = dict(project)
    project["sourceWorkbookName"] = authority_path.name
    project_metadata = dict(project.get("metadata") or {})
    project_metadata["sourceFile"] = authority_path.name
    project["metadata"] = project_metadata
    sources = [
        source
        for source in list(project.get("sources") or [])
        if not (
            isinstance(source, dict)
            and str(source.get("type") or "").casefold() == "workbook"
        )
    ]
    sources.append(
        {
            "id": f"src_{project_id}_xlsx",
            "type": "workbook",
            "name": authority_path.name,
            "path": str(authority_path),
        }
    )
    project["sources"] = sources
    sync = {
        **dict(project.get("workbookSync") or {}),
        "mode": "internal-workbook-copy",
        "workbook": str(authority_path),
        "status": "in_sync",
        "warning": "",
        "pendingReason": "",
        "runtimeLog": "",
        "linkedAt": utcnow(),
        "lastSyncUtc": utcnow(),
        "workbookHash": file_hash(path),
        "authority": "workbook",
        "lastAuthorityAction": "project_import",
        "syncEngineVersion": "V25",
    }
    project["workbookSync"] = sync
    sync["appHash"] = project_hash(project)
    return project


def status_payload(project_id: str, project: dict[str, Any], store: Any) -> dict[str, Any]:
    path, mode = configured_workbook_path(store, project_id, project)
    sync = project.get("workbookSync") if isinstance(project.get("workbookSync"), dict) else {}
    if path is None:
        return {
            "ok": True,
            "status": "not_linked",
            "mode": "none",
            "path": "",
            "message": "No workbook is linked. Choose the project workbook.",
        }
    if not path.is_file():
        return {
            "ok": True,
            "status": "missing",
            "mode": mode,
            "path": str(path),
            "message": "The linked workbook moved, was renamed, or is unavailable.",
        }
    try:
        meta = workbook_metadata(path)
    except WorkbookSyncError as exc:
        message = str(exc)
        status = "locked" if "open or locked" in message else "invalid"
        return {"ok": True, "status": status, "mode": mode, "path": str(path), "message": message}

    baseline_workbook = str(sync.get("workbookHash") or "")
    baseline_app = str(sync.get("appHash") or "")
    current_workbook = meta["sha256"]
    current_app = project_hash(project)

    if not baseline_workbook or not baseline_app:
        status = "review_required"
        message = "Workbook linked. Choose which version should establish the first synchronization."
    else:
        workbook_changed = current_workbook != baseline_workbook
        app_changed = current_app != baseline_app
        if workbook_changed and app_changed:
            status = "conflict"
            message = "Both the workbook and the app changed after the last sync."
        elif workbook_changed:
            status = "workbook_changed"
            message = "The workbook changed and can update the app."
        elif app_changed:
            status = "app_changed"
            message = "The app changed and can update the workbook."
        else:
            status = "in_sync"
            message = "The project and workbook match."

    project_meta_id = str(meta.get("projectId") or "")
    if project_meta_id and project_meta_id != project_id:
        status = "project_mismatch"
        message = f"The workbook is linked to project {project_meta_id}, not {project_id}."

    return {
        "ok": True,
        "status": status,
        "mode": mode,
        "path": str(path),
        "message": message,
        "workbook": meta,
        "baselineWorkbookHash": baseline_workbook,
        "baselineAppHash": baseline_app,
        "currentWorkbookHash": current_workbook,
        "currentAppHash": current_app,
        "lastSyncUtc": sync.get("lastSyncUtc", ""),
        "warning": sync.get("warning", ""),
    }


def _project_name(value: Any) -> str:
    return str(value or "").strip()


def _project_name_key(value: Any) -> str:
    return "".join(ch.lower() for ch in _project_name(value) if ch.isalnum())


def _current_project_name(project: dict[str, Any]) -> str:
    metadata = project.get("metadata") if isinstance(project.get("metadata"), dict) else {}
    return (
        _project_name(project.get("projectDisplayName"))
        or _project_name(metadata.get("projectName"))
        or _project_name(metadata.get("project"))
    )


def _validate_workbook_project_name(project: dict[str, Any], meta: dict[str, Any]) -> None:
    workbook_name = _project_name(meta.get("projectName"))
    current_name = _current_project_name(project)
    if not workbook_name or not current_name:
        return
    workbook_key = _project_name_key(workbook_name)
    current_key = _project_name_key(current_name)
    if workbook_key and current_key and workbook_key != current_key:
        raise WorkbookSyncError(
            "The selected workbook appears to belong to a different project. "
            f"Active project: {current_name}. Workbook project: {workbook_name}. "
            "Select the correct project on the left before confirming the workbook."
        )


def validate_workbook_matches_project(
    project: dict[str, Any],
    path: Path,
) -> dict[str, Any]:
    """Validate required sheets and project-name identity before reimport."""
    metadata = workbook_metadata(Path(path))
    _validate_workbook_project_name(project, metadata)
    return metadata


def unlink(project_id: str, project: dict[str, Any], store: Any) -> dict[str, Any]:
    project = dict(project)
    project["workbookSync"] = {
        "mode": "unlinked",
        "status": "not_linked",
        "warning": "",
        "lastSyncUtc": "",
        "workbookHash": "",
        "appHash": "",
    }
    store.save(project_id, project)
    return project


def _baseline(project: dict[str, Any], path: Path, status: str = "in_sync") -> dict[str, Any]:
    project = dict(project)
    sync = dict(project.get("workbookSync") or {})
    sync.update({
        "mode": sync.get("mode") or "external-workbook-link",
        "workbook": str(path),
        "status": status,
        "warning": "",
        "lastSyncUtc": utcnow(),
        "workbookHash": file_hash(path),
        "appHash": project_hash(project),
    })
    project["workbookSync"] = sync
    return project


def _trim_resolution_backups(folder: Path, keep: int = 20) -> None:
    if not folder.is_dir():
        return
    entries = sorted(
        [item for item in folder.iterdir() if item.is_dir()],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for stale in entries[keep:]:
        shutil.rmtree(stale, ignore_errors=True)


def create_resolution_backup(
    project_id: str,
    project: dict[str, Any],
    store: Any,
    workbook_path: Path,
    direction: str,
) -> Path:
    """Create a matched project/workbook snapshot before either side is changed."""
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    root = store.docs / "backups" / "workbook_resolution" / project_id
    target = root / stamp
    target.mkdir(parents=True, exist_ok=False)

    project_folder = store.dir_for(project_id, project)
    project_json = project_folder / "project.json"
    if project_json.is_file():
        shutil.copy2(project_json, target / "project_before.json")
    else:
        (target / "project_before.json").write_text(
            json.dumps(project, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )

    shutil.copy2(workbook_path, target / f"workbook_before{workbook_path.suffix.lower()}")
    (target / "resolution_manifest.json").write_text(
        json.dumps(
            {
                "createdUtc": utcnow(),
                "projectId": project_id,
                "direction": direction,
                "workbook": str(workbook_path),
                "projectHash": project_hash(project),
                "workbookHash": file_hash(workbook_path),
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    _trim_resolution_backups(root)
    return target


def resolve(project_id: str, project: dict[str, Any], store: Any, direction: str) -> tuple[dict[str, Any], dict[str, Any]]:
    path, _ = configured_workbook_path(store, project_id, project)
    if path is None or not path.is_file():
        raise WorkbookSyncError("The linked workbook is missing. Relocate it first.")
    if direction not in {"workbook_to_app", "app_to_workbook", "baseline"}:
        raise WorkbookSyncError("Unknown synchronization direction.")

    backup = create_resolution_backup(project_id, project, store, path, direction)

    if direction == "workbook_to_app":
        project = sync_project_from_workbook(project_id, project, store)
    elif direction == "app_to_workbook":
        project = sync_project_to_workbook(project_id, project, store)

    project = _baseline(project, path)
    sync = dict(project.get("workbookSync") or {})
    sync["lastResolutionBackup"] = str(backup)
    sync["lastResolutionDirection"] = direction
    project["workbookSync"] = sync
    store.save(project_id, project)
    status = status_payload(project_id, project, store)
    status["resolutionBackup"] = str(backup)
    status["resolutionDirection"] = direction
    return project, status


def sync_auto(project_id: str, project: dict[str, Any], store: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    status = status_payload(project_id, project, store)
    state = status["status"]
    if state == "workbook_changed":
        return resolve(project_id, project, store, "workbook_to_app")
    if state == "app_changed":
        return resolve(project_id, project, store, "app_to_workbook")
    if state == "in_sync":
        return project, status
    if state == "review_required":
        raise WorkbookSyncError("Choose Use Workbook or Use App to establish the first synchronization.")
    if state == "conflict":
        raise WorkbookSyncError("Both sides changed. Choose Use Workbook or Use App.")
    raise WorkbookSyncError(status.get("message") or "Workbook synchronization is unavailable.")


def _record_runtime_sync_failure(
    store: Any,
    project_id: str,
    stage: str,
    exc: Exception,
) -> str:
    folder = store.docs / "patch_logs" / "workbook_sync_runtime"
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / f"{project_id}-{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}.json"
    target.write_text(
        json.dumps(
            {
                "createdUtc": utcnow(),
                "projectId": project_id,
                "stage": stage,
                "exceptionType": type(exc).__name__,
                "message": str(exc),
                "traceback": traceback.format_exc(),
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return str(target)


def _pending_after_local_save(
    project: dict[str, Any],
    reason: str,
    warning: str,
    log_path: str = "",
) -> dict[str, Any]:
    project = dict(project)
    project["workbookSync"] = {
        **dict(project.get("workbookSync") or {}),
        "status": "pending",
        "pendingReason": reason,
        "warning": warning,
        "runtimeLog": log_path,
        "localProjectSavedAt": utcnow(),
    }
    return project


def choose_path_native() -> str:
    """Open a Windows workbook picker in its own STA process.

    Flask handles requests on worker threads. tkinter can display a dialog from
    those threads and then fail during cleanup, which produced the HTML 500 after
    the user selected a valid G: drive workbook. A separate PowerShell STA
    process owns the Windows Forms dialog and returns only the selected path.
    """
    if os.name != "nt":
        raise WorkbookSyncError(
            "Native workbook browsing is available only in the local Windows application."
        )

    script = r"""
Add-Type -AssemblyName System.Windows.Forms
$OutputEncoding = [Console]::OutputEncoding = [System.Text.UTF8Encoding]::new()
$dialog = New-Object System.Windows.Forms.OpenFileDialog
$dialog.Title = 'Choose the Singh360 project workbook'
$dialog.Filter = 'Excel workbooks (*.xlsx;*.xlsm)|*.xlsx;*.xlsm|All files (*.*)|*.*'
$dialog.CheckFileExists = $true
$dialog.CheckPathExists = $true
$dialog.Multiselect = $false
$dialog.RestoreDirectory = $true
if ($dialog.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) {
    [Console]::Out.Write($dialog.FileName)
}
"""
    completed = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-STA",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            script,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=900,
        check=False,
    )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        raise WorkbookSyncError(
            "The Windows workbook picker failed."
            + (f" {detail}" if detail else "")
        )
    return (completed.stdout or "").strip()


def open_workbook(path: Path) -> None:
    if not path.is_file():
        raise WorkbookSyncError(f"The linked workbook was not found: {path}")
    if os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
    else:
        subprocess.Popen(["xdg-open", str(path)])


def reveal_workbook(path: Path) -> None:
    if not path.exists():
        raise WorkbookSyncError(f"The linked workbook was not found: {path}")
    if os.name == "nt":
        subprocess.Popen(["explorer.exe", f"/select,{path}"])
    else:
        subprocess.Popen(["xdg-open", str(path.parent)])
# S360 WORKBOOK AUTHORITY LINK V15
# Workbook is authoritative. There is no first-sync version-choice workflow.


def set_link(project_id: str, project: dict[str, Any], store: Any, path_value: str) -> tuple[dict[str, Any], dict[str, Any]]:
    path = normalize_path(path_value)
    meta = workbook_metadata(path)
    _validate_workbook_project_name(project, meta)
    project = dict(project)
    sync = dict(project.get("workbookSync") or {})
    sync.update({
        "mode": "external-workbook-link",
        "workbook": str(path),
        "status": "workbook_changed",
        "linkedAt": utcnow(),
        "warning": "",
        "workbookHash": "",
        "appHash": "",
        "authority": "workbook",
    })
    project["workbookSync"] = sync
    project["sourceWorkbookName"] = path.name
    metadata = dict(project.get("metadata") or {})
    metadata["sourceFile"] = path.name
    project["metadata"] = metadata
    sources = list(project.get("sources") or [])
    for source in sources:
        if isinstance(source, dict) and str(source.get("type") or "").casefold() == "workbook":
            source["name"] = path.name
            source["path"] = str(path)
    project["sources"] = sources
    try:
        internal = store.sources_dir(project_id, "workbook", project) / path.name
        if internal.resolve() != path.resolve():
            shutil.copy2(path, internal)
    except OSError:
        pass
    store.save(project_id, project)
    project, status = resolve(project_id, project, store, "workbook_to_app")
    return project, status


# S360 LOCAL DRAFT OPEN MODE V23.1

# S360 CONTROLLED OPEN SYNC V25
# S360 EDITOR WRITE-BACK OPEN POLICY V26
def maybe_pull_on_open(project_id: str, project: dict[str, Any], store: Any) -> dict[str, Any]:
    status = status_payload(project_id, project, store)
    state = str(status.get("status") or "")
    if state == "in_sync":
        return project
    if state == "workbook_changed":
        refreshed, _ = resolve(project_id, project, store, "workbook_to_app")
        return refreshed
    if state in {"app_changed", "pending"}:
        opened = dict(project)
        sync = dict(opened.get("workbookSync") or {})
        sync.update({
            "status": "app_changed",
            "warning": "Project changes are saved locally and have not been written to Excel. Click SAVE + WRITE EXCEL when ready.",
            "pendingReason": "project_ahead_of_workbook",
            "lastAuthorityAction": "opened_local_project",
        })
        opened["workbookSync"] = sync
        return opened
    if state == "not_linked":
        opened = dict(project)
        sync = dict(opened.get("workbookSync") or {})
        sync.update({
            "status": "not_linked",
            "warning": "The project is open locally, but no workbook is linked. Link it from Project Home before writing to Excel.",
        })
        opened["workbookSync"] = sync
        return opened
    if state == "review_required":
        raise WorkbookSyncError("Choose the matching workbook/project baseline from Project Home before opening the editor.")
    if state == "conflict":
        raise WorkbookSyncError("Both Excel and the local project changed after the last sync. Review the two versions from Project Home.")
    raise WorkbookSyncError(status.get("message") or "The linked workbook is unavailable.")





# S360 APP-ONLY SAVE CLASSIFIER V22.1
# App-only drawing edits stay in project.json. Excel is rewritten only when a
# workbook-backed field actually changes.

_S360_SAVE_GATES: dict[str, threading.Lock] = {}


def _s360_workbook_projection(project: dict[str, Any]) -> dict[str, Any]:
    pages: list[dict[str, Any]] = []
    for page in project.get("pages", []) or []:
        if not isinstance(page, dict):
            continue
        if (
            page.get("generatedContinuation")
            or page.get("continuationOf")
            or page.get("indexContinuation")
            or page.get("generatedIndexContinuation")
        ):
            continue
        pages.append(
            {
                "id": page.get("id"),
                "include": bool(page.get("include", True)),
                "notes": page.get("notes", ""),
                "renderProfile": page.get("renderProfile") or page.get("layoutProfile") or "",
                "splitMode": page.get("splitMode") or "",
                "parentPageId": page.get("parentPageId") or "",
                "issueStatus": page.get("issueStatus") or "draft",
                "sourceMode": page.get("sourceMode") or "",
                "syncDirection": page.get("syncDirection") or "",
            }
        )
    pages.sort(key=lambda item: str(item.get("id") or ""))

    worksheets: list[dict[str, Any]] = []
    for worksheet in project.get("worksheets", []) or []:
        if not isinstance(worksheet, dict):
            continue
        worksheets.append(
            {
                "id": worksheet.get("id"),
                "name": worksheet.get("name"),
                "grid": worksheet.get("grid", []),
                "formulas": worksheet.get("formulas", {}),
                "styles": worksheet.get("styles", {}),
                "mergedCells": worksheet.get("mergedCells", []),
                "rowHeights": worksheet.get("rowHeights", {}),
                "columnWidths": worksheet.get("columnWidths", {}),
                "rowHeightsPx": worksheet.get("rowHeightsPx", []),
                "colWidthsPx": worksheet.get("colWidthsPx", []),
                "defaultColumnWidth": worksheet.get("defaultColumnWidth"),
                "defaultRowHeight": worksheet.get("defaultRowHeight"),
                "hiddenRows": worksheet.get("hiddenRows", []),
                "hiddenColumns": worksheet.get("hiddenColumns", []),
                "geometryAuthority": worksheet.get("geometryAuthority"),
                "tabColor": worksheet.get("tabColor"),
                "printArea": worksheet.get("printArea"),
            }
        )
    worksheets.sort(
        key=lambda item: (
            str(item.get("id") or ""),
            str(item.get("name") or ""),
        )
    )
    return {"pages": pages, "worksheets": worksheets}


def _s360_workbook_projection_hash(project: dict[str, Any]) -> str:
    return sha256(
        json.dumps(
            _s360_workbook_projection(project),
            sort_keys=True,
            ensure_ascii=False,
            default=str,
        ).encode("utf-8")
    ).hexdigest()


def _s360_transient_workbook_error(exc: Exception) -> bool:
    message = str(exc).casefold()
    return any(
        token in message
        for token in (
            "open or locked",
            "being used by another process",
            "permission denied",
            "access is denied",
            "temporarily unavailable",
            "resource busy",
        )
    )


def save_local_then_try_sync(
    project_id: str,
    project: dict[str, Any],
    store: Any,
) -> dict[str, Any]:
    """Persist app-only work locally and write Excel only when required."""
    gate = _S360_SAVE_GATES.setdefault(project_id, threading.Lock())
    if not gate.acquire(timeout=180):
        raise WorkbookSyncError(
            "Another save is still finishing. Wait a moment and save again."
        )

    try:
        previous = store.load(project_id) or {}
        previous_projection = _s360_workbook_projection_hash(previous)
        incoming_projection = _s360_workbook_projection_hash(project)

        # Always protect the user's latest app work first.
        store.save(project_id, project)

        path, mode = configured_workbook_path(store, project_id, project)
        if path is None:
            raise WorkbookSyncError(
                "No authoritative workbook is linked to this project."
            )
        if not path.is_file():
            raise WorkbookSyncError(
                f"The linked workbook was not found: {path}"
            )

        if previous_projection == incoming_projection:
            # Symbols, legends, images, crops, annotations, and canvas work are
            # app-owned. Save them locally without rewriting the Excel workbook.
            meta = workbook_metadata(path)
            incoming_sync = dict(project.get("workbookSync") or {})
            previous_sync = dict(previous.get("workbookSync") or {})
            baseline_workbook = str(
                incoming_sync.get("workbookHash")
                or previous_sync.get("workbookHash")
                or ""
            )
            current_workbook = str(meta.get("sha256") or "")
            if baseline_workbook and current_workbook != baseline_workbook:
                raise WorkbookSyncError(
                    "The workbook changed after this editor session opened. "
                    "Reopen the project before continuing."
                )

            saved = dict(project)
            sync = dict(saved.get("workbookSync") or {})
            sync.update(
                {
                    "mode": sync.get("mode") or mode or "external-workbook-link",
                    "workbook": str(path),
                    "status": "in_sync",
                    "warning": "",
                    "pendingReason": "",
                    "runtimeLog": "",
                    "workbookHash": current_workbook,
                    "appHash": project_hash(saved),
                    "localProjectSavedAt": utcnow(),
                    "observedAt": utcnow(),
                    "lastAuthorityAction": "app_only_local_save",
                }
            )
            saved["workbookSync"] = sync
            store.save(project_id, saved)
            return saved

        status = status_payload(project_id, project, store)
        state = status["status"]

        if state in {"not_linked", "missing", "locked", "invalid", "project_mismatch"}:
            raise WorkbookSyncError(
                status.get("message")
                or "The authoritative workbook is unavailable."
            )
        if state in {"workbook_changed", "conflict"}:
            raise WorkbookSyncError(
                "The workbook changed after this editor session opened. "
                "Reopen the project before saving workbook-backed edits."
            )
        if state == "in_sync":
            saved = dict(project)
            saved["workbookSync"] = {
                **dict(project.get("workbookSync") or {}),
                "status": "in_sync",
                "warning": "",
                "observedAt": utcnow(),
                "appHash": project_hash(saved),
            }
            store.save(project_id, saved)
            return saved
        if state != "app_changed":
            raise WorkbookSyncError(
                status.get("message")
                or f"Workbook synchronization is blocked ({state})."
            )

        for attempt in range(1, 6):
            try:
                synced, _ = resolve(
                    project_id,
                    project,
                    store,
                    "app_to_workbook",
                )
                store.save(project_id, synced)
                return synced
            except Exception as exc:
                if not _s360_transient_workbook_error(exc) or attempt >= 5:
                    raise
                time.sleep(0.6 * attempt)

        raise WorkbookSyncError("Workbook synchronization did not complete.")
    except Exception as exc:
        log_path = _record_runtime_sync_failure(
            store,
            project_id,
            "required_write",
            exc,
        )
        pending = _pending_after_local_save(
            project,
            "workbook_write_required",
            f"Workbook save failed: {type(exc).__name__}: {exc}",
            log_path,
        )
        store.save(project_id, pending)
        if isinstance(exc, WorkbookSyncError):
            raise
        raise WorkbookSyncError(
            "The local recovery copy was saved, but the authoritative workbook "
            f"was not: {exc}"
        ) from exc
    finally:
        gate.release()
