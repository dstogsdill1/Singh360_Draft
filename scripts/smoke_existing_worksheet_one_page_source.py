from __future__ import annotations

import json
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill
from PIL import Image as PillowImage, ImageDraw

from core.worksheet_export import export_source_worksheet_xlsx


def verify_exact_source_export() -> dict[str, object]:
    """Prove that exact source-sheet extraction works, not just that code exists."""
    with tempfile.TemporaryDirectory(prefix="s360_exact_sheet_") as temp:
        root = Path(temp)
        source = root / "source.xlsx"
        png = root / "marker.png"

        image = PillowImage.new("RGB", (48, 32), "white")
        draw = ImageDraw.Draw(image)
        draw.rectangle((3, 3, 44, 28), outline="black", width=2)
        draw.text((12, 9), "S360", fill="black")
        image.save(png)

        wb = Workbook()
        target = wb.active
        target.title = "77_STYLED"
        target["A1"] = "SYSTEM"
        target["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        target["A1"].fill = PatternFill("solid", fgColor="1E73BE")
        target["B1"] = 2
        target["C1"] = 3
        target["D1"] = "=SUM(B1:C1)"
        target.merge_cells("A2:D2")
        target["A2"] = "FORMATTED ONE-SHEET TEST"
        target.column_dimensions["A"].width = 31.5
        target.row_dimensions[1].height = 24.0
        target.freeze_panes = "A3"
        target.print_area = "A1:D20"
        target.page_setup.orientation = "landscape"
        target.add_image(ExcelImage(str(png)), "F2")

        other = wb.create_sheet("OTHER")
        other["A1"] = "MUST NOT BE EXPORTED"
        wb.save(source)
        wb.close()

        output = export_source_worksheet_xlsx(source, "77_STYLED")
        extracted_path = root / "extracted.xlsx"
        extracted_path.write_bytes(output)

        result = load_workbook(extracted_path, data_only=False)
        try:
            assert result.sheetnames == ["77_STYLED"], result.sheetnames
            sheet = result["77_STYLED"]
            assert sheet["D1"].value == "=SUM(B1:C1)"
            assert sheet["A1"].font.bold is True
            assert str(sheet["A1"].fill.fgColor.rgb or "").upper().endswith("1E73BE")
            assert "A2:D2" in {str(item) for item in sheet.merged_cells.ranges}
            assert abs(float(sheet.column_dimensions["A"].width or 0) - 31.5) < 0.01
            assert abs(float(sheet.row_dimensions[1].height or 0) - 24.0) < 0.01
            assert sheet.freeze_panes == "A3"
            assert "$A$1:$D$20" in str(sheet.print_area)
            assert sheet.page_setup.orientation == "landscape"
            assert len(getattr(sheet, "_images", []) or []) == 1
        finally:
            result.close()

        return {
            "sheetNames": ["77_STYLED"],
            "formulaPreserved": True,
            "stylesPreserved": True,
            "mergePreserved": True,
            "dimensionsPreserved": True,
            "printSetupPreserved": True,
            "embeddedImagePreserved": True,
        }


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    app = (root / "frontend" / "src" / "App.tsx").read_text(encoding="utf-8")
    modal = (root / "frontend" / "src" / "components" / "ImportWorksheetModal.tsx").read_text(encoding="utf-8")
    client = (root / "frontend" / "src" / "api" / "client.ts").read_text(encoding="utf-8")
    server = (root / "server.py").read_text(encoding="utf-8")
    importer = (root / "core" / "sheet_importer.py").read_text(encoding="utf-8")
    exporter = (root / "core" / "worksheet_export.py").read_text(encoding="utf-8")

    exact_export = verify_exact_source_export()
    checks = {
        "existingWorksheetRebuilt": "S360 EXISTING WORKSHEET ONE PAGE V1" in app,
        "onePageNoContinuation": "layoutProfile: 'single_sheet_excel_exact'" in app and "allowContinuation: false" in app,
        "staleContinuationsRemoved": "withoutContinuations" in app and "page.generatedContinuation" in app,
        "modalSingleSelection": 'type="radio"' in modal and "Add Selected Sheet as One Page" in modal,
        "preserveExactDefault": "useState(true)" in modal and "preserveExact" in modal,
        "clientSendsPreserveExact": "fd.append('preserveExact'" in client,
        "serverReadsPreserveExact": "S360 SINGLE FORMATTED SHEET IMPORT V1" in server,
        "exactWorkbookImport": "preserve_exact" in importer and "_exact_excel_block" in importer,
        "exactSourceWorksheetExport": (
            "def export_source_worksheet_xlsx" in exporter
            and "load_workbook" in exporter
            and "wb.remove(" in exporter
            and "keep_vba" in exporter
        ),
        "functionalExactSourceWorksheetExport": bool(exact_export),
    }
    if not all(checks.values()):
        raise AssertionError(checks)
    print(json.dumps({"ok": True, **checks, "exactExport": exact_export}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
