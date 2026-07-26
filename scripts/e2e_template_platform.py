"""Live end-to-end acceptance workflow for the schema-V2 template platform."""
from __future__ import annotations

import io
import json
import sys
import tempfile
from pathlib import Path

import fitz
import requests
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw

BASE = "http://127.0.0.1:8766"
ROOT = Path(__file__).resolve().parents[1]


def expect(response: requests.Response, status: int = 200) -> requests.Response:
    if response.status_code != status:
        raise AssertionError(f"{response.request.method} {response.url}: {response.status_code} {response.text[:1000]}")
    return response


def make_fixtures(folder: Path) -> list[tuple[str, tuple[str, bytes, str]]]:
    pdf = fitz.open()
    page = pdf.new_page(width=792, height=612)
    page.insert_text((72, 72), "Sanitized Singh360 source fixture")
    pdf_bytes = pdf.tobytes()
    pdf.close()
    image = Image.new("RGB", (640, 360), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((40, 40, 600, 320), outline="#F47C20", width=8)
    draw.text((80, 160), "Sanitized source image", fill="black")
    image_buffer = io.BytesIO()
    image.save(image_buffer, "PNG")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SANITIZED_SOURCE"
    sheet.append(["Source", "Value"])
    sheet.append(["Fixture", 1])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)
    return [
        ("files", ("sanitized_source.pdf", pdf_bytes, "application/pdf")),
        ("files", ("sanitized_source.png", image_buffer.getvalue(), "image/png")),
        ("files", ("sanitized_source.xlsx", workbook_buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    ]


def main() -> int:
    health = expect(requests.get(f"{BASE}/api/health")).json()
    templates = expect(requests.get(f"{BASE}/api/workbook-templates")).json()["templates"]
    template = next(item for item in templates if item["templateId"] == "SINGH360_BASE_V1" and item["active"])
    result = expect(requests.post(f"{BASE}/api/projects/from-template", json={
        "profileId": "EMS_FULL", "templateId": template["templateId"],
        "metadata": {
            "projectName": "Template Platform E2E",
            "storeNumber": "TEST-001", "client": "Sanitized Test Client",
            "location": "Test Lab", "purpose": "Automated acceptance fixture",
            "scopeSummary": "Sanitized network controls acceptance fixture",
            "drawingPrefix": "TST", "revision": "0", "drawnBy": "E2E", "checkedBy": "E2E",
        },
    }), 201).json()
    project_id = result["id"]
    project = result["project"]
    project_dir = Path(project["projectFolder"])
    runtime_template = Path(template["absoluteRuntimePath"])
    project_workbook = project_dir / "sources" / "workbook" / runtime_template.name
    before_template_hash = template["sha256"]
    assert project_workbook.is_file() and project_workbook.resolve() != runtime_template.resolve()

    uploaded = expect(requests.post(f"{BASE}/api/projects/{project_id}/sources", files=make_fixtures(project_dir)), 201).json()["sources"]
    assert len(uploaded) == 3
    sources = expect(requests.get(f"{BASE}/api/projects/{project_id}/sources")).json()["sources"]
    assert {item["sourceType"] for item in sources} == {"pdf", "images", "spreadsheets"}

    document = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    network = next(sheet for sheet in document["sheets"] if sheet["name"] == "11_NETWORK_PORTS")
    network["cells"].update({
        "A1": {"v": "RDM / IDF NETWORK PORTS"}, "A2": {"v": "IDF"}, "B2": {"v": "Switch"},
        "C2": {"v": "Port"}, "D2": {"v": "Device"}, "A3": {"v": "IDF-TEST"},
        "B3": {"v": "SW-TEST"}, "C3": {"v": 1}, "D3": {"v": "Fixture Controller"},
    })
    network["styles"].update({
        "A1": {"fill": "FFF47C20", "bold": True, "fontColor": "FFFFFFFF"},
        "A2": {"fill": "FF262626", "bold": True, "fontColor": "FFFFFFFF"},
    })
    network["merges"] = ["A1:D1"]
    saved = expect(requests.put(f"{BASE}/api/projects/{project_id}/workbook", json={
        "projectId": project_id, "expectedRevision": document["revision"], "document": document,
    })).json()
    reloaded = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    assert reloaded["revision"] == saved["revision"] and reloaded["sheets"][document["sheets"].index(network)]["cells"]["D3"]["v"] == "Fixture Controller"
    conflict = requests.put(f"{BASE}/api/projects/{project_id}/workbook", json={
        "projectId": project_id, "expectedRevision": document["revision"], "document": document,
    })
    assert conflict.status_code == 409

    preview = expect(requests.post(f"{BASE}/api/projects/{project_id}/compile/preview")).json()
    assert any(item["family"] == "RDM / IDF Network Table" for item in preview["operations"])
    compiled = expect(requests.post(f"{BASE}/api/projects/{project_id}/compile/apply")).json()
    project = compiled["project"]
    network_page = next(page for page in project["pages"] if page["pageFamily"] == "RDM / IDF Network Table")
    manual = {"id": "manual-e2e-note", "type": "textbox", "text": "Manual E2E object", "left": 300, "top": 220}
    network_page["canvasObjects"].append(manual)
    expect(requests.post(f"{BASE}/api/projects/{project_id}", json=project))
    current = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    network = next(sheet for sheet in current["sheets"] if sheet["name"] == "11_NETWORK_PORTS")
    network["cells"]["D3"] = {"v": "Updated Fixture Controller"}
    expect(requests.put(f"{BASE}/api/projects/{project_id}/workbook", json={
        "projectId": project_id, "expectedRevision": current["revision"], "document": current,
    }))
    second = expect(requests.post(f"{BASE}/api/projects/{project_id}/compile/apply")).json()["project"]
    updated_page = next(page for page in second["pages"] if page["id"] == network_page["id"])
    assert manual in updated_page["canvasObjects"]
    assert updated_page["blocks"][0]["rows"][0][3] == "Updated Fixture Controller"

    mirror = expect(requests.post(f"{BASE}/api/projects/{project_id}/workbook/write-excel")).json()
    workbook = load_workbook(project_workbook)
    assert workbook["00_PROJECT_META"]["B3"].value == project_id
    assert workbook["11_NETWORK_PORTS"]["D3"].value == "Updated Fixture Controller"
    assert workbook.sheetnames.index("00_PROJECT_META") < workbook.sheetnames.index("11_NETWORK_PORTS")
    workbook.close()

    output_dir = project_dir / "exports"
    pdf_response = expect(requests.post(f"{BASE}/api/projects/{project_id}/export/pdf", json={"width": 17, "height": 11}))
    pdf_path = output_dir / "pdf" / "e2e_download.pdf"
    pdf_path.write_bytes(pdf_response.content)
    assert pdf_response.content.startswith(b"%PDF") and len(pdf_response.content) > 1000
    package_response = expect(requests.post(f"{BASE}/api/projects/{project_id}/export/package"))
    package_path = output_dir / "package" / "e2e_package.zip"
    package_path.write_bytes(package_response.content)
    assert package_response.content.startswith(b"PK") and len(package_response.content) > 1000

    report = {
        "health": health, "projectId": project_id, "projectFolder": str(project_dir),
        "runtimeTemplate": str(runtime_template), "runtimeTemplateHash": before_template_hash,
        "runtimeTemplateUnchanged": template["sha256"] == before_template_hash,
        "sourceCount": len(sources), "workbookRevision": saved["revision"],
        "conflictStatus": conflict.status_code, "compileBackup": compiled["backupPath"],
        "manualObjectPreserved": manual in updated_page["canvasObjects"],
        "mirror": mirror, "sheetOrder": load_workbook(project_workbook, read_only=True).sheetnames,
        "pdfPath": str(pdf_path), "pdfBytes": pdf_path.stat().st_size,
        "packagePath": str(package_path), "packageBytes": package_path.stat().st_size,
    }
    report_path = project_dir / "debug" / "template_platform_e2e_report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
