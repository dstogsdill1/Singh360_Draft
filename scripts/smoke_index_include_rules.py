"""Smoke: workbook index include/exclude is law — no output page when include=NO."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _mini_workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "COVER", "Cover", "YES", ""])
    idx.append(["2", "00_INDEX", "Sheet Index", "YES", ""])
    idx.append(["3", "OPTIONAL EMS", "Optional Tab", "NO", "disabled in app"])
    idx.append(["4", "SCOPE", "Project Scope", "YES", ""])

    cover = wb.create_sheet("COVER")
    cover["A1"] = "Project Cover"
    for r in range(1, 6):
        idx.cell(r + 6, 1, f"Row {r}")
        idx.cell(r + 6, 2, f"Tab {r}")
        idx.cell(r + 6, 3, f"Title {r}")
        idx.cell(r + 6, 4, "YES")

    wb.create_sheet("OPTIONAL EMS")["A1"] = "Should not output"

    scope = wb.create_sheet("SCOPE")
    scope["A1"] = "Item"
    scope["B1"] = "Status"
    scope["A2"] = "Closeout"
    scope["B2"] = "Complete"

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "include_rules.xlsx"
    _mini_workbook(xlsx)
    proj = import_workbook(xlsx, project_id="inc1")

    ws_names = {w["name"] for w in proj["worksheets"]}
    page_tabs = {p["sheetTab"] for p in proj["pages"]}
    problems: list[str] = []

    if "OPTIONAL EMS" not in ws_names:
        problems.append("optional worksheet missing from source tabs")
    if "OPTIONAL EMS" in page_tabs:
        problems.append("include=NO sheet created an output page")
    if "SCOPE" not in page_tabs:
        problems.append("include=YES scope sheet missing from output")
    if any(p.get("generatedContinuation") for p in proj["pages"] if p["sheetTab"] == "00_INDEX"):
        problems.append("index sheet got continuation pages")

    index_pages = [p for p in proj["pages"] if p["sheetTab"] == "00_INDEX"]
    if len(index_pages) != 1:
        problems.append(f"index expected 1 page, got {len(index_pages)}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — index include/exclude rules passed")
    print(f"  worksheets: {len(ws_names)}, output pages: {len(proj['pages'])}")


if __name__ == "__main__":
    main()
