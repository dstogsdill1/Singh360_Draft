"""Smoke: workbook import preserves source cell fills in exact worksheet blocks.

Builds a tiny .xlsx in-memory with a highlighted cell, imports it, and asserts a
renderable block carries the fill in its style map. No customer files needed.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402

from core.workbook_importer import import_workbook  # noqa: E402


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "RACK A IO Schedule"
    headers = ["RO#", "Relay Output Description", "Type", "DI#", "Status Input"]
    ws.append(headers)
    ws.append(["R1", "Compressor 1", "DO", "D1", "Status 1"])
    ws.append(["R2", "Compressor 2", "DO", "D2", "Status 2"])
    ws.append(["R3", "Condenser Fan", "DO", "D3", "Status 3"])
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["B2"].fill = yellow  # highlight one body cell

    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "style_import.xlsx"
        wb.save(path)
        project = import_workbook(path, project_id="styleimport00001")

    problems: list[str] = []
    pages = project.get("pages", [])
    table_blocks = [
        b
        for p in pages
        for b in (p.get("blocks") or [])
        if b.get("type") in ("table", "matrix", "excelRange")
    ]
    if not table_blocks:
        problems.append("no table/matrix block produced from workbook")
    fills: dict[str, str] = {}
    for b in table_blocks:
        if b.get("cellFills"):
            fills.update(b["cellFills"])
        for key, style in (b.get("styles") or {}).items():
            if isinstance(style, dict) and style.get("fill"):
                fills[key] = style["fill"]
    if not fills:
        problems.append("source fill not captured in block style data")
    else:
        has_yellow = any(str(v).upper().endswith("FFFF00") or str(v).upper() == "#FFFF00" for v in fills.values())
        if not has_yellow:
            problems.append(f"expected a yellow fill in block styles, got {fills}")

    print(f"renderableBlocks={len(table_blocks)} fills={fills}")
    if problems:
        print("WORKBOOK STYLE IMPORT PROBLEMS:")
        for p in problems:
            print(f"  - {p}")
        return 1
    print("OK: workbook style import smoke passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
