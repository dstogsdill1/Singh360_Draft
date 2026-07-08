"""Smoke: title-block SHEET NO. uses canonical index sheet codes (4I Phase B)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    # Order intentionally disagrees with EMS numbers so a regression that
    # uses Order as SHEET NO. cannot pass.
    idx.append(["YES", 1, "EMS 0.0", "Cover", "Cover / Project Info", "Front Matter", "cover", ""])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""])
    idx.append(["YES", 5, "EMS 0.4", "Scope", "Project Scope", "Front Matter", "text", ""])
    idx.append(["YES", 12, "EMS 1.1", "RDM IDF", "RDM / IDF Network Table", "Network", "table", ""])
    idx.append(["YES", 15, "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "Lighting", "io-table", ""])

    cover = wb.create_sheet("Cover")
    cover.cell(1, 1, "COVER")

    scope = wb.create_sheet("Scope")
    scope.append(["Section", "Scope Language", "Status", "Notes"])
    scope.append(["Executive Summary", "Provide EMS documentation and commissioning.", "Review", ""])

    idf = wb.create_sheet("RDM IDF")
    idf.append(["Port", "Label", "Device / Drop", "From", "To", "Cable", "Notes", "Controller ID", "IP Address"])
    for p in range(1, 49):
        idf.append([str(p), f"L{p}", f"D{p}", "IDF", f"Drop{p}", "CAT6", "", "C1", f"10.0.0.{p}"])

    lcp = wb.create_sheet("LCP Panel Schedule")
    lcp.append(["RO#", "Description", "Type", "DI#", "Status", "Notes"])
    for i in range(55):
        lcp.append([str(i + 1), f"Relay {i}", "NO", str(i + 1), "OK", ""])
        lcp.row_dimensions[i + 2].height = 30

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "sheet_codes.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="sheetcode4i")
    pages = proj["pages"]
    problems: list[str] = []

    expected = {
        "Cover": "EMS 0.0",
        "00_INDEX": "EMS 0.1",
        "Scope": "EMS 0.4",
        "RDM IDF": "EMS 1.1",
        "LCP Panel Schedule": "EMS 1.4",
    }
    for tab, code in expected.items():
        matches = [p for p in pages if p["sheetTab"] == tab]
        if not matches:
            problems.append(f"{tab}: missing")
            continue
        p = matches[0]
        got = p.get("displaySheetCode") or p.get("sheetCode")
        if got != code:
            problems.append(f"{tab}: title-block sheet code={got!r}, expected {code!r}")
        # Never allow bare Order-as-code like "5.0".
        if got and got.replace(".", "").isdigit():
            problems.append(f"{tab}: bare numeric sheet code {got!r} (Order leaked into SHEET NO.)")

    lcp_pages = [p for p in pages if p["sheetTab"] == "LCP Panel Schedule"]
    if len(lcp_pages) < 2:
        problems.append(f"LCP expected continuation, got {len(lcp_pages)}")
    else:
        cont = lcp_pages[1]
        cont_code = cont.get("displaySheetCode") or cont.get("sheetCode")
        if cont_code != "EMS 1.4a":
            problems.append(f"LCP continuation code={cont_code!r}, expected EMS 1.4a")

    # Sheet X of Y stays physical order.
    included = [p for p in pages if p.get("include", True)]
    orders = [p.get("order") for p in included]
    if orders != list(range(1, len(included) + 1)) and orders != sorted(orders):
        problems.append(f"physical order broken: {orders}")
    totals = {p.get("pageTotal") for p in included}
    if totals != {len(included)}:
        problems.append(f"pageTotal mismatch: {totals}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA31 sheet code consistency passed")


if __name__ == "__main__":
    main()
