from __future__ import annotations

import json
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# The installer runs this test inside the full repository.  The standalone ZIP
# validation has only the changed payload files, so provide narrow test stubs
# only when the repository dependencies are genuinely unavailable.
try:
    from core.sheet_importer import import_workbook_sheets, preview_workbook_sheets
except ModuleNotFoundError as exc:
    if exc.name not in {"core.project_model", "core.page_normalizer", "core.page_composer", "core.workbook_importer"}:
        raise
    import re
    import sys
    import types
    from openpyxl.utils import get_column_letter

    project_model = types.ModuleType("core.project_model")
    project_model.classify_page_type = lambda tab, title, source="": (
        "cover" if "cover" in f"{tab} {title}".lower()
        else "index" if "index" in f"{tab} {title}".lower()
        else "data-grid"
    )
    project_model.sanitize_json = lambda value: value
    sys.modules["core.project_model"] = project_model

    page_normalizer = types.ModuleType("core.page_normalizer")
    page_normalizer.normalize_page = lambda ws, ws_id, page_type, title: []
    sys.modules["core.page_normalizer"] = page_normalizer

    page_composer = types.ModuleType("core.page_composer")
    page_composer.page_family = lambda tab, title, source="": "table"
    sys.modules["core.page_composer"] = page_composer

    workbook_importer = types.ModuleType("core.workbook_importer")
    def _parse_index(wb, name):
        ws = wb[name]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        header = [str(value or "").strip().lower() for value in rows[0]]
        cols = {value: index for index, value in enumerate(header)}
        out = []
        for row in rows[1:]:
            tab = str(row[cols["sheet tab"]] or "").strip()
            if not tab:
                continue
            out.append({
                "sheetTab": tab,
                "sheetTitle": str(row[cols["sheet title"]] or tab),
                "sheetCodeRaw": str(row[cols["sheet code"]] or ""),
                "notes": str(row[cols["notes"]] or "") if "notes" in cols else "",
            })
        return out
    def _style(cell):
        result = {}
        if cell.font.bold: result["bold"] = True
        if cell.font.size: result["fontSize"] = float(cell.font.size)
        if cell.font.color and cell.font.color.type == "rgb" and cell.font.color.rgb:
            result["fontColor"] = "#" + str(cell.font.color.rgb)[-6:]
        if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor.type == "rgb" and cell.fill.fgColor.rgb:
            result["fill"] = "#" + str(cell.fill.fgColor.rgb)[-6:]
        if cell.alignment.wrap_text: result["wrap"] = True
        return result
    def _worksheet_payload(ws, ws_values=None):
        max_row = ws.max_row or 0; max_col = ws.max_column or 0
        grid=[]; styles={}
        for r in range(1,max_row+1):
            row=[]
            for c in range(1,max_col+1):
                cell=ws.cell(r,c); row.append("" if cell.value is None else str(cell.value))
                st=_style(cell)
                if st: styles[f"{get_column_letter(c)}{r}"]=st
            grid.append(row)
        merges=[{
            "startRow": rng.min_row-1, "startCol": rng.min_col-1,
            "endRow": rng.max_row-1, "endCol": rng.max_col-1,
        } for rng in ws.merged_cells.ranges]
        col_widths=[]
        for c in range(1,max_col+1):
            width=ws.column_dimensions[get_column_letter(c)].width or 13
            col_widths.append(int(round(width*7)))
        row_heights=[]
        for r in range(1,max_row+1):
            height=ws.row_dimensions[r].height or 15
            row_heights.append(int(round(height*4/3)))
        return {
            "name": ws.title, "grid": grid, "formulas": {}, "styles": styles,
            "mergedCells": merges, "rowHeights": {}, "columnWidths": {},
            "colWidthsPx": col_widths, "rowHeightsPx": row_heights,
            "sourceSheet": ws.title,
            "sourceRange": f"A1:{get_column_letter(max(max_col,1))}{max(max_row,1)}",
            "printArea": str(ws.print_area or ""),
        }
    workbook_importer._parse_index = _parse_index
    workbook_importer._worksheet_payload = _worksheet_payload
    workbook_importer._extract_embedded_images = lambda *args, **kwargs: []
    workbook_importer._safe_name = lambda value: re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    sys.modules["core.workbook_importer"] = workbook_importer

    from core.sheet_importer import import_workbook_sheets, preview_workbook_sheets


def main() -> int:
    root = Path(tempfile.mkdtemp(prefix="s360_single_sheet_"))
    xlsx = root / "styled.xlsx"
    wb = Workbook()
    index = wb.active
    index.title = "00_INDEX"
    index.append(["Order", "Sheet Code", "Sheet Tab", "Sheet Title", "Include", "Notes"])
    index.append([1, "EMS 77.0", "77_STYLED", "Styled Added Sheet", "YES", "Imported exactly"])
    ws = wb.create_sheet("77_STYLED")
    ws["A1"] = "STYLED PAGE"
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:D1")
    ws["A3"] = "Item"; ws["B3"] = "Description"
    ws["A4"] = 1; ws["B4"] = "Already formatted in Excel"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.row_dimensions[1].height = 28
    ws.print_area = "A1:D20"
    wb.save(xlsx)

    preview = preview_workbook_sheets(xlsx)
    row = next(item for item in preview if item["sheetName"] == "77_STYLED")
    assert row["sheetCode"] == "EMS 77.0", row
    assert row["pageTitle"] == "Styled Added Sheet", row
    assert row["listedInIndex"] is True, row

    project = {"id": "test", "sources": [], "worksheets": [], "pages": [{"id": "existing", "order": 1, "include": True, "sheetCode": "EMS 1.0", "displaySheetCode": "EMS 1.0", "sheetTitle": "Existing", "pageType": "data-grid", "blocks": [], "canvasObjects": []}]}
    project, pages = import_workbook_sheets(project, xlsx, ["77_STYLED"], insert_after_page_id="existing", preserve_exact=True, source_filename=xlsx.name)
    assert len(pages) == 1, pages
    page = pages[0]
    assert page["sheetCode"] == "EMS 77.0", page
    assert page["sheetTitle"] == "Styled Added Sheet", page
    assert page["renderMode"] == "excel_exact", page
    assert page["splitMode"] == "none" and page["allowContinuation"] is False, page
    assert len(page["blocks"]) == 1 and page["blocks"][0]["type"] == "excelRange", page
    block = page["blocks"][0]
    assert block["styles"].get("A1", {}).get("fill") == "#1F4E78", block["styles"].get("A1")
    assert block["mergedCells"] == [{"startRow": 0, "startCol": 0, "endRow": 0, "endCol": 3}], block["mergedCells"]
    assert block["sourceRange"].startswith("PRINT_AREA:"), block["sourceRange"]
    assert [p["id"] for p in project["pages"]][:2] == ["existing", page["id"]]
    print(json.dumps({"ok": True, "pageId": page["id"], "sheetCode": page["sheetCode"], "renderMode": page["renderMode"], "printArea": block["sourceRange"]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
