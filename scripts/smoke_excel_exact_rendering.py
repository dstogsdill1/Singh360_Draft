"""Smoke: tabular pages render excelRange blocks, not generic table rebuilds."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Side, Border

from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["2", "00_INDEX", "Sheet Index", "YES", ""])
    idx.append(["7", "RESP MATRIX", "Responsibility Matrix", "YES", ""])
    idx.append(["15", "LCP Panel", "LCP Panel Schedule", "YES", ""])
    idx.append(["5", "SCOPE", "Project Scope", "YES", ""])

    for r in range(8, 12):
        idx.append([str(r), f"TAB{r}", f"Title {r}", "YES", ""])

    rm = wb.create_sheet("RESP MATRIX")
    rm.merge_cells("B1:D1")
    rm["B1"] = "Furnished by"
    rm["B1"].fill = GOLD
    for col, name in zip("BCD", ["Owner", "EC", "EMS"]):
        c = rm[f"{col}2"]
        c.value = name
        c.fill = GRAY
        c.font = Font(bold=True)
        c.border = BOX
    rm["A3"] = "Controls"
    rm["B3"] = "X"

    lcp = wb.create_sheet("LCP Panel")
    lcp.merge_cells("A1:E1")
    lcp["A1"] = "LCP-1"
    lcp["A1"].fill = GOLD
    for i, h in enumerate(["Point", "Type", "Addr", "Desc", "Notes"], start=1):
        c = lcp.cell(2, i, h)
        c.fill = GRAY
        c.font = Font(bold=True)

    scope = wb.create_sheet("SCOPE")
    scope["A1"] = "Item"
    scope["B1"] = "Notes"
    scope["A2"] = "Closeout"
    scope["B2"] = "Final deliverables"

    wb.save(path)


def _assert_excel_exact(page: dict, label: str, problems: list[str]) -> None:
    if page.get("renderMode") != "excel_exact":
        problems.append(f"{label}: renderMode not excel_exact")
    blocks = page.get("blocks") or []
    if not blocks or blocks[0].get("type") != "excelRange":
        problems.append(f"{label}: missing excelRange block")
    elif blocks[0].get("type") in ("table", "matrix"):
        problems.append(f"{label}: generic table renderer used")


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "exact.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="exact1")
    problems: list[str] = []

    by_tab = {p["sheetTab"]: p for p in proj["pages"]}
    for tab in ("00_INDEX", "RESP MATRIX", "LCP Panel", "SCOPE"):
        if tab not in by_tab:
            problems.append(f"missing output page for {tab}")
            continue
        _assert_excel_exact(by_tab[tab], tab, problems)

    rm_block = by_tab["RESP MATRIX"]["blocks"][0]
    if not (rm_block.get("styles") or {}):
        problems.append("responsibility matrix lost per-cell styles")

    lcp_block = by_tab["LCP Panel"]["blocks"][0]
    styles = lcp_block.get("styles") or {}
    gold = any(v.get("fill") == "#FFC000" for v in styles.values())
    if not gold:
        problems.append("LCP panel lost gold controller band fill")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — excel exact rendering passed")


if __name__ == "__main__":
    main()
