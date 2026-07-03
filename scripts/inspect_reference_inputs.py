"""scripts/inspect_reference_inputs.py — local-only reference fixture scanner.

Reads reference file paths from environment variables and prints a summary of
each (existence, workbook sheets/ranges, CSV headers/rows/categories, PDF page
counts). Never fails on missing files — warns and skips. No customer data is
written anywhere.

Env vars:
  SINGH360_REF_SA31_PDF
  SINGH360_REF_SA38_PDF
  SINGH360_TEMPLATE_WORKBOOK
  SINGH360_CARTHAGE_WORKBOOK
  SINGH360_KATY_CSV
  SINGH360_SA31_WORKBOOK
"""
from __future__ import annotations

import csv
import os
from collections import Counter
from pathlib import Path

_WORKBOOK_VARS = ["SINGH360_TEMPLATE_WORKBOOK", "SINGH360_CARTHAGE_WORKBOOK", "SINGH360_SA31_WORKBOOK"]
_CSV_VARS = ["SINGH360_KATY_CSV"]
_PDF_VARS = ["SINGH360_REF_SA31_PDF", "SINGH360_REF_SA38_PDF"]


def _resolve(var: str) -> Path | None:
    val = os.environ.get(var, "").strip()
    if not val:
        print(f"  [skip] {var} not set")
        return None
    p = Path(val).expanduser()
    if not p.exists():
        print(f"  [missing] {var} -> {p}")
        return None
    print(f"  [ok] {var} -> {p.name}")
    return p


def _inspect_workbook(path: Path) -> None:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        print(f"    sheets ({len(wb.sheetnames)}):")
        for name in wb.sheetnames:
            ws = wb[name]
            dims = ws.calculate_dimension() if ws.max_row else "empty"
            print(f"      - {name}  [{dims}]")
        wb.close()
    except Exception as exc:  # noqa: BLE001
        print(f"    (could not read workbook: {exc})")


def _inspect_csv(path: Path) -> None:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
        if not rows:
            print("    (empty CSV)")
            return
        headers = rows[0]
        print(f"    headers ({len(headers)}): {headers}")
        print(f"    data rows: {max(0, len(rows) - 1)}")
        # Category counts if a Category column exists
        cat_idx = next((i for i, h in enumerate(headers) if h.strip().lower() == "category"), -1)
        if cat_idx >= 0:
            counts = Counter(
                (r[cat_idx].strip() if cat_idx < len(r) else "") or "(blank)"
                for r in rows[1:]
            )
            print("    category counts:")
            for cat, n in counts.most_common():
                print(f"      - {cat}: {n}")
    except Exception as exc:  # noqa: BLE001
        print(f"    (could not read CSV: {exc})")


def _inspect_pdf(path: Path) -> None:
    try:
        data = path.read_bytes()
        # Cheap page-count heuristic without a PDF dependency.
        count = data.count(b"/Type /Page") + data.count(b"/Type/Page")
        print(f"    approx pages: {count if count else 'unknown'} ({len(data) // 1024} KB)")
    except Exception as exc:  # noqa: BLE001
        print(f"    (could not read PDF: {exc})")


def main() -> int:
    print("Singh360 Draft — reference input inspector")
    print("=" * 60)

    print("Workbooks:")
    for var in _WORKBOOK_VARS:
        p = _resolve(var)
        if p:
            _inspect_workbook(p)

    print("CSV:")
    for var in _CSV_VARS:
        p = _resolve(var)
        if p:
            _inspect_csv(p)

    print("PDFs:")
    for var in _PDF_VARS:
        p = _resolve(var)
        if p:
            _inspect_pdf(p)

    print("=" * 60)
    print("Done. Missing files are skipped, not errors.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
