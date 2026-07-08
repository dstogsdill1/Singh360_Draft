"""Smoke: generated continuation pages are reflected in the rendered Sheet
Index (FINAL RELEASE CLEANUP 4H+SA38, Phase E).

Uses an oversized LCP-style panel schedule (same shape as
``smoke_sheet_code_mapping.py``) that forces a continuation page
(``EMS 1.4`` -> ``EMS 1.4a``). Verifies:
  - The rendered index grid contains a row for ``EMS 1.4a`` with
    "CONTINUED" in the title.
  - The base ``EMS 1.4`` row is not duplicated.
  - ``pageTotal`` equals the actual number of composed (included) pages.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "Cover", "Cover Sheet", "Front Matter", "cover", ""])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""])
    idx.append(["YES", 3, "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "Lighting", "matrix", ""])

    cover = wb.create_sheet("Cover")
    cover.cell(1, 1, "SINGH360 EMS PROJECT")

    lcp = wb.create_sheet("LCP Panel Schedule")
    lcp.append(["RO#", "Description", "Type", "DI#", "Status", "Notes"])
    for i in range(60):
        lcp.append([str(i + 1), f"Relay {i}", "NO", str(i + 1), "OK", ""])
        lcp.row_dimensions[i + 2].height = 30

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "index_continuations.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="indexcont1")
    pages = proj["pages"]
    problems: list[str] = []

    lcp_pages = [p for p in pages if p["sheetTab"] == "LCP Panel Schedule"]
    if len(lcp_pages) < 2:
        problems.append(f"expected LCP Panel Schedule to split into a continuation, got {len(lcp_pages)} page(s)")

    index_page = next((p for p in pages if p["pageType"] == "index"), None)
    if index_page is None:
        problems.append("index page not found")
    else:
        block = next((b for b in (index_page.get("blocks") or []) if b.get("type") == "excelRange"), None)
        if block is None:
            problems.append("index page has no excelRange block (renderMode not excel_exact?)")
        else:
            grid = block.get("grid") or []
            cont_rows = [r for r in grid if any("continued" in str(c).lower() for c in r)]
            cont_1_4a_rows = [r for r in cont_rows if any("1.4a" in str(c).lower() for c in r)]
            if not cont_1_4a_rows:
                problems.append(f"no 'EMS 1.4a ... CONTINUED' row found in rendered index grid; grid={grid}")
            base_rows = [r for r in grid if any(str(c).strip() == "EMS 1.4" for c in r)]
            if len(base_rows) != 1:
                problems.append(f"expected exactly 1 base 'EMS 1.4' row in index, found {len(base_rows)}")
            # rowHeights must stay in sync with the appended grid rows.
            if len(block.get("rowHeights") or []) != len(grid):
                problems.append(f"rowHeights length {len(block.get('rowHeights') or [])} != grid length {len(grid)}")

    included = [p for p in pages if p.get("include", True)]
    total = len(included)
    bad_totals = [p.get("pageTotal") for p in pages if p.get("pageTotal") != total]
    if bad_totals:
        problems.append(f"pageTotal mismatch: expected {total} for all pages, saw {set(bad_totals)}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — continuation rows appended to rendered Sheet Index")
    print(f"  LCP pages={len(lcp_pages)}, index grid rows={len(block.get('grid') or [])}, pageTotal={total}")


if __name__ == "__main__":
    main()
