"""Smoke: every non-cover page carries the Singh360 standard table profile.

Asserts renderProfile = "singh360_standard_table" and normalizedHeaderStyle =
"orange" on non-cover pages, and that apply_singh360_profile recolors a dark
source title band to orange (#FFC000) with black text and a header row to gray
(#D9D9D9) while preserving gold controller bands.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.table_style_profile import (
    COLUMN_HEADER_FILL,
    RENDER_PROFILE,
    TITLE_BAND_FILL,
    apply_singh360_profile,
    is_gold_fill,
)
from core.workbook_importer import import_workbook


def _fixture(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "LCP", "LCP Panel Schedule", "YES", ""])

    ws = wb.create_sheet("LCP")
    # Row 1: dark title band (should become orange).
    t = ws.cell(1, 1, "LCP PANEL SCHEDULE")
    t.fill = PatternFill("solid", fgColor="1F1F1F")
    t.font = Font(bold=True, color="FFFFFF")
    ws.merge_cells("A1:D1")
    # Row 2: gray-ish column header (should become gray standard).
    for c, h in enumerate(["Point", "Type", "Address", "Notes"], start=1):
        cell = ws.cell(2, c, h)
        cell.fill = PatternFill("solid", fgColor="808080")
        cell.font = Font(bold=True)
    # Row 3: gold controller band (should be preserved).
    g = ws.cell(3, 1, "CONTROLLER LCP-1")
    g.fill = PatternFill("solid", fgColor="FFC000")
    g.font = Font(bold=True)
    ws.merge_cells("A3:D3")
    for r in range(4, 12):
        for c in range(1, 5):
            ws.cell(r, c, f"RO{r}C{c}")
    wb.save(path)


def main() -> None:
    problems: list[str] = []

    # 1) Direct profile behavior on a synthetic block.
    block = {
        "type": "excelRange",
        "grid": [
            ["LCP PANEL SCHEDULE", "", "", ""],
            ["Point", "Type", "Address", "Notes"],
            ["CONTROLLER LCP-1", "", "", ""],
            ["RO1", "DO", "1", "x"],
        ],
        "styles": {
            "0:0": {"fill": "#1F1F1F", "fontColor": "#FFFFFF", "bold": True},
            "1:0": {"fill": "#808080", "bold": True},
            "1:1": {"fill": "#808080", "bold": True},
            "1:2": {"fill": "#808080", "bold": True},
            "1:3": {"fill": "#808080", "bold": True},
            "2:0": {"fill": "#FFC000", "bold": True},
        },
        "repeatRows": [0, 1],
        "headerRowCount": 2,
    }
    apply_singh360_profile(block, "orange")
    st = block["styles"]
    if st["0:0"].get("fill") != TITLE_BAND_FILL:
        problems.append(f"dark title band not recolored orange: {st['0:0'].get('fill')}")
    if st["0:0"].get("fontColor") != "#000000":
        problems.append("title band text not black")
    if st["1:0"].get("fill") != COLUMN_HEADER_FILL:
        problems.append(f"header row not recolored gray: {st['1:0'].get('fill')}")
    if not is_gold_fill(st["2:0"].get("fill")):
        problems.append("gold controller band was not preserved")

    # 2) Import-level: page carries profile + header style.
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "profile.xlsx"
    _fixture(xlsx)
    proj = import_workbook(xlsx, project_id="prof1")
    lcp = [p for p in proj["pages"] if p["sheetTab"] == "LCP"]
    if not lcp:
        problems.append("LCP page missing from import")
    else:
        p = lcp[0]
        if p.get("renderProfile") != RENDER_PROFILE:
            problems.append(f"page renderProfile={p.get('renderProfile')}")
        if p.get("normalizedHeaderStyle") != "orange":
            problems.append(f"page normalizedHeaderStyle={p.get('normalizedHeaderStyle')}")
        xr = next((b for b in p["blocks"] if b.get("type") == "excelRange"), None)
        if xr:
            fills = [s.get("fill") for s in (xr.get("styles") or {}).values()]
            if TITLE_BAND_FILL not in fills:
                problems.append("imported LCP block has no orange title band")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — singh360 standard table style applied")


if __name__ == "__main__":
    main()
