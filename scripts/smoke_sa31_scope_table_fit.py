"""Smoke: Project Scope uses front_matter_narrative_table fit (4I Phase C)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.page_composer import BODY_W
from core.workbook_importer import import_workbook

GRAY = PatternFill("solid", fgColor="D9D9D9")


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 5, "EMS 0.4", "EMS 0.4 Scope", "Project Scope", "Front Matter", "text", ""])

    scope = wb.create_sheet("EMS 0.4 Scope")
    scope.append(["PROJECT SCOPE — LIGHT DIMMING", "", "", ""])
    scope.append(["", "", "", ""])
    headers = ["Section", "Scope Language", "Status", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = scope.cell(3, c, h)
        cell.fill = GRAY
        cell.font = Font(bold=True)
    rows = [
        ("Executive Summary & Scope of Work",
         "Singh360 is providing EMS controls documentation, preconfigured lighting-control hardware "
         "coordination, remote programming, and commissioning support for the light dimming project.",
         "Review", ""),
        ("Project Intent",
         "General contractor-led lighting fixture replacement will be coordinated with Singh360-controlled "
         "lighting/dimming upgrades for LCP hardware and RDM network integration.",
         "Review", ""),
        ("LCP-1 Dimming Scope",
         "LCP-1 supports the localized dimming control loop and dimming modules. Each dimming zone must "
         "be wired back to designated terminal points and labeled to match the workbook.",
         "Review", ""),
        ("Closeout",
         "Closeout requires final as-built workbook/PDF, updated output matrix, and commissioning handoff.",
         "Review", ""),
    ]
    for i, (sec, lang, status, notes) in enumerate(rows, start=4):
        scope.cell(i, 1, sec)
        scope.cell(i, 2, lang)
        scope.cell(i, 3, status)
        scope.cell(i, 4, notes)
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "scope_fit.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="scopefit1")
    problems: list[str] = []

    scope = next((p for p in proj["pages"] if "Scope" in (p.get("sheetTitle") or "")), None)
    if scope is None:
        problems.append("Project Scope page missing")
    else:
        if scope.get("layoutProfile") != "front_matter_narrative_table":
            problems.append(f"layoutProfile={scope.get('layoutProfile')!r}, expected front_matter_narrative_table")
        if scope.get("renderProfile") != "front_matter_narrative_table":
            problems.append(f"renderProfile={scope.get('renderProfile')!r}, expected front_matter_narrative_table")
        block = next((b for b in (scope.get("blocks") or []) if b.get("type") == "excelRange"), None)
        if block is None:
            problems.append("scope excelRange missing")
        else:
            widths = block.get("colWidths") or []
            total_w = sum(widths)
            lo, hi = int(BODY_W * 0.85), int(BODY_W * 0.95)
            if not (lo <= total_w <= hi + 40):
                problems.append(f"scope table width {total_w} not in ~85–95% of BODY_W ({lo}-{hi})")
            if widths:
                section_share = widths[0] / total_w if total_w else 0
                if section_share < 0.18 or section_share > 0.30:
                    problems.append(f"Section column share {section_share:.2f} outside 20–24% band (with tolerance)")
            if not block.get("nowrapColumns"):
                problems.append("nowrapColumns missing — Section labels may stack word-by-word")
            notes_w = widths[3] if len(widths) > 3 else None
            if notes_w is not None and notes_w > int(BODY_W * 0.15):
                problems.append(f"Notes column still too wide ({notes_w}px) despite empty notes")
            font_pt = float(block.get("bodyFontPt") or 0)
            if font_pt and font_pt < 7.5:
                problems.append(f"bodyFontPt {font_pt} below 7.5 floor")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA31 scope table fit / narrative profile passed")
    print(f"  widths={block.get('colWidths')}, profile={scope.get('layoutProfile')}")


if __name__ == "__main__":
    main()
