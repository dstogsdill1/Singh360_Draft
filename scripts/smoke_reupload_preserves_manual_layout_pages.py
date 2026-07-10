"""Smoke: workbook re-upload preserves manual layout canvas objects (Phase E)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook
from core.workbook_reimport import apply_reimport


def _workbook_v1(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title"])
    idx.append(["YES", 1, "EMS 12.0", "Overall Layout", "EMS Controls Overall Layout"])
    idx.append(["YES", 2, "EMS 13.0", "SCOPE", "Project Scope"])

    layout = wb.create_sheet("Overall Layout")
    layout.cell(1, 1, "Layout placeholder")

    scope = wb.create_sheet("SCOPE")
    scope.append(["Section", "Scope Language"])
    scope.append(["A", "Original scope text"])
    wb.save(path)


def _workbook_v2(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title"])
    idx.append(["YES", 1, "EMS 12.0", "Overall Layout", "EMS Controls Overall Layout"])
    idx.append(["YES", 2, "EMS 13.0", "SCOPE", "Project Scope"])

    layout = wb.create_sheet("Overall Layout")
    layout.cell(1, 1, "Layout placeholder")

    scope = wb.create_sheet("SCOPE")
    scope.append(["Section", "Scope Language"])
    scope.append(["A", "UPDATED scope text from reimport"])
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    v1 = tmp / "v1.xlsx"
    v2 = tmp / "v2.xlsx"
    _workbook_v1(v1)
    _workbook_v2(v2)

    proj = import_workbook(v1, project_id="reup1")
    layout_page = next(p for p in proj["pages"] if p.get("sheetCode") == "EMS 12.0")
    manual_objects = [{"type": "rect", "objName": "manual-box", "left": 100, "top": 200, "width": 80, "height": 40}]
    layout_page["canvasObjects"] = manual_objects
    layout_page["pageType"] = "canvas"

    updated, summary = apply_reimport(proj, v2)
    problems: list[str] = []

    layout_after = next((p for p in updated["pages"] if p.get("sheetCode") == "EMS 12.0"), None)
    if layout_after is None:
        problems.append("layout page missing after reimport")
    elif layout_after.get("canvasObjects") != manual_objects:
        problems.append(f"canvasObjects changed: {layout_after.get('canvasObjects')}")

    if "EMS 12.0" not in summary.get("preserved", []):
        problems.append(f"layout not in preserved summary: {summary}")

    scope_after = next((p for p in updated["pages"] if p.get("sheetCode") == "EMS 13.0"), None)
    if scope_after is None:
        problems.append("scope page missing")
    else:
        block_text = str(scope_after.get("blocks"))
        if "UPDATED scope text" not in block_text:
            problems.append("scope page not updated from workbook")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — re-upload preserves manual layout")
    print(f"  preserved={summary.get('preserved')} updated={summary.get('updated')}")


if __name__ == "__main__":
    main()
