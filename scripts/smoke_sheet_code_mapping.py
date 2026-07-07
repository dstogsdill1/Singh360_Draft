"""Smoke: title-block sheet code uses the workbook/index Sheet Code column,
never the sequential output page order (FINAL RENDER POLISH 4G, Phase A).

Verifies:
  - Cover shows EMS 0.0 (index Sheet Code), not "1" (its Order/output order).
  - Abbreviations shows EMS 0.2, not "3" (its Order) and not "EMS 3.0"
    (the old bug: sequential output order dressed up as a sheet code).
  - Directory shows EMS 0.3.
  - LCP Panel Schedule shows EMS 1.4, and splits into a continuation page
    whose sheet code is EMS 1.4a (base code + letter suffix), never a
    sequential code like "EMS 15.0a".
  - "Sheet X of Y" (plain output order) stays sequential/correct regardless.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")


def _section_row(ws, row: int, title: str, cols: int = 6) -> None:
    ws.cell(row, 1, title)
    ws.cell(row, 1).fill = GOLD
    ws.cell(row, 1).font = Font(bold=True)
    for c in range(2, cols + 1):
        ws.cell(row, c).fill = GOLD


def _headers(ws, row: int, cols: int = 6) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c, f"Header {c}")
        cell.fill = GRAY
        cell.font = Font(bold=True)


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    # Order is the plain sequential column; Sheet Code is the real drawing
    # sheet number. They must intentionally disagree here so a regression
    # that falls back to Order cannot pass silently.
    idx.append(["Order", "Sheet Code", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "EMS 0.0", "Cover", "Cover Sheet", "YES", ""])
    idx.append(["2", "EMS 0.2", "Abbreviations", "Abbreviations", "YES", ""])
    idx.append(["3", "EMS 0.3", "Directory", "Directory", "YES", ""])
    idx.append(["4", "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "YES", ""])

    cover = wb.create_sheet("Cover")
    cover.cell(1, 1, "SINGH360 EMS PROJECT")

    abbrev = wb.create_sheet("Abbreviations")
    abbrev.append(["Abbreviation", "Meaning"])
    abbrev.append(["EMS", "Energy Management System"])
    abbrev.append(["LCP", "Lighting Control Panel"])

    directory = wb.create_sheet("Directory")
    directory.append(["Name", "Role", "Phone"])
    directory.append(["J. Smith", "PM", "555-0100"])

    lcp = wb.create_sheet("LCP Panel Schedule")
    _section_row(lcp, 1, "LCP-1 Dimming Panel")
    _headers(lcp, 2)
    r = 3
    for i in range(18):
        for c in range(1, 7):
            lcp.cell(r, c, f"LCP1 {i}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1
    _section_row(lcp, r, "LCP-2 Contactor Panel")
    r += 1
    _headers(lcp, r)
    r += 1
    for i in range(26):
        for c in range(1, 7):
            lcp.cell(r, c, f"RO{i + 1}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "sheet_codes.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="sheetcode1")
    pages = proj["pages"]
    problems: list[str] = []

    def page(tab: str) -> dict | None:
        return next((p for p in pages if p["sheetTab"] == tab), None)

    expected = {
        "Cover": "EMS 0.0",
        "Abbreviations": "EMS 0.2",
        "Directory": "EMS 0.3",
        "LCP Panel Schedule": "EMS 1.4",
    }
    for tab, code in expected.items():
        p = page(tab)
        if p is None:
            problems.append(f"{tab}: page not found")
            continue
        if p.get("sheetCode") != code:
            problems.append(f"{tab}: sheetCode={p.get('sheetCode')!r}, expected {code!r}")
        if p.get("displaySheetCode") != code:
            problems.append(f"{tab}: displaySheetCode={p.get('displaySheetCode')!r}, expected {code!r}")
        # The old bug rendered "EMS 3.0" (Order dressed as EMS-N.0), never allow that.
        if p.get("sheetCode", "").startswith("EMS") and p.get("sheetCode") not in expected.values():
            problems.append(f"{tab}: unexpected sheet code {p.get('sheetCode')!r}")

    lcp_pages = [p for p in pages if p["sheetTab"] == "LCP Panel Schedule"]
    if len(lcp_pages) < 2:
        problems.append(f"LCP Panel Schedule expected a continuation page, got {len(lcp_pages)} page(s)")
    else:
        cont = lcp_pages[1]
        cont_code = cont.get("displaySheetCode") or cont.get("sheetCode")
        if cont_code not in ("EMS 1.4a", "EMS 1.4.1"):
            problems.append(f"LCP continuation sheet code = {cont_code!r}, expected EMS 1.4a or EMS 1.4.1")
        if cont_code and cont_code.startswith("EMS 15"):
            problems.append(f"LCP continuation used sequential-order code {cont_code!r} instead of source code")

    # "Sheet X of Y" stays plain sequential output order regardless of sheet code.
    orders = [p.get("order") for p in pages if p.get("include", True)]
    if orders != sorted(orders):
        problems.append(f"output order not sequential: {orders}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - sheet code mapping passed")


if __name__ == "__main__":
    main()
