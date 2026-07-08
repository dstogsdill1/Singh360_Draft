"""Smoke: blank drawing/layout/pdf-vector pages carry a non-empty
export-visible placeholder marker (FINAL RELEASE CLEANUP 4H+SA38, Phase C).

Mirrors the real SA31 fixture: ``EMS 1.0/3.0/3.1/4.0`` layout/schematic pages
have zero embedded images and contain literal "Use PDF vector embed" notes.

Verifies:
  - Every blank canvas page's ``page["blankPagePlaceholder"]`` is non-empty
    (this is what ``NormalizedPage.tsx`` renders at export time instead of a
    silently blank base layer).
  - "DRAWING TO BE INSERTED" is chosen for layout, schematic, AND location
    blank pages (FINAL SA31 POLISH 4I Phase E).
  - A page with an embedded image gets NO placeholder marker (real content
    already present — never cover it with a note).
  - A best-effort asset match (synthetic ``assets/screenshots`` fixture)
    attaches an image block for a page whose title matches a screenshot
    filename.

The embedded-image and asset-match checks are skipped gracefully (with a
note) when Pillow isn't installed in the current environment, matching the
existing repo convention (see ``scripts/smoke_component_library.py``); the
core placeholder-marker assertions always run.
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook

try:
    from PIL import Image as PILImage  # type: ignore

    HAS_PILLOW = True
except Exception:
    HAS_PILLOW = False


def _workbook(path: Path, embed_png: Path | None) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 1.0", "EMS 1.0 Layout", "EMS Controls Overall Layout", "Front Matter", "layout", "Use PDF vector embed. Do not import as low-res PNG."])
    idx.append(["YES", 2, "EMS 4.0", "EMS 4.0 Location", "Interior Device Location", "Front Matter", "layout", "Use PDF vector embed."])
    idx.append(["YES", 3, "EMS 4.1", "EMS 4.1 Image", "Exterior Device Location (has image)", "Front Matter", "layout", ""])
    idx.append(["YES", 4, "EMS 5.0", "EMS 5.0 Match", "Panel Room Reference Layout", "Front Matter", "layout", ""])

    ws1 = wb.create_sheet("EMS 1.0 Layout")
    ws1.append(["Instruction", "Detail"])
    ws1.append(["Note", "Use PDF vector embed. Do not import as low-res PNG."])

    ws2 = wb.create_sheet("EMS 4.0 Location")
    ws2.append(["Instruction", "Detail"])
    ws2.append(["Note", "Use PDF vector embed."])

    ws3 = wb.create_sheet("EMS 4.1 Image")
    ws3.append(["Instruction", "Detail"])
    ws3.append(["Note", "Has an embedded image."])
    if embed_png is not None:
        from openpyxl.drawing.image import Image as XLImage

        ws3.add_image(XLImage(str(embed_png)), "D4")

    ws4 = wb.create_sheet("EMS 5.0 Match")
    ws4.append(["Instruction", "Detail"])
    ws4.append(["Note", "No embedded image, but a reference screenshot exists in assets/screenshots."])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "blank_pages.xlsx"

    embed_png = None
    if HAS_PILLOW:
        embed_png = tmp / "embed.png"
        PILImage.new("RGB", (40, 20), color=(200, 200, 200)).save(embed_png)

    _workbook(xlsx, embed_png)

    project_id = "blankpages1"
    real_project_dir: Path | None = None
    if HAS_PILLOW:
        from core.project_store import ProjectStore

        ps = ProjectStore(ROOT / ".docs")
        real_project_dir = ps.dir_for(project_id, {"projectDisplayName": "Blank Pages Smoke"})
        real_screens = real_project_dir / "assets" / "screenshots"
        real_screens.mkdir(parents=True, exist_ok=True)
        PILImage.new("RGB", (10, 10), color=(255, 0, 0)).save(real_screens / "panelroomreferencelayout.png")

    embedded_assets_dir = tmp / "assets_out"
    embedded_assets_dir.mkdir(exist_ok=True)

    problems: list[str] = []
    try:
        proj = import_workbook(
            xlsx,
            project_id=project_id,
            assets_dir=embedded_assets_dir,
            asset_url_prefix=f"/api/assets/{project_id}",
        )
        pages = proj["pages"]

        def page(tab: str) -> dict | None:
            return next((p for p in pages if p["sheetTab"] == tab), None)

        p1 = page("EMS 1.0 Layout")
        p2 = page("EMS 4.0 Location")
        p3 = page("EMS 4.1 Image")
        p4 = page("EMS 5.0 Match")

        if not p1 or not (p1.get("blankPagePlaceholder") or "").strip():
            problems.append("EMS 1.0 Layout: blankPagePlaceholder is empty — page would export silently blank")
        elif p1["blankPagePlaceholder"] != "DRAWING TO BE INSERTED":
            problems.append(f"EMS 1.0 Layout: placeholder={p1['blankPagePlaceholder']!r}, expected 'DRAWING TO BE INSERTED'")

        if not p2 or not (p2.get("blankPagePlaceholder") or "").strip():
            problems.append("EMS 4.0 Location: blankPagePlaceholder is empty")
        elif p2["blankPagePlaceholder"] != "DRAWING TO BE INSERTED":
            problems.append(
                f"EMS 4.0 Location: placeholder={p2['blankPagePlaceholder']!r}, "
                "expected 'DRAWING TO BE INSERTED' (FINAL SA31 POLISH 4I Phase E)"
            )

        if not p3:
            problems.append("EMS 4.1 Image: page not found")
        elif HAS_PILLOW:
            if (p3.get("blankPagePlaceholder") or "").strip():
                problems.append("EMS 4.1 Image: has an embedded image but still carries a placeholder marker")
            has_img = any(b.get("type") == "imagePlaceholder" for b in p3.get("blocks") or [])
            if not has_img:
                problems.append("EMS 4.1 Image: embedded image block missing")
        else:
            print("  (skipping embedded-image assertions — Pillow not installed)")

        if not p4:
            problems.append("EMS 5.0 Match: page not found")
        elif HAS_PILLOW:
            has_matched_img = any(b.get("type") == "imagePlaceholder" for b in p4.get("blocks") or [])
            if not has_matched_img:
                problems.append("EMS 5.0 Match: best-effort asset match did not attach an image block for a matching screenshot filename")
        else:
            print("  (skipping best-effort asset-match assertion — Pillow not installed)")

        if problems:
            print("FAIL")
            for p in problems:
                print(" -", p)
            raise SystemExit(1)
        print("OK — blank layout page placeholders + best-effort asset match passed")
        print(f"  EMS 1.0={p1['blankPagePlaceholder']!r}, EMS 4.0={p2['blankPagePlaceholder']!r}")
    finally:
        if real_project_dir is not None:
            shutil.rmtree(real_project_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
