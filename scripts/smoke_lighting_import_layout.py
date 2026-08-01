"""Disposable browser/restart/PDF proof for exact-source import and Save Project."""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
import socket
import subprocess
import sys
import tempfile
import time
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from core.project_store import ProjectStore
from core.sheet_importer import import_workbook_sheets


def free_port() -> int:
    with socket.socket() as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def wait_health(port: int) -> None:
    for _ in range(180):
        try:
            if urlopen(f"http://127.0.0.1:{port}/api/health", timeout=1).status == 200:
                return
        except OSError:
            time.sleep(.2)
    raise RuntimeError("disposable Singh360 server did not become healthy")


def start_server(docs: Path, port: int) -> subprocess.Popen:
    process = subprocess.Popen(
        [sys.executable, str(ROOT / "server.py")], cwd=ROOT,
        env={**os.environ, "SINGH360_DOCS_DIR": str(docs), "SINGH360_PORT": str(port)},
        stdout=subprocess.DEVNULL, stderr=subprocess.PIPE, text=True,
    )
    wait_health(port)
    return process


def stop_server(process: subprocess.Popen) -> None:
    process.terminate()
    try:
        process.wait(timeout=15)
    except subprocess.TimeoutExpired:
        process.kill(); process.wait(timeout=10)


def post_pdf(port: int, project_id: str, target: Path) -> None:
    request = Request(
        f"http://127.0.0.1:{port}/api/projects/{project_id}/export/pdf",
        data=json.dumps({"width": 17, "height": 11, "confirmPreflight": True}).encode(),
        headers={"Content-Type": "application/json"}, method="POST",
    )
    with urlopen(request, timeout=120) as response:
        target.write_bytes(response.read())


def fixture_workbook(path: Path) -> None:
    book = Workbook(); sheet = book.active; sheet.title = "LIGHTING CONTROL IO"
    orange = PatternFill("solid", fgColor="F4B183")
    gray = PatternFill("solid", fgColor="D9E1F2")
    for panel, start in enumerate((1, 16, 31, 46), start=1):
        sheet.merge_cells(start_row=start, start_column=1, end_row=start, end_column=12)
        sheet.cell(start, 1, f"Lighting Control Panel #{panel} — LCP{panel}").fill = orange
        for column in range(1, 13):
            sheet.cell(start + 1, column, "Description" if column == 2 else f"IO{column}").fill = gray
        for row in range(start + 2, start + 14):
            sheet.cell(row, 1, f"R{row:02d}")
            sheet.cell(row, 2, "Generated long wrapped lighting controls description")
    sheet.merge_cells("A62:O62"); sheet["A62"] = "LIGHTING CONTROLS SCHEDULE"; sheet["A62"].fill = orange
    for column in range(1, 16): sheet.cell(63, column, f"S{column}").fill = gray
    for row in range(64, 88):
        sheet.cell(row, 1, f"LT-{row:02d}"); sheet.cell(row, 2, "Generated schedule description")
    sheet["AH90"].font = __import__("openpyxl").styles.Font(name="Arial")
    book.save(path); book.close()


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="singh360_lighting_layout_") as raw:
        runtime = Path(raw); docs = runtime / ".docs"; source = runtime / "Lighting Controls.xlsx"
        fixture_workbook(source)
        source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
        project_id = "1a2b3c4d5e6f7081"
        project = {
            "id": project_id, "metadata": {"projectName": "Disposable Lighting Layout"},
            "sourceWorkbookName": "Authoritative Fixture.xlsx",
            "workbookSync": {"workbook": str(runtime / "unavailable-authoritative.xlsx"), "status": "app_changed"},
            "worksheets": [], "sources": [],
            "pages": [
                {"id": "before", "order": 1, "include": True, "sheetCode": "EMS 11.0", "displaySheetCode": "EMS 11.0", "sheetTitle": "Before", "sheetTab": "BEFORE", "pageType": "canvas", "templateId": "blank", "canvasObjects": [], "blocks": [], "notes": ""},
                {"id": "after", "order": 2, "include": True, "sheetCode": "EMS 13.0", "displaySheetCode": "EMS 13.0", "sheetTitle": "After", "sheetTab": "AFTER", "pageType": "canvas", "templateId": "blank", "canvasObjects": [], "blocks": [], "notes": ""},
            ],
        }
        imported, created = import_workbook_sheets(
            project, source, ["LIGHTING CONTROL IO"], insert_after_page_id="before",
            assets_dir=runtime / "assets", asset_url_prefix="/fixture/assets", source_filename=source.name,
        )
        base_id = created[0]["id"]
        store = ProjectStore(docs); store.save(project_id, imported)
        port = free_port(); process = start_server(docs, port)
        requests: list[str] = []
        try:
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(executable_path=r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe", headless=True)
                page = browser.new_page(viewport={"width": 1920, "height": 1080})
                page.on("request", lambda request: requests.append(request.url))
                url = f"http://127.0.0.1:{port}/app?project={project_id}&page={base_id}&mode=editor"
                page.goto(url, wait_until="domcontentloaded", timeout=60_000)
                page.locator(".ribbon").wait_for(timeout=30_000)
                if page.locator(".panel-rail-right").count():
                    page.locator(".panel-rail-right").click()
                page.get_by_label("Excel Layout").wait_for(timeout=30_000)
                if page.get_by_label("Excel Layout").input_value() != "exact_source":
                    raise AssertionError("Exact Source / Auto layout did not load")
                if page.locator(".np-table-region-layout.side_by_side").count():
                    raise AssertionError("exact-source default falsely rendered a side-by-side layout")
                if page.locator(".np-xr-table").count() != 1:
                    raise AssertionError("exact-source page did not render as one uniformly scaled range")
                with page.expect_response(lambda response: response.url.endswith(f"/pages/{base_id}/auto-layout") and response.status == 200):
                    page.get_by_label("Excel Layout").select_option("keep_one_page")
                page.wait_for_function("document.querySelector('#excel-layout')?.value === 'keep_one_page'", timeout=30_000)
                if page.get_by_label("Excel Layout").input_value() != "keep_one_page":
                    raise AssertionError("Keep on One Page override did not persist")
                with page.expect_response(lambda response: response.url.endswith(f"/pages/{base_id}/auto-layout") and response.status == 200):
                    page.get_by_label("Excel Layout").select_option("exact_source")
                page.wait_for_function("document.querySelector('#excel-layout')?.value === 'exact_source'", timeout=30_000)
                page.locator(".ribbon-appbar-right").get_by_role(
                    "button", name="Save Project", exact=True
                ).click()
                page.wait_for_timeout(1800)
                if any("workbook-link/resolve" in url for url in requests):
                    raise AssertionError("Save Project called workbook mirroring")
                page.reload(wait_until="domcontentloaded", timeout=60_000)
                if page.locator(".panel-rail-right").count():
                    page.locator(".panel-rail-right").click()
                page.get_by_label("Excel Layout").wait_for(timeout=30_000)
                if page.locator(".np-table-region-layout.side_by_side").count() or page.locator(".np-xr-table").count() != 1:
                    raise AssertionError("exact-source layout did not survive browser refresh")
                browser.close()

            stop_server(process)
            process = start_server(docs, port)
            with urlopen(f"http://127.0.0.1:{port}/api/projects/{project_id}", timeout=15) as response:
                restarted = json.load(response)
            base = next(item for item in restarted["pages"] if item["id"] == base_id)
            if base["sheetCode"] != "EMS 12.0" or base["layoutDiagnostics"]["blockCount"] != 1:
                raise AssertionError("stable identity/layout did not survive server restart")
            if base["layoutProfile"] != "exact_source_excel" or base["layoutOverride"] != "exact_source":
                raise AssertionError("exact-source profile did not survive server restart")
            legacy = restarted.get("legacyWorkbookReference") or {}
            if restarted.get("sourceWorkbookName"):
                raise AssertionError("standalone restart retained an active workbook authority")
            if legacy.get("sourceWorkbookName") != "Authoritative Fixture.xlsx":
                raise AssertionError("standalone restart lost the read-only legacy workbook provenance")
            if (restarted.get("workbookSync") or {}).get("mode") != "disabled":
                raise AssertionError("standalone restart did not disable workbook synchronization")
            pdf = runtime / "disposable-after.pdf"; post_pdf(port, project_id, pdf)
            if not pdf.is_file() or pdf.stat().st_size < 10_000:
                raise AssertionError("disposable PDF export failed")
            if hashlib.sha256(source.read_bytes()).hexdigest() != source_hash:
                raise AssertionError("source workbook changed")
            print(json.dumps({
                "status": "PASS", "projectId": project_id, "basePageId": base_id,
                "pageCount": len(restarted["pages"]), "continuationCount": sum(1 for item in restarted["pages"] if item.get("continuationOf") == base_id),
                "pdfBytes": pdf.stat().st_size, "sourceWorkbookUnchanged": True,
                "activeWorkbookAuthority": False,
                "legacyWorkbookProvenancePreserved": True,
                "workbookMirrorRequestsDuringSaveProject": sum("workbook-link/resolve" in url for url in requests),
            }, indent=2))
        finally:
            if process.poll() is None: stop_server(process)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
