"""Smoke: trailing blank worksheet columns/rows are trimmed from the
normalized/export excel_exact render (FINAL RENDER POLISH 4G, Phase B).

Verifies:
  - A sheet whose real content is 3 columns / 5 rows, but whose openpyxl
    usedRange balloons out to 15 columns purely from an incidental
    formatting touch (no value, no fill/border), renders trimmed back down
    to its real column extent in the normalized output.
  - A meaningful trailing fill is never trimmed away, even with no text in
    that cell (a real blocked/shaded cell must survive) — both a column
    that is otherwise blank, and a whole trailing row that only carries a
    border.
  - trimBlankColumns=false per-page override keeps the full untrimmed
    usedRange width when explicitly requested.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from core.workbook_importer import import_workbook

THIN = Side(style="thin")


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Code", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "EMS 0.2", "Abbreviations", "Abbreviations", "YES", ""])

    ws = wb.create_sheet("Abbreviations")
    ws.append(["Abbreviation", "Meaning", "Notes"])
    ws.append(["EMS", "Energy Management System", ""])
    ws.append(["LCP", "Lighting Control Panel", ""])
    ws.append(["IDF", "Intermediate Distribution Frame", ""])
    ws.append(["RDM", "Refrigeration Data Manager", ""])

    # Incidental usedRange bloat: a font touch on a far column with no value
    # and no fill/border — exactly the real-world "Excel usedRange is too
    # large" scenario the trim step must strip out of the normalized render.
    ws.cell(row=1, column=15).font = Font(size=11)

    # A real blocked/shaded trailing cell (meaningful fill) must survive even
    # though it carries no text.
    ws.cell(row=2, column=4).fill = PatternFill("solid", fgColor="D9D9D9")

    # A whole trailing row with no text anywhere, but a meaningful border on
    # one cell — a ruled divider line — must also survive as a real row.
    ws.cell(row=6, column=1).border = Border(top=THIN)

    # A genuinely blank trailing row (no value, no style at all) after that —
    # this one must still be dropped.
    ws.cell(row=7, column=1)

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "trim_blank.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="trim1")
    problems: list[str] = []

    page = next((p for p in proj["pages"] if p["sheetTab"] == "Abbreviations"), None)
    if page is None:
        problems.append("Abbreviations page not found")
    else:
        block = (page.get("blocks") or [{}])[0]
        cols_before = block.get("colsBeforeTrim")
        cols_after = block.get("colsAfterTrim")
        rows_after = block.get("rowsAfterTrim")

        if not (cols_before and cols_before >= 15):
            problems.append(f"colsBeforeTrim={cols_before}, expected >= 15 (usedRange bloat present)")

        # Real content is 3 cols; the meaningful fill at column D (index 3)
        # must keep 4 columns, not shrink to 3.
        if cols_after != 4:
            problems.append(f"colsAfterTrim={cols_after}, expected 4 (3 real cols + 1 meaningful-fill col)")

        # Real content is 5 rows (header + 4 abbreviations) + the meaningful
        # border row (row 6) = 6. The fully blank row 7 must not survive.
        if rows_after != 6:
            problems.append(f"rowsAfterTrim={rows_after}, expected 6 (5 real rows + 1 meaningful-border row)")

        grid = block.get("grid") or []
        if len(grid) != rows_after or max((len(r) for r in grid), default=0) != cols_after:
            problems.append("grid dimensions do not match rowsAfterTrim/colsAfterTrim")

    # Per-page override: trimBlankColumns=false must keep the full bloat.
    from core.workbook_importer import _excel_range_block

    proj2 = import_workbook(xlsx, project_id="trim2")
    page2 = next((p for p in proj2["pages"] if p["sheetTab"] == "Abbreviations"), None)
    if page2 is not None:
        ws_payload = next(w for w in proj2["worksheets"] if w["id"] == page2["linkedWorksheetId"])
        settings = {
            "splitMode": page2.get("splitMode", "none"),
            "allowContinuation": page2.get("allowContinuation", False),
            "minScale": page2.get("minScale", 0.73),
            "scaleMode": page2.get("scaleMode", "fit_body"),
            "trimBlankRows": False,
            "trimBlankColumns": False,
        }
        block2 = _excel_range_block(ws_payload, "override_test", settings)
        if block2.get("colsAfterTrim") != block2.get("colsBeforeTrim"):
            problems.append("trimBlankColumns=false override did not preserve full usedRange width")
        if block2.get("rowsAfterTrim") != block2.get("rowsBeforeTrim"):
            problems.append("trimBlankRows=false override did not preserve full usedRange height")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - trailing blank range trimming passed")


if __name__ == "__main__":
    main()
