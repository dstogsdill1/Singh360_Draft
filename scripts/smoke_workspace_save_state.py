"""End-to-end save-state smoke using only a generated disposable workbook/project."""
from __future__ import annotations

import hashlib
import json
import os
import socket
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from urllib.request import urlopen

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from core.project_store import ProjectStore
from core.project_workspace import WorkbookDocumentStore
from core.workbook_link_manager import set_link
from tests.generated_fixtures import write_workbook


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def workbook_structure(path: Path) -> dict:
    book = load_workbook(path, data_only=False, read_only=False)
    try:
        return {
            "sheetNames": book.sheetnames,
            "sheets": {
                sheet.title: {
                    "state": sheet.sheet_state,
                    "merged": sorted(str(value) for value in sheet.merged_cells.ranges),
                    "formulas": {
                        cell.coordinate: cell.value
                        for row in sheet.iter_rows()
                        for cell in row
                        if isinstance(cell.value, str) and cell.value.startswith("=")
                    },
                }
                for sheet in book.worksheets
            },
        }
    finally:
        book.close()


def free_port() -> int:
    with socket.socket() as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def wait_health(port: int) -> None:
    for _ in range(150):
        try:
            if urlopen(f"http://127.0.0.1:{port}/api/health", timeout=1).status == 200:
                return
        except OSError:
            time.sleep(0.2)
    raise RuntimeError("isolated save-state server did not become healthy")


def workspace_document() -> dict:
    return {
        "revision": 0,
        "updatedAt": "",
        "sheets": [
            {
                "id": "source-sheet",
                "name": "SRC_SANITIZED",
                "cells": {"A1": {"v": "Locked metadata"}, "A3": {"v": "Generated editable value"}},
                "styles": {},
                "merges": [],
                "rowHeights": {},
                "columnWidths": {},
                "defaultColumnWidth": 8.43,
                "defaultRowHeight": 15,
                "hiddenRows": [],
                "hiddenColumns": [],
                "archived": False,
                "tabColor": None,
                "role": "source",
                "sourceSetup": {
                    "authority": "00_INDEX",
                    "sheetCode": "SRC T.1",
                    "title": "Sanitized Save-State Source",
                    "publish": "YES",
                    "purpose": "Generated save-state fixture",
                    "editableStartRow": 3,
                    "metadata": [],
                },
                "protectedRanges": ["A1:Z2"],
                "dataValidations": [],
                "conditionalFormats": [],
                "tableRegions": [{"id": "table-1", "range": "A3:A3", "label": "Table 1"}],
                "tableLayout": "single",
                "annotations": [],
            }
        ],
    }


def beforeunload_prevented(page) -> bool:
    return bool(
        page.evaluate(
            """() => {
              const event = new Event('beforeunload', { cancelable: true });
              window.dispatchEvent(event);
              return event.defaultPrevented;
            }"""
        )
    )


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="singh360_save_state_") as raw:
        runtime = Path(raw)
        docs = runtime / ".docs"
        workbook = write_workbook(runtime / "protected-clone.xlsx")
        original_structure = workbook_structure(workbook)
        original_hash = sha256(workbook)

        project_id = "5a7e57005a7e5700"
        initial = {
            "id": project_id,
            "metadata": {"projectName": "Sanitized Regression Project"},
            "worksheets": [],
            "pages": [],
            "sources": [],
        }
        store = ProjectStore(docs)
        linked, state = set_link(project_id, initial, store, str(workbook))
        if state["status"] != "in_sync":
            raise AssertionError(f"generated clone did not establish a safe baseline: {state}")
        if not linked["pages"]:
            raise AssertionError("generated workbook import produced no pages")

        # Create a genuine app-side workbook-backed edit without touching the clone.
        linked["pages"][0]["sheetTitle"] = "Generated pending title"
        linked["pages"][0]["canvasObjects"] = [{"type": "rect", "name": "generated-save-state-object"}]
        store.save(project_id, linked)
        project_dir = store.dir_for(project_id, linked)
        WorkbookDocumentStore(project_dir).save(linked, 0, workspace_document())

        port = free_port()
        env = {
            **os.environ,
            "SINGH360_DOCS_DIR": str(docs),
            "SINGH360_PORT": str(port),
            "SINGH360_TEST_SAVE_DELAY_MS": "900",
        }
        process = subprocess.Popen(
            [sys.executable, str(ROOT / "server.py")],
            cwd=ROOT,
            env=env,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        request_log: list[tuple[str, str]] = []
        console_errors: list[str] = []
        dialogs: list[str] = []
        try:
            wait_health(port)
            with sync_playwright() as api:
                browser = api.chromium.launch(
                    executable_path=r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                    headless=True,
                )
                page = browser.new_page(viewport={"width": 1500, "height": 900})
                page.on("request", lambda request: request_log.append((request.method, request.url)))
                page.on("console", lambda message: console_errors.append(message.text) if message.type == "error" else None)
                page.on("dialog", lambda dialog: (dialogs.append(dialog.message), dialog.dismiss()))
                base = f"http://127.0.0.1:{port}/app?project={project_id}"

                # Existing project load, zoom, selection, tooltip hover, and Ctrl+S.
                page.goto(
                    f"{base}&mode=editor&tooltipAudit=1",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                page.locator(".ribbon").wait_for(timeout=30_000)
                status = page.locator(".save-state-control .status-pill")
                status.wait_for(timeout=10_000)
                if status.inner_text() not in {
                    "PROJECT SAVED · WORKBOOK SYNC PENDING",
                    "PROJECT SAVED",
                }:
                    raise AssertionError(f"project opened falsely dirty: {status.inner_text()}")
                initial_label = status.inner_text()
                page.get_by_role("button", name="View", exact=True).click()
                zoom = page.locator('[data-help-id="view.zoomIn"]').first
                if zoom.count():
                    zoom.click()
                if "UNSAVED" in status.inner_text():
                    raise AssertionError("zoom marked project dirty")
                status.hover()
                page.wait_for_timeout(450)
                if "UNSAVED" in status.inner_text():
                    raise AssertionError("tooltip hover marked project dirty")

                resolve_before_ctrl_s = sum("/workbook-link/resolve" in url for _, url in request_log)
                page.keyboard.press("Control+s")
                page.wait_for_timeout(1_300)
                resolve_after_ctrl_s = sum("/workbook-link/resolve" in url for _, url in request_log)
                if resolve_after_ctrl_s != resolve_before_ctrl_s:
                    raise AssertionError("Ctrl+S invoked workbook synchronization")

                # Data Workspace: real metadata edit, save failure, retry, and in-flight newer edit.
                page.goto(
                    f"{base}&view=data&tooltipAudit=1",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                page.locator(".univer-host").wait_for(timeout=30_000)
                page.wait_for_timeout(1_500)
                workspace_status = page.locator(".workspace-status")
                if "UNSAVED" in workspace_status.inner_text():
                    raise AssertionError("Data Workspace opened falsely dirty")
                if beforeunload_prevented(page):
                    raise AssertionError("workbook sync pending alone triggered beforeunload")

                right_note = page.get_by_label("Right-side note")
                bottom_note = page.get_by_label("Bottom note")
                right_note.fill("Generated first workspace edit")
                page.wait_for_timeout(150)
                if workspace_status.inner_text() != "UNSAVED WORKSPACE EDITS":
                    raise AssertionError(f"workspace edit state was not truthful: {workspace_status.inner_text()}")
                if not beforeunload_prevented(page):
                    raise AssertionError("genuine workspace edit did not trigger beforeunload")

                workspace_status.hover()
                page.wait_for_timeout(450)
                if workspace_status.inner_text() != "UNSAVED WORKSPACE EDITS":
                    raise AssertionError("tooltip hover changed workspace dirty state")

                def fail_workspace_save(route) -> None:
                    if route.request.method == "PUT":
                        route.fulfill(
                            status=500,
                            content_type="application/json",
                            body=json.dumps({"error": "simulated local save failure"}),
                        )
                    else:
                        route.continue_()

                page.route("**/api/projects/*/data-workspace", fail_workspace_save)
                page.get_by_role("button", name="Save Workspace Edits").click()
                workspace_status.get_by_text("SAVE FAILED", exact=True).wait_for(timeout=5_000)
                if not beforeunload_prevented(page):
                    raise AssertionError("failed local save cleared the dirty unload guard")
                if "simulated local save failure" not in page.locator(".data-message").inner_text():
                    raise AssertionError("actual local save failure was not visible")
                if not any("500" in message for message in console_errors):
                    raise AssertionError("simulated local save failure did not reach the browser")
                console_errors.clear()
                page.unroute("**/api/projects/*/data-workspace", fail_workspace_save)

                page.get_by_role("button", name="Save Workspace Edits").click()
                workspace_status.get_by_text("SAVING PROJECT…", exact=True).wait_for(timeout=2_000)
                bottom_note.fill("Generated newer in-flight edit")
                page.wait_for_function(
                    """() => {
                      const status = document.querySelector('.workspace-status')?.textContent;
                      const message = document.querySelector('.data-message')?.textContent || '';
                      return status === 'UNSAVED WORKSPACE EDITS'
                        && message.includes('newer Data Workspace edits');
                    }""",
                    timeout=5_000,
                )

                page.get_by_role("button", name="Save Workspace Edits").click()
                workspace_status.get_by_text(
                    "PROJECT SAVED · WORKBOOK SYNC PENDING", exact=True
                ).wait_for(timeout=5_000)
                if beforeunload_prevented(page):
                    raise AssertionError("locally saved workspace pending sync triggered beforeunload")
                resolve_after_workspace = sum("/workbook-link/resolve" in url for _, url in request_log)
                if resolve_after_workspace != resolve_after_ctrl_s:
                    raise AssertionError("Save Workspace Edits wrote the linked workbook")

                page.reload()
                page.locator(".univer-host").wait_for(timeout=30_000)
                page.wait_for_timeout(1_500)
                if page.locator(".workspace-status").inner_text() != "PROJECT SAVED · WORKBOOK SYNC PENDING":
                    raise AssertionError("workspace reload produced a false dirty marker")
                if page.get_by_label("Right-side note").input_value() != "Generated first workspace edit":
                    raise AssertionError("first workspace edit did not persist")
                if page.get_by_label("Bottom note").input_value() != "Generated newer in-flight edit":
                    raise AssertionError("newer workspace edit did not persist")

                # Explicit full mirror runs local save first and only writes the protected clone.
                with urlopen(f"http://127.0.0.1:{port}/api/projects/{project_id}", timeout=10) as response:
                    editor_payload = json.load(response)
                if not editor_payload.get("pages"):
                    raise AssertionError(
                        f"saved Data Workspace project no longer had editor pages: {editor_payload}"
                    )
                page.close()
                page = browser.new_page(viewport={"width": 1500, "height": 900})
                page.on("request", lambda request: request_log.append((request.method, request.url)))
                page.on("console", lambda message: console_errors.append(message.text) if message.type == "error" else None)
                page.on("dialog", lambda dialog: (dialogs.append(dialog.message), dialog.dismiss()))
                page.goto(
                    f"{base}&mode=editor&tooltipAudit=1",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                page.locator(".ribbon").wait_for(timeout=30_000)
                write_excel = page.get_by_role("button", name="SAVE + WRITE EXCEL", exact=True)
                try:
                    page.wait_for_function(
                        """() => {
                          const button = document.querySelector('[data-help-id="save.writeExcel"]');
                          return button instanceof HTMLButtonElement && !button.disabled;
                        }""",
                        timeout=10_000,
                    )
                except Exception as exc:
                    raise AssertionError(
                        "editor did not finish loading the generated project: "
                        f"url={page.url}; text={page.locator('body').inner_text()[:1200]!r}; "
                        f"requests={request_log[-20:]}; console={console_errors[-20:]}"
                    ) from exc
                write_excel.click()
                try:
                    page.wait_for_function(
                        """() => [
                          'PROJECT SAVED · WORKBOOK SYNCED',
                          'WORKBOOK SYNC FAILED',
                          'SAVE FAILED'
                        ].includes(document.querySelector('.save-state-control .status-pill')?.textContent || '')""",
                        timeout=30_000,
                    )
                except Exception as exc:
                    current = page.locator(".save-state-control .status-pill").inner_text()
                    error_text = page.locator(".save-state-error").inner_text() if page.locator(".save-state-error").count() else ""
                    raise AssertionError(
                        f"clone-only mirror did not settle: {current}; {error_text}; dialogs={dialogs}; "
                        f"requests={request_log[-20:]}"
                    ) from exc
                mirror_status = page.locator(".save-state-control .status-pill").inner_text()
                if mirror_status != "PROJECT SAVED · WORKBOOK SYNCED":
                    error_text = page.locator(".save-state-error").inner_text() if page.locator(".save-state-error").count() else ""
                    raise AssertionError(
                        f"clone-only Save + Write Excel failed: {mirror_status}; {error_text}; dialogs={dialogs}; "
                        f"requests={request_log[-20:]}"
                    )
                resolve_indices = [
                    index for index, (_, url) in enumerate(request_log)
                    if "/workbook-link/resolve" in url
                ]
                if not resolve_indices:
                    raise AssertionError("explicit Save + Write Excel did not run the workbook mirror")
                last_resolve = resolve_indices[-1]
                if not any(
                    method == "POST" and f"/api/projects/{project_id}" in url
                    for method, url in request_log[:last_resolve]
                ):
                    raise AssertionError("workbook mirror ran before a confirmed local project save")

                # Workbook failure stays distinct after a confirmed local save.
                missing_clone = workbook.with_suffix(".temporarily-moved.xlsx")
                workbook.rename(missing_clone)
                try:
                    write_excel.click()
                    page.locator(".save-state-control .status-pill").get_by_text(
                        "WORKBOOK SYNC FAILED", exact=True
                    ).wait_for(timeout=30_000)
                    if beforeunload_prevented(page):
                        raise AssertionError("workbook sync failure was treated as unsaved local work")
                    if not any("409" in message or "Save + Write Excel failed" in message for message in console_errors):
                        raise AssertionError("missing-workbook failure did not reach the browser")
                    console_errors.clear()
                finally:
                    missing_clone.rename(workbook)

                if console_errors:
                    raise AssertionError(f"browser console errors: {console_errors}")
                browser.close()
        finally:
            process.terminate()
            try:
                process.wait(timeout=10)
            except subprocess.TimeoutExpired:
                process.kill()

        saved = store.load(project_id)
        if saved is None:
            raise AssertionError("generated project disappeared")
        document = WorkbookDocumentStore(store.dir_for(project_id, saved)).load(saved)
        annotations = document["sheets"][0]["annotations"]
        texts = {item["text"] for item in annotations}
        if texts != {"Generated first workspace edit", "Generated newer in-flight edit"}:
            raise AssertionError(f"workspace persistence mismatch: {texts}")
        final_structure = workbook_structure(workbook)
        if final_structure["sheetNames"] != original_structure["sheetNames"]:
            raise AssertionError(
                "full mirror changed the generated workbook sheet identity: "
                f"before={original_structure['sheetNames']}; after={final_structure['sheetNames']}"
            )

        # Direct conflict gate: a local save may persist recovery state but cannot clear conflict.
        os.environ["SINGH360_DOCS_DIR"] = str(runtime / "conflict-runtime" / ".docs")
        import server
        from tests.generated_fixtures import isolate_server_runtime

        isolated = isolate_server_runtime(server)
        try:
            conflict_project = {
                "id": "c0f1c7c0f1c7c0f1",
                "metadata": {"projectName": "Generated Conflict"},
                "worksheets": [],
                "pages": [],
                "sources": [],
                "workbookSync": {"status": "conflict", "warning": "Both sides changed."},
            }
            server.store.save(conflict_project["id"], conflict_project)
            response = server.app.test_client().post(
                f"/api/projects/{conflict_project['id']}", json=conflict_project
            )
            if response.status_code != 200:
                raise AssertionError(response.get_data(as_text=True))
            payload = response.get_json()
            if payload["workbookSync"]["status"] != "conflict":
                raise AssertionError("local save cleared the two-sided workbook conflict")
        finally:
            isolated.cleanup()

        print(
            json.dumps(
                {
                    "ok": True,
                    "clone": str(workbook),
                    "cloneStartingSha256": original_hash,
                    "cloneFinalSha256": sha256(workbook),
                    "initialEditorState": initial_label,
                    "requestCount": len(request_log),
                    "saveStateChecks": 14,
                    "workbookStructurePreserved": True,
                },
                indent=2,
            )
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
