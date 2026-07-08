"""Smoke: 00_APP_INDEX alias-column fallback + 00_PROJECT_META title-block
metadata (FINAL RELEASE CLEANUP 4H+SA38, Phase A/B).

Builds a workbook with ONLY ``00_APP_INDEX`` (no ``00_INDEX`` at all) using
the SA38 alias column names (``Suggested EMS Code`` / ``Original Tab`` /
``Normalized Page Title``), plus a ``00_PROJECT_META`` key/value sheet.

Verifies:
  - When only ``00_APP_INDEX`` exists, it is used as the controlling index
    and its alias columns resolve correctly (Suggested EMS Code -> sheet
    code, Original Tab -> sheet tab match, Normalized Page Title -> title).
  - ``00_APP_INDEX`` / ``00_PROJECT_META`` never appear as output pages.
  - ``00_PROJECT_META`` values populate ``project.metadata`` (Location,
    Revision, Drawing Package File Name, Drawn By) rather than staying blank.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    bom = wb.active
    bom.title = "BOM"
    bom.append(["Item", "Description"])
    bom.append(["1", "Controller"])

    lighting = wb.create_sheet("Lighting-TDB")
    lighting.append(["Point", "Description", "Type"])
    lighting.append(["LT-1", "Lot lighting", "AO"])

    app_idx = wb.create_sheet("00_APP_INDEX")
    app_idx.append(["Include", "Order", "Suggested EMS Code", "Original Tab", "Normalized Page Title", "Family", "Page Type", "Notes"])
    app_idx.append(["YES", 1, "EMS 0.0", "BOM", "Bill of Materials", "BOM", "table", ""])
    app_idx.append(["YES", 2, "EMS 5.0", "Lighting-TDB", "Lighting TDB / LT-1 I/O", "Lighting", "io-table", ""])

    meta = wb.create_sheet("00_PROJECT_META")
    meta.append(["Project Metadata", None])
    meta.append(["Project Name", "SA38 Alias-Only Test"])
    meta.append(["Drawing Package File Name", "SA38_Alias_Only_Upload"])
    meta.append(["Location", "HEB 999 / Test City"])
    meta.append(["Revision", "V2"])
    meta.append(["Drawn By", "Singh360 Inc."])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "sa38_alias_only.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="sa38aliasonly")
    pages = proj["pages"]
    page_tabs = {p["sheetTab"] for p in pages}
    problems: list[str] = []

    for control in ("00_APP_INDEX", "00_PROJECT_META"):
        if control in page_tabs:
            problems.append(f"{control} appeared as an output page")

    def page(tab: str) -> dict | None:
        return next((p for p in pages if p["sheetTab"] == tab), None)

    bom = page("BOM")
    if bom is None or bom.get("displaySheetCode") != "EMS 0.0":
        problems.append(f"BOM sheetCode={bom.get('displaySheetCode') if bom else None!r}, expected 'EMS 0.0' (alias Suggested EMS Code column not resolved)")
    if bom is not None and bom.get("sheetTitle") != "Bill of Materials":
        problems.append(f"BOM title={bom.get('sheetTitle')!r}, expected 'Bill of Materials' (alias Normalized Page Title not resolved)")

    lighting = page("Lighting-TDB")
    if lighting is None or lighting.get("displaySheetCode") != "EMS 5.0":
        problems.append(f"Lighting-TDB sheetCode={lighting.get('displaySheetCode') if lighting else None!r}, expected 'EMS 5.0'")

    meta = proj["metadata"]
    expected_meta = {
        "location": "HEB 999 / Test City",
        "revision": "V2",
        "drawingPackageFileName": "SA38_Alias_Only_Upload",
        "drawnBy": "Singh360 Inc.",
    }
    for field, expected in expected_meta.items():
        if meta.get(field) != expected:
            problems.append(f"metadata[{field}]={meta.get(field)!r}, expected {expected!r} (00_PROJECT_META not applied)")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — 00_APP_INDEX alias-only fallback + 00_PROJECT_META metadata passed")
    print(f"  BOM={bom.get('displaySheetCode')}, Lighting-TDB={lighting.get('displaySheetCode')}, location={meta.get('location')!r}")


if __name__ == "__main__":
    main()
