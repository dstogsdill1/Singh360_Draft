"""Smoke: orange headers replace dark bands on all non-cover pages.

Builds a workbook with a dark/black title band on a schedule sheet and a cover
sheet, imports it, and verifies:
  - the schedule's excel_exact block has no dark title fills left (recolored to
    orange), and carries the orange band fill;
  - the cover page keeps its own style (normalizedHeaderStyle = "source");
  - every non-cover included page reports renderProfile/normalizedHeaderStyle.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.table_style_profile import TITLE_BAND_FILL, is_dark_fill
from core.workbook_importer import import_workbook


def _fixture(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["0", "COVER", "Cover", "YES", ""])
    idx.append(["5", "BOM", "Bill of Materials", "YES", ""])

    cov = wb.create_sheet("COVER")
    cov.cell(1, 1, "PROJECT COVER")

    ws = wb.create_sheet("BOM")
    t = ws.cell(1, 1, "BILL OF MATERIALS")
    t.fill = PatternFill("solid", fgColor="000000")
    t.font = Font(bold=True, color="FFFFFF")
    ws.merge_cells("A1:C1")
    for c, h in enumerate(["Item", "Qty", "Part"], start=1):
        cell = ws.cell(2, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1A1A1A")
        cell.font = Font(bold=True, color="FFFFFF")
    for r in range(3, 15):
        for c in range(1, 4):
            ws.cell(r, c, f"v{r}{c}")
    wb.save(path)


def _dark_fills_in_header(block: dict) -> int:
    styles = block.get("styles") or {}
    header = int(block.get("headerRowCount") or 1)
    n = 0
    for key, st in styles.items():
        try:
            r = int(key.split(":")[0])
        except ValueError:
            continue
        if r < header and is_dark_fill(st.get("fill")):
            n += 1
    return n


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "orange.xlsx"
    _fixture(xlsx)
    proj = import_workbook(xlsx, project_id="orange1")

    problems: list[str] = []
    non_cover = [p for p in proj["pages"] if p["pageType"] != "cover"]
    for p in non_cover:
        if p.get("normalizedHeaderStyle") != "orange":
            problems.append(f"{p['sheetTab']} header style = {p.get('normalizedHeaderStyle')}")

    bom = [p for p in proj["pages"] if p["sheetTab"] == "BOM"]
    if not bom:
        problems.append("BOM page missing")
    else:
        xr = next((b for b in bom[0]["blocks"] if b.get("type") == "excelRange"), None)
        if xr is None:
            problems.append("BOM has no excel_exact block")
        else:
            if _dark_fills_in_header(xr) > 0:
                problems.append("BOM header still has dark fills")
            fills = [s.get("fill") for s in (xr.get("styles") or {}).values()]
            if TITLE_BAND_FILL not in fills:
                problems.append("BOM has no orange title band")

    cover = [p for p in proj["pages"] if p["pageType"] == "cover"]
    for p in cover:
        if p.get("normalizedHeaderStyle") not in (None, "source"):
            problems.append(f"cover overridden: {p.get('normalizedHeaderStyle')}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — orange headers replace dark bands (cover preserved)")


if __name__ == "__main__":
    main()
