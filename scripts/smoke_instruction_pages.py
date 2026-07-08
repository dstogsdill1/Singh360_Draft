"""Smoke: compact Step/Instruction column widths + 9pt body font override
for instruction pages (FINAL RELEASE CLEANUP 4H+SA38, Phase D).

Mirrors the real SA31 EC/DC/EMS Field Instructions sheets (2-column
Step/Instruction tables). Verifies:
  - All source rows render (no rows silently dropped).
  - The Step column stays compact (<= 64px) instead of stretching
    proportionally with the wide Instruction column.
  - The rendered block carries an explicit ``bodyFontPx`` (9pt ~= 12px)
    override for the instruction_table profile.
  - ``clipping == False`` and no giant row heights carried over from an
    inflated Excel source row height.
  - A 1-row edge case still renders compactly (no giant single-row page).
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 2.0", "EC Field Instructions", "EC Field Instructions", "Front Matter", "text", ""])
    idx.append(["YES", 2, "EMS 2.1", "One Row Edge", "One Row Edge Case Instructions", "Front Matter", "text", ""])

    ec = wb.create_sheet("EC Field Instructions")
    ec.append(["Step", "Instruction"])
    long_texts = [
        "Verify all field wiring matches the approved single-line diagram before energizing any circuit.",
        "Confirm the LCP-1 dimming panel is de-energized and lock-out/tag-out is applied prior to work.",
        "Terminate all low-voltage control wiring per the wiring schedule; torque to manufacturer spec.",
        "Label every conductor at both ends using the panel schedule designation, not a field abbreviation.",
        "Perform a continuity check on every home-run before connecting it to the LCP-1 terminal block.",
        "Notify the EMS commissioning technician once all field terminations are complete and verified.",
    ]
    for i, text in enumerate(long_texts, start=1):
        ec.append([str(i), text])
        ec.row_dimensions[i + 1].height = 90  # inflated Excel source row height — must not carry through

    one_row = wb.create_sheet("One Row Edge")
    one_row.append(["Step", "Instruction"])
    one_row.append(["1", "Single instruction row edge case."])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "instructions.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="instructions1")
    pages = proj["pages"]
    problems: list[str] = []

    def page(tab: str) -> dict | None:
        return next((p for p in pages if p["sheetTab"] == tab), None)

    ec = page("EC Field Instructions")
    if ec is None:
        problems.append("EC Field Instructions page not found")
    else:
        if ec.get("layoutProfile") != "instruction_table":
            problems.append(f"EC Field Instructions layoutProfile={ec.get('layoutProfile')!r}, expected 'instruction_table'")
        block = (ec.get("blocks") or [{}])[0]
        grid = block.get("grid") or []
        if len(grid) != 7:  # header + 6 rows
            problems.append(f"EC Field Instructions: expected 7 grid rows (1 header + 6 data), got {len(grid)}")
        col_widths = block.get("colWidths") or []
        if not col_widths or col_widths[0] > 64:
            problems.append(f"EC Field Instructions: Step column width={col_widths[0] if col_widths else None}, expected <= 64px (compact)")
        if len(col_widths) > 1 and col_widths[1] <= col_widths[0]:
            problems.append("EC Field Instructions: Instruction column is not wider than the Step column")
        if block.get("bodyFontPx") != 12:
            problems.append(f"EC Field Instructions: bodyFontPx={block.get('bodyFontPx')!r}, expected 12 (9pt override)")
        row_heights = block.get("rowHeights") or []
        if any(h > 120 for h in row_heights):
            problems.append(f"EC Field Instructions: an inflated Excel source row height leaked through: {row_heights}")
        if ec.get("layoutWarnings"):
            problems.append(f"EC Field Instructions: unexpected layout warnings (clipping risk): {ec['layoutWarnings']}")

    one_row = page("One Row Edge")
    if one_row is None:
        problems.append("One Row Edge page not found")
    else:
        block = (one_row.get("blocks") or [{}])[0]
        grid = block.get("grid") or []
        if len(grid) != 2:
            problems.append(f"One Row Edge: expected 2 grid rows (1 header + 1 data), got {len(grid)}")
        if block.get("bodyFontPx") != 12:
            problems.append(f"One Row Edge: bodyFontPx={block.get('bodyFontPx')!r}, expected 12")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — instruction page compact layout + 9pt body font passed")
    print(f"  EC Field Instructions colWidths={ec['blocks'][0].get('colWidths')}, bodyFontPx={ec['blocks'][0].get('bodyFontPx')}")


if __name__ == "__main__":
    main()
