"""Smoke: EMS 2.0 Sheet Index exports when 00_INDEX has no self-referencing row (Phase A)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.export_qa import compute_export_warnings
from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.cell(1, 1, "PROJECT COVER")

    idx = wb.create_sheet("00_INDEX")
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title"])
    idx.append(["YES", 1, "EMS 1.0", "Cover", "Cover Sheet"])
    # Deliberately NO row for 00_INDEX itself — the real-world failure mode.
    idx.append(["YES", 3, "EMS 3.0", "Guidelines", "EMS Guidelines"])

    guide = wb.create_sheet("Guidelines")
    guide.append(["Topic", "Guideline"])
    guide.append(["General", "Follow Singh360 standards."])
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "index_no_self_row.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="idxself1")
    problems: list[str] = []

    index_pages = [p for p in proj["pages"] if p.get("sheetTab") == "00_INDEX"]
    if len(index_pages) != 1:
        problems.append(f"expected 1 index page, got {len(index_pages)}")
    elif not index_pages[0].get("include"):
        problems.append("index page include=False")

    included = [p for p in proj["pages"] if p.get("include")]
    tabs = [p.get("sheetTab") for p in included]
    if "00_INDEX" not in tabs:
        problems.append("00_INDEX missing from included export pages")
    if tabs.index("00_INDEX") != 1:
        problems.append(f"index page order wrong: tabs={tabs} (expected Cover then 00_INDEX)")

    warns = compute_export_warnings(proj)
    missing_index = [w for w in warns if "index" in w.get("issue", "").lower() and "missing" in w.get("issue", "").lower()]
    if missing_index:
        problems.append(f"QA gate reported missing index: {missing_index}")

    if not any(w.get("issue", "").startswith("Sheet Index / TOC has no row") for w in proj.get("importWarnings", [])):
        problems.append("expected importWarnings entry for index self-row fallback")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — Sheet Index included without self-row in 00_INDEX")
    print(f"  export tabs={tabs}")


if __name__ == "__main__":
    main()
