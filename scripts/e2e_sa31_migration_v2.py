"""Live acceptance for the SA31 schema-V2 migration and Source Library V2."""
from __future__ import annotations

import io
import json
import tempfile
import zipfile
from pathlib import Path

import fitz
import requests
from openpyxl import Workbook, load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / ".docs"
BASE = "http://127.0.0.1:8766"
STAGED = DOCS / "migration_staging" / "SA31_Singh360_Template_Platform_Migration_V2.xlsx"
OLD_WORKBOOK = (
    DOCS / "projects" / "95d85da603864a62__95d85da603864a62"
    / "sources" / "workbook" / "SA31_EMS_Lighting_Workbook_V1.xlsx"
)
BACKUP_ROOT = DOCS / "patch_backups" / "sa31_schema_v2_source_library_20260726-135630"


def migrated_project_id() -> str:
    report = json.loads(
        (DOCS / "audits" / "sa31_schema_v2_source_library_20260726-135630"
         / "migration_apply.json").read_text("utf-8")
    )
    return str(report["newProjectId"])


def expect(response: requests.Response, status: int = 200) -> requests.Response:
    if response.status_code != status:
        raise AssertionError(
            f"{response.request.method} {response.url}: "
            f"{response.status_code} {response.text[:1200]}"
        )
    return response


def sha256(path: Path) -> str:
    import hashlib

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fixtures() -> dict[str, tuple[str, bytes, str]]:
    pdf = fitz.open()
    page = pdf.new_page(width=612, height=792)
    page.insert_text((72, 72), "Singh360 sanitized preview fixture")
    pdf_bytes = pdf.tobytes()
    pdf.close()

    image = Image.new("RGB", (320, 180), "#F59E0B")
    image_bytes = io.BytesIO()
    image.save(image_bytes, "PNG")

    workbook = Workbook()
    workbook.active.title = "FIXTURE"
    workbook.active.append(["Name", "Value"])
    workbook.active.append(["Sanitized", 1])
    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    return {
        "preview.pdf": ("preview.pdf", pdf_bytes, "application/pdf"),
        "preview.png": ("preview.png", image_bytes.getvalue(), "image/png"),
        "preview.xlsx": (
            "preview.xlsx",
            workbook_bytes.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "preview.csv": ("preview.csv", b"Name,Value\nSanitized,1\n", "text/csv"),
        "preview.txt": ("preview.txt", b"Sanitized text preview\n", "text/plain"),
        "preview.docx": (
            "preview.docx", b"sanitized unsupported-preview fallback",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ),
    }


def source_library_acceptance(template_id: str) -> dict:
    created = expect(
        requests.post(
            f"{BASE}/api/projects/from-template",
            json={
                "profileId": "EMS_FULL",
                "templateId": template_id,
                "metadata": {
                    "projectName": "Source Library V2 Acceptance",
                    "storeNumber": "TEST-ONLY",
                    "client": "Sanitized Fixture",
                },
            },
        ),
        201,
    ).json()
    project_id = created["id"]
    data = fixtures()
    try:
        multi = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources",
                files=[("files", data[name]) for name in ("preview.pdf", "preview.png")],
                data={
                    "virtualPath": "Drawings/Construction Set",
                    "relativePaths": json.dumps(
                        ["Drawings/Construction Set/preview.pdf", "Survey Photos/preview.png"]
                    ),
                },
            ),
            201,
        ).json()["sources"]
        folder = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources",
                files=[("files", data[name]) for name in ("preview.csv", "preview.txt")],
                data={
                    "relativePaths": json.dumps(
                        ["Programming/tables/preview.csv", "Programming/notes/preview.txt"]
                    )
                },
            ),
            201,
        ).json()["sources"]
        spreadsheet = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources",
                files={"files": data["preview.xlsx"]},
                data={"virtualPath": "Converted Schedules/Original Excel"},
            ),
            201,
        ).json()["sources"][0]
        unsupported = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources",
                files={"files": data["preview.docx"]},
                data={"virtualPath": "Reference Documents"},
            ),
            201,
        ).json()["sources"][0]

        bundle = io.BytesIO()
        with zipfile.ZipFile(bundle, "w") as archive:
            archive.writestr("Drawings/Bid Set/zip-plan.txt", "sanitized")
            archive.writestr("Manufacturer Data/device.csv", "Part,Value\nFixture,1\n")
            archive.writestr("../rejected.txt", "reject")
        bundle.seek(0)
        zip_result = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources/import-zip",
                files={"file": ("fixture.zip", bundle.getvalue(), "application/zip")},
            ),
            201,
        ).json()
        assert len(zip_result["sources"]) == 2 and zip_result["report"]["rejected"]

        replacement = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources",
                files={"files": ("preview.txt", b"version two", "text/plain")},
                data={"virtualPath": "Programming/notes"},
            ),
            201,
        ).json()["sources"][0]
        assert replacement["version"] == 2

        traversal = requests.post(
            f"{BASE}/api/projects/{project_id}/sources",
            files={"files": ("escape.txt", b"reject", "text/plain")},
            data={"relativePaths": json.dumps(["../escape.txt"])},
        )
        assert traversal.status_code == 400

        all_sources = expect(requests.get(f"{BASE}/api/projects/{project_id}/sources")).json()
        active = [item for item in all_sources["sources"] if item["status"] == "active"]
        previews = {}
        for item in active:
            if item["originalFileName"] in {
                "preview.pdf", "preview.png", "preview.xlsx", "preview.csv",
                "preview.txt", "preview.docx",
            }:
                result = expect(
                    requests.get(
                        f"{BASE}/api/projects/{project_id}/sources/{item['id']}/preview"
                    )
                ).json()
                previews[item["originalFileName"]] = result["kind"]
        assert previews == {
            "preview.pdf": "pdf",
            "preview.png": "images",
            "preview.xlsx": "spreadsheets",
            "preview.csv": "csv",
            "preview.txt": "documents",
            "preview.docx": "documents",
        }

        queued = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/conversion-queue",
                json={"sourceId": spreadsheet["id"], "notes": "Sanitized queue fixture"},
            ),
            201,
        ).json()["item"]
        archived = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources/{unsupported['id']}/archive"
            )
        ).json()["source"]
        restored = expect(
            requests.post(
                f"{BASE}/api/projects/{project_id}/sources/{unsupported['id']}/restore"
            )
        ).json()["source"]
        assert archived["status"] == "archived" and restored["status"] == "active"
        return {
            "projectId": project_id,
            "multiFileUpload": len(multi),
            "folderUpload": [item["virtualPath"] for item in folder],
            "zipImported": len(zip_result["sources"]),
            "zipRejected": zip_result["report"]["rejected"],
            "relativePathsPreserved": all(
                value in {item["virtualPath"] for item in all_sources["sources"]}
                for value in ("Programming/tables", "Programming/notes", "Drawings/Bid Set")
            ),
            "versionedReplacement": replacement["version"],
            "traversalStatus": traversal.status_code,
            "previews": previews,
            "conversionQueueId": queued["id"],
            "archiveRestore": restored["status"],
        }
    finally:
        archived = requests.post(f"{BASE}/api/projects/{project_id}/archive")
        if archived.status_code != 200:
            raise AssertionError(f"Fixture archive failed: {archived.status_code} {archived.text}")


def sa31_acceptance() -> dict:
    project_id = migrated_project_id()
    staged_hash = sha256(STAGED)
    old_hash = sha256(OLD_WORKBOOK)
    backup_old = next(BACKUP_ROOT.rglob("SA31_EMS_Lighting_Workbook_V1.xlsx"))
    assert old_hash == sha256(backup_old)

    project = expect(requests.get(f"{BASE}/api/projects/{project_id}")).json()
    before_objects = sum(
        len(page.get("canvasObjects") or page.get("objects") or [])
        for page in project["pages"]
    )
    before_assets = {
        path.name: sha256(path)
        for path in Path(project["projectFolder"]).joinpath("assets").rglob("*")
        if path.is_file()
    }

    document = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    original_revision = document["revision"]
    meta = next(sheet for sheet in document["sheets"] if sheet["name"] == "00_PROJECT_META")
    marker_cell = "R40"
    previous_marker = meta["cells"].get(marker_cell)
    meta["cells"][marker_cell] = {"v": "SA31 migration persistence acceptance"}
    saved = expect(
        requests.put(
            f"{BASE}/api/projects/{project_id}/workbook",
            json={
                "projectId": project_id,
                "expectedRevision": original_revision,
                "document": document,
            },
        )
    ).json()
    reloaded = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    assert (
        next(sheet for sheet in reloaded["sheets"] if sheet["name"] == "00_PROJECT_META")
        ["cells"][marker_cell]["v"]
        == "SA31 migration persistence acceptance"
    )
    conflict = requests.put(
        f"{BASE}/api/projects/{project_id}/workbook",
        json={
            "projectId": project_id,
            "expectedRevision": original_revision,
            "document": document,
        },
    )
    assert conflict.status_code == 409

    preview = expect(requests.post(f"{BASE}/api/projects/{project_id}/compile/preview")).json()
    applied = expect(requests.post(f"{BASE}/api/projects/{project_id}/compile/apply")).json()
    after = applied["project"]
    after_objects = sum(
        len(page.get("canvasObjects") or page.get("objects") or [])
        for page in after["pages"]
    )
    assert before_objects == after_objects == 51
    after_assets = {
        path.name: sha256(path)
        for path in Path(after["projectFolder"]).joinpath("assets").rglob("*")
        if path.is_file()
    }
    assert before_assets == after_assets

    current = expect(requests.get(f"{BASE}/api/projects/{project_id}/workbook")).json()
    current_meta = next(
        sheet for sheet in current["sheets"] if sheet["name"] == "00_PROJECT_META"
    )
    if previous_marker is None:
        current_meta["cells"].pop(marker_cell, None)
    else:
        current_meta["cells"][marker_cell] = previous_marker
    expect(
        requests.put(
            f"{BASE}/api/projects/{project_id}/workbook",
            json={
                "projectId": project_id,
                "expectedRevision": current["revision"],
                "document": current,
            },
        )
    )

    write_result = expect(
        requests.post(f"{BASE}/api/projects/{project_id}/workbook/write-excel")
    ).json()
    workbook_path = Path(after["workbookSync"]["workbook"])
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet_names = workbook.sheetnames
    assert len(sheet_names) == 38
    assert sheet_names[:4] == [
        "00_PROJECT_META", "00_TEMPLATE_PROFILE", "00_INDEX", "00_STYLE_GUIDE"
    ]
    assert "28_RDM_PARTS" in sheet_names
    index = workbook["00_INDEX"]
    rows = list(index.iter_rows(min_row=5, values_only=True))
    index_rows = [row for row in rows if row[0] not in (None, "")]
    assert len(index_rows) == 29
    assert sum(str(row[0]).strip().upper() in {"YES", "TRUE", "1"} for row in index_rows) == 24
    workbook.close()
    assert sha256(STAGED) == staged_hash and sha256(OLD_WORKBOOK) == old_hash

    project_dir = Path(after["projectFolder"])
    pdf_response = expect(
        requests.post(
            f"{BASE}/api/projects/{project_id}/export/pdf",
            json={"width": 17, "height": 11},
        )
    )
    pdf_path = project_dir / "exports" / "pdf" / "SA31_schema_v2_acceptance.pdf"
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    pdf_path.write_bytes(pdf_response.content)
    pdf = fitz.open(pdf_path)
    pdf_pages = pdf.page_count
    pdf.close()
    assert pdf_pages == 24

    package_response = expect(
        requests.post(f"{BASE}/api/projects/{project_id}/export/package")
    )
    package_path = project_dir / "exports" / "package" / "SA31_schema_v2_package.zip"
    package_path.parent.mkdir(parents=True, exist_ok=True)
    package_path.write_bytes(package_response.content)
    with zipfile.ZipFile(package_path) as package:
        assert package.testzip() is None

    mirror_root = DOCS / "test_external_mirrors"
    mirror_root.mkdir(parents=True, exist_ok=True)
    mirror = expect(
        requests.post(
            f"{BASE}/api/projects/{project_id}/export/external-mirror",
            json={"destination": str(mirror_root.resolve())},
        )
    ).json()
    manifest_path = Path(mirror["manifestPath"])
    assert manifest_path.is_file() and sha256(manifest_path) == mirror["manifestSha256"]

    sources = expect(requests.get(f"{BASE}/api/projects/{project_id}/sources")).json()
    assert len(sources["sources"]) == 2
    source_previews = {}
    for source in sources["sources"]:
        result = expect(
            requests.get(
                f"{BASE}/api/projects/{project_id}/sources/{source['id']}/preview"
            )
        ).json()
        source_previews[source["originalFileName"]] = result["kind"]

    return {
        "projectId": project_id,
        "schemaVersion": after["schemaVersion"],
        "pageCount": len(after["pages"]),
        "included": sum(bool(page.get("include")) for page in after["pages"]),
        "excluded": sum(not bool(page.get("include")) for page in after["pages"]),
        "manualObjectsBefore": before_objects,
        "manualObjectsAfter": after_objects,
        "assetFiles": len(after_assets),
        "workbookRevisionBefore": original_revision,
        "workbookRevisionSaved": saved["revision"],
        "conflictStatus": conflict.status_code,
        "compileOperations": len(preview["operations"]),
        "compileBackup": applied["backupPath"],
        "workbookPath": str(workbook_path),
        "workbookSha256": write_result["sha256"],
        "workbookSheets": len(sheet_names),
        "indexRows": len(index_rows),
        "stagedWorkbookUnchanged": sha256(STAGED) == staged_hash,
        "oldWorkbookUnchanged": sha256(OLD_WORKBOOK) == old_hash,
        "sourceCount": len(sources["sources"]),
        "sourcePreviews": source_previews,
        "pdfPath": str(pdf_path),
        "pdfBytes": pdf_path.stat().st_size,
        "pdfPages": pdf_pages,
        "packagePath": str(package_path),
        "packageBytes": package_path.stat().st_size,
        "mirror": mirror,
    }


def main() -> int:
    health = expect(requests.get(f"{BASE}/api/health")).json()
    profiles = expect(requests.get(f"{BASE}/api/project-template-profiles")).json()["profiles"]
    assert any(item["id"] == "EMS_LIGHTING" for item in profiles)
    templates = expect(requests.get(f"{BASE}/api/workbook-templates")).json()["templates"]
    template = next(item for item in templates if item["templateId"] == "SINGH360_BASE_V1")
    assert "EMS_LIGHTING" in template["supportedProfiles"]
    report = {
        "health": health,
        "profileIds": [item["id"] for item in profiles],
        "runtimeTemplateSha256": template["sha256"],
        "sourceLibrary": source_library_acceptance(template["templateId"]),
        "sa31": sa31_acceptance(),
    }
    output = DOCS / "test_evidence" / "sa31_schema_v2_source_library_e2e.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
