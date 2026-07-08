"""Smoke: hard no-clipping guarantee across a representative export
(FINAL RENDER POLISH 4G, Phase C).

Verifies:
  - A Workflow/Milestones page (text family, splitMode=none) shows every
    row with no diagnostics-reported clipping and a non-negative bottom gap.
  - A dense LCP Panel Schedule page has a real safety gap above the title
    block (bottomGap >= 0), not flush/tight against it.
  - Across every page in the export, ``clipping`` is never true and
    ``bottomGap`` is never negative (the diagnostic surface the editor/export
    pipeline uses to raise "TABLE OVERFLOW — NOT EXPORTED CLIPPED").
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.page_composer import MIN_BOTTOM_GAP, page_render_diagnostics
from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")


def _section_row(ws, row: int, title: str, cols: int = 6) -> None:
    ws.cell(row, 1, title)
    ws.cell(row, 1).fill = GOLD
    ws.cell(row, 1).font = Font(bold=True)
    for c in range(2, cols + 1):
        ws.cell(row, c).fill = GOLD


def _headers(ws, row: int, cols: int = 6) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c, f"Header {c}")
        cell.fill = GRAY
        cell.font = Font(bold=True)


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Code", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "EMS 0.5", "Project Workflow", "Project Workflow / Milestones", "YES", ""])
    idx.append(["2", "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "YES", ""])

    workflow = wb.create_sheet("Project Workflow")
    workflow.append(["Step", "Milestone", "Responsible", "Target Date"])
    for c in workflow[1]:
        c.fill = GRAY
        c.font = Font(bold=True)
    for r in range(1, 12):
        workflow.append([str(r), f"Milestone {r}", "PM Team", f"2026-0{(r % 9) + 1}-01"])

    lcp = wb.create_sheet("LCP Panel Schedule")
    _section_row(lcp, 1, "LCP-1 Dimming Panel")
    _headers(lcp, 2)
    r = 3
    for i in range(18):
        for c in range(1, 7):
            lcp.cell(r, c, f"LCP1 {i}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1
    _section_row(lcp, r, "LCP-2 Contactor Panel")
    r += 1
    _headers(lcp, r)
    r += 1
    for i in range(26):
        for c in range(1, 7):
            lcp.cell(r, c, f"RO{i + 1}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1

    wb.save(path)


def _sa38_style_workbook(path: Path) -> None:
    """SA38-shaped fixture (FINAL RELEASE CLEANUP 4H+SA38, Phase H): BOM,
    TOC, a Responsibility Matrix, and a Rack-style sheet with repeating
    merged-gold-band controller sections — the exact shapes that previously
    risked clipping under the new Phase F section-aware splitter.
    """
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Code", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "EMS 0.0", "BOM", "Bill of Materials", "YES", ""])
    idx.append(["2", "EMS 0.1", "TOC", "Sheet Index / TOC", "YES", ""])
    idx.append(["3", "EMS 0.3", "Responsibility Matrix", "HEB Responsibility Matrix", "YES", ""])
    idx.append(["4", "EMS 2.0", "RACK A", "Rack A I/O & Layout", "YES", ""])
    idx.append(["5", "EMS 4.0", "DLE Controllers", "DLE Controllers I/O", "YES", ""])

    bom = wb.create_sheet("BOM")
    bom.append(["Item", "Description", "Qty", "Vendor", "Part #"])
    for c in bom[1]:
        c.fill = GRAY
        c.font = Font(bold=True)
    for i in range(40):
        bom.append([str(i + 1), f"Component {i}", str((i % 5) + 1), "Vendor Co", f"P-{i:04d}"])

    toc = wb.create_sheet("TOC")
    toc.append(["Sheet Code", "Page Title"])
    for i in range(20):
        toc.append([f"EMS {i}.0", f"Sheet {i}"])

    resp = wb.create_sheet("Responsibility Matrix")
    resp.append(["Task", "HEB", "Singh360", "Electrician", "Notes"])
    for c in resp[1]:
        c.fill = GRAY
        c.font = Font(bold=True)
    for i in range(25):
        resp.append([f"Task {i}", "X" if i % 2 else "", "" if i % 2 else "X", "X" if i % 3 == 0 else "", ""])

    rack = wb.create_sheet("RACK A")
    r = 1
    for name in (
        "SuperPak Controller I/O Panel for RACK A",
        "SuperPak Controller I/O Panel for RACK B",
        "Expansion I/O Device PR0663",
    ):
        rack.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = rack.cell(r, 1, name)
        cell.fill = GOLD
        cell.font = Font(bold=True)
        r += 1
        _headers(rack, r)
        r += 1
        for i in range(20):
            for c in range(1, 7):
                rack.cell(r, c, f"{name[:6]}-{i}-{c}")
            rack.row_dimensions[r].height = 30
            r += 1

    dle = wb.create_sheet("DLE Controllers")
    dle.append(["DLE Controllers"])
    r = 2
    for i in range(45):
        dle.append([f"DLE Point {i}", "AI", f"{i}", "OK", "", ""])
        r += 1

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "no_clip.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="noclip1")
    pages = proj["pages"]
    diagnostics = page_render_diagnostics(pages)
    problems: list[str] = []

    # Workflow: text family, splitMode=none — every row must render, no clip.
    workflow_pages = [p for p in pages if p["sheetTab"] == "Project Workflow"]
    if len(workflow_pages) != 1:
        problems.append(f"Workflow expected 1 page, got {len(workflow_pages)}")
    else:
        block = workflow_pages[0]["blocks"][0]
        # header + 11 milestone rows
        if len(block.get("grid") or []) != 12:
            problems.append(f"Workflow grid rows = {len(block.get('grid') or [])}, expected 12 (all 11 rows + header)")
        d = next((x for x in diagnostics if x["pageTitle"] == "Project Workflow / Milestones"), None)
        if d is None:
            problems.append("Workflow diagnostics entry not found")
        elif d.get("clipping"):
            problems.append("Workflow diagnostics reported clipping=True")
        elif d.get("bottomGap", -1) < 0:
            problems.append(f"Workflow bottomGap negative: {d.get('bottomGap')}")

    # LCP: dense panel schedule — real safety gap above the title block.
    lcp_pages = [p for p in pages if p["sheetTab"] == "LCP Panel Schedule"]
    if not lcp_pages:
        problems.append("LCP Panel Schedule page not found")
    else:
        lcp_diag = [d for d in diagnostics if d["pageTitle"].startswith("LCP Panel Schedule")]
        for d in lcp_diag:
            if d.get("clipping"):
                problems.append(f"LCP page {d['sheetCode']} diagnostics reported clipping=True")
            if d.get("bottomGap", -1) < 0:
                problems.append(f"LCP page {d['sheetCode']} bottomGap negative: {d.get('bottomGap')}")

    # Global guarantee: no page in the export may report clipping=True or a
    # negative bottom gap (the "TABLE OVERFLOW — NOT EXPORTED CLIPPED" signal
    # must never be silently swallowed).
    for d in diagnostics:
        if d.get("clipping"):
            problems.append(f"{d['sheetCode']} {d['pageTitle']}: clipping=True")
        if d.get("bottomGap", 0) < 0:
            problems.append(f"{d['sheetCode']} {d['pageTitle']}: bottomGap={d.get('bottomGap')} < 0")

    if MIN_BOTTOM_GAP <= 0:
        problems.append("MIN_BOTTOM_GAP is not a positive safety margin")

    # SA38-shaped fixture: BOM/TOC/Responsibility Matrix/Rack/DLE — same
    # no-clip guarantee must hold under the Phase F section-aware splitter.
    xlsx_sa38 = tmp / "sa38_no_clip.xlsx"
    _sa38_style_workbook(xlsx_sa38)
    proj_sa38 = import_workbook(xlsx_sa38, project_id="sa38noclip")
    pages_sa38 = proj_sa38["pages"]
    diagnostics_sa38 = page_render_diagnostics(pages_sa38)

    if not pages_sa38:
        problems.append("SA38-style fixture produced no pages")
    for d in diagnostics_sa38:
        if d.get("clipping"):
            problems.append(f"SA38 fixture {d['sheetCode']} {d['pageTitle']}: clipping=True")
        if d.get("bottomGap", 0) < 0:
            problems.append(f"SA38 fixture {d['sheetCode']} {d['pageTitle']}: bottomGap={d.get('bottomGap')} < 0")

    rack_pages_sa38 = [p for p in pages_sa38 if p["sheetTab"] == "RACK A"]
    if len(rack_pages_sa38) < 2:
        problems.append(f"SA38 RACK A expected to split across >1 page, got {len(rack_pages_sa38)}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - no export clipping passed")


if __name__ == "__main__":
    main()
