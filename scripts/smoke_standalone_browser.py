"""Disposable browser proof for the standalone Singh360 drawing-set editor.

This smoke intentionally creates every input it uses.  It starts ``server.py``
on an alternate port with a temporary ``SINGH360_DOCS_DIR`` and never opens,
copies, or writes a customer workbook or the repository's live ``.docs``.

Evidence is retained outside the repository.  The workflow covers full and
name-only onboarding, project-local logo persistence, delayed autosave races,
blank/text/image/CSV/XLSX/template page workflows, continuation-group
archive/restore, multi-page PDF import/partial replacement, component and
saved-assembly insertion, copy/paste, save/reopen/restart persistence, and the
final complete PDF download.  Every browser request is recorded and the test
fails if normal use touches a workbook-link, workbook-quality, or sync endpoint.
"""
from __future__ import annotations

import argparse
from datetime import datetime, timezone
from hashlib import sha256
import json
import os
from pathlib import Path
import re
import socket
import subprocess
import sys
import tempfile
import time
import traceback
from typing import Any, Callable
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

import fitz
from PIL import Image, ImageDraw
from playwright.sync_api import Page, sync_playwright


ROOT = Path(__file__).resolve().parents[1]
COMPONENT_ID = "fixture-standalone-browser-sensor"
COMPONENT_NAME = "Standalone Browser Sensor"
PROJECT_NAME = "Disposable Standalone Browser Drawing Set"
NAME_ONLY_PROJECT = "Disposable Name Only Drawing Set"
FINAL_PACKAGE_NAME = "Standalone_Browser_Evidence_Final"
ASSEMBLY_NAME = "Disposable Browser Sensor Assembly"
CSV_FILE_NAME = "Disposable Browser Schedule.csv"
CSV_MARKER = "BROWSER-CSV-01"
IMAGE_FILE_NAME = "Disposable Browser Image.png"
TEXT_PAGE_TITLE = "Disposable Browser Text Page"
TEMPLATE_NAME = "Disposable Browser Text Template"
WORKBOOK_FILE_NAME = "Disposable Browser Geometry.xlsx"
WORKBOOK_SHEET_NAME = "Imported Schedule"
WORKBOOK_PAGE_TITLE = "Browser Imported Geometry Proof"
WORKBOOK_MARKER = "BROWSER XLSX GEOMETRY PROOF"
INDEX_REQUIRED_TOKENS = (
    "Browser Layout Renamed",
    TEXT_PAGE_TITLE,
    f"{TEXT_PAGE_TITLE} Copy",
    TEMPLATE_NAME,
    "Equipment Summary",
    Path(IMAGE_FILE_NAME).stem,
    WORKBOOK_PAGE_TITLE,
    "Imported Browser Page One",
)
FORBIDDEN_REQUEST_PARTS = ("/workbook-link", "/workbook-quality", "/sync")
STALE_COVER_LINE = "STALE LEGACY WORKBOOK COVER LINE — MUST STAY HIDDEN"
DELAYED_SAVE_NOTES = "Edit B reached the autosave queue while edit A was in flight."
FULL_PROJECT_METADATA = {
    "projectName": PROJECT_NAME,
    "client": "Sanitized Browser Customer",
    "storeNumber": "BROWSER-001",
    "location": "100 Disposable Test Way, Testville, TX 75001",
    "projectType": "Standalone EMS Acceptance",
    "drawingSetTitle": "Standalone Browser Acceptance Set",
    "preparedBy": "Browser Fixture Author",
    "createdBy": "Browser Fixture Author",
    "checkedBy": "Browser Fixture Checker",
    "createdDate": "2026-08-01",
    "revision": "R7",
    "notes": "Generated disposable onboarding metadata; no customer content.",
    "drawingPackageFileName": "Old_Standalone_Browser_Evidence",
}
FINAL_PROJECT_METADATA = {
    **FULL_PROJECT_METADATA,
    "drawingPackageFileName": FINAL_PACKAGE_NAME,
}
WIZARD_OPTIONAL_FIELDS = (
    "client",
    "storeNumber",
    "location",
    "projectType",
    "drawingSetTitle",
    "preparedBy",
    "checkedBy",
    "createdDate",
    "revision",
    "notes",
    "drawingPackageFileName",
    "customerLogoAsset",
)


def utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def default_evidence_dir() -> Path:
    base = Path(os.environ.get("USERPROFILE") or tempfile.gettempdir())
    return base / "Singh360_Draft_TestEvidence" / f"standalone-browser-smoke-{utc_stamp()}"


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def tree_manifest(root: Path) -> dict[str, str]:
    if not root.is_dir():
        return {}
    return {
        path.relative_to(root).as_posix(): file_sha256(path)
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False), encoding="utf-8")


def free_port() -> int:
    while True:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind(("127.0.0.1", 0))
            port = int(sock.getsockname()[1])
        if port != 8766:
            return port


def http_json(port: int, path: str, *, timeout: int = 30) -> dict[str, Any]:
    with urlopen(f"http://127.0.0.1:{port}{path}", timeout=timeout) as response:
        return json.load(response)


def http_post_json(
    port: int,
    path: str,
    payload: dict[str, Any] | None = None,
    *,
    timeout: int = 60,
) -> dict[str, Any]:
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {"Accept": "application/json"}
    if body is not None:
        headers["Content-Type"] = "application/json"
    request = Request(
        f"http://127.0.0.1:{port}{path}",
        data=body,
        headers=headers,
        method="POST",
    )
    with urlopen(request, timeout=timeout) as response:
        return json.load(response)


def wait_health(port: int, *, timeout: float = 45.0) -> dict[str, Any]:
    deadline = time.monotonic() + timeout
    last_error = ""
    while time.monotonic() < deadline:
        try:
            payload = http_json(port, "/api/health", timeout=2)
            if payload.get("ok") is True and int(payload.get("configuredPort") or 0) == port:
                return payload
        except Exception as exc:  # noqa: BLE001 - health polling records final cause
            last_error = str(exc)
        time.sleep(0.2)
    raise RuntimeError(f"disposable server did not become healthy on {port}: {last_error}")


def wait_for(condition: Callable[[], bool], message: str, *, timeout: float = 30.0) -> None:
    deadline = time.monotonic() + timeout
    last_error = ""
    while time.monotonic() < deadline:
        try:
            if condition():
                return
        except Exception as exc:  # noqa: BLE001 - eventual state may not exist yet
            last_error = str(exc)
        time.sleep(0.2)
    suffix = f" Last error: {last_error}" if last_error else ""
    raise AssertionError(f"{message}.{suffix}")


def wait_for_browser_download(page: Page, downloads: list[Any], *, timeout: float = 240.0) -> None:
    """Pump Playwright's sync event loop until its download callback fires."""
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if downloads:
            return
        page.wait_for_timeout(200)
    raise AssertionError("browser did not receive the regenerated PDF download")


class DisposableServer:
    def __init__(self, docs: Path, port: int, evidence: Path) -> None:
        self.docs = docs
        self.port = port
        self.evidence = evidence
        self.process: subprocess.Popen[Any] | None = None
        self._stdout: Any = None
        self._stderr: Any = None
        self.pids: list[int] = []
        self.launcher_pids: list[int] = []

    def start(self) -> dict[str, Any]:
        if self.process is not None:
            raise RuntimeError("server is already running")
        self._stdout = (self.evidence / "server.stdout.log").open("a", encoding="utf-8")
        self._stderr = (self.evidence / "server.stderr.log").open("a", encoding="utf-8")
        self._stdout.write(f"\n===== START {utc_stamp()} =====\n")
        self._stdout.flush()
        environment = {
            **os.environ,
            "PYTHONUNBUFFERED": "1",
            "SINGH360_DOCS_DIR": str(self.docs),
            "SINGH360_PORT": str(self.port),
        }
        self.process = subprocess.Popen(
            [sys.executable, str(ROOT / "server.py")],
            cwd=str(ROOT),
            env=environment,
            stdout=self._stdout,
            stderr=self._stderr,
        )
        health = wait_health(self.port)
        pid = int(health.get("pid") or self.process.pid)
        # When a Windows virtualenv interpreter is launched from WSL,
        # subprocess reports the interop launcher PID while os.getpid() in the
        # Flask process reports its native Windows PID.  The launcher handle
        # still owns and terminates the child tree; retain both identities and
        # use the health PID to prove the serving process changed on restart.
        self.launcher_pids.append(int(self.process.pid))
        self.pids.append(pid)
        return health

    def stop(self) -> None:
        process = self.process
        self.process = None
        if process is not None and process.poll() is None:
            process.terminate()
            try:
                process.wait(timeout=15)
            except subprocess.TimeoutExpired:
                process.kill()
                process.wait(timeout=10)
        for handle_name in ("_stdout", "_stderr"):
            handle = getattr(self, handle_name)
            if handle is not None:
                handle.flush()
                handle.close()
                setattr(self, handle_name, None)


def make_fixture_pdf(path: Path, *, revision: str, page_count: int = 3) -> None:
    """Create a deterministic ANSI-B landscape vector PDF."""
    path.parent.mkdir(parents=True, exist_ok=True)
    document = fitz.open()
    palette = [(0.05, 0.31, 0.55), (0.08, 0.48, 0.30), (0.56, 0.22, 0.12)]
    for index in range(page_count):
        page = document.new_page(width=17 * 72, height=11 * 72)
        color = palette[index]
        token = f"{revision.upper()} PDF PAGE {index + 1}"
        page.draw_rect(fitz.Rect(28, 28, page.rect.width - 28, page.rect.height - 28), color=color, width=4)
        page.draw_rect(fitz.Rect(62, 130, page.rect.width - 62, page.rect.height - 92), color=color, width=2)
        page.insert_text((72, 92), token, fontsize=30, fontname="helv", color=color)
        page.insert_text((72, 122), "SANITIZED STANDALONE BROWSER FIXTURE", fontsize=14, fontname="helv")
        page.insert_text((90, 175), f"Vector detail row {index + 1} / {revision}", fontsize=18, fontname="helv")
        for row in range(7):
            y = 220 + row * 58
            page.draw_line((90, y), (page.rect.width - 90, y), color=color, width=1.5)
            page.insert_text((105, y - 10), f"FIXTURE-{index + 1}-{row + 1:02d}", fontsize=13, fontname="cour")
        page.insert_text(
            (72, page.rect.height - 48),
            f"CONTENT SAFETY BORDER / {revision.upper()} / SOURCE PAGE {index + 1} OF {page_count}",
            fontsize=12,
            fontname="helv",
        )
    document.set_metadata({"title": "Sanitized Singh360 browser fixture", "author": "Automated disposable test"})
    document.save(path, garbage=4, deflate=True)
    document.close()


def make_fixture_image(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.new("RGB", (960, 540), "#f7f9fc")
    draw = ImageDraw.Draw(image)
    draw.rectangle((18, 18, 942, 522), outline="#174f78", width=12)
    draw.rectangle((90, 145, 870, 405), fill="#f4b183", outline="#174f78", width=8)
    draw.text((115, 80), "SANITIZED PROJECT-LOCAL IMAGE", fill="#174f78")
    draw.text((250, 255), "BROWSER IMAGE PROOF", fill="#111827")
    image.save(path, format="PNG", optimize=True)


def make_fixture_logo(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.new("RGBA", (180, 72), "#ffffff")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((3, 3, 176, 68), radius=10, fill="#174f78", outline="#f4b183", width=4)
    draw.text((20, 25), "BROWSER CUSTOMER", fill="#ffffff")
    image.save(path, format="PNG", optimize=True)


def make_fixture_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "Tag,Description,Quantity\n"
        f"{CSV_MARKER},Disposable Browser Controller,2\n"
        "BROWSER-CSV-02,Disposable Browser Sensor,4\n",
        encoding="utf-8",
    )


def make_fixture_workbook(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    metadata = workbook.active
    metadata.title = "00_PROJECT_META"
    metadata.append(["Project Name", "Disposable Browser Workbook Fixture"])
    metadata.append(["Workbook Schema Version", "2"])
    metadata.append(["Help Version", "browser-smoke"])

    index = workbook.create_sheet("00_INDEX")
    index["A1"] = "SINGH360 DRAFT — SANITIZED BROWSER FIXTURE"
    headers = ["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"]
    for column, value in enumerate(headers, start=1):
        index.cell(4, column, value)
    index.append(["YES", 1, "XLSX-101", WORKBOOK_SHEET_NAME, WORKBOOK_PAGE_TITLE, "Generated", "table", ""])

    sheet = workbook.create_sheet(WORKBOOK_SHEET_NAME)
    sheet.merge_cells("A1:C1")
    sheet["A1"] = WORKBOOK_MARKER
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="F4B183")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32.25
    headings = ["Tag", "Description", "Quantity"]
    for column, value in enumerate(headings, start=1):
        cell = sheet.cell(3, column, value)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.append(["XLSX-ROW-01", "Preserved source geometry and readable wrapping", 2])
    sheet.append(["XLSX-ROW-02", "Project-local one-time worksheet import", 3])
    sheet["C6"] = "=SUM(C4:C5)"
    sheet["A6"] = "TOTAL"
    sheet.column_dimensions["A"].width = 18.5
    sheet.column_dimensions["B"].width = 48.25
    sheet.column_dimensions["C"].width = 16.0
    sheet.row_dimensions[3].height = 24.0
    sheet.row_dimensions[4].height = 30.75
    sheet.row_dimensions[5].height = 27.0
    sheet.print_area = "A1:C6"
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    workbook.save(path)
    workbook.close()


def seed_component_library(docs: Path) -> None:
    symbol = docs / "library" / "symbols" / "symbols_markers" / "standalone_browser_sensor.svg"
    symbol.parent.mkdir(parents=True, exist_ok=True)
    symbol.write_text(
        """<svg xmlns="http://www.w3.org/2000/svg" width="120" height="90" viewBox="0 0 120 90">
<rect x="3" y="3" width="114" height="84" rx="10" fill="#ffffff" stroke="#174f78" stroke-width="6"/>
<circle cx="34" cy="45" r="19" fill="#eaf4fb" stroke="#174f78" stroke-width="4"/>
<path d="M60 27h42M60 45h42M60 63h30" stroke="#174f78" stroke-width="5" stroke-linecap="round"/>
<text x="34" y="51" text-anchor="middle" font-family="Arial" font-size="18" font-weight="700" fill="#174f78">SB</text>
</svg>""",
        encoding="utf-8",
    )
    manifest = docs / "library" / "manifest.json"
    manifest.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        manifest,
        {
            "version": 2,
            "components": [
                {
                    "id": COMPONENT_ID,
                    "displayName": COMPONENT_NAME,
                    "category": "symbols_markers",
                    "categories": ["symbols_markers"],
                    "collection": "Disposable Browser Fixtures",
                    "tags": ["sanitized", "browser", "standalone"],
                    "defaultLabel": "SB",
                    "defaultWidth": 120,
                    "defaultHeight": 90,
                    "sourceFile": "symbols/symbols_markers/standalone_browser_sensor.svg",
                    "approved": True,
                    "needsReview": False,
                    "favorite": False,
                    "status": "active",
                    "retired": False,
                }
            ],
        },
    )


def project_for(port: int, project_id: str) -> dict[str, Any]:
    return http_json(port, f"/api/projects/{project_id}")


def page_by_id(project: dict[str, Any], page_id: str) -> dict[str, Any]:
    for item in project.get("pages") or []:
        if item.get("id") == page_id:
            return item
    raise AssertionError(f"project no longer contains page {page_id}")


def sorted_pages(project: dict[str, Any]) -> list[dict[str, Any]]:
    return sorted(project.get("pages") or [], key=lambda item: float(item.get("order") or 0))


def pdf_pages(project: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        item
        for item in sorted_pages(project)
        if str((item.get("sourceImport") or {}).get("type") or "").lower() == "pdf"
    ]


def component_objects(item: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        obj
        for obj in item.get("canvasObjects") or []
        if obj.get("objName") == COMPONENT_NAME
    ]


def collect_object_ids(value: Any) -> set[str]:
    found: set[str] = set()
    if isinstance(value, list):
        for item in value:
            found.update(collect_object_ids(item))
    elif isinstance(value, dict):
        object_id = str(value.get("objectId") or "").strip()
        if object_id:
            found.add(object_id)
        for item in value.values():
            found.update(collect_object_ids(item))
    return found


def project_content_signature(project: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "id": item.get("id"),
            "order": item.get("order"),
            "include": item.get("include"),
            "sheetCode": item.get("sheetCode"),
            "sheetTitle": item.get("sheetTitle"),
            "objectIds": sorted(collect_object_ids(item.get("canvasObjects") or [])),
            "sourceSha256": (item.get("sourceImport") or {}).get("sha256"),
            "projectLocalPath": (item.get("sourceImport") or {}).get("projectLocalPath"),
        }
        for item in sorted_pages(project)
    ]


def assert_wizard_metadata(
    project: dict[str, Any],
    expected: dict[str, str],
    *,
    require_logo: bool,
) -> str:
    metadata = project.get("metadata") or {}
    mismatches = {
        key: {"expected": value, "actual": metadata.get(key)}
        for key, value in expected.items()
        if metadata.get(key) != value
    }
    if mismatches:
        raise AssertionError(f"wizard metadata mismatch: {mismatches!r}")
    logo_asset = str(metadata.get("customerLogoAsset") or "")
    if require_logo and not re.fullmatch(
        rf"/api/assets/{re.escape(str(project.get('id') or ''))}/[a-f0-9]{{16}}\.png",
        logo_asset,
    ):
        raise AssertionError(f"wizard customer logo is not a project-local asset: {logo_asset!r}")
    if not require_logo and logo_asset:
        raise AssertionError(f"name-only project unexpectedly has a customer logo: {logo_asset!r}")
    return logo_asset


def assert_name_only_metadata(project: dict[str, Any]) -> None:
    metadata = project.get("metadata") or {}
    if metadata.get("projectName") != NAME_ONLY_PROJECT:
        raise AssertionError("name-only onboarding changed the required project name")
    not_blank = {
        key: metadata.get(key)
        for key in WIZARD_OPTIONAL_FIELDS
        if str(metadata.get(key) or "").strip()
    }
    if not_blank:
        raise AssertionError(f"name-only onboarding populated optional metadata: {not_blank!r}")


def assert_advanced_tools_collapsed(page: Page, where: str) -> None:
    advanced = page.locator("details.ribbon-advanced")
    advanced.wait_for(state="visible", timeout=30_000)
    if advanced.get_attribute("open") is not None:
        raise AssertionError(f"Advanced Tools was expanded by default {where}")


def open_advanced_tools(page: Page) -> None:
    advanced = page.locator("details.ribbon-advanced")
    advanced.wait_for(state="visible", timeout=30_000)
    if advanced.get_attribute("open") is None:
        advanced.locator("summary").click()
    page.locator(".ribbon-tabs").wait_for(state="visible", timeout=30_000)


def open_properties_panel(page: Page) -> None:
    rail = page.locator(".panel-rail-right")
    if rail.is_visible():
        rail.click()
    page.locator(".panel-right").wait_for(state="visible", timeout=30_000)


def canvas_objects(page: Page) -> list[dict[str, Any]]:
    result = page.evaluate(
        "() => window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []"
    )
    if not isinstance(result, list):
        raise AssertionError("workflow canvas audit did not return an object list")
    return result


def add_canvas_text(page: Page) -> set[str]:
    page.locator('[aria-label="Drawing canvas objects"][data-canvas-hydrated="1"]').wait_for(
        timeout=60_000
    )
    before_objects = canvas_objects(page)
    before = collect_object_ids(before_objects)
    open_advanced_tools(page)
    page.locator(".ribbon-tabs").get_by_role("button", name="Insert", exact=True).click()
    page.locator(".ribbon-groups").get_by_role("button", name="Text", exact=True).click()
    page.wait_for_function(
        "before => (window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []).length > before",
        arg=len(before_objects),
        timeout=30_000,
    )
    added = collect_object_ids(canvas_objects(page)) - before
    if not added:
        raise AssertionError("canvas Text insertion created no fresh stable object ID")
    return added


def logical_code_after(pages: list[dict[str, Any]], anchor_page_id: str) -> str:
    ordered = sorted(pages, key=lambda item: float(item.get("order") or 0))
    candidates = [
        item for item in ordered
        if item.get("id") == anchor_page_id
        and not item.get("continuationOf")
        and str(item.get("managedPage") or "") not in {"cover", "index"}
    ]
    if not candidates:
        raise AssertionError(f"logical sheet-code anchor {anchor_page_id!r} was not found")
    current = str(candidates[0].get("displaySheetCode") or candidates[0].get("sheetCode") or "").strip()
    match = re.fullmatch(r"(.*?)(\d+)(?:\.(\d+))?", current)
    if not match:
        raise AssertionError(f"logical sheet-code anchor {current!r} has no numeric suffix")
    used = {
        str(item.get("displaySheetCode") or item.get("sheetCode") or "").strip().lower()
        for item in ordered
    }
    prefix, whole, fraction = match.groups()
    width = len(fraction or whole)
    number = int(fraction or whole) + 1
    while True:
        suffix = str(number).zfill(width)
        candidate = f"{prefix}{whole}.{suffix}" if fraction is not None else f"{prefix}{suffix}"
        if candidate.lower() not in used:
            return candidate
        number += 1


def install_continuation_fixture(
    port: int,
    project_id: str,
    base_page_id: str,
) -> tuple[dict[str, Any], str]:
    project = project_for(port, project_id)
    base = page_by_id(project, base_page_id)
    child_id = f"{base_page_id}__browser_cont_1"
    if any(item.get("id") == child_id for item in project.get("pages") or []):
        return project, child_id
    pages: list[dict[str, Any]] = []
    for item in sorted_pages(project):
        pages.append(item)
        if item.get("id") != base_page_id:
            continue
        child = json.loads(json.dumps(item))
        child.update(
            {
                "id": child_id,
                "sheetCode": f"{base.get('sheetCode') or 'L-101'}a",
                "displaySheetCode": f"{base.get('displaySheetCode') or base.get('sheetCode') or 'L-101'}a",
                "sheetTitle": f"{base.get('sheetTitle') or 'Browser Layout'} — CONTINUED",
                "sheetTab": "Browser Layout Continued",
                "pageGroupId": base_page_id,
                "continuationOf": base_page_id,
                "continuationIndex": 1,
                "generatedContinuation": True,
                "canvasObjects": [],
                "blocks": [],
                "notes": "Disposable continuation-group browser fixture.",
            }
        )
        pages.append(child)
    for index, item in enumerate(pages, start=1):
        item["order"] = index
    project["pages"] = pages
    saved = http_post_json(port, f"/api/projects/{project_id}", project)
    if not any(item.get("id") == child_id for item in saved.get("pages") or []):
        raise AssertionError("disposable continuation fixture was not saved")
    return saved, child_id


def pdf_page_snapshot(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": item.get("id"),
        "order": item.get("order"),
        "include": item.get("include"),
        "sheetCode": item.get("sheetCode"),
        "displaySheetCode": item.get("displaySheetCode"),
        "sheetTitle": item.get("sheetTitle"),
        "notes": item.get("notes"),
        "sourceImport": item.get("sourceImport"),
        "canvasObjects": item.get("canvasObjects"),
        "assets": item.get("assets"),
    }


def assert_current_pdf_base(item: dict[str, Any]) -> None:
    source = item.get("sourceImport") or {}
    matches = [
        obj
        for obj in item.get("canvasObjects") or []
        if obj.get("pdfBase") is True
    ]
    if len(matches) != 1:
        raise AssertionError(
            f"PDF page {item.get('id')} has {len(matches)} base objects instead of one"
        )
    base = matches[0]
    if base.get("pdfSourceId") != source.get("sourceId"):
        raise AssertionError(
            f"PDF page {item.get('id')} retained stale canvas source "
            f"{base.get('pdfSourceId')!r}; expected {source.get('sourceId')!r}"
        )
    if int(base.get("pdfPage") or 0) != int(source.get("sourcePageIndex") or 0):
        raise AssertionError(f"PDF page {item.get('id')} canvas/source page indices differ")
    current_asset_url = str(source.get("renderAssetUrl") or "")
    page_asset_urls = {
        str(asset.get("url") or "")
        for asset in item.get("assets") or []
        if asset.get("type") == "pdf-preview"
    }
    if current_asset_url not in page_asset_urls:
        raise AssertionError(
            f"PDF page {item.get('id')} preview assets do not reference the current revision"
        )


def sheet_item(page: Page, title: str):
    item = page.locator(".sheet-item").filter(has_text=title).first
    item.wait_for(state="visible", timeout=30_000)
    return item


def open_add_import(page: Page, choice: str):
    page.locator(".ribbon-appbar-right").get_by_role("button", name="Add / Import Page", exact=True).click()
    dialog = page.get_by_role("dialog", name="Add / Import Page")
    dialog.wait_for(state="visible", timeout=30_000)
    dialog.locator(".add-import-choices button").filter(has_text=choice).click()
    return dialog


def wait_for_project_state(
    page: Page,
    port: int,
    project_id: str,
    predicate: Callable[[dict[str, Any]], bool],
    message: str,
    *,
    timeout: float = 45.0,
) -> dict[str, Any]:
    deadline = time.monotonic() + timeout
    latest: dict[str, Any] = {}
    while time.monotonic() < deadline:
        page.wait_for_timeout(200)
        latest = project_for(port, project_id)
        if predicate(latest):
            return latest
    raise AssertionError(f"{message}; latest project had {len(latest.get('pages') or [])} pages")


def save_project(page: Page) -> None:
    button = page.locator(".ribbon-appbar-right .ribbon-primary-save")
    button.wait_for(state="visible", timeout=30_000)
    button.click()
    page.locator(".status-pill").filter(has_text="PROJECT SAVED").wait_for(state="visible", timeout=30_000)


def assert_editor_ready(page: Page, project_name: str) -> None:
    page.locator(".ribbon-appbar-right").wait_for(state="visible", timeout=60_000)
    page.get_by_role("button", name="Save Project", exact=True).first.wait_for(state="visible", timeout=30_000)
    left_rail = page.locator(".panel-rail-left")
    if left_rail.is_visible():
        left_rail.click()
    page.locator(".panel-left").wait_for(state="visible", timeout=30_000)
    page.locator(".sheet-item").first.wait_for(state="visible", timeout=30_000)
    page.locator("body").filter(has_text=project_name).wait_for(state="visible", timeout=30_000)
    if page.locator("vite-error-overlay").count():
        raise AssertionError("Vite error overlay is present")


def rename_and_recode(
    page: Page,
    old_title: str,
    new_title: str,
    code: str,
    prompt_answers: list[str],
) -> None:
    item = sheet_item(page, old_title)
    item.click(button="right")
    prompt_answers.append(new_title)
    page.locator(".ctx-menu").get_by_role("button", name="Rename Sheet Title", exact=True).click()
    item = sheet_item(page, new_title)
    item.click(button="right")
    prompt_answers.append(code)
    page.locator(".ctx-menu").get_by_role("button", name="Edit Sheet Code", exact=True).click()
    sheet_item(page, new_title).filter(has_text=code).wait_for(state="visible", timeout=20_000)


def expand_section(page: Page, title: str) -> None:
    left_rail = page.locator(".panel-rail-left")
    if left_rail.is_visible():
        left_rail.click()
    page.locator(".panel-left").wait_for(state="visible", timeout=30_000)
    title_node = page.locator(".nav-section-title").filter(
        has_text=re.compile(rf"^{re.escape(title)}(?:\s|$)")
    ).first
    title_node.wait_for(state="attached", timeout=30_000)
    title_node.scroll_into_view_if_needed(timeout=30_000)
    title_node.wait_for(state="visible", timeout=30_000)
    header = title_node.locator("xpath=ancestor::button[contains(@class, 'nav-section-head')][1]")
    header.wait_for(state="visible", timeout=30_000)
    if header.get_attribute("aria-expanded") != "true":
        header.click()


def move_page(page: Page, title: str, direction: str) -> None:
    item = sheet_item(page, title)
    controls = item.locator("button.reorder-btn")
    if controls.count() != 2:
        raise AssertionError(
            f"page {title!r} exposes {controls.count()} reorder controls instead of two"
        )
    target = controls.nth(0 if direction == "up" else 1)
    if target.is_disabled():
        raise AssertionError(f"page {title!r} unexpectedly disables Move {direction}")
    target.click()


def import_pdf_via_ui(
    page: Page,
    path: Path,
    *,
    placement: str,
    replace: bool,
    expected_pages: int,
) -> list[dict[str, Any]]:
    page.evaluate(
        r"""() => {
          window.__S360_PDF_PROGRESS_EVIDENCE__ = [];
          window.__S360_PDF_PROGRESS_OBSERVER__?.disconnect();
          const record = () => {
            const node = document.querySelector('.import-progress');
            if (!node) return;
            const progress = node.querySelector('progress');
            const output = node.querySelector('output');
            const entry = {
              phase: node.getAttribute('data-phase') || '',
              text: (node.textContent || '').replace(/\s+/g, ' ').trim(),
              value: progress ? Number(progress.value) : 0,
              max: progress ? Number(progress.max) : 0,
              output: (output?.textContent || '').trim(),
            };
            const evidence = window.__S360_PDF_PROGRESS_EVIDENCE__;
            const prior = evidence[evidence.length - 1];
            if (!prior || JSON.stringify(prior) !== JSON.stringify(entry)) evidence.push(entry);
          };
          const observer = new MutationObserver(record);
          observer.observe(document.body, { childList: true, subtree: true, attributes: true, characterData: true });
          window.__S360_PDF_PROGRESS_OBSERVER__ = observer;
          record();
        }"""
    )
    page.locator(".ribbon-appbar-right").get_by_role("button", name="Add / Import Page", exact=True).click()
    dialog = page.get_by_role("dialog", name="Add / Import Page")
    dialog.wait_for(state="visible", timeout=20_000)
    dialog.locator(".add-import-choices button").filter(has_text="Finished PDF Drawing").click()
    dialog.get_by_label("Choose PDF", exact=True).set_input_files(str(path))
    dialog.locator(".pdf-import-options > span").filter(
        has_text=re.compile(rf"{re.escape(path.name)}.*{expected_pages} pages")
    ).wait_for(timeout=90_000)
    dialog.locator(".pdf-page-grid label").first.wait_for(state="visible", timeout=30_000)
    if dialog.locator(".pdf-page-grid label").count() != expected_pages:
        raise AssertionError(
            f"PDF preview exposed {dialog.locator('.pdf-page-grid label').count()} pages instead of {expected_pages}"
        )
    dialog.locator(".pdf-import-options select").select_option(placement)
    if replace:
        replacement = dialog.get_by_text("Replace Existing Pages", exact=True).first
        replacement.wait_for(state="visible", timeout=30_000)
        radio = dialog.locator('.pdf-reimport-choice input[type="radio"]').first
        if not radio.is_checked():
            radio.check()
        dialog.get_by_label("Existing PDF import", exact=True).wait_for(state="visible", timeout=20_000)
        if expected_pages == 2:
            dialog.get_by_role("status").filter(
                has_text=re.compile(
                    r"2 revised pages will replace 2 existing pages in place.*"
                    r"1 unmatched existing page will remain unchanged",
                    re.IGNORECASE,
                )
            ).wait_for(state="visible", timeout=30_000)
        dialog.get_by_role("button", name="Replace Existing Pages", exact=True).click()
    else:
        dialog.get_by_role("button", name="Import Selected Pages", exact=True).click()
    dialog.wait_for(state="hidden", timeout=240_000)
    progress = page.evaluate(
        """() => {
          window.__S360_PDF_PROGRESS_OBSERVER__?.disconnect();
          return window.__S360_PDF_PROGRESS_EVIDENCE__ || [];
        }"""
    )
    if not isinstance(progress, list):
        raise AssertionError("PDF progress observer returned no evidence list")
    counted = [
        item for item in progress
        if int(item.get("max") or 0) == expected_pages
        and re.fullmatch(rf"\d+ of {expected_pages} pages", str(item.get("output") or ""))
    ]
    if not counted:
        raise AssertionError(f"PDF import exposed no determinate {expected_pages}-page progress: {progress!r}")
    phases = {str(item.get("phase") or "") for item in progress}
    if "validate" not in phases or "complete" not in phases:
        raise AssertionError(f"PDF import progress omitted validate/complete phases: {sorted(phases)!r}")
    if not any(int(item.get("value") or 0) == expected_pages for item in counted):
        raise AssertionError(f"PDF import progress never reached {expected_pages} of {expected_pages}: {progress!r}")
    return progress


def nonwhite_ratio(pdf_page: fitz.Page) -> float:
    pixmap = pdf_page.get_pixmap(matrix=fitz.Matrix(0.45, 0.45), colorspace=fitz.csRGB, alpha=False)
    samples = memoryview(pixmap.samples)
    pixels = max(1, pixmap.width * pixmap.height)
    ink = 0
    for index in range(0, len(samples), 3):
        if samples[index] < 245 or samples[index + 1] < 245 or samples[index + 2] < 245:
            ink += 1
    return ink / pixels


def audit_export(path: Path, project: dict[str, Any]) -> dict[str, Any]:
    included = [item for item in sorted_pages(project) if item.get("include") is not False]
    results: list[dict[str, Any]] = []
    index_text = ""
    with fitz.open(path) as document:
        if document.page_count != len(included):
            raise AssertionError(f"export page count {document.page_count} != saved included count {len(included)}")
        for index, saved_page in enumerate(included):
            page = document[index]
            text_value = " ".join(page.get_text("text").split())
            ink_ratio = nonwhite_ratio(page)
            expected_width = 17 * 72
            expected_height = 11 * 72
            if abs(page.rect.width - expected_width) > 1 or abs(page.rect.height - expected_height) > 1:
                raise AssertionError(f"export page {index + 1} is not ANSI B landscape: {page.rect}")
            if ink_ratio < 0.002:
                raise AssertionError(f"export page {index + 1} appears blank ({ink_ratio:.6f} ink)")
            source_import = saved_page.get("sourceImport") or {}
            expected_token = ""
            source_type = str(source_import.get("type") or source_import.get("sourceType") or "").lower()
            if source_type == "pdf":
                source_index = int(source_import.get("sourcePageIndex", source_import.get("pageIndex", 0)) or 0)
                source_revision = int(source_import.get("revision") or 1)
                expected_token = (
                    f"REVISED PDF PAGE {source_index + 1}"
                    if source_revision >= 2
                    else f"ORIGINAL PDF PAGE {source_index + 1}"
                )
                if expected_token.lower() not in text_value.lower():
                    raise AssertionError(
                        f"export page {index + 1} is missing current vector content {expected_token!r}"
                    )
                if not page.get_drawings():
                    raise AssertionError(f"export page {index + 1} lost vector linework")
            elif source_type == "csv":
                expected_token = CSV_MARKER
            elif source_type == "excel_workbook":
                expected_token = WORKBOOK_MARKER
            elif source_type == "image":
                images = page.get_images(full=True)
                if not any(int(image[2]) >= 900 and int(image[3]) >= 500 for image in images):
                    raise AssertionError(f"export page {index + 1} lost the high-resolution imported image")
            if expected_token and expected_token.lower() not in text_value.lower():
                raise AssertionError(f"export page {index + 1} is missing saved content {expected_token!r}")
            if str(saved_page.get("pageType") or "").lower() == "index":
                index_text = text_value
            if str(saved_page.get("pageType") or "").lower() == "cover":
                cover_tokens = [
                    str((project.get("metadata") or {}).get(key) or "")
                    for key in (
                        "drawingSetTitle",
                        "client",
                        "storeNumber",
                        "projectType",
                        "revision",
                        "preparedBy",
                        "checkedBy",
                        "notes",
                    )
                ]
                missing_cover_tokens = [
                    token for token in cover_tokens
                    if token and token.lower() not in text_value.lower()
                ]
                if missing_cover_tokens:
                    raise AssertionError(
                        f"export Cover is missing current Project Settings values: {missing_cover_tokens!r}"
                    )
                if STALE_COVER_LINE.lower() in text_value.lower():
                    raise AssertionError("export Cover exposed a stale legacy workbook line")
            page_label = f"Page {int(saved_page.get('pageNumber') or index + 1)} of {len(included)}"
            if page_label.lower() not in text_value.lower():
                raise AssertionError(f"export page {index + 1} is missing current package label {page_label!r}")
            results.append(
                {
                    "index": index + 1,
                    "pageId": saved_page.get("id"),
                    "sheetCode": saved_page.get("displaySheetCode") or saved_page.get("sheetCode"),
                    "title": saved_page.get("sheetTitle"),
                    "mediaBox": [page.rect.width, page.rect.height],
                    "inkRatio": round(ink_ratio, 6),
                    "textCharacters": len(text_value),
                    "drawingCount": len(page.get_drawings()),
                    "imageCount": len(page.get_images(full=True)),
                    "expectedPdfToken": expected_token,
                    "packageLabel": page_label,
                }
            )
        missing_index_tokens = [
            token for token in INDEX_REQUIRED_TOKENS
            if token.lower() not in index_text.lower()
        ]
        if missing_index_tokens:
            raise AssertionError(f"automatic Sheet Index is missing current pages: {missing_index_tokens}")
    return {"path": str(path), "sha256": file_sha256(path), "pageCount": len(results), "pages": results}


def forbidden_url(url: str) -> bool:
    path = urlparse(url).path.lower()
    return any(part in path for part in FORBIDDEN_REQUEST_PARTS)


def access_log_requests(paths: list[Path]) -> list[dict[str, str]]:
    found: list[dict[str, str]] = []
    pattern = re.compile(r'"(GET|POST|PUT|PATCH|DELETE|OPTIONS)\s+([^\s]+)\s+HTTP/[0-9.]+"')
    for path in paths:
        if not path.is_file():
            continue
        for method, request_path in pattern.findall(path.read_text(encoding="utf-8", errors="replace")):
            found.append({"method": method, "path": request_path})
    return found


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--evidence-dir", type=Path, default=default_evidence_dir())
    parser.add_argument("--port", type=int, default=0, help="Alternate port; 0 selects a free port.")
    parser.add_argument("--headed", action="store_true", help="Show the browser while running.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    evidence = args.evidence_dir.resolve()
    evidence.mkdir(parents=True, exist_ok=True)
    port = int(args.port or free_port())
    if port == 8766:
        raise SystemExit("The disposable browser smoke refuses to use live port 8766.")

    report: dict[str, Any] = {
        "startedAt": datetime.now(timezone.utc).isoformat(),
        "repository": str(ROOT),
        "python": sys.executable,
        "port": port,
        "evidenceDir": str(evidence),
        "checks": [],
    }
    network_requests: list[dict[str, Any]] = []
    network_responses: list[dict[str, Any]] = []
    network_failures: list[dict[str, Any]] = []
    console_messages: list[dict[str, Any]] = []
    page_errors: list[str] = []
    dialogs: list[dict[str, Any]] = []
    downloads: list[Any] = []
    prompt_answers: list[str] = []
    failure = ""
    server: DisposableServer | None = None

    try:
        with tempfile.TemporaryDirectory(prefix="singh360_standalone_browser_") as raw_runtime:
            runtime = Path(raw_runtime)
            docs = runtime / "runtime-docs"
            original_pdf = runtime / "original" / "fixture-drawing-set.pdf"
            revised_pdf = runtime / "revised" / "fixture-drawing-set.pdf"
            image_fixture = runtime / "imports" / IMAGE_FILE_NAME
            csv_fixture = runtime / "imports" / CSV_FILE_NAME
            workbook_fixture = runtime / "imports" / WORKBOOK_FILE_NAME
            logo_fixture = runtime / "imports" / "Disposable Browser Customer Logo.png"
            make_fixture_pdf(original_pdf, revision="original")
            make_fixture_pdf(revised_pdf, revision="revised", page_count=2)
            make_fixture_image(image_fixture)
            make_fixture_csv(csv_fixture)
            make_fixture_workbook(workbook_fixture)
            make_fixture_logo(logo_fixture)
            seed_component_library(docs)
            image_fixture_sha = file_sha256(image_fixture)
            csv_fixture_sha = file_sha256(csv_fixture)
            workbook_fixture_sha = file_sha256(workbook_fixture)
            logo_fixture_sha = file_sha256(logo_fixture)
            report["disposableDocsDir"] = str(docs)
            report["fixtures"] = {
                "originalPdf": {"path": str(original_pdf), "sha256": file_sha256(original_pdf)},
                "revisedPdf": {"path": str(revised_pdf), "sha256": file_sha256(revised_pdf)},
                "image": {"path": str(image_fixture), "sha256": image_fixture_sha},
                "csv": {"path": str(csv_fixture), "sha256": csv_fixture_sha},
                "workbook": {"path": str(workbook_fixture), "sha256": workbook_fixture_sha},
                "customerLogo": {"path": str(logo_fixture), "sha256": logo_fixture_sha},
                "componentId": COMPONENT_ID,
                "componentName": COMPONENT_NAME,
            }

            server = DisposableServer(docs, port, evidence)
            first_health = server.start()
            report["firstHealth"] = first_health
            report["debugRoutes"] = http_json(port, "/api/debug/routes")
            library_before = tree_manifest(docs / "library")
            write_json(evidence / "component-library-before.sha256.json", library_before)

            dist_index = ROOT / "frontend" / "dist" / "index.html"
            if not dist_index.is_file():
                raise AssertionError("frontend/dist/index.html does not exist; run the frontend build first")
            report["servedBuild"] = {
                "index": str(dist_index),
                "indexSha256": file_sha256(dist_index),
                "indexModifiedNs": dist_index.stat().st_mtime_ns,
            }

            base_url = f"http://127.0.0.1:{port}"
            with sync_playwright() as playwright:
                edge = Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
                launch_options: dict[str, Any] = {"headless": not args.headed}
                if edge.is_file():
                    launch_options["executable_path"] = str(edge)
                browser = playwright.chromium.launch(**launch_options)
                context = browser.new_context(
                    viewport={"width": 1800, "height": 1100},
                    accept_downloads=True,
                )
                context.grant_permissions(
                    ["clipboard-read", "clipboard-write"],
                    origin=base_url,
                )
                context.add_init_script(
                    r"""(() => {
                      const nativeFetch = window.fetch.bind(window);
                      let delayNextProjectSave = false;
                      window.__S360_DELAY_NEXT_PROJECT_SAVE__ = () => {
                        delayNextProjectSave = true;
                        window.__S360_DELAY_SAVE_STARTED__ = false;
                      };
                      window.fetch = async (...args) => {
                        const request = args[0];
                        const options = args[1] || {};
                        const url = typeof request === 'string' ? request : request?.url || '';
                        const method = String(options.method || request?.method || 'GET').toUpperCase();
                        if (delayNextProjectSave && method === 'POST' && /\/api\/projects\/[a-f0-9]{16}$/.test(new URL(url, location.href).pathname)) {
                          delayNextProjectSave = false;
                          window.__S360_DELAY_SAVE_STARTED__ = true;
                          const response = nativeFetch(...args);
                          await new Promise((resolve) => window.setTimeout(resolve, 1800));
                          return response;
                        }
                        return nativeFetch(...args);
                      };
                    })()"""
                )
                page = context.new_page()

                def on_request(request: Any) -> None:
                    network_requests.append(
                        {
                            "method": request.method,
                            "url": request.url,
                            "resourceType": request.resource_type,
                        }
                    )

                def on_response(response: Any) -> None:
                    network_responses.append({"status": response.status, "url": response.url})

                def on_request_failed(request: Any) -> None:
                    network_failures.append(
                        {
                            "method": request.method,
                            "url": request.url,
                            "failure": request.failure,
                        }
                    )

                def on_console(message: Any) -> None:
                    console_messages.append(
                        {
                            "type": message.type,
                            "text": message.text,
                            "location": message.location,
                        }
                    )

                def on_dialog(dialog: Any) -> None:
                    entry = {"type": dialog.type, "message": dialog.message}
                    dialogs.append(entry)
                    if dialog.type == "confirm" and (
                        dialog.message.startswith("Archive page")
                        or (dialog.message.startswith('Archive "') and "restored later" in dialog.message)
                    ):
                        entry["action"] = "accepted"
                        dialog.accept()
                    elif dialog.type == "alert" and dialog.message == "Page template saved.":
                        entry["action"] = "accepted"
                        dialog.accept()
                    elif dialog.type == "prompt" and prompt_answers:
                        answer = prompt_answers.pop(0)
                        entry["action"] = "accepted-prompt"
                        entry["answer"] = answer
                        dialog.accept(answer)
                    else:
                        entry["action"] = "dismissed-unexpected"
                        dialog.dismiss()

                page.on("request", on_request)
                page.on("response", on_response)
                page.on("requestfailed", on_request_failed)
                page.on("console", on_console)
                page.on("pageerror", lambda error: page_errors.append(str(error)))
                page.on("dialog", on_dialog)
                page.on("download", lambda download: downloads.append(download))

                # Project Home: prove the name-only Skip path keeps every
                # optional field blank, then archive the disposable project.
                page.goto(f"{base_url}/app", wait_until="domcontentloaded", timeout=60_000)
                page.get_by_role("heading", name="Drawing Projects", exact=True).wait_for(timeout=60_000)
                page.get_by_role("button", name="New Drawing Project", exact=True).first.click()
                wizard = page.get_by_role("dialog", name="Create a standalone drawing set")
                wizard.get_by_label("Project Name", exact=False).fill(NAME_ONLY_PROJECT)
                wizard.get_by_role("button", name="Skip and Open Editor", exact=True).click()
                page.wait_for_url(re.compile(r"/app\?project=[a-f0-9]{16}&mode=editor"), timeout=60_000)
                name_only_query = parse_qs(urlparse(page.url).query)
                name_only_id = str((name_only_query.get("project") or [""])[0])
                if name_only_query.get("tool"):
                    raise AssertionError("Skip and Open Editor unexpectedly opened an import tool")
                assert_editor_ready(page, NAME_ONLY_PROJECT)
                assert_advanced_tools_collapsed(page, "after name-only Skip onboarding")
                name_only_project = project_for(port, name_only_id)
                assert_name_only_metadata(name_only_project)
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Home", exact=True).click()
                page.get_by_role("heading", name="Drawing Projects", exact=True).wait_for(timeout=60_000)
                name_only_card = page.locator(".simple-project-open.green").filter(has_text=NAME_ONLY_PROJECT).first
                name_only_card.get_by_role("button", name="Archive", exact=True).click()
                page.get_by_role("status").filter(has_text=f"{NAME_ONLY_PROJECT} was archived").wait_for(timeout=60_000)
                name_only_archived_card = page.locator(".simple-project-open.yellow").filter(has_text=NAME_ONLY_PROJECT).first
                name_only_archived_card.get_by_text("Archived", exact=True).wait_for(state="visible", timeout=30_000)
                name_only_archived = project_for(port, name_only_id)
                if (
                    name_only_archived.get("archived") is not True
                    or not name_only_archived.get("archivedAt")
                    or not name_only_archived.get("archivedReason")
                ):
                    raise AssertionError("name-only project was not archived recoverably")
                assert_name_only_metadata(name_only_archived)
                report["nameOnlyProjectId"] = name_only_id
                report["checks"].append("name-only Skip onboarding kept optional metadata blank and archived recoverably")

                # Full onboarding populates every field, stores a project-local
                # customer logo, and opens the generic Add / Import dialog.
                page.get_by_role("button", name="New Drawing Project", exact=True).first.click()
                wizard = page.get_by_role("dialog", name="Create a standalone drawing set")
                wizard.get_by_label("Project Name", exact=False).fill(PROJECT_NAME)
                wizard.get_by_label("Customer / Client", exact=True).fill(FULL_PROJECT_METADATA["client"])
                wizard.get_by_label("Store / Project Number", exact=True).fill(FULL_PROJECT_METADATA["storeNumber"])
                wizard.get_by_label("Project Location", exact=True).fill(FULL_PROJECT_METADATA["location"])
                wizard.get_by_label("Project Type", exact=True).fill(FULL_PROJECT_METADATA["projectType"])
                wizard.get_by_label("Drawing Set Title", exact=True).fill(FULL_PROJECT_METADATA["drawingSetTitle"])
                wizard.get_by_label("Prepared By", exact=True).fill(FULL_PROJECT_METADATA["preparedBy"])
                wizard.get_by_label("Checked By", exact=True).fill(FULL_PROJECT_METADATA["checkedBy"])
                wizard.get_by_label("Created Date", exact=True).fill(FULL_PROJECT_METADATA["createdDate"])
                wizard.get_by_label("Revision", exact=True).fill(FULL_PROJECT_METADATA["revision"])
                wizard.get_by_label("Notes", exact=True).fill(FULL_PROJECT_METADATA["notes"])
                wizard.get_by_label("Drawing-package File Name", exact=True).fill(
                    FULL_PROJECT_METADATA["drawingPackageFileName"]
                )
                wizard.get_by_label("Customer Logo", exact=True).set_input_files(str(logo_fixture))
                wizard.get_by_role("button", name="Import Files Now", exact=True).click()
                page.wait_for_url(
                    re.compile(r"/app\?project=[a-f0-9]{16}&mode=editor&tool=add-import"),
                    timeout=60_000,
                )
                query = parse_qs(urlparse(page.url).query)
                project_id = str((query.get("project") or [""])[0])
                if not re.fullmatch(r"[a-f0-9]{16}", project_id):
                    raise AssertionError(f"new project URL contains no stable project ID: {page.url}")
                if query.get("tool") != ["add-import"]:
                    raise AssertionError(f"Import Files Now did not preserve tool=add-import: {page.url}")
                generic_import = page.get_by_role("dialog", name="Add / Import Page")
                generic_import.wait_for(state="visible", timeout=60_000)
                generic_import.get_by_role("button", name="Close Add or Import Page", exact=True).click()
                generic_import.wait_for(state="hidden", timeout=30_000)
                audit_url = f"{base_url}/app?project={project_id}&mode=editor&workflowAudit=1"
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after full Import Files Now onboarding")
                page.goto(audit_url, wait_until="domcontentloaded", timeout=60_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after opening the workflow-audit editor")
                page.screenshot(path=str(evidence / "01-new-project.png"), full_page=True)

                created = project_for(port, project_id)
                logo_asset_url = assert_wizard_metadata(
                    created,
                    FULL_PROJECT_METADATA,
                    require_logo=True,
                )
                initial_pages = sorted_pages(created)
                if created.get("projectMode") != "standalone_layout":
                    raise AssertionError("new browser project is not standalone_layout")
                if len(initial_pages) != 3:
                    raise AssertionError(f"blank drawing set should start with cover/index/blank, got {len(initial_pages)} pages")
                if initial_pages[0].get("pageType") != "cover" or initial_pages[1].get("pageType") != "index":
                    raise AssertionError("new drawing set does not begin with app-managed Cover and Sheet Index")
                initial_blank_id = str(initial_pages[2]["id"])
                package_candidates = [path for path in (docs / "projects").iterdir() if project_id in path.name]
                if len(package_candidates) != 1:
                    raise AssertionError(f"could not identify one project package for {project_id}")
                project_package = package_candidates[0]
                logo_name = Path(urlparse(logo_asset_url).path).name
                logo_copies = list(project_package.rglob(logo_name))
                if len(logo_copies) != 1 or file_sha256(logo_copies[0]) != logo_fixture_sha:
                    raise AssertionError("wizard customer logo is missing or changed in the project package")
                logo_fixture.unlink()

                # Inject one disposable legacy cover row into provenance, then
                # prove the standalone renderer displays only current settings.
                cover = initial_pages[0]
                cover_blocks = cover.get("blocks") or []
                if not cover_blocks:
                    raise AssertionError("managed Cover has no content block")
                cover_blocks[0]["rows"] = [[STALE_COVER_LINE]]
                created["pages"] = [
                    cover if item.get("id") == cover.get("id") else item
                    for item in created.get("pages") or []
                ]
                http_post_json(port, f"/api/projects/{project_id}", created)
                page.reload(wait_until="domcontentloaded", timeout=90_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after cover-authority reload")
                sheet_item(page, "Cover / Project Info").locator(".sheet-item-title").click()
                cover_view = page.locator(".np-cover")
                cover_view.wait_for(state="visible", timeout=30_000)
                for token in (
                    FULL_PROJECT_METADATA["drawingSetTitle"],
                    FULL_PROJECT_METADATA["client"],
                    FULL_PROJECT_METADATA["storeNumber"],
                    FULL_PROJECT_METADATA["projectType"],
                    FULL_PROJECT_METADATA["revision"],
                    FULL_PROJECT_METADATA["preparedBy"],
                    FULL_PROJECT_METADATA["checkedBy"],
                    FULL_PROJECT_METADATA["notes"],
                ):
                    cover_view.get_by_text(re.compile(re.escape(token), re.IGNORECASE)).wait_for(
                        state="visible", timeout=30_000
                    )
                if cover_view.get_by_text(STALE_COVER_LINE, exact=True).count():
                    raise AssertionError("standalone Cover exposed a stale legacy workbook line")
                customer_logo = cover_view.get_by_role("img", name="Customer logo", exact=True)
                customer_logo.wait_for(state="visible", timeout=30_000)
                if customer_logo.evaluate("img => img.naturalWidth") < 1:
                    raise AssertionError("project-local customer logo did not render on the Cover")
                report["projectId"] = project_id
                report["checks"].append(
                    "full onboarding metadata/logo persisted, Import Files Now opened the generic dialog, and current Cover settings hid stale legacy rows"
                )

                # Project Settings round-trips the same wizard metadata/logo.
                # The Cover can be intentionally hidden only in its explicit
                # advanced option and restored without losing the page.
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Settings", exact=True).click()
                settings = page.get_by_role("dialog", name="Project Settings")
                for label, key in (
                    ("Project Name *", "projectName"),
                    ("Customer / Client", "client"),
                    ("Store / Project Number", "storeNumber"),
                    ("Project Location", "location"),
                    ("Project Type", "projectType"),
                    ("Drawing Set Title", "drawingSetTitle"),
                    ("Prepared By", "preparedBy"),
                    ("Checked By", "checkedBy"),
                    ("Created Date", "createdDate"),
                    ("Revision", "revision"),
                    ("Drawing-package File Name", "drawingPackageFileName"),
                ):
                    if settings.get_by_label(label, exact=True).input_value() != FULL_PROJECT_METADATA[key]:
                        raise AssertionError(f"Project Settings did not preload wizard field {key}")
                if settings.locator("label.project-settings-notes textarea").input_value() != FULL_PROJECT_METADATA["notes"]:
                    raise AssertionError("Project Settings did not preload wizard notes")
                settings.get_by_text("Current project-local customer logo is set.", exact=True).wait_for(
                    state="visible", timeout=30_000
                )
                cover_options = settings.locator("details.project-settings-advanced")
                cover_options.locator("summary").click()
                cover_include = settings.get_by_label(
                    "Include the automatic Cover in the exported drawing set", exact=True
                )
                cover_include.uncheck()
                settings.get_by_role("button", name="Save Project Settings", exact=True).click()
                save_project(page)
                configured = project_for(port, project_id)
                assert_wizard_metadata(configured, FULL_PROJECT_METADATA, require_logo=True)
                if configured.get("coverSettings", {}).get("include") is not False:
                    raise AssertionError("advanced Project Settings did not exclude the Cover")
                if page_by_id(configured, str(cover["id"])).get("include") is not False:
                    raise AssertionError("excluded Cover page was not retained recoverably")
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Settings", exact=True).click()
                settings = page.get_by_role("dialog", name="Project Settings")
                settings.locator("details.project-settings-advanced summary").click()
                settings.get_by_label(
                    "Include the automatic Cover in the exported drawing set", exact=True
                ).check()
                settings.get_by_role("button", name="Save Project Settings", exact=True).click()
                save_project(page)
                configured = project_for(port, project_id)
                if configured.get("coverSettings", {}).get("include") is not True:
                    raise AssertionError("advanced Project Settings did not restore the Cover")
                report["checks"].append("Project Settings preserved wizard data/logo and explicitly hid/restored the managed Cover")

                # Create and rename a blank page. Structural controls must
                # capture the newest live canvas before reorder/include changes.
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Add / Import Page", exact=True).click()
                add_dialog = page.get_by_role("dialog", name="Add / Import Page")
                add_dialog.get_by_label("Page Title", exact=True).fill("Browser Lifecycle Page")
                add_dialog.get_by_label("Sheet Code", exact=True).fill("TMP-1")
                add_dialog.get_by_role("button", name="Create Blank Layout Page", exact=True).click()
                add_dialog.wait_for(state="hidden", timeout=30_000)
                rename_and_recode(
                    page,
                    "Browser Lifecycle Page",
                    "Browser Layout Renamed",
                    "L-101",
                    prompt_answers,
                )
                save_project(page)
                lifecycle_project = project_for(port, project_id)
                lifecycle_page = next(item for item in lifecycle_project["pages"] if item.get("sheetTitle") == "Browser Layout Renamed")
                lifecycle_id = str(lifecycle_page["id"])
                order_before = int(lifecycle_page["order"])
                sheet_item(page, "Browser Layout Renamed").locator(".sheet-item-title").click()
                structural_canvas_ids = add_canvas_text(page)
                direction = "up" if order_before > 3 else "down"
                move_page(page, "Browser Layout Renamed", direction)
                lifecycle_after_move = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: (
                        int(page_by_id(current, lifecycle_id).get("order") or 0) != order_before
                        and structural_canvas_ids.issubset(
                            collect_object_ids(page_by_id(current, lifecycle_id).get("canvasObjects") or [])
                        )
                    ),
                    "dirty-canvas reorder did not save the latest drawing objects",
                )
                order_after = int(page_by_id(lifecycle_after_move, lifecycle_id)["order"])
                if order_after == order_before:
                    raise AssertionError("blank page reorder control did not change its saved order")

                lifecycle_item = sheet_item(page, "Browser Layout Renamed")
                include_box = lifecycle_item.locator('input[type="checkbox"]').first
                include_box.uncheck()
                excluded_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: page_by_id(current, lifecycle_id).get("include") is False,
                    "blank page exclude state did not autosave",
                )
                excluded_order = int(page_by_id(excluded_state, lifecycle_id).get("order") or 0)
                excluded_direction = "down" if direction == "up" else "up"
                move_page(page, "Browser Layout Renamed", excluded_direction)
                excluded_reordered = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: int(page_by_id(current, lifecycle_id).get("order") or 0) != excluded_order,
                    "excluded page reorder did not persist",
                )
                order_while_excluded = int(page_by_id(excluded_reordered, lifecycle_id).get("order") or 0)
                include_box = sheet_item(page, "Browser Layout Renamed").locator('input[type="checkbox"]').first
                include_box.check()
                reincluded_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: page_by_id(current, lifecycle_id).get("include") is True,
                    "blank page include state did not autosave",
                )
                reincluded = page_by_id(reincluded_state, lifecycle_id)
                if int(reincluded.get("order") or 0) != order_while_excluded:
                    raise AssertionError("re-including a reordered page moved it back to a stale stored position")
                if not structural_canvas_ids.issubset(collect_object_ids(reincluded.get("canvasObjects") or [])):
                    raise AssertionError("include/reorder operations lost the newest canvas objects")

                # Delay the response for save A, then make edit B while A is
                # in flight. B must autosave without a click, page switch, or
                # manual Save Project. This also proves Page Properties writes
                # both canonical and displayed sheet-code fields.
                open_properties_panel(page)
                page.locator(".status-pill").filter(has_text="PROJECT SAVED").wait_for(
                    state="visible", timeout=30_000
                )
                page.evaluate("() => window.__S360_DELAY_NEXT_PROJECT_SAVE__()")
                page.locator("#page-code").fill("L-201")
                page.wait_for_function(
                    "() => window.__S360_DELAY_SAVE_STARTED__ === true",
                    timeout=30_000,
                )
                page.locator("#page-notes").fill(DELAYED_SAVE_NOTES)
                queued_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: (
                        page_by_id(current, lifecycle_id).get("sheetCode") == "L-201"
                        and page_by_id(current, lifecycle_id).get("displaySheetCode") == "L-201"
                        and page_by_id(current, lifecycle_id).get("notes") == DELAYED_SAVE_NOTES
                    ),
                    "edit B made during delayed save A was not automatically queued and saved",
                    timeout=60.0,
                )
                if not structural_canvas_ids.issubset(
                    collect_object_ids(page_by_id(queued_state, lifecycle_id).get("canvasObjects") or [])
                ):
                    raise AssertionError("delayed autosave queue lost existing canvas content")
                page.locator(".status-pill").filter(has_text="PROJECT SAVED").wait_for(
                    state="visible", timeout=30_000
                )

                # Add one sanitized continuation record to the disposable
                # project, then archive the dirty base page through the UI.
                # The app must move and restore the complete group atomically,
                # exposing one root archive card rather than a duplicate child.
                _, lifecycle_child_id = install_continuation_fixture(port, project_id, lifecycle_id)
                page.reload(wait_until="domcontentloaded", timeout=90_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after continuation fixture reload")
                child_title = "Browser Layout Renamed — CONTINUED"
                sheet_item(page, child_title).filter(has_text="Continuation").wait_for(
                    state="visible", timeout=30_000
                )
                sheet_item(page, "Browser Layout Renamed").locator(".sheet-item-title").click()
                dirty_archive_ids = add_canvas_text(page)
                lifecycle_locator = sheet_item(page, "Browser Layout Renamed")
                lifecycle_locator.click(button="right")
                page.locator(".ctx-menu").get_by_role("button", name="Archive Page", exact=True).click()
                lifecycle_locator.wait_for(state="hidden", timeout=20_000)
                page.locator(".sheet-item").filter(has_text=child_title).first.wait_for(
                    state="hidden", timeout=20_000
                )
                archived_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: {lifecycle_id, lifecycle_child_id}.issubset(
                        {str(item.get("id") or "") for item in current.get("archivedPages") or []}
                    ),
                    "continuation group did not archive atomically",
                )
                archived_group = [
                    item for item in archived_state.get("archivedPages") or []
                    if item.get("id") in {lifecycle_id, lifecycle_child_id}
                ]
                if len(archived_group) != 2:
                    raise AssertionError("archive did not retain exactly the base and continuation pages")
                archived_base = next(item for item in archived_group if item.get("id") == lifecycle_id)
                if not dirty_archive_ids.issubset(collect_object_ids(archived_base.get("canvasObjects") or [])):
                    raise AssertionError("dirty-canvas archive saved a stale base-page snapshot")
                if any(
                    not item.get("archivedAt")
                    or not item.get("archivedReason")
                    or item.get("archivedGroupRootId") != lifecycle_id
                    for item in archived_group
                ):
                    raise AssertionError("archived continuation group lacks timestamp/reason/root metadata")
                page.locator(".status-pill").filter(has_text="PROJECT SAVED").wait_for(
                    state="visible", timeout=30_000
                )
                archive_ui_state = page.evaluate(
                    """
                    () => {
                      const visible = (node) => {
                        if (!(node instanceof HTMLElement)) return false;
                        const style = window.getComputedStyle(node);
                        const rect = node.getBoundingClientRect();
                        return style.display !== "none"
                          && style.visibility !== "hidden"
                          && rect.width > 0
                          && rect.height > 0;
                      };
                      return {
                        panelLeft: Array.from(document.querySelectorAll(".panel-left")).map((node) => ({
                          visible: visible(node),
                          className: node.className,
                        })),
                        sectionTitles: Array.from(document.querySelectorAll(".nav-section-title")).map((node) => ({
                          text: node.textContent?.trim() || "",
                          visible: visible(node),
                        })),
                        sections: Array.from(document.querySelectorAll(".nav-section")).map((node) => ({
                          text: node.querySelector(".nav-section-title")?.textContent?.trim() || "",
                          archivedCards: node.querySelectorAll(".archived-page-item").length,
                          visible: visible(node),
                        })),
                        visibleSheets: Array.from(document.querySelectorAll(".sheet-item"))
                          .filter(visible)
                          .map((node) => node.textContent?.replace(/\\s+/g, " ").trim() || ""),
                        statusPills: Array.from(document.querySelectorAll(".status-pill"))
                          .filter(visible)
                          .map((node) => node.textContent?.replace(/\\s+/g, " ").trim() || ""),
                      };
                    }
                    """
                )
                write_json(evidence / "02a-archive-immediate-ui.json", archive_ui_state)
                page.screenshot(path=str(evidence / "02a-archive-immediate.png"), full_page=True)
                if not any(
                    str(item.get("text") or "").startswith("Archived Pages")
                    for item in archive_ui_state.get("sectionTitles") or []
                ):
                    raise AssertionError(
                        "Archived Pages did not refresh immediately after the saved archive; "
                        f"current rail DOM: {json.dumps(archive_ui_state, ensure_ascii=False)}"
                    )
                expand_section(page, "Archived Pages")
                archived_section = page.locator(".nav-section").filter(has_text="Archived Pages").first
                archived_card = archived_section.locator(".archived-page-item").filter(
                    has_text="Browser Layout Renamed"
                )
                if archived_card.count() != 1:
                    raise AssertionError("Archived Pages exposed the continuation as a duplicate root row")
                archived_card.get_by_text("Includes 1 continuation page", exact=True).wait_for(
                    state="visible", timeout=30_000
                )
                if "Archive time unavailable" in archived_card.inner_text() or "No archive reason recorded" in archived_card.inner_text():
                    raise AssertionError("Archived Pages UI omitted the archive timestamp or reason")
                archived_card.get_by_role("button", name="Restore", exact=True).click()
                sheet_item(page, "Browser Layout Renamed")
                sheet_item(page, child_title)
                restored_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: (
                        {lifecycle_id, lifecycle_child_id}.issubset(
                            {str(item.get("id") or "") for item in current.get("pages") or []}
                        )
                        and not {lifecycle_id, lifecycle_child_id}.intersection(
                            {str(item.get("id") or "") for item in current.get("archivedPages") or []}
                        )
                    ),
                    "continuation group did not restore atomically",
                )
                restored_base = page_by_id(restored_state, lifecycle_id)
                restored_child = page_by_id(restored_state, lifecycle_child_id)
                restored_ids = [str(item.get("id") or "") for item in sorted_pages(restored_state)]
                if (
                    restored_base.get("sheetCode") != "L-201"
                    or restored_base.get("displaySheetCode") != "L-201"
                    or restored_base.get("notes") != DELAYED_SAVE_NOTES
                    or not dirty_archive_ids.issubset(collect_object_ids(restored_base.get("canvasObjects") or []))
                    or restored_child.get("continuationOf") != lifecycle_id
                    or restored_ids.index(lifecycle_child_id) != restored_ids.index(lifecycle_id) + 1
                ):
                    raise AssertionError("restored continuation group lost its stable metadata or latest canvas")
                save_project(page)
                report["checks"].append(
                    "dirty canvas survived reorder/include and continuation-group archive/restore; delayed save A queued edit B; Page Properties recode persisted both fields"
                )
                page.screenshot(path=str(evidence / "02-page-lifecycle.png"), full_page=True)

                # Create a local text/table page, duplicate it, then save and reinsert a text-only template.
                text_dialog = open_add_import(page, "Text / Table Page")
                text_dialog.get_by_label("Page Title", exact=True).fill(TEXT_PAGE_TITLE)
                text_dialog.get_by_label("Sheet Code", exact=True).fill("TXT-101")
                text_dialog.get_by_role("button", name="Create Text / Table Page", exact=True).click()
                text_dialog.wait_for(state="hidden", timeout=30_000)
                save_project(page)
                text_state = project_for(port, project_id)
                text_page = next(item for item in text_state["pages"] if item.get("sheetTitle") == TEXT_PAGE_TITLE)
                text_page_id = str(text_page["id"])
                text_block_types = [str(block.get("type") or "") for block in text_page.get("blocks") or []]
                if text_page.get("pageType") != "data-grid" or text_block_types != ["paragraph", "table"]:
                    raise AssertionError(f"text/table page structure is incomplete: {text_page!r}")
                table_block = (text_page.get("blocks") or [None, {}])[1]
                if table_block.get("editable") is not True or text_page.get("canvasObjects"):
                    raise AssertionError("text/table page is not an editable local page")

                sheet_item(page, TEXT_PAGE_TITLE).click(button="right")
                page.locator(".ctx-menu").get_by_role("button", name="Duplicate Sheet", exact=True).click()
                save_project(page)
                duplicate_state = project_for(port, project_id)
                duplicate_page = next(
                    item for item in duplicate_state["pages"]
                    if item.get("sheetTitle") == f"{TEXT_PAGE_TITLE} Copy"
                )
                if duplicate_page.get("id") == text_page_id or duplicate_page.get("linkedWorksheetId"):
                    raise AssertionError("text-page duplicate retained the source page or worksheet identity")
                if duplicate_page.get("pageGroupId") != duplicate_page.get("id"):
                    raise AssertionError("text-page duplicate did not receive its own page group")
                if duplicate_page.get("blocks") != text_page.get("blocks"):
                    raise AssertionError("text-page duplicate changed the editable blocks")

                sheet_item(page, TEXT_PAGE_TITLE).locator(".sheet-item-title").click()
                dirty_template_ids = add_canvas_text(page)
                expected_template_code = logical_code_after(
                    project_for(port, project_id).get("pages") or [],
                    text_page_id,
                )
                open_advanced_tools(page)
                page.locator(".ribbon-tabs").get_by_role("button", name="File", exact=True).click()
                page.locator(".ribbon-groups").get_by_role("button", name="Save Page as Template", exact=True).click()
                save_template_modal = page.locator(".modal").filter(has_text="Save Page as Template").last
                save_template_modal.get_by_label("Template name", exact=True).fill(TEMPLATE_NAME)
                save_template_modal.get_by_role("button", name="Save / Update Template", exact=True).click()
                save_template_modal.wait_for(state="hidden", timeout=60_000)
                templates_response = http_json(port, "/api/lib/page-templates")
                template_entry = next(
                    item for item in templates_response.get("templates") or []
                    if item.get("name") == TEMPLATE_NAME
                )
                template_payload = http_json(port, f"/api/lib/page-templates/{template_entry['id']}").get("template") or {}
                forbidden_template_keys = {"id", "order", "include", "sheetCode", "sheetTab", "linkedWorksheetId"}
                retained_template_keys = sorted(forbidden_template_keys.intersection(template_payload))
                authoritative_text_page = page_by_id(project_for(port, project_id), text_page_id)
                if (
                    retained_template_keys
                    or template_payload.get("blocks") != authoritative_text_page.get("blocks")
                    or not dirty_template_ids.issubset(
                        collect_object_ids(template_payload.get("canvasObjects") or [])
                    )
                ):
                    raise AssertionError(
                        f"dirty-canvas template retained identity/source keys or lost current content: {retained_template_keys}"
                    )

                open_advanced_tools(page)
                page.locator(".ribbon-tabs").get_by_role("button", name="File", exact=True).click()
                page.locator(".ribbon-groups").get_by_role("button", name="Insert Page Template", exact=True).click()
                template_modal = page.locator(".modal.modal-wide").filter(has_text="Insert Page Template").last
                template_modal.locator(".pt-lib-item").filter(has_text=TEMPLATE_NAME).click()
                template_modal.locator("select").select_option("new_after")
                template_modal.get_by_role("button", name="Insert Template", exact=True).click()
                template_modal.wait_for(state="hidden", timeout=60_000)
                save_project(page)
                templated_state = project_for(port, project_id)
                inserted_template = next(
                    item for item in templated_state["pages"]
                    if item.get("sheetTitle") == TEMPLATE_NAME
                )
                if (
                    inserted_template.get("id") == text_page_id
                    or inserted_template.get("pageGroupId") != inserted_template.get("id")
                    or inserted_template.get("sheetCode") != expected_template_code
                    or inserted_template.get("displaySheetCode") != expected_template_code
                    or inserted_template.get("include") is not True
                    or inserted_template.get("linkedWorksheetId")
                    or inserted_template.get("sourceImport")
                ):
                    raise AssertionError("inserted text template retained source identity or lost local page identity")
                inserted_template_ids = collect_object_ids(inserted_template.get("canvasObjects") or [])
                if (
                    not inserted_template_ids
                    or dirty_template_ids.intersection(inserted_template_ids)
                    or len(inserted_template_ids) != len(
                        collect_object_ids(template_payload.get("canvasObjects") or [])
                    )
                ):
                    raise AssertionError("inserted dirty-canvas template did not receive fresh recursive object IDs")
                report["checks"].append(
                    "text page, detached duplicate, and dirty-canvas template save/switch/insert persisted with the logical next code"
                )

                # Import a generated CSV as one editable project-local table page.
                csv_dialog = open_add_import(page, "Excel Worksheet / CSV Table")
                csv_dialog.get_by_label("Choose CSV", exact=True).set_input_files(str(csv_fixture))
                csv_dialog.get_by_label("CSV source preview", exact=True).wait_for(state="visible", timeout=30_000)
                csv_dialog.get_by_role("button", name="Import CSV Table", exact=True).click()
                csv_dialog.wait_for(state="hidden", timeout=90_000)
                csv_state = project_for(port, project_id)
                csv_source = next(item for item in csv_state.get("sources") or [] if item.get("type") == "csv")
                csv_imported_pages = [
                    item for item in csv_state["pages"]
                    if (item.get("sourceImport") or {}).get("sourceId") == csv_source.get("id")
                ]
                if len(csv_imported_pages) != 1 or csv_imported_pages[0].get("sheetTitle") != "Equipment Summary":
                    raise AssertionError("CSV import did not create exactly one Equipment Summary page")
                csv_import_page = csv_imported_pages[0]
                csv_table = (csv_import_page.get("blocks") or [{}])[0]
                if csv_table.get("type") != "table" or csv_table.get("editable") is not True:
                    raise AssertionError("CSV import did not create an editable table block")
                if csv_source.get("sha256") != csv_fixture_sha or csv_source.get("importMode") != "one_time_editable_table":
                    raise AssertionError("CSV import provenance is incomplete")

                package_candidates = [path for path in (docs / "projects").iterdir() if project_id in path.name]
                if len(package_candidates) != 1:
                    raise AssertionError(f"could not identify one project package for {project_id}")
                project_package = package_candidates[0]
                local_csv = project_package / str(csv_source.get("projectLocalPath") or "")
                if not local_csv.is_file() or file_sha256(local_csv) != csv_fixture_sha:
                    raise AssertionError("CSV project-local source copy is missing or changed")
                csv_fixture.unlink()

                # Import a generated image as a contained, uncropped project-local drawing page.
                image_dialog = open_add_import(page, "Image / Screenshot")
                image_dialog.get_by_label("Choose Image", exact=True).set_input_files(str(image_fixture))
                image_dialog.wait_for(state="hidden", timeout=90_000)
                sheet_item(page, Path(IMAGE_FILE_NAME).stem)
                save_project(page)
                image_state = project_for(port, project_id)
                image_page = next(
                    item for item in image_state["pages"]
                    if (item.get("sourceImport") or {}).get("type") == "image"
                )
                image_source = image_page.get("sourceImport") or {}
                image_objects = image_page.get("canvasObjects") or []
                if (
                    image_page.get("pageType") != "image"
                    or image_source.get("sha256") != image_fixture_sha
                    or image_source.get("placementMode") != "fit_body"
                    or len(image_objects) != 1
                    or image_objects[0].get("src") != image_objects[0].get("sourceUrl")
                    or not image_objects[0].get("objectId")
                ):
                    raise AssertionError("image page structure or project-local provenance is incomplete")
                local_image = project_package / str(image_source.get("projectLocalPath") or "")
                if not local_image.is_file() or file_sha256(local_image) != image_fixture_sha:
                    raise AssertionError("image project-local asset copy is missing or changed")
                image_fixture.unlink()

                # Import one formatted worksheet from a generated control-sheet workbook.
                workbook_choice = open_add_import(page, "Excel Worksheet / CSV Table")
                workbook_choice.get_by_role("button", name="Choose Excel Worksheet", exact=True).click()
                workbook_choice.wait_for(state="hidden", timeout=30_000)
                workbook_modal = page.locator(".modal.modal-wide").filter(has_text="Add One Worksheet from Excel").last
                workbook_modal.locator('input[type="file"][accept=".xlsx,.xlsm"]').first.set_input_files(str(workbook_fixture))
                worksheet_row = workbook_modal.locator(".op-table tbody tr").filter(has_text=WORKBOOK_SHEET_NAME)
                worksheet_row.wait_for(state="visible", timeout=90_000)
                worksheet_row.click()
                preserve_geometry = workbook_modal.locator(".lib-showretired input[type='checkbox']")
                if not preserve_geometry.is_checked():
                    raise AssertionError("formatted worksheet import did not default to preserved geometry")
                workbook_modal.get_by_role("button", name="Add Selected Sheet as One Page", exact=True).click()
                workbook_modal.wait_for(state="hidden", timeout=240_000)
                workbook_state = project_for(port, project_id)
                workbook_source = next(
                    item for item in workbook_state.get("sources") or []
                    if item.get("type") == "imported-workbook" and item.get("originalFileName") == WORKBOOK_FILE_NAME
                )
                workbook_pages = [
                    item for item in workbook_state["pages"]
                    if (item.get("sourceImport") or {}).get("sourceId") == workbook_source.get("id")
                ]
                if len(workbook_pages) != 1:
                    raise AssertionError(f"formatted workbook import created {len(workbook_pages)} pages instead of one")
                workbook_page = workbook_pages[0]
                workbook_import = workbook_page.get("sourceImport") or {}
                if (
                    workbook_page.get("sheetCode") != "XLSX-101"
                    or workbook_page.get("sheetTitle") != WORKBOOK_PAGE_TITLE
                    or workbook_page.get("renderMode") != "excel_exact"
                    or workbook_page.get("layoutProfile") != "exact_source_excel"
                    or workbook_page.get("layoutOverride") != "exact_source"
                    or workbook_page.get("trimBlankRows") is not False
                    or workbook_page.get("trimBlankColumns") is not False
                    or workbook_import.get("sourceType") != "excel_workbook"
                    or workbook_import.get("sha256") != workbook_fixture_sha
                ):
                    raise AssertionError("formatted workbook page lost indexed identity or source-fidelity settings")
                linked_worksheet = next(
                    item for item in workbook_state.get("worksheets") or []
                    if item.get("id") == workbook_page.get("linkedWorksheetId")
                )
                if linked_worksheet.get("classHint") != "excel_exact":
                    raise AssertionError("formatted workbook worksheet lost exact-source geometry classification")
                local_workbook = project_package / str(workbook_source.get("projectLocalPath") or "")
                if not local_workbook.is_file() or file_sha256(local_workbook) != workbook_fixture_sha:
                    raise AssertionError("workbook project-local source copy is missing or changed")
                workbook_fixture.unlink()
                if csv_fixture.exists() or image_fixture.exists() or workbook_fixture.exists():
                    raise AssertionError("one or more disposable import sources remain externally available")
                report["checks"].append("generated CSV, image, and exact-geometry XLSX imports became source-independent project pages")

                # Add all three original PDF pages in Full Sheet mode.
                original_pdf_progress = import_pdf_via_ui(
                    page,
                    original_pdf,
                    placement="full_sheet",
                    replace=False,
                    expected_pages=3,
                )
                initial_pdf_state = project_for(port, project_id)
                imported = pdf_pages(initial_pdf_state)
                if len(imported) != 3:
                    raise AssertionError(f"multi-page PDF import created {len(imported)} managed pages instead of 3")
                pdf_ids = [str(item["id"]) for item in imported]
                if len(set(pdf_ids)) != 3:
                    raise AssertionError("multi-page PDF import did not create stable unique page IDs")
                if any((item.get("sourceImport") or {}).get("placementMode") != "full_sheet" for item in imported):
                    raise AssertionError("Full Sheet PDF placement did not persist")
                if any(item.get("include") is not True for item in imported):
                    raise AssertionError("newly imported PDF pages were not all included")
                report.setdefault("pdfImportProgress", {})["originalThreePage"] = original_pdf_progress
                report["checks"].append("three-page PDF imported as project-local Full Sheet pages")

                # Duplicating one PDF-backed page must create an independent
                # app-managed canvas page, not another managed reimport entry
                # or a duplicate of the full three-page group.
                # Use the group's numeric base code as the sequencing anchor.
                # The additional imported pages intentionally carry PDF-group
                # suffixes (for example XLSX-102a), which are source grouping
                # codes rather than independent logical-number anchors.
                duplicate_source = imported[0]
                duplicate_source_id = str(duplicate_source["id"])
                duplicate_source_title = str(duplicate_source.get("sheetTitle") or "")
                duplicate_expected_code = logical_code_after(
                    initial_pdf_state.get("pages") or [], duplicate_source_id
                )
                duplicate_source_object_ids = collect_object_ids(
                    duplicate_source.get("canvasObjects") or []
                )
                sheet_item(page, duplicate_source_title).click(button="right")
                page.locator(".ctx-menu").get_by_role("button", name="Duplicate Sheet", exact=True).click()
                save_project(page)
                duplicated_pdf_state = project_for(port, project_id)
                duplicated_pdf = next(
                    item for item in duplicated_pdf_state.get("pages") or []
                    if item.get("sheetTitle") == f"{duplicate_source_title} Copy"
                )
                duplicated_object_ids = collect_object_ids(duplicated_pdf.get("canvasObjects") or [])
                duplicated_urls = {
                    str(item.get("sourceUrl") or item.get("src") or "")
                    for item in duplicated_pdf.get("canvasObjects") or []
                }
                source_urls = {
                    str(item.get("sourceUrl") or item.get("src") or "")
                    for item in duplicate_source.get("canvasObjects") or []
                }
                if (
                    duplicated_pdf.get("id") == duplicate_source_id
                    or duplicated_pdf.get("pageType") != "canvas"
                    or duplicated_pdf.get("sourceImport")
                    or duplicated_pdf.get("continuationOf")
                    or duplicated_pdf.get("pageGroupId") != duplicated_pdf.get("id")
                    or duplicated_pdf.get("sheetCode") != duplicate_expected_code
                    or duplicated_pdf.get("displaySheetCode") != duplicate_expected_code
                    or not duplicated_object_ids
                    or duplicate_source_object_ids.intersection(duplicated_object_ids)
                    or not source_urls
                    or "" in source_urls
                    or duplicated_urls != source_urls
                    or len(pdf_pages(duplicated_pdf_state)) != 3
                ):
                    raise AssertionError("PDF page duplicate retained managed import identity or lost render content")
                report["checks"].append(
                    "single PDF-page duplicate became an independent canvas page with fresh IDs and did not duplicate the managed PDF group"
                )

                # Give one imported page user metadata/order and place reusable components.
                first_pdf = imported[0]
                first_pdf_old_title = str(first_pdf.get("sheetTitle") or "")
                rename_and_recode(
                    page,
                    first_pdf_old_title,
                    "Imported Browser Page One",
                    "PDF-101",
                    prompt_answers,
                )
                move_page(page, "Imported Browser Page One", "up")
                save_project(page)
                first_pdf_id = str(first_pdf["id"])
                page_after_metadata = page_by_id(project_for(port, project_id), first_pdf_id)
                if page_after_metadata.get("sheetTitle") != "Imported Browser Page One":
                    raise AssertionError("PDF page title edit did not persist")
                if page_after_metadata.get("include") is not True:
                    raise AssertionError("PDF page rename/reorder unexpectedly changed its include state")

                sheet_item(page, "Imported Browser Page One").locator(".sheet-item-title").click()
                page.locator('[aria-label="Drawing canvas objects"][data-canvas-hydrated="1"]').wait_for(timeout=60_000)
                expand_section(page, "Components")
                browser_panel = page.locator('.libv2-browser[aria-label="Component Browser"]')
                browser_panel.wait_for(state="visible", timeout=30_000)
                browser_panel.get_by_text("Component library ready", exact=True).wait_for(timeout=30_000)
                browser_panel.get_by_placeholder("Search components…").fill(COMPONENT_NAME)
                component_card = browser_panel.locator(".libv2-browser-card").filter(has_text=COMPONENT_NAME)
                component_card.wait_for(state="visible", timeout=30_000)
                component_card.click()
                page.wait_for_function(
                    "(name) => (window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []).filter((item) => item.objName === name).length >= 1",
                    arg=COMPONENT_NAME,
                    timeout=30_000,
                )
                component_card.drag_to(page.locator(".sheet-viewport"), target_position={"x": 1050, "y": 470})
                page.wait_for_function(
                    "(name) => (window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []).filter((item) => item.objName === name).length >= 2",
                    arg=COMPONENT_NAME,
                    timeout=30_000,
                )
                selected = page.evaluate(
                    "(name) => window.__S360_LAYOUT_WORKFLOW_AUDIT__?.selectByName(name) || false",
                    COMPONENT_NAME,
                )
                if not selected:
                    raise AssertionError("inserted component could not be selected for copy/paste")
                open_advanced_tools(page)
                page.get_by_role("button", name="Home", exact=True).click()
                page.get_by_role("button", name="Copy", exact=True).click()
                page.get_by_role("button", name="Paste", exact=True).click()
                page.wait_for_function(
                    "(name) => (window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []).filter((item) => item.objName === name).length >= 3",
                    arg=COMPONENT_NAME,
                    timeout=30_000,
                )
                selected_for_assembly = page.evaluate(
                    "(name) => window.__S360_LAYOUT_WORKFLOW_AUDIT__?.selectByName(name) || false",
                    COMPONENT_NAME,
                )
                if not selected_for_assembly:
                    raise AssertionError("component could not be selected for saved-assembly creation")
                prompt_answers.append(ASSEMBLY_NAME)
                open_advanced_tools(page)
                page.get_by_role("button", name="Home", exact=True).click()
                page.get_by_role("button", name="Save Selection as Assembly", exact=True).click()
                assembly_state = wait_for_project_state(
                    page,
                    port,
                    project_id,
                    lambda current: any(
                        item.get("name") == ASSEMBLY_NAME
                        for item in current.get("savedAssemblies") or []
                    ),
                    "saved assembly did not persist",
                )
                saved_assembly = next(
                    item for item in assembly_state.get("savedAssemblies") or []
                    if item.get("name") == ASSEMBLY_NAME
                )
                assembly_id = str(saved_assembly["id"])
                saved_assembly_object_ids = collect_object_ids(saved_assembly.get("object") or {})
                if not saved_assembly_object_ids:
                    raise AssertionError("saved assembly contains no stable source object identity")

                browser_panel.get_by_role("button", name="Component Builder", exact=True).click()
                builder = page.get_by_role("dialog", name="Component Builder")
                builder.wait_for(state="visible", timeout=30_000)
                advanced = builder.locator("details.libv2-advanced-workbench")
                if advanced.get_attribute("open") is None:
                    advanced.locator("summary").click()
                assembly_card = builder.locator('[aria-label="Saved assemblies"] .libv2-saved-assembly-card').filter(
                    has_text=ASSEMBLY_NAME
                )
                assembly_card.wait_for(state="visible", timeout=30_000)
                assembly_card.click()
                builder.wait_for(state="hidden", timeout=30_000)
                page.wait_for_function(
                    "(assemblyId) => (window.__S360_LAYOUT_WORKFLOW_AUDIT__?.objects() || []).some((item) => item.assemblyId === assemblyId)",
                    arg=assembly_id,
                    timeout=30_000,
                )
                page.evaluate("() => window.__S360_LAYOUT_WORKFLOW_AUDIT__?.deselect()")
                save_project(page)
                before_replace = project_for(port, project_id)
                component_page_before = page_by_id(before_replace, first_pdf_id)
                placed_before = component_objects(component_page_before)
                if len(placed_before) != 3:
                    raise AssertionError(f"click/drag/copy-paste saved {len(placed_before)} components instead of 3")
                overlay_ids_before = {str(item.get("objectId") or "") for item in placed_before}
                if "" in overlay_ids_before or len(overlay_ids_before) != 3:
                    raise AssertionError("component copies do not have three distinct stable object IDs")
                placed_assemblies_before = [
                    item for item in component_page_before.get("canvasObjects") or []
                    if item.get("assemblyId") == assembly_id
                ]
                if len(placed_assemblies_before) != 1:
                    raise AssertionError("saved assembly did not insert exactly once")
                placed_assembly_object_ids = collect_object_ids(placed_assemblies_before[0])
                if not placed_assembly_object_ids or saved_assembly_object_ids.intersection(placed_assembly_object_ids):
                    raise AssertionError("inserted saved assembly did not receive fresh recursive object IDs")
                all_overlay_ids_before = {
                    str(item.get("objectId") or "")
                    for item in component_page_before.get("canvasObjects") or []
                    if item.get("pdfBase") is not True
                }
                report["checks"].append(
                    "component click insertion, drag insertion, copy/paste, and saved-assembly reinsertion persisted"
                )
                page.screenshot(path=str(evidence / "03-pdf-components.png"), full_page=True)

                # Replace the source with a revised PDF in Fit Inside Drawing Body mode.
                preserved = {
                    str(item["id"]): {
                        "sheetTitle": item.get("sheetTitle"),
                        "sheetCode": item.get("sheetCode"),
                        "displaySheetCode": item.get("displaySheetCode"),
                        "include": item.get("include"),
                        "order": item.get("order"),
                    }
                    for item in pdf_pages(before_replace)
                }
                before_replace_pages = pdf_pages(before_replace)
                if len(before_replace_pages) != 3:
                    raise AssertionError("PDF group no longer had three pages before partial replacement")
                unmatched_id = str(before_replace_pages[2]["id"])
                unmatched_before = pdf_page_snapshot(before_replace_pages[2])
                revised_pdf_progress = import_pdf_via_ui(
                    page,
                    revised_pdf,
                    placement="fit_body",
                    replace=True,
                    expected_pages=2,
                )
                after_replace = project_for(port, project_id)
                replaced = pdf_pages(after_replace)
                if {str(item["id"]) for item in replaced} != set(pdf_ids):
                    raise AssertionError("revised PDF replacement changed the stable imported page IDs")
                for item in replaced:
                    page_id = str(item["id"])
                    actual = {
                        "sheetTitle": item.get("sheetTitle"),
                        "sheetCode": item.get("sheetCode"),
                        "displaySheetCode": item.get("displaySheetCode"),
                        "include": item.get("include"),
                        "order": item.get("order"),
                    }
                    if actual != preserved[page_id]:
                        raise AssertionError(f"revised PDF replacement changed metadata for {page_id}: {actual!r}")
                    source = item.get("sourceImport") or {}
                    if page_id == unmatched_id:
                        if pdf_page_snapshot(item) != unmatched_before:
                            raise AssertionError("partial PDF replacement changed the unmatched existing third page")
                        continue
                    if int(source.get("revision") or 0) != 2 or source.get("placementMode") != "fit_body":
                        raise AssertionError(f"revised PDF metadata is incomplete for mapped page {page_id}: {source!r}")
                    assert_current_pdf_base(item)
                group_ids = {
                    str((item.get("sourceImport") or {}).get("importGroupId") or "")
                    for item in replaced
                }
                if len(group_ids) != 1 or "" in group_ids:
                    raise AssertionError(f"partial PDF replacement duplicated the import group: {group_ids!r}")
                component_page_after = page_by_id(after_replace, first_pdf_id)
                placed_after = component_objects(component_page_after)
                if {str(item.get("objectId") or "") for item in placed_after} != overlay_ids_before:
                    raise AssertionError("revised PDF replacement did not preserve component overlays and IDs")
                all_overlay_ids_after = {
                    str(item.get("objectId") or "")
                    for item in component_page_after.get("canvasObjects") or []
                    if item.get("pdfBase") is not True
                }
                if all_overlay_ids_after != all_overlay_ids_before:
                    raise AssertionError("revised PDF replacement did not preserve the saved-assembly overlay")
                report.setdefault("pdfImportProgress", {})["revisedTwoPagePartialReplace"] = revised_pdf_progress
                report["checks"].append(
                    "two-page revised PDF partially replaced a three-page group in place; mapped IDs/metadata/overlays persisted and the unmatched third page stayed byte-equivalent"
                )

                # Prove the original upload locations are no longer dependencies.
                original_pdf.unlink()
                revised_pdf.unlink()
                if original_pdf.exists() or revised_pdf.exists():
                    raise AssertionError("disposable source PDFs could not be removed")
                project_dir_candidates = [path for path in (docs / "projects").iterdir() if project_id in path.name]
                if len(project_dir_candidates) != 1:
                    raise AssertionError(f"could not identify one project package for {project_id}")
                local_pdf_sources = sorted((project_dir_candidates[0] / "sources" / "pdf").glob("*.pdf"))
                if len(local_pdf_sources) < 2 or any(not path.read_bytes().startswith(b"%PDF") for path in local_pdf_sources):
                    raise AssertionError("project-local original/revised PDF assets are missing")

                # Update the final filename after all drawing edits so stale names cannot pass.
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Settings", exact=True).click()
                settings = page.get_by_role("dialog", name="Project Settings")
                settings.get_by_label("Drawing-package File Name", exact=True).fill(FINAL_PACKAGE_NAME)
                settings.get_by_role("button", name="Save Project Settings", exact=True).click()
                save_project(page)
                after_post_replace_save = project_for(port, project_id)
                for item in pdf_pages(after_post_replace_save):
                    assert_current_pdf_base(item)

                # Browser refresh, Project Home reopen, then a true server restart.
                page.reload(wait_until="domcontentloaded", timeout=90_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after browser refresh")
                open_advanced_tools(page)
                sheet_item(page, "Imported Browser Page One")
                refreshed_project = project_for(port, project_id)
                assert_wizard_metadata(refreshed_project, FINAL_PROJECT_METADATA, require_logo=True)
                if not logo_copies[0].is_file() or file_sha256(logo_copies[0]) != logo_fixture_sha:
                    raise AssertionError("browser refresh lost the project-local customer logo")
                report["checks"].append("saved project survived browser refresh after source PDFs were removed")

                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Home", exact=True).click()
                page.get_by_role("heading", name="Drawing Projects", exact=True).wait_for(timeout=60_000)
                project_card = page.locator(".simple-project-open").filter(has_text=PROJECT_NAME).first
                project_card.get_by_role("button", name="Open Project", exact=True).click()
                page.wait_for_url(re.compile(rf"/app\?project={project_id}&mode=editor"), timeout=60_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after Project Home reopen")
                page.goto(audit_url, wait_until="domcontentloaded", timeout=60_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after Project Home workflow-audit reopen")
                open_advanced_tools(page)
                sheet_item(page, "Imported Browser Page One")
                assert_wizard_metadata(project_for(port, project_id), FINAL_PROJECT_METADATA, require_logo=True)
                report["checks"].append("Project Home reopen restored saved standalone state")

                save_project(page)
                first_pid = server.pids[-1]
                server.stop()
                second_health = server.start()
                second_pid = int(second_health.get("pid") or 0)
                if not second_pid or second_pid == first_pid:
                    raise AssertionError("server restart did not produce a new owned PID")
                report["secondHealth"] = second_health
                page.reload(wait_until="domcontentloaded", timeout=90_000)
                assert_editor_ready(page, PROJECT_NAME)
                assert_advanced_tools_collapsed(page, "after owned server restart")
                open_advanced_tools(page)
                restarted_project = project_for(port, project_id)
                assert_wizard_metadata(restarted_project, FINAL_PROJECT_METADATA, require_logo=True)
                if not logo_copies[0].is_file() or file_sha256(logo_copies[0]) != logo_fixture_sha:
                    raise AssertionError("server restart lost the project-local customer logo")
                if [str(item["id"]) for item in pdf_pages(restarted_project)] != [str(item["id"]) for item in pdf_pages(after_replace)]:
                    raise AssertionError("server restart changed imported PDF page identity/order")
                if len(component_objects(page_by_id(restarted_project, first_pdf_id))) != 3:
                    raise AssertionError("server restart lost component overlays")
                restarted_saved_assembly = next(
                    (
                        item for item in restarted_project.get("savedAssemblies") or []
                        if item.get("id") == assembly_id and item.get("name") == ASSEMBLY_NAME
                    ),
                    None,
                )
                restarted_assembly_placements = [
                    item for item in page_by_id(restarted_project, first_pdf_id).get("canvasObjects") or []
                    if item.get("assemblyId") == assembly_id
                ]
                if restarted_saved_assembly is None or len(restarted_assembly_placements) != 1:
                    raise AssertionError("server restart lost the saved assembly definition or placed copy")
                for item in pdf_pages(restarted_project):
                    assert_current_pdf_base(item)
                report["checks"].append("server restart preserved pages, order, metadata, and overlays")
                page.screenshot(path=str(evidence / "04-after-restart.png"), full_page=True)

                # Export from the latest saved state and retain the real browser download.
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Export PDF", exact=True).click()
                export_modal = page.locator(".export-pages-modal")
                export_modal.get_by_role("heading", name="Export PDF", exact=True).wait_for(timeout=30_000)
                included_count = len([item for item in sorted_pages(restarted_project) if item.get("include") is not False])
                export_modal.get_by_text(
                    f"{included_count} included pages, in saved project order",
                    exact=True,
                ).wait_for(timeout=20_000)
                export_modal.get_by_role("button", name="Export Complete PDF", exact=True).click()

                warning_heading = page.get_by_role("heading", name="Export Warnings — Review Before PDF", exact=True)
                try:
                    warning_heading.wait_for(state="visible", timeout=8_000)
                    page.get_by_text("I understand these warnings and want to export anyway.", exact=True).click()
                    page.get_by_role("button", name="Export Anyway", exact=True).click()
                except Exception:
                    pass

                wait_for_browser_download(page, downloads, timeout=240.0)
                download = downloads[-1]
                if download.suggested_filename != f"{FINAL_PACKAGE_NAME}.pdf":
                    raise AssertionError(
                        f"export used stale/unexpected filename {download.suggested_filename!r}"
                    )
                export_path = evidence / download.suggested_filename
                download.save_as(str(export_path))
                if download.failure():
                    raise AssertionError(f"browser PDF download failed: {download.failure()}")
                final_project = project_for(port, project_id)
                write_json(evidence / "final-project.json", final_project)
                export_audit = audit_export(export_path, final_project)
                write_json(evidence / "pdf-audit.json", export_audit)
                report["pdfExport"] = export_audit
                report["checks"].append("latest saved complete PDF downloaded with correct filename/order/content")
                page.screenshot(path=str(evidence / "05-export-complete.png"), full_page=True)

                # Recoverably archive and restore the whole project from Project Home.
                pre_archive_signature = project_content_signature(final_project)
                pre_archive_package = str(project_package.resolve())
                page.locator(".ribbon-appbar-right").get_by_role("button", name="Project Home", exact=True).click()
                page.get_by_role("heading", name="Drawing Projects", exact=True).wait_for(timeout=60_000)
                active_card = page.locator(".simple-project-open.green").filter(has_text=PROJECT_NAME).first
                active_card.get_by_role("button", name="Archive", exact=True).click()
                page.get_by_role("status").filter(has_text=f"{PROJECT_NAME} was archived").wait_for(timeout=60_000)
                archived_card = page.locator(".simple-project-open.yellow").filter(has_text=PROJECT_NAME).first
                archived_card.wait_for(state="visible", timeout=30_000)
                archived_project = project_for(port, project_id)
                listed_archived = http_json(port, "/api/projects")
                if (
                    archived_project.get("archived") is not True
                    or not archived_project.get("archivedAt")
                    or any(item.get("id") == project_id for item in listed_archived.get("projects") or [])
                    or not any(item.get("id") == project_id for item in listed_archived.get("archivedProjects") or [])
                    or project_content_signature(archived_project) != pre_archive_signature
                    or str(project_package.resolve()) != pre_archive_package
                ):
                    raise AssertionError("whole-project archive changed content/package identity or list membership")
                archived_card.get_by_role("button", name="Restore Project", exact=True).click()
                page.get_by_role("status").filter(has_text=f"{PROJECT_NAME} was restored").wait_for(timeout=60_000)
                restored_card = page.locator(".simple-project-open.green").filter(has_text=PROJECT_NAME).first
                restored_card.wait_for(state="visible", timeout=30_000)
                restored_project = project_for(port, project_id)
                listed_restored = http_json(port, "/api/projects")
                if (
                    restored_project.get("archived") is not False
                    or restored_project.get("archivedAt")
                    or not restored_project.get("restoredAt")
                    or not any(item.get("id") == project_id for item in listed_restored.get("projects") or [])
                    or any(item.get("id") == project_id for item in listed_restored.get("archivedProjects") or [])
                    or project_content_signature(restored_project) != pre_archive_signature
                    or str(project_package.resolve()) != pre_archive_package
                ):
                    raise AssertionError("whole-project restore changed content/package identity or list membership")
                restored_card.get_by_role("button", name="Open Project", exact=True).click()
                page.wait_for_url(re.compile(rf"/app\?project={project_id}&mode=editor"), timeout=60_000)
                assert_editor_ready(page, PROJECT_NAME)
                page.goto(audit_url, wait_until="domcontentloaded", timeout=60_000)
                assert_editor_ready(page, PROJECT_NAME)
                final_project = project_for(port, project_id)
                if project_content_signature(final_project) != pre_archive_signature:
                    raise AssertionError("reopened restored project changed drawing content")
                report["checks"].append("whole-project archive/restore preserved package, page, source, and object identity")
                page.screenshot(path=str(evidence / "06-project-restored.png"), full_page=True)

                # Retain a concise final summary before disposable docs are removed.
                report["finalProject"] = {
                    "projectMode": final_project.get("projectMode"),
                    "pageCount": len(final_project.get("pages") or []),
                    "includedCount": len([item for item in final_project.get("pages") or [] if item.get("include") is not False]),
                    "archivedPageCount": len(final_project.get("archivedPages") or []),
                    "pdfPageIds": [str(item["id"]) for item in pdf_pages(final_project)],
                    "componentOverlayCount": len(component_objects(page_by_id(final_project, first_pdf_id))),
                    "savedAssemblyCount": len(final_project.get("savedAssemblies") or []),
                    "placedAssemblyCount": len([
                        item for item in page_by_id(final_project, first_pdf_id).get("canvasObjects") or []
                        if item.get("assemblyId") == assembly_id
                    ]),
                    "csvPageCount": len([
                        item for item in final_project.get("pages") or []
                        if (item.get("sourceImport") or {}).get("sourceType") == "csv"
                    ]),
                    "imagePageCount": len([
                        item for item in final_project.get("pages") or []
                        if (item.get("sourceImport") or {}).get("type") == "image"
                    ]),
                    "workbookPageCount": len([
                        item for item in final_project.get("pages") or []
                        if (item.get("sourceImport") or {}).get("sourceType") == "excel_workbook"
                    ]),
                    "drawingPackageFileName": (final_project.get("metadata") or {}).get("drawingPackageFileName"),
                    "projectLocalPdfSources": [
                        {"name": path.name, "sha256": file_sha256(path), "bytes": path.stat().st_size}
                        for path in local_pdf_sources
                    ],
                }

                context.close()
                browser.close()

            library_after = tree_manifest(docs / "library")
            write_json(evidence / "component-library-after.sha256.json", library_after)
            library_changed = sorted(set(library_before) | set(library_after))
            library_changed = [
                name for name in library_changed
                if library_before.get(name) != library_after.get(name)
            ]
            unexpected_library_changes = [
                name for name in library_changed
                if not name.startswith("page_templates/")
            ]
            if unexpected_library_changes:
                raise AssertionError(
                    f"browser workflow changed protected component-library files: {unexpected_library_changes}"
                )
            if not library_changed or not any(name.startswith("page_templates/") for name in library_changed):
                raise AssertionError("page-template save/insert did not create its expected disposable library records")
            report["componentLibrary"] = {
                "fileCount": len(library_after),
                "componentCatalogUnchanged": True,
                "manifestSha256": library_after.get("manifest.json", ""),
                "expectedPageTemplateChanges": library_changed,
            }
            workbook_like = [
                path.relative_to(docs).as_posix()
                for path in docs.rglob("*")
                if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".xlsb"}
            ]
            if len(workbook_like) != 1 or "/sources/workbook/" not in f"/{workbook_like[0]}":
                raise AssertionError(f"unexpected disposable workbook files were created: {workbook_like}")
            retained_workbook = docs / workbook_like[0]
            if file_sha256(retained_workbook) != workbook_fixture_sha:
                raise AssertionError("project-local generated workbook differs from its disposable fixture")
            report["workbookFilesCreated"] = workbook_like
            report["checks"].append(
                "component catalog stayed byte-identical; only the expected page template and generated project-local workbook were added"
            )

    except Exception:  # noqa: BLE001 - retain full diagnostic evidence and return nonzero
        failure = traceback.format_exc()
        report["failure"] = failure
    finally:
        if server is not None:
            server.stop()
        report["serverPids"] = server.pids if server is not None else []
        report["serverLauncherPids"] = server.launcher_pids if server is not None else []
        report["finishedAt"] = datetime.now(timezone.utc).isoformat()
        write_json(evidence / "browser-requests.json", network_requests)
        write_json(evidence / "browser-responses.json", network_responses)
        write_json(evidence / "browser-request-failures.json", network_failures)
        write_json(evidence / "browser-console.json", console_messages)
        write_json(evidence / "browser-page-errors.json", page_errors)
        write_json(evidence / "browser-dialogs.json", dialogs)

        access_requests = access_log_requests(
            [evidence / "server.stdout.log", evidence / "server.stderr.log"]
        )
        write_json(evidence / "server-access-requests.json", access_requests)
        forbidden_browser = [item for item in network_requests if forbidden_url(str(item.get("url") or ""))]
        forbidden_server = [item for item in access_requests if forbidden_url(str(item.get("path") or ""))]
        api_errors = [
            item for item in network_responses
            if int(item.get("status") or 0) >= 400 and "/api/" in str(item.get("url") or "")
        ]
        console_errors = [item for item in console_messages if item.get("type") == "error"]
        unexpected_dialogs = [item for item in dialogs if item.get("action") == "dismissed-unexpected"]
        report["browserEvidence"] = {
            "requestCount": len(network_requests),
            "responseCount": len(network_responses),
            "requestFailureCount": len(network_failures),
            "consoleMessageCount": len(console_messages),
            "consoleErrorCount": len(console_errors),
            "pageErrors": page_errors,
            "apiHttpErrors": api_errors,
            "forbiddenBrowserRequests": forbidden_browser,
            "forbiddenServerRequests": forbidden_server,
            "unexpectedDialogs": unexpected_dialogs,
            "servedAssetUrls": sorted(
                {
                    str(item.get("url"))
                    for item in network_requests
                    if urlparse(str(item.get("url") or "")).path.startswith("/assets/")
                }
            ),
        }
        served_asset_files: list[dict[str, Any]] = []
        for url in report["browserEvidence"]["servedAssetUrls"]:
            relative = urlparse(url).path.removeprefix("/")
            local = ROOT / "frontend" / "dist" / relative
            served_asset_files.append(
                {
                    "url": url,
                    "path": str(local),
                    "exists": local.is_file(),
                    "bytes": local.stat().st_size if local.is_file() else 0,
                    "sha256": file_sha256(local) if local.is_file() else "",
                }
            )
        report.setdefault("servedBuild", {})["requestedAssets"] = served_asset_files

        if not failure:
            assertions = {
                "pageErrors": page_errors,
                "apiHttpErrors": api_errors,
                "forbiddenBrowserRequests": forbidden_browser,
                "forbiddenServerRequests": forbidden_server,
                "unexpectedDialogs": unexpected_dialogs,
                "consoleErrors": console_errors,
                "requestFailures": network_failures,
            }
            failed_assertions = {name: value for name, value in assertions.items() if value}
            if failed_assertions:
                failure = "Browser diagnostics were not clean:\n" + json.dumps(failed_assertions, indent=2)
                report["failure"] = failure
            else:
                report["checks"].append("zero browser/server workbook-link, workbook-quality, or sync requests")
                report["checks"].append("zero API HTTP failures, request failures, console errors, or page errors")

        report["ok"] = not bool(failure)
        write_json(evidence / "report.json", report)

    if failure:
        print(f"FAIL: standalone browser smoke\n{failure}", file=sys.stderr)
        print(f"Evidence: {evidence}", file=sys.stderr)
        return 1
    print("PASS: standalone browser smoke")
    print(f"Project: {report.get('projectId')}")
    print(f"Server PIDs: {report.get('serverPids')}")
    print(f"Checks: {len(report.get('checks') or [])}")
    print(f"PDF: {(report.get('pdfExport') or {}).get('path')}")
    print(f"Evidence: {evidence}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
