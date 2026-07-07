"""Smoke: blank drawing/layout/pdf-vector pages still get the dark page-header
band, not just tables (FINAL RENDER POLISH 4G, Phase D).

Verifies:
  - Overall Layout, LCP-1/LCP-2 Control Wiring Schematic, and Interior/Exterior
    Device Location pages classify as the ``canvas`` page type/family and are
    NOT covers, so the Singh360 header band is eligible to render on them.
  - Every one of those pages keeps ``normalizedHeaderStyle == "orange"`` (the
    band's on/off switch) — only the cover page is exempt.
  - Static regression guard: NormalizedPage.tsx's header-band visibility check
    must not exclude image/canvas-type pages (the exact bug that hid the
    header on blank drawing pages) and must not require a table body.
"""
from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Code", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "EMS 0.0", "Cover", "Cover Sheet", "YES", ""])
    idx.append(["2", "EMS 2.0", "Overall Layout", "EMS Controls Overall Layout", "YES", ""])
    idx.append(["3", "EMS 2.1", "LCP-1 Schematic", "LCP-1 Control Wiring Schematic", "YES", ""])
    idx.append(["4", "EMS 2.2", "LCP-2 Schematic", "LCP-2 Control Wiring Schematic", "YES", ""])
    idx.append(["5", "EMS 2.3", "Interior Location", "Interior Device Location", "YES", ""])
    idx.append(["6", "EMS 2.4", "Exterior Location", "Exterior Device Location", "YES", ""])

    wb.create_sheet("Cover").cell(1, 1, "SINGH360 EMS PROJECT")
    for tab in (
        "Overall Layout",
        "LCP-1 Schematic",
        "LCP-2 Schematic",
        "Interior Location",
        "Exterior Location",
    ):
        # Deliberately blank/near-blank so _tabular_enough() is false — these are
        # drawing/layout pages, not data tables (matches the real workbooks).
        wb.create_sheet(tab)

    wb.save(path)


LAYOUT_TABS = (
    "Overall Layout",
    "LCP-1 Schematic",
    "LCP-2 Schematic",
    "Interior Location",
    "Exterior Location",
)

# "Layout" / "Location" / "Overall" map to the plain canvas family; "Wiring
# Schematic" sheets map to the more specific wiringDiagram family. Both are
# blank/drawing page families that need the header band with no table body.
_EXPECTED_FAMILY = {
    "Overall Layout": "canvas",
    "LCP-1 Schematic": "wiringDiagram",
    "LCP-2 Schematic": "wiringDiagram",
    "Interior Location": "canvas",
    "Exterior Location": "canvas",
}


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "layout_headers.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="layouthdr1")
    pages = proj["pages"]
    problems: list[str] = []

    for tab in LAYOUT_TABS:
        p = next((pg for pg in pages if pg["sheetTab"] == tab), None)
        if p is None:
            problems.append(f"{tab}: page not found")
            continue
        if p.get("pageType") not in ("canvas", "hybrid"):
            problems.append(f"{tab}: pageType={p.get('pageType')!r}, expected canvas/hybrid drawing page")
        expected_family = _EXPECTED_FAMILY[tab]
        if p.get("pageFamily") != expected_family:
            problems.append(f"{tab}: pageFamily={p.get('pageFamily')!r}, expected {expected_family!r}")
        if p.get("normalizedHeaderStyle") != "orange":
            problems.append(
                f"{tab}: normalizedHeaderStyle={p.get('normalizedHeaderStyle')!r}, "
                "expected 'orange' (header band must be eligible on blank drawing pages)"
            )
        if p.get("pageType") == "cover":
            problems.append(f"{tab}: misclassified as cover")

    cover = next((pg for pg in pages if pg["sheetTab"] == "Cover"), None)
    if cover is None:
        problems.append("Cover page not found")
    elif cover.get("pageType") != "cover":
        problems.append(f"Cover page misclassified as {cover.get('pageType')!r}")

    # Static regression guard on the frontend header-band gate itself.
    norm_path = ROOT / "frontend" / "src" / "components" / "renderers" / "NormalizedPage.tsx"
    if not norm_path.exists():
        problems.append("NormalizedPage.tsx not found")
    else:
        source = norm_path.read_text("utf-8")
        m = re.search(r"const showBand\s*=\s*([^;]+);", source)
        if not m:
            problems.append("NormalizedPage.tsx: could not find `showBand` assignment")
        else:
            expr = m.group(1)
            if "isImageType" in expr:
                problems.append(
                    "NormalizedPage.tsx: showBand still excludes isImageType — blank drawing/"
                    "layout/pdf-vector pages will lose their dark page header again"
                )
            if "isCoverPage" not in expr:
                problems.append("NormalizedPage.tsx: showBand no longer excludes the cover page")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - layout/pdf-vector/blank drawing page headers passed")


if __name__ == "__main__":
    main()
