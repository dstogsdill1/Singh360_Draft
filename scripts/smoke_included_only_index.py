"""Smoke: normalized Sheet Index lists included output pages only."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["2", "00_INDEX", "Sheet Index", "YES", ""])
    idx.append(["5", "SCOPE", "Project Scope", "YES", ""])
    idx.append(["99", "OPTIONAL", "Optional Full EMS", "NO", "disabled in app until included"])
    idx.append(["6", "WORKFLOW", "Workflow", "TRUE", ""])
    for cell in idx[1]:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True)

    scope = wb.create_sheet("SCOPE")
    scope.append(["Item", "Notes"])
    scope.append(["Closeout", "Final package"])
    optional = wb.create_sheet("OPTIONAL")
    optional.append(["This should not output"])
    workflow = wb.create_sheet("WORKFLOW")
    workflow.append(["Milestone", "Status"])
    workflow.append(["Kickoff", "Ready"])
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "included_only_index.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="idxonly1")
    problems: list[str] = []

    if any(p["sheetTab"] == "OPTIONAL" for p in proj["pages"]):
        problems.append("include=NO OPTIONAL sheet produced an output page")

    index_pages = [p for p in proj["pages"] if p["sheetTab"] == "00_INDEX"]
    if len(index_pages) != 1:
        problems.append(f"Sheet Index expected 1 page, got {len(index_pages)}")
    else:
        block = index_pages[0]["blocks"][0]
        text = "\n".join(" | ".join(r) for r in block.get("grid") or [])
        if "Optional Full EMS" in text or "OPTIONAL" in text:
            problems.append("Sheet Index output still contains NO/optional row")
        if block.get("type") != "excelRange":
            problems.append("Sheet Index did not render as a normalized table range")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - included-only index passed")


if __name__ == "__main__":
    main()
