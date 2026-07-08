"""Smoke: Kyle/SA38-style workbook import (FINAL RELEASE CLEANUP 4H+SA38, Phase A).

Builds an in-memory workbook that mirrors the real SA38 upload structure:
  - ``00_APP_INDEX`` (alias columns) appears BEFORE ``00_INDEX`` in tab order —
    exactly like the real file — so a naive "first sheet containing INDEX"
    rule would pick the wrong one.
  - ``00_INDEX`` (canonical columns) is the real controlling index and must
    win.
  - ``00_PROJECT_META`` is a plain key/value metadata sheet.

Verifies:
  - ``00_INDEX`` is preferred over ``00_APP_INDEX``.
  - ``00_APP_INDEX``, ``00_INDEX``, and ``00_PROJECT_META`` never appear as
    output pages.
  - BOM / TOC get the index's mapped Sheet Code (not sequential order).
  - Rack A gets its mapped code and is *not* misclassified as a blank canvas
    page (real regression found during this pass: a title containing
    "I/O & Layout" was previously swallowed by the canvas keyword rule).
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")


def _workbook(path: Path) -> None:
    wb = Workbook()
    bom = wb.active
    bom.title = "BOM"
    bom.append(["Item", "Description", "Qty"])
    bom.append(["1", "Controller", "4"])
    bom.append(["2", "Relay", "8"])

    toc = wb.create_sheet("TOC")
    toc.append(["Sheet", "Title"])
    toc.append(["EMS 0.0", "BOM"])

    rack_a = wb.create_sheet("RACK A")
    rack_a.cell(1, 2, "SuperPak Controller I/O Panel for RACK A")
    for c in range(2, 7):
        rack_a.cell(1, c).fill = GOLD
    for c, h in enumerate(["RO#", "Description", "Type", "DI#", "Status"], start=2):
        cell = rack_a.cell(2, c, h)
        cell.fill = GRAY
        cell.font = Font(bold=True)
    for i in range(10):
        for c in range(2, 7):
            rack_a.cell(3 + i, c, f"RackA-{i}-{c}")

    # 00_APP_INDEX intentionally comes BEFORE 00_INDEX (matches the real
    # SA38 upload's tab order) — a naive "first INDEX match" importer would
    # pick this one and misread its alias-only columns.
    app_idx = wb.create_sheet("00_APP_INDEX")
    app_idx.append(["Include", "Order", "Suggested EMS Code", "Original Tab", "Normalized Page Title", "Family", "Page Type", "Notes"])
    app_idx.append(["YES", 1, "EMS 0.0", "BOM", "Bill of Materials", "BOM", "table", ""])
    app_idx.append(["YES", 2, "EMS 0.1", "TOC", "Table of Contents", "Front Matter", "index", ""])
    app_idx.append(["YES", 3, "EMS 9.9", "RACK A", "WRONG ALIAS CODE", "Refrigeration", "io-table", ""])

    idx = wb.create_sheet("00_INDEX")
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "BOM", "Bill of Materials", "Front Matter", "table", "Kyle original preserved."])
    idx.append(["YES", 2, "EMS 0.1", "TOC", "Sheet Index / TOC", "Front Matter", "index", ""])
    idx.append(["YES", 3, "EMS 2.0", "RACK A", "Rack A I/O & Layout", "Refrigeration", "io-table", "Controller sections must split only at section boundaries."])

    meta = wb.create_sheet("00_PROJECT_META")
    meta.append(["Project Metadata", None])
    meta.append(["Project Name", "SA38 Test Project"])
    meta.append(["Drawing Package File Name", "SA38_Test_Upload"])
    meta.append(["Location", "Test City, TX"])
    meta.append(["Revision", "V1"])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "sa38_fixture.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="sa38fixture")
    pages = proj["pages"]
    page_tabs = {p["sheetTab"] for p in pages}
    problems: list[str] = []

    for control in ("00_INDEX", "00_APP_INDEX", "00_PROJECT_META"):
        if control in page_tabs:
            problems.append(f"{control} appeared as an output page")

    def page(tab: str) -> dict | None:
        return next((p for p in pages if p["sheetTab"] == tab), None)

    bom = page("BOM")
    if bom is None:
        problems.append("BOM page not found")
    elif bom.get("displaySheetCode") != "EMS 0.0":
        problems.append(f"BOM sheetCode={bom.get('displaySheetCode')!r}, expected 'EMS 0.0'")

    toc = page("TOC")
    if toc is None:
        problems.append("TOC page not found")
    elif toc.get("displaySheetCode") != "EMS 0.1":
        problems.append(f"TOC sheetCode={toc.get('displaySheetCode')!r}, expected 'EMS 0.1'")

    rack_a = page("RACK A")
    if rack_a is None:
        problems.append("RACK A page not found")
    else:
        # Must use 00_INDEX's code (EMS 2.0), never 00_APP_INDEX's (EMS 9.9) —
        # proves 00_INDEX won the preference over 00_APP_INDEX.
        if rack_a.get("displaySheetCode") != "EMS 2.0":
            problems.append(f"RACK A sheetCode={rack_a.get('displaySheetCode')!r}, expected 'EMS 2.0' (00_INDEX should win over 00_APP_INDEX's EMS 9.9)")
        if rack_a.get("pageType") == "canvas":
            problems.append("RACK A misclassified as a blank canvas page (I/O table content was lost)")
        grid = (rack_a.get("blocks") or [{}])[0].get("grid") or []
        if len(grid) < 5:
            problems.append(f"RACK A rendered grid too small ({len(grid)} rows) — table content likely lost")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA38/Kyle workbook import passed")
    print(f"  output pages: {len(pages)}, BOM={bom.get('displaySheetCode')}, TOC={toc.get('displaySheetCode')}, RACK A={rack_a.get('displaySheetCode')}")


if __name__ == "__main__":
    main()
