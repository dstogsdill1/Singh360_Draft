"""Smoke: Sheet Index rows sync with exported pages (FINAL SA31 POLISH 4I Phase A).

Proves:
  - Index row count (included) equals exported page count.
  - Every indexed sheet code exists as an exported page.
  - Every exported page appears in the index.
  - EMS 1.4a continuation is ordered immediately after EMS 1.4.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "Cover", "Cover Sheet", "Front Matter", "cover", ""])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""])
    idx.append(["YES", 3, "EMS 0.9", "Revision Log", "Revision / Issue Log", "Front Matter", "table", ""])
    idx.append(["YES", 4, "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "Lighting", "io-table", ""])
    idx.append(["YES", 5, "EMS 9.0", "Company Info", "Singh360 Company Info", "Reference", "text", ""])

    cover = wb.create_sheet("Cover")
    cover.cell(1, 1, "SINGH360 EMS PROJECT")

    rev = wb.create_sheet("Revision Log")
    rev.append(["Rev", "Date", "Description", "By"])
    rev.append(["V1", "2026-07-08", "Initial issue", "DS"])

    lcp = wb.create_sheet("LCP Panel Schedule")
    lcp.append(["RO#", "Description", "Type", "DI#", "Status", "Notes"])
    for i in range(60):
        lcp.append([str(i + 1), f"Relay {i}", "NO", str(i + 1), "OK", ""])
        lcp.row_dimensions[i + 2].height = 30

    company = wb.create_sheet("Company Info")
    company.append(["Field", "Value"])
    company.append(["Company", "Singh360 Inc."])

    wb.save(path)


def _index_codes(block: dict) -> list[str]:
    grid = block.get("grid") or []
    header_idx = 0
    for i, row in enumerate(grid[:20]):
        low = {str(x).lower() for x in row if x}
        if "sheet code" in low or "page title" in low:
            header_idx = i
            break
    header = [str(x).lower() for x in grid[header_idx]]
    code_col = header.index("sheet code") if "sheet code" in header else 2
    include_col = header.index("include") if "include" in header else 0
    codes: list[str] = []
    for row in grid[header_idx + 1:]:
        include = str(row[include_col] if include_col < len(row) else "").strip().lower()
        if include and include not in {"y", "yes", "true", "1", "include", "x", "on"}:
            continue
        code = str(row[code_col] if code_col < len(row) else "").strip()
        if code:
            codes.append(code)
    return codes


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "index_sync.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="idxsync1")
    pages = [p for p in proj["pages"] if p.get("include", True)]
    problems: list[str] = []

    index_page = next((p for p in pages if p.get("pageType") == "index"), None)
    if index_page is None:
        problems.append("index page missing")
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)

    block = next((b for b in (index_page.get("blocks") or []) if b.get("type") == "excelRange"), None)
    if block is None:
        problems.append("index excelRange block missing")
    else:
        index_codes = _index_codes(block)
        page_codes = [(p.get("displaySheetCode") or p.get("sheetCode") or "").strip() for p in pages]
        if len(index_codes) != len(pages):
            problems.append(
                f"index row count {len(index_codes)} != exported page count {len(pages)}; "
                f"index={index_codes} pages={page_codes}"
            )
        missing_pages = [c for c in index_codes if c not in page_codes]
        if missing_pages:
            problems.append(f"indexed codes missing from PDF pages: {missing_pages}")
        missing_index = [c for c in page_codes if c not in index_codes]
        if missing_index:
            problems.append(f"exported pages missing from index: {missing_index}")
        if "EMS 1.4" in index_codes and "EMS 1.4a" in index_codes:
            i14 = index_codes.index("EMS 1.4")
            i14a = index_codes.index("EMS 1.4a")
            if i14a != i14 + 1:
                problems.append(f"EMS 1.4a not immediately after EMS 1.4 (positions {i14}, {i14a})")
        else:
            problems.append(f"expected EMS 1.4 and EMS 1.4a in index, got {index_codes}")

        # Continuation row metadata completeness.
        for row in block.get("grid") or []:
            if any(str(c).strip() == "EMS 1.4a" for c in row):
                joined = " | ".join(str(c) for c in row)
                if "continued" not in joined.lower():
                    problems.append(f"EMS 1.4a index row missing Continued title: {joined}")
                if "lighting" not in joined.lower() and "io-table" not in joined.lower():
                    problems.append(f"EMS 1.4a index row incomplete metadata: {joined}")
                break

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA31 index/output sync passed")
    print(f"  pages={len(pages)}, index_codes={_index_codes(block)}")


if __name__ == "__main__":
    main()
