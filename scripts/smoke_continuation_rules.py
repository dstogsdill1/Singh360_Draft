"""Smoke: no dumb continuation pages for compact index/scope/workflow sheets."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")


def _sa31_like(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    for order, tab, title in [
        ("2", "00_INDEX", "Sheet Index"),
        ("5", "SCOPE", "Project Scope"),
        ("6", "WORKFLOW", "Project Workflow / Milestones"),
    ]:
        idx.append([order, tab, title, "YES", ""])

    for r in range(4, 14):
        idx.append([str(r), f"TAB{r}", f"Title {r}", "YES", ""])

    scope = wb.create_sheet("SCOPE")
    scope["A1"] = "Phase"
    scope["B1"] = "Deliverable"
    scope["A2"] = "Closeout"
    scope["B2"] = "As-built package"

    wf = wb.create_sheet("WORKFLOW")
    wf["A1"] = "Milestone"
    wf["B1"] = "Date"
    wf["A2"] = "Kickoff"
    wf["B2"] = "TBD"
    wf["A3"] = "Substantial"
    wf["B3"] = "TBD"

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "continuation.xlsx"
    _sa31_like(xlsx)
    proj = import_workbook(xlsx, project_id="cont1")
    problems: list[str] = []

    for tab in ("00_INDEX", "SCOPE", "WORKFLOW"):
        pages = [p for p in proj["pages"] if p["sheetTab"] == tab]
        if len(pages) != 1:
            problems.append(f"{tab}: expected 1 page, got {len(pages)}")
        elif pages[0].get("generatedContinuation"):
            problems.append(f"{tab}: base page flagged as continuation")

    for p in proj["pages"]:
        title = p.get("sheetTitle", "")
        if title.lower().count("continued") > 1:
            problems.append(f"duplicate CONTINUED in title: {title}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — continuation rules passed (index/scope/workflow = 1 page each)")


if __name__ == "__main__":
    main()
