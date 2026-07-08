"""Smoke: Revision / Issue Log included by default for SA31 (4I Phase A)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "Cover", "Cover / Project Info", "Front Matter", "cover", ""])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index", "Front Matter", "index", ""])
    idx.append(["YES", 10, "EMS 0.9", "EMS 0.9 Revision Log", "Revision / Issue Log", "Front Matter", "table", ""])
    idx.append(["YES", 11, "EMS 1.0", "EMS 1.0 Overall Layout", "EMS Controls Overall Layout", "Overview", "layout", ""])

    cover = wb.create_sheet("Cover")
    cover.cell(1, 1, "COVER")

    rev = wb.create_sheet("EMS 0.9 Revision Log")
    rev.append(["Rev", "Date", "Description", "By"])
    rev.append(["V1", "2026-07-08", "Initial issue", "DS"])
    rev.append(["", "", "", ""])

    layout = wb.create_sheet("EMS 1.0 Overall Layout")
    layout.append(["Note", "Use PDF vector embed"])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "revlog.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="revlog1")
    pages = [p for p in proj["pages"] if p.get("include", True)]
    problems: list[str] = []

    rev = next((p for p in pages if "Revision" in (p.get("sheetTitle") or "")), None)
    if rev is None:
        problems.append("Revision / Issue Log not exported as a real page despite Include=YES")
    else:
        code = rev.get("displaySheetCode") or rev.get("sheetCode")
        if code != "EMS 0.9":
            problems.append(f"Revision Log sheet code={code!r}, expected EMS 0.9")
        if not (rev.get("blocks") or []):
            problems.append("Revision Log page has no content blocks")

    # Index must list it too.
    index_page = next((p for p in pages if p.get("pageType") == "index"), None)
    if index_page is None:
        problems.append("index missing")
    else:
        block = next((b for b in (index_page.get("blocks") or []) if b.get("type") == "excelRange"), None)
        text = "\n".join(" | ".join(str(c) for c in r) for r in (block.get("grid") or []))
        if "EMS 0.9" not in text or "Revision" not in text:
            problems.append("Sheet Index does not list Revision / Issue Log")

    # Overall layout still after revision when both included.
    if rev is not None:
        layout = next((p for p in pages if "Overall Layout" in (p.get("sheetTitle") or "")), None)
        if layout and int(layout.get("order") or 0) <= int(rev.get("order") or 0):
            problems.append(
                f"Overall Layout order {layout.get('order')} should follow Revision Log order {rev.get('order')}"
            )

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA31 Revision Log included as real page")
    print(f"  pages={len(pages)}, rev_order={rev.get('order')}, rev_code={rev.get('displaySheetCode')}")


if __name__ == "__main__":
    main()
