"""Smoke: title block SHEET NO. uses canonical index codes, not physical order."""
from __future__ import annotations

import copy
import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.sheet_index_sync import sync_project_sheet_index
from core.workbook_importer import import_workbook

SA31 = ROOT / ".docs" / "projects" / "SA31-102-EMS-Lighting__ce333f83502742d3" / "project.json"

EXPECTED = {
    "Cover / Project Info": "EMS 0.0",
    "Sheet Index / TOC": "EMS 0.1",
    "Project Scope": "EMS 0.4",
    "Project Workflow / Milestones": "EMS 0.5",
    "Responsibility Matrix": "EMS 0.6",
    "Revision / Issue Log": "EMS 0.9",
    "EMS Controls Overall Layout": "EMS 1.0",
    "RDM / IDF Network Table": "EMS 1.1",
    "Lighting Output Matrix": "EMS 1.3",
    "LCP Panel Schedule": "EMS 1.4",
    "EC Field Instructions": "EMS 2.0",
    "LCP-1 Control Wiring Schematic": ("EMS 3.0", "EMS 3.0 LCP1 Schematic"),
    "Interior Device Location": "EMS 4.0",
    "Singh360 Company Info": "EMS 9.0",
}


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type"])
    idx.append(["YES", 1, "EMS 0.0", "Cover", "Cover / Project Info", "Front Matter", "cover"])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index / TOC", "Front Matter", "index"])
    idx.append(["YES", 5, "EMS 0.4", "Scope", "Project Scope", "Front Matter", "text"])
    idx.append(["YES", 15, "EMS 1.4", "LCP Panel Schedule", "LCP Panel Schedule", "Lighting", "io-table"])

    wb.create_sheet("Cover").cell(1, 1, "COVER")
    scope = wb.create_sheet("Scope")
    scope.append(["Section", "Scope Language"])
    scope.append(["Executive Summary", "Scope text"])
    lcp = wb.create_sheet("LCP Panel Schedule")
    lcp.append(["RO#", "Description", "Type", "DI#", "Status", "Notes"])
    for i in range(55):
        lcp.append([str(i + 1), f"Relay {i}", "NO", str(i + 1), "OK", ""])
        lcp.row_dimensions[i + 2].height = 30
    wb.save(path)


def _index_codes(project: dict) -> list[str]:
    index_page = next((p for p in project["pages"] if p.get("pageType") == "index"), None)
    if not index_page:
        return []
    block = next((b for b in (index_page.get("blocks") or []) if b.get("type") == "excelRange"), None)
    if not block:
        return []
    grid = block.get("grid") or []
    header_idx = 0
    for i, row in enumerate(grid[:20]):
        low = {str(x).lower() for x in row if x}
        if "sheet code" in low:
            header_idx = i
            break
    header = [str(x).lower() for x in grid[header_idx]]
    code_col = header.index("sheet code") if "sheet code" in header else 2
    codes: list[str] = []
    for row in grid[header_idx + 1 :]:
        if code_col < len(row) and str(row[code_col]).strip():
            codes.append(str(row[code_col]).strip())
    return codes


def _title_block_code(page: dict) -> str:
    return (page.get("displaySheetCode") or page.get("sheetCode") or "").strip()


def _is_physical_leak(code: str, order: int) -> bool:
    import re
    m = re.fullmatch(r"EMS\s+(\d+)\.0(?:[a-z])?", code, re.IGNORECASE)
    return bool(m and int(m.group(1)) == order)


def test_fresh_import() -> list[str]:
    problems: list[str] = []
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "tb_codes.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="tbcodes01")
    sync_project_sheet_index(proj)

    scope = next(p for p in proj["pages"] if "Scope" in (p.get("sheetTitle") or ""))
    if _title_block_code(scope) != "EMS 0.4":
        problems.append(f"Scope title block={_title_block_code(scope)!r}, expected EMS 0.4")

    lcp_pages = [p for p in proj["pages"] if "LCP Panel Schedule" in (p.get("sheetTitle") or "")]
    if len(lcp_pages) < 2:
        problems.append("LCP continuation missing")
    elif _title_block_code(lcp_pages[1]) != "EMS 1.4a":
        problems.append(f"LCP cont title block={_title_block_code(lcp_pages[1])!r}, expected EMS 1.4a")

    page_codes = [_title_block_code(p) for p in proj["pages"] if p.get("include")]
    index_codes = _index_codes(proj)
    if index_codes != page_codes:
        problems.append(f"index/export mismatch index={index_codes} pages={page_codes}")

    for p in proj["pages"]:
        if not p.get("include"):
            continue
        order = int(p.get("order") or 0)
        code = _title_block_code(p)
        if _is_physical_leak(code, order):
            problems.append(f"physical order leaked into SHEET NO.: order={order} code={code!r}")
    return problems


def test_sa31_repair() -> list[str]:
    if not SA31.is_file():
        return []
    problems: list[str] = []
    proj = json.loads(SA31.read_text(encoding="utf-8"))
    corrupted = copy.deepcopy(proj)
    for p in corrupted["pages"]:
        if p.get("include"):
            o = int(p.get("order") or 0)
            p["sheetCode"] = f"EMS {o}.0"
            p["displaySheetCode"] = f"EMS {o}.0"
    cont = next((p for p in corrupted["pages"] if p.get("generatedContinuation")), None)
    if cont:
        cont["sheetCode"] = "EMS 15.0a"
        cont["displaySheetCode"] = "EMS 15.0a"

    sync_project_sheet_index(corrupted)

    for key, spec in EXPECTED.items():
        if isinstance(spec, tuple):
            code, tab = spec
        else:
            code, tab = spec, None
        if tab:
            page = next(
                (p for p in corrupted["pages"] if p.get("include") and p.get("sheetTab") == tab),
                None,
            )
        else:
            page = next(
                (p for p in corrupted["pages"] if p.get("include") and p.get("sheetTitle") == key),
                None,
            )
        if page is None:
            continue
        got = _title_block_code(page)
        if got != code:
            problems.append(f"{key}: SHEET NO.={got!r}, expected {code!r}")

    cont = next((p for p in corrupted["pages"] if p.get("generatedContinuation")), None)
    if cont and _title_block_code(cont) != "EMS 1.4a":
        problems.append(f"LCP continuation after repair={_title_block_code(cont)!r}, expected EMS 1.4a")

    included = [p for p in corrupted["pages"] if p.get("include")]
    page_codes = [_title_block_code(p) for p in included]
    index_codes = _index_codes(corrupted)
    if index_codes != page_codes:
        problems.append(f"SA31 index/export mismatch index={index_codes[-5:]} pages={page_codes[-5:]}")

    if any("EMS 4.1" in c for c in index_codes):
        ext = next((p for p in corrupted["pages"] if "Exterior" in (p.get("sheetTitle") or "")), None)
        if ext and not ext.get("include"):
            problems.append("Exterior Device Location in index but not exported")

    for p in included:
        order = int(p.get("order") or 0)
        code = _title_block_code(p)
        if _is_physical_leak(code, order):
            problems.append(f"SA31 physical leak order={order} code={code!r}")

    # Page X of Y stays physical
    totals = {p.get("pageTotal") for p in included}
    if totals != {len(included)}:
        problems.append(f"pageTotal should be {len(included)}, got {totals}")
    nums = [p.get("pageNumber") for p in included]
    if nums != list(range(1, len(included) + 1)):
        problems.append(f"physical page numbers broken: {nums[:8]}...")
    return problems


def main() -> int:
    problems = test_fresh_import() + test_sa31_repair()
    if problems:
        print("FAIL — title block sheet codes")
        for p in problems:
            print(" -", p)
        return 1
    print("OK — title block sheet codes and index sync passed")
    print("  fresh import + SA31 repair verified")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
