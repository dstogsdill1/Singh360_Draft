"""Smoke: Guidelines and Field Instructions avoid TABLE OVERFLOW and tiny strips (Phase B)."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.workbook_geometry import excel_column_width_to_pixels
from core.workbook_importer import import_workbook


COL_WIDTHS = (24.0, 118.0)


def _set_instruction_geometry(ws) -> None:
    ws.column_dimensions["A"].width = COL_WIDTHS[0]
    ws.column_dimensions["B"].width = COL_WIDTHS[1]
    ws.sheet_format.defaultRowHeight = 18.0
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title"])
    idx.append(["YES", 1, "EMS 3.0", "Guidelines", "EMS Guidelines"])
    idx.append(["YES", 2, "EMS 17.0", "Field Instructions", "Field Instructions"])

    guide = wb.create_sheet("Guidelines")
    guide.append(["Topic", "Guideline"])
    for i in range(18):
        guide.append([f"Topic {i}", f"Guideline text row {i} with enough words to wrap naturally across the page body."])
    _set_instruction_geometry(guide)

    field = wb.create_sheet("Field Instructions")
    field.append(["Step", "Instruction"])
    for i in range(22):
        field.append([str(i + 1), f"Instruction detail line {i} for EC/DC/EMS vendor sections."])
    _set_instruction_geometry(field)
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "instruction_pages.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="instr1")
    problems: list[str] = []

    for code in ("EMS 3.0", "EMS 17.0"):
        page = next((p for p in proj["pages"] if (p.get("displaySheetCode") or p.get("sheetCode")) == code), None)
        if page is None:
            problems.append(f"missing page {code}")
            continue
        for lw in page.get("layoutWarnings") or []:
            if "TABLE OVERFLOW" in str(lw):
                problems.append(f"{code}: layout warning {lw}")
        block = next((b for b in (page.get("blocks") or []) if b.get("type") == "excelRange"), None)
        if block is None:
            problems.append(f"{code}: no excelRange block")
            continue
        actual_widths = [float(value) for value in (block.get("colWidths") or [])]
        expected_widths = [excel_column_width_to_pixels(value) for value in COL_WIDTHS]
        if len(actual_widths) != len(expected_widths) or any(
            abs(actual - expected) > 0.01
            for actual, expected in zip(actual_widths, expected_widths)
        ):
            problems.append(
                f"{code}: source width map changed {actual_widths} != {expected_widths}"
            )
        if len(actual_widths) == 2:
            actual_ratio = actual_widths[1] / actual_widths[0]
            expected_ratio = expected_widths[1] / expected_widths[0]
            if abs(actual_ratio - expected_ratio) > 0.001:
                problems.append(
                    f"{code}: width ratio changed {actual_ratio:.6f} != {expected_ratio:.6f}"
                )
        min_scale = float(block.get("minScale") or 0.73)
        if min_scale < 0.78:
            problems.append(f"{code}: minScale {min_scale} below narrative floor")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — instruction pages no tiny overflow")


if __name__ == "__main__":
    main()
