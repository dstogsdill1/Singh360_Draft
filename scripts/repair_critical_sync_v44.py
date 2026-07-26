# S360 CRITICAL SYNC V44
"""One-time controlled repair for the live Mi Tienda workbook/project pair."""
from __future__ import annotations

import argparse
from collections import defaultdict
from copy import deepcopy
from datetime import datetime, timezone
from hashlib import sha256
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from core.drawing_page_mirror import (
    DRAWING_MANIFEST_SHEET,
    is_generated_mirror_sheet,
)
from core.page_identity import is_sheet_index_page
from core.project_model import ensure_project_shape, recalc_page_numbers
from core.project_store import ProjectStore
from core.sheet_index_sync import sync_project_sheet_index
from core.workbook_status_sync import (
    file_hash,
    project_hash,
    sync_project_from_workbook,
    sync_project_to_workbook,
)


MANUAL_FIELDS = (
    "canvasObjects",
    "assets",
    "underlay",
    "underlays",
    "background",
    "overlays",
    "annotations",
    "pastedImages",
    "imageCrop",
    "crop",
    "crops",
    "masks",
    "highlightedCells",
    "manualObjects",
    "lockedObjects",
    "connectors",
)


def utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def workbook_tab_key(value: Any) -> str:
    """Excel worksheet identity for verification.

    Excel preserves leading/trailing and repeated internal spaces in worksheet
    titles. Singh360 display text intentionally collapses those spaces. Resolve
    both forms to one case-insensitive key, but reject ambiguous workbooks where
    two physical sheets collapse to the same key.
    """
    return clean(value).casefold()


def workbook_tab_lookup(wb: Any) -> dict[str, str]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for title in wb.sheetnames:
        grouped[workbook_tab_key(title)].append(str(title))

    ambiguous = {
        key: titles
        for key, titles in grouped.items()
        if key and len(titles) > 1
    }
    if ambiguous:
        details = "; ".join(
            f"{key}: {titles}"
            for key, titles in sorted(ambiguous.items())
        )
        raise RuntimeError(
            "Workbook contains worksheet titles that are ambiguous after "
            "whitespace normalization: " + details
        )

    return {
        key: titles[0]
        for key, titles in grouped.items()
        if key and titles
    }


def resolve_workbook_tab(value: Any, lookup: dict[str, str]) -> str | None:
    return lookup.get(workbook_tab_key(value))


def code_key(page: dict[str, Any]) -> str:
    return re.sub(
        r"\s+",
        "",
        clean(page.get("displaySheetCode") or page.get("sheetCode")).casefold(),
    )


def page_order(page: dict[str, Any], fallback: int = 10**9) -> int:
    try:
        return int(float(page.get("order") or fallback))
    except Exception:
        return fallback


def is_continuation(page: dict[str, Any]) -> bool:
    return bool(
        page.get("generatedContinuation")
        or page.get("indexContinuation")
        or page.get("generatedIndexContinuation")
        or page.get("continuationOf")
    )


def continuation_index(page: dict[str, Any]) -> int:
    try:
        value = int(page.get("continuationIndex") or 0)
        if value > 0:
            return value
    except Exception:
        pass

    code = clean(page.get("displaySheetCode") or page.get("sheetCode"))
    match = re.search(r"([a-z]+)$", code, re.IGNORECASE)
    if not match:
        return 10**6
    total = 0
    for char in match.group(1).lower():
        total = total * 26 + ord(char) - ord("a") + 1
    return total


def group_pages(pages: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered = [
        deepcopy(page)
        for page in sorted(
            (page for page in pages if isinstance(page, dict)),
            key=lambda item: page_order(item),
        )
    ]

    base_by_root: dict[str, dict[str, Any]] = {}
    for page in ordered:
        if is_continuation(page):
            continue
        page_id = clean(page.get("id"))
        group_id = clean(page.get("pageGroupId") or page_id)
        if page_id:
            base_by_root[page_id] = page
        if group_id:
            base_by_root[group_id] = page

    continuations: dict[str, list[dict[str, Any]]] = defaultdict(list)
    loose: list[dict[str, Any]] = []
    for page in ordered:
        if not is_continuation(page):
            continue
        root = clean(page.get("continuationOf") or page.get("pageGroupId"))
        base = base_by_root.get(root)
        if base is None:
            loose.append(page)
            continue
        continuations[clean(base.get("id"))].append(page)

    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in ordered:
        if is_continuation(page):
            continue
        page_id = clean(page.get("id"))
        if page_id in seen:
            continue
        seen.add(page_id)
        result.append(page)
        children = sorted(
            continuations.get(page_id, []),
            key=lambda item: (
                continuation_index(item),
                code_key(item),
                page_order(item),
                clean(item.get("id")),
            ),
        )
        for index, child in enumerate(children, start=1):
            child["continuationOf"] = page_id
            child["pageGroupId"] = clean(page.get("pageGroupId") or page_id)
            child["continuationIndex"] = index
            child["generatedContinuation"] = True
            result.append(child)

    for page in loose:
        page_id = clean(page.get("id"))
        if page_id not in seen:
            result.append(page)
            seen.add(page_id)

    for index, page in enumerate(result, start=1):
        page["order"] = index
    return result


def page_groups(pages: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    grouped = group_pages(pages)
    output: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_root = ""
    for page in grouped:
        root = clean(page.get("continuationOf") or page.get("id"))
        if not current or root == current_root:
            current.append(page)
            current_root = root
            continue
        output.append(current)
        current = [page]
        current_root = root
    if current:
        output.append(current)
    return output


def move_group_after(
    groups: list[list[dict[str, Any]]],
    moving_code: str,
    anchor_code: str,
) -> list[list[dict[str, Any]]]:
    moving_key = re.sub(r"\s+", "", moving_code.casefold())
    anchor_key = re.sub(r"\s+", "", anchor_code.casefold())
    moving_index = next(
        (index for index, group in enumerate(groups) if group and code_key(group[0]) == moving_key),
        None,
    )
    if moving_index is None:
        return groups
    moving = groups.pop(moving_index)
    anchor_index = next(
        (index for index, group in enumerate(groups) if group and code_key(group[0]) == anchor_key),
        None,
    )
    if anchor_index is None:
        groups.insert(moving_index, moving)
        return groups
    groups.insert(anchor_index + 1, moving)
    return groups


def move_group_before_prefix(
    groups: list[list[dict[str, Any]]],
    moving_code: str,
    prefix: str,
) -> list[list[dict[str, Any]]]:
    moving_key = re.sub(r"\s+", "", moving_code.casefold())
    prefix_key = re.sub(r"\s+", "", prefix.casefold())
    moving_index = next(
        (index for index, group in enumerate(groups) if group and code_key(group[0]) == moving_key),
        None,
    )
    if moving_index is None:
        return groups
    moving = groups.pop(moving_index)
    target_index = next(
        (
            index
            for index, group in enumerate(groups)
            if group and code_key(group[0]).startswith(prefix_key)
        ),
        None,
    )
    if target_index is None:
        groups.append(moving)
    else:
        groups.insert(target_index, moving)
    return groups


def apply_targeted_order(project: dict[str, Any]) -> dict[str, Any]:
    groups = page_groups(project.get("pages") or [])
    groups = move_group_after(groups, "EMS 3.1", "EMS 3.0")
    groups = move_group_before_prefix(groups, "EMS 11.0", "EMS 12.")

    pages = [page for group in groups for page in group]
    for page in pages:
        if (
            code_key(page) == "ems89.0"
            and "blank" in clean(page.get("sheetTitle")).casefold()
            and "template" in clean(page.get("sheetTitle")).casefold()
        ):
            page["include"] = False
            page["notes"] = clean(page.get("notes")) or (
                "Blank / page-end template retained for editing but excluded from publication."
            )

    for index, page in enumerate(pages, start=1):
        page["order"] = index
    project["pages"] = pages
    return project


def manual_payload_count(project: dict[str, Any]) -> int:
    count = 0
    all_pages: list[dict[str, Any]] = []
    all_pages.extend(page for page in project.get("pages") or [] if isinstance(page, dict))
    all_pages.extend(page for page in project.get("archivedPages") or [] if isinstance(page, dict))

    for page in all_pages:
        for field in MANUAL_FIELDS:
            value = page.get(field)
            if value in (None, "", [], {}):
                continue
            if isinstance(value, (list, tuple, set, dict)):
                count += len(value)
            else:
                count += 1
    return count


def block_fingerprint(page: dict[str, Any]) -> tuple[str, str] | None:
    blocks = [block for block in page.get("blocks") or [] if isinstance(block, dict)]
    excel = next((block for block in blocks if block.get("type") == "excelRange"), None)
    if excel is not None:
        # S360 CRITICAL SYNC V44F — generated Sheet Index pages intentionally
        # inherit the same srcRows metadata when the base block is cloned, but
        # split_sheet_index_pages replaces each page's grid with a different
        # chunk. Fingerprinting srcRows alone falsely labels every TOC
        # continuation as a duplicate. Use both source-row identity and rendered
        # grid content so true repeated slices still fail while distinct index
        # chunks pass.
        src_rows = excel.get("srcRows")
        grid = excel.get("grid")
        if (
            isinstance(src_rows, list)
            and src_rows
        ) or (
            isinstance(grid, list)
            and grid
        ):
            payload = json.dumps(
                {
                    "srcRows": src_rows if isinstance(src_rows, list) else [],
                    "grid": grid if isinstance(grid, list) else [],
                },
                sort_keys=True,
                ensure_ascii=False,
                separators=(",", ":"),
                default=str,
            )
            return "excelRange", sha256(payload.encode("utf-8")).hexdigest()
    if not blocks:
        return None
    payload = json.dumps(blocks, sort_keys=True, ensure_ascii=False, default=str)
    return "blocks", sha256(payload.encode("utf-8")).hexdigest()


def verify_continuations(project: dict[str, Any]) -> dict[str, Any]:
    pages = project.get("pages") or []
    errors: list[str] = []
    duplicates: list[str] = []
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    base_position: dict[str, int] = {}

    for position, page in enumerate(pages):
        if not isinstance(page, dict):
            continue
        page_id = clean(page.get("id"))
        if not is_continuation(page):
            base_position[page_id] = position
        root = clean(page.get("continuationOf") or page_id)
        groups[root].append(page)

    for root, group in groups.items():
        if len(group) <= 1:
            continue
        positions = [pages.index(page) for page in group]
        if positions != list(range(min(positions), min(positions) + len(positions))):
            errors.append(f"Continuation group is not contiguous: {root}")
        children = [page for page in group if is_continuation(page)]
        indices = [continuation_index(page) for page in children]
        if indices != sorted(indices):
            errors.append(f"Continuation indices are out of order: {root}")

        seen: dict[tuple[str, str], str] = {}
        base_fp = block_fingerprint(group[0])
        if base_fp is not None:
            seen[base_fp] = clean(group[0].get("id"))
        for page in children:
            fingerprint = block_fingerprint(page)
            if fingerprint is None:
                continue
            prior = seen.get(fingerprint)
            if prior and manual_payload_count({"pages": [page]}) == 0:
                duplicates.append(
                    f"{clean(page.get('displaySheetCode') or page.get('sheetCode'))} "
                    f"duplicates {prior}"
                )
            else:
                seen[fingerprint] = clean(page.get("id"))

    if errors or duplicates:
        raise RuntimeError(
            "Continuation verification failed. "
            + " | ".join([*errors, *duplicates][:30])
        )
    return {
        "groupCount": len(groups),
        "multiPageGroupCount": sum(1 for group in groups.values() if len(group) > 1),
        "duplicateContinuationCount": 0,
    }


def index_headers(ws: Any) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(int(ws.max_row or 1), 75) + 1):
        headers: dict[str, int] = {}
        for column in range(1, max(int(ws.max_column or 1), 24) + 1):
            label = clean(ws.cell(row_number, column).value).casefold()
            if label:
                headers[label] = column
        if {"include", "sheet tab", "page title"}.issubset(headers):
            return row_number, headers
    raise RuntimeError("00_INDEX headers were not found.")


def verify_workbook(
    workbook_path: Path,
    project: dict[str, Any],
    repo: Path,
) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        required = {"00_PROJECT_META", "00_INDEX", DRAWING_MANIFEST_SHEET}
        missing = sorted(required - set(wb.sheetnames))
        if missing:
            raise RuntimeError("Workbook mirror is missing control sheets: " + ", ".join(missing))

        tab_lookup = workbook_tab_lookup(wb)
        manifest = wb[DRAWING_MANIFEST_SHEET]
        manifest_headers = {
            clean(manifest.cell(4, column).value): column
            for column in range(1, int(manifest.max_column or 1) + 1)
            if clean(manifest.cell(4, column).value)
        }
        excel_tab_col = manifest_headers.get("Excel Tab")
        page_id_col = manifest_headers.get("Page ID")
        code_col = manifest_headers.get("Sheet Code")
        if not excel_tab_col or not page_id_col or not code_col:
            raise RuntimeError("00_DRAWING_PAGES is missing required columns.")

        manifest_rows: list[dict[str, str]] = []
        for row_number in range(5, int(manifest.max_row or 4) + 1):
            page_id = clean(manifest.cell(row_number, page_id_col).value)
            if not page_id:
                continue
            manifest_rows.append(
                {
                    "pageId": page_id,
                    "sheetCode": clean(manifest.cell(row_number, code_col).value),
                    "excelTab": str(
                        manifest.cell(row_number, excel_tab_col).value or ""
                    ),
                }
            )

        pages = [page for page in project.get("pages") or [] if isinstance(page, dict)]
        pages.sort(key=lambda page: page_order(page))
        if len(manifest_rows) != len(pages):
            raise RuntimeError(
                f"00_DRAWING_PAGES has {len(manifest_rows)} rows; project has {len(pages)} pages."
            )

        resolved_manifest_tabs: list[str] = []
        missing_tabs: list[str] = []
        mismatched_tabs: list[str] = []
        for row in manifest_rows:
            actual_tab = resolve_workbook_tab(row["excelTab"], tab_lookup)
            if actual_tab is None:
                missing_tabs.append(clean(row["excelTab"]))
                continue
            resolved_manifest_tabs.append(actual_tab)
            if row["excelTab"] != actual_tab:
                mismatched_tabs.append(
                    f"{clean(row['excelTab'])} -> {actual_tab}"
                )

        if missing_tabs:
            raise RuntimeError(
                "Drawing-page mirror tabs are missing: "
                + ", ".join(missing_tabs[:20])
            )
        if mismatched_tabs:
            raise RuntimeError(
                "00_DRAWING_PAGES Excel Tab values do not exactly match the "
                "physical worksheet titles: "
                + "; ".join(mismatched_tabs[:20])
            )

        relative = [
            wb.sheetnames.index(actual_tab)
            for actual_tab in resolved_manifest_tabs
        ]
        if relative != sorted(relative):
            raise RuntimeError("Workbook drawing-page tabs are not in app page order.")

        index_ws = wb["00_INDEX"]
        header_row, headers = index_headers(index_ws)
        tab_col = headers["sheet tab"]
        page_id_index_col = headers.get("page id")
        index_tabs: list[str] = []
        index_page_ids: set[str] = set()
        blank_run = 0
        for row_number in range(
            header_row + 1,
            int(index_ws.max_row or header_row) + 1,
        ):
            raw_tab = str(index_ws.cell(row_number, tab_col).value or "")
            tab = clean(raw_tab)
            page_id = (
                clean(index_ws.cell(row_number, page_id_index_col).value)
                if page_id_index_col
                else ""
            )
            if not tab and not page_id:
                blank_run += 1
                if blank_run >= 50:
                    break
                continue
            blank_run = 0
            if tab:
                actual_tab = resolve_workbook_tab(raw_tab, tab_lookup)
                if actual_tab is None:
                    raise RuntimeError(
                        "00_INDEX points to a missing worksheet tab: " + tab
                    )
                if raw_tab != actual_tab:
                    raise RuntimeError(
                        "00_INDEX Sheet Tab does not exactly match the physical "
                        f"worksheet title: {tab} -> {actual_tab}"
                    )
                index_tabs.append(actual_tab)
            if page_id:
                index_page_ids.add(page_id)

        generated_ids = {
            clean(page.get("id"))
            for page in pages
            if is_continuation(page)
        }
        bad_generated_rows = sorted(generated_ids & index_page_ids)
        if bad_generated_rows:
            raise RuntimeError(
                "Generated continuations were incorrectly added as 00_INDEX base rows: "
                + ", ".join(bad_generated_rows[:20])
            )

        mirror_tabs = [
            ws.title for ws in wb.worksheets
            if is_generated_mirror_sheet(ws)
        ]
        if not mirror_tabs:
            raise RuntimeError("No tagged generated drawing-page mirror tabs were created.")

        unmatched = "EMS 13.1a IDF"
        unmatched_actual = resolve_workbook_tab(unmatched, tab_lookup)
        if unmatched_actual is not None and unmatched_actual not in index_tabs:
            raise RuntimeError(
                "Unmatched physical worksheet was not preserved in 00_INDEX: "
                + unmatched
            )

        return {
            "physicalWorksheetCount": len(wb.sheetnames),
            "drawingPageRows": len(manifest_rows),
            "baseIndexRows": len(index_tabs),
            "generatedMirrorTabs": len(mirror_tabs),
            "unmatchedSourcePreserved": (
                unmatched_actual is None or unmatched_actual in index_tabs
            ),
        }
    finally:
        wb.close()


def verify_reimport_ignores_mirrors(
    workbook_path: Path,
    repo: Path,
) -> dict[str, Any]:
    from core.workbook_importer import import_workbook

    imported = import_workbook(workbook_path, project_id="v44_reimport_check")
    worksheet_names = {
        clean(worksheet.get("name") or worksheet.get("sourceSheet"))
        for worksheet in imported.get("worksheets") or []
        if isinstance(worksheet, dict)
    }
    if DRAWING_MANIFEST_SHEET in worksheet_names:
        raise RuntimeError("00_DRAWING_PAGES was reimported as a workbook draft.")
    # The codeName tag is authoritative; verify by inspecting source workbook.
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        mirror_names = {
            ws.title for ws in wb.worksheets
            if is_generated_mirror_sheet(ws)
        }
    finally:
        wb.close()
    leaked = sorted(mirror_names & worksheet_names)
    if leaked:
        raise RuntimeError(
            "Generated mirror tabs leaked back into Workbook Drafts: "
            + ", ".join(leaked[:20])
        )
    return {
        "importedWorksheetCount": len(worksheet_names),
        "ignoredGeneratedMirrorTabs": len(mirror_names),
        "leakedMirrorTabs": 0,
    }


def repair(repo: Path, project_id: str, report_path: Path) -> dict[str, Any]:
    docs = repo / ".docs"
    store = ProjectStore(docs)
    original = store.load(project_id)
    if original is None:
        raise RuntimeError(f"Project not found: {project_id}")

    workbook_text = clean((original.get("workbookSync") or {}).get("workbook"))
    if not workbook_text:
        raise RuntimeError("The project does not have a linked workbook.")
    workbook_path = Path(workbook_text)
    if not workbook_path.is_file():
        raise RuntimeError(f"Linked workbook not found: {workbook_path}")

    manual_before = manual_payload_count(original)
    workbook_before = file_hash(workbook_path)
    project_before = project_hash(original)

    refreshed = sync_project_from_workbook(project_id, original, store)
    refreshed = apply_targeted_order(refreshed)
    refreshed = ensure_project_shape(refreshed)
    refreshed = sync_project_sheet_index(refreshed)
    refreshed["pages"] = group_pages(refreshed.get("pages") or [])
    recalc_page_numbers(refreshed)
    store.save(project_id, refreshed)

    continuation_result = verify_continuations(refreshed)

    synced = sync_project_to_workbook(project_id, refreshed, store)
    synced["pages"] = group_pages(synced.get("pages") or [])
    synced = sync_project_sheet_index(ensure_project_shape(synced))
    recalc_page_numbers(synced)

    sync = dict(synced.get("workbookSync") or {})
    sync.update(
        {
            "status": "in_sync",
            "warning": "",
            "pendingReason": "",
            "lastSyncUtc": utcnow(),
            "workbookHash": file_hash(workbook_path),
            "appHash": project_hash(synced),
            "authority": "app",
            "lastAuthorityAction": "critical_sync_v44_repair",
            "syncEngineVersion": "V44",
        }
    )
    synced["workbookSync"] = sync
    store.save(project_id, synced)

    final = store.load(project_id)
    if final is None:
        raise RuntimeError("The repaired project could not be reloaded.")

    manual_after = manual_payload_count(final)
    if manual_after < manual_before:
        raise RuntimeError(
            f"Manual-work payload count decreased from {manual_before} to {manual_after}."
        )

    final_codes = [code_key(page) for page in final.get("pages") or [] if not is_continuation(page)]
    if "ems3.1" in final_codes and "ems3.0" in final_codes:
        if final_codes.index("ems3.1") != final_codes.index("ems3.0") + 1:
            raise RuntimeError("EMS 3.1 was not placed immediately after EMS 3.0.")
    if "ems11.0" in final_codes:
        first_12 = next((i for i, value in enumerate(final_codes) if value.startswith("ems12.")), None)
        if first_12 is not None and final_codes.index("ems11.0") != first_12 - 1:
            raise RuntimeError("EMS 11.0 was not placed immediately before the 12-series.")

    continuation_result = verify_continuations(final)
    workbook_result = verify_workbook(workbook_path, final, repo)
    reimport_result = verify_reimport_ignores_mirrors(workbook_path, repo)

    result = {
        "ok": True,
        "projectId": project_id,
        "workbook": str(workbook_path),
        "projectHashBefore": project_before,
        "projectHashAfter": project_hash(final),
        "workbookHashBefore": workbook_before,
        "workbookHashAfter": file_hash(workbook_path),
        "manualPayloadCountBefore": manual_before,
        "manualPayloadCountAfter": manual_after,
        "pageCount": len(final.get("pages") or []),
        "includedCount": sum(
            1 for page in final.get("pages") or []
            if isinstance(page, dict) and page.get("include", True)
        ),
        "continuations": continuation_result,
        "workbookMirror": workbook_result,
        "reimportCheck": reimport_result,
        "completedUtc": utcnow(),
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(result, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--report", required=True)
    args = parser.parse_args()

    result = repair(
        Path(args.repo).resolve(),
        args.project_id,
        Path(args.report).resolve(),
    )
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
