"""Sanitized multi-schedule shapes inspired only by the named reference layouts."""
from __future__ import annotations
import sys
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from core.worksheet_export import export_excel_layout_page_xlsx
from tests.test_excel_layout_export import layout_page

page = layout_page()
base = page["excelLayout"]["tables"][0]
page["excelLayout"]["tables"] = []
for index, widths in enumerate(([2, 5, 2], [5, 2, 3], [1, 2, 6], [4, 4, 1], [2, 2, 2])):
    table = {**base, "id": f"schedule-{index}", "title": f"NEUTRAL TEST SCHEDULE {index + 1}",
             "y": 60 + index * 180, "columnWidths": list(widths)}
    page["excelLayout"]["tables"].append(table)
wb = load_workbook(BytesIO(export_excel_layout_page_xlsx(page)))
ws = wb.active
titles = {cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("NEUTRAL")}
if len(titles) != 5 or len(ws.merged_cells.ranges) < 15:
    raise SystemExit("independent sanitized schedule shape export failed")
print("PASS: five independently sized editable schedules exported with real merged titles")
