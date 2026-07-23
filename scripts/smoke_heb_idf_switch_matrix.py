"""Smoke: H-E-B IDF switch-matrix import uses the exact seven source columns."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.heb_idf_switch_matrix import (
    HEB_HEADERS,
    PATCH_MARKER,
    build_heb_idf_network_block,
    find_heb_idf_header_row,
    next_available_continuation_code,
    replace_heb_idf_pages,
    sheet_code_key,
)

FORBIDDEN = {"Network", "From", "To", "Cable", "Notes", "Device / Drop", "Port"}


def _payload() -> dict:
    grid = [
        ["IDF #2 TABLE (SWITCH 3 & 4)", "", "Controller ID Legend", "IP Network"],
        ["SC16", "380", "001-050: Rack or condenser", "Location"],
        ["", "", "", ""],
        ["Label #", "Description", "Controller ID", "IP Address", "IDF#", "Switch#", "Port#"],
    ]
    for switch in (3, 4, 5):
        for port in range(1, 49):
            if port <= 44:
                grid.append([
                    f"L{switch:01d}{port:02d}",
                    f"Switch {switch} device {port}",
                    200 + switch * 50 + port,
                    "-",
                    2,
                    switch,
                    port,
                ])
            else:
                grid.append(["-", "Spare - Fiber Channel", "-", "-", 2, switch, port])
    return {
        "id": "ws_heb",
        "name": "EMS 13.2 IDF #2",
        "sourceSheet": "EMS 13.2 IDF #2",
        "sourceRange": "A1:G148",
        "grid": grid,
    }


def _synthetic(problems: list[str]) -> None:
    ws = _payload()
    header = find_heb_idf_header_row(ws["grid"])
    if header != 3:
        problems.append(f"H-E-B real header expected row index 3, got {header}")
        return

    first = build_heb_idf_network_block(ws, "b0", pair_index=0, header_row=header)
    second = build_heb_idf_network_block(ws, "b1", pair_index=1, header_row=header)
    if not first or not second:
        problems.append("H-E-B block builder returned no block")
        return

    if first.get("headers") != HEB_HEADERS:
        problems.append(f"wrong H-E-B headers: {first.get('headers')}")
    extra = FORBIDDEN.intersection(set(first.get("headers") or []))
    if extra:
        problems.append(f"generic network columns leaked into H-E-B table: {sorted(extra)}")
    if first.get("colWidths") != [98, 354, 98, 86, 38, 62, 46]:
        problems.append(f"Kyle column widths not applied: {first.get('colWidths')}")
    if first.get("contentWidth") != 1582:
        problems.append(f"paired page width should be 1582, got {first.get('contentWidth')}")
    if second.get("contentWidth") != 782:
        problems.append(f"single page width should remain 782, got {second.get('contentWidth')}")
    if first.get("layoutMode") != "two_up":
        problems.append(f"first H-E-B page should be two_up, got {first.get('layoutMode')}")
    if len(first.get("leftRows") or []) != 48 or len(first.get("rightRows") or []) != 48:
        problems.append("switch 3/4 tables were not paired 48/48")
    if first.get("sectionTitle") != "IDF #2 TABLE (SWITCH 3 & 4)":
        problems.append(f"wrong title: {first.get('sectionTitle')!r}")
    if second.get("layoutMode") != "single" or len(second.get("rows") or []) != 48:
        problems.append("third switch did not create a single-table continuation")
    if second.get("sectionTitle") != "IDF #2 TABLE (SWITCH 5)":
        problems.append(f"wrong continuation title: {second.get('sectionTitle')!r}")
    sample = (first.get("leftRows") or [[]])[0]
    if sample[2] != "351" or sample[4:] != ["2", "3", "1"]:
        problems.append(f"integer fields gained decimals or shifted: {sample}")

    used = {sheet_code_key("EMS 13.2a")}
    code = next_available_continuation_code("EMS 13.2", 1, used)
    if code != "EMS 13.2b":
        problems.append(f"reserved continuation code was not skipped: {code}")

    pages = [{
        "id": "page_idf2",
        "order": 1,
        "include": True,
        "sheetCode": "EMS 13.2",
        "displaySheetCode": "EMS 13.2",
        "sheetTitle": "IDF #2 Port / Network Table",
        "sheetTab": ws["name"],
        "pageType": "data-grid",
        "pageFamily": "idfTable",
        "layoutProfile": "network_48_port",
        "linkedWorksheetId": ws["id"],
        "pageGroupId": "page_idf2",
        "blocks": [],
        "canvasObjects": [],
    }]
    index = [
        {"sheetCodeRaw": "EMS 13.2", "sheetTab": ws["name"]},
        {"sheetCodeRaw": "EMS 13.2a", "sheetTab": "EMS 13.2a IDF #2 Layout"},
    ]
    replaced = replace_heb_idf_pages(pages, [ws], index)
    if len(replaced) != 2:
        problems.append(f"expected base + one continuation, got {len(replaced)}")
    else:
        if replaced[1].get("displaySheetCode") != "EMS 13.2b":
            problems.append(f"generated continuation collided with reserved source code: {replaced[1].get('displaySheetCode')}")
        if replaced[1].get("generatedBy") != PATCH_MARKER:
            problems.append("generated H-E-B continuation is not tagged")


def _actual_workbook(path: Path, problems: list[str]) -> None:
    from core.workbook_importer import import_workbook

    project = import_workbook(path, project_id="heb829smoke")
    expected_tabs = ["EMS 13.1 IDF #1", "EMS 13.2 IDF #2", "EMS 13.3 IDF #3"]
    for tab in expected_tabs:
        page_set = [p for p in project.get("pages", []) if p.get("sheetTab") == tab]
        if not page_set:
            problems.append(f"actual workbook missing imported page for {tab}")
            continue
        base = next((p for p in page_set if not p.get("generatedContinuation")), page_set[0])
        block = (base.get("blocks") or [{}])[0]
        if block.get("headers") != HEB_HEADERS:
            problems.append(f"{tab} imported wrong headers: {block.get('headers')}")
        if FORBIDDEN.intersection(set(block.get("headers") or [])):
            problems.append(f"{tab} still contains generic network columns")
        if block.get("tableProfile") != "heb_idf_switch_matrix":
            problems.append(f"{tab} missing H-E-B table profile")
        if not ((block.get("leftRows") and block.get("rightRows")) or block.get("rows")):
            problems.append(f"{tab} has no switch-table rows")


def main() -> None:
    problems: list[str] = []
    _synthetic(problems)
    if len(sys.argv) > 1:
        workbook = Path(sys.argv[1]).expanduser().resolve()
        if workbook.is_file():
            _actual_workbook(workbook, problems)
        else:
            problems.append(f"workbook path not found: {workbook}")

    if problems:
        print("FAIL — H-E-B IDF switch-matrix smoke")
        for problem in problems:
            print(" -", problem)
        raise SystemExit(1)
    print("OK — H-E-B IDF switch-matrix import passed")


if __name__ == "__main__":
    main()
