"""Smoke: 00_PROJECT_META is never an output page, even if a hand-edited
index accidentally lists it with Include=YES (FINAL RELEASE CLEANUP
4H+SA38, Phase A rule 7 — defensive/belt-and-suspenders check).
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "BOM", "Bill of Materials", "Front Matter", "table", ""])
    # A hand-edit mistake: someone lists the metadata sheet itself with
    # Include=YES. It must still never render as a drawing page.
    idx.append(["YES", 2, "EMS 0.1", "00_PROJECT_META", "Project Metadata", "Front Matter", "text", ""])

    bom = wb.create_sheet("BOM")
    bom.append(["Item", "Qty"])
    bom.append(["Controller", "4"])

    meta = wb.create_sheet("00_PROJECT_META")
    meta.append(["Project Metadata", None])
    meta.append(["Project Name", "Defensive Test Project"])
    meta.append(["Location", "Nowhere, TX"])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "meta_defensive.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="metadefensive")
    pages = proj["pages"]
    page_tabs = {p["sheetTab"] for p in pages}
    ws_names = {w["name"] for w in proj["worksheets"]}
    problems: list[str] = []

    if "00_PROJECT_META" not in ws_names:
        problems.append("00_PROJECT_META missing from source worksheets (should still be preserved as a source tab)")
    if "00_PROJECT_META" in page_tabs:
        problems.append("00_PROJECT_META rendered as an output page despite being a metadata/control sheet")
    if "BOM" not in page_tabs:
        problems.append("BOM (real content) missing from output pages")
    if proj["metadata"].get("projectName") != "Defensive Test Project":
        problems.append("00_PROJECT_META values were not applied to project.metadata even though the sheet itself was correctly excluded")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — 00_PROJECT_META never renders as an output page (even if Include=YES)")


if __name__ == "__main__":
    main()
