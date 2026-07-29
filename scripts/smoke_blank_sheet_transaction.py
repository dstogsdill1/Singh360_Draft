"""Browser proof for atomic blank-sheet creation and copied-workbook sync."""
from __future__ import annotations

from hashlib import sha256
import json
import os
from pathlib import Path
import shutil
import socket
import subprocess
import sys
import tempfile
import time
from typing import Any
from urllib.request import urlopen

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from core.project_store import ProjectStore
from core.project_workspace import WorkbookDocumentStore, drawing_workspace_sequence
from core.workbook_status_sync import sync_project_to_workbook
from tests.test_sync_order_reconciliation import (
    document_for,
    fixture_project,
    write_fixture_workbook,
)


def free_port() -> int:
    with socket.socket() as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def wait_health(port: int) -> None:
    for _ in range(180):
        try:
            if urlopen(f"http://127.0.0.1:{port}/api/health", timeout=1).status == 200:
                return
        except OSError:
            time.sleep(0.2)
    raise RuntimeError("disposable blank-sheet server did not become healthy")


def file_hash(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def index_record(workbook_path: Path, page_id: str) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        index = workbook["00_INDEX"]
        headers: dict[str, int] = {}
        header_row = 0
        for row_number, row in enumerate(index.iter_rows(min_row=1, max_row=30), start=1):
            found = {
                str(cell.value or "").strip(): cell.column
                for cell in row
                if str(cell.value or "").strip()
            }
            if {"Sheet Code", "Sheet Tab", "Page Title"}.issubset(found):
                headers = found
                header_row = row_number
                break
        if not header_row:
            raise AssertionError("copied workbook has no usable 00_INDEX header")
        for row in index.iter_rows(min_row=header_row + 1):
            record = {
                label: row[column - 1].value
                for label, column in headers.items()
                if column <= len(row)
            }
            if str(record.get("Page ID") or "") == page_id:
                return record
        raise AssertionError(f"copied workbook has no 00_INDEX row for {page_id}")
    finally:
        workbook.close()


def main() -> int:
    evidence = Path(
        os.environ.get("SINGH360_BROWSER_EVIDENCE_DIR")
        or ROOT / ".tmp" / "blank_sheet_transaction_browser"
    )
    evidence.mkdir(parents=True, exist_ok=True)
    print(f"evidence={evidence}", flush=True)
    with tempfile.TemporaryDirectory(prefix="s360_blank_sheet_browser_") as raw:
        runtime = Path(raw)
        docs = runtime / "runtime-docs"
        source_template = runtime / "generated-source-template.xlsx"
        workbook_copy = runtime / "protected-workbook-copy.xlsx"
        write_fixture_workbook(source_template)
        shutil.copy2(source_template, workbook_copy)
        source_hash_before = file_hash(source_template)
        print("stage=generated_copied_workbook", flush=True)

        project = fixture_project(workbook_copy)
        project_id = "c1a2b3d4e5f60718"
        project["id"] = project_id
        project["workbookSync"]["workbook"] = str(workbook_copy)
        store = ProjectStore(docs)
        store.save(project_id, project)
        project_dir = store.dir_for(project_id, project)
        document_store = WorkbookDocumentStore(project_dir)
        document_store.save(project, 0, document_for(project))
        synced = sync_project_to_workbook(project_id, project, store)
        if not synced.get("workbookSync", {}).get("verified"):
            raise AssertionError("initial copied-workbook baseline was not verified")
        print("stage=initial_sync_verified", flush=True)

        port = free_port()
        server_stdout = (evidence / "server.stdout.log").open("w", encoding="utf-8")
        server_stderr = (evidence / "server.stderr.log").open("w", encoding="utf-8")
        process = subprocess.Popen(
            [sys.executable, str(ROOT / "server.py")],
            cwd=ROOT,
            env={
                **os.environ,
                "SINGH360_DOCS_DIR": str(docs),
                "SINGH360_PORT": str(port),
            },
            stdout=server_stdout,
            stderr=server_stderr,
        )
        console_errors: list[str] = []
        page_errors: list[str] = []
        api_responses: list[dict[str, Any]] = []
        failed_requests: list[dict[str, str]] = []
        blank_page_id = ""
        try:
            wait_health(port)
            print(f"stage=server_healthy port={port}", flush=True)
            with sync_playwright() as playwright:
                edge = Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
                launch = (
                    {"executable_path": str(edge), "headless": True}
                    if edge.is_file()
                    else {"headless": True}
                )
                browser = playwright.chromium.launch(**launch)
                page = browser.new_page(viewport={"width": 1500, "height": 900})
                page.on(
                    "console",
                    lambda message: console_errors.append(message.text)
                    if message.type == "error"
                    else None,
                )
                page.on("pageerror", lambda error: page_errors.append(str(error)))
                page.on(
                    "requestfailed",
                    lambda request: failed_requests.append(
                        {
                            "method": request.method,
                            "url": request.url,
                            "failure": request.failure or "",
                        }
                    ),
                )

                def record_response(response) -> None:
                    if f"/api/projects/{project_id}" not in response.url:
                        return
                    record: dict[str, Any] = {
                        "method": response.request.method,
                        "url": response.url,
                        "status": response.status,
                    }
                    if response.status >= 400:
                        try:
                            record["body"] = response.text()
                        except Exception as error:
                            record["bodyError"] = str(error)
                    api_responses.append(record)

                page.on("response", record_response)
                editor_url = (
                    f"http://127.0.0.1:{port}/app"
                    f"?project={project_id}&mode=editor"
                )
                page.goto(editor_url, wait_until="domcontentloaded", timeout=60_000)
                page.locator(".page-tab").first.wait_for(state="visible", timeout=30_000)
                page.screenshot(path=str(evidence / "01_editor_loaded.png"), full_page=True)
                print("stage=editor_loaded", flush=True)

                initial_count = page.locator(".page-tab").count()
                page.locator(".page-tab").last.click(button="right")
                page.get_by_role("button", name="Add Blank Sheet After").click()
                dialog = page.get_by_role("heading", name="Add Blank Sheet").locator("..").locator("..")
                page.get_by_label("Sheet Title").fill("Disposable Blank Page")
                self_code = page.get_by_label("Sheet Code (optional — set after Renumber)")
                if self_code.input_value() != "":
                    raise AssertionError("optional sheet-code field was not blank")
                page.get_by_label("Page Template").select_option("canvas")
                with page.expect_response(
                    lambda response: response.request.method == "POST"
                    and response.url.endswith(f"/api/projects/{project_id}"),
                    timeout=30_000,
                ) as created_response:
                    page.get_by_role("button", name="Add Sheet").click()
                if created_response.value.status != 200:
                    raise AssertionError(
                        "blank-sheet create failed: "
                        f"{created_response.value.status} {created_response.value.text()}"
                    )
                page.get_by_text(
                    "PROJECT SAVED · WORKBOOK SYNC PENDING",
                    exact=True,
                ).wait_for(timeout=15_000)
                if page.locator(".page-tab").count() != initial_count + 1:
                    raise AssertionError("blank page did not appear in the page strip")
                new_tab = page.locator(".page-tab").filter(has_text="Disposable Blank Page")
                new_tab.wait_for(state="visible", timeout=10_000)
                page.screenshot(path=str(evidence / "02_blank_created_saved.png"), full_page=True)
                print("stage=blank_created_saved", flush=True)

                page.get_by_role("button", name="File", exact=True).click()
                page.get_by_role("button", name="Save Now", exact=True).click()
                page.locator(".page-tab").filter(has_text="Gamma").click()
                new_tab = page.locator(".page-tab").filter(has_text="Disposable Blank Page")
                new_tab.click()
                if "Disposable Blank Page" not in new_tab.inner_text():
                    raise AssertionError("navigation back to the blank page failed")

                page.reload(wait_until="domcontentloaded", timeout=60_000)
                page.locator(".page-tab").filter(
                    has_text="Disposable Blank Page"
                ).wait_for(state="visible", timeout=30_000)
                persisted = store.load(project_id)
                blank = next(
                    item
                    for item in persisted["pages"]
                    if item["sheetTitle"] == "Disposable Blank Page"
                )
                blank_page_id = blank["id"]
                if blank.get("sheetCode") != "":
                    raise AssertionError("blank optional sheet code was not preserved before renumber")
                if not blank.get("sheetTab") or not blank.get("linkedWorksheetId"):
                    raise AssertionError("blank page did not receive a durable worksheet identity")
                print("stage=navigation_reload_verified", flush=True)

                new_tab = page.locator(".page-tab").filter(has_text="Disposable Blank Page")
                target_tab = page.locator(".page-tab").filter(has_text="Gamma")
                with page.expect_response(
                    lambda response: response.request.method == "POST"
                    and response.url.endswith(f"/api/projects/{project_id}"),
                    timeout=30_000,
                ) as reordered_response:
                    new_tab.drag_to(target_tab)
                if reordered_response.value.status != 200:
                    raise AssertionError("page reorder did not save")
                print("stage=reorder_saved", flush=True)

                page.get_by_role("button", name="File", exact=True).click()
                page.get_by_role(
                    "button",
                    name="Renumber Sheet Codes",
                    exact=False,
                ).click()
                renumber = page.get_by_role("heading", name="Renumber Sheet Codes").locator("..").locator("..")
                renumber.locator('input[name="scheme"]').nth(2).check()
                with page.expect_response(
                    lambda response: response.request.method == "POST"
                    and response.url.endswith(f"/api/projects/{project_id}"),
                    timeout=30_000,
                ) as renumber_response:
                    page.get_by_role("button", name="Apply order & codes").click()
                if renumber_response.value.status != 200:
                    raise AssertionError("renumber did not save")
                print("stage=renumber_saved", flush=True)

                with page.expect_response(
                    lambda response: response.request.method == "POST"
                    and response.url.endswith(
                        f"/api/projects/{project_id}/workbook-link/resolve"
                    ),
                    timeout=120_000,
                ) as workbook_response:
                    page.get_by_role(
                        "button",
                        name="SAVE + WRITE EXCEL",
                        exact=True,
                    ).click()
                if workbook_response.value.status != 200:
                    raise AssertionError(
                        "Save + Write Excel failed against copied workbook: "
                        f"{workbook_response.value.status} {workbook_response.value.text()}"
                    )
                print("stage=copied_workbook_sync_response_200", flush=True)
                page.get_by_text(
                    "PROJECT SAVED · WORKBOOK SYNCED",
                    exact=True,
                ).wait_for(timeout=30_000)
                page.screenshot(path=str(evidence / "03_synced_after_reload.png"), full_page=True)

                if page.locator(
                    "[data-nextjs-dialog], .vite-error-overlay, "
                    "#webpack-dev-server-client-overlay"
                ).count():
                    raise AssertionError("browser error overlay was visible")
                if not page.locator("body").inner_text().strip():
                    raise AssertionError("browser rendered a blank page")
                browser.close()
        finally:
            process.terminate()
            try:
                process.wait(timeout=10)
            except subprocess.TimeoutExpired:
                process.kill()
                process.wait(timeout=5)
            server_stdout.close()
            server_stderr.close()

        final_project = store.load(project_id)
        final_blank = next(
            item for item in final_project["pages"] if item["id"] == blank_page_id
        )
        final_document = document_store.load(final_project)
        workspace_blank = next(
            item
            for item in drawing_workspace_sequence(final_document)
            if item["pageId"] == blank_page_id
        )
        workbook_row = index_record(workbook_copy, blank_page_id)
        workbook = load_workbook(workbook_copy, read_only=True)
        try:
            physical_tabs = workbook.sheetnames
        finally:
            workbook.close()
        if final_blank["sheetTab"] not in physical_tabs:
            raise AssertionError("copied workbook is missing the blank page worksheet")
        if workbook_row["Sheet Tab"] != final_blank["sheetTab"]:
            raise AssertionError("copied workbook 00_INDEX tab does not match project")
        if workbook_row["Sheet Code"] != final_blank["sheetCode"]:
            raise AssertionError("copied workbook 00_INDEX code does not match renumbered project")
        if workspace_blank["sheetTab"] != final_blank["sheetTab"]:
            raise AssertionError("Data Workspace tab does not match the project page")
        if file_hash(source_template) != source_hash_before:
            raise AssertionError("generated source template changed; only its copy was authorized")

        unexpected_failures = [
            item
            for item in failed_requests
            if item["failure"] != "net::ERR_ABORTED"
        ]
        if unexpected_failures:
            raise AssertionError(f"unexpected browser request failures: {unexpected_failures}")
        if page_errors:
            raise AssertionError(f"browser page errors: {page_errors}")
        if console_errors:
            raise AssertionError(f"browser console errors: {console_errors}")
        failed_api = [item for item in api_responses if item["status"] >= 400]
        if failed_api:
            raise AssertionError(f"browser API failures: {failed_api}")

        summary = {
            "ok": True,
            "evidence": str(evidence),
            "disposableDocsRemovedOnExit": str(docs),
            "generatedSourceTemplate": str(source_template),
            "copiedWorkbook": str(workbook_copy),
            "sourceTemplateHashPreserved": source_hash_before,
            "blankPageId": blank_page_id,
            "blankPage": {
                "order": final_blank["order"],
                "sheetCode": final_blank["sheetCode"],
                "sheetTab": final_blank["sheetTab"],
                "worksheetId": final_blank["linkedWorksheetId"],
            },
            "workflow": [
                "add blank sheet with empty optional code",
                "Save Now",
                "navigate away and back",
                "reload",
                "reorder",
                "renumber",
                "Save + Write Excel on copied workbook",
            ],
            "apiResponses": api_responses,
            "consoleErrors": 0,
            "pageErrors": 0,
            "unexpectedFailedRequests": 0,
        }
        (evidence / "summary.json").write_text(
            json.dumps(summary, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
