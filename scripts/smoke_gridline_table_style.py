"""Smoke: normalized tables use gridlines and no zebra body rows."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook

BODY_ALT_FILL = "#F4F6F8"


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["3", "ABBREVIATIONS", "Abbreviations", "YES", ""])

    ws = wb.create_sheet("ABBREVIATIONS")
    ws.append(["Abbrev", "Description"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True)
    for r in range(2, 10):
        ws.append([f"A{r}", f"Description row {r}"])
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "gridline_style.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="style1")
    page = next(p for p in proj["pages"] if p["sheetTab"] == "ABBREVIATIONS")
    block = page["blocks"][0]
    problems: list[str] = []

    if block.get("bodyRowFillMode") != "none":
        problems.append(f"bodyRowFillMode={block.get('bodyRowFillMode')}")
    if block.get("gridLines") is not True:
        problems.append("gridLines not enabled")
    styles = block.get("styles") or {}
    if any(st.get("fill") == BODY_ALT_FILL for st in styles.values() if isinstance(st, dict)):
        problems.append("zebra BODY_ALT_FILL was generated")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - gridline/no-zebra table style passed")


if __name__ == "__main__":
    main()
