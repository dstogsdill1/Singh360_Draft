from __future__ import annotations

import argparse
import copy
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.project_data_compiler import apply_compile, manifest_recipes
from core.project_store import ProjectStore
from core.project_template_service import ProjectTemplateService
from core.template_platform import (
    ProfileRegistry, SourceLibrary, TemplateRegistry, WorkbookDocumentStore,
    atomic_json_write, sha256_file, utcnow,
)

DOCS = ROOT / ".docs"
OLD_ID = "95d85da603864a62"
STAGED = DOCS / "migration_staging" / "SA31_Singh360_Template_Platform_Migration_V2.xlsx"


def object_count(page: dict) -> int:
    return len(page.get("canvasObjects") or page.get("objects") or [])


def normalized(value: object) -> str:
    return " ".join(str(value or "").split()).casefold()


def rehome_project_assets(value: object, new_project_id: str) -> object:
    if isinstance(value, str):
        return value.replace(
            f"/api/assets/{OLD_ID}/", f"/api/assets/{new_project_id}/"
        )
    if isinstance(value, list):
        return [rehome_project_assets(item, new_project_id) for item in value]
    if isinstance(value, dict):
        return {
            key: rehome_project_assets(item, new_project_id)
            for key, item in value.items()
        }
    return value


def read_metadata(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["00_PROJECT_META"]
    metadata: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = str(row[0] or "").strip()
        if label and len(row) > 1 and row[1] not in (None, ""):
            metadata[label] = str(row[1])
    wb.close()
    return {
        "projectName": metadata["Project Name"],
        "storeNumber": metadata.get("Store / Site Number", ""),
        "client": metadata.get("Client", ""),
        "location": metadata.get("Location", ""),
        "address": metadata.get("Address", ""),
        "purpose": metadata.get("Purpose", ""),
        "scopeSummary": metadata.get("Scope Summary", ""),
        "drawingPrefix": metadata.get("Drawing Prefix", ""),
        "revision": metadata.get("Revision", ""),
        "drawnBy": metadata.get("Drawn By", ""),
        "checkedBy": metadata.get("Checked By", ""),
    }


def migrate(keep_on_failure: bool = False) -> dict:
    store = ProjectStore(DOCS)
    old = store.load(OLD_ID)
    if not old:
        raise RuntimeError("The audited SA31 project is unavailable.")
    old_dir = store.find_dir(OLD_ID)
    if old_dir is None:
        raise RuntimeError("The audited SA31 package folder is unavailable.")

    profiles = ProfileRegistry(ROOT / "defaults" / "project_templates" / "project_profiles.json")
    templates = TemplateRegistry(DOCS)
    validation = templates.validate(STAGED)
    if not validation["valid"] or validation["sheetCount"] != 38:
        raise RuntimeError(f"Migration workbook validation failed: {validation}")

    service = ProjectTemplateService(store, profiles, templates)
    project = service.create({
        "profileId": "EMS_LIGHTING",
        "templateId": "SINGH360_BASE_V1",
        "metadata": read_metadata(STAGED),
    }, source_workbook=STAGED)
    new_id = project["id"]
    new_dir = store.find_dir(new_id)
    assert new_dir is not None
    try:
        document_store = WorkbookDocumentStore(new_dir)
        document = document_store.load()
        recipes = manifest_recipes(document)
        if len(recipes) != 29:
            raise RuntimeError(f"Expected 29 detailed index rows; found {len(recipes)}.")
        compiled, preview = apply_compile(project, document, profiles.get("EMS_LIGHTING"))

        old_pages = list(old.get("pages") or [])
        used_old: set[str] = set()
        migration_pages: list[dict] = []
        manual_fields = {
            "canvasObjects", "objects", "assets", "notes", "underlay", "underlayUrl",
            "pdfSource", "pdfPage", "pdfCrop", "imageCrop", "cropState", "shadows",
            "borders", "rotation", "customLayout", "sourceRange", "linkedWorksheetId",
        }
        destination_titles = {
            normalized(item.get("sheetTitle")) for item in compiled["pages"]
        }
        for page in compiled["pages"]:
            available = [
                item for item in old_pages
                if str(item.get("id")) not in used_old
            ]
            code_matches = [
                item for item in available
                if normalized(item.get("sheetCode")) == normalized(page.get("sheetCode"))
                and (
                    normalized(item.get("sheetTitle")) == normalized(page.get("sheetTitle"))
                    or normalized(item.get("sheetTitle")) not in destination_titles
                )
            ]
            title_matches = [
                item for item in available
                if normalized(item.get("sheetTitle")) == normalized(page.get("sheetTitle"))
            ]
            candidates = title_matches or code_matches
            source = max(candidates, key=object_count, default=None)
            if source:
                used_old.add(str(source.get("id")))
                for key in manual_fields:
                    if key in source:
                        page[key] = rehome_project_assets(
                            copy.deepcopy(source[key]), new_id
                        )
                migration_pages.append({
                    "oldPageId": source.get("id"), "newPageId": page["id"],
                    "sheetCode": page.get("sheetCode"), "manualObjects": object_count(source),
                })
            else:
                migration_pages.append({
                    "oldPageId": None, "newPageId": page["id"],
                    "sheetCode": page.get("sheetCode"), "manualObjects": 0,
                })

        for source in old_pages:
            if object_count(source) and str(source.get("id")) not in used_old:
                preserved = copy.deepcopy(source)
                preserved["id"] = f"legacy-{source.get('id')}"
                preserved["include"] = False
                preserved["order"] = len(compiled["pages"])
                preserved["migrationLegacyManual"] = True
                compiled["pages"].append(preserved)
                migration_pages.append({
                    "oldPageId": source.get("id"), "newPageId": preserved["id"],
                    "sheetCode": preserved.get("sheetCode"), "manualObjects": object_count(source),
                })

        source_assets = old_dir / "assets"
        if source_assets.is_dir():
            shutil.copytree(source_assets, new_dir / "assets", dirs_exist_ok=True)
        library = SourceLibrary(new_dir)
        original_workbook = old_dir / "sources" / "workbook" / "SA31_EMS_Lighting_Workbook_V1.xlsx"
        original_pdfs = sorted((old_dir / "sources" / "pdf").glob("*.pdf"))
        with original_workbook.open("rb") as handle:
            library.upload(handle, original_workbook.name, {
                "virtualPath": "Converted Schedules/Original Excel",
                "originalLocation": str(old.get("workbookSync", {}).get("workbook") or ""),
            })
        for pdf in original_pdfs:
            with pdf.open("rb") as handle:
                library.upload(handle, pdf.name, {"virtualPath": "Drawings/Construction Set"})

        compiled["migration"] = {
            "schemaVersion": 1,
            "oldProjectId": OLD_ID,
            "oldProjectFolder": str(old_dir),
            "migrationWorkbook": str(STAGED),
            "migrationWorkbookSha256": sha256_file(STAGED),
            "pageMap": migration_pages,
            "migratedAt": utcnow(),
        }
        compiled["sourceLibrary"]["count"] = len(library.load()["sources"])
        compiled["projectSettings"] = {"externalMirrorPath": ""}
        store.save(new_id, compiled)

        old_manual = sum(object_count(page) for page in old_pages)
        new_manual = sum(object_count(page) for page in compiled["pages"])
        if new_manual != old_manual:
            raise RuntimeError(
                f"Manual object count changed from {old_manual} to {new_manual}."
            )
        result = {
            "ok": True, "oldProjectId": OLD_ID, "newProjectId": new_id,
            "newProjectFolder": str(new_dir),
            "workbookPath": compiled["workbookSync"]["workbook"],
            "workbookSha256": sha256_file(Path(compiled["workbookSync"]["workbook"])),
            "sheetCount": len(document["sheets"]), "indexRows": len(recipes),
            "pageCount": len(compiled["pages"]),
            "included": sum(bool(page.get("include")) for page in compiled["pages"]),
            "excluded": sum(not bool(page.get("include")) for page in compiled["pages"]),
            "oldManualObjects": old_manual, "newManualObjects": new_manual,
            "assetFiles": sum(1 for item in (new_dir / "assets").rglob("*") if item.is_file()),
            "sourceCount": len(library.load()["sources"]),
            "compilePreview": preview,
        }
        atomic_json_write(new_dir / "debug" / "sa31_migration_report.json", result)
        return result
    except Exception:
        if not keep_on_failure:
            shutil.rmtree(new_dir, ignore_errors=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--keep-on-failure", action="store_true")
    args = parser.parse_args()
    if not args.apply:
        profile = ProfileRegistry(ROOT / "defaults/project_templates/project_profiles.json").get("EMS_LIGHTING")
        validation = TemplateRegistry(DOCS).validate(STAGED)
        print(json.dumps({"dryRun": True, "profile": profile["id"], "validation": validation}, indent=2))
        return 0 if validation["valid"] else 1
    print(json.dumps(migrate(args.keep_on_failure), indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
