"""Smoke: compact/wide tables fit one page; split only when truly required."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.page_composer import BODY_BUDGET, BODY_W, EXCEL_MIN_SCALE, plan_excel_range
from core.workbook_importer import import_workbook


def _lighting_matrix(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["14", "LIGHTING-TDB", "Lighting Output Matrix", "YES", ""])

    ws = wb.create_sheet("LIGHTING-TDB")
    headers = ["Circuit", "Description", "From", "Offset", "Load", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True)
        ws.column_dimensions[chr(64 + c)].width = 18
    for r in range(2, 22):
        for c in range(1, 7):
            ws.cell(r, c, f"R{r}C{c}")
        ws.row_dimensions[r].height = 18
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "fit.xlsx"
    _lighting_matrix(xlsx)

    # Direct plan on a representative block shape.
    grid = [[f"h{c}" for c in range(6)]] + [[f"r{r}c{c}" for c in range(6)] for r in range(20)]
    block = {
        "type": "excelRange",
        "grid": grid,
        "rowHeights": [24] + [18] * 20,
        "colWidths": [120] * 6,
        "repeatRows": [0],
        "splitMode": "auto_rows",
        "allowContinuation": True,
        "minScale": EXCEL_MIN_SCALE,
    }
    plan = plan_excel_range(block)
    problems: list[str] = []
    if plan["pages"] != 1:
        problems.append(f"20-row lighting matrix should fit 1 page, plan={plan['pages']}")
    if plan["bestScale"] < EXCEL_MIN_SCALE:
        problems.append(f"bestScale below minScale: {plan['bestScale']}")

    natural_h = sum(block["rowHeights"])
    natural_w = sum(block["colWidths"])
    scale = min(BODY_W / natural_w, BODY_BUDGET / natural_h, 1.0)
    if scale * natural_h > BODY_BUDGET + 2:
        problems.append("scaled height exceeds body budget")

    proj = import_workbook(xlsx, project_id="fit1")
    pages = [p for p in proj["pages"] if p["sheetTab"] == "LIGHTING-TDB"]
    if len(pages) != 1:
        problems.append(f"import created {len(pages)} pages for compact lighting matrix")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — table fit passed")
    print(f"  plan pages={plan['pages']} bestScale={plan['bestScale']}")


if __name__ == "__main__":
    main()
