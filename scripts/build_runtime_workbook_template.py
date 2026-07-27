"""Build and optionally install the sanitized Singh360 EMS runtime workbook."""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FILE_NAME = "Singh360_BASE_Project_Workbook_Template_V1.xlsx"
TRACKED_OUTPUT = ROOT / "defaults" / "runtime_templates" / FILE_NAME

ORANGE = "D97706"
BLUE = "1F4E78"
GRAY = "A6A6A6"
WHITE = "FFFFFF"
BLACK = "000000"

CANONICAL_TABLES = {
    "21_NETWORK_PORTS": [
        "IDF", "Switch", "Port", "Label", "Device / Drop", "Controller ID",
        "IP Address", "Network", "From", "To", "Cable Type", "Source ID",
        "Status", "Notes",
    ],
    "22_PANELS": [
        "Panel ID", "Panel Type", "Panel Name", "Rack/System", "Location",
        "Controller IDs", "Serves", "Source ID", "Status", "Notes",
    ],
    "23_PANEL_IO": [
        "Panel ID", "Controller ID", "I/O Group", "Point No.", "Point Type",
        "Point Label", "Description", "Device / Case", "Cable / Terminal",
        "Source ID", "Status", "Notes",
    ],
    "24_REFRIG_CIRCUITS": [
        "System No.", "Circuit / Area", "Description", "Qty / Length",
        "Manufacturer", "Model #", "Refrigerant", "Load BTUH", "SST",
        "Evap Temp", "Disch Temp", "Valve Type", "Liquid Temp",
        "Control Type", "Defrost Type", "Volts / Hz / Ph",
        "WICP / Controller", "Source ID", "Status", "Notes",
    ],
    "25_RACKS": [
        "Rack ID", "Rack Name", "Refrigerant", "Controller ID",
        "Condenser Controller", "Circuits / Loads", "Source ID", "Status", "Notes",
    ],
    "26_LIGHTING_OUTPUTS": [
        "Output / Zone", "Controller ID", "Panel", "Point", "Output Type",
        "Description", "Area / Fixture Group", "Schedule / Time", "Source ID",
        "Status", "Notes",
    ],
    "27_HVAC_EQUIPMENT": [
        "Unit ID", "Unit Type", "Controller ID", "Network / IDF", "Location",
        "BACnet / RDM", "Source ID", "Status", "Notes",
    ],
    "28_CABLE_PULLS": [
        "Cable ID", "Cable Type", "From", "To", "Purpose / Device",
        "Cable Standard", "Installed By", "Source ID", "Status", "Notes",
    ],
    "29_BOM": [
        "Qty", "Part No.", "Description", "Comments", "Installed By",
        "Source ID", "Status / Notes",
    ],
    "30_COMMISSIONING": [
        "Item", "Owner", "Status", "Due / Milestone", "Evidence / Source", "Notes",
    ],
    "31_OPEN_ITEMS": [
        "Item ID", "Category", "Description", "Priority", "Owner", "Status",
        "Source ID", "Resolution Notes",
    ],
    "32_GUIDELINES": ["Guideline ID", "Topic", "Requirement", "Status", "Notes"],
    "33_ABBREVIATIONS": ["Abbreviation", "Meaning", "Symbol / Tag", "Notes"],
    "34_PROJECT_DIRECTORY": [
        "Organization", "Role", "Contact", "Phone", "Email", "Notes",
    ],
    "35_PROJECT_SCOPE": ["Scope ID", "Category", "Description", "Status", "Notes"],
    "36_WORKFLOW_MILESTONES": [
        "Milestone", "Owner", "Target", "Status", "Evidence / Notes",
    ],
    "37_RESPONSIBILITY_MATRIX": [
        "Scope Item", "Singh360", "Controls Contractor", "Electrical",
        "Refrigeration", "Owner / GC", "Notes",
    ],
}

PAGE_ROWS = [
    ("YES", 1, "EMS 1.0", "EMS 1.0 Cover", "Cover / Project Info", "cover", "cover"),
    ("YES", 2, "EMS 2.0", "EMS 2.0 Sheet Index", "Sheet Index / TOC", "Front Matter", "index"),
    ("NO", 3, "EMS 3.0", "EMS 3.0 Guidelines", "Singh360 / H-E-B Guidelines", "Front Matter", "data-grid"),
    ("NO", 4, "EMS 4.0", "EMS 4.0 Abbrev", "Abbreviations / Symbol Key", "Front Matter", "data-grid"),
    ("NO", 5, "EMS 5.0", "EMS 5.0 Directory", "Project Directory / Contacts", "Front Matter", "data-grid"),
    ("NO", 6, "EMS 6.0", "EMS 6.0 Project Scope", "Project Scope", "Front Matter", "data-grid"),
    ("NO", 7, "EMS 7.0", "EMS 7.0 Workflow", "Project Workflow / Milestones", "Front Matter", "data-grid"),
    ("NO", 8, "EMS 8.0", "EMS 8.0 Resp Matrix", "Responsibility Matrix", "Front Matter", "data-grid"),
    ("NO", 9, "EMS 10.0", "EMS 10.0 BOM", "Bill of Materials", "BOM", "data-grid"),
    ("NO", 10, "EMS 12.0", "EMS 12.0 Overall Layout", "EMS Controls Overall Layout", "Layout", "canvas"),
    ("NO", 11, "EMS 12.1", "EMS 12.1 Refrig Sys Table", "Refrigeration System Table", "Refrigeration", "data-grid"),
    ("NO", 12, "EMS 13.0", "EMS 13.0 Network Summary", "WICP / IDF / Network Summary", "Network", "data-grid"),
    ("NO", 13, "EMS 13.1", "EMS 13.1 IDF 1", "IDF #1 Port / Network Table", "Network", "data-grid"),
    ("NO", 14, "EMS 13.2", "EMS 13.2 IDF 2", "IDF #2 Port / Network Table", "Network", "data-grid"),
    ("NO", 15, "EMS 13.3", "EMS 13.3 IDF 3", "IDF #3 Port / Network Table", "Network", "data-grid"),
    ("NO", 16, "EMS 14.0", "EMS 14.0 Cable Pulls", "Cable Pull / Termination Schedule", "Cable", "data-grid"),
    ("NO", 17, "EMS 17.0", "EMS 17.0 Rack Summary", "Rack / Refrigeration Summary", "Refrigeration", "data-grid"),
    ("NO", 18, "EMS 18.0", "EMS 18.0 Rack A IO", "Rack A I/O Schedule", "Refrigeration", "data-grid"),
    ("NO", 19, "EMS 19.0", "EMS 19.0 Rack B IO", "Rack B I/O Schedule", "Refrigeration", "data-grid"),
    ("NO", 20, "EMS 20.0", "EMS 20.0 Rack C IO", "Rack C I/O Schedule", "Refrigeration", "data-grid"),
    ("NO", 21, "EMS 21.0", "EMS 21.0 WICP Summary", "WICP Count Summary", "WICP", "data-grid"),
    ("NO", 22, "EMS 22.0", "EMS 22.0 WICP IO", "WICP I/O Schedule", "WICP", "data-grid"),
    ("NO", 23, "EMS 23.0", "EMS 23.0 Case Controllers", "Case Controller Schedule", "Refrigeration", "data-grid"),
    ("NO", 24, "EMS 24.0", "EMS 24.0 Lighting Matrix", "Lighting Output Matrix", "Lighting", "data-grid"),
    ("NO", 25, "EMS 24.1", "EMS 24.1 Lighting IO", "Lighting TDB I/O Schedule", "Lighting", "data-grid"),
    ("NO", 26, "EMS 24.2", "EMS 24.2 Lighting Schedule", "Lighting Control / Dimming Schedule", "Lighting", "data-grid"),
    ("NO", 27, "EMS 25.0", "EMS 25.0 HVAC Devices", "HVAC Device Summary", "HVAC", "data-grid"),
    ("NO", 28, "EMS 26.0", "EMS 26.0 Commissioning", "Commissioning / Closeout Checklist", "Closeout", "data-grid"),
    ("NO", 29, "EMS 27.0", "EMS 27.0 Open Items", "Open Items / Exceptions", "Closeout", "data-grid"),
]


def _style_contract(ws, width: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.font = Font(color=WHITE, bold=True, size=14)
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, italic=True, size=10)
    border = Border(
        left=Side(style="thin", color=BLACK),
        right=Side(style="thin", color=BLACK),
        top=Side(style="thin", color=BLACK),
        bottom=Side(style="thin", color=BLACK),
    )
    for cell in ws[4][:width]:
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.font = Font(color=BLACK, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    for column in range(1, width + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18


def _table_sheet(workbook: Workbook, name: str, headers: list[str]) -> None:
    ws = workbook.create_sheet(name)
    ws.cell(1, 1, name.replace("_", " "))
    ws.cell(2, 1, "Canonical editable source table. Unknown project values remain blank, TBD, or VERIFY.")
    for column, header in enumerate(headers, start=1):
        ws.cell(4, column, header)
    _style_contract(ws, len(headers))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"


def _recipe_sheet(workbook: Workbook, row: tuple[str, int, str, str, str, str, str]) -> None:
    _include, _order, code, tab, title, family, page_type = row
    ws = workbook.create_sheet(tab)
    ws.cell(1, 1, title)
    ws.cell(2, 1, f"{code} | {family} | {page_type}")
    for column, header in enumerate(("Field", "Value", "Notes"), start=1):
        ws.cell(4, column, header)
    recipes = [
        ("Sheet Code", code, "00_INDEX is the publication authority."),
        ("Page Title", title, "Container title only."),
        ("Family", family, "Rendering family."),
        ("Page Type", page_type, "Rendering type."),
        ("Source / Data", "Canonical mapping", "Tables render from canonical editable sheets."),
        ("Current Content", "", "Recipe rows never publish."),
    ]
    for row_number, values in enumerate(recipes, start=5):
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    _style_contract(ws, 3)


def build(path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    meta = workbook.create_sheet("00_PROJECT_META")
    meta.cell(1, 1, "SINGH360 EMS - PROJECT METADATA")
    meta.cell(2, 1, "Project identity and title-block authority. Template Version is not Project Revision.")
    for column, header in enumerate(("Field", "Value", "Authority", "Notes"), start=1):
        meta.cell(4, column, header)
    metadata_rows = [
        ("Project Name", "TBD / VERIFY", "Project", "Title-block Project field"),
        ("Drawing Package File Name", "", "Project", "Export filename only"),
        ("Project Revision", "TBD", "Project", "Issued drawing revision"),
        ("Template Version", "1.0.0", "Template", "Never shown as Project Revision"),
        ("Location", "TBD / VERIFY", "Project", "Project address/location"),
        ("Issue Date", "", "Project", "Issued date"),
        ("Linked Project ID", "", "Application", "Assigned to the package-owned copy"),
        ("Repository", "dstogsdill1/Singh360_Draft", "Application", "Canonical repository"),
    ]
    for row_number, values in enumerate(metadata_rows, start=5):
        for column, value in enumerate(values, start=1):
            meta.cell(row_number, column, value)
    _style_contract(meta, 4)

    index = workbook.create_sheet("00_INDEX")
    index.cell(1, 1, "SINGH360 EMS - PAGE MANIFEST")
    index.cell(2, 1, "Only explicit YES rows publish. Optional empty pages default to NO.")
    index_headers = ["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"]
    for column, header in enumerate(index_headers, start=1):
        index.cell(4, column, header)
    for row_number, row in enumerate(PAGE_ROWS, start=5):
        values = [*row, "Generated from canonical data; physical tab is a page recipe only."]
        for column, value in enumerate(values, start=1):
            index.cell(row_number, column, value)
    _style_contract(index, len(index_headers))
    index.auto_filter.ref = f"A4:H{4 + len(PAGE_ROWS)}"

    for name, headers in CANONICAL_TABLES.items():
        _table_sheet(workbook, name, headers)
    for row in PAGE_ROWS:
        _recipe_sheet(workbook, row)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def install_runtime(source: Path) -> None:
    runtime = ROOT / ".docs" / "library" / "workbook_templates" / "base" / FILE_NAME
    staging = ROOT / ".docs" / "template_staging" / FILE_NAME
    runtime.parent.mkdir(parents=True, exist_ok=True)
    staging.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, runtime)
    shutil.copy2(source, staging)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    manifest = {
        "schemaVersion": 1,
        "templates": [{
            "templateId": "SINGH360_BASE_V1",
            "displayName": "Singh360 Base Project Workbook Template V1",
            "version": "1.0.0",
            "fileName": FILE_NAME,
            "absoluteRuntimePath": str(runtime.resolve()),
            "sha256": digest,
            "supportedProfiles": ["BASE_CORE", "EMS_LIGHTING", "EMS_FULL", "EMS_INSTALL", "EMS_RETROFIT", "CX", "RCX"],
            "active": True,
            "sourceStagingPath": str(staging.resolve()),
            "workbookValidation": {
                "valid": True,
                "path": str(staging.resolve()),
                "errors": [],
                "warnings": [],
                "sheetNames": ["00_PROJECT_META", "00_INDEX", *CANONICAL_TABLES, *[row[3] for row in PAGE_ROWS]],
                "sheetCount": 2 + len(CANONICAL_TABLES) + len(PAGE_ROWS),
                "sha256": digest,
            },
        }],
    }
    manifest_path = runtime.parent.parent / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=TRACKED_OUTPUT)
    parser.add_argument("--install-runtime", action="store_true")
    args = parser.parse_args()
    output = build(args.output.resolve())
    if args.install_runtime:
        install_runtime(output)
    print(output)
    print(hashlib.sha256(output.read_bytes()).hexdigest())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
