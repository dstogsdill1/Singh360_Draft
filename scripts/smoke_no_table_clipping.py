"""Smoke: instruction tables render all current rows without clipping warnings."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.page_composer import page_render_diagnostics
from core.workbook_importer import import_workbook


def _add_instruction_sheet(wb: Workbook, name: str) -> None:
    ws = wb.create_sheet(name)
    ws.append(["Step", "Instruction", "Responsible"])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.font = Font(bold=True)
    for r in range(1, 6):
        ws.append([
            str(r),
            "Verify field wiring, label conductors, coordinate with controls team, and document completion before turnover.",
            "EC/DC/EMS",
        ])


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    for code, tab in [("16", "EC Field Instructions"), ("17", "DC Field Instructions"), ("18", "EMS Remote Instructions")]:
        idx.append([code, tab, tab, "YES", ""])
        _add_instruction_sheet(wb, tab)
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "instructions.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="clip1")
    diagnostics = page_render_diagnostics(proj["pages"])
    problems: list[str] = []

    for tab in ("EC Field Instructions", "DC Field Instructions", "EMS Remote Instructions"):
        pages = [p for p in proj["pages"] if p["sheetTab"] == tab]
        if len(pages) != 1:
            problems.append(f"{tab}: expected one page, got {len(pages)}")
            continue
        block = pages[0]["blocks"][0]
        # header + five instruction rows
        if len(block.get("grid") or []) != 6:
            problems.append(f"{tab}: not all five instruction rows rendered")
        d = next((x for x in diagnostics if x["pageTitle"] == tab), None)
        if d and d.get("clipping"):
            problems.append(f"{tab}: diagnostics reported clipping")
        if pages[0].get("layoutWarnings") or block.get("layoutWarnings"):
            problems.append(f"{tab}: layout warning present")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - no table clipping for instruction pages passed")


if __name__ == "__main__":
    main()
