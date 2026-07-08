"""Smoke: SA31-style title-block metadata inference from cover key/value
pairs when no dedicated metadata sheet exists (FINAL RELEASE CLEANUP
4H+SA38, Phase B).

Mirrors the real SA31 cover layout: label/value pairs across a row
(``Address`` / ``Package`` / ``Revision`` / ``Prepared By``), no
``00_PROJECT_META`` sheet. Verifies:
  - Location is inferred from ``Address``.
  - Revision is inferred from ``Revision``.
  - Drawing Package File Name is inferred from ``Package``.
  - Drawn By falls back through ``Prepared By`` -> "Singh360 Inc." (never a
    dash / blank when the workbook actually names a preparer).
  - A workbook metadata sheet, when present, overrides cover inference.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from core.workbook_importer import import_workbook


def _cover_only_workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "EMS 0.0 Cover", "Cover / Project Info", "Front Matter", "cover", ""])
    idx.append(["YES", 2, "EMS 0.1", "00_INDEX", "Sheet Index / TOC", "Front Matter", "index", ""])

    cover = wb.create_sheet("EMS 0.0 Cover")
    cover.append(["LIGHT DIMMING & EMS INTEGRATION (SA31 - HEB 102)"])
    cover.append(["SINGH360 EMS CONTROLS WORKBOOK"])
    cover.append(["Project", "San Antonio 31 - HEB 102", "Store Name", "H-E-B SA #31 - HEB 102"])
    cover.append(["Address", "8503 NW Military Highway, San Antonio, TX 78231", "Project Type", "Light Dimming & EMS Integration"])
    cover.append(["Package", "SA31_EMS_Lighting_V1", "Revision", "V1"])
    cover.append(["Prepared For", "H-E-B / Field Subcontractors", "Prepared By", "Singh360 Inc."])

    wb.save(path)


def _metadata_overrides_cover_workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Include", "Order", "Sheet Code", "Sheet Tab", "Page Title", "Family", "Page Type", "Notes"])
    idx.append(["YES", 1, "EMS 0.0", "EMS 0.0 Cover", "Cover / Project Info", "Front Matter", "cover", ""])

    cover = wb.create_sheet("EMS 0.0 Cover")
    cover.append(["Address", "Cover Address (should be overridden)"])
    cover.append(["Revision", "V0 (should be overridden)"])

    meta = wb.create_sheet("00_PROJECT_META")
    meta.append(["Location", "Metadata Sheet Location Wins"])
    meta.append(["Revision", "V9"])

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())

    xlsx1 = tmp / "sa31_cover_only.xlsx"
    _cover_only_workbook(xlsx1)
    proj1 = import_workbook(xlsx1, project_id="sa31coveronly")
    meta1 = proj1["metadata"]

    problems: list[str] = []

    expected1 = {
        "location": "8503 NW Military Highway, San Antonio, TX 78231",
        "revision": "V1",
        "drawingPackageFileName": "SA31_EMS_Lighting_V1",
        "drawnBy": "Singh360 Inc.",
    }
    for field, expected in expected1.items():
        got = meta1.get(field)
        if got != expected:
            problems.append(f"cover-inferred metadata[{field}]={got!r}, expected {expected!r}")
    if not meta1.get("location") or meta1.get("location") == "-":
        problems.append("Location left blank/dash despite cover Address being present")
    if not meta1.get("revision") or meta1.get("revision") == "-":
        problems.append("Revision left blank/dash despite cover Revision being present")

    xlsx2 = tmp / "sa31_meta_overrides_cover.xlsx"
    _metadata_overrides_cover_workbook(xlsx2)
    proj2 = import_workbook(xlsx2, project_id="sa31metaoverride")
    meta2 = proj2["metadata"]
    if meta2.get("location") != "Metadata Sheet Location Wins":
        problems.append(f"metadata sheet should win over cover inference for location, got {meta2.get('location')!r}")
    if meta2.get("revision") != "V9":
        problems.append(f"metadata sheet should win over cover inference for revision, got {meta2.get('revision')!r}")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — SA31 cover-inferred + metadata-sheet-precedence title-block metadata passed")
    print(f"  location={meta1.get('location')!r}, revision={meta1.get('revision')!r}, package={meta1.get('drawingPackageFileName')!r}, drawnBy={meta1.get('drawnBy')!r}")


if __name__ == "__main__":
    main()
