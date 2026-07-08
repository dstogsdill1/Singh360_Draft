"""Smoke: LCP / Lighting continuations split by logical sections, not tails."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")


def _section_row(ws, row: int, title: str, cols: int = 6) -> None:
    ws.cell(row, 1, title)
    ws.cell(row, 1).fill = GOLD
    ws.cell(row, 1).font = Font(bold=True)
    for c in range(2, cols + 1):
        ws.cell(row, c).fill = GOLD


def _headers(ws, row: int, cols: int = 6) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c, f"Header {c}")
        cell.fill = GRAY
        cell.font = Font(bold=True)


def _workbook(path: Path) -> None:
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["14", "LIGHTING-TDB", "Lighting Output Matrix", "YES", ""])
    idx.append(["15", "LCP Panel", "LCP Panel Schedule", "YES", ""])

    lighting = wb.create_sheet("LIGHTING-TDB")
    _section_row(lighting, 1, "Controller I/O Table")
    _headers(lighting, 2)
    r = 3
    for i in range(28):
        for c in range(1, 7):
            lighting.cell(r, c, f"IO {i}-{c}")
        lighting.row_dimensions[r].height = 30
        r += 1
    _section_row(lighting, r, "Relay / Contactor Schedule")
    relay_row = r
    r += 1
    _headers(lighting, r)
    r += 1
    for i in range(26):
        for c in range(1, 7):
            lighting.cell(r, c, f"Relay {i}-{c}")
        lighting.row_dimensions[r].height = 30
        r += 1

    lcp = wb.create_sheet("LCP Panel")
    _section_row(lcp, 1, "LCP-1 Dimming Panel")
    _headers(lcp, 2)
    r = 3
    for i in range(18):
        for c in range(1, 7):
            lcp.cell(r, c, f"LCP1 {i}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1
    _section_row(lcp, r, "Expansion I/O Device PR0663")
    r += 1
    _headers(lcp, r)
    r += 1
    for i in range(10):
        for c in range(1, 7):
            lcp.cell(r, c, f"PR0663 {i}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1
    _section_row(lcp, r, "LCP-2 Contactor Panel")
    lcp2_row = r
    r += 1
    _headers(lcp, r)
    r += 1
    for i in range(26):
        for c in range(1, 7):
            lcp.cell(r, c, f"RO{i + 1}-{c}")
        lcp.row_dimensions[r].height = 30
        r += 1

    wb.save(path)


def _page_text(page: dict) -> str:
    return "\n".join(" | ".join(r) for b in page.get("blocks") or [] for r in (b.get("grid") or []))


def _merged_section(ws, row: int, title: str, cols: int = 6) -> int:
    """A Kyle/SA38-style gold section band as ONE wide merged cell (matches
    the real workbook: openpyxl only carries the fill on the merge's anchor
    cell, not on every covered cell) followed by its own gray header row.
    Returns the row number of the gray header (row + 1).
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row, 1, title)
    cell.fill = GOLD
    cell.font = Font(bold=True)
    _headers(ws, row + 1, cols)
    return row + 1


def _sa38_style_workbook(path: Path, big_section_rows: int = 0) -> None:
    """RACK A/B-style sheet: repeating (merged gold band -> gray header ->
    data rows) controller sections, big enough to force >1 continuation.
    When ``big_section_rows`` is set, the LAST section alone is oversized so
    it can't fit on a single page (forces a mid-section hard split).
    """
    wb = Workbook()
    idx = wb.active
    idx.title = "00_INDEX"
    idx.append(["Order", "Sheet Tab", "Page Title", "Include?", "Use Source"])
    idx.append(["1", "RACK A", "Rack A I/O & Layout", "YES", ""])

    ws = wb.create_sheet("RACK A")
    r = 1
    section_names = [
        "SuperPak Controller I/O Panel for RACK A - PR0650CD-SUP",
        "SuperPak Controller I/O Panel for RACK B - PR0650CD-SUP",
        "Expansion I/O Device PR0663",
    ]
    for name in section_names:
        header_row = _merged_section(ws, r, name)
        r = header_row + 1
        for i in range(12):
            for c in range(1, 7):
                ws.cell(r, c, f"{name[:6]}-{i}-{c}")
            ws.row_dimensions[r].height = 30
            r += 1

    if big_section_rows:
        header_row = _merged_section(ws, r, "Oversized Final Controller Section")
        r = header_row + 1
        for i in range(big_section_rows):
            for c in range(1, 7):
                ws.cell(r, c, f"BIG-{i}-{c}")
            ws.row_dimensions[r].height = 30
            r += 1

    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "logical_splits.xlsx"
    _workbook(xlsx)
    proj = import_workbook(xlsx, project_id="split1")
    problems: list[str] = []

    lcp_pages = [p for p in proj["pages"] if p["sheetTab"] == "LCP Panel"]
    if len(lcp_pages) != 2:
        problems.append(f"LCP expected 2 logical pages, got {len(lcp_pages)}")
    else:
        if "LCP-2 Contactor Panel" not in _page_text(lcp_pages[1]):
            problems.append("LCP continuation does not start with LCP-2 section")
        data_rows = len(lcp_pages[1]["blocks"][0].get("grid") or [])
        if data_rows < 8:
            problems.append("LCP continuation is an orphan tail")

    lighting_pages = [p for p in proj["pages"] if p["sheetTab"] == "LIGHTING-TDB"]
    if len(lighting_pages) != 2:
        problems.append(f"Lighting expected 2 logical pages, got {len(lighting_pages)}")
    else:
        if "Relay / Contactor Schedule" not in _page_text(lighting_pages[1]):
            problems.append("Lighting continuation does not start with relay/contactor section")

    # --- Phase F: Kyle/SA38-style repeating gold(merged)+gray controller
    # sections must only ever be cut at a section boundary, and a forced
    # mid-section split repeats that section's own gold+gray header. ---
    from core.table_style_profile import is_gold_fill

    def _gold_rows(page: dict) -> list[int]:
        block = (page.get("blocks") or [{}])[0]
        grid = block.get("grid") or []
        styles = block.get("styles") or {}
        merges = block.get("mergedCells") or []
        wide: set[int] = set()
        for m in merges:
            if m.get("startRow") == m.get("endRow"):
                fill = (styles.get(f"{m['startRow']}:{m['startCol']}") or {}).get("fill")
                if is_gold_fill(fill):
                    wide.add(m["startRow"])
        found = []
        for r in range(len(grid)):
            if r in wide:
                found.append(r)
                continue
            ncols = len(grid[r])
            gold_cols = sum(1 for c in range(ncols) if is_gold_fill((styles.get(f"{r}:{c}") or {}).get("fill")))
            if ncols and gold_cols >= max(2, ncols // 2):
                found.append(r)
        return found

    xlsx_sa38 = tmp / "sa38_sections.xlsx"
    _sa38_style_workbook(xlsx_sa38)
    proj_sa38 = import_workbook(xlsx_sa38, project_id="sa38sections")
    rack_pages = [p for p in proj_sa38["pages"] if p["sheetTab"] == "RACK A"]

    if len(rack_pages) < 2:
        problems.append(f"SA38-style RACK A: expected >1 continuation page, got {len(rack_pages)}")
    else:
        for i, page in enumerate(rack_pages):
            block = (page.get("blocks") or [{}])[0]
            grid = block.get("grid") or []
            gold_rows = _gold_rows(page)
            if i > 0 and not gold_rows:
                problems.append(f"SA38 RACK A page {i} ({page.get('displaySheetCode')}) has no gold section header — orphaned data with no section label")
            # No orphan tail: every continuation page (after the first) must
            # start its body with a section header very near the top, not
            # buried after a run of unlabeled leftover data rows.
            if i > 0 and gold_rows and min(gold_rows) > 3:
                problems.append(f"SA38 RACK A page {i}: first gold row at index {min(gold_rows)} — split landed mid-section, not at a boundary")
            if not grid:
                problems.append(f"SA38 RACK A page {i}: empty grid")

    xlsx_forced = tmp / "sa38_forced_split.xlsx"
    _sa38_style_workbook(xlsx_forced, big_section_rows=60)
    proj_forced = import_workbook(xlsx_forced, project_id="sa38forcedsplit")
    forced_pages = [p for p in proj_forced["pages"] if p["sheetTab"] == "RACK A"]
    big_section_pages = [
        p for p in forced_pages
        if "Oversized Final Controller Section" in _page_text(p)
    ]
    if len(big_section_pages) < 2:
        problems.append(
            f"Forced mid-section split: expected the oversized section to span >1 page, got {len(big_section_pages)}"
        )
    else:
        for i, page in enumerate(big_section_pages):
            if i == 0:
                continue
            if "Oversized Final Controller Section" not in _page_text(page):
                problems.append(
                    f"Forced mid-section split: continuation {i} of the oversized section does not repeat its own gold+gray header"
                )
            gold_rows = _gold_rows(page)
            if not gold_rows or min(gold_rows) > 3:
                problems.append(
                    f"Forced mid-section split: continuation {i} does not repeat the gold header near the top of the page (gold_rows={gold_rows})"
                )

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK - logical section splits passed")
    print(f"  SA38 RACK A pages={len(rack_pages)}, forced-split oversized-section pages={len(big_section_pages)}")


if __name__ == "__main__":
    main()
