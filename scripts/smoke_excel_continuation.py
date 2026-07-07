"""Smoke test: Excel exact range continuation rules."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Side, Border

from core.page_composer import EXCEL_MIN_SCALE, _split_excel_range_block, plan_excel_range
from core.workbook_importer import import_workbook

GOLD = PatternFill("solid", fgColor="FFC000")
GRAY = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _compact_matrix(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RESP MATRIX"
    ws.merge_cells("B1:D1")
    ws["B1"] = "Furnished by"
    ws["B1"].fill = GOLD
    for col, name in zip("BCD", ["Owner", "EC", "EMS"]):
        c = ws[f"{col}2"]
        c.value = name
        c.fill = GRAY
        c.font = Font(bold=True)
        c.border = BOX
    for r in range(3, 40):
        ws[f"A{r}"] = f"Task {r-2}"
        for col in "BCD":
            ws[f"{col}{r}"] = "X"
            ws[f"{col}{r}"].border = BOX
    wb.save(path)


def _tall_schedule(path: Path, rows: int = 120) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "LCP Panel"
    ws.merge_cells("A1:E1")
    ws["A1"] = "LCP-1 CONTROLLER"
    ws["A1"].fill = GOLD
    ws["A1"].font = Font(bold=True)
    for i, h in enumerate(["Point", "Type", "Addr", "Desc", "Notes"], start=1):
        c = ws.cell(2, i, h)
        c.fill = GRAY
        c.font = Font(bold=True)
    for r in range(3, rows + 3):
        for c in range(1, 6):
            ws.cell(r, c, f"r{r}c{c}").border = BOX
        ws.row_dimensions[r].height = 24
    wb.save(path)


def main() -> None:
    tmp = Path(tempfile.mkdtemp())
    problems: list[str] = []

    # Compact matrix → one page (scales to fit, no split).
    mx = tmp / "matrix.xlsx"
    _compact_matrix(mx)
    proj = import_workbook(mx, project_id="m1")
    rm_pages = [p for p in proj["pages"] if p["sheetTab"] == "RESP MATRIX"]
    if len(rm_pages) != 1:
        problems.append(f"compact matrix: expected 1 page, got {len(rm_pages)}")
    if not proj.get("paginationLocked"):
        problems.append("paginationLocked not set on import")

    # Tall schedule → multiple pages, header repeated, no duplication.
    lcp = tmp / "lcp.xlsx"
    _tall_schedule(lcp, 120)
    proj2 = import_workbook(lcp, project_id="l2")
    lcp_pages = [p for p in proj2["pages"] if p["sheetTab"] == "LCP Panel"]
    if len(lcp_pages) < 2:
        problems.append(f"tall LCP: expected multiple pages, got {len(lcp_pages)}")
    data_rows = set()
    for lp in lcp_pages:
        b = lp["blocks"][0]
        src = b.get("srcRows") or []
        repeat = set(b.get("repeatRows") or [])
        for r in src:
            if r not in repeat:
                if r in data_rows:
                    problems.append(f"duplicate data row {r} across continuation pages")
                data_rows.add(r)
        if not (b["grid"][0][0] or "").startswith("LCP"):
            problems.append(f"header not repeated on {lp['displaySheetCode']}")

    # splitMode none → never splits.
    block = {
        "type": "excelRange",
        "grid": [["h"], *[[f"r{i}"] for i in range(200)]],
        "rowHeights": [20] * 201,
        "colWidths": [800],
        "repeatRows": [0],
        "splitMode": "none",
        "allowContinuation": False,
        "minScale": EXCEL_MIN_SCALE,
    }
    parts = _split_excel_range_block(block)
    if len(parts) != 1:
        problems.append(f"splitMode none: expected 1 part, got {len(parts)}")

    plan = plan_excel_range(block)
    if plan["pages"] != 1:
        problems.append("plan_excel_range wrong for splitMode none")

    if problems:
        print("FAIL")
        for p in problems:
            print(" -", p)
        raise SystemExit(1)
    print("OK — continuation rules passed")
    print(f"  matrix pages: {len(rm_pages)}")
    print(f"  LCP pages: {len(lcp_pages)} (unique data rows: {len(data_rows)})")


if __name__ == "__main__":
    main()
