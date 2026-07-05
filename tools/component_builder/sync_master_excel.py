#!/usr/bin/env python3
"""sync_master_excel.py -- turn the editable master Excel workbook into a clean CSV.

The Excel workbook is the human-editable master. This script regenerates the
derived CSV (and extracts any pasted images into sources/<category>/) so the rest
of the workbench has a stable, clean manifest to read.

  Input :  .docs/component_builder/master/Singh360_Component_Master_Catalog.xlsx
  Output:  .docs/component_builder/master/Singh360_Component_Master_Catalog.csv
           .docs/component_builder/master/sources/<category>/<clean>.png  (if pasted)

Rules (deliberately simple -- no AI, no web, no guessing):
  * displayName is taken exactly as the user typed it.
  * category / manufacturer / templateType are used as-is when filled.
  * blank category -> a few keyword rules fill a best-guess AND needsReview=true.
  * pasted workbook images are extracted where practical; curated files on disk
    are preserved and never overwritten unless --replace is given.
  * part numbers are never invented.
  * a row with no usable image is kept but flagged needsReview=true, unless its
    templateType is specific enough to draw procedurally.

Usage:
    python tools/component_builder/sync_master_excel.py [--replace]
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _catalog  # noqa: E402

TOOL_DIR = Path(__file__).resolve().parent
TAXONOMY_PATH = TOOL_DIR / "component_taxonomy.json"

MAIN_SHEET_CANDIDATES = ("Component_Master", "Component_Catalog", "Components", "Master")

# Canonical output column order (stable contract for downstream scripts).
COLUMNS = [
    "rowNumber", "componentId", "displayName", "manufacturer", "category",
    "partNumber", "aliases", "sourceImageFile", "sourceWorkbook",
    "sourceWorksheet", "sourceAnchorCell", "sourceType", "templateType",
    "defaultLabel", "addressFieldNeeded", "topTerminals", "bottomTerminals",
    "leftPorts", "rightPorts", "widthUnits", "heightUnits", "labelPosition",
    "useInDrawings", "symbolStatus", "needsReview", "priority", "notes",
]


def _load_taxonomy() -> list[dict]:
    try:
        return json.loads(TAXONOMY_PATH.read_text(encoding="utf-8")).get("categories", [])
    except Exception:
        return []


def _guess_category(text: str, categories: list[dict]) -> str | None:
    t = (text or "").lower()
    best, best_hits = None, 0
    for cat in categories:
        if cat["id"] == "custom":
            continue
        hits = sum(1 for kw in cat.get("keywords", []) if kw.lower() in t)
        if hits > best_hits:
            best, best_hits = cat["id"], hits
    return best


def _pick_sheet(wb):
    for name in MAIN_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    # fallback: first sheet whose header row mentions componentId/displayName
    for ws in wb.worksheets:
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        if "componentId" in header or "displayName" in header:
            return ws
    return wb.worksheets[0]


def _extract_images(ws, source_root: Path, rows_by_index: dict[int, dict],
                    replace: bool) -> int:
    """Save pasted workbook images into sources/<category>/ where practical."""
    extracted = 0
    for img in getattr(ws, "_images", []):
        try:
            anchor_row = img.anchor._from.row  # 0-based
        except Exception:
            continue
        row = rows_by_index.get(anchor_row + 1)  # 1-based data lookup
        if not row:
            continue
        cat = _catalog.slug(row.get("category") or "custom")
        name = _catalog.slug(row.get("displayName") or row.get("componentId") or "image")
        out_dir = source_root / cat
        out_dir.mkdir(parents=True, exist_ok=True)
        target = out_dir / f"{name}.png"
        if target.exists() and not replace:
            if not row.get("sourceImageFile"):
                row["sourceImageFile"] = f"sources/{cat}/{name}.png"
            continue
        try:
            data = img._data()
        except Exception:
            continue
        target.write_bytes(data)
        row["sourceImageFile"] = f"sources/{cat}/{name}.png"
        extracted += 1
    return extracted


def sync(workbook: Path, out_csv: Path, source_root: Path, replace: bool) -> dict:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(workbook)
    ws = _pick_sheet(wb)
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    categories = _load_taxonomy()
    rows: list[dict] = []
    rows_by_index: dict[int, dict] = {}

    for idx, xlrow in enumerate(ws.iter_rows(min_row=2), start=2):
        values = {header[i]: (c.value if c.value is not None else "")
                  for i, c in enumerate(xlrow) if i < len(header)}
        if not any(str(v).strip() for v in values.values()):
            continue
        row = {col: str(values.get(col, "") or "").strip() for col in COLUMNS}

        if not row["displayName"] and not row["componentId"]:
            continue
        if not row["rowNumber"]:
            row["rowNumber"] = str(idx)

        review = str(values.get("needsReview", "")).strip().upper() in {"TRUE", "YES", "1"}

        if not row["category"]:
            guess = _guess_category(f"{row['displayName']} {row['notes']} {row['aliases']}",
                                    categories)
            row["category"] = guess or "custom"
            review = True
        if not row["manufacturer"]:
            row["manufacturer"] = "Generic"
        if not row["componentId"]:
            row["componentId"] = _catalog.slug(f"{row['manufacturer']}_{row['displayName']}")

        row["needsReview"] = "TRUE" if review else (row["needsReview"] or "FALSE")
        rows.append(row)
        rows_by_index[idx] = row

    extracted = _extract_images(ws, source_root, rows_by_index, replace)

    # finalize needsReview based on image availability + template specificity
    missing_no_template = 0
    for row in rows:
        src = _catalog.resolve_source_image(row["sourceImageFile"], workbook.parent, source_root)
        has_img = src is not None
        specific = row["templateType"] in _catalog.SPECIFIC_TEMPLATES
        if not has_img and not specific:
            row["needsReview"] = "TRUE"
            missing_no_template += 1

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    with_img = sum(
        1 for r in rows
        if _catalog.resolve_source_image(r["sourceImageFile"], workbook.parent, source_root)
    )
    return {
        "rows": len(rows),
        "withImage": with_img,
        "extracted": extracted,
        "missingNoTemplate": missing_no_template,
        "sheet": ws.title,
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", default=None,
                    help="Master .xlsx (default: master/Singh360_Component_Master_Catalog.xlsx).")
    ap.add_argument("--out-csv", default=None,
                    help="Output CSV (default: alongside the workbook).")
    ap.add_argument("--source-root", default=None,
                    help="Sources root (default: master/sources).")
    ap.add_argument("--replace", action="store_true",
                    help="Overwrite existing extracted source images.")
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print("[error] openpyxl is required. pip install openpyxl", file=sys.stderr)
        return 3

    args = parse_args(argv)
    workbook = Path(args.workbook).resolve() if args.workbook else \
        (_catalog.PACKAGE_DIR / "Singh360_Component_Master_Catalog.xlsx")
    if not workbook.exists():
        print(f"[error] workbook not found: {workbook}\n"
              f"        expected the master at {_catalog.rel_to_repo(_catalog.PACKAGE_DIR)}",
              file=sys.stderr)
        return 2

    out_csv = Path(args.out_csv).resolve() if args.out_csv else \
        workbook.with_suffix(".csv")
    source_root = Path(args.source_root).resolve() if args.source_root else \
        (workbook.parent / "sources")
    source_root.mkdir(parents=True, exist_ok=True)

    stats = sync(workbook, out_csv, source_root, args.replace)

    print(f"[ok] read sheet '{stats['sheet']}' from {_catalog.rel_to_repo(workbook)}")
    print(f"[ok] wrote {stats['rows']} row(s) -> {_catalog.rel_to_repo(out_csv)}")
    print(f"[ok] rows with a source image: {stats['withImage']} | "
          f"extracted from workbook: {stats['extracted']}")
    if stats["missingNoTemplate"]:
        print(f"[note] {stats['missingNoTemplate']} row(s) have no image and no specific "
              "template -> flagged needsReview=TRUE.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
